from flask import Flask, request, jsonify, session
import anthropic
import os
import requests
from datetime import datetime, timedelta
from functools import wraps
import json, re, uuid, time as _time, threading
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "rjgrooming-secret-2024")
client_ai = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

WHATSAPP_TOKEN    = os.environ.get("WHATSAPP_TOKEN")
WHATSAPP_PHONE_ID = os.environ.get("WHATSAPP_PHONE_ID")
VERIFY_TOKEN      = os.environ.get("WHATSAPP_VERIFY_TOKEN")
INSTAGRAM_VERIFY_TOKEN = os.environ.get("INSTAGRAM_VERIFY_TOKEN", "mydeal13")
INSTAGRAM_TOKEN   = os.environ.get("INSTAGRAM_TOKEN", os.environ.get("WHATSAPP_TOKEN"))

conversation_history = {}
MAX_HISTORY = 20
_msg_log = []          # recent messages: {ts, phone, channel, direction, text}
_channel_counts = {"whatsapp": 0, "instagram": 0, "facebook": 0}

def _log_msg(phone, channel, direction, text):
    _msg_log.append({"ts": datetime.now().isoformat(), "phone": str(phone),
                     "channel": channel, "direction": direction, "text": str(text)[:300]})
    if len(_msg_log) > 100:
        del _msg_log[:-100]
jarvis_enabled    = True
instagram_enabled = True
pause_until       = None
manual_mode       = False
clients           = {}

SYSTEM_PROMPT = """You are Jarvis, the AI administrator of R&J Grooming in Tallinn, Estonia.
- Respond in the client language (Russian, Estonian, English)
- Keep answers short: 1-3 sentences
- Greet only on first message
- Booking: https://n1425894.alteg.io"""

# ── WA funnel SYSTEM_PROMPT (loaded from whatsapp_bot.js) ────────────────────
def _load_wa_prompt():
    try:
        p = os.path.join(os.path.dirname(__file__), "whatsapp_bot.js")
        src = open(p, encoding="utf-8").read()
        m = re.search(r"const SYSTEM_PROMPT = `([\s\S]*?)`;", src)
        return m.group(1) if m else ""
    except Exception:
        return ""

WA_SYSTEM_PROMPT = _load_wa_prompt()

_FUNNEL_RULES = """

═══ ВОРОНКА /test-chat — СТРОГО ПО ШАГАМ, ПРОПУСКИ ЗАПРЕЩЕНЫ ═══

Шаг 1 → Анна: «Здравствуйте! Я Анна, администратор R&J Grooming 🐾 Чем могу помочь?»
Шаг 2 → Клиент написал запрос: ответь тепло, спроси ТОЛЬКО породу.
         Пример: «Замечательно! Подскажите, пожалуйста, какая у вас порода?»
         ❌ НЕ спрашивай вес на этом шаге.
Шаг 2б → ВЫПОЛНЯЕТСЯ КОДОМ: если порода требует уточнения подвида (такса / пудель /
          шпиц / чихуахуа / йорк) — код спрашивает подвид автоматически.
          ❌ НЕ называй услуги и цену до уточнения подвида.
Шаг 3 → ВЫПОЛНЯЕТСЯ КОДОМ: порода (с подвидом если нужен) известна, вес не известен →
         код автоматически делает комплимент породе и спрашивает вес.
         ❌ НЕ спрашивай сам — дождись кода. ❌ НЕ называй услуги.
Шаг 4 → ВЫПОЛНЯЕТСЯ КОДОМ: порода + вес известны →
         код показывает услуги для породы от простого к дорогому + рекомендует комплексный.
         ❌ НЕ перечисляй услуги самостоятельно.
Шаг 5 → Клиент выбрал услугу: жди подтверждения («да», «подходит», «хочу»).
Шаг 6 → ВЫПОЛНЯЕТСЯ КОДОМ: цена «от N€» + дисклеймер + «Хотите записаться?»
         ❌ НЕ называй цену самостоятельно.
Шаг 7 → Клиент «да» на «Хотите записаться?»: ответь ТОЛЬКО «Одну минуту, проверяю расписание 🐾»
Шаг 8 → После расписания: жди конкретную дату и время.
Шаг 9 → Спроси имя владельца и кличку питомца (можно в одном сообщении).
Шаг 9б → Спроси номер телефона для SMS-подтверждения:
          «Укажите, пожалуйста, номер телефона в формате +372XXXXXXX 🤍»
          Если клиент дал номер без +372 или в другом формате — уточни:
          «Пожалуйста, укажите полный номер в формате +372XXXXXXX (например, +37253112233) 🤍»
          Код проверяет формат автоматически и попросит исправить если нужно.
Шаг 10 → Карточка: Порода | Услуга | Дата | Время | Владелец | Питомец | Телефон.
          «Всё верно?» → «да» → «Запись принята! Ждём вас 🐾»

ЕСЛИ порода не известна → спроси породу (шаг 2).
ЕСЛИ порода с разновидностями (такса/пудель/шпиц/чихуахуа/йорк) без подвида → жди кода (шаг 2б).
ЕСЛИ порода (с подвидом) известна, вес НЕ известен → жди кода (шаг 3). Сам не действуй.
ЕСЛИ порода и вес известны → жди кода (шаг 4). Сам не действуй.
НЕЛЬЗЯ переходить к услугам пока не известен вес и подвид (если порода требует).
НЕЛЬЗЯ называть цену пока клиент не подтвердил услугу.
НЕЛЬЗЯ показывать слоты пока клиент не ответил «да» на «Хотите записаться?».

ОТМЕНА ЗАПИСИ:
Клиент: «отмените», «хочу отменить», «не приду» →
  1. Если телефон неизвестен — спроси (+372XXXXXXX).
  2. Уточни: «Вы хотите полностью отменить запись?»
  3. После «да» — ответь: «Хорошо, отменяю 🤍 [[ACTION:cancel:+372XXXXXXX]]»

ПЕРЕНОС ЗАПИСИ:
Клиент: «хочу перенести», «другое время», «изменить дату» →
  1. Если телефон неизвестен — спроси его.
  2. Покажи свободные слоты, спроси удобное время.
  3. После выбора — ответь: «Переношу на [дата] в [время] 🐾 [[ACTION:reschedule:+372XXXXXXX:DD.MM.YYYY:HH:MM]]»

Маркер [[ACTION:...]] оставь в ответе — система исполняет его и удаляет перед отправкой.

ЗАПРЕЩЕНО: пропускать шаги | спрашивать породу и вес вместе | два вопроса в одном сообщении | слово «малыш» | бронирование через сайт."""

TEST_CHAT_SYSTEM_PROMPT = WA_SYSTEM_PROMPT + _FUNNEL_RULES

# ── Test-chat in-memory sessions ─────────────────────────────────────────────
_chat_sessions = {}  # sid -> {history, state}

def _blank_state():
    return {"breed": None, "weight": None, "service": None, "date": None, "time": None,
            "ownerName": None, "petName": None, "master": None,
            "clientPhone": None, "confirmed": False,
            "service_confirmed": False, "booking_intent": False,
            "price_shown": False}

def _state_context(state):
    labels = {"breed": "Порода", "weight": "Вес", "service": "Услуга", "date": "Дата",
              "time": "Время", "ownerName": "Имя владельца", "petName": "Кличка",
              "master": "Мастер", "clientPhone": "Телефон для SMS"}
    filled = [f"{labels[k]}: {state[k]}" for k in labels if state.get(k)]
    ctx = ("\n\nТЕКУЩИЕ ДАННЫЕ КЛИЕНТА (уже известны, НЕ переспрашивай):\n"
           + "\n".join(filled)) if filled else ""
    breed = state.get("breed")
    if breed and not state.get("service"):
        top_svc, top_price = _lookup_most_expensive_service(breed)
        if top_svc and top_price:
            ctx += (f"\n\nСАМАЯ ДОРОГАЯ УСЛУГА ДЛЯ ЭТОЙ ПОРОДЫ: {top_svc} — от {top_price}€"
                    " (предлагай именно её, если клиент не знает что нужно)")
    return ctx

_SVC_KW = {
    'базовый':       'Базовый уход',
    'гигиенический': 'Гигиенический уход',
    'комплексный':   'Комплексный уход',
    'spa':           'SPA-уход',
    'тримминг':      'Тримминг',
    'экспресс':      'Экспресс-линька',
    'линька':        'Экспресс-линька',
    'вычес':         'Вычес',
}

# Order for listing services simple → complex
_SVC_ORDER = ['Базовый уход', 'Гигиенический уход', 'Комплексный уход',
              'Тримминг', 'Вычес', 'Экспресс-линька', 'SPA-уход']

_SVC_DESCRIPTIONS = {
    'Базовый уход':       'мытьё профессиональными средствами + сушка',
    'Гигиенический уход': 'купание, сушка, стрижка когтей, чистка ушей, глаз и лапок',
    'Комплексный уход':   'всё из гигиенического + модельная стрижка ✨',
    'SPA-уход':           'глубокое питание и восстановление шерсти',
    'Тримминг':           'выщипывание старого слоя + оформление силуэта',
    'Экспресс-линька':    'мытьё, маска, сушка — эффективное удаление подшёрстка',
    'Вычес':              'вычёсывание + стрижка когтей + уход за глазами и ушами',
}

import random as _random

_BREED_COMPLIMENTS = [
    "{breed} — отличная порода! 🤍",
    "О, {breed}! Такие красивые собаки 🐾",
    "{breed} — просто замечательная порода! 🤍",
    "Как здорово, {breed}! Одна из моих любимых пород 🐾",
    "{breed} — такие умные и красивые! 🤍",
]

def _breed_compliment(breed):
    name = (breed or '').strip()
    if not name:
        return "Отличный питомец! 🤍"
    name = name[0].upper() + name[1:]
    return _random.choice(_BREED_COMPLIMENTS).format(breed=name)

def _lookup_price(breed, service):
    """Look up price from WA_SYSTEM_PROMPT price list by breed + service keywords."""
    if not breed or not service or not WA_SYSTEM_PROMPT:
        return 0
    service_lower = service.lower()
    canonical = next((name for kw, name in _SVC_KW.items() if kw in service_lower), None)
    if not canonical:
        return 0
    # Breed words longer than 3 chars, skip digits/weight tokens
    breed_words = [w.strip(',.()кгКГ') for w in breed.lower().split()
                   if len(w) > 3 and not re.match(r'^\d', w)]
    if not breed_words:
        return 0
    best_line, best_score = "", 0
    for line in WA_SYSTEM_PROMPT.split('\n'):
        if ':' not in line:
            continue
        ll = line.lower()
        score = sum(1 for w in breed_words if w in ll)
        if score > best_score:
            best_score, best_line = score, line
    if not best_line or best_score == 0:
        return 0
    m = re.search(rf'{re.escape(canonical)}\s+(\d+)', best_line)
    return int(m.group(1)) if m else 0

def _lookup_most_expensive_service(breed):
    """Return (canonical_service_name, price) with the highest price for breed."""
    if not breed or not WA_SYSTEM_PROMPT:
        return None, 0
    breed_words = [w.strip(',.()кгКГ') for w in breed.lower().split()
                   if len(w) > 3 and not re.match(r'^\d', w)]
    if not breed_words:
        return None, 0
    best_line, best_score = "", 0
    for line in WA_SYSTEM_PROMPT.split('\n'):
        if ':' not in line:
            continue
        ll = line.lower()
        score = sum(1 for w in breed_words if w in ll)
        if score > best_score:
            best_score, best_line = score, line
    if not best_line or best_score == 0:
        return None, 0
    best_service, best_price = None, 0
    for canonical in set(_SVC_KW.values()):
        m = re.search(rf'{re.escape(canonical)}\s+(\d+)', best_line)
        if m:
            p = int(m.group(1))
            if p > best_price:
                best_price, best_service = p, canonical
    return best_service, best_price

def _get_all_services_for_breed(breed):
    """Return ordered list of (service, price) for breed, simple→complex."""
    if not breed or not WA_SYSTEM_PROMPT:
        return []
    breed_words = [w.strip(',.()кгКГ') for w in breed.lower().split()
                   if len(w) > 3 and not re.match(r'^\d', w)]
    if not breed_words:
        return []
    best_line, best_score = "", 0
    for line in WA_SYSTEM_PROMPT.split('\n'):
        if ':' not in line:
            continue
        ll = line.lower()
        score = sum(1 for w in breed_words if w in ll)
        if score > best_score:
            best_score, best_line = score, line
    if not best_line or best_score == 0:
        return []
    result = []
    for canonical in _SVC_ORDER:
        m = re.search(rf'{re.escape(canonical)}\s+(\d+)', best_line)
        if m:
            result.append((canonical, int(m.group(1))))
    return result

# Breeds that require subtype clarification before pricing/services.
# 'coat'+'size' means BOTH must be present; 'any' means at least one must be present.
_BREED_CLARIFICATION = [
    {
        'base': ['такса'],
        'coat': ['гладкошерстн', 'длинношерстн', 'жёсткошерстн', 'жесткошерстн'],
        'size': ['стандартн', 'миниатюрн', 'кролич', 'карликов'],
        'question': (
            "Уточните, пожалуйста, подвид таксы:\n"
            "• Тип шерсти: гладкошерстная / длинношерстная / жёсткошерстная\n"
            "• Размер: стандартная / миниатюрная / кроличья 🤍"
        ),
    },
    {
        'base': ['пудель'],
        'any': ['той', 'миниатюрн', 'средн', 'большой', 'королевск', 'карликов'],
        'question': "Уточните, пожалуйста, размер пуделя: той / миниатюрный / средний / большой (королевский)? 🤍",
    },
    {
        'base': ['шпиц'],
        'any': ['немецк', 'японск', 'той', 'миниатюрн', 'средн', 'большой'],
        'question': (
            "Уточните, пожалуйста, подвид шпица:\n"
            "• Тип: немецкий / японский\n"
            "• Размер: той / миниатюрный / средний / большой 🤍"
        ),
    },
    {
        'base': ['чихуахуа', 'чихуа'],
        'any': ['гладкошерстн', 'длинношерстн'],
        'question': "Уточните, пожалуйста: чихуахуа гладкошерстная или длинношерстная? 🤍",
    },
    {
        'base': ['йоркширский', 'йорк'],
        'any': ['стандартн', 'мини', 'той'],
        'question': "Уточните, пожалуйста: йоркширский терьер стандартный или мини? 🤍",
    },
]

def _check_breed_needs_clarification(breed):
    """Return clarification question if breed requires subtype info, else None."""
    if not breed:
        return None
    bl = breed.lower()
    for entry in _BREED_CLARIFICATION:
        if not any(b in bl for b in entry['base']):
            continue
        if 'coat' in entry and 'size' in entry:
            has_coat = any(c in bl for c in entry['coat'])
            has_size = any(s in bl for s in entry['size'])
            if not (has_coat and has_size):
                return entry['question']
        elif 'any' in entry:
            if not any(a in bl for a in entry['any']):
                return entry['question']
    return None

def _extract_price_from_history(history):
    """Find last 'от N€' price mentioned by Anna in conversation."""
    price_re = re.compile(r'от\s+(\d+)\s*€', re.IGNORECASE)
    for msg in reversed(history):
        if msg.get('role') == 'assistant':
            m = price_re.search(msg.get('content', ''))
            if m:
                return int(m.group(1))
    return 0

def _extract_state(history, state):
    recent = "\n".join(
        ("Клиент" if m["role"] == "user" else "Анна") + ": " + m["content"]
        for m in history[-8:]
    )
    prompt = (
        "Извлеки данные бронирования. Верни ТОЛЬКО JSON без пояснений.\n"
        "Поля: breed (название породы включая подвид если указан — например "
        "'такса гладкошерстная стандартная', 'пудель той', 'шпиц немецкий миниатюрный'; без веса), "
        "weight (вес питомца — например '30 кг'; null если не назван), "
        "service, date, time, ownerName, petName, "
        "master (null если клиент не называл мастера), "
        "clientPhone (null если клиент не называл другой номер), "
        "service_confirmed (boolean — true если клиент явно согласился с рекомендованной услугой: "
        "«да», «подходит», «хорошо», «давайте», «хочу», «окей» — это ДО вопроса о записи).\n"
        "booking_intent (boolean — true если клиент ответил «да»/«хочу»/«записывайте»/«yes» "
        "именно на вопрос «Хотите записаться?» — НЕ путать с подтверждением услуги).\n"
        "confirmed (boolean — true только если клиент явно подтвердил итоговую карточку записи: "
        "\"да\", \"подтверждаю\", \"всё верно\", \"yes\", \"jah\").\n"
        "ВАЖНО: если поле уже есть в «Текущие» и в диалоге не изменилось — "
        "оставь его значение, не возвращай null.\n"
        f"Текущие: {json.dumps(state, ensure_ascii=False)}\nДиалог:\n{recent}"
    )
    try:
        r = client_ai.messages.create(
            model="claude-haiku-4-5-20251001", max_tokens=300,
            messages=[{"role": "user", "content": prompt}]
        )
        text = r.content[0].text.strip()
        print(f"[extract_state] raw: {text[:400]}", flush=True)
        m = re.search(r"\{[\s\S]*\}", text)
        if m:
            extracted = json.loads(m.group())
            # Merge: only overwrite with non-null values so existing fields aren't cleared
            merged = dict(state)
            for k, v in extracted.items():
                if v is not None:
                    merged[k] = v
            return merged
        else:
            print(f"[extract_state] no JSON found in response", flush=True)
    except Exception as e:
        print(f"[extract_state] error: {e}", flush=True)
    return state

# ── Availability helpers ──────────────────────────────────────────────────────
_RU_MONTHS = {
    'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04',
    'мая': '05', 'июня': '06', 'июля': '07', 'августа': '08',
    'сентября': '09', 'октября': '10', 'ноября': '11', 'декабря': '12',
}

def _parse_date_to_iso(date_str):
    if not date_str:
        return None
    if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
        return date_str
    m = re.match(r'^(\d{1,2})[./](\d{1,2})[./](\d{4})$', date_str)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    dl = date_str.lower()
    for ru, num in _RU_MONTHS.items():
        if ru in dl:
            day_m = re.search(r'(\d{1,2})', dl)
            year_m = re.search(r'(\d{4})', dl)
            day = day_m.group(1).zfill(2) if day_m else '01'
            year = int(year_m.group(1)) if year_m else datetime.now().year
            month_int = int(num)
            now = datetime.now()
            if year == now.year and month_int < now.month:
                year += 1
            return f"{year}-{num}-{day}"
    return None

def _to_booking_date(date_str):
    """Convert any date string to DD.MM.YYYY (format Google Script expects)."""
    iso = _parse_date_to_iso(date_str)
    if iso and re.match(r'^\d{4}-\d{2}-\d{2}$', iso):
        y, mo, d = iso.split('-')
        return f"{d}.{mo}.{y}"
    return date_str  # fallback: send as-is, log will show it

# TTL cache for Google Script availability data (avoids hammering GS on every message)
_avail_cache: dict = {}
_AVAIL_TTL = 300  # 5 minutes

def _cache_get(key):
    entry = _avail_cache.get(key)
    if entry and (_time.time() - entry[0]) < _AVAIL_TTL:
        return entry[1]
    return None

def _cache_set(key, value):
    _avail_cache[key] = (_time.time(), value)

def _gs_url(params: dict) -> str:
    """Build full GS URL with params for logging."""
    import urllib.parse
    return GOOGLE_SCRIPT + "?" + urllib.parse.urlencode(params)

def _gs_cancel(phone):
    """Cancel booking by phone via GAS action=cancel."""
    if not GOOGLE_SCRIPT:
        return {"success": False, "error": "GOOGLE_SCRIPT not set"}
    try:
        r = requests.get(GOOGLE_SCRIPT, params={"action": "cancel", "phone": phone}, timeout=15)
        return r.json()
    except Exception as e:
        print(f"[gs_cancel] {e}", flush=True)
        return {"success": False, "error": str(e)}

def _gs_reschedule(phone, new_date, new_time):
    """Reschedule booking by phone via GAS action=reschedule."""
    if not GOOGLE_SCRIPT:
        return {"success": False, "error": "GOOGLE_SCRIPT not set"}
    try:
        r = requests.get(GOOGLE_SCRIPT, params={
            "action": "reschedule", "phone": phone,
            "newDate": new_date, "newTime": new_time,
        }, timeout=15)
        return r.json()
    except Exception as e:
        print(f"[gs_reschedule] {e}", flush=True)
        return {"success": False, "error": str(e)}

# Marker: [[ACTION:cancel:+372XXXXXXX]] or [[ACTION:reschedule:+372XX…:DD.MM.YYYY:HH:MM]]
_ACTION_RE = re.compile(r'\[\[ACTION:(cancel|reschedule):([^\]]+)\]\]')

def _process_action_markers(reply):
    """Execute GAS actions embedded in a Claude reply and strip the markers."""
    m = _ACTION_RE.search(reply)
    if not m:
        return reply
    action, args_str = m.group(1), m.group(2)
    args = args_str.split(':')
    if action == 'cancel' and args:
        res = _gs_cancel(args[0])
        print(f"[action-marker] cancel {args[0]!r} → {res}", flush=True)
    elif action == 'reschedule' and len(args) >= 3:
        # args = [phone, DD, MM, YYYY, HH, MM] after split on ':'
        # Reassemble date=DD.MM.YYYY and time=HH:MM
        phone    = args[0]
        new_date = args[1] if '.' in args[1] else '.'.join(args[1:4])
        new_time = args[-2] + ':' + args[-1] if len(args) >= 3 else ''
        # simpler: rejoin from index 1, split last two as time
        parts = args_str.split(':')
        phone     = parts[0]
        new_time  = parts[-2] + ':' + parts[-1]
        new_date  = ':'.join(parts[1:-2])   # DD.MM.YYYY — no colons inside
        res = _gs_reschedule(phone, new_date, new_time)
        print(f"[action-marker] reschedule {phone!r} → {new_date} {new_time} → {res}", flush=True)
    return _ACTION_RE.sub('', reply).strip()

_MASTERS = ["татьяна", "алиса", "кристина", "анна", "александра", "ксения"]

def _fetch_slots_for_date(date_iso, master=""):
    """Fetch slots from GS for date_iso (YYYY-MM-DD) and a specific master (lowercase).
    GS receives date as DD.MM.YYYY — same format as booking form."""
    cache_key = f"slots:{date_iso}:{master}"
    cached = _cache_get(cache_key)
    if cached is not None:
        print(f"[slots] cache hit {date_iso} master={master!r}: {cached}", flush=True)
        return cached
    if not GOOGLE_SCRIPT:
        print("[slots] GOOGLE_SCRIPT env var not set", flush=True)
        return []
    date_gs = _to_booking_date(date_iso)
    action = "slots"
    params = {"action": action, "date": date_gs, "master": master}
    print(f"[slots] action={action!r} master={master!r} GET {_gs_url(params)}", flush=True)
    try:
        r = requests.get(GOOGLE_SCRIPT, params=params, timeout=20)
        print(f"[slots] → {r.status_code}: {r.text[:300]}", flush=True)
        slots = [str(s) for s in r.json().get("slots", [])]
        _cache_set(cache_key, slots)
        return slots
    except Exception as e:
        print(f"[slots] error {date_iso} master={master!r}: {e}", flush=True)
        return []

def _fetch_available_days(master=""):
    """Fetch available days from GS for a specific master (lowercase)."""
    now = datetime.now()
    cache_key = f"days:{now.year}-{now.month}:{master}"
    cached = _cache_get(cache_key)
    if cached is not None:
        print(f"[days] cache hit master={master!r}: {cached[:5]}", flush=True)
        return cached
    if not GOOGLE_SCRIPT:
        print("[days] GOOGLE_SCRIPT env var not set", flush=True)
        return []
    action = "available_days"
    params = {"action": action, "month": now.month, "year": now.year, "master": master}
    print(f"[days] action={action!r} master={master!r} GET {_gs_url(params)}", flush=True)
    try:
        r = requests.get(GOOGLE_SCRIPT, params=params, timeout=20)
        print(f"[days] → {r.status_code}: {r.text[:300]}", flush=True)
        days = [str(d) for d in r.json().get("available", [])]
        _cache_set(cache_key, days)
        return days
    except Exception as e:
        print(f"[days] error master={master!r}: {e}", flush=True)
        return []

_NUM_TO_RU = ['', 'января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
              'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']

def _iso_to_ru_date(iso):
    """'2026-06-03' → '3 июня'"""
    try:
        _, m, d = iso.split('-')
        return f"{int(d)} {_NUM_TO_RU[int(m)]}"
    except Exception:
        return iso

# Keywords that indicate the client is asking about availability/schedule
_SCHEDULE_KW = re.compile(
    r'когда|свободн|занят|место[сть]?|дни|дат[ыею]|расписани|'
    r'слот|ближайш|доступн|прийти|записа|выбрать|в какое|какие',
    re.IGNORECASE
)

def _build_schedule_for_days():
    """Synchronous: Татьяна + Алиса, remaining days of current month, timeout=20s per GS request.
    Returns formatted string with dates and time slots."""
    _SCHED_MASTERS = ["татьяна", "алиса"]
    now = datetime.now()
    last_day = 31 if now.month == 12 else (datetime(now.year, now.month + 1, 1) - timedelta(days=1)).day
    month_remaining = {
        f"{day:02d}.{now.month:02d}.{now.year}"
        for day in range(now.day, last_day + 1)
    }

    print(f"[schedule] step 1 — available_days for {_SCHED_MASTERS}, {now.day:02d}.{now.month:02d}–{last_day:02d}.{now.month:02d}.{now.year}", flush=True)
    avail_days: set = set()
    for master in _SCHED_MASTERS:
        days = _fetch_available_days(master=master)
        matched = [d for d in days if d in month_remaining]
        print(f"[schedule] {master}: {matched}", flush=True)
        avail_days.update(matched)

    if not avail_days:
        print("[schedule] no available days", flush=True)
        return None

    print(f"[schedule] step 2 — slots for {len(avail_days)} days × {len(_SCHED_MASTERS)} masters", flush=True)
    day_slots: dict = {}
    sorted_dates = sorted(avail_days, key=lambda d: _parse_date_to_iso(d) or d)
    for day_str in sorted_dates:
        iso = _parse_date_to_iso(day_str)
        if not iso:
            continue
        for master in _SCHED_MASTERS:
            slots = _fetch_slots_for_date(iso, master=master)
            if slots:
                day_slots.setdefault(day_str, set()).update(slots)

    if not day_slots:
        print("[schedule] no slots found", flush=True)
        return None

    lines = [
        f"• {_iso_to_ru_date(_parse_date_to_iso(d))}: {', '.join(s for s in sorted(day_slots[d]) if s.endswith(':00'))}"
        for d in sorted_dates if d in day_slots
        and any(s.endswith(':00') for s in day_slots[d])
    ]
    result = "Свободные места:\n" + "\n".join(lines)
    print(f"[schedule] ready:\n{result}", flush=True)
    return result

def _fetch_schedule_bg(sid):
    """Background thread: fetch schedule and store result in session avail dict."""
    print(f"[schedule-bg] starting for sid={sid[:8]}", flush=True)
    try:
        schedule = _build_schedule_for_days()
    except Exception as e:
        print(f"[schedule-bg] error: {e}", flush=True)
        schedule = None
    if sid in _chat_sessions:
        avail = _chat_sessions[sid].get("avail", {})
        avail["full_schedule"] = schedule   # None if failed/empty
        avail["schedule_loading"] = False   # signal completion (set AFTER data)
        print(f"[schedule-bg] done: {'ok — ' + schedule[:60] if schedule else 'empty'}", flush=True)
    else:
        print(f"[schedule-bg] session {sid[:8]} gone before fetch completed", flush=True)

def _fetch_full_schedule(max_days=5):
    """Fetch available days + slots across all masters, merge results."""
    # Collect available days from each master (union)
    all_days: set = set()
    for master in _MASTERS:
        for d in _fetch_available_days(master=master):
            all_days.add(d)

    if not all_days:
        print("[schedule] no available days across any master", flush=True)
        return None

    # Sort by ISO date so earliest days come first
    sorted_days = sorted(all_days, key=lambda d: _parse_date_to_iso(d) or d)
    print(f"[schedule] available days (all masters): {sorted_days[:max_days]}", flush=True)

    lines = []
    for day_str in sorted_days[:max_days]:
        iso = _parse_date_to_iso(day_str)
        if not iso:
            continue
        # Collect slots from each master for this date (union, sorted)
        all_slots: set = set()
        for master in _MASTERS:
            for s in _fetch_slots_for_date(iso, master=master):
                all_slots.add(s)
        if all_slots:
            sorted_slots = sorted(all_slots)
            lines.append(f"• {_iso_to_ru_date(iso)}: {', '.join(sorted_slots)}")

    if not lines:
        return None
    return "Свободные места:\n" + "\n".join(lines)

def _avail_context(avail):
    if not avail:
        return ""
    lines = []
    # Full schedule (days + slots) shown after client confirms booking intent
    if avail.get("full_schedule"):
        lines.append(avail["full_schedule"])
    elif avail.get("available_days"):
        days = avail["available_days"]
        lines.append(f"Ближайшие свободные дни: {', '.join(str(d) for d in days[:10])}.")
    # Slot check for a specific date (when client named one)
    if "slots" in avail:
        label = avail.get("date_label", "")
        slots = avail["slots"]
        if slots is None:
            pass  # fetch failed — say nothing
        elif not slots:
            lines.append(f"На {label} свободных слотов нет — предложи другой день.")
        else:
            slot_list = ", ".join(slots)
            req_t = avail.get("requested_time", "")
            if req_t and not any(str(s).strip()[:5] == req_t.strip()[:5] for s in slots):
                lines.append(f"Слот {req_t} на {label} ЗАНЯТ — не подтверждай его, предложи свободный.")
            lines.append(f"Свободные слоты на {label}: {slot_list}.")
    return ("\n\n📅 РАСПИСАНИЕ (актуально):\n" + "\n".join(lines)) if lines else ""

# ── Price disclaimer (code-level, breed-aware) ────────────────────────────────
_SMOOTH_COAT_KW = [
    'такса гладк', 'джек-рассел гладк', 'джек рассел гладк',
    'французский бульдог', 'чихуахуа гладк', 'доберман',
    'боксер', 'боксёр', 'далматин', 'бигль', 'мопс',
    'бульмастиф', 'немецкий дог', 'американский стаффордш', 'питбуль',
]
_PRICE_RE = re.compile(r'\d+\s*€')
_DISCLAIMER_SMOOTH = ("Окончательная стоимость будет озвучена мастером после осмотра "
                      "при приёмке — может варьироваться в зависимости от состояния шерсти 🤍")
_DISCLAIMER_OTHER  = ("Окончательная стоимость будет озвучена мастером после осмотра "
                      "при приёмке — может варьироваться в зависимости от состояния шерсти "
                      "и наличия колтунов 🤍")

def _is_smooth_coat(breed):
    if not breed:
        return False
    b = breed.lower()
    return any(kw in b for kw in _SMOOTH_COAT_KW)

def _add_price_disclaimer(reply, breed):
    if not _PRICE_RE.search(reply):
        return reply
    if "Окончательная стоимость" in reply:
        return reply
    disclaimer = _DISCLAIMER_SMOOTH if _is_smooth_coat(breed) else _DISCLAIMER_OTHER
    return reply + "\n\n" + disclaimer

def track_client(phone, channel, text):
    if phone not in clients:
        clients[phone] = {"channel": channel, "timestamps": [], "last_seen": None, "last_text": "", "mode": "jarvis"}
    now = datetime.now()
    clients[phone]["timestamps"].append(now)
    clients[phone]["last_seen"] = now
    clients[phone]["last_text"] = text
    _log_msg(phone, channel, "in", text)
    _channel_counts[channel] = _channel_counts.get(channel, 0) + 1

def should_reply(phone, channel="whatsapp"):
    if manual_mode: return False
    if pause_until and datetime.now() < pause_until: return False
    if not jarvis_enabled: return False
    if channel == "instagram" and not instagram_enabled: return False
    if clients.get(phone, {}).get("mode") == "manual": return False
    return True

def get_ai_reply(sender_id, text):
    if sender_id not in conversation_history:
        conversation_history[sender_id] = []
    history = conversation_history[sender_id]
    history.append({"role": "user", "content": text})
    if len(history) > MAX_HISTORY:
        history = history[-MAX_HISTORY:]
        conversation_history[sender_id] = history
    response = client_ai.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        system=SYSTEM_PROMPT,
        messages=history
    )
    reply = _process_action_markers(response.content[0].text)
    history.append({"role": "assistant", "content": reply})
    return reply

def send_whatsapp(to, text):
    url = "https://graph.facebook.com/v18.0/" + (WHATSAPP_PHONE_ID or "") + "/messages"
    headers = {"Authorization": "Bearer " + (WHATSAPP_TOKEN or ""), "Content-Type": "application/json"}
    data = {"messaging_product": "whatsapp", "to": to, "type": "text", "text": {"body": text}}
    requests.post(url, headers=headers, json=data)

INSTAGRAM_ACCOUNT_ID = "17841479914115449"

def send_instagram(to, text):
    url = "https://graph.facebook.com/v18.0/" + INSTAGRAM_ACCOUNT_ID + "/messages"
    headers = {"Authorization": "Bearer " + (INSTAGRAM_TOKEN or ""), "Content-Type": "application/json"}
    data = {"recipient": {"id": to}, "message": {"text": text}}
    requests.post(url, headers=headers, json=data)

def handle_message(sender_id, text, channel):
    track_client(sender_id, channel, text)
    if not should_reply(sender_id, channel):
        return None
    reply = get_ai_reply(sender_id, text)
    _log_msg(sender_id, channel, "out", reply)
    if channel == "instagram":
        send_instagram(sender_id, reply)
    else:
        send_whatsapp(sender_id, reply)
    return reply
# ── EMAIL CONFIRMATION ──────────────────────
@app.route("/confirm")
def confirm():
    import requests as req_lib
    import urllib.parse

    booking_id = request.args.get("id", "")

    if booking_id:
        try:
            r = req_lib.get(GOOGLE_SCRIPT, params={"action": "get", "id": booking_id}, timeout=10, allow_redirects=True)
            print(f"DEBUG confirm id={booking_id} status={r.status_code} body={r.text[:200]}", flush=True)
            data = r.json()
            email   = data.get("email", "")
            name    = data.get("name", "")
            date    = data.get("date", "")
            time    = data.get("time", "")
            service = data.get("service", "")
            master  = data.get("master", "")
            breed   = data.get("breed", "")
            pet     = data.get("pet", "")
            phone   = data.get("phone", "")
            lang    = data.get("lang", "ru")
        except Exception as e:
            return f"Ошибка получения данных: {str(e)}", 500
    else:
        email   = urllib.parse.unquote(request.args.get("email", ""))
        name    = request.args.get("name", "")
        date    = request.args.get("date", "")
        time    = request.args.get("time", "")
        service = request.args.get("service", "")
        master  = request.args.get("master", "")
        breed   = request.args.get("breed", "")
        pet     = request.args.get("pet", "")
        phone   = request.args.get("phone", "")
        lang    = request.args.get("lang", "ru")

    if lang not in ("ru", "en", "et"):
        lang = "ru"

    # ── Email (optional — only sent when email address is provided) ────────────
    email_result = "Email не указан — пропущено"
    if email:
        resend_api_key = os.environ.get("RESEND_API_KEY")
        if not resend_api_key:
            email_result = "RESEND_API_KEY не настроен — email пропущен"
        else:
            td = "padding:8px;border-bottom:1px solid #eee"
            tl = f"{td};color:#666"
            if lang == "en":
                heading  = f"Thank you for booking, {name}!"
                _ms      = f" Your groomer: <b>{master}</b>." if master else ""
                subhead  = f"Your appointment at R&amp;J Grooming is confirmed.{_ms}"
                labels   = ["Date", "Time", "Groomer", "Service", "Breed", "Pet's name"]
                footer_t = "We look forward to seeing you and your pet!"
                address  = "Address: Allveelaeva 4, Tallinn<br>Phone: +372 587 35456"
            elif lang == "et":
                heading  = f"Aitäh broneeringu eest, {name}!"
                _ms      = f" Teie meister: <b>{master}</b>." if master else ""
                subhead  = f"Teie broneering R&amp;J Groomingus on kinnitatud.{_ms}"
                labels   = ["Kuupäev", "Kellaaeg", "Meister", "Teenus", "Tõug", "Lemmiklooma nimi"]
                footer_t = "Ootame teid ja teie lemmikut!"
                address  = "Aadress: Allveelaeva 4, Tallinn<br>Telefon: +372 587 35456"
            else:
                heading  = f"Спасибо за запись, {name}!"
                _ms      = f" Ваш мастер: <b>{master}</b>." if master else ""
                subhead  = f"Ваша запись в R&amp;J Grooming подтверждена.{_ms}"
                labels   = ["Дата", "Время", "Мастер", "Услуга", "Порода", "Кличка"]
                footer_t = "Ждём вас и вашего питомца!"
                address  = "Адрес: Allveelaeva 4, Tallinn<br>Телефон: +372 587 35456"

            values = [date, time, master, service, breed, pet]
            rows = "".join(
                f'<tr><td style="{tl}">{l}</td>'
                f'<td style="{td}"><b>{v}</b></td></tr>'
                for l, v in zip(labels, values) if v  # skip rows with empty values
            )
            body_html = (
                f'<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;color:#222;'
                f'max-width:600px;margin:0 auto;padding:20px;">'
                f'<h2 style="color:#c9a84c;font-family:Georgia,serif;">{heading}</h2>'
                f'<p>{subhead}</p>'
                f'<table style="border-collapse:collapse;width:100%;margin:16px 0;">{rows}</table>'
                f'<p style="color:#666;font-size:14px;">{address}</p>'
                f'<p style="color:#c9a84c;font-style:italic;">{footer_t}</p>'
                f'</body></html>'
            )
            if lang == "en":
                subject = f"R&J Grooming booking confirmed - {date} at {time}"
            elif lang == "et":
                subject = f"R&J Grooming broneering kinnitatud - {date} kell {time}"
            else:
                subject = f"Запись в R&J Grooming подтверждена - {date} в {time}"

            try:
                r = req_lib.post(
                    "https://api.resend.com/emails",
                    headers={"Authorization": f"Bearer {resend_api_key}", "Content-Type": "application/json"},
                    json={
                        "from": "R&J Grooming <booking@rjgrooming.salon>",
                        "to": [email.strip()],
                        "subject": subject,
                        "html": body_html,
                    },
                    timeout=10
                )
                if r.status_code == 200:
                    email_result = f"Email отправлен на {email}"
                else:
                    email_result = f"Ошибка email: {r.status_code} {r.text}"
            except Exception as e:
                email_result = f"Ошибка email: {str(e)}"

    # Отправка SMS через Twilio
    sms_result = "SMS не отправлен"
    twilio_sid = os.environ.get("TWILIO_ACCOUNT_SID")
    twilio_token = os.environ.get("TWILIO_AUTH_TOKEN")
    twilio_phone = os.environ.get("TWILIO_PHONE", "+37266922128")
    if phone and twilio_sid and twilio_token:
        try:
            # Форматируем номер (убираем двойное кодирование %2B от Safari)
            import urllib.parse as _ul
            p = _ul.unquote(_ul.unquote(phone)).strip().replace(" ", "").replace("-", "")
            if not p.startswith("+"):
                p = "+" + p
            print(f"PHONE RAW: {phone!r} → DECODED: {p!r}", flush=True)
            if lang == "en":
                _m = f"Groomer: {master}. " if master else ""
                sms_body = f"R&J Grooming: booking confirmed! {_m}{date} at {time}. Address: Allveelaeva 4, Tallinn"
            elif lang == "et":
                _m = f"Meister: {master}. " if master else ""
                sms_body = f"R&J Grooming: broneering kinnitatud! {_m}{date} kell {time}. Aadress: Allveelaeva 4, Tallinn"
            else:
                _m = f"Мастер: {master}. " if master else ""
                sms_body = f"R&J Grooming: запись подтверждена! {_m}Дата: {date} в {time}. Адрес: Allveelaeva 4, Tallinn"
            sms_url = f"https://api.twilio.com/2010-04-01/Accounts/{twilio_sid}/Messages.json"
            sms_resp = req_lib.post(
                sms_url,
                auth=(twilio_sid, twilio_token),
                data={"From": twilio_phone, "To": p, "Body": sms_body},
                timeout=10
            )
            if sms_resp.status_code == 201:
                sms_result = f"SMS отправлен на {p}"
            else:
                sms_result = f"Ошибка SMS: {sms_resp.status_code} {sms_resp.text}"
        except Exception as e:
            sms_result = f"Ошибка SMS: {str(e)}"

    return f"{email_result}<br>{sms_result}", 200

# ── BOOKING → GOOGLE CALENDAR ──────────────────────────────────────────────
GOOGLE_SCRIPT = os.environ.get("GOOGLE_SCRIPT", "")

# ── Кеш полной истории календаря (общий для всех страниц админки) ───────
_full_history_cache = {"data": None, "fetched_at": 0, "to_date": None}
_full_history_cache_lock = threading.Lock()
_FULL_HISTORY_CACHE_TTL = 180  # секунд

def _fetch_full_history_bookings():
    """Общая история броней с 01.01.2020 по (сегодня+180дн), с кешем на
    несколько минут — избавляет от повторного тяжёлого запроса к календарю
    при каждом открытии Клиенты/Напоминания/Поиск/Экспорт."""
    today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
    to_date = (today + timedelta(days=180)).strftime("%d.%m.%Y")
    now_ts = _time.time()

    with _full_history_cache_lock:
        if (_full_history_cache["data"] is not None
                and _full_history_cache["to_date"] == to_date
                and now_ts - _full_history_cache["fetched_at"] < _FULL_HISTORY_CACHE_TTL):
            return _full_history_cache["data"]

    params = {"action": "stats", "from": "01.01.2020", "to": to_date}
    r = requests.get(GOOGLE_SCRIPT, params=params, timeout=30)
    data = r.json()
    bookings = data.get("bookings", []) if isinstance(data, dict) else []

    with _full_history_cache_lock:
        _full_history_cache["data"] = bookings
        _full_history_cache["fetched_at"] = now_ts
        _full_history_cache["to_date"] = to_date

    return bookings

# ── 35-ДНЕВНОЕ НАПОМИНАНИЕ (первый визит → напоминание админу в Telegram) ──
try:
    from zoneinfo import ZoneInfo
    _REMINDER_TZ = ZoneInfo("Europe/Tallinn")
except Exception:
    _REMINDER_TZ = None

def _send_reminder_telegram(text):
    tg_token = os.environ.get("TELEGRAM_BOT_TOKEN")
    tg_chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not (tg_token and tg_chat_id):
        msg = "TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID не заданы"
        print(f"REMINDER: {msg}", flush=True)
        return {"ok": False, "error": msg}
    # Telegram режет сообщения длиннее 4096 символов — бьём по границам "\n\n"
    chunks = []
    remaining = text
    LIMIT = 3800
    while len(remaining) > LIMIT:
        cut = remaining.rfind("\n\n", 0, LIMIT)
        if cut <= 0:
            cut = LIMIT
        chunks.append(remaining[:cut])
        remaining = remaining[cut:].lstrip("\n")
    if remaining:
        chunks.append(remaining)

    results = []
    for i, chunk in enumerate(chunks):
        try:
            resp = requests.post(
                f"https://api.telegram.org/bot{tg_token}/sendMessage",
                json={"chat_id": tg_chat_id, "text": chunk, "parse_mode": "HTML"},
                timeout=10
            )
            body = resp.json()
            results.append({"chunk": i + 1, "http_status": resp.status_code, "telegram_response": body})
            if not body.get("ok"):
                print(f"REMINDER TELEGRAM API ERROR (chunk {i+1}/{len(chunks)}): {body}", flush=True)
            if len(chunks) > 1:
                _time.sleep(0.5)
        except Exception as e:
            results.append({"chunk": i + 1, "error": str(e)})
            print(f"REMINDER TELEGRAM ERROR (chunk {i+1}/{len(chunks)}): {e}", flush=True)
    return {"ok": all(r.get("telegram_response", {}).get("ok") for r in results if "telegram_response" in r), "chunks": len(chunks), "results": results}

def check_35day_reminders(target_date=None, send_telegram=True):
    """Два напоминания на клиента после его последнего визита:
    #1 — на 35-й день, #2 — на 42-й день (35+7), если он не записался повторно.
    Если между визитом и сегодня есть ЛЮБАЯ более поздняя запись (прошлая или
    будущая) — значит клиент уже перезаписался, и напоминания не шлём.
    target_date позволяет посчитать на любую дату (например, на завтра) без
    реальной отправки — для этого передать send_telegram=False."""
    try:
        today = target_date or (datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date())
        from_date = "01.01.2020"
        to_date = (today + timedelta(days=120)).strftime("%d.%m.%Y")
        r = requests.get(GOOGLE_SCRIPT, params={"action": "stats", "from": from_date, "to": to_date}, timeout=30)
        data = r.json()
        bookings = data.get("bookings", []) if isinstance(data, dict) else []

        by_phone = {}
        for b in bookings:
            phone = (b.get("clientPhone") or "").strip()
            if not phone:
                continue
            try:
                d = datetime.strptime(b.get("date", ""), "%d.%m.%Y").date()
            except Exception:
                continue
            by_phone.setdefault(phone, []).append({
                "date": d,
                "name": b.get("clientName", ""),
                "pet": b.get("petName", ""),
                "master": b.get("master", ""),
                "breed": b.get("breed", "")
            })

        due = []
        for phone, visits in by_phone.items():
            past_visits = [v for v in visits if v["date"] <= today]
            if not past_visits:
                continue
            last_visit = max(past_visits, key=lambda v: v["date"])
            has_rebooked = any(v["date"] > last_visit["date"] for v in visits)
            if has_rebooked:
                continue
            days_since = (today - last_visit["date"]).days
            stage = {35: 1, 42: 2}.get(days_since)
            if stage:
                info = dict(last_visit)
                info["stage"] = stage
                info["days_since"] = days_since
                due.append((phone, info))

        if due and send_telegram:
            lines = ["🔔 <b>Напоминания клиентам</b>", ""]
            for phone, info in due:
                lines.append(
                    f"{'1️⃣' if info['stage']==1 else '2️⃣'} <b>Напоминание #{info['stage']}</b> ({info['days_since']} дн. без визита)\n"
                    f"👤 {info['name']} ({phone})\n"
                    f"🐾 {info['pet']} — {info['breed']}\n"
                    f"Последний визит: {info['date'].strftime('%d.%m.%Y')} у {info['master']}\n"
                )
            _send_reminder_telegram("\n".join(lines))

        print(f"REMINDER CHECK ({today}): найдено {len(due)} клиентов", flush=True)
        return due
    except Exception as e:
        print(f"REMINDER CHECK ERROR: {e}", flush=True)
        return []

def _reminder_scheduler_loop():
    while True:
        try:
            now = datetime.now(_REMINDER_TZ) if _REMINDER_TZ else datetime.utcnow()
            next_run = now.replace(hour=10, minute=0, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)
            sleep_seconds = (next_run - now).total_seconds()
            _time.sleep(max(sleep_seconds, 60))
            check_35day_reminders()
        except Exception as e:
            print(f"REMINDER SCHEDULER ERROR: {e}", flush=True)
            _time.sleep(3600)

threading.Thread(target=_reminder_scheduler_loop, daemon=True).start()

_REMINDER_STATUS_PUBLIC_ID = "rjgrooming/reminder_status_index.json"

def _reminder_status_url():
    return f"https://res.cloudinary.com/{CLOUDINARY_CLOUD_NAME}/raw/upload/{_REMINDER_STATUS_PUBLIC_ID}"

def _load_reminder_status():
    try:
        r = requests.get(_reminder_status_url(), timeout=8)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return {}

def _save_reminder_status(data):
    import hashlib as _hashlib
    timestamp = int(_time.time())
    params_to_sign = {"overwrite": "true", "public_id": _REMINDER_STATUS_PUBLIC_ID, "timestamp": timestamp}
    to_sign = "&".join(f"{k}={v}" for k, v in sorted(params_to_sign.items()))
    signature = _hashlib.sha1((to_sign + CLOUDINARY_API_SECRET).encode("utf-8")).hexdigest()
    payload = json.dumps(data, ensure_ascii=False)
    try:
        resp = requests.post(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/raw/upload",
            data={
                "api_key": CLOUDINARY_API_KEY,
                "timestamp": timestamp,
                "public_id": _REMINDER_STATUS_PUBLIC_ID,
                "overwrite": "true",
                "signature": signature,
            },
            files={"file": ("data.json", payload, "application/json")},
            timeout=15
        )
        if resp.status_code != 200:
            print(f"REMINDER STATUS SAVE ERROR: {resp.status_code} {resp.text[:200]}", flush=True)
        return resp.status_code == 200
    except Exception as e:
        print(f"REMINDER STATUS SAVE ERROR: {e}", flush=True)
        return False

# ── АБОНЕМЕНТЫ (цифровая карта посещений) ────────────────────────────────
_MEMBERSHIP_INDEX_PUBLIC_ID = "rjgrooming/membership_index.json"

def _membership_index_url():
    return f"https://res.cloudinary.com/{CLOUDINARY_CLOUD_NAME}/raw/upload/{_MEMBERSHIP_INDEX_PUBLIC_ID}"

def _load_memberships():
    import base64 as _b64_mod
    try:
        auth = _b64_mod.b64encode(f"{CLOUDINARY_API_KEY}:{CLOUDINARY_API_SECRET}".encode()).decode()
        meta_r = requests.get(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/resources/raw/upload/{_MEMBERSHIP_INDEX_PUBLIC_ID}",
            headers={"Authorization": f"Basic {auth}"},
            timeout=8
        )
        if meta_r.status_code == 200:
            version = meta_r.json().get("version")
            if version:
                versioned_url = f"https://res.cloudinary.com/{CLOUDINARY_CLOUD_NAME}/raw/upload/v{version}/{_MEMBERSHIP_INDEX_PUBLIC_ID}"
                r = requests.get(versioned_url, timeout=8)
                if r.status_code == 200:
                    return r.json()
    except Exception as e:
        print(f"MEMBERSHIP LOAD (versioned) ERROR: {e}", flush=True)

    # запасной вариант — обычный запрос без версии
    try:
        r = requests.get(_membership_index_url(), params={"_": int(_time.time() * 1000)}, timeout=8)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return {}

def _save_memberships(data):
    import hashlib as _hashlib
    timestamp = int(_time.time())
    params_to_sign = {"invalidate": "true", "overwrite": "true", "public_id": _MEMBERSHIP_INDEX_PUBLIC_ID, "timestamp": timestamp}
    to_sign = "&".join(f"{k}={v}" for k, v in sorted(params_to_sign.items()))
    signature = _hashlib.sha1((to_sign + CLOUDINARY_API_SECRET).encode("utf-8")).hexdigest()
    payload = json.dumps(data, ensure_ascii=False)
    try:
        resp = requests.post(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/raw/upload",
            data={
                "api_key": CLOUDINARY_API_KEY,
                "timestamp": timestamp,
                "public_id": _MEMBERSHIP_INDEX_PUBLIC_ID,
                "overwrite": "true",
                "invalidate": "true",
                "signature": signature,
            },
            files={"file": ("data.json", payload, "application/json")},
            timeout=15
        )
        if resp.status_code != 200:
            print(f"MEMBERSHIP SAVE ERROR: {resp.status_code} {resp.text[:200]}", flush=True)
        return resp.status_code == 200
    except Exception as e:
        print(f"MEMBERSHIP SAVE ERROR: {e}", flush=True)
        return False

WALLET_STRIP_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAMCAgMCAgMDAwMEAwMEBQgFBQQEBQoHBwYIDAoMDAsKCwsNDhIQDQ4RDgsLEBYQERMUFRUVDA8XGBYUGBIUFRT/2wBDAQMEBAUEBQkFBQkUDQsNFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBT/wAARCALACEADASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD8qqKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACjtRQKANyG3haNQyAnA5xUn2eEk/ux+VJGAkUfP8I/lQ7lTjoSOKAFFpb5xsBY+1WEsIyvywKfqBVnRNHTWtT0u0lmMMd1cLC8o/hBPWvurQ/+CaPhvWbG1uIviTboZY1cxs4yMjpSuOx8GLpwIz9nUj6ClGlqAc2oGenAr9Drf/gl14WkI/4uVFx1AlX/ABqRv+CWvh0ZK/EiMgDj96vWi4WPzsfSQhG63A9OBzTG01TjdAq+m0Cv0JuP+CYWgwNGH+I0fPcyr/jS3P8AwSw0khGg+ItuysCeZV/xouFj8+I7Hbwtqjf7wFH9mIBlrZST6AV+gUX/AASsspIQ5+IVtk5/5arUUP8AwS504YMvxEtwvtKtK6Cx8AtpoX/lgvPsKYlgidbdD+Ar9C1/4JfeGxjzPiRED/10X/GpR/wSz8OTfd+JELemZFppoLH54GwViMW68+wpPsSFh/ow/IV+h7f8Eq9GyAvxItyf+ui1EP8AglZpjnCfEOBiTj/WLRdBY/PV7FT922QDvwM0o01NuWgHtgV+hj/8EobVAP8Aiv4Bk9fMWnD/AIJT2GMD4iQGTONvmLQmgsfniul7P+XcEH1ApDZRht4gDeowOK/Qj/h1dZpJ+8+I9sBnBzKv+NOn/wCCVelhh5PxFtGyef3q0XCx+fIs4cHNuu0+wyKiawjyVEIKn2Ga/QuP/glRp+WDfEK3C5x/rVqX/h1To5baPiJb9cH94tFwsfnd9hiUbRCG+oFTCxSOPAtlcnvgV+guo/8ABKvSLV40i+ItqGY95FzVuL/glDp0jMB8QoSQBwJFouFj87309VXIt159hUb6crDIt1H0Ar9GZf8AgkzaRqzHx/GE9TItIv8AwSh007/+LhxEDHSRaLhY/OcaVg5EIOfYUrWG1Cv2VfckCv0Nn/4JPxln8rx9AQOQTIv+NNuP+CT32ayMzePIN3b94vP60XCx+eP9n8YNuA3bgYpz2JChWtFB/wBkCv0Ls/8AglVaTxFpPiDbhwOglXj9amm/4JTWMMRI+IkBbHeVf8aLjsz88FsFQ826kY9BxUjWEOAFgXI56Cv0Cj/4JXWAI3/EO34GTiVf8akk/wCCVukKcj4hwYxnJkX/ABouKx+fQsISPmgQH6Cj+y4x0gQ9+gr9CV/4JWaLs+f4iW4OMgmVf8ao/wDDrCwlnAj+I1oygZP71c/zouFj4G/s6AdYE56cClFikS7Taxtn1Ar9BZ/+CVeipDlfiRbCXbnBlXH86XTf+CVmhzWxe5+JNt5nYJKuKLhY/PU6eknIgUEDsBiiLT0WMn7MrE9OBX6GH/glPpLlfK+JFsMjvKv+NN/4dS2iuCPiNabf+uy/40XCx+eo0wKQwtVIHXgU9rBGxm0RT2wBX6Gn/glHaY+T4i2pz1/erj+dVh/wSng+Yv8AEO0CjpiZf8aLhY/PgadGXBNsuP8AZAofSUA+WFSx7YFffkP/AASzspJcf8LHssennD/Gib/glZCFHl/Eay3EngzD/Gi4j4GGnQnANugK84wOaVdLTdxbIWPbA4r9ALX/AIJS28kaM3xEtOePllX/ABq7bf8ABKXSwwE3xFhUlsArKv8AjTHY/PF9KjbGIFDA88CozZI+Q1qgx6AV+kd1/wAEl9GW0E0XxBA2/fZ5FwP1rJT/AIJP2k5zF8Q7aRM9RKv+NAWPzyGnRMP+PdcjrwKRrJAwYWy7Qe4Ffol/w6Wt23bPH0BPr5q1Um/4JNpF1+INsF95VpXCx+fS6asmf9HT8hSy2KYwLZd3rgV+hCf8EorQoMfEW1znnEq/41KP+CTtjt5+IkGc/wDPVf8AGi4WPzyXTomTH2dM/QU3+zYx/wAu6cewr9B4f+CT9ubllk+IlqIvVZVz/Op3/wCCUWlIx3fEe3wPWVf8aLhY/PP+zYW+YQJ+Qpv9nqG+W3TH0FforH/wSi8P7CW+JMCnPaRagm/4JYeGElK/8LPt1HHLSr/jRcGj88m02Jj80Cj6AUDToRnMKHHsK/RR/wDglP4Z2Fh8TIMDv5q0if8ABKzwmynPxOt89/3q0XCx+d62UStk28ZU9OBmlOlRhsrAp9sCv0Lb/gll4TWQhfihbBcc/vV/xok/4JceFIkJHxUtl47zL/jRcLH57f2Ikh/491UY9BSHRkOD5Cjb6gV+hP8Aw618OPGTH8UrU8f89lx/Ooj/AMEtfDwPzfFK0Ax/z1X/ABouFj8+jpcZyWt1DY4GBSLYxqMNbRk+oAr9BZ/+CXvhVEwPitabsd5lx/OoIP8Agl94YKsZvitZYA/hmX/GmFj4CGnx42/Z42z7ClOkxMV/cLkegFffq/8ABMLwkXAX4sWXToJl/wAasj/gl94T4x8VrQ8c/vl/xpXQWPz5fS4fM3GFCo/ugU0afHnH2aPnocCv0JX/AIJeeDSwB+KtrjHP71f8asD/AIJd+BxHkfFa3Bx2lX/GldBY/O3+y0JA+zoMewoGnp/BBGW9CBX6DJ/wTB8HM7A/FmzIA4AlX/Gq0/8AwTA8LhlEPxUsD/vTL/jTTQWPgYaZDKB/o6qw68ClOlJIRm2QAdNoFfoBB/wS/wDC5VS/xVsie+Jl/wAaef8AgmJ4QQjPxXs19f3y/wCNFwsfnydKVeBbo31AqZdKikXyzbKrdQ2Bivv7/h2V4KbhPi1Zk/8AXZf8ajm/4Ji+FY4ww+K1ltzyfOX/ABptpILHwT/ZMKhc28eAeTgVG+lxKxIt42B9hX3rZf8ABNnwLO5ST4uWZIPP75f8a0E/4JlfD043fFu1Bz2mX/Gp5kDR+fJ0uNSrC3Q885ApJNKgY5ECA+mBX6Fy/wDBMz4cRx7z8XrcDPJMy/41Vi/4JpfDmVsr8YLRhnAxKv8AjTTQWPz6OlIP+Xdd2eDgUPpSIP8AUJj3Ar9DW/4JnfDaBMzfF63U54JmX/Gox/wTa+FrttHxktC3p5y/40XCx+eJ0uMNnyRs9MDNONhHnBt029iAK/RKH/gmb8MpH+X4vWzD/rqv+NEn/BND4WLlf+Fw24fsPNWi4WPzvOnJECRboxPTIFPXToGG1oE3f7IFfoLcf8E0fhqvMfxgtfqZV/xpq/8ABNj4ZRxsZPjHZg+0y/40XCx+fqaWjEo1vHg9wBUraVCFKfZ0/wB7A4r78i/4Jr/DeYny/jHZMo/uzL/jVgf8Ezfh27kD4v2zewlX/GmFj89W0uNBg2yEeuBTW0mPO5YEwOcYFfoof+CYnw7RCW+LUAGP+eq/41JB/wAEx/hiUPm/FyEH/ZlX/GgLH50f2fGjF1t0YH+HAo+wQupzaor46YFfonJ/wTG+GYJK/F6Dpx+9X/Gobn/gmZ8NbUB5fi/boGHBaVeaAsfnr/ZsG3eLdCuMdB1qRNPihX/j1ickfxKK/Qlf+CZHw2ZMj4uwFcZBEq/407/h2n8LIsCb4wwKccjzV/xoEfnj/ZUJO42yD8BTjpMYyBbx4I9Bmv0ST/gml8KH6fGGErjtKv8AjTv+HZvwqABPxhiUdv3q/wCNAH50jR0VQUtoyO+QM0hsEV8yWkZ46IBX6Mn/AIJm/Cll+X4xRFv+uq/41DJ/wTH+GRx5fxfhP1lX/GgD86U0xDyLaNlPbAyKU6PFIPkgVSOvAr9D4/8AgmP8Od5x8XYCMdpV/wAagl/4JofDdG2p8YbXd6ecv+NK47H58DR0dMLBGvqSBSnSYyF22yE99wFff8n/AATU+HsZUn4w2ef9qZf8aY3/AATU8BBA4+MFls7kzL/jRcLHwINJQYY20e3vwKY+nRIQ32ZNvsBX3qP+Cc/w5BIb40afx289f8arTf8ABOPwDIf3Hxl0xh3zOv8AjRcGj4RjsIMEG2XcenAp7abFKojNqquehAHNfdlr/wAE1fB0rD/i7+mMn+zMv+NaK/8ABNDwFFgzfGCyX3Ey/wCNFxHwDNpCRPGr2yBc9QBzU39kxOQTbJsHoBX31J/wTT+H0iAx/GKyZs9TMv8AjUZ/4Jn+DVGV+L1htz189f8AGi47HwkPD6KA62qMD6gVE+hxpJk20ZB46DivvrSv+CbHge6Z1m+MFm6j/nlMuf506b/gmh4DUnHxgtQM95l/xphY+Av7HA4FpGy55OBS/wBgxANm3VWPTIFffKf8E0/A3UfGG0Iz2mX/ABqGb/gmZ4UdiY/i3ZMvbMy5/nSuFj4Kj0WL7r26KQfQc01tIjDFhax+y4HWvvhf+CZfhBeZvi7ZY/67Lx+tOX/gmp4Dwd/xis8/9dl/xpiPgF9KR0wbVA45bgdKVtLikT5rVADwpAGa+/pP+Cavw+A4+MdmG95l/wAarD/gmn4LkYiP4wWDDt++XP8AOgdj4GfTFiyDbKT0zgUf2IqswMCHI9BX3xP/AMEz/CMYP/F3dPz/ALU6/wCNVU/4JveDwxB+MWmE/wDXdf8AGgLHwk+lxoADbLvHPQU9NNiDEtaoCRxwMV94n/gmt4MIyPi/ppPf9+v+NSw/8E3vAI4n+Menj2WZf8aAsfBf2KGEbRaRsW7lRUR0aN5DuhjXI42gV9+y/wDBNn4dbGK/GKxHH/PZf8azZP8AgnN4BSbavxk0wDGf9ev+NK4WPhV9IiiO9IkYqOVIHNOWxiKfLaxnPUYHFfc7/wDBObwEzDb8ZtLBPX9+v+NSL/wTe8CjlfjJpnI5Pnrz+tFwsfCTaVCz7o4FzjkEDrQbBCAv2WPPU8Cvu6T/AIJveBQmR8Y9MHHOZ1/xrOk/4J2eCoWP/F5tJGP+m6/407oLHxCtjEnS1Vs9iBSR6ZAT89qoPYACvt8f8E8/AsgHlfGfSC3/AF3X/Go5/wDgnR4YUBk+MWjFSDy1wP8AGi6HZnxO2lxORut0HptAp40mIMN9qgB6YAr7Stv+CcegOoLfF/RWGDgi4H+NS23/AATu8KZCy/GLSCfQXA4/Wpugsz4sfSY4Vz9libPsKSHS40yz2kZVuBwOK+25v+Ccvg4gFPjLpQHfNwP8aqyf8E9fBkLhX+Mukbf+vgf40XQWZ8VrpNuDtNuh75wKkXRrcHc9uhT6CvtJv+Cd3hEIpj+MekFST/y8D/Gktv8Agnp4QkOJPjJpQOf+fgf40XQWZ8USaTbp/q4E+hApPsMMgH+iRrtPQAc19un/AIJ0eDJOnxl0nr3nX/Gnf8O5fBgQEfGXSs5/57r/AI07oVj4e+wwAtttlO7gAgcGkbTY/L2m2TeD1AFfbjf8E5fChIC/GPSDz1M6/wCNPH/BOLwpzn4yaOf+26/40roLHxJLo8bqpSBeRjoKRNFQja1uobPDADFfcSf8E4fCWP8Aksmk7c87Z1/xqKb/AIJ0+E4iQvxn0cHPT7Qv+NO4WPiL+z0Q7XtI9o6nAzS/YIkfetojr6ECvt5P+CcXhmVR/wAXh0gk+lwv+NX4P+CavhJky/xd0s89p1/xouFj4OOnR+b5qWsZ7bCBilXSkO4C2QyE8DAxX3fL/wAE1fCKnj4v6YB7zr/jUY/4JseFmPyfF/Sifadf8aYj4cWyhRRixjaQeqjFStpcU2C1pCgHXaor7oh/4JmeGnOT8WtOb6Tr/jU5/wCCZvhJF+f4tWCn3nX/ABoHY+EP7FtjIHjtkIH8JA5qb+wYWbzBbRrgdMDFfcjf8E1fCe7K/F3T/wDv+v8AjUp/4Jq+GnGE+LWnsP8Aruv+NK4WPhWXw1HONyW6K/pgYpsmgKqbRZxkgZzgV91j/gmfoY/1fxX0/A5/16/40g/4Jn6IPv8AxX08Z/6br/jTCx8GSaLESCtsq4HOQMVWk0yMybkto9oHIwK+9pP+CZ/hxSS3xa08DH/Pdf8AGiP/AIJm+Esgv8XNP2+06nP60BY+CGs4GBIt0VQOeBSx2FvKVxAi/UCvuvxj/wAEzdI0f4f674l0Lx7Frh0qBp3SFgRwM4NfDNtunU5JVlYr+XFANDRb2rSeW1vGCe4UcVT1+0ghs2McKpgjkDHcVrTW8bxIykqRxz61Q8QGVdKZGHAIyfxFAjk6KKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoHWijtQB0CKuyNj02CiWFujHkjipoFzHHu+7tH8qc8eevHHX0oC5ZsZDBFpzjJKXS+3Oa+1v2wvCsnw9+D3w58W6DqF1pt5f2qpOkbna3HU818YQw4tNPPpdL/ADr74/4KBAf8Mq/Cw5x/o6fyFKxTPi+Lx54zTHl+JZgNveU/41YX4jeNVUAeJ5/+/p4/WvPXYqynceg4pQ4DZOSfrTsK53svj7xlPtEvii5J7HzT/jT4viF4yjIX/hK7kDt+9P8AjXAM20jcSc9OafHgjB4J6c0rBc73/hZfjmJdi+K7nHP/AC1P+NV2+I3jmNRjxPPISe8p/wAa4xcHapJLE+tOZUHykHA6tmnZBc64eP8Axq/TxDK3r+9P+NPT4jeN4j8niGY/9tT/AI1xMiBFDJuA9j1qaIqY9/K4p2QXO6j+KPjldpPiGXg9fNb/ABqdfil443q6+J50YHtKf8a8+ZgjAkEKen1psibI943ZzyT2osgueiv8WfH8g58V3OzPeU/41FH8UfH9pJ5sfie4ds5+aU4/nXArIJFzkhf50oYLkOTtP8NGgXO8b4r+O2cyS67I5J/57H/Go5Pij45BaQeIJl5GCJW4/WuHR0HGT1pGbeSMnZ6etILnbH4r+P2P/IxTgZz/AK1v8aRPij49Rj/xUVx8xyT5rf41xecMATn8elOBJJAYkH9KVgudvL8S/HFyQ0viCZpEIKt5rcfrUyfFj4ggnZ4ouFY9SJT/AI1567MufnJFBYt90nb3osFz0mT4xfEuZSp8W3GwgA5lP+NV/wDhZ/xCQtjxdcjJ/wCep/xrz9HycEnIp5l2AlmJXvRYLnoKfFX4iMSB4xueRjHmn/GpJviz8R5Yyj+MLkqOMeacfzrzsEY3DJB6fWms52lCxPeiw7s7j/hYvj/kp4tuFzwf3p/xp3/CxfH7rtPjC4K9/wB6f8a4RwAhwT06ZoiCEZ5Jx09KLBdneL4/8bngeLZ8Y5JlP+NNuPHnjBCM+MboAjp5p/xril8sdQSo96Ro42JyhJxxRYE2do/jnxm2Nvi66YbeQZT/AI1Wh8YeMInMkfi67DEcgyn/ABrkInRcoFJHc5p6qpkB2kgCiwXZ2q+NPGchy3i24Y4wQZT/AI1NF4x8ZwqNniqbb3Xzj/jXEKiknapHHPNKQqldudvc55FFgudsPGvjrhl8VTD285v8aH8dePcc+LJx/wBtm/xriyqDgFgOwz1pkfzjDZx9abQXO3X4gfEJQNvi64x/12P+NC/EHx+2V/4S2fa3X98ef1riMBWCKSSfegIoTdznvzSsFzsYvFnjJTuTxTOrdyJm5/WpD4w8afKzeK7oJ3/et/jXGBAyqfmC+xpPNKttJLA9AKYXO5/4TrxxAgEHjC5APbzT/jULePvH8pX/AIq+5AB5IlP+Nck8aKgOSD9ajZwU2pkZPPPWi1xXO+b4n/ECS0a2Pje6eE9VMp/xqnb+PvH0TbIfGl1Gg/6an/GuIiUFwvKnPPPWgkJKyKTn0zRyhc9CT4m/ENMAeN7o4P8Az1P+NLP8Q/Htyv7zxncn6yn/ABrz7MZIIzkdeajfbjKEnnkZ6UNDTR3S+PvHMfTxhd/hKf8AGj/hYnjwkg+M7sA9vNP+NcO33Mgnn3pOB3JalYd0dqPiD40Qk/8ACbXwb/rof8aa3jvxZKcnxrfE9/3p/wAa4ohWO3HPfmnbY933elFhNo7I+OvFbja3jS+C+0h/xqCfxFrk20yeLr1n/wCuh/xrlCIlcfKdx6YNK7L90Ifc0WC52Q8W+KNjIPGF6yYHBlb/ABpqeKvEQZl/4S28T6Sn/GuQaWNYvLRSWakwqrjGX9KLCbOqfX/EErMP+Eru2XH3vNP+NVrnX/EUp8tvEt00WOD5rf41z/7sA8bT6UjBXHzcEdKLBc6eDxTr8cRQ+KLtBjhfNbn9aefEevy4z4ouvTPnN/jXHgCRyGzux19KkKquPlOBy1FgudPJrGqyHbJ4ku27bvOb/Go31PUFJT/hJb0JjnErf41z58oAN0WkEiMwLJz2osFzd+2XyuDH4kvVOOplb/GpV1fU+Avii9P/AG1b/GufkZDGSY8t2qKOSNVKlaaSC5041bUkbB8TXvI/56t/jU66pfomH8UX5yOF81v8a5ZHiHLKSKdI64+ZTn+ECnoFzdW+vEky3iS+Gf7srf41L9ovHGT4pvlP/XVv8a5lWUr8wwR2FCTqT86lfSk0gudUlzf7CP8AhKr1cdvNb/GoGur2QgN4mvs9/wB63P61zwnjwS6tntiibywqs2S36UrBc6Am5XlPEl8p7/vW/wAacNRu4UAbxLfyIeCvmNj+dc2kwcDepx2qQyLgbV+tFguaf2qWGYsms3oU9xK3+NSm9uABnXdQGTwTK3+NZTlVlA5Kkcj0oWRNjM/3R2ppILmq088keyXX74oT081v8aiinljYKusXaJu6+a3P61mK4kQlk+XPFMZ1XO9ScdKGkFzauNRnmYo2sXcuOgMzf41XBUZKajc7u/75v8azSyDkjk/xelOWaE8dD3PrSsFzat7qSF939sXKp3xM3+NNmlLTGVdVuzz181v8ax1MQJP8NOeZVGByD0FFguazzEcvq94Qf+mrf41FI0RBJ1K8ZfTzW/xqhvUL8y4Y0KyyHAXjuaLBc0IiqZaPU7uJe+JWz/OpY7wrICuuXy+/mtkfrWL5oSQp1J6Uo2Zbcp3d6LBc6g6i7ow/4SO+cEdDK3+NRidipK+IL4cc/vW/xrm4po1bDJgH+KpWljjbKruXuaLBc6AO0hB/4SG+24wf3rf4068uPPiCPrV5Oij5cytwfzrnS4O8hdiYqUTobfEak5phc2I9UnhTb/bd30+75rf41FLfCQ5k1G9kJGM+a3H61i+UI2+dCxxkYp7OGIYEhQORigLmol60JCpql6F7nzW/xp1zqm6RCmq3rIo5/et/jWN5ihlXOQenrSbl3EEFKAbN4XobDrrN8gxz+9b/ABoGpSYyNavsD0lb/GsOKQbiipuXFOEy7Sy5CjrinYRuwamyKzprV+T6GVv8ahNxGXD/ANr3wPr5rf41j28qKTuGQ3Q055Qh/vihodzSmaNwN2s3rH3lb/Go3+dAo1m9I/u+a3+NUMxkgomc04rEAG+4fSkFywbWz7X11kdcytz+tCxWRAEd3eA9/wB63+NUizSHBjA9Ce9SB1wCmUHfjmgLmjFHFGyumqXsUeeR5rc/rVryUlOH1q+fPIHmt/jWU8gCoG+ZTUcMyoxQnYCfvUDubSWsDAAazerg85lb/GnOluy7Rrt915Blbn9aw5ZSvyr8+TwabgSHaBtcc0rBc3bd47Qu0Gs3kI/ixK2T+tQzXglHzavfHnoJW5/WsQDYxZlyo6j1qRdpdWB2DP5UWFc0UuFDMBrmoI3oJG/xq1DebiP+Kg1JfUeY3+NYJIF2ZCvyjvVkOp4H3jyKLBc33uIlUlvEGouD28xv8aSNrSYEf2vf595W5/WsJZRuG9Nyg8mknkVJA8ec9qYXNa4WGTkanfbV6/vW/wAaSKK2JyNav4W7fvW/xrLmmLRh88/3aiMm8ZYHJ6DtQFzoHjgl4k1y+ceplb/GoZLOxbhNSvA398yt/jWQJDIu8L8nekaYIMM2QegoC5pJDbhyp1W/I7t5rf41I0Onkn/T74nHynzW5/WsdnIbJBB7qPSnmUGNmLBRjgH1oC5srDZFGzqF6Gx3lb/GoWttKL83V4wxyfNb/GseO5aTIJLY/iHamyu4mBU7lxyPWgLmu9ro6E/6VeMcf89G/wAakig0to/+Py9V+gHmN/jWSp2nD5DH7tKrOJUibJBPWgLmk4s1YRme6Ydf9c3+NKIdLmBKi5YgdTM3+NZV2zG48tGbA6k9qA5HzKSAo596aSC5bnt9PBUwpOq/xETN/jT0XTiP3slyFHQCVv8AGoIwkkJlU4A6r61ChCqSwIz09qLILmnBFYH5WvL2IHoBK3+NSC10c4AuLvzP73mtz+tY75IwGJY9KEc42ZOe5pBc1JIdNJCC6u8+vmt/jUcltoy4ZmumA4P71uf1rPdlxtU4Pc+tIXadVVcZGeD3oC5oG307gJNdFT0Hmt/jSJHpak72uhjqfNb/ABqlasQPLyQ/XmpZXXgnPHXAp6BcufYtMwrfaLsAn/nq3+NXI9M0thjz7xuc/wCtb/GskM2Mbjz90Ukd5LAThiSeMYpBc2IbfRAzKZ7xn95W4/WgwaKFx593jPP71v8AGsVZQjBy2eeaGlIkyeh6D1oC5sJZ6UzjF3eKpOBiVv8AGo7iy0UPhZ7uSQHkmVuP1qhDJtY7j8p6fWnROHJQnbJnhhQFzQjtdIYHGoX6DPOJW/xqwltpmf8AkL6iiZ/56t/jWQ/7q4CqNxPWlacK5AP1GKAua5tNMdiX1XUWXt+9b/Gniw0gNldV1Bf+2rf41ipdZkxn5PpUTXUolJI/d+tNILnTRw2kbHbr2pID0xI3+NMns7R8s3iPVPxlb/GsFpSy5LEE9DjpSzXbSIEILAdT7U2guax06zKEp4g1HH/XRv8AGqrWpQnb4j1JY/8Arq3+NUo51eUBSyRd8Ch5mDMG5iHelYLl7zVQkHxVqe3HaRv8anEkLJn/AISjVD9ZG/xrGaOFm4ACHqTSO8afIi7h600gubBFvJ/zMmpsO+ZG/wAabHpvnlvsviG/YgHGZW9PrWQqEA5+UfzrT0Pa9wzKu0KpoaFfU+3/ANhNprj9lr4wvcXM1yyQyIGlcngKa+ALC3Pkyk/xSuQfTmv0H/YMiEn7KXxiIOcpIP0NfAtvFi0mxllEsgP/AH0akple5gVJBvck1R8RgHRGIOeR/MVpMEH3gTxWR4hVRpLlT/EOD9RQScjRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFA60UDrQB0sDZiiyTt2irI+ZTk8dqgRt8UYIwNoz+VPWTblScCgHuauTHplkR/z9L/MV93/ALf5J/ZP+FDHr9mT+VfCVqyvp1mCc4u1/nX3h/wUEkUfsp/CpQf+Xdf5UFM/PSbquBzgc1HvKnpk1NOPLVSDkkdKYreYQG+SnckbuMhG7jFPiGMg9R39KaWKkjGT2NLnP0PWi4EhYkqfu47+tSNKXKgDr3qMqDhev+zSx/IwXrnpSAliOA3f+lNjUICQdyn9Kc7BV2p989aZjcihD0PIoAXGQFHJzkH0pTI8oCg4Cnn/AGqFftng9PrUf8OQcgHmgB0RDFgBx6HtTi5kjcjoB970pryBio6c9fWljcQuwJ69vWhoBpctGpPy8Yz61LCgeMhjtAOQajjdQjbjnHOPSmB+FYn92TigCUxlict8v970oTMBIPXtSBwpIJ+TsPU03lSQzZf+VAD1b5jk5B4PtTF+STAbIp/m7B93JPeo/NCnOMP/AEoAc0rHJTOB1p20GI4PWofPeFmZRmM9RTw3mBnQ/L6elACo8kaEDkUJKcEk5IoSXK53fL0+po534H3v7tAJCtMDGxPU1IiFYC27GaRpM5xHwBy1IFPIJ25HAoHbzHRIOFLdakkungPlCPKY5b0pm0+XgnDdhSRzSQgq671NAWGxoUzjowzUgl2BVXk9jio0Zo3YY4I4FNDEZz8rAcCgTRbQ4mVifmx931pu0TSuOjnoKht51ZlDnLZ6058+ax6SY4psLIdIzoVV13EcZp4kXZ1BqOOUheuf71KXjJ+QY9aQWQ9VDDKnaB3NP3KQOoHeoTOJAdoIA6imJKEGWPHYU2KxKspgynLIacESRSyt0/SoRdFlAC8HvSRLliAcA0gsS7xONjErjv60zcWI2n7tBYCTCndjqKQvvYHG1R1ancYOxX588inSBdocHB7mmO+8gfdIPB9aRipGCcA9aLgO88g/IvzHimjk45DZ5FI6PEAyHcfX0qVJ454iQCki880hWGqu3PNKF3Z2kA1AHJYyE4XvSvhsMBx60BYkMW4EbvmpzSbAqg5J6mkKKOQ3zelCxvu+Y/Mfu0FIjbhTj5nB6U9J9mQQSxHBpZAyYYLnH3jRN++KuPkxQS0NBGCxGGp8kpiZXxuLcGmnBbcT8vehkLdOR2oGPlAnJz1ApuxgMnpQxMY5OfeoxLnIJNACsmHBHA9KdNLJkZXAxTWIBByc0gYkHPOelADiyyclcbR92nKxkThtlNKNGvLZaggMpHIOOlAEglMY2bQ4I+96UCRthYR8AdaamJlwhxgc0oYmAxr+LUAAxKowdi/1oacJ8ije3c01VUOoPzccD3oGFfaPvHqaAIgzN8p4NSQXBSNkZdzHofT3pUgK7izfLRD+76nGelAEg2hBnt/F600klgzcp6VIM2yneuQe3pTFJwC52jtQAoifglsLTWZgvy8DP3qRkZXXc2UNSbXhRmY7l7LQAksnCNnafUd6a0vnY3fIo/WotxjiMhOVz+VTblhVWzuDc49KAHCU7RxwDwtAkLMVxuf0PSoTIeS2QCeMU6SRmXCct/OmwFklExHBRF4K+tJ9mBViWwOopzO0keCu1l/h9ai5kQFyVOegpAEaI0e7ByD09afwVJ6DsDSSKbch2cFD2FLI+5NwX9360APMrGPaw49adFKFcEHt09aEuFCDd8yCm+YDnAxjpQBGw2s7M2GPSrMd0rKYzye7YqvMOAW4U9WpxRYuhyhoAc6BQcDKHiiD/XYJJTFJHK7sQ3boKWO7DhlRNoH8XpQBIZGlDcfuwKkhUEgA4QioFlAyR8w/u+tLGzMrBvlU9BQBYztYZbIHp3qKTeJg33Rj7vrSfKmOflPFNedgy7RuA6+1ADplSXDFSsgHGKY12JI9oGWHVqlEnmt5qHGBjHY1Aj4kJKbMj7o70AIGAYFWIz1pwmIVkAwD1NIWVc7VySOaZ5hjxjle4oAkIBVSrZUcn2pEIiO7cTvqMzBZFxlENOQ85B+QUATRtLBwmMN0z2p6sLg4fgjt2qENuYZYr6VGsrbsLz60AXDC+3BkB9KRH2nj5m71EkiDGfxPpUg+Q5jPz9/cUAP82POTyh6g9qcUhbCs+709qjluIjtwmW/u1G2FAduhPT0oCyHlTHgK3GaWVgMKuSxqKR2Kgn14NPhUM6n+LPPuKAshzKGYIG5prSOCAFO0e1JPJmbhdqDvQbl9pxwB7UASAoZQy9+qmpBFszluvf0qJrgI8bgfKerUrSbic8IeQaAJGkkC7goZQemOtG9LvKoTG/v2qP7S8bFUO9j0HpUUjNI+4/LjqfSgCxLtiAGMyenrTXkeWLaBhe9RZDfMevY0sTbZcZoAdF5kLYHzRnjBp0ig7kyeBn6U2e5Z9wRcYxmkMgDZDZ96AAfMMCTr39af9nSbKvkNjpUefMJB4PYU5izcg/MKAH+WI12xceq+tBIxkfKfT1pqSMVJc/N2pxlLqV6tj71ADhKzIRjJHf0phldWDD7w6H1oB4wDjjp60IFlIUtiTsKAHeaZGyx2OR2oSYoCgXdkZ3CmSROJPLY4fFDDy1BDY9adgEjV0Jw3ykdKkWUKQCN2Kj3HkE/NiguwjEg5C9aLASGUuenJ6U0zNnAXn1p6lZowyH5vSnxxk4IOD3FFgIoV88HacEdaa7vHzIpXHClRU32XefmkEY96QebCy7mEq54PpRZALG3loJFbzD3BqZSJlGDsJPT1qt5zGQ/LsX09alilCqc846e1IBWSVWC/eyfvDtQGa1kwVEpbsO1IjNGzhWzvHftTQ2FGHKuDyx709AF3x8kDbzyDSvntwvY+lOdY53GxtxH3iKQSiM7PvR54osA1bkt+7UfNnril3gAqCQR1qMnYGJO0E8GphMgQYXL0WAFOVwGO7sfSgMzNsIG/tUaS7Zeu1s8rTRGrM5DET5pASMsxJUrwOSaX7PJsLb8D+7THVkQEzfNnkd6QSBs7mb6nvTTAkmaSS32xrwp60sM0hQxsvzDv60pzFHkNhfSofMZeWbBPQ09QJheEo21MAcH2qTf8hOd6dxVdj8pcfd6FfWhLjz/lRfLA6mjoBMkaSozjnHal8osOOGx0pgPlyAq2RU28kls4bFNIB4n+zptZc+9WNOB8yVwdo2E4/Cq0bK46cHirFj8s0yls/If5UPQVtT7n/YJk8r9kL4wytwMSf+gmvgOxuT9llAOQZXP6mvvj9hf/AJMx+MWP+mg/8dNfn1ayn7KQpxiR8/nU2KZc+1bmbA61jeIy32B85wSOPxFaexmP7ttp71n+InP9lMhAOCDn05FFhHJUUUUgCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoFFAoA6W3n2RR9/kHFL5gfJ+4cdaZCwMcYzghR/KlJbk54AprYGzXtQTplmc4zdqCfxr72/wCCgKJ/wyh8KmV9xW3X+VfAdtJu0q0UHrdr/Ovv/wDb5t/+MQ/ha7HkQKP0qOo2fnpMhfYR1wOKaFDZPQin5x5bZ7CkaTIwFwT3qhEQOHXnNPMuWC4xmkdMYOeDxStkLknDDp70AJIgyCD81OWfBCkcdjS+aJABjDevrR5oHylN1AAH3kDPzZ60qRMrkq+B/FRDOPmUJ1/SmYYgj7w70ASNMVbgZJ6GnyG3miwPlkB6etQ4+RSp5B6GpZJIyq7Yvmz196AI0jKjDHr0PpT5YGGz5vnB4akdkcKOgz8woZVX5SSV/h+tAEyvGsbEriXv71Csi8Ap8pPI9Kdj5PvZcGgYmyEO1+9AD/K84YBCKDUcxDOqjqp5PrSzY2jBwV61G+1NvOVNAEvnbshxgjpUbuoyw+bHWmMQ8owcEU4gISQcg0ASLgJuR/k7imOBHko2A3UUixhcsvI9PSlcjyzjg0AAGFIJznoKarsj43dvvelIIsjk4fsP60hCBwpJx3HrQBaZjEuc5UimtKwiwOc/xelMc+WQQSU9PWpGDEbkOUI5FACo2VKu3OOvtSAs5CBvoaWHcsZaQY9KSadQAyrgY5oAWTKp8jbmFIG3gE/eqRQsg3Rt/D0qNQQ3JoAbPb5dCh5/rTwzs4BGWHWkaF5nUI3ApBzIw38gcn0oAkkmLnai4H8RpI32DbjNRq+T8r7cDk+tPhlABb72P4aAJpLkQYxHx61BLibBkGB2xQ8olw2fl9KRWMYBPIPagAyQAAcgUsY3En+IU1Yo1YNuwDSLvhmx1B6UAPACMSnXvSrIDg7sLnlaJZEVgFHzd/emYHG07mJ6elADJGzIMnC5zmpbgbUURneG5pPK4/eH5c9PSo1i8twY2yc9KACIMoLjnnkHsKsMUyHU4XvSCQNJleF/ipxXa2M+4+lADdm456A9qeY9g9RngelI26YgoflXg06WBliLB8N2oAgWILLuzznpUs8jSFWzjFRIrFSc/vBUq3C+XnG8D7woASO4bkfezQ2H5PyY7UqyB1LD5V7H0pCcjcp+Yd6AEEPmqSDQXLKUzgdzSmUSe4FN3KD1ytACRMA+xzuX1oZ/nI+8AOtPZ1fgLtHeo3Qx9DxQAvnhFP7vdxSAs7jC/NilyuMZ7c+1OhkMbgFuOoNADnIliw33x0FNZpAN23kDp606SECVnRshhzTjM3liMHBHf0FAEXmeagC/un71JtB9sDp60EoeBwcdfWm7spsP3vWgBfNESHjJqJ2fb/t9jTpDtKnOfWlMuDyck9D6UANgImySx46ipAVwqjnmlVkT5kOT/EPWnRyLIjOpwaAFlUhlLNle9NijM2d7fKv3T60wFpflB+tSJGYoSJG4PSgBryNKQJB8g4FQgtCwG7epP5VI77I1AbJzSrcqoOE3E96ACPa5feflx09KiVPLwH+aPORUkSGRypOCakaPygFc7gTxQBH8+4DAKk8E9qSZSuELbXzwR0NIAUbBf90x5pZUeEjI3Ifut6UAPYFgMNhgeT602QMhJLfKelK0iuo3cKDyfU0x3MRAJ+Un5aAGFEKNuO5qVJGhAydynoPSnKUc/K3zZ5pZHIJAbLZ4oAUEJliPkPUelK0kagBfuHq1OkuGZQCP941EE8rcCcxN+lADzIPLCv07UvlbUJLbs9BURVihycqOh9KmiciPPU+tACM+0BeWY96UJgHPyr3qSGVoZS7L5ntRNd+e+4jYB1FADOR9wfLSpN5bgsMx9xT8k5wcAio/tDPGVROV6mgB75lO5D/uimszJ87dem31pAyvF5iNhl/SoxcSOctyf4T70AOJkmGB8uP4amR45YWKtll4qLzpGHIw3ejaM7l9OvrQA1CW5zhgKfkyEOo5X+H1pTvcA7cds0jMyHCHmgBGlL/ej7VHGzIwyuUqWJnVsOM5HWnOVG0/w/yoAeGUHGzJHIqupWV+BtbvUxlEUZA+Zj/F6VF5asoZT8w6mgAOdwjI2qe/rVhAyRkkgMKrO7zAAthR0anJ+8Ta3DDv60ASqkb4IOHPWl8sJlXfdnpUBchdvQdqXzGjVcHLZoAmjKqCH+bHT2pIbeVSXifK5pvmbyCOBnmlLbCSSVjP86AHSS7+GPINQmQsCw+TB6U9YV2/IfnJzg1CWDSEMeRQA+ZgwVC2M9KQTlAUYFlFNlgxHuV8EVYRd0Csj5b+KgCHeoGUzvJ6+lWInWXPO3HWovOJJ2JuanMwB5Tb60AMkicklThKdBHJHKXkOcdKkkBjHmA7ox2pjEsNwbcD29KAGtMSSynHPNOecnD7cbeQPWkC5bjBTvUksuEwoyB/HQAPOs0e4JgjrTo5UYcj5u1RhGlICdO31pXV1k3EcgcmgBxf96QRjjlaeSrr14A/KldkaMZ5bs3pUDq6sGZiCRwaAJV81RzGCw5BNQzQtPIHY+XKBkVLJZSSKJJpsED5aeyhoQGk3uO/pQBHJMSgRj+8H8VMZdzrvO0Ypwtw2RntT4I4yv7z5sc0AIrbcE/ex+lSWkqxpL5n+rIqMqDKCTwegp64UFH6NQALEmzzonxjtQs2Vw4Kk09ESMjHyr60sl0sICuu89j6UAOjWJ8CXLDHDelS2cUMIcBiG7bqiWRAu6PnPUelRtKvDyH93QAjPlzuO054p20zYVce5onjdlVu38Jpu5IrYhnxKxxQANC4ID/dz/DREyRSCO4Hyk8EUpjaJEKyZBp6wi4UqWw4PegBRZy2shaNg0b9cUjSDJReX7Goot8GULnbnqe9KiSFsjrnr2oAsRzmVWWRM7ec1EkiOriMZY9D6U+SXYAvAbPzEd6hYeS+9G2jNABb3UZdknTbIOhoTcJWw+ZBz+FS3CJEiMx8x3ORQs7DJEYDtwSaAI5EhllEoJ460m3EwVThGpMCLKhuv3qGAZQM4UH73pTQC3UBt5UDNkN0p727xxksQ6dhUbgDO9sqelLNDJZlS5LI/wB007gCNMnRdw9PSle7eLpH8p6mkVzGTk4B6UrTkNh+h60wJEyo3MNwxx7VLuIGcgFhgVWEn2RtzNuQ9B60PGksu8vgkcChMC2JDEQBzgVPp6g3Mj785Q8fhWbjYcB+TWho5/fzEnonWjcD7n/YNLSfsgfGOHttkP8A46a/PqyUeVKDx+8f+dfoZ+wEd/7JHxjGc/LL/wCgmvz6sFVlnBOMSv8AzoAsfZyF3LIN3pWR4ij2aY+RtbIz78itiWKN2GMqwFZHiFn/ALLkBO4ZHP4ilcDkaKKKkAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKO1FAoA34pQI0G3+Ac/hUoxIvtiool2xoc5+QcfhUodXHAwQKa2BmpahRptkO4u1P6198ft/Syf8ADJ3wsX/lmYE/lXwPan/iW2pJ/wCXpf519+ft8MG/ZF+FhDZ2wJ/SpW5XQ/PRyFxk/LijzCR15pkzZKA9wKZvwOBkjr7UySV23YAGfeozIXwpG7FOaNjt8tvmpCrKwX7r+tAD1JbHOGp2eR8wR/501WCYycmnCNJRuJ+egBC5UAY59aZtYAMrYGeRT1wvIbB70FJHUGNgxPWgBS7HGBtI/WnSzSNGExtyelMZmSMKxBOc8UMMKoLZDHPNADQrdCcjvmjzTvG77o6fWnsNoy4yO1IJI1HzLxQAocudxpsj+SwkHNOYx4DJzTpBE20gEnuPWgCIylkJz8x7UkZ8xGC/eHQU9ihOFUhT1b0oQ7H4OB2oAYXilIB4lB+anM4kOAMIOp9KcI45M5G1vX1pUcKSjDKetAAGYNnOUprj+IDjPFJIobJVtqr0FKQyIecKaAJpd21ckFz0pjFofvJknvTNkTKQ7kN2PrSI7chm6dM0AEkpz8wx7URSSFSVbC96meQmPLbSx4470sTHcF2AKRytADE3iVWkffHj7tPdhcyZxsQDmg4LHf0HQU3bnJB+XHSgBY0VNwTIyOKXaAAykkgfNmkXJ+Xdt46UNIVZVZSQByKAJPNxFtyFYjjFMRBHhgd3940648oRiSJsjofaohJt+XdhSKAJGfzeDDhR0ahICzKFwQetIZioVQee1Nds9GIPpQA8AxXaKfmUHtTnINw47Y4NQJkkFTuI60LcMzBgMEdaAHxusG7Pzn+VIk4Vssc5oRlZtwO09yaZMwwAqc9zQBIjDcT37GlkwuGByc81HuDQjbwRSouFU55zyKAFUg5LDvUZGyXch/CpY1+c4NDOFYEDew7UAOMigAg4X+IU9gwxICFHpTC0T9Bg+lJPIZmUZwooAnAG4DO1DyaiwTKSXwoOKRFYoST8oPBprDcTvOR2xQBOSsakKdxbrQiLbgjO7dTIo0wzA/N2p7OI05Pzk0ARmHYDtbg84o8yMRlXU7u59qaP3bnn5T1b0qVz5EeQBIpPXvQBCQI1LRjKdx60K6SKfl/D0pwYsDt4Y9jQGDE9B6kd6AHRy+WM7dxP60H5slhjPGKaOCSD07UM21C5P4UAITsGCPlHf1pTNGWxjGaRfn6sM44FPWSEA5XJ9aAGIWL4QiljYCRiWxjhqbmJmJUmM461GEUMSHyCME0AWmmD/IBjA4NMf5kGG5HVqQzbY8AZI43elJHjGM/IevvQASKGCknjt9aQxFmX5sHvT2mAAjaP6Goyrjq34+lACoNrnnAFSR4ORnC1HtO35hj/AGqkLkqFChB3PrQAhkdyIkXav971pUd0j8qY8fwmkExcYY4Ve9LGS6FXbeD0oAPLkBAyCTSnyYwQq/P3NM2QqMOxwOlOLhVUIAxzxmgBSMDJzntT48SIQpIf3p8kwcAEYb2qDO05ZqAHdMqfmPoKZHI0fyH5lJ5PpQQFTcr5Oep708SZwfujPJ9aAEYrGSCMg9KSeXzQpZMKtMkJkyUOXzTnDIqrI2VPpQA5mMceUjwzcA0xC2NrffFPjct8u7HoKJEdfvEBvX1oAeMt05U0eUeVLAE9BSY2rw3y98UxUDFiGYuO3rQA6IlYpI2OVpbflOTnHQUW6j5i/wA0fcd6VyrAtENuOpPYUAOdyFwvX+9URkUt842t2HrUiAhhzhW6H1NNmHmMQ/3wMBqAFYYRt5y38LDtT9jhAxYE46jvUH2hocoRuGODUiEbWLN8+OKAGSxFIw6cKTyKfsJ2kN8p6VLCf9FZZSCO1KjhYghXcD39KAImXnrzjBPrTXJVx5a9ByakaIqd2d4xSAlnUo3IHIoAYWMkiAPgnqKUybHIU7vWmpt3szUqEQt8pGW7etABG8gf5jtXtx1qUtwQq4J6ikSQAEn5sdhSRPvJx8p70AM37QeMmkt8SKecDuKkkkLEFFBx1FOWeMAFo9rdjQAx3MKAL8+OcU3LS7S3ytnjFSFvmBzhz0FORGZG3tj09qAIywdMY+ZTTmkWUo4wABg0xEkA2PgDPBPejByykUAS/MqHGMGomAjTDNu56UgTan3+/IqMog+ZPmOeaAHrdbuFjwoPLUjSEsdq/MTwPWpd6+UOhQnoO9QqhClkIDZ70AG5nyWOCDyKEiePPlHgnpU8I+RnxuNMjDMC47HpQAiO5JyfKI6+9PM24HsDwT60twxuQMjZ6mligdWBIzFQAjcxbM4b+GmBASFThu9TFldiPKOB0akLLyMbf50ARvCnKvLg+lKhNsrYbcnpSPGi/N/F6etP25TcDx3oARvOA3RNlT29KlWOdRuJAXvmowxKHn5qZJb9H8wsoHJB6UASgh8jOOKf9peLH/LQAflTC6EAdePvUsa4zjuOnrQA5lEgEm44PUUhKxrvHAPakVWhcOPm9qHlSVx5ibPQUAOgRnA52knpUk0Sx3ChWznqKQyiPrhmxwRSK4yGzlscUAOnf5xHGO3Wm/MMOp3460pmwAQfmpxWRhvhIbj5l9aABZVKkHn0HvTRJEADjZIePXNIlyrqP3WHFOjlIbhQ475HSgAkAtXVFXk9R2NIJBKcOmSOqdqjErbzvk69Kj85idhABPegCzJcvJtWTiIdAKZLtON4znoabtyApbp0zQsqrxINw7GgBY2MJwG3Ke3pT/O3suOPf1pi4JwvJPWlVwcqWGAaAHLhXIZsr1IpLlpGK7XxDnr6UyW4h2Aqp3BuTT59kqRvASAD8wPegBI2ljBBBkUngnvQZlOQYTx1FDSMgwWBz93HrTo2bg9JCefpQAhyQHJ5zxntSAvO+VOZBSrxIyn5gaWDDzlC20etADpt6ABwNx6gU37p/wBXlT/CaJYfLJzJvYHINI4YYZ2wD3poBwhZ1ZX4RuntSFpEj8t33bOmaCgiXCS+Yj+vrTUHyEbhvHahgOLZUN1WmO2ZA6gHHUUxtyvu/h/u1O/lbQ+dx9O4ppgRSOlw+4jawHA9KdNGUUOTu460Msa5kRDJkc00PIVIYbEPbFJgSLEeCp3Bh1q1pkxjuJIlO75M1QXz4i2DlMVd0hSs0hXqUOaSY0j71/4J9OB+yp8YQW42Sk/98mvgG2hkEVwyFSvnuf1Nfef7BzCL9kf4xszbQY5B+hr4LsExbOUbOXfI/GncGWAWfB6ACsLxFKp0+RR1yP5itl5XkOHGGHT3rF8SOhsmAXa2RnjrzTYjlaKKKkAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKAaKBQB0VtKvloD/AHBUhK549KhQfImOuwfyqVhtGfamthPc07c7tNtEXALXKjP4197ft6JJF+yP8LkZs4hX+VfA1vKVs7QqQCLhT096+7P277+a4/ZW+GCSDaDCvP4VHUvofAc8oKRqy/NgYNIke5ht4Pc+tPmUMVBbOBSNJgYXg1RIpQLjnHvSBSjgE5z3poABDZ3H9Kc8hDBRjmgBNhik2ls56GnjOzb/ABUqgzJ8rAOtNhimfOCC3rQAZGQAOe9CusU6srbc9aRsMcKwVu9BQgbSV5oAmngAlXY33ud1JHxlTyc0jviNUU5H8qjePKhkbJHXNADmUOcM3ANL5fysGIx2p73CiHbty3qKiXZKoUN82eQaAGxMY2IY5A5pxUj96vPqPSlkUo/TnsachMJ3E8GgBqurgA8AmjduYrnj1pCQrHC7kbmk+VgWBwRQBMYx0HzGm7tr9eBTsq6ht2PpTXZTx0B6n0oAUus6nC+WvrQHEcnzncgHFRBghxu+U09YiqMSRg9KAE+WVizH6UFhHKueQelMhZUfDHj0HepMZznkdqAHY3kgnb3FOQkgsT0703LMPmGPen+aYUCkbkPXFACRvwd3zClDeWxMZyuOlOYpwy4Kdx3pfPifhUwh6mgAZhcLmPhwOlQRTSgkbN5xU0Q25MWMUB/MxhgrD0oAjaTeRiPYR1HrSMxk4RctinO7E44LY7UscLupKH5scigBkSOjBlX5h1ND/wCtD/yo89kGxxlj6UocKvyjd60AOiTEbFT8p6+1Rr8oIJ3jsRSCUAADJz29KdIjRgEsoU+lAAYd0YYvz6UIxIDgg4POaRxvC4O2kCFnAyMDrQA6NxvLA7VPWlBPmBgcCmFgMq3zem2liOUOfmx29KAH+YQTt60HcoBUjcT3pkb5cdvc9qc5VBzwSeKAFaIthlcbgeRSsS8e0AMc9BTY4xGdwbJJ5pTMDJgfKM8sP6UAOfJQLuwnfNMBMfAIYdqUgRqe4zSbmflRtAoAnVjEMkhm7LSOw4YnJP6VBn5t6nc1SsTgtjcD1IoAUI3OWBQ+tNBeM4PK9h6Uirgffy3YU93ZMNkEjtQBHNKSQBwT1PpTsjGScDv71J58ci7SuD3IpjLuVhkBccGgALLId24Bh/D6iklzt56dhSrHGqZ3bpsfLTIxIGYuQGxwKAFSQKrE/fxge1NVfLGN2d3NOjIjJLjLGmttYFvu+1ADXORnsKRcFSV49qemZDgACoXLRT8fdxzQBOpYjhOPT1pykAHC7h2P901DkknLFcjgVJHuC8ErjrQBIZSMHI9we9NQkyKd2VJpd6O6sUzikZw7AqNo/u+tADgxSd9xzGRwKeVMa/fBx0FM3bRuYfL3podJGBAI/lQAobLrggA9aQymRtm3GO9IXicYjHzdx6U1iRty23PegCaFQwZSMjsabbglHGRuB4NM2nIw3yHqKcoMQbH8XFAEquoTDH5j3pJI2dMqQT0/CoVBBBPOO3pUyRsRuVsgnkegoAjUO0LLIMKOlOinVYwjDf6U+O4KMyON8fSmHYDgcE0ABZwDtXnPWnI2Mkje3pSwswb7wBHIFNEom37QFbPNACCQM5CqF+tMZVJJ3ncPWkb1Bw4qTKSLkfe7g0AKCUXK/N7VKZkdc7fLmH8qiM6k+Wi4NJjcCCRuoAekhhySu+PvinE+YflwM8YoSUoM9V70xXy5OcHsKAFZNhCSNx/KiQ/ZxuB8xPT1pqIGDAn5hzzTx9zKsNtADRMpBLLjjgEdKcJFC5wGbsR2pCu4Fo8MO9NgBBO1h9KAHA7jknC4+7Uyq8o+8FXGDTInWVnWVdsmMD0pHkOcA5wME0AP8v5lXf06Y7025RlKgED1xRnyypX5ifWnGQLJww5HO6gCPaqjAOVPUmkzgYzuP970pwkZ8xFduemacIsMAuNi9c0ANO3G3Py96Qkw8scHt70El5Q8YyBQZhMcSLhRTQCrJlQc7TUgO9Cj4y33Se1R71DAEf8A1qeSHXKkEjvTsAkZxEyyf60dGpUiE+0liuO/rSxSmQDOBg80Y2qc/Kp6GpAekjBWimIJJ+Q1G0cseMtu560gtywBzuHrTRuXhXz6+1ACmSLByCT3pkaR9I3wSehqUkWyqTh8nn1pHihnjZo/kcdqAATfZwQygrntTQyyZdVx7etNMgkjHPQ0+JCDyfl7UAWoLcrGzIQR6GoBlC2OPWmxFg+UbgdVpstwkj7WBX3oAlDDeOcoeDQvmGYKHyvpUdu8NvJ82XU9BT/JhecmOUxv2BoAmNw4kKgCmygPnJxJTP3xk2cAf3qVm2NtOC9ADCvJycv2pjsduB8p71I0uCSMF/SozOwB3x9etABHIF5U8d6VXWJyQ2UI5WmicSLxHhfWnKvc4FAEhkiZcovH8qmD7VGfnbtgdBVdVKkhcYI5qe2khjJErZ44NACAOjbi2VIp08hZVGzc2OvpQ/lOeH57D0pQ+FLK2VA5J70ARZIXbt34GSaAwdRtIJ/u+lTLeqoAWPgjk1H5qLc4iXjHNACqRCuZR9KSJnyXRth7e9REAyHc3B6DtVmOVEX5xjjgUAQmUq43Aq5/WnhisiqCFB6+9NEgnbEuBnhW9KUOLeQQyDKnkNQAr5Mm0gEjkEVFMC7KykAnvUiAo5ZXBDcEntSOvylCMdwaAGb8ERuN59RUseCMMQ6L2HaqyRLEQwYl/epVALBg20d8UANE6gEJHtOetKhVDvX5vUUO5ifJG5fT1pY5hOOF2ehNAEr3ImAxb7VB5oDlBtEWFY9KYryRkbsMueAKlN3lcEgnPGKAGeSzk4OGHam4Ze4FTWzl1cjiSq6M3O45bNAFkSrFEUxmZzw3pURXepVn2kdTTQrrcDGCT0p08ZeTYTj1agBVl5K7d2OhNMcs33+f6UBdp2bht7GlJCoV3Ar3PtQAkKoGIUk5+79aWYuj4K8d3qWRALMbceWT98dqbEj4CJKrIeu6gBCJI/mK5jPekW3Mr44XuM04zmIlVYOh4x6UpkIJyRgjqO1ACtJKjBdo2j72PSpTJkH5Qy44qF2xlQ4KEdaXz2RBGuCKAEUybhk7R6Yq1pUnlzzyFgAUxtquLkk4ZRkelXNNhWSaZ8qV8ugaPtT9ifzZf2P/AIv7DgAP/I18J2spjtlAbPzPn86++P2HyIP2M/jLJxnEnH4V8DWkg+zjOBl3OMe9AMe8rMxfOCOlUvEFxFLopDJtm3DB9eRWi0yEkbOfXtWRrvOnSEnJyP5igRy9FFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUCigUAdCjrHCmRyUFJvypz1xShT5Ue7rsGPypWXeD644NNbCe5pWbYhsDgEi5Xj8RX3h+3rfGT9lr4YxtFtBgXB9OK+GraNRpunNkbhdL/ADr7p/b1B/4ZZ+GTcH9yn8qlbl9D4CLByPlAIXj3ppdDjCEnvipJHHylwFJAA/KomV0fGcj1pkjGBKkphcdzSA8Bhgt6elOZtqkZyexpi/uzuPftQA8EcqATkdRSIdnIY8dqIirthTt9KVYmckZx/WgCN2BwwXB/nT0DghmwQeM+lKMow4Hy8ipspOPTPWgCFpRnKjGOCfWnxkHcyEe4p4KxjDrkDpUKuoDYUpk96AEBJO9fvDtUokBTcIfnJxmo1+VCSefap3cpbglhgnj1oAajM6kMwzmmSwsgJZtwp6jdCT/F2NMhGV6gnPegBI2IIA5GackypIdyZXPNIi/MVfGPajYjkhWw46A0ADq0LmSPlCeVpzxRKA6sAD1DUM/lEJwWPpTkUSA7sBx0HrQAwoo4xlP71GTIPvcDpTHZkyDyueR6VLJKgjGxeRyTQBH5kZB3r83rSiGZRuT5xSjk4GCW5yaDK6naGx60AKVMx3B8diKURkHiQMncU0YB4PWm7SrblPHpQBLITnagAo8zYDzkAc04HgvuAWmMMjPAB6CgAWXMf7scHqKQqqj5TuOOTTvLkBXYozjpQ7KWwqYIHPpQAi8Mu1s5HJ9KfGclyMrgdajJckAL+NKqzKcHGw9aAGkCLHPmK3U0oiJAOdq0uQOAM/ypAAOJG+gFAD98cQAiXdnqaa8KyYycr3FLyGVVxk06RSsi7cEd8UANIKYEnH90VG7FT8o2k09o2Yhs5Hp6Ugl3HnBPbNADBIpwQvzDrSKSGLZ21IrkEk4NDOoBKjk0AI0TkZHB9PWlQsV5TOD3p2xtgYHHNRhpULIDkN3oAEY7jtOOcGpQ8JBVRufNRKdqFQM57+9CKevG4UASDc5xtxz0p6K7ZBXAFNWYnKucZ7ikKGEAh9wNACzO6j91Hn1zRllUcYwfu0Biy5zjPagclgW60AKJvmLmM7u1KZPMBbaQw9aT/VZXI2f1pTIMEsMkdfpQAsjKUGRkmmiQhckYQdR60EeYuVH0pCJACQBu9KAFMcTncsm1z0FISzNhvzpscYDZPD9hUjfOSGIVgKAE8tic9cClJDqdw5Apfn8skkDinRxDy/mcbqAKxVi3BxxxTtpB6hlI5NSNDIoJYg8cEVFtPBY89hQAf8tlVSGXuasNHtztYEVAFEcgORjHIqQsW77R2PagBBlXATBB606Z1DqSuaaxIdVHX2pGZkHPWgBScq2TuGOlM+cRgMAmemaFYMwEZx65qS4b7QVJOB6d6AIo1C4ZW2kdaVWRnG/p2pRCW+VSMDrmnH5YiXXgUAICg4HSpBMEABHXoajjfftBwuTxUrB9vCgyA5FADZYgQrI/BPIpTL5ZwGwDxTnjPlBuEOeRSB0ZMMuW7GgBFAK9fmzz9KRNwzgbiemaACcdDzSzoZJAAyqo7igB24W6bipOf4j2prqSNykAmh4/4XY80xIlV9m/Ge/YUAD8/KB83en5Qsq/dXu1MM37wouA396pVkEqFPlDe/egAk5bnAx90nvTXYDp1qWJUYYcgYPSo3IMh6bOx9TQAR9MZH0pd4BLAc9M0kYIkHA3mgqyMVYYjPVqAEbK5J70hmMnRNqd6WTarbd2BjhqFZhzgFO4FACj/RCHUjYeopx2yOZF+XA6CozIspI+6AOAafvCR5Ix/u0ADzCQ5+9xyadt28xjepHNNiYqdyqCMU5pFnwgIiPr60ALCRGjbxjPQ+lGC5AYYz0prI9uMOd0ZoS43kbkx6E0AOkOx0yee2ac0pySvyDHOabPEdwbO9fakGcbieDxj2oAYkhg+ZOVI5qUyRybcpyarqMNtB2pUiHI5I29qAJGMeRtHzUpkRBtXgnqfSocssgYfjUjYQgjBzTuAoO3jpUoBAAGG9aiiKO+S3/1qGcPKArBQKQDwjFhscD2qOTcpwdqn1qQRqxBZtrUkkvTd82O9AEOCSvOWz07CnEYx82OecUqfJ0G7PWn7FkHyMAc8hqAI1QIhz8wJ4pip/EH6HoakcuHAVdoHU0srRNDkgqc9RQAB/JLELlz0qJt5jJIBf8AuU+OXjPB9jTlJd9ybeOpNACQOGP3QsnoelSMFY7nUfUU0uEffwR2pAcLubH0oAfIsc6HEhTFMRNuADkf3qJYGMYkwOvT1pCkiYbIKn9KAB5kRsBcn1pruWOC2aLsfZ9uMNu9KRHjdvp2NACxRkP1+T0p7qHJx2FIqCSVSDgDnFWnWOJic7mIoAjto3bJBCnvu9KGj+c+am6PHG2lMQk5dyD7U4Zi+UvvX1oAcTCAFCEtjih2DARqNnHeoPMMM2AcqRyTTnb5lTIKHox9aAJN4Qqpxu6Z7UklpLFOJI8NGR82KZL8g2HmQj8vejLSw/JIfNX+H1oASUHC4T8+1NFvKw38MOmKN7AqQ24Hhj6U7BfI34GPzoAjAMnBOWHQVM5+1Q7XwrLSxRrKBtYKV6midslVUfOO/rQBH5e0AP8AKew9aejFWBkYAdiaUEunzEFhSOLdkALHd7UAMnQo5I2qp6k96jiLKewHYHvViYCa3j2HLqaQxmWNWTBIPI9KAGqRjIOZB69Klj8pk+VSz56D1pCohAYruPbbSBiCGhTbzyKAEWNj84+8DyKnNzGFAEW5wc596heQ3MbDHlno1JGrFAq44ONxoAUsd/mL8rH7x9KUsNp3Y2Z4PrTWJIKtwAevrUsoQxoI8DB/ioAjBMoIYhSehHalTgkZ3R9yaeEiyX6OP4R3qKNTljkrz0oAfIAiEYJHqKYkiLG2UzGevrVi2byGZZRmE9CPWkVEKnDDdnhTQBCsizLhSVK9EPekOxpOM4H3sVM0e4luNp4OKiYwMSu4qB+tAAyBVYpjHvSrEzwnjCmkfG3y0PyHqw7U1bdcHExNAD45jApjKb8859KfH8oKkjcRkZpi4Q7cgg9c9aWR1kGEAyKAH+XIctIMYHX1q3okjRPMQRgqeKpCSSQ7WfIFW9MYCSYgAkIc0DR91/sLW6y/sb/GMvwCJPw+U1+f1uh+zngEh3C/nX6D/sPRlf2KvjBIWChhIP8Ax01+fcTEW6Edmfn8aAZKqPjkgHvWTr7Aae4zzkfzFaojVizbjuI49KzPEMIj0w455GfzFAjlaKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoooFAHSxL50MQ/wBkfyp7phcDqtR25JgQggEKP5VIOVJyM4prYDZhj3adpxz/AMvS/wA6+8v2+7cQ/sm/DEjGTEn8hXwnajfpumIoBc3i/wAxX3h/wUFgki/ZS+F4YAKIEz+QqOpfQ/PCdjLtDHgL2pAxXAyCnvSvnK9+KixyQBgH1qiCQsMZwufU1GyDOFI2kc5qT5yOFwRSyIqoBjcT+lACrBmNTkbR37inBxtBbkjhTUUaBABvBX3qWOdIxtmA2HoRQAphZ+eTnmof4wqjDVJ5yRNtgkJHvTRK0UwMgGSaAHrksMHkdaa0gyAeRmkZi10STtGOKaNxztHNACpKtuzFo9/p7U5irxl+pJ+5SAMUGMbs9Ke0e5cqME8E0APRCyAfdUc1AzglgmBk9aUJLGpU/Mv60yTaibH5B/u9aAJFO1MghmU54oSWOUN+7IlPemr5bbfLbn1PanCVk3LhTg9u5oAUxjbncN/bNG8LySC4p63EUgPmrsbpxUShF3NkEHpQA1ju3SHqKHdZV+Tr/FUrwqse4tg1ADtVsjC0APjIRsfqKP4myc8Uz5ChUHHvUqxRqmS2TQBHkhgCuM1MG6qAMEdab+83NgAjFCu4ADAKKAJBsSLBI9qXBUq3y7fX0qFriNMgKGGOcUqyR4wvMZHK0AOXzIpS2/5SPzoEpDH0Yc0AjB5GMfKfSmsY343YbHPpQAcseuAOlNSVw+1zkdqlW4UJhl3EdCKYWSTB9OnvQA9g0WGHQ9RTUKMCHHPaiRjkE8e3rTvMRlwB83agBJXVgCnUdcdqRXKsDuzuoIK5YEMO9CLnOcBewoAfIjwOAGB7moEcOxbAGKnHJDZBI60Sxg7W4Ud6AIUcnnaMZ6etKBg7xyO49KmSBm+cjCA9aY6xDIRsE/rQAi4bJRvl7j0pUHkMMkOpNNWNoxxjb3qRXVcAgbSetACrCwYsMbKbJgnK/jSSwYOBJhDSgE8dh3oAQKuOeR6elEYKHBxgUbkxlfvE4pQhZjnGRQArSBmAXA9aFXc5AINKCEIDjr0PrQsq7sFdvvQAO25trYX3o27WIK7h2NPnj+X5iCnY1GgkZSAw9gKACFS29d3zdhQ6OeC2DSyJnDkrGR1IpjXC42n5x60ADgngYDY61Ki+YnBUtjkmokZcnBzQFDRttbbjmgBuGRyrHK9wamPlthj1xwKryzNKFHHHfvUmCTj260ASu7MgUjgdvWkBRozI/Cr3prA7cg4YUecqpsI69Qe9AC+ZHMRuXZxwfWkWQK21wTkfLTsrLgtjYvYUhkQj5Vz9aAAqFTI5btmhAcbs7m9DSgFj8ozSASoTxt9aAGq8Tn7pRx0zSlucYGacVyVyAB+tOd4imAuPegBV4B3HaTSGF/LODvB60yMbhgEH60rMYUOw8+lAEYXKgYyR0PpUn2k7doGGH8VRq5eNRnjJzT7fknkE9s0AKSCRufcDwab5pib1XOM1LIiRLgDdk9qi8wsAONuaAHqN+WU7T3psah43U5C9sUnRuWAAORT0kZGMqjOeNtADF5HDDA/vUBSWymCueaRpg+cRnnrSRl1P3QFPrQAhUCXlsrUqo7HAADA9fWkkA8s5A9sU6IebGDuIbv8ASgB8wwhAUCQ9SelJGqumxj05NPDeSNr/ADx9RTI9rTH+Hd+lACBl3krw3TPtUgQI2M5iPWmn92Cu3cCeDSlWXO4DAHGfWgCNmVXKSLhW+6RUgXyVIyM9qdcsksCZUeaD1FQnDA7+CO/pQAgyHIbHSpgCiMDgDHWol+8TsyuOtO3NNxxtHQUAEZ3qQMKB0qRgGADFVPvULttPzDkelPSSOeRVkyrEYBFADguX2li2BxUu4RDDj5yOg9Krbi0jLFkIvBpXQ/fJ4xzQBK0QdAwfYR0X1qNlEyfK3zD+GjzSyAjGQOKYyGPDLyT1FAC27AHZJhqUjBx1So1Q793DAj5qnzGABjccfhQABS3JIwOlMEmxsHnNMjcu5Enyr2IpVXBIccdjQBJJbooEkbjB6ihWjVMkBj6CmLEq8lsUREI5KkEe9ADlVMAsxYk9D2FSBGLZXn0pieXIQAdrE05tzPs3bf8AaoAVvlI2cN3FOkijljGD5bg1EzcbWYDHQ+tScvEDxkHIoAFVtmCwZe9KLnyxteMFc8GmZHGMZJ60rFkBBwQaAI5IVJLoQCeoohYliqHnvTMeYwwRnPSnSwSQNnAC+tAExQZPy4YUoDRtuYru7ZqEKz/Oz5IPAqb52OWAJoAa/wC9yd2MdqQ58rBO0dzSSShWLIAHx09akyIo1dvmz1WgBBGJhxgbenvUUi4kLqOR97FPLMvz/dB7U4gqMjn2HegB0kcciLKjbT0NPERjG1fnT196iXG4oCMMOQaWQOrCPdtBoAnRkBO8/NjkGleSLeBt4xwtVijZwPm4+96Ueb82xiBkdaAJpIWOdihkI5PpUKsFKrt3Ec0PEyNgynZjPFJvLYAwB6mgCbB37hht3UntTY91uzEgeYR8v0oNwGTaMLSbH4LMCMUAMKSQH5wBG4yT6GlVwVG4gL/DSPCrHY0xKkdGp1q/2X/WxiWLnBHUUASPAQoJOwevrTgzKg4Ge1MEkRAPmERHsaTzFUbg+fTNAErGOAbmUuzdh2qNJY0ILRcnp7UZ+UEZyfWg3hTC7N7dsUAOUbTlB16ipPs7OnmxNtKnkVEk7sMgBQOtI6s5Uo5x3C0APi3Z35G/OMGpY5GZzsAznnFRoduTMu0f0qJ5GRgyEJHn86AHy3ixS4ZdyZ5Iomi3puRsK/ahsSx5UANnoe9OW1BI2yDdnoaAGRcAxSDco5BpzyII9mMpnk0rIyMd+CR0pChdgij5yc4oAcIHXDqQ7Z4FKrN5jfKGY9faojI9oxLEAntTjMSODtJPJ9aAJHV/ukgjr9KjRcErlcHnJpqucuucikiceaynGex9KAHNP85UriPjOafcTQxYDxF4j1YUj5TKkbk9ajDtgjjb/doAdPF94wkNEe3cUyQ+VEFUA59KaflbcrcHtTgu7JbAHrQA5ABw2GPtUku6KNSQAG7+1N2xKPlO5u1Ti2kuIt24Fh0U96AGCAqA64MfrUumk+fLwMbDx61X3Sxtt2HOOFq1pu6SWUsuCE7UDR92/sYO0f7D3xcZSPvSD9K/P6zy1rgH+N/519+fsc/L+wx8XCTj944/SvgSwjaW2U4AAZufxoBlhhKeVXgVmeIG/wCJS4YYYsMfnWlMGY7TIQe22snxDGY7BgX3ZI/mKBHL0UUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABSjtSUUAdJGAscbPyNgx+VPUK2TwvFQxPuiTABwg/lS7lZCW644FNbAb9u7Cz0vGBIt4pX8xX33+3/AOZL+yP8MnmIJMcYz+Ffn9azD7DpxYLuW7Xr9RX3x+3rqC3f7JPwzQbT+6TkfSo6ldD89X+6sY+/gYpjQkKAOV7/AFp044Xb1Cj+VNZiyAg4fHaqJBS7HDPhh096jYssg5ye9SshYIQoY96dLHtZdwXmgCKNPMJ2429waeP3Q+dQVPamhdrYU7RQzbSCSCfSgBWLSAHYEI+7SyXG9FWSPPqab5vA3YbPSkDliAT8tAD5RsKEfOp4+lOdssI1Iz3b0pocxKRxg9/SmyMikArjP60AIp52Fsc8kVMN8TBjgjOAo6GoNm8DnYtPR3QYyCKAJnlSY8HYB1NI3l7CGwRnt1qMEE7GwF60ssYiYSKBnsPWgBqyxoCFQhSep70sYeRiwADA8CnFw7KuB83X2qJ9ysVBwfWgCVhzmT72abtC7gy5Q0oRUB3fO3rTfMJGcAntigBHjBjJJ4HQUgwEJc5HpU29WjOPlaoGPOT97sKAHJITkbeMcVIkasuY/wARTEYhsFeD2FKPlbK8e1ADfNbO0/KxpjghiGbdxT8YzuHB/i9KR2jQFQd2RQA+NNoAADZ70KdkrLjqKWCXy8jgjFIrhyxwMY4oAPMWNfLK5Bpo2A4xkY60JMHyhXDetCJ1X+GgAjUN93lal6DG0b+2aYAeoIyO1PVt3DABvUUAEO6VjkDzB0qQoY2G7knqaiaXyz8i4J6k0/ztuCOc9aADG3KLwO9RoAgy4+lS74pMBvlPakYbuCBx2NAEYnJI4pQ25gB93uDSkqQAQFx0NNlD4DDbkUAPVmGV/wCWZ7U1sYIC5NPCkor5yc9PSlaMq4dBnPWgBiqxAzg59acYSq/OAF9fWnbgvzKOT29KY28FXP7zn7vpQA5c7NrqNpPy05YWII3AUxsu2B8xPUHoKkIK4VuB60AR5UJtdcc8GliT72fw96DE5yCPMGfyoBKSHPK+npQAJMrvtcbSOlK5JBBGV7GmOyg/MAR2NOZ1Xo24nt6UAMYbQcHI/u+tJna2Qdp/lQrEMTSuCW3H8vWgBpAK7mOaVRsBKqMelI2ADjrSK24EHrQApkDMUCAZHWlADKQThh0pMgctjaOvrUTvmQFAOOme9AD0ZVyCMEjmnNkoRnjHFLIUuAHZdrDimgqw6YYUAO3sVAzxinEgDJG4YpkTqAcrkd6kR1zjGUxQA1VCOAPusKdteNymBtPQ0GRWymzAxwaRpT5KjowPBoAejsFI+63YetKwJwWfd6+1QDIbO/J71ICGXn/9dAAWHBT56eLnJxtG6o42xgIAPUGn/K2NuCe9ACOm5QRgHvilXJG5fmI65okkEShgvzClVw0ZMfDt1FADI0VWLjn1HpSoobmEZH8QNEaFCNh6nmnKnlsWHyH0oAVGweFxk4INNZAflYBRnjFPDFiGPI9PSmNGhGdw+hoAJ1JIUBQwpI/MdhtG1uhxTSY8Y+Y47GnKxIAQ4BOMntQAK7BirfeB4NBYksScn0pGUg7SR16+tBAU5U7W/SgBxjcjIHI6CpECofmADN1FIhmIDP8ALg9u9PMfnKxOPMFACKGJycFhyB605fmfkDDHmo0dpQCBypwfSgMqyl35HfFAD/vMykjavPFJjdyxJ9jUbDaTKuMDsaczeYMjHvQA4oQeACp7ijaVY71wn86aXcMEjO1TwTUzswBR2DDHBoAjMjbSy4wB9z1pfOZot4QLJ3HoKcpK/vFCnAwQaZ5O6QmJshhkigBFhO0hSG3DJPpSOwUbcKWA+96Ch9yHbHwoHzUKw3YUAsRQA0AjAXqeM+tPMiqQmzIxzTChSUJt567aeu6Nm3AAEdfSgA8xVIIXg0jsVO7GFIpglaQkHbjHFOLiRNrnHpQAgPmLmPn1pEZ1OdpHrn0psMTxHKHbQXmdxvOB6CgB7dQeCo6Uqq/DZDevtShDEu5unYUyIZO7pTsA/Gcj+A9aVDHEw3AbR0pSpGABwaUiLIDjNDARisjqVwuO/rUhDSc44FRSyIzKIkIx37GpDIQg+XHNIBrPHMyqQBg/e9KRn3Ng8IP4qH2A/MMA/nTtsKIGzuXP3aAEjwNxP3e3vS5UANncM9PSo1clyF5T37UwHa5UdD3oAeWV2LINrA8095C8eGPyg9PekRCMBRx60oQE4A5zQAsZXy2I+/SxOwRg/OaZ5WWOMZ7e1SloIF3P80np2oARXiRzwHb3pxBYMwwMUihJwZAoVxSlfLj3sc+1ACSjdDuY/hT41XaH3AEDoaQhZI8kDA7U0rE3CjDe9ADH/esxxsI6GlWcv8ki7j61LsBO0gbe5ph4kwAuwdTQAi/vMgHbx0olBdNjpg9iO9NeRfN3IuccZqQhi2d+WI4PpQBGWbKxugG3n8KeeSd67RjinSTNEyIyCQkfepjRMz4Y54yooAbvjDDK4yME+1LcQNG8ZgPmRd6cRvXBUdOlM2ywx/IflHJxQA53yATHk9qRFkJURn1yKRHlkGQeMckdqf8AvdnTaw6H1oAYTuABiAx1U96U+SgBMRyeijoKVJnZSzEZFG6RSG4YmgBqOSdsrc9sVIvmQEbkyT0200qHj3rjA6kUkXmswCPkepoAlQYcGQ7UP92pZYmwhtsAZ59arvvZSjjb/WkXcFXY/INAEyxu+WYh2HVaYXjYbZRkZwAO1EUxYt/e/vUI23LYDyDpmgBzyCRBGg2qv8R601YE3H58tTpJRjDoN3Xio98b8xrtzwaAJYyckO2T2pYbgw7gifP/AHvSoF4OGOB2apN5HXCjPX1oAhfewLPlnp8avt5jO3NPc71Ow4I9KWCN3VsXBLnt6UANM6iQhYmB9aUA7W3JgevrS+Y8aNFvDg/xGkSSRARJh17CgAjMjHnBUdFpZcMx8zCtxipCV3jZy571XlhJlJbDnjigBJGUE5A2dyKeGSVdsf5mllYJwUDGmKd6HhQB2oAlTERwQpHfFKFZpN6ybcDgGo18osBjJqwskWSzLhgOKAJRcyo21wG461LYhmmlIbKbOfyqON0kUhQB/WrFiys0yldrbDigaPt79jRWuf2Hvi4gIwrucf8AAa+BLcFbYYbb8zZ/Ov0A/YtiI/Yg+MDqBkb/AOVfAEQ22q4AO5nz+dAMmCYOVO5T1NZ2vRkabITjGRj8xWjGDbMMLvB6/Ss3Xpll0+THGCOPxFAjl6KKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoFFFAHQRriGM4z8o/lUqIpVsjHFFvE4gQlfl2ClAXZ82M46U1sNmrBbxvplhGAAz3agtntmvvL9urSV0r9kb4brvEpKIR+Qr4OtzusdPGAP9KXofcV94/t6bv+GS/hkO3lJyfoKlbj6H5+zD5QRzlear5VGBBz61K4ZEyo3ZFV1BxwM0ySUTbnAQ7V70yUFmChst3NNYkrjbjFOTcybsfMO3rQA5X3L6sOlROxGGIw1KWUYYcnuPSpw8bqu+gBiwoU3L070kaR5BJwuacSoX5ORSFB5e/v6UANMqgkbdwzwaUIrNknPsegoKsygYyDSYLYBXAHY0AOAUDL8jtQBucZGB2oZ9rgbSzelK5XAJoAdwMjGWpWViuCuGpizNkFalFyzrtYY5+9QBGjxCMhhl800qyKQTnNSYTZ5eBknhqEypZXHA6GgBm8BNrfdoEwDAfwe9KYyzFgwIFRSAE8dfagCUr5gJHGOlN+8cgDNIGwMj7wqVTFIjH7rigCAqwk+Y4Papo2VeSvPrUYbzPmYY7VPk7MADHv3oAikxcAleUH8NEKdQVHI79qEC7iB8r9l7GgRsXO/k46DtQBK8KqnHI/vUyNVRt+MCnDzUGVweOlSF55Ix+5G31oAhlRSwdfyo6EcdaVj8w+TBFLJblxkNx3oAHZYiCFzmomUmXjHAzUkI25D/ADUhA8wKo96AJVJIDlQOOadF5ZRiACKjYMT0HNG4qy7R9aAAgFdzjDdqarqBtxuJqaUAygY4NMQFWYEfN2oASZdgXzACe1NSTeeQM9qVoiTvkffjtSRtGxORtoAkEiR8/wAXcelTQj5cg7snpUARSc9aI5kUhVBX3oAkkjZcOUPB5HrUe7JxF8rdwasFpYl3eYXHoaYJEZgGhwSfvUARwlwCGAck9RUjhu/Pt6UFikvlp8qdc01iQDk7TmgASR0Hpz1pmSxPHekDyBDnDDPSlctG6nblT1b0oAFZWn2kfL3PpSvC28iNQM01UHzDse9B34OH24oAWQbAMD5h1xUfIOf4qQ5CnPJPekLEkEDJB4FADkCh9p4DdT6U6a32gFevbFKgLSfONq+vpTDJJGW7p0BoAV0RU+YfP3xTQgx8oGKekbkEjke9Iw7YwaAGgBR05qQKVYYXt1pm0uenA6mnpIyuVP3aAGp95gKcZth2gU0v5TEAZBpAM5I5GKAHl8fMfTpTEYMhCj8KFhXySS2W9KaQRtOMN2FAEgjWX/ZK9RSOwYqQOnamFT5m77vHSpAu7BAwaAHqBkMB9aRtvG04IpI5FVeOvOaWONWXcvXvQBKsisNzLQojQ7l4J7UELgbRubuKV4Vki+Q4IoAjBU8x/iacGyMsckUiHYgUqPf3pdgb7o3g9c0APxCMANljSKi4YuoJXlaRYgI2ZhgpyB60qbWUSg5ycFfSgCHeSu7A5OMelIx5AzhScH3qQFIZGUrlG71G8flAqvzI3f0oAXcJGAYAAHAp6IDuVuF7GoSNwAC4I7+tTM64VMYPr6UADqzrkyklei05JFSUbhhsc02YEKAo79aTy3OSRlvWgB+XUSeWeG7UkcRMbbhjHehCRlQOT+lA3RcsdyDqKAGoYwTg5B4qQRLHlQdynrUjNAynYu1jxTE+X5SMj1oARJBE+zHynuKHjZzgfdHI9zSz+W5AHyuOlORyBlR8w60ANm/dgZGGI5WnbSVzu2tj8qjmYtMpYZ/2qJPm4AGe/wBKAHIm0HGPMI6+tIYnUAkBXHO70pACGBxhezelSHewJJ3ADrQAzY8zgZGSPvUNgIYpe3IPrQcjaQOPSic79rBQSKAI9oYgqvNMYBWG4E+vtUyRCUF0kxJj7vrULiSMjeAff0oAcm4N8vIxQGO9T1qNZ2Y8DYMVJF+8A52r6+tACuWjkUH5kNPwp5I246D1pZAjSKq9B/FQdvmAOPoaAJYwZIyx7dKSIgcbQSfWkkHljZFznk0uUktwcYIPagCQS7UAKDg5yajdmPzqMue3pSo44Mg69hTU2k/Ic0AMmTZGGAy2eR6VKI4pod0fUHkUIrfM23d7CljUgkhNozyKAEBGNuB1pmzJORx6VLvh7Nls9KhQlJCW5GenpQA9Sm35G2nOCKRECsyuxx2Ip7tD1X79JO6uoCDa3fFACP5aN8pZm9TTlEYYtne3vTBKoUhlxj9ae0iFQEj59aAFYKoLEDcewpWBnh3HgqeB60qtEoyw+Y0LFL8zKAfYd6AERo5uFOJe4pGiaF8FRg96AqmXiPnuRTnLNnbkgdj3oABE8pKgDHrUYClyo5wKlBZYyA2wntSCMIPkxubqaAGsdgK7QFpyRiMYOMGkkEaNiQ8gZ4prAOcryAOOaAHzSboPLVAWz970pNnK7jlQPvUEx7cOdhIpBEpAVH+Qc0ADAM2QOQM5qMech5YBWHIqR1DgknBAyAO9Q24K5Djfu7+lAAGKEY49APWnqxLYkbLfoKasZjkJU7hjnNT+ZAEBC8+lABHGMEtxj0oUBsELhO/vSlC8TKoz3oRkljXopXjb60AKvlYeNRsY+veolBkG37jr0qfG0B2VTVWZRNOHRtrUAPy24eYu5vWkaKJCGLkH0FEeSwXOH/vU9oGhwSoYE0ARN+7TcvQ1NBEblMjG0HmmSxhSNnIPUelOiVoztXgMaAHi6SH5THvYGoxDvZpYgGJ6oakUY4ZQBnr61GrBZHEZ2E/xUANZCBhlwM08g5CSAlM9qEYrlJBuyeDQjsGbBzjpQBIwjdSsCkEHnNJHGIywIwD1amON3zLwR1p7yvOg+T5QelABIofJCgKOme5pyCJgXdirL1A9KYMmXbIhVP73pTIk/evkb4xzgUAOUjcZYsbR606RjncAPdqVIzI7FcbOpWmSHe20cIO1ADnQFd6kVGsyh/ubXNAJibJXI6U4w4DMF3KwyW9KAJFHlMMAEt1NTSbR0AOB1qJHQxlPbk1O5iEARTu9zQA1Ck2AOHHSrNghWSZsZfYefSqqARN6571f0p9rzqMEFD1oGj7h/YxkMf7DfxfKjqzg/wDfNfAVsV+xJlctub8Oa/QD9ipQ/wCw98YFOPvyf+g18A2y7bUPwcFx+tAMmRHQAEZXrWV4iRTp7OBjkD9RWj57YAUcd89azvELb9OY9gR/MUCOWooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACjpRQKAOni8w26ZfjYOKYyErnqcUluS8Cj0QfypxHHvincDWs4z9gsRt5N0vf3r7u/bzjkX9lL4ablwvlL/IV8LWTEwacdv8Ay9Lx+Ir7u/b8ut37Knw0hCYzChz+AqepXQ/PaeXaqBeu2oWY/Kd201NLE0ZC4+Zl61GSi/K659TTJGFtrDzBk9qf83PljqOab5PP3sr2oUE5xJjFAAqNGAcZzwadEseeTtPpTP3kTBt3WpMAMN4zu6N6UANYruIVcYp6/ukWQjcueaR08oqpHyn+Kn/KF2r8ynqKAFwCwMZxupkgdep696EAVCmNvPX0pVm2jaw3DOM0AI2fLAxk560kke5Qu3mpiqgYU59KWPEi4k+THf1oArpwAv3WB4qw7tgEqAc9fWotkOSSScHgilk3PgjkZ+7QApVCTuGCaaJecPhlH8NSRqXJDDbjnNQFcSMycE0ATSeTsLKpBPak8vEJYZqQcRFWUEt3qKOQpuU/OtACjGAdoI70pj+YqVyPWmlkXJXkn+Gnq5gBx37+lABJEMblGT/cpkWYjyMA9vSnkPuLKd3HWkymQXO4nr9KAGshMpKr5hA4NPkDsAQAj4708hsgRYjQ9TTXyhw3znHXNAArMuN/BI60K0hOFmIUdFx1pNrEY25HrUiMoxlRketAA4aMgs2Q3U0xmCJhck0su6ZtvQCmMzRryuD2FACoB90p856UpQH2YDOaFllbqvbg0xMq4/iDdTQAqwsGBkbcD0p5wAcjA/nSFS5wOQO/pTom3g70yB09qAIo2kEoUjA6inRKzSM7jFP8lOTv3HqKIxvGX4PagAxsbHXNIkbEEkDb+tO/iwOaYQ4HTigBWQDG3pUx2yx/dAIqPHCkDH0pxLBwysD6igB67mwVH51GrYLKflA5NStcFgABtzwaiHlr/rCSwPBoAbuI2nG5c8ileQO/ynkHpSMiOxEbEZ7H1oePK8DDDqR3oACYhknO79KbvyMEZQ07axjDGMsAelMMb9VUnPagAMgVtigEH1qVo2VflAY+tMWBzwYiM1OkUxwvRR29aAIGUMP3X3v4g1MGVOMYB6n0q9LbSSIWEZTHamLZyuuSny91NAECxs+Mn5R0HrTi4Dk4BJ7elSm2kHT5SOh9KDbHBBAye9OwETeaVJBxxUag/d796srFlGEnyqvAOetZ8m9WPYdj60gJQGbKqcADmnY43t0AqMOZAWwBgdPWgK8nTkelACrhm45BFGwhgP0p3kMzbtpUAU5lLOrbcYGcUAR7d+QBgClVCVKryKmWMPkldopo5yU4x1HrQBGhTcBJ26VJuMhKjgYoEPzBtuc8UrxCJh2AoAcnlBQXXBFNCoHyvftSApNLnbn0oK5b+6RQA/7rdMZ7014ZCMo272oWNnYYfPtTT5lucoMmgB6yZQ712OO9NjbIKngHoaWQM6h9uQeqiogfMQgcgUATDBG1zwe9NMbRj5Dkd6bjzcLnYKWPKkA9c8H1oAbg43jkA8j0p2xGOY2yc8il2jfk8HPK+tNaFS26M7X9KAHDczYUbTml+baVZcc9TTcFgCH+YGlUmQkSNkigAaQRHGCw96ldjxs+VO59KiLBVZSu7PIp6MkUWAN27qPSgB55jwD83rnrU/lKYRt+93qBmjKAIMAdRTnfzY9qnYFoABEYlIYcHvTUfB6fSmFyVw3IHPFPCqSSxAU+nagBVhDA8jefXtTZVaKRURgwPrTHiV8uD07E9aUqRIu4Yz1PpQA4J5bsCA5xR8mCCMPjnHpSnG8gjAx1pAAnGMn1oAEYyxeWo2gdKeofHuvY01AUIB4btSMpkbDNhsYBoAVC0rFmXYAO9N3xO/ysRgciiYsQIpX+QD7woBWMqcfu/WgAbYSChxx+VDRkRhT8+e9JhfN+7mPHJpzbgCq/KpHFAESKvHmnjtT1i8wgJwtJDC0vAXBH61aFs2AMBD9etADGj+6QMAVLdAMEYLtIpyxsRuJXI/hzSPEzMMbTntnpQAKV2BccddxqJMySBsYxxildXQ4UZFRbWQDbxz19adgJrgKqhEAJJ5poKvGAwC4Pal2AgEkZ+tCW5GTlfzpANGWG4OV7YqQBnXZIQMn7wpgUxkupX86fsRIizn5ieKAIxFDvK8hl5APeo5oz5g8thluSfSnndlZGG4g0snkkZXnJ5HpQAxQ2AoXHP3qfIVJCMu0n+KmMXePltig4p3ndFlXegOM+lAEu3gBow6g8HvSvIGUqn3s/lTVjXcVWU4boT2pHj2gqVyfX1oAdKgaJQRyOc09A8aBw+d3FQAybGDDIXmnkqFBcY9BQBLh1OUYMf71NkuJNpDfK/r60xASu5G2g8UvltGdrHep70ANRt5IP3+p+lK5x/q+hprRMJM7cAd/WnhFUMzHaD2oAa+OVdflxyaURBAHjHy0xn3I46jHBp7W7C0DhsHsKAEd4c4dTtI5xTisbKPLO0DpUO4sACOT2p8eQeR0oAejFDgLkkc01doIVflLdQacgl3B0wCPWklkxMPMQEMMEigB3ygYByw7+tNUbGD7Rk8EelNlEa4ZPpg0752wu3rQA+GQLKFTqecnpSfZn88yCPYO+e9JhTGA68L6dqeZmuEAd8KvQigCKTCNy5ZDTdiuoCsGx6mkWQxEkKGXvS4gdB5atG/8AOgBGCsVwOQeR6UuXeQF+namYdJASmM9R61O8USgMWz/s0ARHMZxgkmpPNwgG3HNLL5rmN41G0dqSSUyMC0e0g9PWgBVZ5Pv89gKWRCEKMm3byDTPLaU4jfa1SeVOFHmsHweBQAgO8BhgEHvSSHZklSD7UjmOMlw5LDsO1KoYoHDBwTzmgBuDtHAPsKlDHaNikNnn2pUEG85bDfpTEWdN0iOrYPI9KAJdzF8OTIewNK8yR5bbhxxgU0PI+ckZ7mmvuK4HX19aAIyu3MqPtJ6ihj5aFiuQaU7GyTwe49Kj3Ki7DJmM9fagCZTugO/hexpPPcRbE+5QYy2DuDoBxzTlWPBypA9BQARNk4jXce9TRkbihABIqCKMs5MDBPrUn2aRHyw3ZHrQBZhgCcMMg/xelS2g2XEvlkFdlVxDLtIAyuOlWdPQyCVtuAE5oGj7i/Yv8xf2Ivi6ViLAu4JH+7XwPb7TaqOpDvz6c196/sgW+tJ+w/8AFOSyZPs5lfcD1xtNfBdsUjtELLu5bd9c0Aydog5DN8hHSsrxAgGnSGM/JkZ+uRWgd85BJ2r2FZuvhl05+y5H8xQI5eiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKO1FA600B0kaAxREf3BTn3D7gzxTIDsiQYyCg/lSl8EgDjFFtR9DTsA7jTl+6zXK8+nNfdX7eej3Vr+zH8NriS4WSHyUGz3xXwpZHy7ezk2/duVI/Ovub9uzWvtP7Lvw0tmTDPCpz+FQ/iKWx8FSF3KjrkDFQyFQDlcmprgCMKVHYVEAc7gM561RBBEcqV/h7n0qZAiqSR06U0hVIA5J7U5oHeRdqZH92gBrRlCGPO7t6VINsYKuM56UTh4CuR/9akC9yck0AEc5iTZjcp7mjzUZPl4we9MQFWwwwtOCsSG28enrQBIBvHJ5pWiTYMnnNMK7FG316UigsDQAK6glD07GpAwYbeoBpFZCNhG3/bpFcIcDgevrQBNBIsYZGUFc5yaUoRlgMZPAFV3Ib+efWrKESRE45A/KgCFC7Od3JFJtjOednrT7ZhiQH5mI4b0pkSgxncu455NADhGGQZfvSCMDcAcj2ocryHyfSlEUYjLI2WHY0AAjCA7l60YVlIPC+vpSo32iP7v3aFVAjbuPagBskRCgIeDSsq7cd/WlK4Q7fungUzouG4HrQA5DglTypp8kZc7Su5MdaYqfMdp3DH51IG2YAOCe1AEbSy23y7flI705QswyB83cUYZJQrDeTTp0VWHlDBxyKAGzEgAIuHApcmRQG5bHQd6cOuCPmIpZbUxKokU7j0OelAEbJKkYZRlRyaeBvTzOmeMUK7+WUHA7j1pmJEHpkUANC+U3HzZqRi0TKpOA1Juwoz1FMeRhgtyTQA8oI2AA3j1p6oS2McU2L5gfyzUkeFTYRluxoAjDLkKV+YHrTpHd5AGH7vt70ZBIYLgjg0kg2EY+bPagAjcqxBXaop4EW4AZqDzHWXPX1FHmoT8p59KAJ5FjQb+SOhFMjnUEL5eSTxupAWUhsZJ4p0hLqFYYIPX0oAR3DZUDAz970p8UZlJVhkBSRTFyy7cbVzyfWrFjL5F0spXeqnp60AfQv7Of7HZ/aE8KXmsJ4zsNBe2k8v7PdShWPv1r1y0/4JXX1yzbfiRpHXgrMDn9a+M4NREEsj2Oo3enbzlo4JCoPuea2dO8Q+IYXB07xVfLIDwGnbn9aAPsaH/gk7qjDn4iafn2bIq3H/wST1Mjf/wsSwAHfP8A9evnHwT+0d408K38UGralc3doGAZvMJ4r7X8D+MR4y8ORajFe3H2aZQzHzDxQCR5oP8Agk/ftJsX4k2Dt6A//XqZP+CTFyWIf4k2W70zXodzazx3DXVjqdwXY8RmQinLqd4srxT3VwjYz5nmGgdjgYv+CRvmMxf4k2YUdSDVlP8Agkfpqj978Tbf8P8A9daXxE+LafDzwzdaxcS3ckUPAUMfnNfKOu/tm+LdTvJJNNiuYbY8qGkNO4WPpqf/AIJL+H14k+KMKnHRh/8AXqlN/wAEm9Bdfk+Kdlx6/wD66+U9R/ab+IOqgl7iZUA6eYax7n4+eOpBuF7cAAdPNNIND6/j/wCCSmlNyPihYsPYE0S/8EkbUA/ZviZaMfQoa+U9K/al8baJIrSPcSoByC5NeveBf29ZFlRNZt5h2HzEUNjSO51H/gkl4qjiLab4zs7444UcGvKvGn/BOD42eEQ8lrpS6xbqMlo2Gfyr7G+GX7RvhbxvaI1prElpfkcI0te6+EfibrmmzxxNejULSQZ35ztHvSTCx+JHinwF4l8E3j2viHQL7TpUOP8AUsV496x0hWUK0TIwHbPI/Cv3k8TeKvBvimCXT/FWgWd1ayqd04jUnFfNnxQ/4J0fDH4r2U2o/DXVY9E1UoX+zs4Ac+mKYmj8rJEYtlflb1qB5QgLMuSBjHrXoXxl+CXi74F+I20nxbpk1soJ8m72kpIO3PQV5+UUvnAYEcEdKBEKndgsNmeRSbiz8/e7VI8BDLxxUTriTOOO9ACxsFcccdzUpDS8bto7VXfJYEjA9aepDLjGQO9AD0EqrtPC/wA6DFswU79RSBJMBpDtQdKUyM23AoAPK3AbByDzTX8wjgYxU5fdH8h2tmmqzsPm5I6UAQqjSDldrZpywDO3dg/3qcdqLkfezQEyMtz7elADfLI7YGevrSffJ4xjv6058sAuMjNCMEYiRcAc80ANZy64A5/lUifIQPL4NMKxu52D5eualEk+3gZFABKSTwNrdqdIFjTceJO49qYZG7pk+vpTyGZdxXOKAGxSKVZCMq3empG1uxx88ZpTtIJA2n09aAsign9KAH5AlV/0o2eY7EcDt9ajwuCD070bjkMh5HSgCQzqhyyHjikEm4skkZAI4anLK0RLOA2aFuHkkywynp6UAMZWRhkbhShtxxjNRSNJ5hcHK+lODPuV1GM8YoAkbAfayZBFNTEalGG4GkclpgjDBIpXVRjccEdKAF2hsD+DHX0q1aRrdSpEwwTwD6+gquiqFz27rU0GIpI2wTtORQB9Cfs6fsyeDvjJbagfEvxDg8IajbttjtZhjePXNe+2f/BMXwJe7DF8ZbCYH0df8a+F2ayv5hJLczQTAY3RMQalhhlJIs9fvY5OwM5/xoA/QK2/4JJ+Gbq33x/EqGcY4kjIb+Rqnef8ElvDtum5vifFCB1LLn+tfGvw6/aO8dfCLXoA2q3N3ZK4DwySFsrmv0h+H/iy3+Ifg+y12IySw3UYaUbvunvSuOx5Daf8EmNEvF/cfFKKUDusf/16nT/gkjohba/xPhPsF/8Ar17x4QXet3Z27NFySpJ61dh0eSAJPLLL5qscjf1ouOx4JH/wSL8NxIWk+JCEf3tv/wBemSf8EkPDMUPmH4lQxR55aYhP5mvZviz44tPhB4AuPFOtSPHjItLdnwZm9BX5rfEb4/8Ajf4o6hNd3Wtz6dp8hPlW0MhTavamJo+rz/wTC+HtvcJbyfF6wWVjgIZV5P518gftBfB+3+B/xRl8K2erxa5bJH5q3cJyp9s1xUtxceYHl1i/Z85D+e2apX1zLcSCWa4lvJBwHlfc2PSgRGkYIKj5ZG/IVEU+zZ3R49/WjO9cHkZ6elOErADzV3IDgZoAMJIwfdsB421HEvlSMWGAOR706aCJj8uc9cURRLtLynHoKAG581SwQKM8pUomJjCJgNnODUaYV2zyT0NByhI29aAF5EhJPzEcCptkKxAyffzwPSokOH5XJHelaMsS56+lAABsOQcqeo9KAV3Y54pF4fhevWlnjEZUA5FAA8zIePu0MuIy4O9e4pGTacBd2aFjdgdny4H3KAGbhtJUc/3aRWLH721ccg1LsjZCH/dzf3vaog0Sn5ucDg0AOZl3rk4bFMbzZGyT90ce9KdiZLDPFCHcu5eAP1oAtbi8CyjgjgilUKQPM5U9T6VFG5DHK9R0p8LkbhwwxQA3zYXkxtwB0NLIskZBAzGOd1SskNwmGXy2A4aoHV2Aj3fKOfrQAGRGG88E8YqIrnKDgdqUKGbJTg8GpPJUnah5HQ0AQurqAcbcVNHPvUZRdo796Yx87qnA7etOMKtGCpxjtQArSsRnHHamxAh8kbmPQelMWTB3bSwPGPSpbUncSnzNQBJtcrtPyAnrmhplMXlsnIOA3rUbiQ/f6NxihQxPTBXjHrQBGVaJQw9aeZN4BHUGpEZXGFXLZ5HpQscByd2GHagCNW2Od6ZFAeCOXHKI3U+lNlTOW6j09aVQuQZMeUeCO9AD3KAbF5GeG9aZGVYuFfBFEtyHPlRp8o6NTkWNwQRsA5J9aAIkVk3HOSeoqyk5VNoUc1FtUyhlO7PAqN0eKXLAkUAWEA5Dfiwp0z26n5Yjn17VDHgtuIJA6Zq1uZ0O1FA7Ed6AIUjMjbBwp6nNTlETCxrn1pqPCcnH7wdVpgZkkY7TkjtQBKVCuDyuan8p2yFO5SOSagD5jIVeT1zT45XRCF+XigB8cU1qCS+5fSr2nzCRJ228bDmsuNWLFmbPtVixJVpii5XbyPSgaPur9j6aeL9hj4sNEDgO+efavgiz2NZIrp95nOfxr78/YuTP7CnxgJ6bnx/3zX5+2TH7IE7FmOfTmgGWD5UmADtx0rN8Sb/7NYEfKCOfxFX8h/4fpWb4glcac0Tcjg5/EUCOWooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACjPFFHamgOiijKwpkdVH8qURliTjOKSGYmGNcfwjn8KftPUdMUdQWxp26N9mtMKf9evA+tfbn7ctsq/sv/DKQoRIIV5P0r4kikK21m6rnbOvH4ivt79t+e4u/2VfhpJNF5aiJQPyqH8Ra2Pgy5BRlJG4kDgduKiYOOvANTTf6wZ44HHrTWjBHLZPpVEAqB/mHVak3SMwZWxio0PJXO3NO4OM5wKAHMzSfeG41C2YwFxuNSKVU5FMAx8xoAcEAALUMxxx0pm/jBHXpSDepII4oAlTJByOlMIIPHem78nb+tSeUMYLYbtQAyPIO1hUjBchQOvUU0IQM456AUZGPu4I60ADfMu3HI6U9B5YwOjdjSSR4UMOWqRjGyhpBhhQAJCmeuzB/OkYbXJUcelNIL/P1Ap28KhBBAbpQA3HzEFM5pSOMY4Hf1pB90o+V776c/CDA3L3oABthUsgxntRtVoixPzHtTYcNlMcmlVQVYEcj0oAMIBlwc9sU5drncR0/hoEm/CDH+8aY2Ek2kEcdRQA8BY5MsNoPf0pwiXc2BlcdaZKsZIBbOOakR1IO04OOlACpH5gGeCOjUqmORWwMEdR60Ha4O75BimKighkBPuaAHPj5Qfmb+E0EyFiJDuGKjyVc7gWJ6VNHKoU7clgOc0AR5hY4D4IFOCvBCTneT0pvlKh+ZMlx2oMEh4QkqBzQAzaQRuPXj6USCOFfL37ie9MV+OnzelACoykJkt1oAdA0MyhSdpGakkj2oqK+71qHaIJhlchqlW3BfCnGeaAGkKuOzehqb76gsMOOlRrgttlHze1IwZ/kPAHSgAMJRgTySaijjAkJ24FPUMUKA5pY03rg8YoAM7QWbODxn0pRMrqF6YPX1pdqxAc7uelKm0MCy8Z6UAI8wDAfwVJg7Puk0MqPlCu3PQ+lN3tEvlgZGaAJ4RuTKrhhUyDaSYwVf0FVhIoX5gcjmpFnJ+jcH1oA2vtEt/prh/meP88V9jfsgay134GubZ8yLEeAf5V8X6e2xZU/2T1r6i/ZGluV0u8ktVJZH+Ze1S2NH0ZcalY22o7bhjCZDwR0WpLyGU2sjLiW27MOpqvcTaZqsM0V2ohn4y5GMVnQXx0aKUNMbixJwDnpSL6FD9t3RrfRv2UdKljt1WS5lGZCOa/ONJJYbeNMcFetfpr+3uRN+yH4blHCGbIr8zHCm3iDA42DmrSIbFFy6rkMQveo2ndpOWPShFjdWCtz6VHHHtyh/OnYRKbiYHcDnA71J9thniK3VsJP9oDGKjRMghj0FNVRISR8oA5pNAtDX0W4vfD0y3+j3T7F5KKTxX17+zh+1O5lh0vVZS5cBWLHpXxjbTPaSCSI4UdV9a0I5306ePVLHKMG/eKh6Ggq5+qHiGxm1BF1bRbsvZlcyRFsg1n6TqepaVPDf6fcSRyjqqNjBryX9m74vjxP4bW1kkJlhG0oT1r1gsum3vnKpK3HGOymgZ62NS8JftHeFJ/Bfj6ygkupUKQ3bqAyt0BBNflz+1T+zFrf7MfjZrK7V7jw5duTY3qj5cE8DP0xX3tParcotxZ7oLuDkODjca9E1jw3pX7WHwS1fwTr8KvrtnCXtJGGX3ryMflQJo/FqWQhAB98frURIYAhcHvWx4k8I33grxRqmgalE0V9p8rRurdSAcCsmSJkOACR3oJC3jEjHdwtI6oDhBwOc0Rt8vCmo9+BgigCUh5o8MPl6imBxt2gYxRklfLXjPf0o8vdxnpQAp/d4OOO49abOSrDA+U0sZXBUHcTT4WWSOSJ+GXke9AEakdX+XFSLCr48tt2etQhiWBK8A4qxGiFwEPJNAAkWJCNpDevrTQrFyXXkdh3qyis8uxucc0y5kjSQKQ5kJwEQZJoAhlgBBKE57inGVlUBchvevTPh3+zV8Ufisgfw54UupLZz/x8uhCj869os/8Agmj8T/s6yanqlnpjHnZLIB/WgD5KJbad/wA/sKUzFowADnsDX1vd/wDBNb4n+TJJpuq2Woso+7HKCT7Yrw/4l/s5fEr4R75PE/hS7t7QNgXaIWU0AecKdhyVwT3p2RLnzH2gdKajxyqQh3beCpGCKVkXhtuPrQAjLHGTtbLHoKdsjCktw2KURxKTxhz0qG5jAUfxe9AAAWBPXipI3yoB4puQiFtpxiowd5D420ATSFFPB600qyjA/OmoQ56c1OkZVDn5v6UAQz5baT94dGoPJDdT3FLlmVkXG3HINORsLnYVIHegBFzyqjOetTICowvTvSRzRiEmMYkNMWJ4iHB3E9qALMeAMOPoaeqCXJDbSOmPWqZuGYhWXcPWrUD5IwOB3oAu6jbNf6MZH5liHfqa++f2Cdc/tf4WzadI5cxZ5PavgmCZTaXIOSNtfc3/AAT10s3Hgm9mQMFDndg0mi0z6Y06yNrcwlFKeWxz6mursrX+1NasYxGwEzBWQ+ma8/8AH9jrtne2t7pR3W6/fT6V3Xwr1yfXvFek/bbcwSjqOxqeo3sfGX/BVTxq0vj/AMO+B7dilnYRCdoweMn1r4fud2doHyjhR6V9Kf8ABSrUGuP2udYDAkQ2sYUelfMzzCZRtO04rSxFysCWfbICSKd/qzkjC5odWf7p+YdqjTdJ94dDSEO4icPGCQTyaGleRyJPmQ+lOKKincCozmmSIhQBVJyc0AIhCnGPmHSiQF23AZVeSPWkZCozgnHepMbRkDk0AIACMdN3OfShZX5AGQO9Ii5yufmFOQBXLbflPUetACyyFsIn4inkegwB1qEoY2JAyD+lTFI1GVfMnoaAFCDht3I6j1pG8qRCd3XjHpSNuUbtuMdfekXyjk457pQA2OYIxSM5U8bvSnFZATlssBwPWlESq3C/Kf0ocrE2M5btQAkpZlGcbqR0OdoXIx1pzDaTwS7CkWRk/d8ZPrQAwHdER0QdaSLCkYGB6elOcMGyRkkY4p+XVCPLzxywoAafmztO4YqZIwwBJ8sD9agLhk3qpX/ZHenIoPLq20jvQBNkSShFG5T1PpSXCLARtbefX0oZECBgdmOnvSKys4Z/lY8UAN2mbaOhPQ1Iv7xSNmNvFVG+WQxxtuXPX0qdZpTESF6cbqABVMrcLjHb1o8va4VDkHqKjKyEKd+PepYE8w7CNp/vetADVEjSbFwh9TRCm2Tj92R1PrTvKKnbyzUSgLt5570AI6EHfyQeDSbslU9+GNOZvL5UEk/lUDSy5wVxGeuKAJYgDIQ2QV5PvSSLG8gki+ds/dpwZ/L+VMD0PWhQm04+U5oAa+FG4Nz1IpgxGfMPzKe1OeHaRLt3e3rTmIK8D5ifu0AOHQjZgNzn0oO112AfLnk0hjk8s5OSO1OjQbT656UAEUYicof9Wf4venHcc7OVXqPWkd85G3G3v609cNuONpOKAGmR3jJX8UpEYPhUJWT0NPxBG5OT5nr60jRLIQ8Y2nuT3oAia0O5hz5hHUVJG00C7VIJPrU2UVdyku69qYgidm5ILDNABGXRvmbLe1OkkZXxkMhHJFRs2D8ozt705iA28HgjlaAJliDnckgHH3T3q1p8RxOS3zlcYqinkrkg/M3GKu6cwR5QQc7DzQNH3V+xhLGn7DPxejYY+Z//AEGvgCGE/Y0PJGXP61+gf7GNmj/sK/F6Urk7pBn8DX5/WxIsY8DkM386AYYVhuHy1m6+S9g52nAI5/EVrmdH+/Hhzxj1rL8QSv8A2a8ZxtyP5igRytFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUCigU0B0cSN5KFRkbBTwQgPHJHSooJHiRCp4KDj8KesmQSwO4jijqBqW3EOnFwQftS5+ma+8/wBvILJ+yj8NXRSFEagD8K+FLOFJLDT+rE3SjdnpzX3d+3nD5X7KHw3XP3Y0/lUNe8X0Pz2li/eqGHzEDB9KFRSxUjkd6lumEroGyGA6+tRFMnBBHv61RA0x+ZGzKOVprZ8gHFPb90mwZBam7T5eznmgBZUMcaMPunrQrDOe1RhmYeW3QUKgB2E7R2NAC52EN1U9qkzhfnPye9Jt8sjA3kUBTKN0mQe2KAG7FZQB+FKu2Y7WyCO/rSJCCCz5QA8Ugi3jIJOT0oAlP7oe1MB3AE8DNDuzKCy/MpwBT41LKRjluAPSgA2DdhX69R7U9zuA+XcAeDUbCGJdpOH/AL1OV2SIso4zxQAB9rEn8qGLyrvGBg4ApEYTje4wQaHCrJ8rHB7UAPkWcjDKCKVJZI+CABT4pcowkP0qPd50bM2SVPFAEiSidGTHln+9VfcUbae3Q+tSbVHO7Ge1AkErbVGHHSgBqgSZz8tKWCvkNuIpMs7lHTGOfrStjllGDjG2gBCd2QU5NHlqhIPXHWn5IXPekVRuIYHcRxQAu0SgKF3Ec05A0YGBgHtUZUCQAEq1OHmJKC4J+tAEkW77QAy4GKYwCSvjv3p8zsz5B5x970pjuuwc/N3oAI921gpGKZloSJE3KPQ0rRiNlbfheuPWnvOsrDywXBGCPSgCN/nO/YdzcVHyOFzxUm50G0ZI96amYTh+d1AC8ja7c08qFHJ+XqT6UwBc7VYmnRqoyWPHcetACs24KXGf7vrSNhVPf1zQrKJNgJ57+lKGCBgwbjoaAGKp3gDqangKTEq+cA9ai2ARFzwe1SIS0IAXgnrQABU3MEG4Z60m1I+WbBzQOQq/dAPWkuFU7cDcAeaADaGB+bK+tOEiyDawPB4NN2g/eyo9PWnqik4U7uePWgBFAB3feOeKMLvPVc9aSKJRI3mlh6U4xkuGwdmadgLFhMUMic+Wf4q+uf2LZ4xo+qQ4Kyl9yua+RbFVWdyoJGPu19Wfsb2udP1UfMX8zK+1Sxo+i9UVL4s5iG6LGcfxVVutIt7rSrm4iRlBxuhz0PrT7t1SWVsOrx4yn94Vp+RCLdJ45DGbkY2k8UrF3MH9uclP2OfDCsTnzuAfSvzSLnyogRkbBX6d/wDBQC0W3/ZL8NJnLGXjHSvzJMOIYgTj5BVGYwlACwXDYpEBf7v40xkOW3ZIHelDcblOMdjQBKdrxnnBHWhVBXIPSoEYAlSDk1IhwMfdFAFiIBjuLVPZzrHM0TZaGUY/GqbqH5Q/KByKYr8oRnCmgaPWv2bPFLeH/iCtmzlY2YADP3q+6LnUkjuo45cmKdchj2Nfmz4RvH03x7pN0hwS4r9MNPk0jxD4LtPNBW9eMfMPpQU2PvL42FrFboSWY5WQdK3/AAP4kuPCXi3StYjYo4kVJF7Mua5MaRLqWhyWwdkubc7o89WAqrpHiVoojZ38LJcKco30pWQXPFf+CoHwyh8HfF3S/GFkm228RxiWTb93dxXx3I4IABGa/SP/AIKW2Ees/s5eCtXZd0sIADt1Ffmg4LeW3P3Aee/FMgeZXQAquBULNv5ccUpZpAOMUm8IAMbiaABRvACHp2pCyuNv3T6+tKm0AZ+U0kqEyBuSvQUAKsQxtHXPX1prA537TlTTkTa3Vs+p7U13YMGbkenrQBIAsuDtxuqeGEMNn8QPWoggI4PWrCzeXGMD5h0I7+1AHSeCPAmp+PvEVvoekQl7qZgDKORGPU1+iHwd/Zo+GvwL0yDUta08eJvFRGdsvzoG9MV53+yB8Nl8OeGDrjwYv7vlJW/hWvpfT4La10PxD4o1BP8ARdGtndN3RpAOP5UDR8/ftL/ty+L/AAaP+Eb8KLaaBM648mzQRmIdMHFfFWvfFXx94qvJLrVfGmqNKzZKi5YAZ/GsTxh4tuPH/jLWfEN25aS8nZo8ngLnjH4VlhVyQ2fSnYR2mi/Eb4g+HJRdaX4u1WR0YNt+1Nj+dfWPwJ/4KJa4tzB4Z+JVrba/pUxEQa9QOVPTqa+Io7iS0KyW8hRk5C54P1qTVLtNZszOqiO7j5Ozg5HQikCR+ifxy/Yl8EfHjQ7vxf8ACkDRvESx/aJtMB+WYYz8or84dd0HUvC+s3ekavbSWepWjlJYZOvB619+fsU/F7UbjwfbXAnLX+lOFZieWX0PtxS/8FLPhBpmt+GtF+LOgWiwXEuF1Hyl+Un3oG0fnvkNnjP86cieYpTBB7ZpoAOHwcuMjHQUkm5zyxBXuO9AiVBLGxRiCKVt7H7gIFRyuxUZP40rPhVAYnNAEgAxu24b2qTapj3I3PcVEPmUHBNRyRlWJXIBFACozbjtXBApx3sf3o4x2pgVmAzndjin+c7ARthe+40ARBUXlMvnrjtSu5hwc5B7U4h4yDGmG/Q0NJvYOyc9MCgBwBbG7hSOlKs6RsFTle9NGFGN3zdjSRqjjkbD/OgC9byKIbgLlhtOcfSvtj/gn9qN9b+C9Qa1Y+SGO4V8PWjFUnUDOUJP5V93f8E+ALf4dahIEJ3Oc0DR9Q/2rNrNs0abkYcbfU1r/BXUrtfiNaWd8mQpPltWNNbpb3KSwqwRvvAdq9B+HukRS+NNIu4zkrnJHapvqW9j8xv+Ci6Y/a98RDG7/R0r5sCqijHPPNfTP/BQ9w/7YHiHjGLZAa+aZ4oRyrHPpVszInRc5WT8fSmsuJFI55prY3bguB0x60qHFIB8sbNz1XNMCNyqjBPApwlZMHBYE9KkJ2jj5mY9+1ADHEjKsa/wnOaa8bquWPemqxLFVOwjndRIGb5znA/WgBEw0jKflbP3qleNol5Oec4prLFMhYkxN9aBEGH+sz/M0AIJJFBOCyntUsbgkgr1/io8rCkg7fXNIpXyio6d6AHNM0xICnA70KFXIC/Meh9KjJaL5h930p29QhOfmPQelADypTLg4x1HrUIMasx3ZLDvTx82d+TjnjvQNpOPK4bg+1ACDLgIDluzVMHjDBJBtkxwfWowoglGBkDp7VIFBdiy5JHBoAV0OOfSo45XUEKcjFEkp3BTnaBimsoGCoJQ9TQAjKCwI64qctI6BdwKY5xURUquFO4kdaSLBOOaAF2iMHIJX09KVowygMvB5z6UjMWbKjKgcj1pCJJVBU8DtQAm1GIP3QvWl2Yy5kwOwpVYbQxXgDp60LsAYhN+7oKAGlNnLEsp6U15nCqEHAqRI2hiY5yT0BpFPmKu4Mh9qAHo7SBcAhh1PY0ufMcFlyR2qMTSswVThB+dPR/myvDd80AKsW5sZ3D0pwcoeAFOejUzAJBiby5M9+9N3NuDuScHGRQBPuKnod5PaoeMkkbOae0Y3L8zAk+tEoMLKSDIM9qAGl3jYBTkmkLoxO5D5n94UspDkOVwvcGlaFXwudgPegBEbdyqnIPI9acsYYsCce/pSgpEhVMs1JHKg+VuB3PrQA4kqoMfzdqVmMpG8bMdalWFXBaE5J4x6VVdW3jzM5FAE5IClHT5D/F6VHjyziPn+7UmdkZ8wEx+tNktgsRljbKHjHcUAR/PDJnbtk9anknaZcLGqEdc96Z5knk5wHRfzofa6hnBU9jnigBRNKhVdu4Hp9aDId5EijdjqKXyQFIUkj+8KXYvkl1BOPWgCTZCq7sYzxTrFyDMuOi9apqflO7LDt7VZ05sLOvU7ODQB95fsgzSW/7BXxZdTwZXH6V8AWDlrZUbP3mOfTmvvn9kyb/jAb4qIAc+e3T6V8DWSk2ajJBLNg/jQBNKROx4wQMZ9azNdJbTHDcFSB9eRWk5jB+YlCKzfED7tPIxnkc/jQBy9FFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUCigU0BvxqDEnzc7Rx+FKQ2DjkUQKuyM4/hH8qfvIzhaL6gbdkGNnYhAQRcKcevNfcP7dTTt+y58NTISUMS/yr4fsJvksH5CLcLkfiK+4P25rxLn9lf4clGJwicfhUN+8X0Pgm5KsynpgUx2LBU79QadIAWUuMDbzmo8cjac5qiA3pjDHDDOD60xJAyKGJz6iiQB24XcF7VHxgsv4igCV0DHG78aV0VlCn5sd6YBu2kLihwVUAcUAOUrAAckkc0Z3jeQSCfypUVMjcCfehn2sAvzL3FACAFwVOS3akVyCFCnINSCHcA27GD2piswZioII9e9ADmnaQ5IAI9aRt6gMp5zUUvzgMAQQeadHkgMDyD0oAlykinK7m9KbG2zIOQM9fSgZ3lwOR1pu5yxZh8vdaAJQAzH+FfUd6DCnZsY/ipAu7g52noKF2qSTnb6GgB6tkbV+ahjjJCnI61GoVSWB2+hHapCyKN4Zs/zoAFVHBJyD/Ko2iLSZBKY7058tlgMDuKaXEnyvkL60AEkkgYKTj39aeOWx39aaxAPOWHagF5M44xQBJJG8I3/eHcVGG8zIYbWxwTToleMtvO5WGKJFYHI+YAdKACVyMKpDEjr3pcyBNrNx+tRYVTnB2txx2pQNp8tskY4JoAkzleTtXGKQkxgZ+Ze59qSRN8GwfeFNgVWDMWKkDGD3oAejIwORlex9KfCVwSqlfpUMUvmHYy7QOlPI5AOQfagBsb5YhwQTTwVHykZY9Ka67gAQS3Y0Ky55+8KAHlVCnA2mkjbewzxQJfNbbjpTo1IPPagBkjr/AHSCPSniTegQKTnuajXc5Jxux29KkVGGGdtg9KABMoQkg3dxUiRFXx5gUHkLSqkZYnzM+/pTPMSRgWHI4zQA2US7vnXilSNnHynb9ae7SR4KkOD2NMlxIPvYf0oARgFIDkSc9qag+c5JQZ4JpQkaKCzHI5wKRm8wcgkds00A4gM/zfMo71IGaM8HcD/DUcQH3WJVSakNs0GWLb09utNgXdKhLSuepx0r61/Y8jcWV+sB2OzdG/nXyfo5Ll89Mda+r/2R51t4b1X3YJ+96VDGj3nxDcXenSlru3V7fODKg61Ua+heydfnkjbmNgfuV1V1fW7wPBeI0lof4vSuP1PQpdIklubBzc2j8+XnOKSZdiz/AMFAHaX9kzwg6btvm4JFfmvOuIokJ5KCv0l/bxuC/wCx74SwjLul5B7cV+asjho4kJ+YoMNVGZDKwGRzt7monA9eO1SSu6EggFcfjUByQM/LQANwvB5qQDepHJ4qLB64wKlX91gZzQAqKpAAyuKXYGZGGdoPIprzNIfk4AHNPRslOuKANCzUnxDprrkMrrj86/Sj4ewRjwjpks0RZnhCg++K/OnwvYHVfFOnBFYqrjOK/R3wxaXWnaBp0UqssTRDYB6460mV0NWawlLkwb1uY1JGOlZtl4hsrom01e0xeHhJkGCK0otav4Y2gSIGcDO4jqKgis7DxjJbW8yNYamJlC4GN/NMT2Kf/BR3yrP9lXwrb7sFtpUHqa/MdlCeWX6bBj24r7u/4Kg+No2sfAvgiCTfJaRf6Qobjp3r4NnkPyggsMBcelAgZFx8rcjv61SYbH+boanZgMAcH3prBXID5FACKq7uTjNSGLyY9ofgnn2qBo1MoO41ONmQDkmgCM4c4Y9O3rTXdU4BLGnuNsvA4pjRhwQoIegBU3E5Ga09Osftmo2NttZjLIBj8aygXVcA855rpfC1zHZ+JdHmJJUTLuX15oBI/Rv4aatqGgeC9OsUtg0SRquR24rt/ivfPZ/sf+L7mAsssjMr49DmsDwEftWjQSW4LxPECUPbjrXf2GjwfEL9nT4geF1G69jhknjU9ScHigs/GbTnP2OMA4OT/M1cycE7iMVBHZvYSy2sytHcW8jRsrdQQTUv2VmOGzxzTuQSA4XPOTU1hGFuGGD84xTRGqjOfarVphZFAB3gjkelIEz6b/YiLxa1rdiN3kyAHbX2p8TPDsXjX9kbx7plwDmyjEkRPYg180fsd+BjbXM+qqpAkUfL619MfEnVU8M/sy/E65nZo1khCoG9TSsU2fj/AG8SxW5jctujYx5PtTiAyjPT1qSCEyWXn4J3uWIPcU2QKVJUEACmSN2IEIPOemagRRG3lt36GnrjAz+vamyMuSu7JxQArt5UigA460u8MSDwDUbP9wkEkelBKyP8uduOaAJ1CJkhyTjpTGzJjPHpmmCNQd6g8CrAwse7+I9BQBH85O0t2oUMqkfeahRjO4YJFKCEQ7j9KAGRHdnPX3pUQMwGCQaMDOSMGrETu7BRHhqALFhbGTzsnJCHHqeK+8P2D0uLH4ZX1xBbPON53IOcda+F9LUB53GQVQ5H4V9x/sEeJNW0DQBcWluLzTN5+0RMO2aGyon0jLqKSeXJHFJHLzmMjFd78H/EcN/4usbYRtDMhOR61g/EvxJpmox2V5oNsiTA5ljUd6634bta6lrejX8dqLe45VwgxnpUrcp7H5f/APBQyYj9rvxIT0+zp/Wvm6QJIFcDBHWvoP8A4KDT+Z+1p4kY/wDPFP6186gnhjkA1bMxoLmQhBketPkgdAu5weegqRTI67IuBnk96Yi+TyfmOetICRmkOERT9TTUjYHaSS+afv2epB9Kem0sXyePWgCu6jJYZAB5NIXx2bJ6Cnh3Lb9uUzwvvSITN5ibCGHJoAicNK2ccDrmpETghenr6UwESICCwwe1SZ44OWzQA4uJF2sSSO4oAaM5OD9KTy1D4BIY9RSyZx5KnGf4qAEIDZLNt9qZiN8ruyf7wpxjjH7pjub1oyg4VSqj+L0oAcMOhXcUcdAe9ORJJAQ3ykdKZgSjA4YdG9aGUSv98qy/lQA8l2cYX5R1NO3uVOOR7U9YjgFZMr3FG0qDn5R6UAN/dCBmJwcdPWhJXjh+UZQ849KUlG+XZ83YnpTCSOoOfTsaAGqwPzJyTyRTkuhICuzacc0xQI2zzk9QKVoAzBl3LjrQAiyJGw2kD2NP2ea4ZDsA6qO9KGhcYZOfWomV4eh3R+vpQBIH82Q7OAvb1pYXddwjGGNRhhhT+oqRgBjaTuPcUAMXcrgtlgevtSzRKsQJZmyeg7UrR5XIOxR96nBI2KuSSg9KAEeMKq7SfwqMpnaQrbc81KFVW+U59qHlkTnZ8negBJIuVCtmlc7tq980nysAQ21velMmFHy/Nng0ARPKspxKxBB4o8yYcIdwz3p8IjdyJUyaiWRQWVSQM9fSgCVTIrbpQNnoO9STqjgDJwegquZOcc/U0+NiGIHLHuelACoqAFMlWz1PapI9rkqfmHqKRnYnYIwCerHvSbSowvykenSgB5jETbkbZ7VI87MuWUHHeoXOR+8B20gZ0BCAnPTNAFlJxgnG5T/AaUJGqNjIVuoPaqyNk7WyGP8AEO1SGXPyHLY7etACRqUkCjLRevam3AkmLAg7R0Ap6KJwyK/l/wAhTQxYFCTuT+Je9ABGGQ55wRjFPfzSASPlH92mm6YdEyB196UTMqjIJz2HagAaLzBuU/L3qfTYwr3DHIAjOBVcLIj7sjHoKu2GDHctzymMUAfeH7J9gg/YA+KEwyC0znr/ALNfn7YrutFO/Aywx+NfoT+yZH5n/BPz4oh9wUSvj8jX56WkavaKwLZDMNv40ATSwKmDKc+lZWtnOnPxxkY/MVoSgsQrEgjms/Wzu058cDI4/EUAc1RRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFAooFNAdJCoMMRzgbRUnkqckPnA6VWU/uos5xtFKrAZxml1H0N2xUKlkDnDXCjH419vftvW0cH7Kvw8253bE/lXxDp7K7aYAT/x8r/MV9zft4xCH9l74dAEjMafyqX8RS2PgGbJCjJztH5Uw7cDbkmiU8qMEnaKR2SRQF+VvWqIFJ8zDRnAH3qjdPLGc9akRBIo2fLzyPWmTjMgHagBRuxgGnRjepGee9MTpxTkkCEqB1oAcI2XADZBpUjAIG7k/xUm3ygHDZHpTA4Yg4wuaAH/PG55JHb60mQVyc780NIZGwDz2pGy47lh0oAaS2/AFSREkNkfNUQctwQc1IuOTnawoAVWfnbjHcU3cSSDwO9Ctsz704xAHg5oAQSErhSSBTsqBlwStRuGXBwQBTpJRIgKjC9xQBP5UYXcTgHtTQY2yB93saaFSSP52wuelIFKvgEgH0oAGIEm12wRSOxlUgjAHf1plwp3L8pOPvGnjapPzZz2oAeMBMhhjspoA3HrhqiMe4EkHcOlKHVVyQT7UASbSzYVs46mkMghf5eSetEcomfOCmKUIoY7icnvQA1XDM2RjP6UsrAkYbdgdaWaHYQckjFNKqgDZ68CgCXYhQSbtp6VFIAihT370oAHyOCc9MUEKIyHO70oATyi6fIckU95S0IEgw9Qorwtwcg1LMzSAcbT6+tADkDIAyHPqDShlH3k5qONiikg8d6IwWB3HNAEhlLsAFxRkqQecd6a2QuR19KlUM6qxYL7UAJ5OCskbbQOtPcxyuCzc+lMQBcgEE9xTEKjjbtHvQBMixt8gG0+1Em0qMJtI7+tNhjXzBgke571O64wWcYzxQBWcsqA9yaauJMBc7s/nUjYkO0tx60ZS3YbcsTQAL5XmbZcoR3pzlE+WLL56UOVu0KOuH7GmR4iXbna4PDUASxxByVxlh1B7VBuJY8twelPDFmzyGB6jvTpLsN8qx/MO9AGnox2zMf4cV9cfshaf9qsL92VmjDjIHpXyRozAOxC545U19W/sm65Lpen6jLb5kIkAMZ6UmNH0vruk7rMPbN5ik48p+orkUuZ7GSXySzOuMwvXWv4mhuYHNxCYZ2wVKjiuf1O607VZWFtI0V4n393cUkUVv28bt7z9kXw07x+W3ndB2r8zpmJhjXnOwfN6V+nX7dNoR+x34eb5uJvmr8zmtsRQ5zgoKoTKEbMdyvnp1p5XEeWyRUtxtG5VBxTWwFAP3cc0EkQO5CRkCmnMYxywNKcxtnPyelTRBecjKmgCPyxkMCelTxHykaVgcAZxUkcC5xk5PStPQfDd14q1200ixjeWaRh5m0fdX1NA0ekfADwvNquvWdwYyUaYNjHbNfoVZ2uy1/s+bcHaMGMntXhvwi+GUfhowNHGd9qgBGOpr3ux1WLxJbm0wYdRhGUOO1S2UjnTNClythdO1uTnbMOoNdpomk6RYWr61qEmI9Jia4FwOjlRnH6VweqXEdxqsNs8TNPny3BHc968X/bD+PEHhXwPH8O/Dt15l7d/8fkytyg7rmmkJs+Wfjz8Vrn4w/FzV/EMzsYDIUgU9Aorz+Vy2DyT7UiRBFCLkk9WPc0FAh7n6UyRpZuNwNORt4Oc4HrSoA56n8aRkYsGB6dvWgBVXGMdD39KkJDjHcd6BGFTk7mPb0oVlKdMEHrQBG2I1G4/NninlhgbBls02bLAMRkUgYRLuB/CgBSpKsQDu9KnnleO1tpowyyQyBsiovNZhwoX3FXrKNLuF4NxLnkH39KATP0Y/Zk8bWvinwDZKZf9JhUKdvXGOle4eBNYg8FfEKO6aZH02+HlTwnGCD61+an7N/xWl+HHidYZ5D9lkYK4J+77192z29trtnbaxaTs9s+CdnJB9RQWfJX7f37Ptx8Gvi7deIrK3L+F/ELm6gmiX5I884Poa+YlnG3G7JbvX7O2elaF8f8AwBefDnxii7njP9n3co5Q4wDmvyZ/aC+BniL9nr4g33h3WoJEtVkP2W7ZTslTsQaCDiGlU5UfeHeuz+Eng2++IvjK10qxheQhg0rKOAPesP4ffDbxL8UtYh0/w9ps115jBXudh8uPPcmv0F+Fvwx8O/sw+FDDcXMN54quVD3FwCCIj6ZoGmdX4H0N/h7Fb29irFgFQqR0J4JryT9vj4/2svhTTfh5ocwae4O7U8dMe9YXxc/awtvCcNzHpjrdavICqBDkA+tfGmpeILvxBq1zqeoytcX1wxLM3bNA2WSRCiwgkqq44qrJKqoWHGKYbgZ6nPeo5JBsYuwUDu1BI1SHBbPOOlOQRsucYY9qqxzxRyE+fGeKsC4g6+fGG7UAPeIeVgg8UpiUpuTIx1phuoSpBuIyfWlhuYu1zGAR60ACriPcM7fT0ogUqGDZYn+VTiaCSMoLiMuegXrSLECM5wR+tAERcowyCy9s0pyxzIMccEdqlaT58OCOOKYu2IZ3b89jQAqLvHo46VJC8m8Iev8Aeqs0kiHJGPQUqz5ZeooA2bRhtnC/LlCCT34r7k/4J9azDZeBdRjcfIXIYN35r4Qt7gOkudwIQ9K+0P2DSs3gnUd4P3yABSZUT6W1hf7B8RLd7W+wzdEHrXqfwmuYH8QactvKZFck4P8AD0rgTYnX9Le0DhrqP7gPpWp8E7C40z4gWUMokVlJyX6Ui+h+df8AwUDhI/az8SADkQIf1NfOoZztBXAzX05/wUFiA/a58SZ6G2XmvnCa2jUKQ+OelUZFNWMUu5ScE81NOR56KmcNzUcpUg7c5B5xTyouIlCHa4PegB6QFhndgg80LbkSFvMBj7+tIp2rlm5Bxj1pqgkkKMknp6UACFA7YbJHQGl3uSWYAA8ZFRfKpKkHcDkmpC3lxlicg9qAGiMIcIc/WmyS+Z/CUUdcVN5McsXmLJhx2pgnIH3AR6mgBqFZTsbKr/eNOkhWMnLfKelN80uSJUCr7d6dFNAPMEoJYjgelADWRQp4J96VvliOPmPp6UkUxSJsgsO2aGlDDcSR9KAIsMyHII+lPbKKH5IHXNK/95TmkUvKSxPygfdoAnVUX50JGf4aUNk7SdxPVj2pIY42BWUkE/dPpTmCq2BuyB+dAC/JC3lMGKsOGPakQCMkEnaB1NLvaU4lxGhGAfSmSJGjBQxkXFADi+1wSN3HGKbJK0kioRtPf6ULI8RwiAgjq1IyEtknDY60AATqAMrikUIuASWU9fapAFK/u5BkDnFMEanBQEHvmgBzIqjbEdwPY9qFHI2kgrzg0xSQ2SOB1pZDuwwOFHSgB0spdwzA7TxtWkjUISgY5PYU+3lKIzBNztwB6VEkTtIuQTIegFAAdpQ7SVIPzetSmWVkAQcd6SSQrwY8MOtMBIxjg0AKsanAwcmj5o2KIcntUiALlt/zntUZTI3A4kHf1oAYjDcS2c0x8AHgmpZSY2UEbmbstDRtEdw+f/Z9KAIM7VyAR/vVNGqzR8NtOaY5Mp8yQFQP4cdakVEQFgCfY9qAJTEFU7pckdqFYsnPTPBqNNpbJB2Z/GpEiUzFdxVPegAZ9pOeRUiWwYEpKAT2NI0bxkliGQ+lRzKiMCFKxHrQA6SPaQC+5/8AYqLIBK4Ic9zSyRYUtE2R6dxSOX8vLDPv3oAcqxSybXfBHpT9oTe8bZwMbfWoVWPOSpyelOAEM33SWIoAljcheVJP8qbIHlO49B/dpzY3A5Kt3Bp5yE3ZwB2oAiVunXA64q3ZMAZ95KqU+WqYG87sFQO3rVmxnVY5gVzleD6UAfdv7KWqzL+wX8ULcI/l+a/zY46V8BWeHs927aVLce+a+7f2T70v+wt8U4TuwJXPt92vg6yCtApJK8tz680AWt6SIBI2wjn61ma5zYMecZHX6itF9szAFd23nNUdeIbTmPQ5HH40AcxRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFL2FJQPypoDoI5iIY8jPyj+VNcs/QYFEJ/cpzn5RSO+0/NwMUuo+hr6fM0a2W0nctwp/Wvtv9uXWvt37Nnw3hYFT5CH9K+IrNgHsGX7vnqMV9p/twSxSfs6fDjacOsC/wAql/EUtj4dLZ4BycUgVWwV+U96CdjK3UbRmmj7+5Tx6VRBICjZYDAHf0pr4YADP19aEZTA5PDUKhIXDfLQBHu+XbgimpgAgMQasI4TIk+YnpTPJQLuL4NAETbnVRHwM85pw+c/Kc47elJu3YzwPX1pspAYKv36AHxkgNzmgTgnbyp+nWmJ5ZT5m+anlFxndQAjttZQM9c0csd3JGaFIkwScsD39KCzbtw+6DQBJs3LuztOeBQWLgMBgA80pQToWJwwNIrCMEj5gaAH7scKdwPamKSrEY+b0pw+YbsbMdxTCRjPH4UADEvndkKOuKWTGBsb5qGGAADweqnvUaghiGBHv6UATb5JAQRn3ppXHb56QMyMGU5NSF2K5K/Kf4vegADsib8/8Bp6r5ylgAT71GuVPIy3ZfWnpJ5UnJIz2FACq5zhwAf7w7UjYTOehH50ruDnblkPUHtTQSzYfpjj2oAcWMCBXJYHpUcygKpJJqUwBxw+fSmHfHHtYZNACuhOMHPHeo1YgEN3HA9KT5/4s+2KQsZGDL2HzUATtC6x5J4I4NCZWMjO5iOBUY5HzEle1KOPlYHpwaAEznAYYNGHMgUZH+1TyMAKSAT0Jp4TZHgMrIepHagBjCRQMbWpQvmY8wlfQUyOJIc/Mc/pUsbbTubnPSgAaFQvUqRQki5AlB2joaazmVgZB9F9aIxvOGPA7UATKyyOFjPBNRuvluQykjPFLCFUnnpSSSvLIob7gPWgBW2FeRtI7mkD5OMYx3pzEudrdu9IQQ23t60AK2T0bk9KjOchnJKj7w96eyCNht5oztPrmgB64ADZIHpUoWJgcNh81EqqQcn5vX0p26Pb3L9iKAL+ngBpAjEv619TfseQJqGk65A7lZmPysegNfLGmfdk2jDqMmvp79kic6bYahcSMRbSvtc+h9aTGkfQmkrPqdjd6NeK0N5A2YpsffHauTmdzcXlpcRyWl9F0k2/frt57y4tEOEzkhopwOarXXiWDVYporywUXar/rsc4pIuwn7Zt2X/AGLfDwkZmlEvO6vzWlnX7PEucZQV+nf7QXgjXvjT+yjbWHhe2+33GnSbpYVHzn6CvzxuvgH8Ubcqp8Fam2wY4tm/wqiGcTkAHJ7U2SLK8HrXcQ/Af4pSkgeCdU6d7Zv8K6DRf2W/i5rDhIPBt2m4Y/ewkUCPIgjZ28lcdakK7Pv/ALuP1NfUXhr/AIJwfG7xUymTSYNNh7tKduK9e0D/AIJo+HvAsMWo/EbxWh8vDNZwNu3e1AHxD4I8Ia78QNWi03w7p02oXEjbfMjjJRPqa+8fhT+y7afCHRoH1F47nxHeqDLIpyIgf4c16lol54L+GmmLY+C9KttMttnzXhUB3964Txz8dPD/AIWYm71FJHZc7VbPPrQUekw+H7XTLYIsqiVVyx9a4vxR4itNGnS+s7lLa4iX5izbQRXzl45/bHF1G1ro8LNJggORXgnij4neJPFjsbu9eBTx5at1FKwNnvXxW/atj0Ga5i0PF7qs6lWk/hQ+oPrXyjqmqXer6hPf6lO9zd3DFnkamSrGhOCSW+8xOcmq5Ta2VO5aYmBkIxhiX70JIc/ICfWkACnipYpHAKhcCgQ1WaQfdIFALpjaDipEXJGGINPP7vGWOKAFiVchicGkACt13e1Rs4GDzT1ZF246nvQBIyAAEPnn7lQbyCfMT5f507IDAj161HPMQdvU9qAFEm08fN6e1TwXZjw8Z2sOc1Qd92AvEnpTvM8pA0jqo7oe9AGzJN5gW6gJ3DhwvWvqH9mL9odtInj0LWpy1rJhYhJ2rxb4Rfs/ePfi/frH4e0eeOxY/vLu4jKxqvrk19EX/wCyp4Z/Z60f+3vGGvLeajH88VtE+Ru/Ck9yj6TTUAZ47ixlZZ1O+Nhxgda6Pxr44+Hnxe8P2ui/EbRPt2o25CW1yseXr5X+FH7Qum+KdWGmPMITuxHuPUV77d21pdQiWGMC7X5kkxmqWwmjgvjt8VoP2YfD1ppegeFE0KzvU3Q3iR/NIPc18O+Nfjr4l8Z3MxMzRq/Ifdk4Nfqf4i8N6F+1n8J7nwBrvlweJbNCbC6kAByBwua/Jz4lfDfXfhF42v8Awt4ktHs7+1cqhdcLKmeCD34xSEcjMjzyNNJI0kxOS5qMphcl8H19auSFUVscNnkVXkCsOck9qAGpKAvzHntXf/BSXwJ/wl6f8LIluIvD+OWt13NXn+I1GMbmp8Zw2GXf7UAfaVhb/sbX07xnUtWhGeGe3IGK9E0n9mz9mDxNZpdabrl5PAwzuRDxX56x6dbX6uyoEkUfdNfU37D179tuL/w7cxFohmQMR0NAHu8P7JP7OeMHUr454B8s1xniP4bfsh+DLxoNQ8QXssyDmOJCxBr2P4522j/Bv4M6h4hmb/SLqMwWu8YIfHWvy3yL2Wa8v0M15cuZSz+pNAH2TrWjfsoa74N19PDGpX0evwWzSWwuIyoZ+wr42tG82Js5ZlYgZ+tSOYtwxB5T4xu6VE5+zlducHqKAGySbMHOeKqs+2UNg/T1q01wpyVTLYqus7BvmGAaAGtJvYYOQOop3nq524x6VDJGyvuQ/WpY8O4yuSe9AE1o7eVOcnIFfZf7Cl9NYeFr1kBaNmO4Yr46t0EUE4b5iVNfa/7A0kNp4QuzOm9Cx/Ckyon1ZbQpdQrdabLi4XlkNdr8JPFltr/jGzsp4fKvISfm29a5JNOtZ5Rc6Pcqs4+9CWxmtX4YKB8T9MMkRhuctuHrSRfQ/Pb/AIKFOo/a48RdsWqV8z3G0/MTj2r6O/4KFyN/w1t4lyefs6Y/WvmqTDKBu+aqMhsiZ5z+FKmSw2ttao1BzjdzUmPKIJb8BQBPFGsrjaCXB5J6Ux43jkZ8lWB4Jp8AkkUiH16UjzuSVmHSgBjTGUEFdmO4700sWyMHj+GllOxdw5HYUhudyjK4/wAKAJ4I0Kk5ZTmoZD5TMwzsJ/CpZ8tCrZ+TPb0qOR3iO5huiPRRTsAq4YbXOPf0pZLcyD5AC3qaPvITgnP6U1iVXbuIJ6e1IB0YYZ3Hp1FIJskgRYUdaYNkjgb2DihzsO3kn2oAc8gP3RgUqsFcf3aj8xgD8vHpRvLgtjgdqALOwShjuwcUkfDqpbpRkSQqVbYR2pp65OQ2Pu+tAD7hGuXwCQFHT1pVjTgA4wOhpMkDfuzximSKJwuH2svP1oAl3l1Iz0FMiJ2kA7hT1CyjKYDAcg0Rq8xIXarAcUARIkSfNylOZTjLHGehp7QzB1DID6GnsHjyCgZv5CgCLYSAFOCOpPehZkXAVNx7mnQlFRg3zK3AqNN8eVUbloAlibksDyKGkcSB4m57moVySS2R7CnHCbd/Trn1oAezO7B5WJpIlLS4kYIvbNMDPG2WHynoPSiZl8xc8g0AOe3RTnO4f7NNmUldrEkHpila2XYHjlxz0p8YDJnOGB496AERR5OGLbweD6U8IJlG1trZ5amH96wJ4IPNBxDkqSMnp2oAZJL5blSN3tRIu5ASx256+lP875cFdxz1oUkO3y7l/lQAsYaTgHb6Gn+SgyHkyxqJ/MZMLwBSgBxgH5u5oAUKUJUsSe1KZW2kbCy/xE0wyGGQHd81Oea4KklQE9KAGRqqsVLYY9KsKjJyv7wUyGSOVSsq4fsfanqfKJMWWFACec5yHTHvTvNjRODvc9KgaQu53McgdPWplVX5HDdqAECAtl87u1IqkOQT2705T820nLUOkZcqZAGIoADHIDkDI9afZxsDOScEL0pBbyw8ibcMUtorSmcsSWCGgD7W/ZVCD9hr4pZnCMZm+U/7tfC1tlraJGB25b5vxr7n/ZksrOT9hX4ks25ZhO3T6V8O2S5sl5wMt/OgCZYyfu8KOpqhr8YGnu27ccj8ORV1XlibDrlKztamElhIMYORgfiKAObooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAClFJRQB0SIn2ZGBwdo/lTAyrw3zDFQxkGMZJ3BafkKmF7imPoa1jkvYc8faF/mK+xf21Vx8Afh4Of8AUrx+FfH2mHmw9PtC8fjX2B+2xvX4GfD3Iwvkr/Kpl8RS2Pi6SRY3BZSy7RxURkJI2jBPSrFwhkdWB+Xb+tQN+5YH7zUyAVQed3yfxU9UWTOG6dKYSVHPyk9hSqMZwcSUAJtXgh+R94UyQLv3Mcp6etBhB5VvmHU0OSpzuyT1oARztTpkH9KGhY42H61HjcOD35qYsEKhfugc0AEezbjGWzzSRYyR905peJDkAhe5oJUcdT6mgBHicPnOD6UBiCAPXkUpQ/e3ZOeKaMM2O+aAHySBjwOB1pygu3HApeAxXrS7lYgrkbetACIpG5g2cdRTQ6j5gvPegqEYuCdrcGnAmHCqN2emaABHyThcSDofUUOxJ3EfLTQ/zdcNUpHy7t+V6GgCNgGX0JpSSkbfNSKnmKxyeOlN2NgknmgCeMqBljyelMkQlssfm7UhIlVdh4Tk09Zw2d43EdKAIjkZK+lSRzBwVAIOOtK0yycRrsXvSCUxqdg696AGZwSM5bHWpGIRTht5I601UDZLnB7UGJesb5bHIoAC7Rxhgck0oVCQwOSf4aD+8hCZxIOfwpVixzn5qAFSQRgq3zZ7+lSiPzcCR9gHSoWQMevajAkj+fPHegCSZVkkAI3bejUxVjOSFwo7etOYhUG3kUzcQM44NAArnpnI7e1SLIsGNxznpUaEx52jGe9KYCU3RnK96AHNIoILHIPQ+lJPASU2tg/zqNSxUYOQDyaDmZwz5DDoaAJU5+o6mhmCsAeQajUMxI35qTYVxuxg9KAJON4DNtXtTJIwrbVk356U7YGx5nI7UnlJ0HyHPBoASHchIPLCnoyliNvNMLiPJI/GkYgRkoeD2oAfMoXALA59KchTbtU4Yc1WjTONv3vU0TDaMbue9AGrp84jSTa24sMNX1b+x+LW90fU9PvX2o8gOTXyNpp2pKN3zkfLX1P+yNGL7Rb1fMKzCTjHrSY0fS9zp8mnStaSXamNgDCrEZPpXM63bXckjly1vNEPvY4YVP4n043xeSaSSPUrYrtcZ6VLba3d3OizLNGLiNF5fHPSpNDjvEHxr8ZfAeyg1eOWa00u5O3eEyrVRH/BSPWE5bU2wRkL5A5/Suu/bMmN1+yH4elKqQ02AdoyK/Pq3m2QxARLu8sc4ppCufa7/wDBSjXcFk1MlgOhgA/pXWeA/wDgorrmu6ilnc36RCUgLLsAIr4AW7iZiHhUgfeIHaorgf2VPFe2ZItmPPbBp2Fe5+tGo/Ffxfr9gJ4tZkeF14CHGfyrldsuq3ZGrSyyTSrwCSRXz5+z18YW1XT49Lu7jMoAWMk5r31NYvrSVZTF9pRFBL47UWA6j4Z+E/D+o+JZ9B8TDfDeoUgkbgIT0r4N/bD/AGf9W+AHxaurTUWludE1AmayuGJK7c9M19mnxF9suVkaTyZwd0LL1U16h448FaP+2F8FLnwpqgQeKdMQtYzNw7EDjn0pks/HCW6VMRqAgA4Iqi17lsMS59a2PHPhHVfh74p1Lw3rkD2+o2ErRMHXG8DoRXPbfk+U8UCLEjK6nBpgYquQPwpiEY5NTJnqTzQBESzN92pQwLDJ2mgTEnB6UrhXXBagBA+JMA/jVhnI6DcarMdicNQrOuCD1oAVy0g6Y9qaBtUbqQ/MfmJzTkfa+1uRQAxgxweQM8H1ocFx339qtoy9/mHanMruUWJS0rttXA70AVtN0u61jU7ex0+Frq/mcIkajPX1r7p/Z9/Y38L+E7W28R/EZxf35UOmlnBX8aw/2Yfghb+ErZPEeqRB9UmGYS652A96+h4baTVtRitPMdppDhn6gLQUb3ij4r23hzwjcS2sUHhfwlZpjZEu0ykdgfevzg+Mfxjvvij4iluAXi0dGIhh3ZDL6mu4/bA+Ln/CV+L/APhEdOlaPR9LbZOiHiSQdz+NeAi4Gzg7VoBsm+0S6TPHqenSECNtx2nkGvtf9n7452vjnRbfT7u5VL6IAbieTivhwXRsZTIq+ZEfvx+oqfTdTuvCepw6xpErrEG3MqHG32oFc/UyLzI7lL6xnNpfwEOkoON2K6D4wfCDw/8Atn/DuS2uI47H4g6XF/o9zgK82BwM9814R8B/i/Z/Ezw/DDNOBeoANmeSe1esWeo6n4e11L21na3v7dt28Hhh6UFH5f8AjvwNrXw28V33hzxFbSWepWrldsi43DPUVg5y2fSv2B+OXwH8M/tqfD6a9tIYtN+IOmRZSRBtaXA6e9fk/wCLPBur+APEt74d1+zez1K1coVdcbgO9BBiC3ViT3qSOHsfvdqnCdQTytSF8ocHn1oAXT0C3Yz1FfYP7AfheXWdavr5AQZm8uM46mvkG2VnXamWlchUwK/Sb9lLSbX4J/Bi/wDFmoERpYwG6TeMbnI/xoA8x/4Kc/EqO68TaB8PbCZZLawgSacIeBJ3BFfEjOrtgN07f3a1/iF4+u/iP491zxReSvLNfXLyRl+yZOBXOl84AfnuKALDHjDHJqA3BjO3G+n/AHlypyBVdjvf0xQAxpMElsikDFz8tSKpdss/HpUUgcvuiOQOtACy/uh8vJ7ilt2J5HUc/SgMNvmA4I60iPmQEfLmgC9Z/OkzudybcV9t/sPQn/hCbjAJXee3aviOzmBjmB4TaePWvtb9iDXbrSPBc8kUPn2287+PuihoqJ9Fa9oE22O90e4aGeLLMgON1dh8H/F7eJ/GGjC4hNvqUDFJDtxuFQ+bpur28dxbz/Z2x8wPrVz4X2X2T4laWZQGZmOJFqS+h+eX/BQncv7W3iTJ+Y26Z/WvnBlBC55r6S/4KHps/a48RjPBt0/rXzeHCqAPSqMiIgA8HjNCuvU881HJJkEgnAp8aM8JcHIFADwZFbMfFK8jyAeZyRTQd6ccUrMUQEjOKAFLsigYyueBSQDzJSGXbzxSF2I3Mdo9KeIkdcq2PQUANaXaWXkPnG32qQnyl2gbs/pQJjtwYhuHAamRDaWXPJ7+lADhJsXbkgmmsuHCsSc9xUnnKgO9d/v6UCMsDsIKGgBl1AkCAq3mE+vampHnjflu1P8A9WSN25j0pkrZX+7jmgA2Mr8tmmSMTJx0xTkI2E8sp6mneV5inbyoHJoAVYRKC27aQP1pyOXU5bc4HFRhgo2M2/0pfKTrEdpx1oAQuVbLtsOOnrS+ehOZE2nHB9aC5JG9NxA6ijz1YH5N2BQAuVbDMStWFg3r5ok2KBzUBl8uMP1JGNvpSsDLEM5z2WgCUSOVwrFl7mkE0iHDMSD0Ipi7toaLjHUURXC5/nQA8oVPJ4PXFLkwJkHg9aiWXbISx+UjipXIKA54PWgBvmCTbzgHoaHjdCFB3j+VNTao6EimIvlPuVsigCUsIxzJuJ7elJgFQSeM0wKhfcw+X1oVtjDcMr2oARlXORyvp61KnQANjJ6Uk7fMrIuF9PeogUYjB780AWXkCIQB89Ks6zptf5SKjljZACD78UxiUcN1zQA7zSFKheO31ojdskZw2eajYsQDnLZ4pZCMjA+c0ASqQxb5vlHU+lLLMCmyNcjOc1FGnlPkcqTytStEQWKtj/Z9KAGBFzgNuJpGMkkmG4A7inB3TkDK96cVEiFlbaPSgBjAcsPu96lgm8mYNnI/u0khWGME8ue9Sogl+ZTg46UAMmnWRmZU2nHWo+q5Zsr2pxuVQnKdOtJC6ySElc+i+tAD0xGwIPHoaRhBuYMpUkUSEiQlsqwHAFEu9lz1XuaACJoomCZ3E/pV602os+1t7FetZwCBcL8w/vehrRsfkt5iGw22gaPtT9mhQf2FfiMN2CZn/ka+GrNA1lHu4ALc/jX3X+zbOY/2DvH4FsWUytmQDpwa+FLXD2CKSdu5v50CHPMFBCtk+tZeuKn2FyPvZGfzFaXlRI2UPGKz9eUrpzZ9R/MUAczRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFA4xRQKAN6IZjjOf4RSiHklTwO1LAo2rg/KVoclBtB+tMLmzpjAiyPpcL/Ovsb9t+VW+AXw5A6+Sv8AKvjfTGX/AEMjoZ1H619i/tvMq/AL4dAdRCv8qmXxIrofFNy7S7AGwQKhJ2Eg80Tk+apJ7ChkOc5q2SGxihGd3vSKPQ7j3oBPO1setES7m4O31qQHiKM4PmY9acOFZUUOP73pTXQBgFOT6UkZAVwvGeooAZCokO0fI386UxgnDDaRTkUSbQPlIPWpJHAwpXdz1oAYsjYwBjFIGCAt1PekY7icdR1pqne/AoAXcCNx5U8FaTcImyfmHYelPbJXjjPFOBGQCeR1NADVIbBU4JPSlIcNx+dKdm8KrYJ6Gh42ZhiXNADQGIb5ssOcU9pFmCMPlPSomQ7id3HQ0v3YSDyv8J9KAJGzGSv3iejULGCGy2DRbndHgn589aeyEq2Dk0ARLujJKHcR1p7y7sgjk9TTVR1Bw+1jTmk2qcNnFADABE4+bC0St8+c7V9KRWWQ5J59Ke0j42hN9AAse85Hyr3pVOGKk7kx1pI5WB2ngn+Cn5CM2D25FACHkYHPqaEiJc+W23A5FSLG0qkxsAPT3poj3E7iA4FACvk5Kj5gOTTYyzn5j82OKUMo4Xg45pUfDYY/NjigBAuAAWx60+Rh5ZQfKO59aY0To4YNggU8fvdq5COT3oAhK7QGQ/L3zTwwYYz8p70ydmWbCkHb1Io8xcbieO4oAliBZtrn5PWmBnyVVtq+lMjIZsdARxSBsDax+bsfSgCZJf3ZQdaiEvlthmznt6UL8wyOo7+tSQSI7MskWT6+lADSGUhw2Aaau4yjdnnpSfZTJ/y04zxmlSV7clHG4nocUATLMBlX/Kkdy6jJ2jPBqPpg5yc0kkpkA3dj0oAXz2Xh+R6etAYjHlnGaIxuOM8ZpCqAspyxzQA6Xeq8NljTNpccE7804IApAbafU054yFyD+NADrMMFlIOG9q+sv2Mdp0zUR5gSQtkFvWvlWzTfBIgODj71fS/7KNuZ9GvofMMRMg2uKTGj6c1LXVjhKXFusrof3jY5I9qpWr20sc8lhLhCuTbmq90HldLG7/0e5GDFO3Rx6VkXenNBeTSqxtLlV5C9HFSX0NT9tJNv7InhzcoU/aO1fnjuBhhBPGwV+hv7aTE/seeF33bh5xzX54iX91FnpsFUjNoYyhc4PBFSwzRqjQP81vIuCD296ryTAkgcL3puVY4LfKRimNG54M1658I63HslKKGyje1foN8FvHVn428PIPPDSRriVD1avzkYG+tjAW/0pBmNvX2r034AfFS48I69FHLOVjQgSxk8Ghso+3tW0Gza/ZrG6MBPRG/vVa8N+JNU8D69aajBKyz27ZfafvjPQ1Vml07xTp0V9bP5ckqh0KnvWdb380D/AGa8zkD/AFhGcj0pXFY2P26f2e7D9oj4ZRfFTwdbIfEenx/6fbwjBdAMkkDqa/LZInRmEqlJEO142GCD6V+svwU+LLfDvxkdM1FGn8Oan+7ljkHy8/0r5i/4KCfstP8ACvxkPHnhuAz+DtbJkleJflhkPOKYmj41JcthVwBUkmSASKsvFtIZTkjoRULRGMFupNAiI5ZeBgU0BcDJoDluCMUmNo69e3rQA/ZtGVOaWNsnDHFCkohBOQe/pSqI5cZO1/73rQAuwqmS3HYU0ncBtyGB/OkdWUgE7vQ05Ld2XcZMYNAAHDN8ny4616V+z/4Sbxb45haRd1vbkOymvN4rVpnZkbavIavqX9kjwwYbS9v87ZDwrY7UDsfUWhX1hZRx206nYvyoMYArsLKOPRvhr438aEBorOxkWE46Pjg1zuiS6ZdWLQXsYWb+GTHeul8dWX2D9irxx9nl3I2/dx2oBn5CX2sT63qd/qVw5ee7maQsTnJJqITFgG3U6zgVbOJvvZz/ADNOEcYOM8d6BCg/xbs+lS2b/Z5Gb78b8PGelQvycoeBUy8c560DR0PgzxLqHw38QQappsr/AGdmyUB4Ffox8Gfifofxf8ORRmdI9TQANkjJNfmxYXMYBhuDmCXgj0NdH4K8Y6n8LdfgurSd1tiwZdp6igo/UqyTUfBWtw6jp915M8J3Ag8SKOoxWd+09+zvoX7X3gOTxH4ehisPiBpUe5o0AU3JHUHHWuU+DnxY0r4peH4o7qZU1DYApDdPeu9t5tU8B6ymp6VcEeUQSQcBx6EUEs/JjWdI1Dw/q13pOq2zWWqWjmKaFxtPB6iqRCjOMkH72a/UL9qT9mnQ/wBqjwjP448GRR6d4802LfdWcQx9qA6jHrX5lyaHqQ1dtGa0eHVhL5D27L8wbODxQI9C/Z5+HVx8RviDZwRxF7OB1ZjjjrX0f+3v8UoPBXgzSPhZoc4WTYHu2i67T2Ndp8J/Celfsq/BybxPrBjOo+R5qLIACxI4Ffnn4/8AH9/8SvGeqeI9QlaSa6kZkVv4EzwKS3HYyvKCxqiPtCDGaY+EGQ+TUXmFhsL549KFiQY5piJopnZPlPHpTmZj0NUyHgYYPXtVtInZgznAxQApPGM5PeliBVvlPy91pWlVvlUYOOTTkIT7hzxyaAI2ZTIQDhTREq7trH6VIiDae2aX7Nk5zlu1AE9v8qSrjcxUj9K+zP2Fp2fwPqkcRzIjH5MZzXxpp8Z2zMDggEnivsD9iRJ4PC97cQMdwYkgDqO9JstI+qbK4stRs8uv2cglXA45rQ8AG+8PePNH8l2uoHk5yM4Fc488bXYkL7La54XjGGHWuq+EOpS2PxE02zmlEnz5QsuflzSK6Hwn/wAFFIgf2tNbbGwtZxtivmR2weV5/nX0x/wUT1EXn7XGut/ctY16Yr5te6XAIXLDvVGRXQhyGIwueRUSborgnd+7PapPNV3JY7QeopCu9Sudo7UAKVKncp+lAfzRweO9MRmchQdoB5p6xDewVvk9PegByttXB6Z4piIpJJJBzQkn/LM+tISeVJ5zz9KAHszM4YNxTgUX5y4PtTN4xjsOlIiCR85PmdhQBMuzzgxOU7qaTcpuHVWIiOM0bdjEscP7UoA57A/w0AMMTW7MM5jY8Gld0kjwDkjvStnLZbIA+7UbIFOUb5T1oAYAQSAeT0oDjled+O1PDKhLONx/lSxxF3JVg3HAoARlXI29Mc090zGFBwOtPAcMqngDnPpUcikSEliF70AOWPaCVb5SOahkheJlUfMGqYQMyjy2Cr706eLyQqvKCx6YoAZLA6lfmypH3fej7QyfL1bHT0pnl7PlMuSe57Uh+QjI+Yd/WgB32h4yNrfKfvU8qjASxcr3qF5VlOFG0DrT7UjOFbbGe3vQA4Sq+AB1prpgjD8iljIhYj73vTXBl5zgdqAJEYuOMUrhgRjr3pgj3Q5zhhTFRiMb/moAshWcDYBt6GmEsilCmSOc02CKNCwaQjPSnwK0PO/cM0AQq5KAKcgnmn71C7QOT3pxAZ9ynA70xyYjknOelADolkAIVvn9D0piMZGIH3x3p00Jwh39aSH77r1B70AOQJA+5jvOfu1IhEwaROWU9KgMSJkiT5hzSrIYZRKh4PUUATwOJJS4G09xSTRgOWBJXvSB8bpR37U03BfP8Kd+KAHgEg45B6e1I8ePlZjjvikI4LE4j/rSRMSGXOW7cUAOjI2MNxLDoDVkI0cW8nr2FV3mTbtI/eUiSeWThi+eMGgCw0ykZ2bl7io2UOMIdo9fSkmnkmUqsYjQDlh2qJI0cBN/X170AStuRcdSe9RRl14lY7KlMDx/fbaewpQCU4HNAAsLbv3b5TuDVu1I+zXSlsNsOBiqpjcKcjA9qtWKt5FyG5+Q4OKBo+5f2btduYP2APH8CWwdTM4Lkexr4MsoWksEYMBktkfjX3V+zzetb/8ABP3x8Fkzm5YEY9q+ErSNZNPiO8qctn86BEohZc/N+VZ2vBxp77ueRz+IrSWFAeX+WszXwy2LjOVyOfxFAHM0UUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRR0oA34FdYFwOCo5/CiTK9uMUyKVhEoJwuBz+FOO0KfnyMU3sBp6a2WsNvQ3Cj9a+w/24FK/BD4dZPSBf5V8e6Yf3mm46faV/nX2B+3RMf+FOfDpPW2X+VS9yuh8YSDdIpY9AMUySMgkls0kzlmUDsB+FNLMv8W4VTJECMRvJ6VImJv49pHf1qJjtPP3TSn5CrKaQEgVgQc5xSp8xZsYqXZ8okzmoWbDZxhPWgAR2jyG5B6U7j5fmy4piwPjKuD7GgFh94Yx3oAWV8gGnJE2Mg8Zp00aC3V9245qFHLgZ45oAdyzDBxg09XCthiMZ44qKT7454ofDc7qAJSVyQPvU1im3HfNMU7FyvJzR5ikZ796AHfKeh+YmlQFGKuQVJ7Ux5FTtye9KSNp7mgCXDhWxwM8Uqqx3ZbpTFdnUr90jr9KdbRPliGyB096AEaKQAtu4qVPLaMseT6UPFkHLcU0w7PmTkCgByMiuW2cjpRG+WOOQaZIzNhtuAOpoKHcNrcnoKAFkO8kgbGH8VJ5o2N/fx1pzDO7tx1pNytAynrj74oAfHAzR7w/HtSf6s9c5pnltFCpL8Uu4DkHK460AKzmI7lGcilbll3fKcZpjcAY49Kcr78hjlscUANz5pO8ldvSnxwmXLMxGOmacqMcEkNjrQ0aOPlf60AR7zuK7gze9NkG0g7dx7ilMYH3Gz/jTRGHYHzcOByaAFWRF4xz2pIyrSDAz609I3ZcBc+9M83a4wnTrQA/bh/mGB/D9adhkPq3enmZZcBk2gdDTbjCsm07qAGiTOVzx2+tOjAlVtxxt7nvSeeDwI8nsaai7mPzcdxQBGhByD09aepGQg5oJKAnG5T0pQ67fl4c0AOBUnYKRgnI/Wo1BB4Pzd6lDZyr9O9ADBiMeq9j70+E7H+Y5z1PtTVmVDj7yUrLGBvU5bslAGjauI4ZQOYiDj619U/sgRRro9357bVaTgkdDXydpp81ZQeAB0r65/ZMurWPQLuC54LNgHHSky1sfQPibRvtOmutzKszL80NwpwV9qxvKnu9DcXG13gHyTjqw561el08We8NcsVkI8tzyoFYOpC90KeS4hb7RGy/6vsfpQwRoftnRmf8AYw8Ny5ztnORX50BR5UYIw2wbRX6P/tjapDqP7FehOIDbS+dhkPrX5uNKFEIOGfYNtCEwllUEq0WDjrUJcIBjoaJ5WEnzMM+lQ+YfulsUySQ3Dk5iO1l5zUc0zQyRahbEh0PzoP500s+AAABToZBDLufmNvlce1A0fWv7OPxYi1O1j0y8nyjcJnqGr6JntReWjI7qZk+6QP4a/Nbwrq0vg7xLBKkrLbuQ0ZHY5r9BPhJ4xg8ZeH7eZpk+3RqAB3celSXc37G5sp/9D1GLccfJNjGK93+Htxo/xa8B6l8MPFRS6huYm+xTSDOzjjHvzXjd1JZ28n+nWrKxHYdPerei+ILWyv7e4sZjDNA29McMSKES0fn78evgpq/wF+Juo+FdViYQiVmspyvEqZ/wrzeYBRiv1/8A2pPgtYftafA99X06NYvGuixeargYZ1A5H41+Pt/Fc6fez2d7E0F5bsY5UYYIIqiSBwFO5Rn1qNsKML1NODYGR0NRyMuM5oAasZY8nkUjrngjLenpTd2SGBqXYfvD5ieooASNSowh+tSKvzD5tq96YF29O9SwsJW2scCgDW0aIP8AaMkbQuR9a+zv2Lba31fwrewzOFCsTkdjXxrojCM3UbHlkwpr6P8A2NvF39nXV7YtKqoxIIPegaPqbU9Cm0y5CyENas+RIo7121vYS+KP2YviL4diJmnjtJbhUA5OBmuF1e51e3hSSH/S7djkKecV2v7Pfjmz0nxxLpmskJY6tEbV1YYALAjFNDZ+QVsjW8PkynyniYq4I6HJ4p7/ACg4TcPWvdv21vgTffAj41atEYNmh6pK1xaShTtwTkLnpnmvABK2f6elIlslZvN4CbajQFXODkU+Ms75LYApV/eyH+FR1oBMeJRGNh/eBu/pV+0u0njNreDdE3Ct/dNUhbhXyBx2qRYdwIJwe2aCkzs/h38QdW+F2vxTRTN5YI2t2I9K/Qr4U/FTT/ih4YiZbpRqKqNyelfmjDMLiA2t0Mp/C/cV2Pwf+I+q/DfxXAbQvdxu21kjBYYzSuM/S3QdT1TwP4kj1fSZi21v3qZ+WQdwRVy7+FPwy+JPxjsvF+m6WieJkHmXNiBhHI5Jx71w3hfxRL4i02O4EqwmRQ3lN1rS03Wrjwl4ktdb0p1W7gILgrkuO4oTCx8Uftv/AB21b4jfEO88NC2k0fSdKlMYs8bckHFfNYGPudQK/Tz9sH9mHTf2kPDr/Ej4fxxx+LLaLdqGloMGbHUgetfmbdafc2N3Pa3UL2t5AxSWGVdrBhwRTIII1EpznGB+tTBWwMHcO5p0cSsCVPbsMc1ZjiKx7s59RQBTZxuAIyp708HPysdx7U+RQxLKOg+7SIpkXA+8P0oAFcKeVwMc0octyo2LTtrRt/f45NPVAV3MwHpQA6HErcttWlTeknLYXsaZGVfknBHQetTJgn94celAFyyG6OZgdrbTX2L+xBdmy8JXc8fJVjuUjqO9fG9vcjZJGOuDg+tfZX7DSxXPhO9idgjEsOlJoaPp2+uLXV7I7YBbKwzG23gt3qX4RWVzL8UNGkMqSFCQykdhjFYula2z6hLoWoQlACfs8qrW58GdDudO+MemLLMWwzEHsakpo+Bv+Cg8/mftY+I2Tj9ygP6186xk7FYPk+lfQn7f6ND+1h4nyc/u1NfPSFWORwD1qyBWfY37xeOxpZMkZ/hNPlciPnDD1pjRsYg6tlR1FADcZA52gHg+9SbirfNyp61EjAx5IyAcj60933YJX5e9ADkRlb5SG54p0m5yCQA3QmoWUNhlJVgadkZI3fKerUASPEWcYfO2kPyndnjvSGFUXKtj+tOSRQPmXPt60ASFcxkg7vQULIqjnk+nrTFLPypG7sh7055WTmSPD9uKAHbkdiMc9qZJMm0qIyvrUUjMMMrAN/ED6VKbjdjdjYe+KAGrtQYI4PQ00bN3ysRJ7U4qXJBIKdqa0cm7DAAY60APlkbhCdynqe9L5bdFIKEdTTmk2R7Rhqj3+YNpPJHagB6xfvFV5CFPcUTJGl0oLkrjhjRs3AZ+bHejCEEAg8cUAJMjjknJ7UJI5XgZPcUwMZRhztYDjFBiZQMybeOlADjhjlCI39KiLmJgu3aT1NSNFHIv+s+fHBpFYFPLc7z60AGwxjrvVu9PBKx7W79DUURZX8vPyHv6UgYxy4J8ygBV+TDK+fUUvmscEryehpBJGrcJnPWl8zcAN20joKAH7URlVhkN3p5zHj58ikA3xEdGqNQSNwO4jt6UAPLbRkN83aoxtXq2ST+VK5eMq23vUgXyBukH3jxQAzaAxAfA9acHKqVUAA9x3p8ExVXDw71bofSmJGrBsPnH6UAIkYdPkOGppIjfG/J7iniZoJAwUBelOnjVmEq4IPJAoAarDOAcev0pQm5wqkYPajzbdhwpDUI6ZxtIY96AJHV4t0crbVPTFQhxCSq8+lKSwbY7bj2pgfEhDfe7UAT28+0sXj3N2FMdSW3DhuwoEuH64Ip/2kYzsyaABXYjk4/vVIFDc4xjoajTMzc/IPSneYQ5UH5aABiZSXY/dHAqaMOULbsD0NVy5WQMnzAVZeQz8v8AJgcigBiopJBkPTgVLbMyrcAHP7uq5VRjyzgVZtPmt7rkY2feNAH2v+z/AJT/AIJ8eP26/wCmN29q+FreYjT4VHzKd2fbmvur9nq4Wb/gn34/hDYYXbHge1fCliwgtY2HzKd2RQBNGmw537h2qrr779Lc/wC0OPxFXf3bnONnoKoeIVxYMSfTj8RQBy9FFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUdhRQKANyOLMSbT2GaChjPBzxSxZ8pNpxlRmhlKY2tnIpvYDb0Tas2mksCPtK/zr67/AG6yD8Ivh3tP/Lsv8q+QtGG660wEgZuV/nX2B+3TGB8G/h4Qc/uF/lUS+JFdD4qdSCuPQZpjqTyvSnzLll7cCmN83Gcf1rRkjAwY9acV3HIPApm0ZGcVIwZflGOakBxk565FPaYeWBjmoijKnJFERJwxIzQA4D5OuM9qRVYqw3DHoaHkUdOT7UwYLcnbQAsSBhjJwDzQSyMQxyD0Ip6pwNpwpPJ9KecKwjJGD0PvQBEkLN8xPI6UOu7/AHe9PZip2sw4pWfaMhRmgBsDCE5b7hpGQRsefkammQvkuML6ClctgYwR6UAJ5hUYAyM9aG3SHpilTI4BpUcknjNAEgUg/eAAoUHcTGTu/SljALc9fSpWumR/kixQA1ZNjbX5HcetOmkzhV4Q0krJPyRtemhGUjI4oAckm3g8p60mOfQLyD603se4/u0kh3KP4SKAHiPILOfm6getHnl9yCLy48cn1qNCQvXcexqTa7DJYAdwaAImOwZLZHYUkcpycjt0oMW5s55HamyKV6+nagCQ3A4BGKedqgMpycdKhiRXTIPzinofMOFAHHWgBTOXjxkL7GmRghsqfwqSGNS+0kFjTmjMUgOQDjoKAGLIZDhl2etDWsTncHx6rTnkaTAK7W7UwLtG4nmgBVlkg+VW3A9Mdqa87q3zEZ70uO6nHr70rTK4C7OR3oAdHJ5i8YCnsaULLGPmxTF2x8ty1NDEfPI+4dhQBODtG4EOe9IBtBboajdlKDZlWzQrLL/EUI9aAHFy2D93PGKjbI4H3qlxgc4IHeomZsblI68igB0h2Y2tuPen43qNhyc85pETePkwCTzTHDBsb9o70AAXbJipBHGyn5tsgPBNNWGQjggrTM4+Uev3qANDTCWlkVsKQOT619U/st2ZutDu9zBFVsb8V8n2ab5GwcYH519ifsl3K23ha/Lp5ql+Rikxo9vtdR+xK0N1ia2OAH2/dqrqFsbG5ZxMrwOMqTziriaxZjIubNhaZ6kd6q6xa7YnuLJxc2zDIT+7SuWloTftqOH/AGNNDYhQ3ndhivzLeUiKIsfl2DB9K/Tb9tJc/sWaC5GCZgMV+Y0lviKLccZQVRmRguzNn73anZxnIzJig2pUkhu1KImVeevrQAqsT1PzY6U4Bc4Y5XHI9KarfNjHzetOWJw2UXcDwaAL8US39v8AZXwJFG6Nv6V6t8B/iZP4S1yG3uZD5eQvJ+6fWvI1LROCCBIOh9KuNI4Md/AeYziQCgaZ+oul+Ooxo8U9/pov7eWMfv4xkAe9QjxD4Wub2CWOxkgUfxds14F+yt8b4dRsz4a1N1mSXhPM7Cvb9S8Lxw6vmwdTAAT5Z6n6UluXbQ9L074vW3wz1Kxv0YGxnIjliP3WU9c/hXy//wAFEv2aLW2WH4weB4hNouoYe+igGQjHvxXpuu6XFrtmthfMsS7TtbOMH0r0P4E65bJpt78N/GAW70HVUaGMyjcELcD+dMhn48CXzCDuBU+lPMYOMEYPrXtf7Xn7OOofs3/FK508ws2hXrmSwucfKVznHp3rxaIlxyee1Ahirt5JXaKA23PPB6Urx9G3cdxThEwjDMcr2oAkV1X5ZBuHapY3jIAVdsnYmqSfM2XOE7VMm8MDkFc8E0AaVhKtveRmRso3DV1Hw68YSeAPGizI2I5GDAHowzXEKwyQT3yD71oFV1GAROwW7j5jb1oGj9KvBHjWDxX4etrrT51MmB5kbdjVu6k0+9bKOLe9jbcrjjaw6GvhD4PfGvUPh9qq207ERE4cP0Ir7K8Nat4f+I2nwXOnX8cV02N6lsc0my0rnt2s+FfDX7Wvw0bwR4xkjtvENshFhqTgZLY4Jb0r8xPjj+zR43/Z78RXFh4g0m4l01ZCsGpwrujkTseK+87XQtW02dJYZyzRtmOSI85r03RvjnFLaDRfiHocWu6d/q1aWMOVH1xQhOJ+OieXLkxMNo4Ifg/lUhh2AkYJHpX6m+M/2K/2fPi0Z9Q0TWR4c1GbnynbCgntivG9X/4JYSefI+jfE3Sfs+eFd+RRcmx8KoTvAJ+bsCcVK9xDFjzJFV+yr8xr7XtP+CXV3bz+ZrXxN0dbQH59jfNivUfDH7Nn7PfwZCX95cP4u1eH+BTujz7ihsaR8RfCP9n7x38cL5Y9A0iW301W/fapcLsjRe55r7F8N/Ab4c/szeDdSvtZ1C21vVpITuu2wUjfHQZqr8Zf2ztI8L6dJpmgxWul2arhbHTxtk47tivhv4m/GfxD8Url/tlw8GmZ4gU4yPekkDZ0lx+0fr2leKrl7WcNpfmnyyowAM8V9g/CT4raZ8QNHibz4kvio3Ke9fnLZeXsNvcLm3PBYdRXReGfFWp/DvUUu9PnaS16gqTwKdhpn6l6BqN/4J1WPVdIuQzDHmRA/LIPQiuY/aP/AGQ/D/7Tnh+fxt4AWHSvGsEZku9OA2idh1wPUmvHfgz8crDxtbxRT6glvfgACORuGNe/6Lrms+GL+LU9OuPJuFwQY2+RvrTBo/LjX/Deq+D9auNH16wl0nVYG2vBOu3JHvVb7PgnPBIziv1t8aeGfhb+1PZLYeOtNXQvEqD5NThUJubsSa+dfFP/AASg8VpftL4W8Yaff6a+TEC4LKvagmx8NTQAoS5CYHrXovwt/Zx8b/F3wnrXiPw/ZM1jpSEyOy43gDtX074d/wCCbUHgm9j1D4k+MbJrGL5zaW7gO/tU3xw/bC074aeFj8P/AIT2kNjD5ZhuJ1XhlPBzRcGj4NbzUkkhkXyZYmKSqexBqFwHIJJwOlXLuWS4ubmeXDTzuXkPYk9cVUdUYjYTv9DQIRcAcHB75pVkUHDnI7E1FIBH1GT3xQrRnvnHIoAuWu7bL/eAJGPSvsP9hqd00C6k6oWO4dxXx3pjjMxJz8vX04r7B/YmuPsuhXBXBJYkLSZSR9cX/k3Vql1byRG5j6rj5xXRfB6/S6+I2lOGV2BIPHeuC1fTZFkTUNNcebj97Ee9dB8E9Wtb/wCJmlCP/R5gW3oe54qSnsfCH/BQpgf2uPEm3acwJx+dfN7SK3Cphe9fQH7e9/Hf/tXeJ5InVgI1Xj1FfPcsxQAgfgO5qzMdISVAx8uahA8vIB4Pang71BDZbPIpuVyc9fSgCRHCggjimsxdcHAHrSBgTjGB2oIBwCMgnGaAHSL5WwNyD3pwZCpUj5c9aiCDPyvuZTwKUSxmQ4G3+9QA6JQJTk89h6VJIRuJB+Yd/SmyRkglCB7jvUioxhJGMCgCMxiQblfYw6mpVjlEZJwy9qhO1iW3cd/anMu7IWQ7fXPAoARv3iEEc/3qRRuXacAdqeoZFYb1K+/ekilxuRlxGepNAD45dqGNvm9DTVXAYM/JHFKqBCVVgo7bu9LmNc+aeccUAJGdpyBkDrU3mQMvyJg+tQPOM7UGOKeFKyKo5yOtACI5ikKno3FG5JJfLC7Nozk0ryKG27d7dj6UwyZkGfmJ4oAJcI6kstRuEZ8sTnFSLboG6/N6Uxz5b4fGT0oARSpQ7aUDaPkxlutKw8sAgDB9KarxxA7Mhj0zQA9DsU5G7HOKYrtEQzLlD0PpT1HlANnc3tQzCTGCNnf60ARu+ecbRSiURLwMynpQYJH2qcY7GkaUh+R8y96AHoXXDMRz19qRnwxKcD09aaX46jn71PE5BG1d47UACZlxtkyc9KVC0j4Z+AelKhV2HHln+dPKBXBkHfgigBTNIqlQeM8VCFERyr8E81YMTk5TBOelQOXDYEeD3p2Am2kAhvmjPQ+9Rj9yfvbge3pTSny/M/y08RvKu0YK9jSAJHULhRuPbFNSVlGONx7CkijIyM/OKc52ndjJ70AMLEg5IDDpmpEdZWVf+WnqKUeU6lmQnPWowygkBPl9aAFlXJ+UgBep9aEY4znjt6U5IfNDMpAbsD3psaSAsrjHuOlACuxZ/vfNintG8Y3Ehh3zSeQ6A42nPcUhXd95zuHRfWgCaOPzF3xsMDt708sWUgjLYqGQlRgMFJ6gVIj4BG7NACK7xfK8ZKnuKtWW1orsodoCHr16VAsrA8MGxU1o3mw3bZCny+KAPtz9niWFf+Cf3j4cbzct29q+E7EYsowxG3LY47192/s6qF/4J/eP2+Un7Sw/8dr4XsVaSxjXK7QWJ+uaAHtOEdd/JrN8QDNizZyMjH51pvIhO0Dc+OtZOt/JYMucjIP60Ac5RRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFAoooA34OYUGPm2jH5UvlsFJxx3pYomVI2H3doz+VNYMGLE8dhTewmbelI2/TiCM/aFx+Yr66/bdglHwV+HskhHEK/yr5E0tismmuMEi4Xg/Wvsf8AbnufO+Bfw8JVRiFf5VL3L6HxFeKWZeQPlquSVYLkZ65q1drvKNn5QoqEyoVwUz70ySEsd4IbJqVpGBBGCe+e1RiLadynIFTRjcOo96AGhUmbl8H9KY6/MFyD9KVSobbxikMYHJPFAChMfd69805WGCMZoy23J5btikRcEkcnvQBKi5GAw29xTpW3gdMr0qIAu3yg88GlLc7QOlADWcsw46c5p0YMhGB1NKpDcYw3ek3ZQ7Dz2oASVGQNuGRSIySIBzmnpIdp3EZHX3pDGDyPlTuBQAI0cKNjk96I2YnAHJ6/Sh9m0cYpY1AQ4PPrQA4goPftT1kyMdW71FGzBSc7mqaMKhB4JPUUAM8wIxwNzdqGLHLE8jtUjR+YCyfe9KhaQLww+buaAJMkLuI4p6bXOFI56g0wYKkFuOw96QTFDkJh+maAHOUjk2qQBUUqmQsWboO1SOmSQSORmmvcKBswA2KAI1DFd26lOW5Y84pdoXHTHenMNzBeOlAEa7i4IOKkadeFCY96ZIpj25PPpUgUuMcCgB3lFcMvI9aQMM53ZPpQQ0Y2h845IprSRmPcVKv2oAlckruf5T2qDO84HGKfyVDNyKQSlThlGaAGwxtcNjcFC09QobDY9qjJJcHOPanMo+8vJoAfIhLDPI70ZESDcu8VC0hQZB5qVCrFSDkd6AH73dfmUAj7tRuwiIbg5Pb1qRGMbEk5WmsRI+QAVPGfSgBgQk4V+epFN3Cb5eI2X170AOzBNn3TndSyR78byBzzQAbWZMEhfp3p8UZxtHzNS7GQA/fXPFOViWwMLnqaAGOedofA707aMYYfTHrUvlgEDACnuO9CxHJz8p7E96ALlgqlWKgCQDmvq/8AZfZtN8KXVyMCNpOS3TFfJlkryTxwxDM0jbAo5JzX3V8DfBz6d4SGlzhI3mTeVYYOcUmUkd9d3i/ZAYtlxC4B2471x99rU2lyXDpvibGBE3Q1YhW60a7ltpW2vE/7oHvWprmqQSWEck1opumGDkdRU2NGiz+1/qE19+xR4fdwpBm5I7c1+cDYNrCpwQUGa/VD4s6AvxV/Yb1230+ITX2jkSiJBkjnmvythY/Y1VwFeP8AduD2IqzJojNvjJD4Hao2DIDu6GrUTqpw/wCFMmjEjkn0oEQIPkOzninRq5wqna3pSvbmNdwOOKarkYZfzoAepZWJOOOoqxBfi3Y5x5Dja6/1qgJGydxGD696TOTkn5cfd9aANTRPEN14L163vbWQrGrBlcHj6V96/DL4yf8ACwPCdv5e1tQiQKSvXFfn1AyXCG2nx5TfcJ7Guw+FfxG1D4Y+JIR5h8hm5B6FaTRSZ96alb372n2mCTzQo+cd1NU9F8YTXY8q5Bjv4jmN8Y5FbPgfxVZ+M9DS806SIq0eZUHXNOitbLUZwhgW1vAflcjAYUrlWPV/E3gzSP2xfgpd+FNUaP8A4SvT4t1nO338gcAH36V+Rvi7wTq/w98W6h4a1q1e01SxlMbJIMFgD1Ffp54av7zwT4gttT01ngnhILYP3x7Vs/tSfs16R+154EHjLwlHFbeOrCLdLEnDTYHO71ppktWPyUeE56Lljkj0pXjaVcA5rb1XQ7/w7q9zo2rWkmn6tbMUmgmQhhiqksIjwB8uOtMVzNji3LslGE9aVFJXCH5exqWfDgLnI70wjbjnA9KBXuR7RHjnnNKX+YYOCOjDtSNPDtA2kHNN81QCMZz0oA0Fkh1FBFdnYy9JFrY0PxTrfhaRWs55RbqcgxseRXMLlmySGQdvSrVveSwncjjjop6UNDTsfRfgv9sDU9JijttU3yQqcAj71emWn7X3h6+gxdxo49D1r40XWI5FxLbITnqops/9lzjd5Loe+KVi+Y+4If2kfAN1Dnz0tZM55OKwNT/ao8NabOwgvXeLPOxiTXxx9k0oLko7emetNaSziXdDbbzno/NFg5j6U8ZftZWN9Ew0iG7uXzyTnbXjfiv41eJfELOqSCxiPJ2H5iK42XVppQ0ccUcGeyDBqmeScnc3fdQkS3cSaV7iQ3EztNMx5dznFKUZvmBGG5pypsTcpylMXIUnPPpTJuTA4ABKlT1qeG58lSpIkg7qapiRV7c9xT0cODgAHsKALEKXNlci70idoZBztRiDXsvgD9rLxb4NjS11OJrq1Qcb+TXiiFhJlG2SY61o2+qyoRvRZQOpIoBM+qv+Gz9G1JV+0adJFMBjcFIOamm/bkutMtjHpL3qErtwCa+XDrBKZ+zRZ7HFRx644PyQxqO7Y6UDuen+Pv2hvFnjtSlxezCJxyZGO4V5TNIwbLyeY7cl2OTmoZLpjMWkbg9NvSo5Zdq/wtn9KAbHMxI6gGq8iKzbo+o61HG7cgsDnv6Uod2BCjFAhxzIjZxkDnNRxeWQAE+YcigPsDAjcxHGKltyQV+YBqAJ7SFjazycA46V9i/seWE9r4KkvY4xIyEk4HavkqNFTS5yBgsa++P2N7KKw+G1vu2/vhhtwpMqJ6zomvadrxVEcW0q/Kwbua7L4L+ErS78eXOpiePzNPhkkKIeenWuXvtI0u3klLxrDMBuTZ3qfwg1t4N+HfxB8XpeGOWG1KKScdiKSKbPy8+P3iJ/E/x18Y6gXDgXbxqT3ANcBEdygnk+npVzUbltX1bU7/OZLm7eUMe+TmqchMRyACf4sVRmKfkbnIB7jvSIxVt/GKQzMgBKZHrTpEDlSpwvegB7uc7VxzQpJ7gsO1IUJwBjI70058zOeB1oAeCY5NwXr1qQNEwYOuFPQimhXVCSf3ZPWnQnDeXw2emaAHhugU55+7TJmKOcHHqKfKBGxLDc3bbUeTywIPrQAeZEueNxPpSb1dCOg9KRnXB2pye46U2QpHHufqfSgCYRIASpLjtik80BgGIP07Um7CKFICmnxAB8kgr60ALIFlba7AcfKw7UjKqZUfNx1PSmNGdxJI9sU/J2HeAT/s0AKyfISWwCPlb3oVnVAp4cdKZ53zbSM5H5VY8gxoHzlD/FQA1v3cilcNnrTZxGZMpknHOKcWXG1cY65NR7THjkfNQApmyQoTL+tNcKW+bk45PpUxWVRwAwxzimrCXU5XauOc0AQqB0HK1KIo7hApG1hnFEcsYOAOBTJAJGBB246UABBAWNuDSZMY+YYU8YPenlGWPJwc00AqAQQ69/agBkkbRKpL/KaUnADRkEd/WmOMyAs+Y/anrlWBGMDtQA8RrIAUwD3BpqZRt3AHp61LiO7C5PksOnvSNG6MC5BPagBVO9fm+fnj2ojmIOGXeuetClA2AeT1ApV81CdoDbuMDt70AShoQCA/lv1HvTCJVXeDvY9hUbRkLzhufxqRR5aFkOcnmgBpR5BiMAH+L2FIWPl4RsAH73vSmQKu7nr0FNJXG88c/doASPLcj72eanwynnAHvUaAq4cHgnkCpQWLsQQ2eimgBkrqDgcL3NJlsbV+bPQetPnAQbRjJ6rTDGT91tp70AEgZQcKVYdT2oYMUw79elII5WyEbfjrUWxxIVzgd80ASLEyAFn3MOg9KSQMG5I3Ypyhycgr9RSujqSD83HWgADKi5I3KevrQNgJIbt0pQ5CjcnynqKlRoXBYIQy9KAGb04KjJ7j0qa2YNFcgEfc6+hqJI1BJTjPWprSM/ZrvIC/LxmgaPsr4Ea61n+wN45gCZD3jDP4V8V2RD2EQBw2Sf1NfavwZljP8AwT78YoFTeLwknv0r4lsWKWcTjp82aAZYZNihlw2Rg5rM1ps2D9DyP5itVSJGCxYIPWsrXh/oj9iCOPxFAjnKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoAoo7UAdDbEhEXOBtGM/SnSZwTjmo0ciGMDuo/lS+Ydxy2eKroC2NzSc79OBA5uV4/GvsH9uQY+BHw/wAL0hX+VfHuhgNLp+SOLgfzr7A/bklB+BngJB3hX+VRL4i1sfEUrH5Qfm+XOKPmwGBGPTvTirFlCgbsCkMalxk4PpVMgRHUAjGTQg2Zxzmj7mRnmlRHJGxuD1pARCIK2QuQevtUg2grv69qew8tcLhj3pNizIFbGfX0oAY2GYAHC0q/OdvQ/wA6GGwYYA46e9NMYCh92X9PSgCxEZIuAABnvTnUt8x61F1wd+f6UvVQu7Iz1oAVGTJA5bvUCnY7Djy+oPvUhTB+TGTwfpTATE5V1DLQA5Y2bLnkDvTnbePlHNMVZOWXlT2p6tlgCu0d6AHR+XtJlGT2ppYFsKuBQxLHtsFPwzKSANvY0ANY7AW4oV2PXr2ojKnKuMrmiSKNXwjHnv2FACA4kyjYbuKeWByWwV9aDFg5GM+vrTTHI2Wxlf7tADlXOCOT2p5lYHlR9KiTcjA7cGiQlyWLc46UASSShc5Tg9aa4RuQARigEhcggZppG9uuzH5GgBAAw46f3aXYd2d3FGSHGMKehqQ4J2BccdaAIigZhuJOOmalwVXNJMoiVVHfv6U9JMkK3A9fWgCIAg9BzU7EEBSBu7VFI3zYGOO5pY1aQ7sbsUAMZ2U+WwAz0p2DGpdjnHf0odAzDeMmgttj2YDr6HvQAwyLJwo2j19aRiykHoKVhkA4UY7UkgPytnn0oAljw8ZYrlR1ohdWVlRCPrTFfGADjPano5TgdaAEYmNRkigQlhuRxjutNYlj83NMCKDwce1AEilpDhm2r2+tK0exQMbsmhTvXpimoZCwGM80ASRSNEMHlSenpUgRcnGMH73tTW3IxVxyaksyqXG2XAQjrQBu6H4Q1zxJCW0rQ7vVYEPMltGWA9jxXU6T+z78T/EkyxWHgrUgxPys8JGB9a7j9nz9rPxT+zlouoaX4e0jTtTtryTzTJdxBmU+g4rttZ/4KWfF3UEaO1t7PTjjrbxgYoA0Phj+xnqHw8uI/Efjwxx3iDdBY5GQe2RXrVhqaW+otNtEb4ABHQV8e+Iv2rPiV4ouGm1G5FzIT95z0rFg+O/jYMcurH07VLVyk7H29dTw6xcyG4VUmQ/Ix71oappn23RlK7HuUHGPSvhqT48eNWOH2exFA+PnjiAMI7ghjjIU1PKx8x+hHwE8fR+BtYvdL1iMPoeqIYLqJuVAPGcV86ftIfsB+J9A1u/8SfDdV8TeGL+QzrBbfPJFnkjAr5zvPjl44mkZnkOW6810vgj9sT4teAJ/+JXq8sMR6wbsqatKxLZxt58H/Hmmv5V74L1WN1wDmB+v5Vha1oWq+H3CatpdzpjvwouEK7vzr6gtf+Ck3xUBQ3umabfMpGTNCDnFcT+1T+1LaftJ6XoJOhwaRq1mQLkwpsVqYjwVz5vJPQfd9aiwQ2AAOOlKZAHYDv8AxU3duGScH+dAEfDPg4zil2MeGXJ7e1I2C3PDYpyls8+nWgCNotpAPJPSr6pDqMAt5ziccRyVn8u/Jxjv607G0g5+boKATPTfhF8XtR+GesJb3Mjpbscc9GFfbfg/4i+HvHlnbzSXUMVyq/Lhhya/O2DUYLq2+y6jF5i/wSj7w9jVzTXv9IxJpepMqjou7kUNFJn6jwXFtLZL9onhbaMKysKg8O+OtV8A60uoaZPtQNyqkbW+tfnDD8SvGVtGF/tWYJ6BjT5Piv4y27TqU23+6WpJFcyZ+m/xJ+GPwk/bDsRJrCxeGfGAX/j9jxH5jdsnvXx78Tf+Cb/xW8Fl5dBEPiTS+TE9t8z7e1fPk3xY8arIjjUX3joQ5BWun0D9rT4y+GAqWPi29EQ6ReZkYpkaGZqH7N/xY01/LuvAmq7gcDbAcfyrhvEHh/VfCt89hrdhLp16g3eTONrY9celfSnhz/gor8bdBnhnub6PU41YFop1DZHvXl37R/xquv2hfiDD4uvLKPT7jyFikijG1Sw70CPKJMDaVwzZ5ApbeRXba45zTyjI5OAue/rUbSY5I5zQBKbVlYlfuZ5okG8gY9s0wGSVOX2jNSw+jdM8GgBUJxgDODSeYcE5/Ckj+WbcDnB5FIzh5SVA+hoHck88JjbjJOKY8mCcH5h0FN3qisSBu7U1D8pYAF/T1oEJKm4hs8nrSuo2cjIHWpD0zxmmqpYk4yPT1oASNWlOFOFHQU2QbHyVwe9KWG4/w1MA209GHrQBXjdHyxGGFTRKGO9McdQaVVDjfgZHakChDuX7vcCgBEjLOSpHPUVKrAEjv6VGzneNp2k9KGXG7njHNAEkkrKAejDpSAuUP909aZvDRgtyw6UKwl4B24/h9aAAvkYOAO1LDvXIC7kPftSFY1YZBPHTtTi7Tjah8tcfnQA0wyAkhd0fcim/O+W3bcdKkQy+XtVsY6r60KhiIYgFe4PagBIY2AJHzE8Gp0Kbdq4yOc1FvznBCfTvSSSgqAq4OcUAa8IeSyijyN0jgYr9Df2ebM6V4I0y1nURo6ht3viviD4S+EZfG/iqws44wbeNg0jEdK/RvQPCyr4ft7O2ZFkiQeWPU4pMov60ZvNJfaZD8sbe3avPv2x/Fsfwt/ZpstDjKpqPiCQiVVPJFeheELW+1fWFs9VgFvb2g82SeToAvNfBH7Yfxq/4XF8Y7pLSUnQtIH2a3Td8u5TgkflTBs8OFsILeOEdQPm+tV5FwACBn3qeW43bwB+8J61AGPSUD2NBJGud21scUi8McYweKcFIJyBupYowz44Cjk0AIQchcfLmkb7pDJkZwKCxkDID8vrSoJ414w4HFADiGlVcEDbxtNOYlEwwAz3FPliKgYXAbqaYkZAyXyQeKACNAnzI2589PSmFCS2MZJ5FSJlXLAAZ70x8xsxAyp6tQAA4BBH/ANalSRGBSSP5aTeSM7OPWgHL4zhaAJUKKpDxnb2pJCNgQLtX1pwc/wATc9hTPMYSESfMPSgBAyHjOf8AapAF3YJ+TvUpuACQsACe4o85XU5TaaAGl4iSqrxjrT45pFTY43RelRqSh2nG096kdtrgMcrjigAHlqTjnI/KkdVkG0HjHJpzyKq8AAGmBXRwIxuD9aADy3EYUS5T3pE85zsLn296Vv3Z243cUxnIZePpQA8JtbaSFYDmkZs44GBQ0OPmQ72/ioKRlASfm7igBFkZsdCo6gUxpI4iBGvynqDQV29OPTbSsrMg3YUj0oAbK6qR+7wP4aVWE4Uj5AOopEkYfe+YevpSFkxjHXpQBNGAW+cDaPWmISZMseOwFNRCOrAg8H2okl24GApHQ+tAE2fLb1c/yppV1O7f17U4SK0XmYy2cYpFm80fcxzQA8I20Ec80rfuoyQvzd6VGy2FPvx608xsFyG5J5FAEYDbd6rn1X2ojwPnkAKA9B1FNLSIdpyFJ4amMMhsnJ/nQBOMNl4vmOeFpXieUnawV+5Paq0L+V0wq+venySGRWIOwj070AL5ckZO5d/+161JG6kZfgeg60+MyXEIDPtA7Cq5yk5wF/GgC0Jo04i+Unq3pUL5BJB3CmIyFihBUnqaWeNQQkb5oAVmGBs4B60mWbvwOlNAKMVxkGnFQF+YbRQBKsjMvbdSorMSVYDjkVHH5YQsDk00xu7bkbBx0oAsKyAglfnUZ4qe0kN1bXkb4UhSwqvHKzELImPf2qxboClww5TYRQNH2X8EoFH/AAT78ayEDP2w8/hXxLYRkWMRC5T5icfWvtP4LmUfsAeNcEbftp/lXxhYOV0+EKezdPrQIk+0Ow/dxiPFZevHdYuT1yM/mKvyTOxwRhu1ZmtbvsD5OOR+PIoA52iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKXtTQG9Am2NOOqj+VDY2tt7dqjQEwpyd20fypxbCHPYZyKOoGvozFWsjj/AJeFP619eftsxsPgd8PWJ3Awqf0r5J0IA3emI2Cr3Ck59M19k/tzWwi+AngJlRdoiXGPpUP4jTofEc21kUL8rY6+tVAoGckZqaZCEVvUVV5xtx171RmTR4jU7sH0ppBGH3bRTl2kYI+lI6lQCw47UAKGychc+ppRsf7gx9aaHLcAHafSnQq5xu2hQeDQBJ5e0ep75qGN0DFQCpPc0pbcx6DHp3pWIVcuKAEdBg9vpSqQq4HOO9OAEq4U5Pv6UwHaeRgDoPWgBQzRncAMVHuYZc8jPT0qR92AQQc0FRgDtnkUAPJZE3RNgnrSiUSL+8GPWmqrRHJHy04ukhyAN1ADTjbgEFfelbLIFXhAaHCKmBwzHHtQp2AoxBX1FAAfnJH8XYUqNsUqwoEkZ+ZFJYdzSBZHfdtDN2FADlA52nj0pdxGdvBqMq6AluHNCK0bHGDmgCT98WGMNQysBkqAfSlGWGd2Md6QKJDnfyO1ADcCRMHg0uRyJF3ACnLA5OTgr3zTSxjkPRlxyaAHLIqHLJuXHFLJcmYgGPbxwaR2Bwc5FNcHknrjigBVXtkFTQ8Zcbf7vNRBWLDnipd4XG45NABCd45HzihDyQG2+1NEuGBA4NSPsJDKOaAGFCo65pQcHIGVHWghnOR93uKIXDPscbUPf1oACh++hBPrSvsQAsMvTCAj7UbbgcGkGGA7sKABWjyCRyaeOSO1MJL4zinR7SR5nGOlADXOGA60u5WGduCKUsG5PBHSkBBzigBq5Zjg4akDOTjdtcdKaAd4waldgMZGWz1oAd5jbfnOSDS7vN4IxzUaZcngBqJG8rAYc0ATrLIpKrKQPSia5n+6ZT9aiDhwdoGfejhOpyaAHLLN3kP1p6zyh/8AWnPrUMgyAelLHuIPrQBZN1KiEmYmohdzb8+cQT0FQO5bK5prlhtyOexoAnN3MSczEmhbiYHPmEmo1U4OBz605QRwQCaAJlu7iMELJuyOTTGkZwS2MmmIxicrt+U9ae7/AMWBxQAkT5JU4wRT9hIORg44NRBGdvkA6VI4kZNvpQA1ozxjB96QB+R1GKQ7oUwRwe9IWZvunHFACcK4LjHpTjKWOABj3ppXDDcc1IF+YHA20ACSCI+x7CpFdQ2UYq1QOCJfu5z0px+VgWUA0AWGup0BJlLAD8qDeSsu7zSag+Y5YDcajG4NtAxmgCTzJQc+YWz3pWaW2cbZCc96QqsZCsfm608jDK33vagBUlmhG5ZGLH1qyrNKoJw7DqtQxOACz9M02Ob94SnyjufWgBzET9TtA7VG4DIAcKAevrT2dJ8cBSDUM+JHAxgCgBPNRRtZScnjFPJ+XODgfw0KTGu3AxnrSby+cGgB6SRnttaoyjDJHDZ496mwAuWGTULHnPRfegBzyoAFkTB65pFAPMYwx6UoRmOQox6HvSs5UlVABPcUAR7XRs9SKcJDuGFw/alVX7ckd6d5ZzwuD60AILd2DGTp60oiG07WI+tI5OSGY5pH/dpuGCfQ9qAHg7CMdab5gDkMPkPelJzGCvJPX2oAyhXjnvQAOiqck5B6Uh5BUcGhV8xdrNgjpilClo2LYVh0oAVGQpgDDDsaeRnrtz7U1mEkAUptlHRl708BGUcbcfeoARHUHLDJAqaNlkYKUwx6VCZUAwgy2KdFIQy7j8/UH0oAuw2EtzfWlvEAbmeURR8Z5PTNbvj34deIvhtcww+JdJnsUuF3RXBQhGHsaqeHoZtR8T6Jb2xEc7XSFX7hs9a/ZDR7Pwv8Rvh1pPh/4m+G7bVUgtVSO72jcvHXNA7H4m74WQESJgdMVNpVlc69qMFlpsDXd1Kdu1BnHvX6eeJ/+CeHwGvbl76LxJfabC2WEC4wo9K5+z+Gvwp+DLsvhGP7deqCDeXgAYfSk2NLucN+z38ID4Hs4mkjEt5Ku6Y4yUr2nVdOu4DDPY3ZjlXoN2AK8v1T43aH4WVpmu0ilbJcRt1r5/8Aiv8AtZarr8Ulh4cBs0fIa4GQfwqSmes/tKftXXPhnwhceCNGuEuNevQRfX0ZyYh6A18QQMQNjtvLMWaQ9SfU02WWWeZ555GmuZSWkmkOSTQGVcKRz2A71ZDLTq/yttXpggdxTMBl+YDaDx60qmTy2OQWPQ+lNClMEkN65oENbPmggYFMcmM7lGAetTA7fkJBB6ZpMMcg4IoAjZdyrs+6TzUicRspbCHvTV/dqwAHsKSM8fOAR/dFAEypIIhtbeM96bu3ZUBQfUUgj3fcl2+xprjY23HPrQAh3KcZzUiBmyFx9DTJGEX3jh6lW3dl3I/PtQAhllizHsG0+lKF4O8BcVG+5j5buQe+aeyyxD96A8Y7rQA5ZEBy69OhqNjufzCPmHSnFgVAAyDTA4WUd170ASiVsFhj3pqEs2ONnemTxlpNwYAdsUwElimFD4+8KAJ41WSYj+BexpJMSMccBegpiSiRgGXEo4GO9Kp2ZJC5PWgBVkHPmJ8uKIpN0gWP5M/dzQZ1U7Y1yp67qPL3thiFbHysO1ADiDHIVkA3kck0qsrY3dqj5dwrHPH3qdJiOVRjGRyPUUABcEkJ0xTwkciZfKqP7tBReCp4ppcL8oBbPrQA1jHkNCSdvY96kAZAGlTJPT2qEYGNw246EVJE7KcM25TQANETEysBg9KhgUxptIBbtUmS8gBPy9j604kyYUJz60AQyx7gOcNSCQNhWUZB4ocNvVSPmpXzgYTI7mgBN2MgDbzUokI+7ye4qLA2ZPGKdF8rb8DPYUATK6PgFdjA5zUsal1YM2M9D3qqXEmCDtIPanSth0K9Cfmz0oAk3tEvlt86Z+8e1GxRwCAM0fdBUqDGaJFDx7ANiZ+9QAqW8cZMkhLIOw7mml45X2qp57ntTl3IMAZIpwaMqVxtLHk9xQAjsyKdq/KeuKQkAfdJT3ojEqBsEFc96S4cmPBbHpjuaAHSyokW3H7w9DUCoVUnPzdqkQl4trKFcdqkXYOXGWHO3tQAhkLxbAu1v71KY3MeNwdB396Yzl2LMQinggdqaIgGJVywHQH1oAk+QOCOncU1iJGyD5YFNK/N8wCE8VI4VV6ZboBQAoUnOJARjpVq0cxxzqSDlOn4VnrbEDd5mSe1W7NiYbjgb9tA0fY/we1SKH9gXxjA8f3r1sH8K+MrQgafAONxDY/Ovs74MxpL+wL4zDIhYXrYbuOK+KrTb9iiB5+9QDLW8kgMRuPH4Vna8u2xcEdxg/iKuGMEgqwPv6VR1qUnT3RuuR/MUCObooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiigU0BuxnbHGeGG0fypZW3IwHANJENsSZGRtGaV8SKeOlHUT3NjR4Xml09V4Pnrj86+v/ANta3ubb4CeAxNL5g8pcDOccV8k+H/8Aj707IHEwPX3r61/bggmb4J+BJj/q/JUbc+1Q/iNL6HxXMDKgDttXbxUSkbfm4A6e9WZNgRPMG75eKjURvgjHHQGqIIVIccdf5VIAzqAx3inBVGSwA+nemCTI+RcJ3oAcjtCwXGUPWnZ+bao2xnvTNxwAOPrT0KlcHt3oAY7LgEDpShS3zYFNxggqN2al8sqv9KAIlVozlflNPZmlIOPmFN3EEbumfypWJVhg5560ALvjVQrqRuPUU0rGgxk4znJpyAgk8MfentEWA6HnoO1ACCJ8ZLbl7CmDBJDLgelKshDMoOCOh9KRlIUknBPf1oAUkhTvXIpqsrdCePWlCyCIMSCR0HrSNs2jcMA9aAHiVSRyMClAJBKN5YqNdmOBkDpTmXepxwKAHphQSx308IBl92Pao0VQMjqKfFGrAmV9o7UAIGBU7OCeoqJgYmyFyalRQkhP8FPmkeNS0aBxjvQBEZ2bjdijGz5SMof4vSkaRHXlCJKeql49wOVH8Pc0AJC6xPt4ZT0pzFnkJIGzFIxgAHHzHjB7UgjZuF5UCgBQhkGR8pHpTVQufn646UrbVIIO04pGU7gc5OOtADlGCAwp0S+a5QYpqyFxtYU5W+TgfMKAI2R7aY7eRjpSysJNuRtx0p6yEMDjnvStsGGPI9PSgCNnLrkjp0pkas5woB+tOaNpGHdaRI33fLxigBNpLbSAKCSeoAxRJvBDOOKTcpZSBQA8gOqkYBB600fLlc5Y96EVIyWbO00uwY3A8elACA8YVee5pAxBwVyacdyxArjJPNMCtgkfM3p6UALKfLwyjBNJuOeeaeF3LjaCe/NEkI8s4HNACIQafnAJAGc1Eowi4HOacrdc4P40AKXLA7lzQoKtkde9Bfe2AMfSlWMqCCeaAEk2jkfe9KTe8hG7oKeAudpHzetJg5xjOOaAHs7KMLznrTgNowR1/ipu0g5x1/SlaFlU7myvagAGQjEdBzSp+8II79aHxsBX8R61ExaKUFBxQA8EiVgpwKcQVU4bNMIIJbHWkWRTlcUAKHdsjqMUiDHXgHv6UElm2gdKY6P0OMGgB+MsMAYFH71wdvAFMKF0wgxipF3YHGMds9aAABtoJbFGGA3Od9IyFSHHzevtS5J5U49aABCR0+Whnc9UyfWkPyqXxQpknH38CgBQC7A43H1p7nyyq4z/AEphYw8LQjh8Efe75oAmQbUH8WTSOQAABnn8aYreWSEHXrTzIIsGOMSOexoATABBcbTngetNZI1cMzHaeuKkeRpQBKo3k8Fe1QBNhKHBLHHFAEztA+0ZOc0iruJC/Nj1oKBWCkDcvIzTQ2DvXg0AOUfNyOKUqBwBkHr7U9DkFjwxqFEKZOckmgBWV24DYFOWQBSjLkdzUKks5QycetPLFSRjp0oAed2Pl4PakEjMu3dz3pWdPlJ4B601xkFgeB0x1oACdoJxuxSowbkjB9KaJztxgY9DS/OM5GT2NAD9m0E/dFNSSOT5elCupyXbfx0p7um0BUwT+lADSOfu9KQgudx7UuWVTk7hS4DruAx65oAcHxBuwDSsA4DA4OOlLIsYtRj71NCkhSB83YUAICx/gxxzS8gHaBjHNIfMdjg7MDn3pFBk7YA6+9AHQWtykBtLywk8u9tiGUe9eqWX7WfxJsII4PtTOiLgZ9K8Mwd26MmMgdala8uyoHn0Duev337RvxA14MJLkR577sVy2p+OPFGqg/atZKDuqmuDWSdyS05PtmlgfzN2csw7k0rFcxqXMySA/aLuW6Y9cnjNZ8qq2MHao6UwSFztPA70SxMiglcr29qZLY35FGXX6VGgIfcqZqUB3wFIYenpTY97sUVtnqRQIVFKBgBjf1pYnZl8tgMZ4NNKHAJbcQetEK+Zuzg+3cUATLCCvzjoetRk5farYGfvU4O5HTzADjaajkaMLtwSc9PSgCTarD958rA/KfWmxxlGLBst6U9SWjXODjp60Axucr8r+9ADUK727NRGSVYEZIORQ64kBUD3qMyAP8o+b1zxQA94sjeOTnkelMXKyFkbae4qVkk2lkAL96jRt4KsmPU0AOYAk7juJ7inI0kSsQQV/u1GFTdtA6U5iRmgBXm3YAG1j39KXA2leNvc+tIF6qR8p6n0qONRKWUHGKAJFlTbsYDHtTWMaDYqlmPQntSEEHhRmnAFM55zQAH5uAoDY6mlCnhdu4mlx5gIcYA700ybHAj546mgBWY/dP38YowQAowwNIQQdzD5scNSRsrE87Tjn3oAeGTcCwwR0oIO7MmGyMCmtGoxtf5scChsrgn5noAQMsQzIcdgKaH8rofMJ7U8Yfl03cc/SiIo/wBxdu3pQAojJUMTu9vSkibzMqOF9fWhmMSHA+UjkU+NVkjGAAo9aAGFSQAR07095WK4TjFIT5wCk4ApyqHJU8YoAhyzDnmpkQtHsVxn1pkSAyYZwAP1pNoafaOFHv1oATygmdzeYe49KBsLAKcmghiW2AK3v3oJMaZIxmgBTxx5efenAvKnK5UH7tLEJJEADhVz1NOIkiyTgj1oAYqM/wB58JnpTgHjLLwyHt3p5CjDqcOTyKFjZt25fL9GoAazsoCkbB/eoXPQgY7E96EY7SJhuUHgjrT2jVcmNs+3pQASLu6ny36A01NyKUwGYcg+tK7K42t96gK7YKLkigCQMp5cYek27skCoZELS7vut6VKjlwcH5+1ADWDFwoUH1pGOx8IuCO/YVJhg+4nn+dKu4EptGGoAZjzlYt94DtSx7yMIAf6USeWMqB89IkhDYzt96AHIpDfOoU9cetXLSHdb3L7VyF7VTeNnflvMbHFWbNGFtdEn/lmRg0DR9i/B9fK/wCCffjCQYUtfFc568V8U2IxaRrJwOSGr7R+FWqWp/4J7+K7do380agcEfSvjC2iZ7KEbe5II7UCJIjGjD5cr6Vn67hrR2yOowB9RWhv248wfMOgFZutOJLNj05H86AOeooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACl7D6UlHamgN+NH8lB2IFIxxgFc06Pd5UeOflFKy5Ukrjil1A1tLcxy2TBdw85e/vX1n+2RdTXHwG8CB0wvlLj8q+TtCCyTWSbcjzl/nX11+2wY0+BHgGNF24hU/pUv4i+h8Wy/KYx6ih4059qZM+3YPvZApMFj16VRAqgNkg4AHSl8tiFkVcAUAhiMfI1SeTKADu49KAGSHcykLtz1oVFZgrHC0qqXBXGc96EXd+7ZcMOlAA/UBeMd6bvKnJ+ajdghdmCOtPEYVd4GQeq0ARSYRctyD0pWQwqmTuRjSSIIQN53bjwPSmglTtJyp6D0oAfks2FOB6+tJEpO4BsHNKNqqMnjNCjAz78GgByIyufl59aCI+QpL+opxDZBJ4/nRtfqq7Wz19aAD5h977tIxEYKsMhqduGDuH1FMdvNUhhx2oAFABwBuXsKWRwp243A0zdtAwKcEAHzDGaAFC7R8ozR5LTZKjOO1IrfIdvKjrT0d0yVGKAFSJRnedh9KZIWOSD8g7U9iJgWPLComdXOQOnUetAD4n2EkcqRyKd5m1t8Yxx0qvP8oGOO9ShsKGbgUASZWUZ24Y8GkjJjY46U7duHA4PFKF80cDletADWTBLHoRSLgEADjFBk2ttx1pTuBwOmKAHCONR8gy57UxUZGJA3kinwsHBYjLAYpE3GQKooARCeQB8x60CNlkBYABuM56UNH5T/NyT3pnlMCSXyp6e1AChgJCoHApWgZzvDbSO1NIDYGMEd/WpAThc9qAIgzjCsOD196QxtjHl7QO/rUwjMk0ahgiMcFyOgruIPAnhacRF/F4jYrlxtPyn0oA4YIZVC7MLSqrEBSmMHg+tehx/DrwgenjZc/7pqRvhn4WwCvjZDzx8tAHm/kM7HKlcd/WkERiBOw5PevRk+G3hon5/GaY9dtTj4beE3wG8brjP92gDzFI3ABC4IPNPJO1iF/CvTR8MPBo5PjhR/wGpB8J/B8gOPHUf4igDykROR8q9e1SfY5NhLRkGvRdH+Gvhe9uJ45vGaW6xnCuy/erZj+EPhKU5fx9EBn0oA8dSF0YjYRTzG7nbtP1r2CX4Q+Ckznx/H+C1EPhR4I5/wCK/jz/ALtAHlP2dvu7STTCskP/ACzPFeu/8Kr8Djj/AIWJGp/3DTD8KvAjD/kokZP/AFzNAHkhjlkJypHHSpCjlQNnTtXq5+FHgboPiDHn/cNVbv4X+DofuePY3b/dNAHmgjIJymDjr6VE4flQh2+1emwfDDwg+fM8eRqP901bX4VeBMjPxCjX/gBoA8maGWTjGAOxpEtXJz5Z2j9a9ef4TeAypK/EWNj/ANczUP8AwqvwT5mwfEGMKR18s0AeUGNon5UkHpx0prQyF9uwvkV7C/wi8DLgf8LEjPH/ADzNSR/CXwFgZ+JCf9+zQB44I2UYZSuKVImZSccelez/APCovh4+PM+Iy4HfyzVqH4M/DWUf8lKVR/1zNAHh21lGNvy1GiNE3TINe4z/AAb+GUbAf8LKRh/1zNNf4N/C5Vz/AMLMX6eWaAPEypQgBSwanCBtuADivZ0+EfwxYfL8S1H/AGzNTJ8JPhgF5+Jq/wDfs0AeJC3diBtyPWo7i3aIgqhOf0r21vhL8NSRs+Jq4/65mopPhN8Nkx/xcxG/7ZGgDx1YXKKdmT61Ksbx4O3rXrQ+G3w3iIUfEXI9ozVhfhp8MXXafiIQf+uZoA8aaN1YlUJz7U1oSADg8HO6vZ0+FPw4fPl/EkDHrGaqzfC74crwfiOME/8APM0AeOne8vmFCe1TLE0fzBc5PQ9q9a/4VV8OSPl+Iy9f+eRqCT4bfDqI4PxDD/SI0AeWCJkbfjqaa8TZJC4HevWE+HPw6Cgn4hcE/wDPM0j/AA4+HBDEfEPr28s0AeTC3IBVlxnkPQoG8JJ09fWvVz8Nvhtt5+IOf+2Zpr/DP4a5/wCR/wA/9szQB5c1u/LY3AdFpjx45VeT29K9b/4Vx8MxGCfiGxYdMRGpI/hn8MXQk/EQrn1jNAHjm3AYtTYy5Y45Udq9kHwy+FYBL/EEn6Rmmn4b/CgNlfH8g/7ZN/hQB5LFnJO38KmZXccJjPBr1N/hv8Ll5j+IL/8Afpv8Kjk+Hfw42/u/iCST2MZoA8nkcx/KikkdanVZJEB2c98V6Qnw3+HasTJ4/Az6RGr8fw7+GKjP/Cw2HH/PI0AeWraHbjbz1pRAT8+OVr1yL4dfC5Bz8RX5HP7s0P8ADn4XyKTH8RGBH/TM0AeP+XIzFiucCoZQ0iYx5f8AWvV3+HXw6R/+R/baB/zzPNV5PAXw2LD/AIrt/r5RoA81gXzEGRjb39abNhvuDGOtemr4A+Gnfx62CP8Ankaf/wAK4+GOMj4gMv8A2yP+FAHk5+d+Pkx+tEO584G3Fesf8Ky+GZGf+FgE/wDbI0xvhx8OYyCvjpn/AO2ZoA8vQbnG75SKtAySLtxwe1ekD4efDx8MfHDL/wBszU8Pw9+Gp+98QXjb/rmaAPKnjYgKo2ep9acsXmAYG3Fepv8ADv4Z4/5KGx/7ZGkb4ffDQAAeP2J9oz/hQB5ZIpLBVGMd/WmWoaKfKjnPNerL4C+GTgKfHrj38s/4U2T4ffDFGUL8QHz6+U3+FAHmA3tMcLsoERJYlcEd/WvVH+HPw0VAR8RW5/6ZGrkHwr+GsqBj8Rj9PLNAHjhhZcMBtJNG3Ody4zxXsc3wu+GqEY+IZJz18s1WHwx+GzuQ/wAQj1/55mgDyVjlDGOcfxelMhh3Apj5f79esX/wz+GsEJMPxALt6eWayU8A+BgOfGp8sn/nmaAPPsSRt5a8jPFPyytuCZYd69DHgDwKSQfHBB7fIasRfDrwI7Ff+E66jj5COaAPM2jG47Fy5pu0jcXGAOpq9q1lFp+rz21tcfbLdT8s/wDeFUmLPub7wB+7QAm3fGVJ46j3pH8kYLAq46YqQgom4Dj+VMRxkgcg9T6UANO0sGLYFTMYQuEkyT1X1qJkO/OwFD0B9aWW1aLBKfe9KAFVkRsHJU9fapjBCFyHz7GokJAxtz704DYhG3JoAa2QvoKbIVbBI5xQpYZDcimqxJ27OPWgCYwIMO3HHWkTBQ7cPnvTEk3ZQj5e5qQ+SuDG+D6UAREMmFzgH1p5O0q5XCCmEDk43kfpTvJacDDbh6UAPJWViRwnv3pIflDDbgetNQBEbzBsA6U1CH434FAEmzDfuzk0i+YnEi9e9IxVhhW2Ed6VGkVRu+cd/agBGt9u3Z84Ocn0pfLO0BU3KP4qfKIYVADk7vTtQgmIBDZweKAGlHKgsM88Y60ICWJOCw6A1K1vOF8xV2k9RTQ8Mgxt/eLQANtTBKHBPQUm8nIZSo7Zo+1zDggbc9KXM0n31yM8UABXaucBWHOR3pw8x13k5TPSnQqikiT7w5FMUum50Gc0AM2sj5C4X3p6IFfKgD1oeSYDcV/D1pybnXBiypPJzQA2Yb/mZCGHTHenozhN8fyv/dpftEifJgFRURDgmRmz7UASZDErJ36t6U1UZlKAYHXPrQwV4TvGB2NCXESRgZywPFADRcAgrIMkcU9Wxw5+U9KC6OxDLxjmmtEJlIA4HagB4aE9+O9RPh3yvQCnpGowuwCP1okiiU7gegoGhA4Ujb949D6VatiWguTIfmCHC+pqpgg/Op2EcN6Gp7RvNtrobfmVcg0B1PsD4aXht/8Agn/4nQW6lmvz83fpXx5ZyeXpsQxkkMfpX2T8LvLb/gnx4oymW/tA8/hXxfaEGyjw2OtJAxW3NIDgbqpayQ1g5P3sjj8RWgUjAUM3br61n600bWL7F7jn8RTEc7RRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFL6U0Bt2+VEYI+XaOafI+/IJ4FNj3NCg7bR/KlOCnPbpQ9xPc2NDfy5bTCYAmU/rX1l+2ZJ5/wK8CSAceUv8q+U/DEazXtkkv3WnUH6Zr6z/bRt4bf4F+C1hbcAi4/Ko+0Xc+LZgdiYHYZodgqjA5qYthIw3cUhi2nOc1RJGJPMALLyO9OJZvao/Xj5e9SAgJ/KgAjBTOeMdKDl/mblhTAWJ+Zc+lCMAw8wYHagCRmbhmX5u1C5QbsYJNJL8xHHHambH2hiPlzQAxUOS55fPSnqojySMg9aFIxnHehY/3mSuc/w0AIBtfcPu0oZl5HPNOeNO/X0pFePG0DnNAErqdmRwTSneYvvZ5qMOeDjvSFlUFu+elACgsrDcOM0jyAElRn1prEgjjg09gF6jg0AIVBG5eo600kuMn7npQVEeSBwaFA5AGA1ABjyzlBgnpTsOhz1Hc0pKoPLcYA53UpCrgl856UAMjG2TIPA52+tP3KZd2NpbihsqQ5HTtTXAkO+Mc9x7UAKYdhIxlT3pWUsuBzio23Ffl4XuKF3AcDigCRN2cY+b09advMcgMYye60zJL5B2yDpSkOr+56mgB5Qlvu7t/X2psyCMFAdzAZNCPKG4PA5p4/eFmI2nHNAERyoVh36ipEfMZZOCO3rUWzdwDuXpTgUi+Yj7vGPWgBySmNsSfMG4pHhPVecdqkjRXO4jjtTDI27bGKAAfPjcNvoKUqRgY2tSGB8hsUkr+dIu7gjpQAcpwVwD+tNZE6FcE1L5o2YlXHpRJCske/+EUARGNePlwaCqbM4II/hzTSMD5Oc96Ta7oCx5z970oAekUYwfmOT0z0pwhQuFGeT0psbB/lcf8AAvWnMRgKOmelACvBEXxg5HvTDbRFc8k56Z6VIigAjHI55pnlMp3nnJ6UACQpKuCDtB5Oab9nXLMC20dcmpHyo3KNrZ6UwOrMdo2t3oAkjt4pU3YLc+tNNrGpI5596dCGjJcKAe+KN+cjGc0ARvapnAUmhIomJXZgjvUhkVRsI5NNdDGeRx60AIbaM5ABz65qM2kaOAc/nViRhtwo5qIxFj8w5oAjkt1xuGePeiO1jmUgrk1KknBDIeO9IpVckKdpoAiS0RyVGQAOeafHDGrAbSc1JhVYZ+UnpQ0beZjHWgB626LkYJFIsUWQu35DUTM6nYRwe9SBAmFxz60AWDDG4ACcDvTltrZhgjmqyuyKVXvSRwEnOcvQBNJawKR8u7PamNaxAcJkU3BJxjJHWkIOGwMcUAV2t48javB7U9beJuduMdqmCpsAIwT+tMEJfkjaR0oAY1rDt3bcH0psUKEnK4B6VOYz1xz3pm3sRjvQA+NUPybAv4danTZIAuwKR39arKqOcs2KdgEjac470AWbiKFNpC/Wo0SMncyBhnp6VHu2kuw6Uh3bSU70APKxJJtCgA96ZJEhXKxAbeTx1pgRuhHHenRSGUlCMFeR70APQRSKCYxj0qLy03HCAe1PZGX5lXn0pB8w5QgdzQAwwxqSqIMnqKRY1VwpjWrBCv8AKzYH60GNTkDlaAGrEmeIhipUhg5Yxg+1MI2jCjBpVUg4b5TQA/NsBxCAfTFKUjILCNRj2pFBYlSvy/3qh2srkBuO1AD3hiK7mjVD6YqrsAfcEHHarEhYsC55H609YRJ87fKRQBCyKmH2DJ7YqVMKuRGoJ9RSOhdgSp2ikZQg3D8BQBOgRNxMSNxnOO1PV1jIIRSuM9KrfOqEEc4yBT4mBPzHBx0oAuPLFNj90gyODiq7BHHyxoQOpxTEiDkhDnigo8Rww4xQBXmjRTnaAcfdxwaeYFKAsig49KsZ3c7cjFI4QAKTwe1AFaGTyjtaIMCOuKtxurYCquPpUTEovzDjtUfBxztoAv8A2oEf6tVx7VC8qsQdigHtiq6zeY2zHzHpTcbn8sg8cmgCV3Dk5iQAe1Ma2WRRJtQr6Y6VIF/vDntQbfaTsHynqKAKsMYL4KDPbipzAqnhFL/Sl2/MF8s49aeVQd8tQA8+UiLuiTfn0p4BhUF41ZSeKgKO4G/14HenRtKmVCkqT0NAEyTxn5VjBJNBWNt3yDPrio0jTkbdp9qb9mYkBH+QnBoAe6RxjcYguOnFSKEkGSgH+zimiBwORvAPFNC/vCrfK3agBzRIwDbAcHpio5oYuSyBc9MUpmAchfvU2RweJOvrQA0AKpHOPenBon6HJ9KYxTOAcn0pAoGdw2+lADmR0BG7KdSKWQ/LuUYNMiIkYlu1SognJHmbAeooARN7ZGcMRjd6U1IZQxVnyg5JoYxqSiZbHBNLuBGxeQeooAUshbCZakUkg9x3pYx5LbCuFP6U5ljXJjbcehPpQBHx5gJGfQ0u53LAt8oHFKqEggcep9aNuHGYyTjr2oAap3cONox19aVhEoGBz2qbZuIJGajeSON8SL260ACFsAMv41GVKHC8ZqX5I1zG+SRUZUKQz8e9ABMGkg55I6mkOxo13cYqQJHGp+bO6oud2zOB296AFzHkcbvepoVKNuj+aoCcLgpkU63cQsQe9AEwWSOQuE3buNlPEbHaSuADyKhVZPM3I4HsaPPkDkNIOfSgCZ53ZtmdoqONY0P70YGetNJjYDPBHOakaIJGrN+8DHgUARsqM22MlsnqalSSdSI0O456jtQZYmGwphxSAKEJHy47nvQAGKTcWY4cHrT/AN8vzodzH+GmlQ+CxJ9vWgOidznP3aAHB3XLs3PcVGNxJJbGe1AcO5DDHpSSoqHBB56UABjKqcDd70oG0jA5NNKsEypxiljbB3M3zUAIdzE9wO1KCjMMJg+tNAm37kbcPSnS3E0bKAuM9aAFdNxyONvf1qSJiylwMEdqjERGSy4BHWlBV0CxckdaAHxI8qHC7XPXNNkEWdvJbHU0hJY7h8uO1PQq24su0YoBbixSMnyN8y9gPWp7fbMtyxTawjNV4XVD8ibs1chxKtw23kRmga3Prr4XIw/4J8+KyUwPt5+b04r4xs8fYIsDBw3HrX2r8KJYB/wT88XQzDazXxIPrxXxZaRtLp0ZQcLupIGOSNVf5hyR0qnrQ/4lznbjkfzFWsEDcw2n1qhq0ry2L5PAI/nTEYFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUCijtQBtxnbEnU/KKfu2ZyvWmRyFIkGM5AxT/NJ+VuvrTS1A2vDOPttqpzzMtfWX7YaBfgB4PBTawVcflXyp4aEa31iJOIjOMtivrb9tK2ig+BHgzypPMVlUj8qiS1K6HxVKFlhT+FgOKRsggE5GOtOkTCjPdajJIA9KokEw/HX29aPLVOuT7elNQEZI6UocOcE896AB0HGw9KahGQX5qbCKox1NQghmx0/rQBKcMM7celNE0jLsP3e1I2SAMHijcFGCDmgAVF4zxTl+UgL81AMb4DcUHZEwCtwTQA1lMcgY9DTyAWxj5j0ND7SPkbcfSkkUkAE4ftQAocRtkdc0mc7mK9O9PiYrFtdcnPBpjGdRnZlCeRQAGHfwfunmpGZsBdvyDvTYGZGJPr0pw80livAzQBE/J4pVGAWPNIfmOTnI60hKqcjNAEhc7ckdOlOChlLDkdwe1ML74iSMCkQEgc8e1ACu4PK9KXydsZK8E9KHQBScdaBkxA+hoAi68dB3pwZs/KMUrRoDg9DSrGQc8kjpQAjZBAZTn1pxlIfDcDFKsrOD8uKdjeCD6U9AFmTbskBJx0pJNswyG2nHIp7TBYPLI3GoyBtC/xfzpAA8sqCDjFGFdxxyBSJt4G3J9KdKhTBA5xQA6MbRtxwaGUDBUYpuS8e0cGgcr06UAC3DOcDj1FIzRkDOcijG4DA6US5KhgBtHWgAzGwDEbgeMVIYwybFbg9qRE8yMNtyKZsUHCHrQA5IU+50I5z61GYXgJOMq3f0qZE3MBnntV/S9J1DX9QhsdMs5b+8c7VhjQnn8KAMorlQFBPPX0NEeBlTww5z617pov7F/xZ1i2Sf+x002NwCqyuAT71B4j/AGN/i14ftWuH0D+0IF5Z4HBOPoKAPFEJJG4cU8B1JZBn+lGpWd1ot49rqFpNZXMZwYpVK4/GoFuGIK9M/rQBPJnbuPbuKjDCQjA2805pAQuRgd6TBVsLwD60AKXIcheo600NjKn5T6+tKHKuTjp+dbngvwPrHxE8RQ6HoFv9p1Sb5khP8QoAwz86ldp+tEjfKEGa90f9h/4xgkNoSRn+6ZADSr+w58adnHh6M+h81aAPCsfw4PHemjJOewr2jXf2P/i74atHubzwo00KLuZ4XyR+VeSy2ctpdPDcQyW11ESHhlXaRQBWDseAnNKQ4BIXGO9TKm8MTwM8jvSSusa4TkY70AQKHJORuOKkWOTcoI6Ck3DI/hOKctwVYKefegBsoA6rk1AxZCBj5RzmrDSAseOvvU+n6Vdaxex2NhaS3t5KcLDEhbP5UAZ+7bxgnPeiIlHyCSD3r3bQf2KvinrFut1d6dFo1uygoblx8wqxqP7D/wATbWEyWENvqu3kxxOMmgDwhsswYDGBQJBIMOu01p+LvB+veA9S+weItLn027XosiHB/GsYzjOAKAJvLGOeR605dpG01CJwOp/Cngh16YzQA5gScL2/Wl3DjzExmmqy9BwfWl3EZY8qePpQAz7EJOVO7PFONuIlAPGD1pY/3Wdr7VHJY/0rtvAvwe8b/FVynhnw/PeRg4M7qVX8+lAHDFeuQT7DvTctwADzX0nYfsG/Em5jU3N3Z2c/aJ3XOa5zxh+yH8VPA1pNd3Hh/wDtTT4xue5tTuwPwoA8RMhi428HvTljVsECpJINs7oY3SRDhoZFIZT9DTmhAwSMD0oAbl/4cAevrRvlYBQBjNMZuMHO3PrSSBlPBOKAJQAx2lOfWkdQjhgOR29arqWGSc4NdB4M8Ial8QfEtp4f0hQdSuTiLPc9qAMR/kzxyaYLl2Ugr8o6+9b/AMQvAWvfDHxVc+HfEtq9pqNvgspyN4PpXMmcxruQYHoaAJ9/luCec9vSpAuTtYHJ5DVSMsgO4dD1zVmGcqjFz+6GCe5oAm2fxFeF70yWCWQFv4RXot38D/GVl8MIvH81g8fhmRsLMy4zXDFlZQ6jAYZH0oAq+XJ5YOenal2iRDxzUj9Ce/aoZWCISMgmgBuBHglst2pm4PKQeMioGUq4yC2fetjwh4Q1Px54mtdC0aIz6lckLHH6ntQBTUMeR8uBj609UnTGW355qTxNoOp+CfEl7oeswtBqNk/lzRnsaqfatoOOPWgC0ckg7cHFIEz95c54zSRSF49zDjFPTLAADCmgCH513L1A9agKbh3UHrWg8SSRFx/ACT9BXYH4MeLZPh4fHcWmST+GUOHnQZ2fWgDzwRkKGCnirAiZpA68mpEjBwR8wbkMO1TxRBMAnB/nQBAvLAMNpPf0qbaAu3n61LJ5bRBmX5g20H3Ndb43+E/if4eeGNE1/WbbZpmsZNrJ60AcekMsa/KQwJ6GmsmFyUGQe1PDrGh5wc1WJZycvjv1oAQgghsZ56elBlkYkA4pCkhIZfWgRq+QwOe/NAEa5iYHO85yVqzguflBw3YVW8vyxuHy47n0rvPhx8GPHPxVkb/hFfDtxfqDg3DjYv59KAOSjEoXaqspHc96JBvBBTB9fWvo2H9gr4ttCDMLSKXr5Jdc1w3jv9mH4pfDy2lutY8Myy2CDJuYDvwPXigDyYxgZBXBpk4RyuzqOtK8oL5KSDZwyOMHNJ5auMA7GY5B9KAGYj5z8p9KahD5ULQY/KJXG9/71OGYwcnLfyoANkYfezbcdqe4X76rgfzqNOGbzF3selPxnc2ccfdoAZLJ5bhlj608p5bjOTu7+lMRVILEHNPKybSIxgEdfSgAlYMCq9hyfWlRU+z5xt7Gmq0igAx5PrS+YEfEsZCGgByrtIBb6NT41lcthug4z3qHesrkLlY1FSI/Rudy9KAHb2AC469j2pxXaDlQ2R1NMGN5ZhjcKc4Cx7d3J6UANVFQhduQetJJEHIUjjtSqsafxEnHQUquCcEHp0oAieIRgLtzTkUbPnIA9akAATKvuI7+lRBhIMSHHt60AOWPcv3sjsfSo0iZPvfNnv6U1t5IRB8p60u2TZx0XrQA6SIoRlcjPUUIkUhG9sgHjFRxecy7Q3PWlwYsHZ16j1oAsMsDYUDkHrQm5A6DofWolkXKhFyx/SliZSzZ+73NAEkYWTIAyV5Oaa7g4OMHP3aQygrviPKnmni4Z/mMeOaAFUckkcCmj5H3lcg0jzOpyVyncU+JYyrZYn0FAA1vFKhKvjvimsCcGQ7gvSmqgCh2U4B9akja3Lbjnr0oABbscSIdo9Ka0JMu8rwKUqZHLrnaOi5qVBIARjAPf0oAj8t0PmRjjpj0pyKzlg457MKBayxbgkmQ1NCzQna+DH3oAavm5IZflFKyDBdfkYfrUjRxyEsHK8dM1XkBYfLnA/WgB6ShhnZinMN7Z7elMilXPHy47U9mG7cVK8UAOVhuBHGOtXbcRfYbpw5DhM8VRJL8qPlA6d6mtkAs7oqv8POfrQB9n/DD7Gv/AAT78SlkO9rxjk/Q18S2gMljEUO0rur7g+Hvkn/gnlr48ra5vD/I18P2qg2MTc5GRj1oGx5dycSAHjqKztYVRZNgEHI/nV2QvnCnFZ+ql1smVueRz+NAjDooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACjtRRQBswAgIrDIKg5/CpHRYx97dnt6UxCzRxgA4Cjn14oAXJIUj39DTW4nudD4dTN1ZRMpcPMOPxr6o/bJmWL4M+DbcIwZY1NfL3hLyxqelxuxIe5UFq+qP23rZIvhP4RK7gFjUc9+KmXxF9D4ze4BIx04o2rtzn8Ka6qwB2npzTQhC4waZI4sO3ShY4yCRwaacY+YlcdM0KP72PbFACxOpfa6EE9M1HJHsbB4qZd5U7Bkj1pgYsAzj5vegAWWT+FM4pI5XZiJFwam+cAEDA96a8iFTuzn1oAbINp4FG5SM7afFGJY9wOcGk2HdnHAoAYY8gEfKacnXLHODTmQDr37UnGOeGzxQBJHjcc/hU9tLIqOTyPeqpGTuYEgVIJDt6HaTQAufmLbcj+VNkBGHQk+oFABORzg0pfYpC5LUANlHAYDA7ioxw2MZzT1YjJ2kg9VphRSxKsf92gBfusSMnFKpAO4jg9qXLKMAY9RSLuKknkdhQA5/3i5UfKvOKOGBKnHHSkSUMuWBBHYd6Ciqd4bBPagBWA2hQCc0igr8objrRsGDgnP6Uh2jcCe3FAD1Du4I49RRK5jbAGKZ85AwSCOlKhEg5zu70AOUAjPVe9NLqvyk5z39KNzI+1eQajZFX+HrQBIpPSMcjvQSzEFsnFMV/LQYOVB5FSKTJnC9aAFQrnc2cdhSZaRtvQU4BIwC45pGXcQOd3agBpQscDIYVOArxc8EU1CHXDghh39aXHByCW7UACsyrsXjPUUioM9Pw9aVAGPzEgipUC+b8pJ4JoA6n4Y/DXVfir4z0/wzoyZubpx5kpHyxL3Jr6c8bfFjwb+yjYHwV8NtLtdY8ZKoXUtcu1Dqj91T0Oab+xvYw+Cvgt8UPiKyD+0YrXybOU9Y25zivkWW/n1iW4v7hy93dSmaR36sSeaAOr1/44/EPxDeyXV94t1KORmJ8uCcqq/QA1J4d/aB+I/he+jubLxdfTmM58q5nZ0P1BrhZmUDeTwOx5qo5RiPlkU57KaAPpzxn+0P4K+OPwxvbPx34fgsfHFqN1nqenII1lPYMB1r5agYld2akkKMp3bmA6ZQ5pgt2Zd0e7d6UAW1XCj+Mk9fSnxLuJU8c8VWRmXAIIyeamMqrwnX1oAsi2yGYnBFe0/sevJbftDaG8TtEfJPzJweteJxyEAliete3/sgSrJ+0NohGQBA2QfxppgSftG/FHxpp/xy8U29r4t1e3ijlBVEuWUKPTANcDB8ZviCqAL431oD3um/xrc/aXCv8fvFxJGfMGK84WNSCS6Aj3pNgkeo+E/2l/il4Uv1vLfxVd6gIyGNvfSGSNx6EHtXuXxc0rSf2nfgC/xS0jTINJ8XaDzrUVom2OVPXaK+P/Mhhf558ccKozk4r7F+AsZ+F37HXxM1nxEfscHiGDyNOhm4M7ewoGz41i/eL5o/iXJA7U2RMDOMH2qKGby7cYG3cc0jyndzkgc8UCGMcckc1FIS5xypqZ33AMB14NNVCzAFTxzxQBpeFPDF/wCMPEWn6JpiGbUb6URRIOcZ4ya+v/E/jDwp+xJ4bt/DXhiytPEHxNuohJqGpXiiSOzYj7gB6EVxX7COk28WveM/GVzAJDoOnNPbF1yVkA6/pXzf418V3fjTxdrOvXzs9zfXLSlifU0AdP4q+OXjvxjfS3mqeK9RDSMWEVvOyoufQA8Cs/Tfip4y0S4jnsvF2rQyjkA3LY/nXI5CEBhgEd6eAgUH7+ehoA+xvhd+0VoPx701fhz8Y9PtpGu1EWm6/CgSaGbopdu4zXzx8a/g1q3wU8d3fh7Uf30AO60vFGFnjJ4IPeuGjR4IjKjlZYiJY2U8hh0xX2X8d7RPiv8AsdeDvG1wS+raFstZJwMs6k45P40AfFywEnJBGPWrcMBXAYk55qyUACkgspUfN60nljODnmgCB2iLhcGh12jaemM59BTzbqGDYJpz25MSgZMjuIwfTJoA9z/Zb+AFl8S7u/8AFfjGY6f4B0NfNupScG4I/hU1ufGP9tPWNWeTwx8N7eDwl4QtMQ28lsmydwONxYdciuq/aGupfhd+yR8PvB+m/wCi/wBplri9ZeDNu5wa+NsDcqP8oX8zQB1UnxH8XXUvmXHizV2lP8Qum6/nXoHwx/al+Jfww1OO5sfEEur2ob95Zao5mjkHcYNeMI6mQIPu05pXjIXHGeD70Afb/j3wJ4R/bB+HF78QPh9aR6B8QtIj87VtGj4S4QfedFr4ukZw7LLGY5EYo8Z6qfSvWf2TfiNffD347+Grm3ldbLUZls7yEH5ZY24Kn2qx+194Di8BfH3W7e0jEVjqLNdwxqMBQecUAeNCLcxUAnvSIPKc8ZFTrIYVxGMs3BqMgKGyxzQBI9sgTzBXpf7M0bR/HXQTE5idyNjrwQ2eK81BV4gecDtXqP7M+3/hfnhkjkB1OD65oSA+qP2nfDdn+0/aeI1srZIfiH4JVY5IkGHvYsAl/fGK/POWFlmkWWIpIjFHjbqGHUfnX1R8R/jHqvwh/bN8QeJNMG6GOVIry3bpLCQN4x9KoftefCXT0m074qeC4xL4O8TJ5ssES5+xzfxK2PUk1TQHzBtIZhyd3SvW/wBnD4Jz/GbxukNwxs/DumqbrUrxjhERRnBPv0rzzQ/Dl94n1yy0rTIWuL+9kEUSAZ698e1fUPxt1qL9mr4W2vwq8NzRt4n1KES+ILyP+4RkID/OkkB7F8ZvipY/ED9knxHYaNbCz8M6R/olnGgxvZeN344r897dv9HgB4YxjFfU+g3Qf9hjW4gMESfM3rzXyUhYwRhW5KAAelNoDQJVc85bHUVFtD7u4I59qrxymMlWOeOtWkUFdw54qQITZADlz9K9X/ZJla3/AGk/BxjYq/2kDP4V5ckJdsFT7GvWf2UbfH7SXg35T/x8CgCh+12pk/aW8a53Em6JyfpXkahwCGGa9q/a5tT/AMNMeNDg/wDHz/SvIJLdweATnigBkLOseOSKuxHeAQCMetVVjeHAPNXonKEEDjvQA+U/6NOQoB8s8D6Gv0B+BHxktfhF+zHpL69paa14O1I/ZtRtZFB2BuN49CM/pX5/SyZt5y3Tyz0+hr6v1C6Vf2BRG4BdpF257CgDjP2pP2bYvhLeW3ivwlMdV+HGuj7RZXcfPkBhnaxHTHSvAZBsCL3POexr6g/ZZ+P+mWumS/Cj4kZ1HwLrSiOGaU7mspGGAVPYZNeUftH/AAK1b9n/AMazaRcn7bolyfO0vU4+Y5oTyoB+h/SgDy2W5DxbRw3mrzX15+2BeSzfs0/B9GdnjVWIDdjxXxuD/DjnzBX1/wDtbnP7NnwiU+jf0prYD5DaYu5AyPenxnzR8qkEUgt239PlJ4+tWow8K9M+tICWKFZIgVbawPIpfswkPp1yaSIEYYKRg5NPuCXiKxglpWVAR15OKAPoP9lz9nPSviDb3/j3x3O2nfD7Q/3spJ2m6Yc+Wp9TjFXfi/8Atq69ru7QfhraQ+CvCFsPKtxaL5dxIg4yzdSa7r9r/UF+GX7Ofw6+HmlZtYNSgjv7oJx5jEZOfzr44nUISvRF+6PSgDbl+K3jeWUyyeM9ZaXPJ+1t1/Ou++HX7X/xO+HN+kraz/wkdln57LVz5ysvpg14pKyo2CSSeiqMmmF0BBUyEd8xmgD6A/ah+JPw3+KVrofibwno/wDYXiaePOqWsI2w788lR2rwd9oUgbtzdR6VWBUt8qEH1KGp4rbcQoOcnJPpQAsBTa6seexpFKqG3nPeicCKTYvPvThtBwV59aAFjJA8xcYPQUjAjIZfnpEISbg4+vSlkjdgeuP1oAjD4fO0gj8qeZWY53YWmsWA29U9D9405FG7kZbsPWgBrGUOCsmR2p0s8jECTnNIVIJBGDTmCtCNx+c9KAEPzNwMKB19aVSAAcc0wSMi7UGR605gw2knp+tADgwY4I6jinMdmCyZOOKEdX+Z1wwFDF3IYA7RQAxMrksnJ6VJHyCVOcdTTDOVXAX86b5ZRg/zAGgCQFXwAu1e5qOdEVgoJYGiUt5gGcqR2oRMsFVGx6mgCRo41Rdr7f605ZA4BPy44x61XDLC5QDfn17VLtzwDkH1oABGEJK5NNkDMoDAkHvSvHIEyoyB1HrQsrbOVypoAREXGCNhHf1ojZVB42nPApwcy4VhtSh0BXJ429KAGLtDFsEDuKlRDgtvwh6g1GGkznA2+nrSjdEcyjIPQDtQA+XbtChs57+lJEVb5W4x+tK0Y2BSpIY5+WnC2DneDlB1UdaAFK7X3BSUzyvY1E4CylwoP+zUsu6Ebozuz+lRyKZF3YO/2oAc46sr7Se1Oimm6SDcvp61DuVhnByOtK0iqQ6HLDtQBO8qrKChxn+E0plBk+bIH6VCfKny7ZjcdTTW5HyscDv60ASzojYaMEgdh3pHGFBUfUU7YjLkMVHcUiOG+7k47igBBGm4H+L0p0ksh4ZMgVEyb5CCTkjAxT2MvlhQNxXrQARSI/BOwirEch+x3S4JGzgiqeCRhoyc+lTWzP8AZ7qLa2zYetAH2h8OnX/h3v4hyrki8OD+Br4ptHKWMLbeSGFfYvw316eH9gDxJapbGRft5G89uK+OrR9tjEDltwPXtQNiuxDYZSWPSs7Vj/ojAdAR/OtRmSEAq296y9XJFq3HBI/nQIw6KKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAo7UUo6UAbUYzHGR1AFPkJk4YBfTFNhJeBeuQBRK+4A85FNbCe5saJgXmnMgO5LgHH419T/ALZOp/afhF4PiYMJDGp/Svl7w/Es95pylimZx83pzX0v+2ZaR2vwz8HFXZ28pcH8Kjdl9D5NYjYCxPFRKSvzZ4p8jAKB60OvyZHaqJAqj/f+7TMDoOg6UMcryM04qGjAB247UANJYH5uh9KFUt0BP1p+Ac5yAO1NjKuxG7ysfrQAu44xu/CkGBk4zTXGDx83vTkG8Y6YoAfEu87gSo9BQJAGIORQTtUbDk96lRlaP5hzmgCEqFbfgkdhQ4Kx7sc9hUu+VCSqhhTShXLP19KABXdlDFenb1pQwhUnqp7elMV9vO0ge1OWRXHINACBsnOTipAiMCQeahKhcjJ57UiDacgH60APBIfpx60CIPIVwd3XcKdjJHzYFNYkHgkj1oAbvYHZjI9TTlOXyuVC/lTsApuAO2mEuAVUEA+tAAcBjJg49BQ0aOd3IJ4pNoQHBOe9NPGDyKAHbPK4J+XvR5eeQN1IYi+Rk4pVxDIuwsw74oAc5MQGO9IQTtwCretO3ne27JyKTJDc56UAOdQCrjgDrQ8ol4K49qHcS7dueOtNkTI3KDx1oARUGc4+XvUqIUkXGce1NjGfm/nTmYkdxigBJgrybuSgHPtQqggEnpTVYcso6dR606MBTknPtQA9fMlGRggetICeo+X/AHqWNcZJO0HtSMpVgSCw/lQAoTHzEEg1KkeBhcqcHBpMKkWVbd7GmCR1TeRhQenrQB9dfs83R179kH4maTbbmubRPNkjXqQa+QrSTzLWJR8pUbWJ7GvdP2RPi/Y/DD4ivZa2N3hfxCn2S/Q9EB4BP50/9pv9nDVfg74rn1bS7ZtW8D6m32qz1K0G+JFY52MR0IoA8W0qUWGq2V9LALiG3lDtC3RwK+lW/a78INb2kTfDLTt0KBC3kjLe/SvmOO7tyfluVz3BzSSalCnH2lSOuBk/0oA+/P2X/HXw9+Pnjq/8P3ngGytD9keVHWIcECviX4oaZDovxL8UW1vH5Vtb3jokQ6AbjxX0l/wTk0jVbr413WpR2c0emCxdXunG2NeO9fPfxpdG+L3jB1fzE+3yfMvQ/MabA4Wbew29FPQ9xVXY0eRzmroj3bmJOD0FQyRZPekBAly3KkHjmvcv2QHz8ftGYkhRES2PSvDmhLsQgII5Ne7/ALGNt5/x+0hGBANu1AHtfxOn/Zvu/iXrr+JYtc/tYSfvfIxtJ74rBhtv2Tbm8ittniKFJXCea+NqZ7nntXiH7QCpD8c/FiKucTcfnXAh08tlZNykc+tAH3b8Vfhf8F/2cNF0rxJZ6HqPi3TdSQPZagAHgVwM4c9q+TfjV8fNe+MM9tDeJHp+h2f/AB56ZafLFGOxx3Ne/fsjfG3QvE+h3vwT+JMit4Z1lSmmXs2D9kl/hOT0BOK+e/j/APBHWfgT4/vfDmsQsIN2+yu15SaMngg/SgDzVLso/Td7HsKla58zgggdqgWEZIxjPapVjKKWx8ooAtW4AO09xxmrXlBdoIwxH51RRmXnGKVrqRWycHjoaAPq/wDYYQav4Z+KWgRg/bJdLd41Xq3B4r5Kmsnsr25tp42EsErRlCOQQa9O/Z9+Mc3wV+JuneI0R5LBmEV/B/fiOM/pXsv7Vv7OJuZG+K3w7Q694N15ftVxHZ4drKQ/eVh2oA+SHhMjhSCcVJHGRnj5APu1OJIPu+cEdeCGU5B9OlD3FtgbpwD7A8/pQA4BUt5if+eZwM19j+M8eC/2BbDT71jHcazcJNBGxwSAe1eQ/s5/s2a98Z/Elpd3ls2keDrFhcX2q3Y2ReWvJUH1xWn+2L8ZNO+IXi3T/DHhsg+E/DKfZrYr92QjgmgDw5ZAESM90FGzC4zuHr6VnC4LjGD9akS4YAnOB70AWhKQAEG6ke6EUcTfN8kys3ocHNU/tRYgoNoFK8wkOGBUMMfWgD7E/bFh/wCEk/Z5+FHii2JeyeP7OxHIDLjOa+QYIRIW+Xd3z7V9hfsseLND+M3wj1X4IeL7qKxu2zLoN5cEBVlPJGa+c/iZ8J/EnwX8Rz6L4p0u5sPJcrFeMv7qdM8Mpx0NAHGC3jYDA2GnNbFAu4ZycU43tsxBNwnB7Z/wqawMmpXkdrp8Ump3crbY7eFCzEn04oA7P4FaDNr3xq8F6bbKzzPqMZ+Xkgbq9Q/b01211b9oKS2tm8w6bbfZpSvZgK9A+DHgqy/ZO8GX/wAU/iB5cPie5gaLRNGfHnpIR8shXrjpXyD4r8VX3jDxHqmu6k+/UNQmaaQ+hJoAotIVGcHbnnFQtcE5+Xgd6hEzHknikeUr9D0oAd9pKDIzjPQV6l+zLOT8evDTJ1LqMfjXkjo0eWOcV6z+yygk+PfhjsPMX+dAF79qWUx/tF+MN+f9aoKjuMCvSv2UPixpH2XWPhb42dZPCXiGLNvPIeba4Gdm30yTXnH7XAC/tJeMsc4lQZ/AV5hbXDQvE4O1oXEiODghh0p3A+1/CvwasP2LtJ8QfEDxi0Vz4gZWi8MWZwRIrfx/gMV8ZeKdf1Dxbrmo6zqtw8+oXrPJI5PIB/hrrfiP8ZfFPxWtNHg8TXjXcOlJstctnaK4JiWWQ44IP8qLgfTvhQBv2I/Eeezf1r5QjKrDGTnOwdK+r/C0e39iPxE/P3+n418nxwb4omOQAgouA4yscALn69atQO8RyAcY6HtVdoQu0hjipVYkHZlhSAvxyswXPT1r179lNgv7SPg3BJP2gV4tHP5Qzg5/u163+ybP5n7SHgxjkAXQoAtftcSEftLeMzg/8fP9K8mCHOCD81eo/tXXayftM+N2B+UXeP0rzE3Gxc5yTQBXuIzGVySaTzAoBwSO9JMzy/M2Riq3mnyW25LCgCe6l22k/wA2cxnFfTmuTlv2JLRckfvBkdutfKQJaGbgsAjZr6m1DB/YmTceTKoA/GgD5u8wMkUfO0Kp3/3TjqK+sPgj8U9I+PHgFvg18SLlFukUjQNduOsLdkLenIFfIcJdoUUcEKv8qs2rSwzRSRu0M0bBo5AcFW9aAOg+Jnwx1z4S+L7nw74gtZILm3mHlzN9yZM8Op7gjFfSn7WwD/s8fCMEZwD/AErT8L+IdN/bC+EK+EvEMsdv8TNBUHS76TAN8o/hJ744rP8A2u9Mv9C+APww0rVoGtdWs55IZoHGDwcZHtQB8o7VDsD0JPHoaUbYhvDBjnp6VE7EM2Ou6qzk7sg/hQBYEkkTF8g5PC1ILxI/s82MeVMrED6isyV3xlSd2elRGRgrI5IDdaaA+zv294pNU8M/CrxHApk0xtJihZx91X2ivk2aTcR/WvrT9n7xHon7SvwRvPgv4n1BLDxDZ5uNEvJyBvcfdj/HFfMPj/wBr/wq1+50PxVp02mXcDlFlkX5ZAD94HHIxQwLXwt+INv8MvFD6rd6Jb69Ey7fs1yoKj3r2G3/AGwNASR2k+GWkBRyP3I/wr5xF1alsrcr8vsef0p8LG+Z7a0D31zKu1I4UJLZ7dKQH318Rx4V+JP7Deq+OLPwhYaJq4uwqy28eDj2r8/7bclugBO8iv0CvPDWr+Gf+Cbd/aarZy2M8l2GSGYAPjHXHWvz7tWVoEG5s+p7UAKVO5s9e5pWbC7QcseAaQTMCcjMY6mm52Et2Pb0oAUABirElvWp4XI4Jyai3A9OU9afuGzcByOlAC+WJJSG4fHFL5XzA5JYdaRS0uNwPmdqkPHIBPrQA07VBGCT15phTcp3Lz7VK0ErYONwP6UPCLZTubczCgCAKDjkjHb1oaUKeBkkflSujZDr98fw0hYXIwRskAoAcjZXBppd4CGU5HpTmi8vBJwoHJ9aQSs3KrlR3oAlZhcAELtPekVkf5S5x3qNCWV8ZziltwhB3YBx3oAkSAIMDJ70ea5Gz7oP50+NI1XmTn0psieaATlQOlADPJUrtwQf71D2/ACkuR6UAGX92CeOvvSRZRHPKKv60AOSZwwDAkDqPWlkBwWiyfY9qaJzKFCoQ3r60pUqRtUqT96gBke6TJft6VKf3xHtSoFc/LxjrSbQsmUz70AM4jJJyR/Kg/Ou4gsO1PJIfeqlh3U1I/lOgw5RiegoAI32JlSC2ehqMh0yy5Bz0FHklAWYE+9JIgiQPvJ56d6AJEcBThSwJ5qPLxSFkBPqDTo5GRCVGQeueooHn4MnGKAFklIUvsANQZKgnbnNSEu67mxj0HekERkJw21O9ACsR5WAuSeppFAK4AKj19KjLlCVXJTu1SRvkFcHFAClMSc52gfnU6Rq6lk+XH8Pc1BnYMA5Wglkbcp+f+7QA7bliDxx19KVQVGUbJpFkZztIz60wgK4OdhFAEqPKMggVJHIzRziTI+Q/dqMPJjJyVHPFSQp+4uWILbkyB6UDR9e/DseT+wD4kyCN1+f5V8a2LAWUWeetfZngFif+CfviFWVv+P48/hXxlaJusoiMhQDnFAMmkKoytjiqOroTZOwbIyP51cwJCDn5O/tWbqieVbuqPuQ4/nQIxqKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAo7UUdqANiNsLGTxwMe9STSl0IIwfamI7eWi4yMA05zhQw5B601sJ7m7oMe7UNKibOGuF5/Gvpn9tyBbXwD4JjR2K/Zgf0r5t0Bd17pce/G6ZcN+NfTX7b9jHbfDnwO6ziVvs4yPwqVuX0Pj7JOOMginOwIC96QMYHGAcMO9OMgJGR1pkicbcnqP4fWk2AMMZOf0pCvbPPZqcrAKeTkUASKrRDdgsPemHZOPnwh9qcr4PXimkruGc80ANOI+Ac0jtgYAqQwgjIJpYtioWbrQA6KLKjHpSLhcnBqLadwIJx2qZZGRQTknNAADjaCxVSaeW/eAN93PWoGbzMuxIUHpUkhH2f5icZ4oAfIPKVjyc9MVDG5ByynBPFKJSMbvu04y7ecZU/d+tADGXJycg0okPQ0khZmGc5HWhlX86AJDsAO7n+lLvEa4X589vSmcEdM4p21cgrQA+YMQG7DqKhZ9wKnIzUzqAN27PqKg8zcxBBQdjQBI4ESDGS/XB71Hz8xbOSO1SRKHJVs5ppOyQq2c4oAchKRbRx3JPakaMKvy5BPOexo4BOfTqakZHRePmGKAGHpg9QKbJyoI64qQ4jUNz9DTWIbJI4xQAiEKoJBFSRMOQM7aYJMx7Am4DmkQleV9OlADjGFJZW+Ufw0omduCMD+dNQDIJ709wQwIGeKAFBEhwRsPb61H5Z81VJw3enk5Iyc0NGZTuHDCgBThZQFOaeYw3AbGajGCw4we9DoGwd+B6UAAHlkANuFNjfO4jLY7dqRm4wi4FKncr170AJA4BI656+1e8/CL9rXxT8NNJ/sLUreHxb4Xbrp9/h1T6ZrwUIVbK8+ooLHcNvyHPO2gD6+uPi5+zV40hFxrnhe70e+fl47KL5M96zZPix+zx4NQz+HPC93rF4pyqXsQ2H2r5ZjlMhwAG2nPzCk8wqSwGDnpigD2vxx+1r4o8TW5tfD1nb+DbA/KV00bC6+hrxGSea6uJZZpWkmkYs8jHJY+9DxAgjPJ7UeUdoAbnPSgB2fQ5IPSrGI5oy33XHUVDEFJ2g7W/vVIyr64I7+tAEaYY7yNqg816n+z7480/wCFXxZsPE2qq7adFAyMIxk5NeYllB3HOB/D604XJCktyueFoA6L4q+LLfxp8SNe8Q2KvHZ3su6MMMHrmuV35XIolxKCeVB4AqtIWVtnNAFm1uXWdGV2idGDJKhwVbsRX0drn7R2j/Fj4Gv4S+Idu9z4l0tB/Y2rIuXcgcK7elfNIcgY5FPMuMKfnPb2oABCxi+YE7eM1KEUISTgdqakjRyAOcr1GKWQ7sjp3oAid+PWjyw5DBuO9K6BUyBTkiXqeKAHJGCSx5TGOelet/A/9pPxZ8Dp5Y9NlGq6FOf9J0i7IMLDucGvJiz+XsBypppIQgjp0NAH2Lc/Hn9nXx2ou/EXgyfSdQkG6VLGL5d3ciqT/Gf9nTwiPtXh/wAI3Wq3ScpHeRfITXyUXZ2weDjg460wzSoP9kDnigD3P4sftZ+L/iZpf9iabBD4S8OEY+xafhA498V4PsCnAzg9Qe9WEYMFyPmPSnMm7A6HvQBVMYUAAk/SlMfmMMnH1qxGm3O3g0mRuVpBtXsKAK5i+UY3ZHenxKXbnDH0NWBcfNwuFpiSAtuC7TQBYs7uewnhnt5pLe7ibfFMhIZD7V9PeCv25r5NEttA+JXhyz8ZaXENi3tygedF7YyK+WpJZNvSoQ7EZHB78dqAPrm9+Lf7NGpP9ok8L31tM3JRIvlz6VXf9r3wV4CjMfw7+H9mbz+C+vYhvj9xXyjEzSehAPGRTnZyNvAOe3pQB1HxE+JPiL4oa+2r+JdQlv7hjhI2Y7Ih2AHSuVMJYksSOvNPjZVGAdvPekYruzv3LmgCLy9p+6dvvQ1qzjIDYqfJGSWyO1SCdyNq8ZoAprGyIdw+UV3fwe8awfDP4j6R4puIfPt7JgWiHJOOa49ZSrbZBz/e9KR7napULj39aaA7H42+P7b4o/FbxB4qt4mt7bUZA6IRjGAB/SuJWQJnJyvY0BkkOH4FNnhMYJUhoz90HtSAla4RVPzE5pUmV2bnBKYqr5e1eeSaQAq3c0Ae6aP8a9F039m/Vvh/JBJ/atwxKyAfLXikNsoRFZjlQBxSGXaQGGQae691OQKAI5IkHPUU0JkbkU8VIDub1p5nZQQq4FAEbKsp+YbWx1rufgV4xsPhf8V9E8T6sHls7GQSMsfLVxHmAthhnjrRJIq/KPnDDrQB1Xxo8a2fxC+LniPxPpyMthqNyZkDjBAwK5COctwOV9aYFUkgbgB1IqTywSMYx6DvQBIykkEsQtRlG3YVcDFN/eowByV/lVgSbAfXHPvQBHHZ+fFIoO3cpBNezX/xh0y+/Z3TwCLdk1GJgwkxwa8hScSYyCuBTTMT90AEdfWgCJY1/dknbtABPqalSTJ2qNx9aglG/oxHtTEdgcBuncd6AN/QPEF54Z1mx1bTrl7W/s5A8cicHIOcV7b+07+0zB+0V4S8KC5tmtdc0lfLuDt4kx/FXzt5ofCMSppEuNpDSfNzjNAE0soxkHk9ahkyQCv3qHTLbozuz29KFJgGcZz2oAY5DDgEEd6idPLGB8zH1qy8m8AldqmmmIDBHzZ70AJpuoXelX0F9ZXElndwMHjnibDIR0Ir6i8HftpfbdNt9I+J/ha18XWiAINRlQPcAetfMDJtI2/PmnLclCBnaQc0AfX0vxQ/ZbvP9Il8M6jBLnPlLF8v0plx+198N/AkLJ4A+HVnPefwXN/EMr7jivkczyMctjB54FMeVyMS4DdsUAejfEr9oTx18Tjcx6xq0kenStn+zY2/dKPYV5ixwcKCAKe/PzE5x1pz5kxk7c9KAEjZWfklB3U96aR+9ODkelPWHn5iG+tAjVW9KAFU+UcgZ9qldvNUEgJim7EYffxnrSojknDgqKADzCJVOecdKlLrHEzM2HY9D2FVmlIk3BMEd6eZNzEuueOKAJFGw5E/BprMm47nLZHWmb0BB2cd6dkL0XBNACsY2xh+fWmyLEpGXy2ODTTsUZA47j1qMICQWGQentQBJukBCEbwRTWlCY64H8NSo7CMuo3Y4prfeBYAEjGBQBIk2FJVfzqPowLDOelES7fmZsbfWpDMz4KqCKAEZVBG5CGPQrUoKqgDtj0qNJnjI3n5T/F6Ukka71ZySDyDQAokAIJG1u2Kmln3R+VIvDfxDtVdjjaTnnofSp1WMx5Vsr3oAjth5MmR8wHSkectJls59BTlDxHJ+63QUrRGT5ApBznNADGcrynfqtITuUspII6+lSwIHJR+GHRqjHmI52LvHpQBKtyCo2DkdR61GSM7yuT6elMglDSFWG1vX0p+7AZcnPr60AObMI3hsg9vSkdVG1sFgTSRIX+TYxJOM1NAfIdkYb6AIflTLZPX7vrSkux5J200BSTITjnpUhcqMdieKAGFtpyc7BTS3z7xkp3HrT9juxXPWl8poCQDz6UAI0cbgnO1T2pVEDHY7EDsaTaVJY8juKUbHUtt49KAI5F2OzIxZMUKflJcFR2Jp23Cll47AU53eRf3g4WgBiNg5Ocd/pUsk4iwCm5T0NJuC4dTlfT0pdyh23nJI4oBbjxhSMAsD1FW7QCG0uwzFyycD0qnExjOYzuB6irdsoCXBYlcpxigaPsPwOUP/BPnXowGMn24nB78V8UW4eOyhkAOzByBX2l4IkSP9grX8sxb7Yev0r4zQE6fCYzk4OQe1AMib9583IBGMCqWrwKlkWXtj+dW28xCGAqhqu5rcsc9v50CMaiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKO1FHagDXUnykBO0YFPJDIQ1NjcFE45CiiQlkbcMHtTWwnudHoDbbzTAT/y2X+dfTf7aAgT4aeC/L3bvJXJb6V8waEha80oO+1jOuD+NfUv7btvFH8MfA7LP5r+QvH4VHUvofIMp2OD94kVDIGXB6j+VTA/u+meO9RRZYkN+VUSMwFHB+U0uzCZ3fQUpYg7QMigpjkc0AIG5Af8AOp/JyOD09agEeeh/GncqOSSfXtQBKD5uVJ2+9IgUNy2abIh2D5uKWKIKgwe/WgBCcHjpSEFm+Y4xUgj6kEnFJuDcc5oAasgPB9enrSEb+ScqD09KcAMDC/NnqaN3zc0ACZIwemeKFByQSQAeKkMLqhP3h2FRZIHzc89PSgAYnacsTTUXc3XipmwEyo3e3pTFkbdnZgdM0APB2OD1HpVh8tIAqfKeuKqhtoK4JNSZcQjDfNmgBSArYK89jTYpBvIkHXoaUS7shufemzFSo5JoAAcSErnPalYF2zRC5+70B/ipzkDI7+tACONuFDbietBnMQIUk4FMk/dnIOWI60olbb0ycUASqv2mMuTgimKo2kF807eWixyBSFQACuaAFUMpBXn2pshEvzAYIHNAfMikEimSsFuM8rxQA8H7pB+alkmOQAflxyaTzEVQ2OfSmZU84yCPu0ASRjAwDwe9DDaVKvjHagBRHtB564phIGHydo9KAJGkY8gcGkKKQGXjFNV8YPPNCsQ3BO09aAHqM4pNwQMFPJ6048EYPXpUZAc5B+YdaAFXr1O7vSM5JC7ec9aCfLII5apI5FmULL8hB60ARqdpHPOetBJ35J4z1omj82UFXwo/WlVA3yhsYNAAY++TnsRQsBfHJU56nvS7AG+U7cGiSSRjyOPagB4tctlWyc9KfJGVUhjtPpUe8EY3baUy5UBufegBfMyg4wBURdmYnGMetP3+XhgKTe0md3yjNAAkTOc7sMKiljJmJ3ZwOTTnkRgQcrg0r5UDHSgBhQlOWwe1CxkjCnJ9aVl3Lk5I9PSmeSQ2EbANAFn7O7Djkjr9KTcHHoo/Wli8yUbEfaTwTQEaKTG4HHagBoTcCDwxHApSMIEbNSuWcnIGAO1M3grgNkY70ANb5F4OBTWl7he3Sgudw4Deg9Ka80p4xxQA9ZwyhcEewpOGO45GOvvTDksCcjiphcPLH5aqGA/iPagAt5MBjj6Zpcl2wfvHpmnQymVDGy8rzmkDlW2k/Meh9KABWOcHgd6UYXAYbj2qORvnBGc9CPWlibHzls7exoADGA2ScYpGU5GGyaWVjK2eQKjTCkc/KKAJ2XKDb261GYFlGdwU9OadLNJL0XCjp71EQHI3jHpQAGNgdrE49RSsD0J+YUbzERjgZ6etIV2ZJbPNACllyDyR3oUpu/ug05pBGQOu6mK4VuQWWgAI8sja2QT1qRSVHB61HnucopPShSVyDwp70AK0jFsc8UZCff5Y/dqRZERTg7qjcKRk8mgBrQcYY5JoAIBVWz6n0qQSMEyPmJ6CkCbckf6w0AIoEhxnZj170m7bnjAHb1pZEWQDJ+YU4r8oDkBR39aAGCTk5J59aVZSilSfk6+9PLgDlc46GkSVS+5kO4dBQA0sVIKnA96njlWQ8/KfSopW80lwNoxzUbOQo+Y4x94UASygs54CgCm4Drlc8UbvMQMSdnqaeVKEMpyuOlAESymLjblT604SEtyuD2NOV2DqGHDdhQ7MGKkYOOKAJUJzsJOTSBwcxnr6mmxyMy7R/wB9VGXw209T3oAlUlTg53dm9KY0eSWVue9BkaIjjPvQWL4P3cUARvAOBuJB60pjAwF4x39acspZS38XTHrTA5SPGc7u3pQAeapwG496a7BmARsqOtBUpgkZBpfLGCV+X2oAfD22nbg806YhnUA5XuaiyCdrZTPFSMmxRGOvY0AG3eMDkenpRGuOOtINpwAxVs8n2qQgIxOd3oKAGOSuMAj6UGTJGVPX0qRZ3TnHANK9w0nO3PPWgCNSAxKngnk01wQSHbJPSnI7K5CDfk0w480uxII7UAJu56fKOpp0gQsM5weh9KeFyNoGd3NSHzQoVYxgc0AQGFsZDfLTkAHy/ebtTmhaUEk7D3FIhCEheT3NADQV8zDDI70rpFKSYiVx2prMd+GyB/OnELncflxQAj+uTtHUHvSA9NrEj2pJsP8Ae5FEa4BK8L6UAS5G4NklR/D605pokOGQ5PIzUW2QyB1BDDovrVhpnY4ljAOKAImYKd2MCnGJZgHjYLxjHvSyEspCr0HNMj8vaOCDQBIIJFP3gqgcgUyViEwMtj+L0oKMQdrMeOaRndQuOQKAFBV4wzZzQ4EQHPymhHOOfunt6U7aqxlmfJ7CgBAuV/eN8vanRSiNezJ6VGknnDDjntTtiHocDvQAI5JLfd9j0NIfkKkjhv4RQSAu1fmYdKdExlB3cMvegB6SEtjO4jpjtSGeXdjsO9JHKnII2mjvgnFACxushO8nb60KxDkKTkc5qOMl2IwcD0p+8q2VOCOooASQrIoY7lOe1LLJwpC/KPSn/aTIBtTDd6JJWIHyYTuKAHG/lMW3AGemKckDSAFXAI65qIAS4K8EGmuoJyrnAPzYoAR0ZZSWwUB5A71KYyqF2JC/w00+WF5k79KBKkx+ZyqCgBojeQFlY8dzSqpYkM2WpQMttU/Ie9IyRrkd/WgBzMEBA5J70sQ3EqTspoVghI6djTimyI7n/eHoKAAxGJsFsg0yWby5MqMrjvRDncVPL9jSox81lZTvI4NAEgZdu9U46Y9KPtAZSBHu4pShjjPJD+lPilj8v/nm3f3oAWBoyQFyOORVm2/49L0ZJbYcZqsJUByqcjqamhm3Wt0VwMpQCZ9aeDvJj/YI1pmlbebw8H6V8c25kNjbmMkcHNfYPhuCMfsEaqWbDG9PX6V8gWhP2KBlOAAc0FMesxJGQV+ves/W3JtioHHGcfWtATkMCyZz0rP1pg1oxHXIz+dBJgUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRR2oA1Y8iNCTt460+Rxsw3XtSBt0UYC5OBzQ4O3HUetNbCe5s6W5E+m7sj96v86+k/20WjHgXwCIXJBtBuFfOOmRbrrTo93PmAj86+jP2v7QjwB4HnD5VbcA5qVuX0Pl9lyoOdvFQuxHG3B9amuBtYEHII6VCzsPvLz2NVckTggEHjvTioEZ9D0PrSLhx97A7ikALAjdxRcBCG8oAcHsKUOwix1pA7Y5OSOlKh3t1wvekAwHaAW79Kl8xj8o6VHnacZ3U8DJ+9igBsUpifA+6eo9KeZfLOc5B6U7ej/JjHvULKFIQNxmgCYyO4GBjmot2wHPrQCFOC3Q0odH4GSc84oAe7skYaNvmJ6UxwNu4t8xPOKa3Mg+bCjrUkv7tQEGVY8fWgBMbQcHrRGzMQrfdHekB2AgsSRTkO59xOFoAUuYn9QaAxZiTx7UjEl89RSu69AuCaAAnkYNBbCnPIpHJBHYevpTVOWwelACpIXYIx208kgEg/J6etRlk5Kjp/F6U7J3Kw6enrQAFAR8x4p4fK7RTVPUD5s9vSmlinFAE+/wAobgcn0pokYtlvlXFMiYHJY/SpHyPmPPHSgBTIqOJFGcDpSSSC45K7WxxSJKZSBjaewpjArJkffxzmgB0fz/KeWHSnHEbDJBY9x2phy2GTt1FPPzjcmBgc0AKfkOYRvJ61GSWcbuBTlYRpxwaiUlnGTQA9JQPlxketSHCnAOaiCHqDgD9alQrMMdKAGq5B3dhRKmWBj6mnlVT5S2RSFMEFGw1AEZQqcknPpSqxb5mGam35UZ5bvUJl2nPT2oAG+VM56HNNBDbTnvT3/fJknA701U2gc8Z4oAmlbOAPm5oLHPI3Ugw/G7BB5pRKYZAIznJ5oAMrJww20skKgcPmi4lVmwQQ3emxRBlOCS1ADRgoecU3Oz/aFKf3Q9/SmgeYxyce1AEjIpQtnKnj6UzcGRlJJI5FPU+WShO4elIGCHheO9ACJujQPu78j1pY0E7MQ20+lKE83/ZoWEJnLkmgAEiKSqgh+5pAwTr19aZsI6HLds0qyHuOR1NAEiTlH2jkGh5QpwF61ESvb86epDgZPNAAyqnzdzSEE89BUjR5xk80FSowT1oAjaQgYA/GlAC4wSaUow4J4oDC2PrmgB4bcQwOAKf52fmdfm7VFk9zgnoKkXjHINAAYzKPM3bMdvWq68sGIwKkkJkY7sgjsKBJjG9celACMxicEHKjnFKyCVg549FpAxL/ADDjtThnp2P6UANErF9pGfSn7RHhmO7mlEIA4bNIQGU57dKAFmUMwbOaBGGI3t8tImfIO4/So1wNpbPXigCRrc7s78HoD6UKXUYAB96YziUc9KeH2JyO9ACOWkGCMU0szqFPalDj8c0bCM4PBoARcYwmeOuKaDlvYdaerPAMqobnmmmQykkjZQBIIf4x0H6UoHfdmodwXK7sA0qsIznoO1AD9ib85INNIDZDfMBRnaT3z0pyqzHkjIoAFKnqcegqOYNk5Pb9KkKuDkc570H92MM27PagBiueCDkdhTspH/DtXutLMMjI4OP++feozIQuPvcdfWgB0bFpCyjKY/KpmjKAlmzkcD0qBHaPAxgGpCQSdxJ4oAEGxCGOe4NIg8vJLbt1PSMEZTjikDIpGeWoAFVh8m7YDzSeaEOMbyO9PcLJy4w+Plpke3G0jGepoAY0nzAsfl9KdkYDZOPSh08ogE7s05R5Xz/f9vSgCMMYyGzkdhSsdxBIwT29aUOZjll2jtQRswc0ANcGMDJz7UscjBhlePSkdSf3melKH4BzQA+WTzFCsmD2NNLEqMnkU7zQoAz83amh/lPHPegAOHQADLU1AR8oPzdSaQNsPAznqPSpY0Rjw2zPX3oAUgOBuPfrRGpgJJbg96axaPPGfSlIV2wx+U/zoAe0A+8knzZ6U0uykgKPMphABwpO7PamYKsT/FQA/LjLHI+lWDMdvBJNVw5fkNhe9PR9gba+4noPSgBzt5qEbsj1pibFJHVaag2MSTyacY8gkt+FACledwbePSmOQ4wxpGBGWGdvejk8jgUARtGI2BDcHtT1jyMkkA8Uj7WGTlj2pUchdrH6UATbWVPv/N2ocuI8s2XqLIOCTginGRsZz9aAJFlLR5x07+tNYl0IAwccCmoAxyrcd6VslgN2DjrQBNbOXjIz93rTXmCOCDlaWMqCOoYfw+tNLLk5G3I4oAXeD97oelMG3d8w69KcspYbcB8dx1phb5htGW96AH+WQeeB6UxtrOARtHbFKGlzl+vanyqJ0BVtsgoAaE2sATxThKELD+E96RATFgnkU7y9yfKfnHrQA0RodhJzg5xT3mEkgyuFHemwgb8MeDSnIOwnrQApEmcxnBp3mgcSJz60mTH94kZ44pxUupHUdjQBHIz5DIOOmaVpDLjLEEdcd6ZC7BsZ3e3p708zAvtUZPp60ARvI0bgL90/zpyFZeAcHPNIVLZycMOaazf3PvmgAkRA+0Dd61LGqSgg8KKiEg+4SQ9C3JiyBHwe/egCRPkk2n7vYelKTuJCnKjrUQUSKwaQ7+v0FPhmVVPPHQe9ACsx6EnbRHGpDAnL9s0mQN2WJyKft3oE3En1PagBiuc8Hp3qYTsMkcnHWotnzbcnFOZtmdnzADmgCVZXkXLn5u1H7uYHAJYdaaGSVMxviTvTkjdPuyY9fegB2MZKH5R1pbc/JctyQUxTURzJ98bPSr1tEI7S8L8nZ8tAup9U6Cm79gzUzzj7cf5V8hWoxZRNuwuMfjmvrvQzexfsI6gHQmJr3rj2r5FgKrZxBSSDnPtSRTElPOGPOOKoatGPsJZecYz+daCSJEuTy3as7V5WezbjAJH480xGFRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFHamgNWMbVT3HSpHcugAGKRG+RBnHAp4XKkbqL20E9za0V/LvdMYH5hIO3vX0V+11qDTeBPBcecx+SCePavnrRRu1DTQWwPNHOOvNfRv7YccaeAvBYRgSYBx+FZvcvofLNwB5w28LioX3E/O2PSnT8v6YFRAsxO/p2rV2JBiUZW6e3rTi2ZA2cKe1KhK5L84HFNTccBuPapAcBtz/FmnI7BeUx70ikhgTyaeSQvPIoAjC4G4GjgnHXPal/1XzZyvegsGHAw3agBssZVQM7j/Kl27lG08ij5sKc4OeaQSbiB9znmgB6iKVeG2sDz701mA4X5Dnr60KqMxBPTkmhXVWJYbl7UAIU4yTuIqVlJj549qNpcZjxRJ5hGH/OgCNGDfeHHc0HJ6fd7Gjbgbex70sZGNhbAoAAWRgeq96GAkBPQ07Zt/i4prkZ+Rs0ALFIMFJBv9DTMFiVzxT1cbSG60hyF4H40ANKCMEDp3NOC7kwTx6+lO8wRpyMmmbixII4xQA5R5LblORTRKSSxHWnEFV/pTgjNHuJGwdqAGRgAkjmnEljndTCCGBH3O9SIA/figBrF8gL+dNO9CCxqQHIYKaa52rg8k0AODFmGOB3pJ0AkGGJ4pAzcYG2nhdpBBznvQBEPQHinoozyaR1z9w/Me1KiuuMj8fSgBUYg7B85Pal3kdTg+lCuyMGACsO/rT2+dgxbazcUAN3qG3dQevtSElCPm4o5Q7SeKa8TIylSCO+aAJBKAu1Rh/WmlQ4GG+bvTcAOChyO+akMiKoCj5s80AKBsQr96kVwo6biO3pTEy3zJ98HpS7Gc7yQMHDCgBC+WzjbzQoLsVzhjSA7WKnkHoaVoiSOfxoAlKEHGQT6mmu7RNkHDe1McFcHdjFOwXXIfLUAOwrHJbLd6UsqD3NRkq/Hf1p6koGycgdKAEHzBscSGkKNypGHIpnmEKzD5vfFBUkb9/zUAObhQpJyOtSrEQcI2B6+tRkb03B8EdfepEGUHzc+lAETPtbHp3ochGGeQakuFKplvy9aikY+XjORj8qAELruw42mnt0GOtMLKV2ueccH1qeKdVG0rk0AOiO1csfoaY8e4HLfMelI/wC9bg8AcU49ArHB6g+lAESpwRuyR+lGdufm+Ud6VQQS6kDHXPekPz9wF74oAV5MKrHk9mpTJlRuXmml1JCg08HacE89qAHC4IGWXmosNI2X/CnxyAsVY07eqdeSelADJhtwSfwp27931xQAVxk7s9Kd5LsCGfA7UANUGQfK3FIzeWMN8w7VEisDtDYOaljVc7S2WoAYAQuSevGKVJBtIYZoL+UwU/NTiBHGX3ZxzigBjMWAASlWXfww20iSM/zE7c09cSNtJx/WgBgGSBnA7U6VTIAc4UdTT5SqjhcU9YQ6je21DQBGEaNOXyCeKbvZCf4hShBG5+bemaC2M88dsUANDjccr17elNdlBOTlR2x0o3FvcetOEm1DxkdzQAh3lTt5FPQuR8y5I6Got7uQFOE71KAykFXyfSgCRIePvnPfNRsiq2A2fT61IX38Z2nvTJCPu547mgB28hOeW9PWojknDDYT2qRo2jxzx2PpTRySSNxHSgBPLZgcnntTwSF2t+dJHvnzuO2QfdpgRtxDNyKAJBhkPJxSIAPuru/pSwpIQz9FHamDcsm9EOMc0ASbjnruOKchyu5eQB8xpVRW+YfKSOlOVjAQoTKnnHrQAxWEa5Dbl9Kaj4cE/JnoKSaREkVgm0NxilLh+D8xHQelAAGJOH4PbFI0ZTBdqUYK5HLe9MJDICpzjrmgB8j4Ubeaj5kQ/wAvSlDB0BUcjrQV27WB69RQBGpUMPmyRUyHIyDjHWkeRYudmc96ah2Dd/CetADmba2VXr3pCBL/AKzpUsg2RAn7ueKgRick9O2aAJ2QKAVkyewpqEIcNyDSBEbkja3rQck4Rgy+p9aAB2BIOMc0jSZ/hyfWnSRlFBLhiTzimhRjIP8AwGgCNiDnceP7vrTkC4LJ8uKOCdzGjG3JHegBTIxyAMj+9QsmCRnmnBc87sUgKyPwOlADiAD8rfKeoppjEmSflUUcFy2eB1FPZSwLbwB6UARhhGMryB2o3JN91cMKdvGCFHb71NDDP3dpx09aAJI9pfnlaDsMmGXC/wA6URtt3MdtBYOMD5j/AHvSgALK27y12qByaQSq2Bndx970pygKpw2KjRgh3EcYoAkafYowuRj73pTHGAGQ5X+IGlMg3D09MUBiW+U4HpQAkMqh/wB38vrTpDuICcH1pplL5VY+ncU+FgeCMH3oAI5GHDHNIpBfGdo7VKmJGCHg+vY0Tx+WF3rnHTFAAQUQ569h60kWVGScoTgj0polYEMVIx2pVmWQEBfLTvQA+OGRpOQNg6HvSbhvyxx6VGAV+VnwnYU8ogAJbDUAPJ2gh/nB6U2N2AZR34x6U1WKgE/MM9fSl8wSdsZP50AJFAwbKtg56+tMlYh/3f3hSA5JIbGD0pwYPwfkOevrQACYuPmGDnk+tIAYjvHAzzT5RtAXq2eDSRgyHbu280AWVCpHvVQ5P51V3tGxY8r/AHfSpVgKOWWTLCiXL5bsepoAbK6TgkDb6mkWMkYJ69KPLeMdmTPSkibE2Cc+ntQAsTiKQiQbm/pQ0mXORtX0pPKfzCGYEnoac2IiRKdy96AEErA4U/Ie9NOY24fnHpUj7QNyNmP09aZK6/w9MdfSgBWCYVz8hNPKjJKscYpBGXQnjBxgGnhGjT5iDxQAoRQR8xwat252WV4rNxt4qrFE7uMkYxwKkDmOyvQ/GFxn0pMZ9l2Grmy/4J+zqwDI93tzj2r4uVkiso3z94cCvsmOFf8Ah3S7scH7ePx4r4qs5c2sW7oKYMfj5iXPBHAqnqo/0M5PTGPzrSwAgYDzOOKz9VkLWTBkxyOfxoEYVFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUdKKXtTQGwj7UQ4z8opWbKn+HNNLBYU29MDNMPzqaHuJ7nQ6Gd+oaapOGMwA4r6N/bETZ4B8GEvlhCM/lXzloJ8nUNLdjys4PP1r6O/bHnE/gXwewI5gXp9Kn7RfQ+V5DucEc+tI6k/MxwB0pHwHBPSlb5WznI9KpkjgN6klgmOlR5KjJOSOhFKBvUg8e1KAMbW6dqQD+SoPr3pvb72KQlgMH7vakWPIOaAHLgZyMKaQLtPXimo2z7xpzSeg/GgA3kH5etOZ1dQWXaRRCgl+UHDetIP3b7XO7nigBGBbBzinxyCMEFdw7VG+XJ7CnMNp6igBFlGTn5cnrUnlFlJ37/amhfNxgjOaQho265XPNAC9TtBz6n0pv8W3NStLgEbc571HGwckAYNADgSDtPIpXODhFpjKQc56UhY84oAMEH5jilAYfxYWk6D5vwodWIBI+SgBw+dgoI9ifWiUbD1Bb2pgztJ7CljcMMkZxQAAM3LNRj3yvekMxkJyu0U+JDn5CMe9AAwIHXKUi7nXCDHrQV+bDkEe1PBEf3up/lQA1U2j71MkX585zxUvlMwJLbV7VGh+bHXAzQBKSHUEHn0oQMM88VExCMDnrQ0jlgF4XFAEwxncDyKQK033Tkd6ZvAHHTvSrMwQ7DgUAO+4MuOlIZVPY+xpUlOcZ3DvS+cTkYGKAEVlJUsdxpSxeUc4U9KInUH5hTXlLHAGPSgBxXy8c5bNKHQqxVN0g6io1V8ZB3N3oUSI27G31oAAqjlW2k9R6USqGxh8J3PrTncY+Vfm70isDhW4HpQAzbuwgPPY0pzHwWyD1FOd/s0g43bulRlWB3dT6UAKGDDr83ahST8meTQqOSSoAPpSOFXnp60ASSKYIjz83rToyZIsA/Me9RIjucZ3HtUqI65BG0d6AHMRbgZ+YUhlJXO3GajOXOM5I6U9eIzu5OelADhHvXIIFNMrxvjqPWmKxMh7e1SMcAsPm9RQAFlkbGd/vUZby2KnlD1oDBuQu32pRgZGRzQAnlK5JDZTFIHx8opckLsHApyAEBc80ARgkNwan83amCOfWopAQ2QelODL1LZoAco3ggthTTYovLf5j8lHmICMKSMc05pA5G7gAcUAKxWVvlXbgcGmoG4Y/MVoDhBjrmkDu6+mP1oAXh33n5ewFSeYFXGPmpiSBxyvSmu+HBzgUAOVlYkKd3rTVk3Z3NwOlSl1ZgkY+YjmmCIlgp+UDv60AJsbCuOWqRpE3KNvzdzTZV80BUbaBTAG2HbyR1oAccEHaRnvSiJnTCsM56VCZFfqpzQnXCsUPvQA9RsJ8z5R2p29QAdvGabJLvABOSD2pqE559elAFgmIYbBPPT0pjyiRuD8g/hNKrNGMlMr60+Z4wV2Lhu9ADGcohwcqe1K8qyQ4CYx3oK+YSFbD/pUb7oJOuT3xQA+Da6gBuR2psyMuR/DUpmjjUOFw56mmmKQ/vAc5oAhCOqlv4R1qcKWXzFxgdqbGsjEgDHsaa+5TgkY74oAe8gk9jTWAzjOR2qMhg3y9DTypQHPAoAfvIXa3BpQWjwcg5pqKXQ7TkepoihyWJcn60AKVyxKvg96TbsbIOR3pWG4kK2DjmkAbjB470AIWkfgDKetKkskJwDuzS7gJPlbjHSlMnzjaRjFACtmTGRtPp60+J2GGX52FNeTI5YbsdKEDY3K+B3FABJL5r73GwDtQhWbG35X7UpnZ8KwGP71MmwOI+GHOaACSN4xk8etRqhVMqcjvRsGAxfJ9KjwwwTwvpQA9fkGMYBp627EZU7vamxIxBOc+1RpI+4/w4oAmUkEqTkfypsQJcru+Q09HIBPHvSApIM7tpoADmLH8a56UiyxODjg5+7SqpjP3uCelNG0PkDcM0AOygGc59qQzBlCgbB60vmRs+3YV9aAoB25Cr2BoARU8sZVsnrj1phPzBs/Ke/pUkpM6jaQuOtOjMbKVY/N6etADIsndmnhXCseoph4RsHB9KRYjtOWwx7UAKiZbltvtTnTyzu3Ae1BUAH5sn2pikMSWbJHrQBMyB8srgMe1MZCOrZpksSMpaMketPjyEJHK+tACFjHnDbfapWKsuWPI7VCMtkk4pzDI+981ACyN5ycE7h2pqssY3KcAdafHgoVU892oYGFSWGce1ADFmDnJ+XFTogx5juAB/DVdXEg3Lwe4p4jPlsyHcO49KAHowkLN93HShNkmQpw+OD6VHAPNX5WCsOlH2dvx9aAFMckbDnacdfWlZckZODSZcjB6Acmms2GXecL/AA+9AExleOPCr83qKWORio4Ln1PahJCDv3bcdqYZZH4UYHegCVJTECSd57e1NR1cZX5fWo1YNwDx3pD+5IZTu9qAFWYK3TcKUyo+ARiPt9aiSYMSWTappzShFA2HGeKALKxB1G2QAdxTI1KyYk4UHiq/liQA7intUySADbLkjPymgBzbC5I+XFMYo68jPPBp4UqMnBAOQtJ55cEsm0ZoANxAwCGprNlSpPzHpQSrHCnr1PpT5TFEgXbk9moAiXjjndUsTY3bW+oqOMq7/K2HoZwj4ZcA9aAHklizbvlHWhCkhwTsHZvWnAZG5SMdMetIsCDo3Pb0oAQQK7EFjxR5cSg7smnFyTg8e9IM525GD3NACqsaAsuQPQ0OmwFsjYR0oYlImAAf6VEhLg5OD6UATLG7ADeFU9Ke+1PlL7jUaKoPzHJ/lSOkang8460AOB+bKMVIqWGTfp15u+ZtvU/WoVcMQNwBA6+tTRLv0+6weNvWhsaPrqW/c/8ABO/yy2R/aAx+VfHNqoFpEynJxyK+xXG3/gnixBDD+0ce/Svji1ybSNQM8UkDJ45SBgfKp/WqGruTasOoyOfxq554wuUwRxmqOqtm0b6j+dMRi0UUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRQKaA1xgwoB12ik4xnNOXCxIf8AYH8qaTlCtD3E9zpfDoE2qaUrHgzKOfrX0X+2XH5HgjwhH8pAhGCPoK+c9BzJfaUM4YTrj86+iv2yHLeCfB6MwIEIz+VT9opbHypKPnXJ4pjZVsg81LcffXkcDioWk2E4UnPXFUxEisCM9x29aAQTnP4VGrK5BUbcU8tnjbtpASKCykHpTEfa2CcelPWTeuD2ppTec0ANOT945pysc88CmkggA03cC209DQBIpAJ2jB7mk3bjgEbu9MIZOvA7GnqQ3cD1xQA592AM/lSBzkDO45pXR2VQWCjPBPekb5vlXg9zQA8R8HBwueaYoxkq2cGn26ES7Q2TjimnKOwJ5oAekm8kDr3FMdxnGMH1pFUrnaRup7DKEswJFADG+Yfex7+tAHynsKMnb82NvajcCpHagA2s2QBup7GQR7WFQZZCdvC09HJ5Y89qAHqOCxPHpTY5FU9ODSEAZKnpTywaPcSOKAEcgcDpQcKOTgUisTyTkCkYhu2KAFVV7Nkd/apQBjaDuJqIIEB560o4GAaAHNwdpfimhvLbrwetH3WwcZ9aa6lzgngUAOXDHLH8acI2AyrZFMOBwTgCm7SDlTgYoAlXcvCjOetNeFyPkOc9aRcsMg49qcEA/jx7UAOCmNQx5xRv3jIOKULhSD92kjCbcyH5h92gA++OeD6Uisp+V+D2qRUP32YZ9O9RkgtuxtPagAjiIJKPx3pzqyY3NmmhgTwMetNkYqBuOaAJxIYwGyMelML7huPHPWoyhZRk5HenKSF6gY6ZoAkLbzuJCkdKjyGOFO096AqyJ8x5pdoHSgBHAPAJ3d/pTOhAzwO9Kr84NPIRVzu70AIg2EjdgHqaeiGIsxfI7VHGyFiMjZ61KwDLweQeKAGA7RvJzinuVxvDc0hUE8HjvSGHIwOo5FACHEnPR6UAhTzQJOoPXvT3dQpI496AIt3zgjrTjCx+daRduCOg70qAqSAce1ADA5c4H3/T1pWGzofmp0qlCCRg1HvIk+Y9utAD3bYVOc5pHYsMAdaSQ5OQ3GOKZ5mcbutAC+YydR0704YkIJPHYUsIZgd5ASmgFWwGBQ0AOAY4OelPMm0jJ/CkJZl69KRYyQf50AOU4Oc4zTPMAOOo7U5GMY+Zd1NWQOD+7x6UASBjuDD5WFSORcncTimCQgDcPwpsmGZW5HsO9ADxHtHDYpwJHQgep9aiEabsliR3pxACkZIHbNACzycKY1+UfeNI7CVMAAe/rSMwYKN2MHkjvSBsYzzz8tACRo/AyN2akPykKTzmojjO0nknk051SNMZ79aALAYoQCw68VGxkJzke1RogPzKdzZ6UTIc8Nj1PpQA7d++CscA/wAVPaMqG7+hqFQQu1mBFSMwkQeWeVPQ0ADkEYPJqSMArw+DTFw4LKwyOopjsZGwDgetAFiaN1QksCPUVFvAXhhjuD1pqZyVL5FCupLKetADduXzu+WnffcruBHahCFyp+6aJcqAByn970oASPIJBzn2qdYCBkdKgTAcLnr0NPyVJG/kUALjEu5eTjkU+TEgwPlb+760keJiQB8wHWhk2nDn5uzUANEbICWX5ccmmjI+Ycj1FLHNIJCjHMeKlWJgw2Y2Ht6UAIHjHzFctio3LZUj5F7/AEqYgB9oIJIprxOAoJ3CgBC4KAdQelOYhcAMAe9RzSAbQBjHamvcKCNw+btQA10Vz6UxMjK7smgyEjAp8Lh+eAy0ANB3EbTg0qyBTh+femKQJSf4e+Kkykh64AoAbIdzDnCU4KOo5o5VgNvHapVEq4wmQaACNsqS/wA2O1R+dghgnljNP3HnK7DSF3UfMu4dqAEYhnyePcd6c4D4XO0dvemZ5+bkH9KFBLbWYH0oAApOV3ACk8uQDcB+8H8qJ3EmFI2lf4vWhRI5+VsEdfcUAIzYxxg9z6VMkodWI+Vx0PrUA3xkjjaafMcbdoyO+KAGiGRnJb/61SkYU7mApnmFshjtHagwDBYvlu1ACvLtAwMetN8xgflPXt60iuGQjgfXvTfmQ/LjkdKAJlkbJ3R8+tNZ137gMMBUhupBDt4BqMSY5PJI5oAlEu5eEz60F3KbSR7VF5oBIUYJFOZkQjd3oAdjkMDkD71OHzSZjJxj7vrUW6PJCnHFIZC20AYA6kUATSASMP8Alm54GKd9ncLw/Sq5Ib7pJHdjRkzdW24oAl2MRguMUx3BAD/NimsEcYZ/mx0pYiI1w/I9aAHo6nC5G2nK3qePakEiRRsEX52FRxN8u0nc56YoAUkFgR8opVYAjadx7/SgowGG+8KaGXG1Plb19aAJWUOobIZB2FJsZUD5DAHjHamp8oBB57+9LI2T12qeooAc0gONpwT1pqAs3HPP5U3cseFIA9KIpGYnJwB39aAHs218Zyc800yYPzfMM9PSlbcBkkZ7E9qau2QHHX9KAJJALgYUgdsU7Y0YHmEFAcVXd2HCDGOpFOjbK/Oc0ASmNJm3Idp6Z9KV5Ps+EI8wHvUbKrfxbAOtSnayg7gWHSgALBsbeB6UisEOAQV96Q/KCSQCetNBUjap5NAErqdjBeU7n0qPAZCA3y96ezyCMRgjHemKRGD3XvQAbdikod1ORo5VIcbZDwKb97ocegpy/IcsuTQAqISzAt0H507MWeRgd6i6EsDhvSnBxyDjGO9AEiCHdyuPepYGC2V2qYwR0P1ql57Z28e1W4GK2s5UgqV5otoC3PrhNNZP+CfE0y3KuP7SwY/wr48gZltU4+fHHtX13awFv2B5m83j+0vuj6V8kjm3iP8AHjH4UkU0MzxuzkHrVHVRm1YjHUfzq8yMhO4FeKpap/x5N+H86ZJiUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABR2ooFNAa8Y3KAcH5RgUnOxicKcYp6qhiQr94KOaR13RndyaFuJ7nSeG1Mur6QmQCZV5/GvoP9sqB7Xwx4QRnDZgH8hXgPhpM69o6hv+Wq/hzX0D+2Za7fCfhGTfvfygB+VS9y+h8rXGSwzw3aoxkkVNM2HBIAao2254Ix3pkjeCcEdOlOWTepDDDdqTjHue/pSKhJz3HegB4jZQDmlJ2pxxSZKsB/DS7gpII4oAY0ZZQQaaAC3J/GnrHk5Q4HpTGYAcrgUASh2UYYgg0okCA7Y9x9aZGyquXGR2oDkj+4uaAJQxkHzkcHgUzeQTuHy0gZRkMMZ6H1pzN0GOlAD1m8oghevQ1EwKkuCHyamVgR8pGfQ1GZI1ztXPr6GgBgfceuB3pRDuJOaUFAcAcH9KVVZSQXBHagBEY/MpANGOvP4Ujt8p3dqZEC3zUASOM9T07UgO7oOaNpbkkA+9PjiI6MKAGBSpORgGnMmTgEY9aftKg7yMUwjAIBFADWYROVzkYpTMgAyOaMBuuPalGEwXwaAGs29eO3alDEj7vPrTAWDbl6elPSXAJJxQAuXXgjrTdxznjFN2ksQHzntSMckKePf1oAkLK7ZyDjtTPOJYDHTtQ21SAOtNHDdMmgB5bPVcGnxsSeQDQsoP1FEOTuJwBQBJkl+fm4p8amEZbDj07io4m35xxRt25IOfagBBGrHcrknrg1JlJ1yxwR2qMEZxjbSqnO7oKAHB1QDjD5pjYP1NEY3H8etPYiIkjr60AN8oqoBOAaNjLxIvH8NJKS4Bc89qV5X8sK7bj/AAmgBMIvLcewpSy5+X9KR3UKP3eT3oSWPGCuDQA7YCAe9IT/AA7RjvSomWG3pSyL5fWgBg2gEbfkpzOu3j5f60nmK67V+X3pEk5YNgj3oAfHMFBBTIPU0eYd3GAKaHO1iuAB2pC/AzzmgAIZiSOc0jJhPm4HYUrSPD070h81iWxn2oAfglMg4I6CgZkGTw1IWbksu009GKjecGgAkjKgFmppXzEPQmn4O7czA5pjAHJyA3tQAghIXcRgU1Y0AJxx3zTeehfHpmmpJmbCrxjmgCSIB2wfwFOkVkAwOnIpCGV+VzxxjtTjI2MMPmx19KAHKC4DDAqIlw2AeD6UEb8ENtHejecjaQAPWgBw3J3pwy7DnApXkEqdACKYXbGD19aAHnGfmNEbkAhuV7U0IQNxOfamlxtOOG9KAJI2BbaB+PrUkpWUAEqoHUCmIQ8Q7EU1os7XU5NACs4WMjYCB0z600ZKgk/MDUrzKihcc5qNhnIHBPf0oABJHjbjqeaSQdN5AHahFyMEjOaesir8sg389aAGdSCOMHr61I670LEgEelNMak4jPPpQ4YAf3gelACmUy43Ko20plXkLw3r601QWyMYGeaeUIOMA+9AB5MRGVPzd1oaNU7/AC9xSNvj+YL81OeVduFAyetACLIoPyrluxpr8tjGHPf0pwYD7uAT60PcHG0oB9aACeRCojHDevrT0OQUyAQOKieMypzxjpTF3owZhjFAD5OECkYenKRF8x5FJnqT84PekVwRg9R09aAJFcb9y/InvRIxDbuNmO9Lu3ffHOOMdKZIkjHkKRjjHagCUzqI9zD5h0psEr53E7RUaA7fmxxStJ9obC4CgfrQA5I0EhIOCRTi/lEFWz3xUWHBCk5NBUd8Y96AHvKszqxADdjTHiOezUinfGQ2DjpimpIZAVJ2nt60AIgZTngilaPY6lTkd6dGpIz0Udc0iryNvA70AOWEEEtgA0zfGCARx2pXDbgDzn9KaEI6jcKAHlizDPyjtRJLOuMfgaA/GCMilUkt97j0oAUTNIo3EZFHmM3A/LtimMRvBH4kd6kMybSAvzUANViCQQM9qaVBbrz60pUkdRu9KVSrA5I464oAYG4IPPpSvG+3722mMCxHcA85qTErnI+ZR2oADC+0Z6UfxcHjvSFTuI3Z9vSlwo7/AI0APKNt5G6o3BY8ninZdT97Kd6V0PVelADVj+cY+YClkK7vl/GpUR1znAHvTSYxkAAg9TQAwqRlgdy01SWc8g8UuwlCAdoz0pzLuKsmFwKAHyBTFuJGR2FRrMr8MueOKVwJBwV4pwjZhkEcCgBrSoYsbfm7UqgADBHPrSDBHH3h1ppVm5BG0DvQAsyNFheFbrx3pdm4DBGO/rTvNEqFZPvAcGnRxukRwNwPU0ANZ4R0XeaYXZ8g4GOlORM8RrinfIAQ3JoAZFKAcOA3sO1SCaNcCNMk9zRE0ajLIc+tNXyydy5GO1AA7YGfvZpvmoygbfoacsu44UCMd896VQznIGCOw70ANMZbG7j0xQuWByfmHShnVCMgv7DtSGVJcbkOM9qAD5Tww59aXCou3GPeml0wVByO5pySbWxjPpQAisCPm+ZR2qRpRINqJsFRsVJOBgd6VVLcbuKAERf4d3Hr601WDAhR0NSH5mC8YFMVfLYkYKd6ACOQbsEgr3qba55QYA71GNhfBXCmgsSSm/alADg6JuMpyKWJAz5j7+tNLKp2sm5PWnXDCRAEHl4oAsbNhPmFWpvysSFqCMKo+Zs1MjbDuOAPegBEGHYe3FMaOVHBOTTpgHcMDnFNM0oGFYH3NACl2c8J8w71GTIkm1kyDSq0nOW5pxDseTmgCQwGPDMQRjoadGD9kuMcDb0pqRhsgyZOKliytncZI+7/AFpvYXU+t7NPK/YGk6bTf/j0r5EGPs0ZPpgV9aLayr+wY04kGz+0cbPwr5JVitvGFIyRzUJ6FsY2ZPvuSAOM1R1Tctmw6g45/GrrkFgvGO31qlqhJs2yecjP51RJi0UUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABSikozxQBsooWJCv3cDNJI22FttEZIiUDHQUFtsTcCmhPc6bwoxXWtIbjPmD+dfQ37YzF/BfhU8bfLXNfO/hnA1bSsY/1q/zr6N/bDdD4A8LABeY16VF9S+h8qTOFcbl3ConWPcOMKalcbxtUADbUSvs+RsEGqJGMqqpwflNNEhOOeKlkjUjC8CmR4ZiVAAFAEo2snJ5pqqc7cZFKHTbgDJPcUsRBYj8qAGumQADg0gjZVzkdae4cHLLx601QH53d+lADAN2QTg/pSg8fOcgdMVICAdrKGWmmDHIIC9s9aAEHzfMMEjoDSu3mEAHr39KAuzBP3PehiJDsTCj+9QBGVAO1j78dTUkfz4BwopoVxwVBI6GngiNwCck+lAArhSwO3bTJQJDnsOlLIPmDbab8zZAXAoAN275T0pFl8s4PI7Uhyp7YpduR2waAFK7zkdKcvXnio0R0OR92pNwJx0oAfhSDk01sAZp2VI6fjSoBjBxigA+8AR1FNIO4cAml4wQvGO9IjkcZyPWgBApzz8tRFdxIznFSD5GOTvU96kEKschqAINrY5GD60KQ4Cueexp/zM2GOVpXVF4PBoAjaBkb72RTmUleegpPLI43celOUZGc9KABMfxdO1Oz5nbalIpVic9aAX6DpQA9FIbKEADtSI4UsOmTUZUq/wB7HtTlK7tx5HegCRiI26bt1N3lGB6pnvQX+bKgAUwburY25oAfzvJwMdgKASQSRj607aysCCMUj7jJlTn1oAR8Yw34Cmhg3ysPoakcKh+Ybj1pJB8ozgE9KAFUOD1p0nzjkDNRqGUckZ709XBJ6UAJHnk5xineWTkhsk02Jhhs+tIp3MSKAFKfLinHDEBcb/enEZXjt1qKb76uvTvQA47AMEfMfTpTQQhzt3expWD5Dr+FBLE7hye9ACmRZR83B7UjO6dGx7Ux3GckYNPGT6bu1ADSz53Mc+1OBLDLEKR0odm68bhQuSDxk0AKQ/U4H9aaGH8QAPtQE3Ny3IHSnkbgScZFAEJEZc5Pbign0O1qUghs8E0roSofjI7UARBcfMJCSeDUoQMuwtkn+KmRoxJxtx39ackyICqoScdaAFjwvyMKTYXbaOM9/ShBk9eoqR/3aALwe5oAjVyj7McetPZkB5z/AIUzbtU4+YnvSxnyyON4PWgCRPLcYD/jT9yp1XOO9Ry4+XaFUDkZp6K7Df8ALu9KAGxsEcsFPPY1Kswc8ALUSljwuAe9Pj+dguOR2oAMDHJBOacGSRCjLg5qP5mf5gEI6A015BnGOaAFyGyvGM8GpAFQfPgjPGKhB3LnA4NSBxLgEADNAAWG0sg2jPU0gIJz/FSsjuhGRjNMUAHDglR6UASLIuD3oZfMU5YoB0xTSMsWG0HsB3qTG4AgfN6UARqzAEbiV96b5WDkENUjIVPICluMGldCFwFGR0xQA2SU9ABn+VDEyoQ2PrTVzJhQnzdzSyjauwkbv50AKP8AVn5unSm4diWZgf8AZPejcpwo4YdqApWTJ5HrQAfvGypUKO2KYjclSuCeMmldsyEq3GKVpTuHy8etAD1RuVDYHvTWBQ/LJ/8AWoBaTluCOlMP7s9Mg9aAHg4XrljT9iqPl6+lRK6Dk5JPGPanq3mPheAOhoAcSZMDHAokmUrjbmmncjZ9qVZhtxjk0AEZwdpxtPf0psxVj/ddBxjvTDyw/u1NyhGVDe1ADVZpYxzRn5MbsGjy1VCxbC01Gj6NnJ6GgARHBHzc0/bIvGflPU0k8ao64b5TToU8zdub5R0NADIxIpbjK9jQqnf/ALJ4qXYflVGyvemYKcEd6AE4Rhu4XOKl8hz83Cio8oSCwyAenpTwu4/PJhc/doAGKwg5U5PcdKjEYY+YpG6rQbep6MqngetVmXcS0fBzyKAEcInO47j1ApUIOdrYXuKaoIJYrjBp6R8EkcHv6UAPeMOMqQP60zeAcFf8KV2Lj5SFFP6qVIBFAAqeblVOU74poyGKhsL3NNLY+VD5fr700rtB/u+tAD5LdlJIk3AjpSAlUC4AHemkFPutwe1GG25BoAlHHzcHFQqSSSf++aYAVO7Ofb0qc7hh1APrQALy+B0I5p/3RsU9uTTQSknOCjCidvLIVSCSKAEACsGHzY601piSdgwuKdvwMqBnHNIbjy8cDFABG2xgCMoetPdTCf3cmUI6GmB0kPAwtSDGwjAI7H0oAQGQAMKaxKfPwSetMUMSQrbf604O2QJE2gd/WgBSzKuQM/ypFbz8H7pFG8Hp8opEJBHGKAJP9cACMY60rSOqBVxxz70hztyBtpqKQAwHJPNADtxGGTB9Qe1J5hD5XHPWhUKsSO/XFGFUccr/ALNADndNoAVQc/dphbafr2ppxuBOC38JpWVkAJxuzxTSAVwwGCv40ik9GOAe9OBYgEtuwaViGYoE4bqfSkA14o1UDeSc9KG5AUYC55zSyQNHH82Ac8H1pkaux4Xcc0APjPlghjkE8YpQ8auQwytIFMT7pBwOgpzy7iXWPaKAHBTklSCvoe1MZyp5AZalCrNyjbW74qMxKX4OKAAgv3Ge1I8BjH7xvwp2I4n2nk+tK8atlt5OOxoAXy32t5eCMUPkKo4U96RSvQE5PWkaPqA244oAcX2OAQCD6UpYNlRwh60xiMAOuCKeGUKMg/hQARQKG45xVqJiLK5yASFqujROflGMVYtRttrvocpRfQXU+sLi1kP7AccwcLnUgMfhXyOFEcKgY5HNfX9++P8AgnjAwUD/AImYH6V8grMI4oy4BUjtSSNHsRuiou4Abe+KztTO+1Ygdx/OtCaWOThVwPes7U2/0YjHcUyDIooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACgUUdqANeIDy1+gokGEJ4pVUCJM9wKc+DEcjpTWwHSeGZY4tS0vegb96pOPrX0b+2N5LfDrwo0aAboh1+lfOvhYrHqulsQD+9FfRX7ZTK3gDwiAoGYh/IVHUrofJ8pyFC8fLzTCgVQFIJ96e7AYXgYpixsxO3FUSGxjwuAaCS67NoVxQ+Y1wetRhWPIoAkA2YCge+aJXxgr1pYsHrgsOlLIGEgIX5qAG7pJVwzcelKijOB1pCAOn3j1FAdd2wfLQA5Z9hxs3YpXG7DUwbj90CpA5jGCKAGSksB3FNUYOCOvb0p28VHkq+aAJVkYLtBBNOYqFAKgmoBlSQB8570q+43GgCfYwhL7hgVFuYAgHAPpQrZ+XPXtTSCvFAAYS3zEjA9aAABk04cA0pHAyOKAGcg5Ug+1O278Fht9KX5Ow5pVYbTmgAWQgEOvHrTsCRfl/KkTBbHGKbgpJlelACqNnBAxSPsXtTpMMMk4pqlVY7+RigBofIIA4NAUMdu7b6UrMHxtHHamyJsQHqf5UADgqwBY0bNrbsk/WnDgYPJ9fSkKhVIagBRtB+ckHtilJAPyfd9aRWQjBHNKAzdcYoAMAn5etIASMZxS444/OkVSp45zQAigxk5+f3pQWB4G4GkJ28A5z1pctAPlXKnqaAHOvI3cUhjLJg9KVFUneTkUK6qW28g0AKgKgYwcevalA4OTTTIF+71NAyq8YyaAEPy/MDuPcelOd1ZMFc560p4AHc9qZHFuY9sUAA6cnIp6hQ2B+VBhiOcPz6U5YkHAPHrQAh8vGB97NMBZfal2jcSTnFOUo3B60AKoIHXIo3clcZU/pSH5DxTi+1Tn71ACeVKqEpg4PA9KjAdH3H8afEHDMxPz+lMYlyd3FAEkirIOB81NJ2HpSFd33T0od9vDc+9ADmcBc4FNJYrnP4UpKEfL19KaJNh/mKAFGWIbjPvT5AOp6+gpC6rjyxgnrSqNsgDYNAEbypgYHNDEsQTgYqSRAHYKARio/lKgDrQA0yYyduDTxmMDbjDdR6U3ynznipIysQJYZNACO/klSuD60/IkywwB3FRLIHJ+UbvenhEHfaaAIwcscdMVNC2xcnHpTWRQBngUuA4wBigB0g3r6BeQaaofZuHSkTdEcEZFSSES4CHb7UANj/AHhwThhTlJVhkbT61EEIbJOGHSnFWJBDbye1ACzREuJA4Zu1IwZhuZfm7/SjHl44AOaYfMBLbs89KAFOdowOM1M/lhQH4B700kbBkU5mSZQuKACVfJZVVtwPNIuDncPl7jvT5o2RkRwN/wDDmox948c0ANVkjkIwWU9DRGytKfnYHsaeAbfJIHNRsd53YwRQBJ5O9iXk3HtSM/lAqjb2pjgkZB/DNNEfGeBQBKkrLgfKM9zTTGQxYHeaQiNI8HLE+lSY3AMjbh3B7UAMkCH5kGZKQys5xgrSMNhJT8akM4eNk2846+lAETcZ7cdqcu1lAJxUcWZM5HK9Kl6Y460AL8ipgsSe1M3A/SkdGVyNoHHX1pwVmGWGAPSgByImMsfbHrUkceNwUjp3qAHfjjipMbWDLj6GgByQswJY0xbby2OSOasq24Dd1pssYcjntQBXkRYiuT8vepQPNI/u9j3pcJEmZF3CnFQACmAp/hoAYsQztb5l96lkeFMAr06GmKMEH739KcZFwDjdQA1sSLwoIPrUUn3gikbD1xT0GSxi/EUxEI5H3qAJFTbGF6c9TSREg5J3EGgkMBzz6UkfLEHjPSgAndSRsG0k80qbVJEgzzSS5jUbk47GohJwCfm56UAWBKp4RcLnmmMpikJBwp605AcnGNvehnDqQBx7UAROrrhgdy+lKd5jwGyD2pNpUnZ0701WiDcsQ3pQA9ZAq7SOlLkOCoODUSptYtnI/WpfLXG4jCn86AF39Y2wfRqbs2r8549KeChU7h8o7mozgEv95R/DQA/crAhBx60YxwVzQfmIC/dPWkJZlGOi0AIw4Py4A6UIjSKSOMfw+tK75QjOM06PcPmzwP4qAEKvgrtwcdKVI9wGB83vTopHO7fyPWmNKHOUGMCgAlkCfKAMd80ilTgcEH9KdwwA4OetIHijYBU3GgBvyRSDGGWpCu3kcqeo9KYxxyAOaaW7A0ALtKNkUbJWZTu3Dml3cYP3vWmsGVQVfC0AINyEArzTiHJ44PvSEkpy2X7VIHYRjcAW9qAERmLASZ46U4HHJanbn+UstIhQHEi5zQBGrujbx81SGVGUELtpCVZti9BzSMpIFACFkmHAwR0p0YZmAfB549aaitkYUbvenFliIL8PngDpQApdNxVAQT94noKVEIOAdvPB9aazCUENtAz1FETbZCrNwehoATzGZz5mDt7GnqryPuTC+wpsiqpO5st6etLCcjajYGeT6UABkPmfvMEr3NSFyODjB7CoGcAMNu/nrU0a+bGV4DHv6UABjXbn7vPNJGTG5R1DxnuKSMOGZWG7H8NOlEkYzxt9fSgB21HyY+g7GklkVcFlAPtTUPl8MMqehqYspHGCnegCNE8xiQQB6d6jlPznbwcVK7IjZA5xxSCYYOV60AMQhlYMw3AdfWnxl9nABApv7pmPGBSxKZ9+1tmB931oAAw3BgMH0FTRS7YboHG1kx/Kq0SODtHJqSKNhb3O4D7tJi6n1/qmY/8AgndHkLgaoMflXxzDIZbdOBkCvrHU7iV/+CfsUXG3+1Af0r5MCMIIjjnHBplsm37cEqCfSqequTaMGAByP51ehBZs4G4Dk1Q1Xi2cHrkfzoJMaiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKOxoo7UAbCEPGigcgUrsPKOR0oV/3UYA5wOaaR8pzjmncT3Oq8NqDqWlnAK+aOp96+hf2yjs8E+DVABzAMflXzpoCMmoaX6+atfQ/7ZSlfBXgssuD5A/lUdTS2h8rzEKR5g5NQhijErU0xAcHqO9IRGTk8Y6CqIEBbILUqIDuJ6VHsdgcfc71IoATa3APSgAVRtyOo6UGUkgrywpNpUj1HQUrIVIk79xQOxGco+8de4qRnVz93BpMrnpyaX5R1oEOXgbu9IgIJY80iOn8NG/5vl696AF2jcARgGmupZcLj+tPYBgATt96YW5ATjnr60AC9MEZx60jSr0K7ad8xYqRSOofA209AGnBG5eCKRd7jJ4qTd8v3fu80nMq5zjPajQCTy8J1BJpiqScYzigJj7wxjv60u0tyOKQChTyCuPek24+UfMaXMj8E5ApQyocjg09AAJwTnpTCjA/LzSANkt1FDSNHyBjNILMehKg7lwKbtEmSMdKcSWU56dvrSBxzu+/jgetAWYFFwoxg4poUt8p6jv608joD0NMk4wQelPQBuCDjGDStlxtIzUiycZxn3pom3NjbijQBRjHzAAilEbBSQN1N2lif7uKdHlFyrYFIBApHJXFIcg4UZFDszADOVpUyevFPQBoUowPX3qRXMeV4aM1EzFXwvIpyldp7rSDURo8nhsJQI37DA9fWnoR9QOlJvkcgfdHpQFmKB5i5PUUgGxsHvSsjld23GKRI2fHGc0APkVH27DyOtKCJEKggP7nrTVC5xjGP1pzgHBVeRQAxF2HkYNIynd8pxUqo0hztxTJQIzwOe9AWYhBjPTJ7GgBiM4pyuWHyjBpku4pjNAWZJGVlOGO0057f0fIHSo4dsylC2CKcAFOGOQO9AWYLhW5OX7elIFbcQB1/KlZxJlAMJ60Kyhvn5A6GnoFmNlVs5HH0oBwuDzmnDB3Y5Q9ajmIOAvTuaTsAEDdkDtSt5YUk/ex0pBuKbOw70MUCkHrQrDsKqlgMDBpHKLwOTQIz95fSjyynG3Oe9AWFjJXj160m1d3y4prYjbG3PvU6wqfmX8RQFho3EcAH6UmcEYHPcGnGZQw8scgUxyfMUtxnv6UBYHUNz0IpHQlefm460ZAY4OP604MRkgYx3oCw5lUQqXbNNLBgKQNuIJOadvVFyeCKBWY9Zdo2hN1KpDrgjYaFlZlLA5pHZsqwXcTQFmKvDYkIpmdkp24zSTZcZK/NSIA68jpQAoHmuCx79KSXaOA2OelJIASAop7ANHyORQAnmJkDHNPU7/lPy+9QBslTjvUwIHJHfigB4Ys3LlgPWg4VuRg9qYTz05zxUhdFQkrzQFmRFmyWbnFNM6HjZg0qOj5B454oKkE5G6gLMaAp6nC05gm3EZ3k9QadlfKG5MDNM82KLJUE+/pQFmOtyEl27Tt7k1IV4JXqaiAYgknHpnvT3YNg5wfSgLMGQgHNMVDuwRhT/FThMSdueaQ5OQTQFmEjbZeBnjg0myQnkYBHWiMkjaePSgmXJU8gUADOwwc7hQ8jSZwBwOlNDYIGPxp7DZjgAnv60AMUFQAEz9alYeUR8uc/pQpkD7ge1Bn5xnJPWgCZkKx/L81RjKn1GKVWOcLwMUCXPGMCgB7HcFG3IoyQuXXbjtTZGZGGe/Qio03xkmRt5PSgCyJYxFgoc+lMJjEeAuH9KQXBQc/e7VGs2OvU9aAsxNpjGV6nrUrSDjA5pjOqgY5pFIyMcselAWYGM5z0JofcuCwx7etPZSMEjA9PWiQebjafmoABOXwuNwqNo0IPOPenB0PDEqRTSwmOOiigBwRVjzn8KWIg8lce1JmMACTIGeDSptLfuzuOeaB2CP77Fx8vbHWml0bgR4A6H1pUl8t2ULtz+tOJc/wDbnrQFhsa4+YKA2elDRucnv6U5G2FmZdxpqs8iuR+A9KBCrLgbHUfjQ6pgvwuPSogh37CNwPJPpUmAz7cYVefqaAEjZCcBsE9qaqshYE/gO9SyBGGT8r9qGGY92MMvT3oCzI1yDkr9acJCwO37ndfWkDhwM8N6U7GcEJjFAWYg3A5HT0pdpHK456jvTQh3ZH3ac7ITheDQA07QMAcntTdsvdMAjrT9h3ZU7jilBlbhm4x0oAUJlOmRUayLEdhTKnvTjGw43cY6UAEjBUdKAEO2ThRxjr6UuQoAAobau0Ec+nrSxjc4BAJ9KAGfKOWHFOWRR0Xr608L5Z5j+Y9Kbv8oghcnuKAI1eSNuCGU1KgMbcjdnvTXZLkjI8sVLFHIg+8GX3oAY+I2AHU+lNaViQrED0pd7GRuPanGSMLtdeaB2BWBwZDtAPGKY7xls4Yt2p8yDapHAzTWHmAEcEH86AsJEYmY4XGeoNI8qoxGzjsacqu5+7tAPWnPOsjYZenegVmRiJh+8xk9h6U7aemMZpXG1/lOB6UgLK/WgLMcilTtKg9808uEbjqeaiG7nnI9KU5X5iNtAD2k53rgPRhtpLHd/s+tMZ0kACjB9fWgPt+8MH1p6APdXdNu0KDTI0aM7T0pCHclmk4p23ahLNlfWkFmSZwpG35fWo03FyGHbikMx27VoMxDAPwf73pQFmT7flJcbT6VBGpLluhH8NSlvX5h60qug5C5oAhZHDk7uSKuWyM1lNlQWx1JqNHAJyv3hirkDA2FwoQABeTQCR9Q6jJH/wwekJVQ39ojn8K+T4o/KgUP8ANkV9cWz2E/7DEqygGQXwx+VfJqRmW2QBcHPy0DbKuCDhMYqnqSbbR+O4/nV6WN4JhuG1sZqDVsf2e5zkkjP5igRz9FFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUCijtQBspATCrY/hH8qZMuVAAqWL/AFSDPVRSyBQg5wM8+9NbCe50vhmI3OraYF4YOuPzr3v9sq3uoPCng3zl+TyBj8q8H8OMF1rTHPVZAVFfQv7adx9q8E+CWMYQ/Zx39qzvqadD5Plf5unaoQQSTVh0ye3TpULRlQOB+daaECI7JnBx7VImepHWogSD0A/GlWQgDkUaBqWuVX/a9KjBZx83DUiyYYHg/jTxy27ikGoxsKvTBpAM981J5e75+MUjW+MMMc8daA1GjC4wvNOxtOPummhGf5eAKkEO5NuR9c09B2Edd+CfujtSOhK5xgU5Bj5CQR60jjzGwMAChtCIlZjww49akZFQZbpTtpfK46Uxos/MccH1pAM+8CV+6OtKBhuBgntSBdrdME9BT0yrH1oACGGA1OYsFFBDN8p5NMYO3XtQLUkT51PqKYqF/lYZamgsnzDmpt6yDOBmgNSMu0MgBH4U5iRuLLuBHFDyBmwRjikMpC7B1p2GADeXsIwD0pFjPpuNO3FOGpiy7G4GKNB6ilHUZbpQqeYu4du1MklaTOSNtAcqNykcCkLUVQycY460fe6jFPaRplB2Yx+tLsJG4rj2oFqMB2KcnK0oYlcDAFLs7EcHpS7CVyRgjpRdBqRk/MDtwRSsQvfJPanqkgwX5HYUmGU8p1p3QaggViN3A7Co1JMhwuBUoBxjGRSBSR06Uhq4wHaSOhNPBUEK3X1oYBiDjBFLsEjL3oAdnIHO4VFvYfKBtFSomBtHPNOVc/LjvQPUhC4GTz6GlY+Xy3GacIwo4GfmpzoJADjkUA0xqRyLyOA1OMRxnb0705WdAMjilOU6D73amkGpAgjkcDdtIp86DIVTmlaxaXIVfemJGQdpXp3o0DUTYoJ4wfWmFM8/eHapShJORxTmj3LkLsNDsGo1Vzyy7VpZeGDY3IO9NVm3bZl47U9soCFGQaWgg2kIWTnPUelRqBnaeh6n0p+DEpCjOetMERYZUZxQ2h6hIhUYXpSAYXpVhPmGSuCKhn7/AC9aaSDUTzNqcdaMsFJ3fhUewoche3SpBGduQvPpQ7BqIoxz1BoLAcKOvagApyV6igrsYADPHWkGogGD93mgpIQePlHWpsKVwRzSLIBkbeaA1IslucYpFlY8EcetOZmY5K9qUkum0gLjuaA1F2gKCBhe5oG0dtw9abu8wAdPpTgoZR8uCKA1HEKGyoxTdsrEP94/WkCbTn1pxPlt8oyTQGojlmK59aUqfvHgCgIw2gjOT1pZWEp44A4oEIrow+7jNOG0gEfdzzTgV27QoznvUWTknHtigdhAi+aQPu087g2F5pAhwRtxTCx27QOQetAhzPIo59etSIxcDvzUPmNJgFcgGnbsMdowaB6jpcMTlcGjDDqOPWgOGzuGTTsZOTwKBDFOVwDzTvMZOANwNARUyc9aaTtBUUD1F3rIGA6/yqMKVJX7x7UquBkcfWhZR8x64oDUDycAfPTtwX733qjDgNwME0/cMkHr2oC4gBJIY4JoDsB5Z/ClBDAEn5qXa3ORu4poTuMMJOA3NOb+6RipSGI2Y4x1pGgO3kZwOKegakLBtuAccUsS4GSMt2pUOxcOPpTdjBtw+5jpRYCVWSP5fvE08EEjPTsahWLPzFe33aeYShUbuvf0qR2HhyDt6g011VcA8H0pxAyB+tLtAIJAYUBYYycDIyaYeRjGMVOuyNg+aikbzJA6j5aAuIqeXgnnNPUBPmX5jUXmFSNvOetODCE5BznqKA1H/aGkbaRjPc96GQxfPzgdjUZIlAGAqjmpS7uo6ZzQTqIrK7DGN3cU2YhCCy//AF6PKKN82BUvyyjBA29KA1ECK0eVIJzwKVTGBt+62e1IYo0A+bgHgU3y0DF8gn09KCtRWwxO5cqKa25x8p/djtT9gPOR1pQoPyk8ZoDUFy0fy9e4pi/KTkdacqCNywORSF1bLjA9qA1HbinyqMbu/pRJGUGGXC+tNAB75NSNJvXGQ23mgNRrSKY9qDK929KbjC9MY7+tSmJSpYYXPvTA6qCCw45HvQGohOz5iMmkxJkkNkH+GkLbjubr2FSllwGH3vSgNRkjtvVQuAaTYcZxx3p52suMDfQuEU5xxzignUN0QHyHBpo5P3eSMUFSpAUfeoUGNuBkjmgNRrMscgEnDY4oJdWBI4PQ+lSYwGLYfI/KiNkQgffX2oDUjwI2w4xx1pwQRqGU596dIonbBHy44psY2deQOvtQPUF3nlm3ZpFAJ2jr3pwZIQQgDbqiKLgYIB70BqSBRIQrLxng0xVyzDdjFPFwu3y+Mnv6Ug8uRQOAV6+9AtRokUHDDGOnvTJGDsG29O3rUuFkIPFOBj6HGaCtRqs0g20sSGQehBpyFFGQQMn1pPNCA4IFAhEkLMyyfKB0prRZBwOexqQbZF5x19aSXKsMEZ+tA9SNQfXBpz5JDbenb1oLIFIJGfrTQ+TywA+tAaku8BMjrR5zMuGUMKSOVN21iDjnOalaOMuJN4GeMZp6CK+B025zT5trJjGSP4aURhSQCv501lySuV/OkA4bSATwR2qNwXOTT0jDHGQPxpZSgIRCCfXNAEedowwyT0piynJWVeO1WAy+XtO0H60OglAOVyvvQBGyOjAqfkx0qxg/eHXHSm7RgkkHj1pVZUB3YHHrQBLEwIAcDJqaI4sbtQuMLn61UaZOckAY656VYtYzLoV44Kk+x7UDR9Y2unp/wwU8xQbzfcc+1fJtqrNbRsOGXkV9bPE7/sB4RcAX2ST34r5FtZC1pDhcDHOKBEjlpmZn5bFZurACxYYx0/nWjIr/ACkjjsazNYDfZWyO45/GgDDooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACjtRSgUAbSK32dSBxgVHI52Dj5qpC/mVQobAHHSo3uXkGC3H0p3A7XQJf8Aic6WNoKhwSM+9fY/j7wf4F+Onh7QIdQ8XroVxp0IjaM454r4JGq3SPEyyFWj+7tGKvN4v1RjzcZ/4CKmxVz66T9kX4VD73xNTp7VIP2RvhJxu+JgP5V8hL4w1RRxcD/vkU0+LtTJ/wCPj/x0VNmFz67f9kf4SD7nxJB9ORVc/slfC0H/AJKOoH1FfJn/AAl2pj/l4/8AHRTJfFGpTDDXBx9MUWYXPrU/sjfC9iNvxMjXP0qyf2OPhokIkHxSjX0Hy18fr4j1BMEXB49gae/irU3PNy2Kdh3R9f237G3w9miaRPifE23+HK81Qg/ZG+HtxIyN8SouG4ORxXyanifU48hbp1zTT4j1FsH7U+fY07BdH1437GfgIkbfiTDjPqKpSfsj/D1bhbcfEuESk4xla+VV8V6qu3F3JlenNV5NevpZfMa4cv1zmlYLo+tb39j3wNaRBh8TINwP+zWZH+yd4MeQL/wsaEZPXK18vnxFqRHN3IR7mm/23fls/aX/AO+qLDuj6gvv2UvCdoVMPxEhlYnHUVrJ+yB4O8mMy+P4Qz8n5hXySNdv1/5epfwY1IfE2psoU3spA/2jSsxXR9bN+xz4HI+T4hQA5/vCsu8/ZC8LW04RfiFBtY8HK18s/wDCQaiDkXkw/wCBmnf8JBqRUg3spz6saLMLo+rG/Yz8PPEXh+I1uRx3WqR/ZC8PrNsf4h2wGfVa+YE8R6nGMLfTgem803/hINRJyb2Yn13mnYLo+qZf2PfCwHy/Em2Deny1Xi/Y98PtLj/hYtqFzzkrXy62vag+M3k3/fZo/t3UD/y+z/8AfZpWYXR9ct+xT4WSIPJ8SbXB91qvJ+xz4RVTt+JFuxA7Fa+T21/UnGDez4/66HFNGt34BH2yb/vs1VmF0fXVn+xV4XvshfiRah8ZGWWqFx+xz4btL0wSfEO2x67lr5WTXtRRsrfXAPtIaRtd1B33NezsfUyGizC6Pq6T9jnwioyPiPbBcc8rUMf7IHhJmx/wsa2AP+7Xyv8A25qB/wCXyY5/2zS/27qGB/pcox/tGlYV0fVN5+xroNvF5kPxHtmUjplaZB+xxpEsW9/iJbLx0ytfLf8AwkepFdpvZsf75p3/AAkuqYwL6fH/AF0NPUd0fUp/Y40AKC/xItvzWq0n7IXh6OQL/wALHtuR1ytfMI8R6ljm+nP/AAM0067qB/5e5j/wM0ahdH1XH+xx4aYAt8T7VTjOCVrOm/ZM0GOcJ/wsa2aPON2Vr5jOs3563c3/AH2f8ab/AGte/wDP3N/32f8AGizFdH1X/wAMc6E6gx/Em1YHrytWLP8AYm0m7BJ+ItoFH+0tfJw1u/UYF5Pj/roaeniDUkHy39wo/wCup/xoswuj6xX9iXQMkSfEm0X/AIEtU7j9jfw/BKqJ8SLVwTgnK18sHXNQbre3B/7aN/jTRrN+D/x+Tj/tof8AGizHdH1pP+xZ4fjjRo/iRaknkjK1gzfsqaDbXYgf4hW4Un73y4r5t/4SHUwMfbp8enmGmtrl+ww13Mfq5o1C6PrSL9jHwtLbiQfFG1XPUZWs1f2RPDrz+WPiXa7QcA5WvlsavfY/4/Jh/wADP+NJ/bF6OPtc3/fw0WYXR9af8MaeHCuR8ULUke60sH7GHh+VsP8AE+0VfXK18ljWb7/n8nH/AG0P+NKNavx/y+z/APfw/wCNFmK6Pr9P2LfCyYX/AIWtbAn0K0J+w/4fuGOz4oWjfVlr4/8A7Zv85+2T5/66H/Gnrr+pRn5b+4H0lP8AjRYLn1+f2G/Dwyp+J9oG+q02T9h3QEXJ+J9p+a18gtr+pM2TqFyT/wBdW/xpx8Q6mV2m/ucf9dW/xoafcLn1w/7EnhxVJf4pWgx7rT0/Yi8MyjKfFWz/ABK18hHW79hze3B+sh/xpo1i+Xpdzf8Afw/40a9xXPrqX9ivwvC20/FSz3f8Bpo/Yt8M84+Klpt+q18jHVr0n/j8m/7+H/Gj+1b3p9rm/wC/h/xo17j07H15Y/sS+H7248s/FGzRezErV/Uv2EPD1lB5ifFSymGOgZa+NRrN8vS8nH0kP+NOOu6gRj7bcY/66H/GizHdH1tb/sS6HOuR8TbMY9WWpR+w9o3/AEU6z/76WvkMa7qC9L24/wC/rf407+39S/5/7j/v63+NFmK6PrS5/Yk0aBQx+JlkR671ptx+xLoUcatH8TLKTIyRuXivk067qDDm9uD/ANtG/wAaT+2r/H/H7Pj/AK6GiwXXY+s0/Yq8PtGCfiZZ7vTK1C37FuhhwP8AhZNoRjrla+UP7YvgeLyb/v4f8aX+2r//AJ/J/wDv4f8AGiwXR9Xn9jbwyksaS/E61G7vlat3/wCxX4Qto0ZPilaux6jK18hnVr1jk3cx+sh/xobVr09byb/v4f8AGiw+Y+ubP9i7wncgf8XQtVfnHK1BdfsZ+GoCAvxMtSfXK18mrq16v3bub/v4f8aDq96et3Mf+2hosK6PrG3/AGOPC7H978UrSM/8BqVf2MPDErfuvinaN+Kivkj+1Lo/8vMv/fZpV1W8UfLdzD/toaLBddj63X9inQmJB+JtoAO+5aqzfsdeGbcjzPibar/3zXyodav/APn9n/7+H/Gmtql45+a6mP1c0ILo+uLX9jLwfPHub4qWqn/gNVz+x34T8/YvxTtTz1+Wvk7+1LsDH2mbH++aBql2P+Xqb/v4aYrn1xc/sb+EIE3N8V7Utnp8tZNv+yd4RmuTG3xQtVUHGflr5cbU7t+tzMfq5pP7Quf+e8g/4GaAufWUn7IHglCM/Fa2I/4DUkX7H/gWRD/xdW2Bz/s18kHULk9biT/vo0C/uR0uJP8Avs0rDufVGp/smeDbDG34o28mT1+Wrlp+yP4IliSSb4p26A8fw18lnULlutxIf+BGl/tG6Ax9olx/vGiwXPrKb9lD4dRybD8WIMk88LU7fsjfDnZlfi1BkdeFr5EN5OxyZXJ92NAvZx/y2f8A76NMVz60P7I/w/PC/Fe3bPqFoh/Y+8CuWP8AwtS2C+uFr5LF7OOk8g/4EaX7fc/8/En/AH0aVgufVlz+yd8P4J1QfFS3KnqcLV+T9kb4aC33D4tW4lA4GFr5CN3MT/rpD/wI0fapv+ez/wDfRosFz61s/wBkj4fzuwk+K1uq9j8tWdT/AGQPAdlB5kXxZtnOM/w18hC8nUY86T/vo0G/uGGDPJj03Giw7n1jH+yT4Le085vivbgn+E7anh/ZD8FSxgn4s22fT5a+SPt1xjHnyAf7xoF/cgcXEmP940WC59ZN+yJ4NaUL/wALUtnGOvy1YP7HPg4Rlv8AhatttA6fLXyMNRuh/wAvEn/fRpx1W8xj7TLj/fNGoXR9Y2n7H/g+6bn4rWqgd/lq5/wxl4IGN3xctfyWvkEaldr0uZf++jSf2ldHrcyn/gZpiufXQ/Y68Blgv/C3LX64Wi7/AGPvAFrtz8XLY59lr5E/tC5/57yf99mkN7cN1nkP/AjQFz6/sP2OPAF42P8AhbdqPwWrP/DF3gFSFX4u2vPstfHC39yn3Z5R/wADNL/aF1ni4l/77NKwXPsh/wBi74eRYLfF21H4LTG/Yy+HRXKfF61P4LXx02oXTdbiU/8AAzSf2hcj/l4lH/AzRYdz69b9jr4fKDj4uW35LUB/ZI+Ho4Hxat8/Ra+SDfXB6zyH/gRpPtc3/PaT/vo0WC59bt+yT4CyP+LrwMPX5asQfsg/D6br8V7ce/y18gi9nH/LeT/vo0o1C5H/AC8SY9Nxot5iufW2m/sl/D3Up54k+K1vEYs8kLzVaD9lX4dtcNHJ8V4cq2M4XmvlFL64jJKzyKT1wx5pPtU2c+a2fXJosO59Y3f7K3w6tgWX4pw4HstMtP2XfhzeA/8AF1IQc/7NfKRu5yOZmP4mkW5mU5Ejj8TRYLn1hcfsufDW3yrfFeLP0Wpbb9lv4VNGWf4txKc9MLXyUbyb/nqxpDcyk5Ln86YXPrGf9mL4Wq5WL4sxBT/srU8f7L/wm8lmPxdjUj2Wvkbz5MEbjS/aZB/FS1C59cQ/s0fCEuwl+LiFAP7oqK2/Zx+C9zM8bfFxUIOAdor5OW7lXkMOfUCmNMzHPAPsMUwufXlx+y78JI1Jj+LUb8f3VqK3/Zg+FNwf+SsRD8Fr5H81/wC8fzpwuJF6SMPxpWC59c6h+y18LbcAw/FmLfj0WnQfsr/C54d0nxfhRsdCFr5D+0SH+NvzpPPkPV2/OiwXPryP9lH4ZzuAvxkgC4/urVqT9kf4YxR5/wCFy2+cei18dCeQdHYfjSm4lb/lqx9smmFz6/i/ZO+Fjj5/jNbg/wC6tI/7JXwtjBb/AIXPbnj+6tfIHmvn77fmaDPIf42/OgLn1r/wy98MA6r/AMLgifP8QVeKnn/Zb+F9sm5vi9ESR6LXyF5r5++QfqaXzpP75P40rBc+t4P2W/hg67m+LkQXGei1nzfs6/CW2uFjk+KiOCeTgV8sefJ03t+dL9ocDGR+WaLBc+s7z9mr4OLaCa3+LCPJz8u0VR0r9nT4T3fEvxRRH6DgV8siZgQQfypRcODnNFgufVP/AAzn8JFl2v8AFRFGcdBWg37NHwY2qW+LSjP+yK+RzdOepB/Cj7XJkdD+FGorn16v7M/wR8rd/wALaQt9BVSP9mv4OtIR/wALYUenAr5LM759KTz39aYXPrOX9nP4P2rLv+K6sCfQVYP7OfwWZA5+KoJz6CvkUTsB1H5UvnOf/wBVKw7n2An7OXwKC7n+KeTn2qF/2fvgTFIcfE7cM+1fIvnuOOPxApRcP6L+QosPmPq6/wDgJ8Elt5JIPiZ93+EgdasaP+z78DLy1D3PxQKSdMccV8km4c/3fyFIZ2z2/Kiwcx9fzfs+/AePcq/FAkj6VX/4Z6+BrZ/4ultH0FfJBnY88fkKTzm/2fyFFg5j66vf2cvgklk8sPxVDuBwMCo9L/Zs+Dt1a+dN8U0jbp0Ga+ShO4PBH5U1pHJzk0WFc+w4v2aPglIxB+K4U+vHNTD9l74KtkL8WlH5V8a729TSiVx/EfzosK59oR/so/ByVCw+LiADpnFQP+yp8Ht3zfF2LHqQK+OPtEv/AD0b8zTfNf8Avt+dMLn2S37KnwXXr8YIhx6LUL/sw/CEW8sVp8Yo2mZflQhdpNfHnmP/AHjQJXUghjmlYLn3s+paFH+yXr/gyw1aLVrzT7ssrhgC6Y6ivjK1UC2iwNuQRj8axLDXb/Tt/wBnuXjDjDLngj3FM/ti6yf3mM/7IpiOhlLx4BORWfq64sWx0yP51nHWLrGPMH5CoZr+aeMo7ZU0AV6KKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAP/9k="

_LOGO_B64_CACHE = None
def _get_logo_b64():
    """Извлекает base64 логотипа R&J из уже закоммиченного BOOKING_HTML_B64 (первая data:image/png в разметке)."""
    global _LOGO_B64_CACHE
    if _LOGO_B64_CACHE:
        return _LOGO_B64_CACHE
    try:
        import base64 as _b64_mod
        html = _b64_mod.b64decode(BOOKING_HTML_B64).decode("utf-8")
        m = re.search(r"data:image/png;base64,([A-Za-z0-9+/=]+)", html)
        if m:
            _LOGO_B64_CACHE = m.group(1)
            return _LOGO_B64_CACHE
    except Exception as e:
        print(f"LOGO EXTRACT ERROR: {e}", flush=True)
    return ""

@app.route("/assets/logo.png")
def asset_logo_png():
    """Отдаёт логотип R&J как обычную картинку по стабильной ссылке — для email-рассылок
    (base64 data: URI ненадёжен в почтовых клиентах, обычная <img src> работает везде)."""
    import base64 as _b64_mod
    logo_b64 = _get_logo_b64()
    if not logo_b64:
        return "Logo not found", 404
    png_bytes = _b64_mod.b64decode(logo_b64)
    resp = app.response_class(png_bytes, mimetype="image/png")
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp

@app.route("/assets/wallet-strip.png")
def asset_wallet_strip_png():
    """Баннер-полоса для Apple/Google Wallet пасса — фирменная визитка R&J с кличкой
    питомца и счётчиком посещений конкретного абонемента (?id=RJ-00X)."""
    import base64 as _b64_mod
    import io
    from PIL import Image, ImageDraw, ImageFont

    base_bytes = _b64_mod.b64decode(WALLET_STRIP_B64)
    img = Image.open(io.BytesIO(base_bytes)).convert("RGBA")

    mid = request.args.get("id", "").strip()
    if mid:
        memberships = _load_memberships()
        m = memberships.get(mid)
        if m:
            W, H = img.size
            draw = ImageDraw.Draw(img)
            try:
                serif = ImageFont.truetype("assets/fonts/IBMPlexSerif-Bold.ttf", 40)
                sans = ImageFont.truetype("assets/fonts/Jura-Medium.ttf", 22)

                card_left = (W - 1212) // 2
                card_right = card_left + 1212
                inset = 60
                bottom_y = H - 56
                label_color = (205, 199, 181, 150)
                value_color = (238, 233, 220, 255)

                pet_label = "ПИТОМЕЦ"
                pet_value = m.get("pet_name", "") or "—"
                lb = draw.textbbox((0, 0), pet_label, font=sans)
                vb = draw.textbbox((0, 0), pet_value, font=serif)
                label_h = lb[3] - lb[1]
                value_h = vb[3] - vb[1]
                draw.text((card_left + inset, bottom_y - value_h - label_h - 8), pet_label, font=sans, fill=label_color)
                draw.text((card_left + inset, bottom_y - value_h), pet_value, font=serif, fill=value_color)

                used = m.get("used_visits", 0)
                total = m.get("total_visits", 0)
                vis_label = "ПОСЕЩЕНИЯ"
                vis_value = f"{used} из {total}"
                lb2 = draw.textbbox((0, 0), vis_label, font=sans)
                vb2 = draw.textbbox((0, 0), vis_value, font=serif)
                lw2 = lb2[2] - lb2[0]
                vw2 = vb2[2] - vb2[0]
                draw.text((card_right - inset - lw2, bottom_y - value_h - label_h - 8), vis_label, font=sans, fill=label_color)
                draw.text((card_right - inset - vw2, bottom_y - value_h), vis_value, font=serif, fill=value_color)
            except Exception as e:
                print(f"[wallet-strip] render error: {e}", flush=True)

    buf = io.BytesIO()
    img.convert("RGB").save(buf, format="JPEG", quality=90)
    resp = app.response_class(buf.getvalue(), mimetype="image/jpeg")
    resp.headers["Cache-Control"] = "no-cache"
    return resp

WALLETWALLET_API_KEY = os.environ.get("WALLETWALLET_API_KEY", "")

def _wallet_pass_payload(m):
    """Собирает JSON-тело пасса WalletWallet из данных абонемента."""
    mid = m.get("id", "")
    used = m.get("used_visits", 0)
    total = m.get("total_visits", 0)
    return {
        "passStyle": "storeCard",
        "logoText": "R&J Grooming",
        "stripURL": f"https://rjgrooming.up.railway.app/assets/wallet-strip.png?id={mid}",
        "organizationName": "R&J Grooming",
        "description": f"Абонемент {mid}",
        "headerFields": [
            {"label": "ПОСЕЩЕНИЯ", "value": f"{used}/{total}"},
        ],
        "secondaryFields": [
            {"label": "АБОНЕМЕНТ", "value": mid},
            {"label": "ВЛАДЕЛЕЦ", "value": m.get("client_name", "")},
            {"label": "ПИТОМЕЦ", "value": m.get("pet_name", "")},
        ],
        "auxiliaryFields": [
            {"label": "ДЕЙСТВИТЕЛЕН ДО", "value": m.get("expiry_date", "")},
            {"label": "ТИП АБОНЕМЕНТА", "value": m.get("plan_name", "")},
        ],
        "backFields": [
            {"label": "Тип абонемента", "value": m.get("plan_name", "")},
            {"label": "Дата покупки", "value": m.get("purchase_date", "")},
        ],
        "backgroundColor": "rgb(10,10,9)",
        "foregroundColor": "rgb(242,237,226)",
        "labelColor": "rgb(201,160,90)",
    }

def _wallet_create_pass(m):
    """Создаёт Apple/Google Wallet пасс для абонемента. Возвращает (serial, shareUrl, googleSaveUrl) или (None, None, None)."""
    if not WALLETWALLET_API_KEY:
        return None, None, None
    try:
        r = requests.post(
            "https://api.walletwallet.dev/api/passes",
            headers={"Authorization": f"Bearer {WALLETWALLET_API_KEY}", "Content-Type": "application/json"},
            json=_wallet_pass_payload(m),
            timeout=15,
        )
        if r.status_code == 200:
            data = r.json()
            return data.get("serialNumber"), data.get("shareUrl"), data.get("googleSaveUrl")
        else:
            print(f"[wallet] create failed {r.status_code}: {r.text[:200]}", flush=True)
    except Exception as e:
        print(f"[wallet] create error: {e}", flush=True)
    return None, None, None

def _wallet_update_pass(m):
    """Обновляет уже созданный пасс (пуш на устройство) — вызывать при изменении счётчика визитов."""
    serial = m.get("wallet_serial")
    if not (WALLETWALLET_API_KEY and serial):
        return False
    try:
        r = requests.put(
            f"https://api.walletwallet.dev/api/passes/{serial}",
            headers={"Authorization": f"Bearer {WALLETWALLET_API_KEY}", "Content-Type": "application/json"},
            json=_wallet_pass_payload(m),
            timeout=15,
        )
        if r.status_code == 200:
            return True
        print(f"[wallet] update failed {r.status_code}: {r.text[:200]}", flush=True)
    except Exception as e:
        print(f"[wallet] update error: {e}", flush=True)
    return False

def _next_membership_id(memberships):
    nums = []
    for k in memberships.keys():
        try:
            nums.append(int(k.replace("RJ-", "")))
        except Exception:
            pass
    n = (max(nums) + 1) if nums else 1
    return f"RJ-{n:03d}"

@app.route("/api/mark-reminder", methods=["POST"])
def api_mark_reminder():
    body = request.get_json(force=True) or {}
    phone = (body.get("phone") or "").strip()
    date_str = body.get("date") or ""
    stage = body.get("stage")
    done = bool(body.get("done"))
    if not phone or stage not in (1, 2):
        return jsonify({"success": False, "error": "phone и stage (1 или 2) обязательны"}), 400

    status = _load_reminder_status()
    entry = status.get(phone, {})
    if entry.get("last_date") != date_str:
        entry = {"last_date": date_str}
    entry[f"stage{stage}_done"] = done
    status[phone] = entry
    ok = _save_reminder_status(status)
    return jsonify({"success": ok})

# ── ФОТО АНКЕТ ПИТОМЦЕВ (Cloudinary, постоянное хранение) ───────────────
CLOUDINARY_CLOUD_NAME = os.environ.get("CLOUDINARY_CLOUD_NAME", "u35xfusf")
CLOUDINARY_API_KEY = os.environ.get("CLOUDINARY_API_KEY", "555146516498372")
CLOUDINARY_API_SECRET = os.environ.get("CLOUDINARY_API_SECRET", "AQTPeJ2sfE0XhjtlCwz_PcM2ZrQ")

def _normalize_phone(phone):
    """Добавляет +372, если у номера нет кода страны."""
    p = (phone or "").strip()
    if not p:
        return p
    if p.startswith("+"):
        return p
    digits = re.sub(r"[^\d]", "", p)
    if not digits:
        return p
    if digits.startswith("372"):
        return "+" + digits
    return "+372" + digits

def _pet_photo_key(phone, pet):
    raw = f"{(phone or '').strip()}|{(pet or '').strip()}".lower()
    import hashlib
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()

def _pet_photo_public_id(key):
    return f"rjgrooming/pet_{key}"

def _pet_photo_url(key):
    return f"https://res.cloudinary.com/{CLOUDINARY_CLOUD_NAME}/image/upload/f_auto,q_auto/{_pet_photo_public_id(key)}"

@app.route("/api/upload-pet-photo", methods=["POST"])
def api_upload_pet_photo():
    import hashlib as _hashlib
    phone = (request.form.get("phone") or "").strip()
    pet = (request.form.get("pet") or "").strip()
    file = request.files.get("photo")
    if not phone or not pet or not file:
        return jsonify({"success": False, "error": "phone, pet и photo обязательны"}), 400

    key = _pet_photo_key(phone, pet)
    public_id = _pet_photo_public_id(key)
    timestamp = int(_time.time())

    params_to_sign = {"overwrite": "true", "public_id": public_id, "timestamp": timestamp}
    to_sign = "&".join(f"{k}={v}" for k, v in sorted(params_to_sign.items()))
    signature = _hashlib.sha1((to_sign + CLOUDINARY_API_SECRET).encode("utf-8")).hexdigest()

    try:
        resp = requests.post(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/image/upload",
            data={
                "api_key": CLOUDINARY_API_KEY,
                "timestamp": timestamp,
                "public_id": public_id,
                "overwrite": "true",
                "signature": signature,
            },
            files={"file": (file.filename or "photo.jpg", file.stream, file.mimetype)},
            timeout=30
        )
        body = resp.json()
        if resp.status_code != 200 or "secure_url" not in body:
            return jsonify({"success": False, "error": body.get("error", {}).get("message", "Cloudinary upload failed")}), 500
        return jsonify({"success": True, "url": _pet_photo_url(key)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# ── РУЧНЫЕ ДАННЫЕ КЛИЕНТА (email, Instagram) — Cloudinary raw, единый индекс ──
_CLIENT_DATA_INDEX_PUBLIC_ID = "rjgrooming/client_data_index.json"

def _client_data_key(phone):
    import hashlib
    return hashlib.sha1((phone or "").strip().lower().encode("utf-8")).hexdigest()

def _client_data_index_url():
    return f"https://res.cloudinary.com/{CLOUDINARY_CLOUD_NAME}/raw/upload/{_CLIENT_DATA_INDEX_PUBLIC_ID}"

def _load_all_client_data():
    """Один запрос — все сохранённые правки клиентов разом (ключ = хэш телефона)."""
    try:
        r = requests.get(_client_data_index_url(), timeout=8)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return {}

def _save_all_client_data(data):
    import hashlib as _hashlib
    timestamp = int(_time.time())
    params_to_sign = {"overwrite": "true", "public_id": _CLIENT_DATA_INDEX_PUBLIC_ID, "timestamp": timestamp}
    to_sign = "&".join(f"{k}={v}" for k, v in sorted(params_to_sign.items()))
    signature = _hashlib.sha1((to_sign + CLOUDINARY_API_SECRET).encode("utf-8")).hexdigest()
    payload = json.dumps(data, ensure_ascii=False)
    try:
        resp = requests.post(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/raw/upload",
            data={
                "api_key": CLOUDINARY_API_KEY,
                "timestamp": timestamp,
                "public_id": _CLIENT_DATA_INDEX_PUBLIC_ID,
                "overwrite": "true",
                "signature": signature,
            },
            files={"file": ("data.json", payload, "application/json")},
            timeout=20
        )
        return resp.status_code == 200
    except Exception as e:
        print(f"CLIENT DATA SAVE ERROR: {e}", flush=True)
        return False

def _load_client_data(phone):
    """Совместимость: одна запись из общего индекса."""
    key = _client_data_key(phone)
    return _load_all_client_data().get(key, {})

@app.route("/api/save-client-data", methods=["POST"])
def api_save_client_data():
    body = request.get_json(force=True) or {}
    phone = (body.get("phone") or "").strip()
    if not phone:
        return jsonify({"success": False, "error": "phone обязателен"}), 400

    name = (body.get("name") or "").strip()
    phone_override = (body.get("phone_override") or "").strip()
    email = (body.get("email") or "").strip()
    instagram = (body.get("instagram") or "").strip().lstrip("@")
    comment = (body.get("comment") or "").strip()

    key = _client_data_key(phone)
    all_data = _load_all_client_data()
    all_data[key] = {
        "name": name, "phone_override": phone_override,
        "email": email, "instagram": instagram, "comment": comment
    }
    ok = _save_all_client_data(all_data)
    if not ok:
        return jsonify({"success": False, "error": "Cloudinary upload failed"}), 500
    return jsonify({"success": True, "name": name, "phone_override": phone_override, "email": email, "instagram": instagram, "comment": comment})

def _get_reminder_dashboard_rows():
    """Общая функция: последний визит на клиента за всю историю,
    с расчётом стадии напоминания. Используется дашбордом и отчётами."""
    today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
    bookings = _fetch_full_history_bookings()

    by_phone = {}
    for b in bookings:
        phone = (b.get("clientPhone") or "").strip()
        if not phone:
            continue
        try:
            d = datetime.strptime(b.get("date", ""), "%d.%m.%Y").date()
        except Exception:
            continue
        if d > today:
            continue
        if phone not in by_phone or d > by_phone[phone]["date"]:
            by_phone[phone] = {
                "date": d,
                "name": b.get("clientName", ""),
                "pet": b.get("petName", ""),
                "master": b.get("master", ""),
                "breed": b.get("breed", ""),
                "email": b.get("clientEmail", "")
            }

    reminder_status = _load_reminder_status()
    all_client_data = _load_all_client_data()
    rows = []
    for phone, info in by_phone.items():
        days_since = (today - info["date"]).days
        if days_since >= 42:
            stage_status = "done"
        elif days_since >= 35:
            stage_status = "stage1"
        else:
            stage_status = "pending"
        date_str = info["date"].strftime("%d.%m.%Y")
        saved = reminder_status.get(phone, {})
        stage1_done = bool(saved.get("stage1_done")) if saved.get("last_date") == date_str else False
        stage2_done = bool(saved.get("stage2_done")) if saved.get("last_date") == date_str else False
        client_saved = all_client_data.get(_client_data_key(phone), {})
        display_name = client_saved.get("name") or info["name"]
        display_phone = _normalize_phone(client_saved.get("phone_override") or phone)
        display_email = client_saved.get("email") or info["email"]
        rows.append({
            "phone": phone, "display_phone": display_phone, "name": display_name, "pet": info["pet"],
            "breed": info["breed"], "master": info["master"], "email": display_email,
            "date": info["date"], "days_since": days_since, "status": stage_status,
            "stage1_done": stage1_done, "stage2_done": stage2_done
        })
    rows.sort(key=lambda x: -x["days_since"])
    return rows, today

def send_full_40day_report():
    """Полный список клиентов за всю историю (по последнему визиту)
    со статусом по каждому — сколько дней прошло и на какой они стадии."""
    try:
        today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
        from_date = "01.01.2020"
        to_date = today.strftime("%d.%m.%Y")
        r = requests.get(GOOGLE_SCRIPT, params={"action": "stats", "from": from_date, "to": to_date}, timeout=30)
        data = r.json()
        bookings = data.get("bookings", []) if isinstance(data, dict) else []

        by_phone = {}
        for b in bookings:
            phone = (b.get("clientPhone") or "").strip()
            if not phone:
                continue
            try:
                d = datetime.strptime(b.get("date", ""), "%d.%m.%Y").date()
            except Exception:
                continue
            if d > today:
                continue
            if phone not in by_phone or d > by_phone[phone]["date"]:
                by_phone[phone] = {
                    "date": d,
                    "name": b.get("clientName", ""),
                    "pet": b.get("petName", ""),
                    "master": b.get("master", ""),
                    "breed": b.get("breed", "")
                }

        rows = sorted(by_phone.items(), key=lambda kv: kv[1]["date"])
        if not rows:
            tg_result = _send_reminder_telegram("🔔 <b>Отчёт за 40 дней</b>\n\nНет визитов за последние 40 дней.")
            print("FULL REPORT: клиентов не найдено", flush=True)
            return [], tg_result

        lines = [f"🔔 <b>Отчёт по клиентам за 40 дней</b> ({len(rows)} чел.)", ""]
        for phone, info in rows:
            days_since = (today - info["date"]).days
            if days_since >= 42:
                status = "✅ напоминание #2 отправлено"
            elif days_since >= 35:
                status = "1️⃣ напоминание #1 отправлено"
            else:
                status = f"⏳ ещё {35 - days_since} дн. до напоминания"
            lines.append(
                f"👤 {info['name']} ({phone}) — {days_since} дн.\n"
                f"🐾 {info['pet']} — {info['breed']}\n"
                f"Визит: {info['date'].strftime('%d.%m.%Y')} у {info['master']}\n"
                f"{status}\n"
            )
        tg_result = _send_reminder_telegram("\n".join(lines))
        print(f"FULL REPORT: {len(rows)} клиентов отправлено", flush=True)
        return rows, tg_result
    except Exception as e:
        print(f"FULL REPORT ERROR: {e}", flush=True)
        return [], {"ok": False, "error": str(e)}

@app.route("/api/send-full-report")
def api_send_full_report():
    rows, tg_result = send_full_40day_report()
    return jsonify({"success": True, "count": len(rows), "telegram": tg_result})

@app.route("/api/check-reminders")
def api_check_reminders():
    due = check_35day_reminders()
    return jsonify({"success": True, "due_count": len(due), "clients": [{"phone": p, **{k: (v.isoformat() if k == "date" else v) for k, v in i.items()}} for p, i in due]})

@app.route("/api/check-reminders-tomorrow")
def api_check_reminders_tomorrow():
    today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
    tomorrow = today + timedelta(days=1)
    due = check_35day_reminders(target_date=tomorrow, send_telegram=False)
    return jsonify({"success": True, "date": tomorrow.isoformat(), "due_count": len(due), "clients": [{"phone": p, **{k: (v.isoformat() if k == "date" else v) for k, v in i.items()}} for p, i in due]})

@app.route("/book", methods=["POST", "OPTIONS"])
def book():
    if request.method == "OPTIONS":
        resp = app.make_default_options_response()
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        resp.headers["Access-Control-Allow-Methods"] = "POST"
        return resp
    data = request.get_json()
    try:
        print(f"BOOK DATA: {data}", flush=True)
        r = requests.get(GOOGLE_SCRIPT, params=data, timeout=10)
        print(f"GOOGLE SCRIPT RESPONSE: {r.text[:500]}", flush=True)
        resp = jsonify({"success": True})

        # Email уведомление о новой онлайн-записи
        resend_key = os.environ.get("RESEND_API_KEY")
        if resend_key:
            try:
                requests.post(
                    "https://api.resend.com/emails",
                    headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                    json={
                        "from": "booking@rjgrooming.salon",
                        "to": ["myrnj1@gmail.com"],
                        "subject": "Новая онлайн-запись — R&J Grooming",
                        "html": (
                            "<h2>Новая онлайн-запись</h2>"
                            f"<p><b>Клиент:</b> {data.get('name','')}</p>"
                            f"<p><b>Телефон:</b> {data.get('phone','')}</p>"
                            f"<p><b>Email:</b> {data.get('email','')}</p>"
                            f"<p><b>Порода:</b> {data.get('breedDisplay') or data.get('breed','')}</p>"
                            f"<p><b>Услуга:</b> {data.get('service','')}</p>"
                            f"<p><b>Мастер:</b> {data.get('master','')}</p>"
                            f"<p><b>Дата:</b> {data.get('date','')} в {data.get('time','')}</p>"
                            f"<p><b>Стоимость:</b> {data.get('price','')} EUR</p>"
                            f"<p><b>Кличка питомца:</b> {data.get('pet','')}</p>"
                        )
                    },
                    timeout=10
                )
            except Exception as e:
                print(f"BOOK EMAIL ERROR: {e}", flush=True)

        # Telegram уведомление о новой онлайн-записи
        tg_token = os.environ.get("TELEGRAM_BOT_TOKEN")
        tg_chat_id = os.environ.get("TELEGRAM_CHAT_ID")
        if tg_token and tg_chat_id:
            try:
                tg_text = (
                    "🐾 <b>Новая онлайн-запись</b>\n\n"
                    f"<b>Клиент:</b> {data.get('name','')}\n"
                    f"<b>Телефон:</b> {data.get('phone','')}\n"
                    f"<b>Порода:</b> {data.get('breedDisplay') or data.get('breed','')}\n"
                    f"<b>Услуга:</b> {data.get('service','')}\n"
                    f"<b>Мастер:</b> {data.get('master','')}\n"
                    f"<b>Дата:</b> {data.get('date','')} в {data.get('time','')}\n"
                    f"<b>Стоимость:</b> {data.get('price','')} EUR\n"
                    f"<b>Кличка питомца:</b> {data.get('pet','')}"
                )
                requests.post(
                    f"https://api.telegram.org/bot{tg_token}/sendMessage",
                    json={"chat_id": tg_chat_id, "text": tg_text, "parse_mode": "HTML"},
                    timeout=10
                )
            except Exception as e:
                print(f"BOOK TELEGRAM ERROR: {e}", flush=True)
    except Exception as e:
        print(f"BOOK ERROR: {e}", flush=True)
        resp = jsonify({"success": False, "error": str(e)})
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp

# ── BOOKING APP PAGE ───────────────────────────────────────────────────────
BOOKING_HTML_B64 = "77u/PCFET0NUWVBFIGh0bWw+CjxodG1sIGxhbmc9InJ1Ij4KPGhlYWQ+CjxtZXRhIGNoYXJzZXQ9IlVURi04Ij4KPG1ldGEgbmFtZT0idGhlbWUtY29sb3IiIGNvbnRlbnQ9IiMwYTBhMGEiPgo8bWV0YSBuYW1lPSJ2aWV3cG9ydCIgY29udGVudD0id2lkdGg9ZGV2aWNlLXdpZHRoLGluaXRpYWwtc2NhbGU9MSI+Cjx0aXRsZT5SJkogR3Jvb21pbmc8L3RpdGxlPgo8bGluayBocmVmPSJodHRwczovL2ZvbnRzLmdvb2dsZWFwaXMuY29tL2NzczI/ZmFtaWx5PUNvcm1vcmFudCtHYXJhbW9uZDp3Z2h0QDQwMDs2MDAmZmFtaWx5PVBsYXlmYWlyK0Rpc3BsYXk6aXRhbCx3Z2h0QDAsNDAwOzAsNjAwOzAsNzAwOzEsNDAwJmZhbWlseT1Nb250c2VycmF0OndnaHRAMzAwOzQwMDs1MDA7NjAwJmRpc3BsYXk9c3dhcCIgcmVsPSJzdHlsZXNoZWV0Ij4KPHN0eWxlPgoqe2JveC1zaXppbmc6Ym9yZGVyLWJveDttYXJnaW46MDtwYWRkaW5nOjB9Cmh0bWwsYm9keXttaW4taGVpZ2h0OjEwMHZoO2JhY2tncm91bmQ6IzBhMGEwYTtjb2xvcjojZmZmZmZmO2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZjtmb250LXdlaWdodDo0MDB9Ci5zY3JlZW57ZGlzcGxheTpub25lO21pbi1oZWlnaHQ6MTAwdmg7ZmxleC1kaXJlY3Rpb246Y29sdW1uO2FsaWduLWl0ZW1zOmNlbnRlcjtwYWRkaW5nOjQ4cHggMCA2NHB4fQouc2NyZWVuLmFjdGl2ZXtkaXNwbGF5OmZsZXh9Ci5jb257d2lkdGg6MTAwJTttYXgtd2lkdGg6NDAwcHg7cGFkZGluZzowIDI4cHh9Ci5iYWNrLWJ0bntkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo4cHg7Y29sb3I6I2ZmZmZmZjtmb250LXNpemU6MC44cmVtO2xldHRlci1zcGFjaW5nOi4yMmVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtiYWNrZ3JvdW5kOm5vbmU7Ym9yZGVyOm5vbmU7Y3Vyc29yOnBvaW50ZXI7cGFkZGluZzowO21hcmdpbi1ib3R0b206MzZweDtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZjtmb250LXdlaWdodDozMDA7dHJhbnNpdGlvbjpjb2xvciAuMnN9Ci5iYWNrLWJ0bjpob3Zlcntjb2xvcjojZmZmZmZmfQoubG9nby1yantmb250LWZhbWlseTonQ29ybW9yYW50IEdhcmFtb25kJyxzZXJpZjtmb250LXNpemU6Mi41cmVtO2ZvbnQtd2VpZ2h0OjYwMDtjb2xvcjojZmZmZmZmfQoubG9nby1zdWJ7Zm9udC1zaXplOjAuNjYzcmVtO2ZvbnQtd2VpZ2h0OjYwMDtsZXR0ZXItc3BhY2luZzouNGVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtjb2xvcjojZmZmZmZmO21hcmdpbi10b3A6M3B4O3BhZGRpbmctYm90dG9tOjE0cHg7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyMDEsMTY4LDc2LC4yKTttYXJnaW4tYm90dG9tOjIwcHh9Ci5ob21lLXJqe2ZvbnQtZmFtaWx5OidDb3Jtb3JhbnQgR2FyYW1vbmQnLHNlcmlmO2ZvbnQtc2l6ZTozLjI1cmVtO2ZvbnQtd2VpZ2h0OjYwMDtjb2xvcjojZmZmZmZmO2xpbmUtaGVpZ2h0OjF9Ci5sb2dvLXRhZ3tmb250LXNpemU6MC43NXJlbTtsZXR0ZXItc3BhY2luZzouMTJlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Y29sb3I6I2ZmZmZmZjtsaW5lLWhlaWdodDoxLjV9Ci5sb2dvLXJvd3tkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6ZmxleC1lbmQ7Z2FwOjEycHg7bWFyZ2luLWJvdHRvbToyOHB4O3BhZGRpbmctYm90dG9tOjE4cHg7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyMDEsMTY4LDc2LC4yKX0KLmxvZ28taW1nLXJvd3ttYXJnaW4tYm90dG9tOjI4cHg7cGFkZGluZy1ib3R0b206MThweDtib3JkZXItYm90dG9tOjFweCBzb2xpZCByZ2JhKDIwMSwxNjgsNzYsLjIpfQoubG9nby1pbWd7aGVpZ2h0OjkwcHg7d2lkdGg6YXV0bztkaXNwbGF5OmJsb2NrfQouaG9tZS1nc3Vie2ZvbnQtc2l6ZTowLjY2M3JlbTtmb250LXdlaWdodDo2MDA7bGV0dGVyLXNwYWNpbmc6LjRlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Y29sb3I6I2ZmZmZmZjttYXJnaW4tdG9wOjZweDttYXJnaW4tYm90dG9tOjIycHh9Ci5ob21lLWgxe2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZjtmb250LXNpemU6My4xMjVyZW07Zm9udC13ZWlnaHQ6NjAwO2NvbG9yOiNmZmZmZmY7bGluZS1oZWlnaHQ6MS4xO21hcmdpbi1ib3R0b206NnB4fQouaG9tZS1oMSBlbXtmb250LXN0eWxlOml0YWxpYztjb2xvcjojZmZmZmZmfQouaG9tZS1zdWJ7Zm9udC1zaXplOjAuOHJlbTtsZXR0ZXItc3BhY2luZzouMThlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Y29sb3I6I2ZmZmZmZjttYXJnaW4tYm90dG9tOjI4cHg7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWZ9Ci5vcHR7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6MTZweDtwYWRkaW5nOjE2cHggMDtib3JkZXItYm90dG9tOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4wNyk7dGV4dC1kZWNvcmF0aW9uOm5vbmU7Y29sb3I6I2ZmZmZmZjt0cmFuc2l0aW9uOmNvbG9yIC4ycztjdXJzb3I6cG9pbnRlcjtiYWNrZ3JvdW5kOm5vbmU7Ym9yZGVyLXRvcDpub25lO2JvcmRlci1sZWZ0Om5vbmU7Ym9yZGVyLXJpZ2h0Om5vbmU7d2lkdGg6MTAwJTtmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWZ9Ci5vcHQ6aG92ZXJ7Y29sb3I6I2ZmZn0KLm9wdC1pY29ue3dpZHRoOjM4cHg7aGVpZ2h0OjM4cHg7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtqdXN0aWZ5LWNvbnRlbnQ6Y2VudGVyO2ZsZXgtc2hyaW5rOjB9Ci5vcHQtaWNvbi1pbWd7d2lkdGg6MzhweDtoZWlnaHQ6MzhweDtvYmplY3QtZml0OmNvbnRhaW59Ci5vcHQtdGV4dHtmbGV4OjE7dGV4dC1hbGlnbjpsZWZ0fQoub3B0LXRpdGxle2ZvbnQtc2l6ZToxLjUxMnJlbTtmb250LXdlaWdodDo2MDA7Y29sb3I6I2ZmZmZmZjttYXJnaW4tYm90dG9tOjJweDt0cmFuc2l0aW9uOmNvbG9yIC4ycztmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWZ9Ci5vcHQ6aG92ZXIgLm9wdC10aXRsZXtjb2xvcjojZmZmfQoub3B0LWhhbmRsZXtmb250LXNpemU6MC44ODdyZW07Y29sb3I6I2ZmZmZmZjtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZjtmb250LXdlaWdodDozMDB9Ci5vcHQtdGl0bGUtYm9va3tmb250LXNpemU6MS4zOHJlbTt3aGl0ZS1zcGFjZTpub3dyYXB9Ci5vcHQtaGFuZGxlLWJvb2t7Zm9udC1zaXplOjAuNzhyZW19Ci5vcHQtYXJyb3d7Y29sb3I6I2ZmZmZmZjtmb250LXNpemU6MS4yMjVyZW07ZmxleC1zaHJpbms6MDtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZjt0cmFuc2l0aW9uOmNvbG9yIC4yc30KLm9wdDpob3ZlciAub3B0LWFycm93e2NvbG9yOiNmZmZmZmZ9Ci5kaXZpZGVye2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjEycHg7cGFkZGluZzoxMnB4IDB9Ci5kaXZpZGVyOjpiZWZvcmUsLmRpdmlkZXI6OmFmdGVye2NvbnRlbnQ6Jyc7ZmxleDoxO2hlaWdodDoxcHg7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wNil9Ci5kaXZpZGVyIHNwYW57Zm9udC1zaXplOjAuNjg4cmVtO2xldHRlci1zcGFjaW5nOi4yMmVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtjb2xvcjojZmZmZmZmO2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmfQouaG9tZS1mb290e21hcmdpbi10b3A6MzZweDtwYWRkaW5nLXRvcDoyMHB4O2JvcmRlci10b3A6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjA2KTtkaXNwbGF5OmZsZXg7anVzdGlmeS1jb250ZW50OnNwYWNlLWJldHdlZW47YWxpZ24taXRlbXM6Y2VudGVyfQouaG9tZS1mb290IHNwYW57Zm9udC1zaXplOjAuNzc1cmVtO2xldHRlci1zcGFjaW5nOi4yZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2NvbG9yOiNmZmZmZmY7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWZ9Ci5mZG90e3dpZHRoOjJweDtoZWlnaHQ6MnB4O2JhY2tncm91bmQ6cmdiYSgyNTUsMjU1LDI1NSwuMTYpfQoucHJvZ3Jlc3N7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjttYXJnaW4tYm90dG9tOjQwcHg7b3ZlcmZsb3c6aGlkZGVuO2NvdW50ZXItcmVzZXQ6c3RlcH0KLnBze2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjVweDtmb250LXNpemU6MC42NjNyZW07bGV0dGVyLXNwYWNpbmc6LjEyZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2NvbG9yOiNmZmZmZmY7d2hpdGUtc3BhY2U6bm93cmFwO2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmO2NvdW50ZXItaW5jcmVtZW50OnN0ZXB9Ci5wcy5kb25le2NvbG9yOiNmZmZmZmZ9Ci5wcy5hY3RpdmV7Y29sb3I6I2ZmZmZmZn0KLnBkb3R7d2lkdGg6MThweDtoZWlnaHQ6MThweDtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2p1c3RpZnktY29udGVudDpjZW50ZXI7ZmxleC1zaHJpbms6MDtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjEyKTtmb250LXNpemU6MC42NjNyZW07Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWY7Zm9udC13ZWlnaHQ6NjAwfQoucGRvdDo6YmVmb3Jle2NvbnRlbnQ6Y291bnRlcihzdGVwLGRlY2ltYWwtbGVhZGluZy16ZXJvKX0KLnBzLmRvbmUgLnBkb3R7Ym9yZGVyLWNvbG9yOiNmZmZmZmY7Y29sb3I6I2ZmZmZmZn0KLnBzLmFjdGl2ZSAucGRvdHtib3JkZXItY29sb3I6I2ZmZmZmZjtjb2xvcjojZmZmZmZmfQoucGx7ZmxleDoxO2hlaWdodDoxcHg7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wNyk7bWFyZ2luOjAgNXB4O21pbi13aWR0aDo2cHh9Ci5wbC5kb25le2JhY2tncm91bmQ6cmdiYSgyNTUsMjU1LDI1NSwuMTgpfQouc3RlcHtkaXNwbGF5Om5vbmV9LnN0ZXAuc2hvd3tkaXNwbGF5OmJsb2NrO2FuaW1hdGlvbjpmdSAuMzVzIGVhc2UgYm90aH0KLnNsYmx7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmO2ZvbnQtc2l6ZToxLjkzOHJlbTtmb250LXdlaWdodDo2MDA7Y29sb3I6I2ZmZmZmZjttYXJnaW4tYm90dG9tOjIwcHg7bGV0dGVyLXNwYWNpbmc6LjAxZW19Ci5zYm94e2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjEwcHg7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMTYpO3BhZGRpbmc6MCAycHg7dHJhbnNpdGlvbjpib3JkZXItY29sb3IgLjJzfQouc2JveDpmb2N1cy13aXRoaW57Ym9yZGVyLWJvdHRvbS1jb2xvcjojZmZmZmZmfQouc2l7b3BhY2l0eTouMjtmb250LXNpemU6MS4yMjVyZW07ZmxleC1zaHJpbms6MH0KI2JJbnB1dHtmbGV4OjE7YmFja2dyb3VuZDp0cmFuc3BhcmVudDtib3JkZXI6bm9uZTtvdXRsaW5lOm5vbmU7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmO2ZvbnQtc2l6ZToxLjUxMnJlbTtjb2xvcjojZmZmZmZmO3BhZGRpbmc6MTJweCAwfQojYklucHV0OjpwbGFjZWhvbGRlcntjb2xvcjojZmZmZmZmfQouY2xye2JhY2tncm91bmQ6bm9uZTtib3JkZXI6bm9uZTtjb2xvcjojZmZmZmZmO2N1cnNvcjpwb2ludGVyO2ZvbnQtc2l6ZToxLjE1cmVtO2Rpc3BsYXk6bm9uZTtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZn0KLmNsci5zaG93e2Rpc3BsYXk6YmxvY2t9Ci5id3JhcHtwb3NpdGlvbjpyZWxhdGl2ZTttYXJnaW4tYm90dG9tOjIwcHh9Ci5kcm9we3Bvc2l0aW9uOmFic29sdXRlO2xlZnQ6MDtyaWdodDowO2JhY2tncm91bmQ6IzBmMGYwZjtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjEpO2JvcmRlci10b3A6bm9uZTttYXgtaGVpZ2h0OjIwMHB4O292ZXJmbG93LXk6YXV0bzt6LWluZGV4OjUwO2Rpc3BsYXk6bm9uZX0KLmRyb3Aub3BlbntkaXNwbGF5OmJsb2NrfQouZGl0ZW17cGFkZGluZzoxMXB4IDE0cHg7Zm9udC1zaXplOjEuMzYzcmVtO2NvbG9yOiNmZmZmZmY7Y3Vyc29yOnBvaW50ZXI7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDUpO2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZjt0cmFuc2l0aW9uOmNvbG9yIC4yc30KLmRpdGVtOmhvdmVye2NvbG9yOiNmZmZ9Ci5kaXRlbSBtYXJre2JhY2tncm91bmQ6dHJhbnNwYXJlbnQ7Y29sb3I6I2ZmZjtmb250LXdlaWdodDo3MDB9Ci5ub3Jlc3twYWRkaW5nOjE0cHg7Zm9udC1zaXplOjEuMjg4cmVtO2NvbG9yOiNmZmZmZmY7Zm9udC1zdHlsZTppdGFsaWM7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmfQoubm8tc2xvdHMtcGFuZWx7Z3JpZC1jb2x1bW46MS8tMTt0ZXh0LWFsaWduOmNlbnRlcjtwYWRkaW5nOjMycHggMjJweCAyNnB4O2JvcmRlcjoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDgpO2JvcmRlci1yYWRpdXM6MTZweDtiYWNrZ3JvdW5kOnJnYmEoMjU1LDI1NSwyNTUsLjAyKTttYXJnaW4tdG9wOjRweH0KLm5vLXNsb3RzLWljb257Zm9udC1zaXplOjEuOHJlbTtvcGFjaXR5Oi4zMjttYXJnaW4tYm90dG9tOjEycHh9Ci5uby1zbG90cy10aXRsZXtmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWY7Zm9udC1zdHlsZTppdGFsaWM7Zm9udC1zaXplOjEuMnJlbTtjb2xvcjojZmZmZmZmO21hcmdpbi1ib3R0b206MjBweDtsaW5lLWhlaWdodDoxLjM1fQoubm8tc2xvdHMtZGl2aWRlcnt3aWR0aDozNnB4O2hlaWdodDoxcHg7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4xNSk7bWFyZ2luOjAgYXV0byAyMHB4fQoubm8tc2xvdHMtY3Rhe2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246b3BhY2l0eSAuMnN9Ci5uby1zbG90cy1jdGE6aG92ZXJ7b3BhY2l0eTouNzV9Ci5uby1zbG90cy1jdGEtdGl0bGV7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmO2ZvbnQtc2l6ZToxLjA1cmVtO2NvbG9yOiNmZmZmZmY7Zm9udC13ZWlnaHQ6NjAwO21hcmdpbi1ib3R0b206OHB4fQoubm8tc2xvdHMtY3RhLXN1Yntmb250LXNpemU6MC44NXJlbTtjb2xvcjpyZ2JhKDI1NSwyNTUsMjU1LC41NSk7bGluZS1oZWlnaHQ6MS41NTttYXJnaW4tYm90dG9tOjE0cHh9Ci5uby1zbG90cy1jdGEtYXJyb3d7Zm9udC1zaXplOjAuODJyZW07Y29sb3I6I2ZmZmZmZjtsZXR0ZXItc3BhY2luZzouMDhlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWZ9Ci5uby1icmVlZC1iYW5uZXJ7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6MTJweDtwYWRkaW5nOjE0cHggMDtib3JkZXItYm90dG9tOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4wNik7Y3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjpjb2xvciAuMnM7bWFyZ2luLXRvcDo0cHh9Ci5uby1icmVlZC1iYW5uZXI6aG92ZXIgLm5vLWJyZWVkLWJhbm5lci10aXRsZXtjb2xvcjojZmZmZmZmfQoubm8tYnJlZWQtYmFubmVyLWljb257Zm9udC1zaXplOjEuNTc1cmVtO2ZsZXgtc2hyaW5rOjA7b3BhY2l0eTouM30KLm5vLWJyZWVkLWJhbm5lci10ZXh0e2ZsZXg6MX0KLm5vLWJyZWVkLWJhbm5lci10aXRsZXtmb250LXNpemU6MS40MzhyZW07Y29sb3I6I2ZmZmZmZjtmb250LXdlaWdodDo2MDA7bWFyZ2luLWJvdHRvbToycHg7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmO3RyYW5zaXRpb246Y29sb3IgLjJzfQoubm8tYnJlZWQtYmFubmVyLXN1Yntmb250LXNpemU6MC44ODdyZW07Y29sb3I6I2ZmZmZmZjtsaW5lLWhlaWdodDoxLjU7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWZ9Ci5uby1icmVlZC1iYW5uZXItYXJyb3d7Y29sb3I6I2ZmZmZmZjtmb250LXNpemU6MS4yMjVyZW07ZmxleC1zaHJpbms6MDt0cmFuc2l0aW9uOmNvbG9yIC4yc30KLnNiYWRnZXtkaXNwbGF5Om5vbmU7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDoxMnB4O21hcmdpbi1ib3R0b206MjBweH0KLnNiYWRnZS5zaG93e2Rpc3BsYXk6ZmxleH0KLmJuYW1le2JvcmRlci1ib3R0b206MXB4IHNvbGlkICNmZmZmZmY7Y29sb3I6I2ZmZmZmZjtwYWRkaW5nOjJweCAwO2ZvbnQtc2l6ZToxLjQzOHJlbTtmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWZ9Ci5iY2hne2ZvbnQtc2l6ZTowLjhyZW07Y29sb3I6I2ZmZmZmZjtjdXJzb3I6cG9pbnRlcjtsZXR0ZXItc3BhY2luZzouMTJlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWY7dHJhbnNpdGlvbjpjb2xvciAuMnN9Ci5iY2hnOmhvdmVye2NvbG9yOiNmZmZmZmZ9Ci5zdmJ0bntkaXNwbGF5OmJsb2NrO3BhZGRpbmc6MDtiYWNrZ3JvdW5kOnRyYW5zcGFyZW50O2JvcmRlcjpub25lO2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjA4KTtjb2xvcjojZmZmZmZmO2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZjtjdXJzb3I6cG9pbnRlcjt0ZXh0LWFsaWduOmxlZnQ7dHJhbnNpdGlvbjpib3JkZXItY29sb3IgLjJzO3dpZHRoOjEwMCU7b3ZlcmZsb3c6aGlkZGVuO3Bvc2l0aW9uOnJlbGF0aXZlfQouc3ZidG46aG92ZXJ7Ym9yZGVyLWJvdHRvbS1jb2xvcjojZmZmZmZmfQouc3ZidG4uYWN0aXZle2JvcmRlci1ib3R0b20tY29sb3I6I2ZmZmZmZn0KLnN2cHtmb250LXdlaWdodDo2MDA7Y29sb3I6I2ZmZmZmZjtmbGV4LXNocmluazowfQoubWFzdGVyc3tkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOjFmciAxZnI7Z2FwOjFweDtiYWNrZ3JvdW5kOnJnYmEoMjU1LDI1NSwyNTUsLjA3KX0KLm1idG57YmFja2dyb3VuZDojMGEwYTBhO3BhZGRpbmc6MjJweCAxMnB4O3RleHQtYWxpZ246Y2VudGVyO2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YmFja2dyb3VuZCAuMnM7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmO2JvcmRlcjpub25lfQoubWJ0bi5hY3RpdmV7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wNil9Ci5tbmFtZXtmb250LXNpemU6MS4xNXJlbTtmb250LXdlaWdodDo1MDA7Y29sb3I6cmdiYSgyNTUsMjU1LDI1NSwuODUpO2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZjt0cmFuc2l0aW9uOmNvbG9yIC4xNXN9Ci5tYnRuLmFjdGl2ZSAubW5hbWV7Y29sb3I6I2ZmZmZmZjtmb250LXdlaWdodDo2MDB9Ci5tdGl0bGV7Zm9udC1zaXplOjAuOHJlbTtjb2xvcjojZmZmZmZmO21hcmdpbi10b3A6M3B4O2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmfQouZ2J0bntkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO3BhZGRpbmc6MTRweCAwO2JhY2tncm91bmQ6dHJhbnNwYXJlbnQ7Ym9yZGVyOm5vbmU7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDgpO2NvbG9yOiNmZmZmZmY7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmO2ZvbnQtc2l6ZToxLjQzOHJlbTtjdXJzb3I6cG9pbnRlcjt3aWR0aDoxMDAlO3RyYW5zaXRpb246YWxsIC4yc30KLmdidG46aG92ZXJ7Y29sb3I6I2ZmZmZmZn0KLmdidG4uYWN0aXZle2NvbG9yOiNmZmZmZmY7Ym9yZGVyLWJvdHRvbS1jb2xvcjojZmZmZmZmfQouY2FsLWh7ZGlzcGxheTpmbGV4O2p1c3RpZnktY29udGVudDpzcGFjZS1iZXR3ZWVuO2FsaWduLWl0ZW1zOmNlbnRlcjttYXJnaW4tYm90dG9tOjE2cHh9Ci5jYWwtbXtmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWY7Zm9udC1zaXplOjEuOTM4cmVtO2ZvbnQtd2VpZ2h0OjYwMDtjb2xvcjojZmZmZmZmfQouY2FsLW57YmFja2dyb3VuZDpub25lO2JvcmRlcjpub25lO2NvbG9yOiNmZmZmZmY7Y3Vyc29yOnBvaW50ZXI7Zm9udC1zaXplOjEuNTc1cmVtO3BhZGRpbmc6NHB4IDhweDt0cmFuc2l0aW9uOmNvbG9yIC4yc30KLmNhbC1uOmhvdmVye2NvbG9yOiNmZmZmZmZ9Ci5jZ3tkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOnJlcGVhdCg3LDFmcik7Z2FwOjJweDttYXJnaW4tYm90dG9tOjEycHh9Ci5jYWwtd3JhcHtwb3NpdGlvbjpyZWxhdGl2ZX0KLmNhbC1sb2FkaW5ne3Bvc2l0aW9uOmFic29sdXRlO3RvcDowO2xlZnQ6MDtyaWdodDowO2JvdHRvbTowO2Rpc3BsYXk6bm9uZTtmbGV4LWRpcmVjdGlvbjpjb2x1bW47YWxpZ24taXRlbXM6Y2VudGVyO2p1c3RpZnktY29udGVudDpjZW50ZXI7Z2FwOjEycHg7YmFja2dyb3VuZDpyZ2JhKDEwLDEwLDksLjg1KTt6LWluZGV4OjV9Ci5jYWwtbG9hZGluZy5zaG93e2Rpc3BsYXk6ZmxleH0KLmNhbC1zcGlubmVye3dpZHRoOjI4cHg7aGVpZ2h0OjI4cHg7Ym9yZGVyOjJweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4xNSk7Ym9yZGVyLXRvcC1jb2xvcjojZmZmZmZmO2JvcmRlci1yYWRpdXM6NTAlO2FuaW1hdGlvbjpjYWxzcGluIC43cyBsaW5lYXIgaW5maW5pdGV9CkBrZXlmcmFtZXMgY2Fsc3Bpbnt0b3t0cmFuc2Zvcm06cm90YXRlKDM2MGRlZyl9fQouY2FsLWxvYWRpbmctdGV4dHtjb2xvcjpyZ2JhKDI1NSwyNTUsMjU1LC43KTtmb250LXNpemU6MC44NXJlbTtsZXR0ZXItc3BhY2luZzouMDVlbTtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZn0KLmNkbnt0ZXh0LWFsaWduOmNlbnRlcjtmb250LXNpemU6MC42NjNyZW07Y29sb3I6I2ZmZmZmZjtwYWRkaW5nOjRweCAwO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZjtsZXR0ZXItc3BhY2luZzouMWVtfQouY2R7dGV4dC1hbGlnbjpjZW50ZXI7Y3Vyc29yOnBvaW50ZXI7Y29sb3I6I2ZmZmZmZjtib3JkZXI6MXB4IHNvbGlkIHRyYW5zcGFyZW50O3RyYW5zaXRpb246YWxsIC4yc30KLmNkOmhvdmVyOm5vdCguZGlzKTpub3QoLnBhZCkgLmNkLWlubmVye2JhY2tncm91bmQ6cmdiYSgyNTUsMjU1LDI1NSwuMDcpIWltcG9ydGFudDtjb2xvcjojZmZmZmZmIWltcG9ydGFudH0KLmNkLnNlbCAuY2QtaW5uZXJ7YmFja2dyb3VuZDojZmZmZmZmIWltcG9ydGFudDtjb2xvcjojMGEwYTBhIWltcG9ydGFudDtmb250LXdlaWdodDo3MDAhaW1wb3J0YW50O2JvcmRlcjpub25lIWltcG9ydGFudH0KLmNkLnRvZCAuY2QtaW5uZXJ7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4yOCk7Y29sb3I6I2ZmZn0KLmNkLmRpc3tjb2xvcjojZmZmZmZmO2N1cnNvcjpkZWZhdWx0fQoudGd7ZGlzcGxheTpncmlkO2dyaWQtdGVtcGxhdGUtY29sdW1uczpyZXBlYXQoNCwxZnIpO2dhcDoxcHg7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wNyl9Ci50YnRue2JhY2tncm91bmQ6IzBhMGEwYTtib3JkZXI6bm9uZTtwYWRkaW5nOjEzcHggNHB4O3RleHQtYWxpZ246Y2VudGVyO2ZvbnQtc2l6ZToxLjMyNXJlbTtjb2xvcjojZmZmZmZmO2N1cnNvcjpwb2ludGVyO2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZjt0cmFuc2l0aW9uOmFsbCAuMnN9Ci50YnRuOmhvdmVye2NvbG9yOiNmZmZmZmY7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wNCl9Ci50YnRuLmFjdGl2ZXtjb2xvcjojZmZmZmZmfQouc3Vte2JhY2tncm91bmQ6dHJhbnNwYXJlbnQ7Ym9yZGVyOm5vbmU7Ym9yZGVyLXRvcDoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDgpO3BhZGRpbmc6MjBweCAwO21hcmdpbi1ib3R0b206MjBweH0KLnNye2Rpc3BsYXk6ZmxleDtqdXN0aWZ5LWNvbnRlbnQ6c3BhY2UtYmV0d2VlbjtwYWRkaW5nOjhweCAwO2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjA1KTtmb250LXNpemU6MS4zNjNyZW07Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmfQouc3I6bGFzdC1jaGlsZHtib3JkZXItYm90dG9tOm5vbmU7cGFkZGluZy10b3A6MTRweH0KLnNse2NvbG9yOiNmZmZmZmZ9LnN2e2NvbG9yOiNmZmZmZmY7dGV4dC1hbGlnbjpyaWdodH0KLnNwe2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZjtmb250LXNpemU6Mi40MzhyZW07Y29sb3I6I2ZmZmZmZjtmb250LXdlaWdodDo2MDB9Ci5mZ3ttYXJnaW4tYm90dG9tOjIwcHh9Ci5mbHtmb250LXNpemU6MC43MTJyZW07bGV0dGVyLXNwYWNpbmc6LjIyZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2NvbG9yOiNmZmZmZmY7bWFyZ2luLWJvdHRvbTo4cHg7ZGlzcGxheTpibG9jaztmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZn0KLmZpe3dpZHRoOjEwMCU7YmFja2dyb3VuZDp0cmFuc3BhcmVudDtib3JkZXI6bm9uZTtib3JkZXItYm90dG9tOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4xNCk7Y29sb3I6I2ZmZmZmZjtmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWY7Zm9udC1zaXplOjEuNTEycmVtO3BhZGRpbmc6MTBweCAwO291dGxpbmU6bm9uZTt0cmFuc2l0aW9uOmJvcmRlci1jb2xvciAuMnN9Ci5maTpmb2N1c3tib3JkZXItYm90dG9tLWNvbG9yOiNmZmZmZmZ9Ci5jYnRue2Rpc3BsYXk6YmxvY2s7d2lkdGg6MTAwJTtiYWNrZ3JvdW5kOnRyYW5zcGFyZW50O2NvbG9yOiNmZmZmZmY7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWY7Zm9udC1zaXplOjAuODYycmVtO2ZvbnQtd2VpZ2h0OjYwMDtsZXR0ZXItc3BhY2luZzouMjhlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7dGV4dC1hbGlnbjpjZW50ZXI7cGFkZGluZzoxNnB4O2JvcmRlcjoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMjUpO2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YWxsIC4yc30KLmNidG46aG92ZXJ7Ym9yZGVyLWNvbG9yOiNmZmZmZmY7Y29sb3I6I2ZmZmZmZn0KLnNibG9ja3t0ZXh0LWFsaWduOmNlbnRlcjtwYWRkaW5nOjUycHggMjBweDtkaXNwbGF5Om5vbmV9Ci5zYmxvY2suc2hvd3tkaXNwbGF5OmJsb2NrO2FuaW1hdGlvbjpmdSAuNXMgZWFzZSBib3RofQouc2kye2ZvbnQtc2l6ZTozLjZyZW07bWFyZ2luLWJvdHRvbToyMHB4O29wYWNpdHk6LjR9Ci5zdHtmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWY7Zm9udC1zaXplOjIuNzI1cmVtO2NvbG9yOiNmZmZmZmY7bWFyZ2luLWJvdHRvbToxMHB4O2ZvbnQtd2VpZ2h0OjYwMH0KLnNze2ZvbnQtc2l6ZToxLjA3NXJlbTtjb2xvcjojZmZmZmZmO2xpbmUtaGVpZ2h0OjEuOTttYXJnaW4tYm90dG9tOjI4cHg7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWZ9Ci5oYnRue2JhY2tncm91bmQ6dHJhbnNwYXJlbnQ7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4xNik7Y29sb3I6I2ZmZmZmZjtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZjtmb250LXNpemU6MC44NjJyZW07bGV0dGVyLXNwYWNpbmc6LjIyZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO3BhZGRpbmc6MTNweCAyOHB4O2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YWxsIC4yc30KLmhidG46aG92ZXJ7Ym9yZGVyLWNvbG9yOiNmZmZmZmY7Y29sb3I6I2ZmZmZmZn0KLmxvYWRpbmctc2xvdHN7Y29sb3I6I2ZmZmZmZjtmb250LXNpemU6MS4yODhyZW07cGFkZGluZzoxMnB4IDA7dGV4dC1hbGlnbjpjZW50ZXI7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmO2ZvbnQtc3R5bGU6aXRhbGljfQouY2R7ZGlzcGxheTpmbGV4O2p1c3RpZnktY29udGVudDpjZW50ZXI7YWxpZ24taXRlbXM6Y2VudGVyO2hlaWdodDozNnB4IWltcG9ydGFudDtwYWRkaW5nOjAhaW1wb3J0YW50fQouY2QtaW5uZXJ7d2lkdGg6MzJweDtoZWlnaHQ6MzJweDtib3JkZXItcmFkaXVzOjA7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtqdXN0aWZ5LWNvbnRlbnQ6Y2VudGVyO2ZvbnQtc2l6ZToxLjE1cmVtO2N1cnNvcjpwb2ludGVyO2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZn0KLmNkLmF2YWlsIC5jZC1pbm5lcntib3JkZXI6MXB4IHNvbGlkIHJnYmEoOTAsMTgwLDkwLC4zNSk7Y29sb3I6cmdiYSg5MCwxODAsOTAsLjY1KX0KLmNkLmJ1c3kgLmNkLWlubmVye2JvcmRlcjoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDcpO2NvbG9yOnJnYmEoMjU1LDI1NSwyNTUsLjMyKTt0ZXh0LWRlY29yYXRpb246bGluZS10aHJvdWdoO3RleHQtZGVjb3JhdGlvbi1jb2xvcjpyZ2JhKDI1NSwyNTUsMjU1LC4zMik7dGV4dC1kZWNvcmF0aW9uLXRoaWNrbmVzczoxcHh9Ci5jZC5zZWwgLmNkLWlubmVye2JhY2tncm91bmQ6I2ZmZmZmZiFpbXBvcnRhbnQ7Y29sb3I6IzBhMGEwYSFpbXBvcnRhbnQ7Zm9udC13ZWlnaHQ6NzAwIWltcG9ydGFudDtib3JkZXI6bm9uZSFpbXBvcnRhbnQ7dGV4dC1kZWNvcmF0aW9uOm5vbmUhaW1wb3J0YW50fQouY2QudG9kIC5jZC1pbm5lcntib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjI4KTtjb2xvcjojZmZmO2ZvbnQtd2VpZ2h0OjYwMH0KLmNkLmRpcyAuY2QtaW5uZXJ7Y29sb3I6cmdiYSgyNTUsMjU1LDI1NSwuMjIpO2N1cnNvcjpkZWZhdWx0O2JhY2tncm91bmQ6bm9uZTtib3JkZXI6bm9uZTt0ZXh0LWRlY29yYXRpb246bGluZS10aHJvdWdoO3RleHQtZGVjb3JhdGlvbi1jb2xvcjpyZ2JhKDI1NSwyNTUsMjU1LC4yMik7dGV4dC1kZWNvcmF0aW9uLXRoaWNrbmVzczoxcHh9Ci5zdmJ0bi1yb3d7ZGlzcGxheTpmbGV4O2p1c3RpZnktY29udGVudDpzcGFjZS1iZXR3ZWVuO2FsaWduLWl0ZW1zOmJhc2VsaW5lO21hcmdpbi1ib3R0b206NnB4O3BhZGRpbmc6MTZweCAwIDB9Ci5zdmJ0bi1uYW1le2ZvbnQtc2l6ZToxLjUxMnJlbTtjb2xvcjojZmZmZmZmO2ZvbnQtd2VpZ2h0OjYwMDtmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWZ9Ci5zdmJ0bi5hY3RpdmUgLnN2YnRuLW5hbWV7Y29sb3I6I2ZmZmZmZn0KLnN2YnRuLXByaWNle2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZjtmb250LXNpemU6MS43MjVyZW07Y29sb3I6I2ZmZmZmZjtmb250LXdlaWdodDo2MDA7ZmxleC1zaHJpbms6MH0KLnN2YnRuLmFjdGl2ZSAuc3ZidG4tcHJpY2V7Y29sb3I6I2ZmZmZmZn0KLnN2YnRuLWRlc2N7Zm9udC1zaXplOjFyZW07Y29sb3I6I2ZmZmZmZjtsaW5lLWhlaWdodDoxLjc7ZGlzcGxheTpibG9jaztwYWRkaW5nOjAgMCAxNHB4O2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmO3doaXRlLXNwYWNlOnByZS1saW5lfQouc3ZidG4uYWN0aXZlIC5zdmJ0bi1kZXNje2NvbG9yOiNmZmZmZmZ9Ci5zdmJ0bi10YWd7Zm9udC1zaXplOjAuOTc1cmVtO2NvbG9yOiNmZmZmZmY7Zm9udC1zdHlsZTppdGFsaWM7ZGlzcGxheTpibG9jazttYXJnaW4tdG9wOjJweDtwYWRkaW5nOjAgMCAxNHB4O2ZvbnQtZmFtaWx5OidQbGF5ZmFpciBEaXNwbGF5JyxzZXJpZn0KLnN2YnRuLmFjdGl2ZSAuc3ZidG4tdGFne2NvbG9yOiNmZmZmZmZ9CkBtZWRpYShtYXgtd2lkdGg6NDAwcHgpey5zdmJ0bi1uYW1le2ZvbnQtc2l6ZToxLjM2M3JlbX0uc3ZidG4tcHJpY2V7Zm9udC1zaXplOjEuNTEycmVtfS5zdmJ0bi1kZXNje2ZvbnQtc2l6ZTowLjkzOHJlbX0uc3ZidG4tdGFne2ZvbnQtc2l6ZTowLjg4N3JlbX19CkBrZXlmcmFtZXMgZnV7ZnJvbXtvcGFjaXR5OjA7dHJhbnNmb3JtOnRyYW5zbGF0ZVkoMTBweCl9dG97b3BhY2l0eToxO3RyYW5zZm9ybTp0cmFuc2xhdGVZKDApfX0KLmxhbmctYmFye3Bvc2l0aW9uOmZpeGVkO3RvcDoxMnB4O3JpZ2h0OjE0cHg7ei1pbmRleDo5OTk7ZGlzcGxheTpmbGV4O2dhcDo2cHh9Ci5sYW5nLWJ0bntiYWNrZ3JvdW5kOnJnYmEoMTAsMTAsMTAsLjkyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjEpO2NvbG9yOiNmZmZmZmY7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWY7Zm9udC1zaXplOjAuNzc1cmVtO2xldHRlci1zcGFjaW5nOi4xNWVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtwYWRkaW5nOjVweCAxMHB4O2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YWxsIC4yc30KLmxhbmctYnRuOmhvdmVye2JvcmRlci1jb2xvcjojZmZmZmZmO2NvbG9yOiNmZmZmZmZ9Ci5sYW5nLWJ0bi5hY3RpdmV7Ym9yZGVyLWNvbG9yOiNmZmZmZmY7Y29sb3I6I2ZmZmZmZn0KLmNiay1idG57YmFja2dyb3VuZDp0cmFuc3BhcmVudDtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjE0KTtjb2xvcjojZmZmZmZmO2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmO2ZvbnQtc2l6ZTowLjg2MnJlbTtsZXR0ZXItc3BhY2luZzouMTZlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7cGFkZGluZzoxMnB4IDIwcHg7Y3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjphbGwgLjJzO3dpZHRoOjEwMCV9Ci5jYmstYnRuOmhvdmVye2JvcmRlci1jb2xvcjojZmZmZmZmO2NvbG9yOiNmZmZmZmZ9Ci5tYnRuLC5zdmJ0biwuZ2J0biwudGJ0biwuY2J0biwuaGJ0biwuY2JrLWJ0biwubGFuZy1idG4sLmJhY2stYnRuLC5vcHQsLmRpdGVtLC5jZCwubm8tYnJlZWQtYmFubmVyLC5iY2hne3RyYW5zaXRpb246YWxsIC4xNXMgZWFzZX0KLm1idG46YWN0aXZlLC5zdmJ0bjphY3RpdmUsLmdidG46YWN0aXZlLC50YnRuOmFjdGl2ZSwuY2J0bjphY3RpdmUsLmhidG46YWN0aXZlLC5jYmstYnRuOmFjdGl2ZSwubGFuZy1idG46YWN0aXZlLC5iYWNrLWJ0bjphY3RpdmUsLm9wdDphY3RpdmUsLmRpdGVtOmFjdGl2ZSwuY2Q6YWN0aXZlLC5uby1icmVlZC1iYW5uZXI6YWN0aXZlLC5iY2hnOmFjdGl2ZXt0cmFuc2Zvcm06c2NhbGUoMC45Nil9Cjwvc3R5bGU+CjwvaGVhZD4KPGJvZHk+CjxhIGhyZWY9Ii9hZG1pbj9wYXNzPWFuemExOTg1IiBpZD0iYWRtaW5CYWNrTGluayIgc3R5bGU9ImRpc3BsYXk6bm9uZTtwb3NpdGlvbjpmaXhlZDt0b3A6MTRweDtyaWdodDoxNHB4O2ZvbnQtc2l6ZTowLjlyZW07Y29sb3I6I2M5YTA1YTt0ZXh0LWRlY29yYXRpb246bm9uZTt6LWluZGV4Ojk5OTtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZjtiYWNrZ3JvdW5kOnJnYmEoMTAsMTAsOSwuODUpO3BhZGRpbmc6NnB4IDEycHg7Ym9yZGVyLXJhZGl1czoyMHB4O2JvcmRlcjoxcHggc29saWQgcmdiYSgyMDEsMTYwLDkwLC4zNSkiPuKGkCDQkNC00LzQuNC9LdC/0LDQvdC10LvRjDwvYT4KPHNjcmlwdD5pZihsb2NhdGlvbi5zZWFyY2guaW5kZXhPZigncGFzcz1hbnphMTk4NScpIT09LTEpe2RvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdhZG1pbkJhY2tMaW5rJykuc3R5bGUuZGlzcGxheT0nYmxvY2snO308L3NjcmlwdD4KPGRpdiBjbGFzcz0ibGFuZy1iYXIiPgogIDxidXR0b24gY2xhc3M9ImxhbmctYnRuIiBvbmNsaWNrPSJzZXRMYW5nKCdldCcpIj5FVDwvYnV0dG9uPgogIDxidXR0b24gY2xhc3M9ImxhbmctYnRuIiBvbmNsaWNrPSJzZXRMYW5nKCdlbicpIj5FTjwvYnV0dG9uPgogIDxidXR0b24gY2xhc3M9ImxhbmctYnRuIGFjdGl2ZSIgb25jbGljaz0ic2V0TGFuZygncnUnKSI+UlU8L2J1dHRvbj4KPC9kaXY+Cgo8IS0tIEhPTUUgLS0+CjxkaXYgY2xhc3M9InNjcmVlbiBhY3RpdmUiIGlkPSJob21lU2NyZWVuIj4KPGRpdiBjbGFzcz0iY29uIj4KICA8ZGl2IGNsYXNzPSJsb2dvLWltZy1yb3ciPgogICAgPGltZyBzcmM9ImRhdGE6aW1hZ2UvcG5nO2Jhc2U2NCxpVkJPUncwS0dnb0FBQUFOU1VoRVVnQUFBVU1BQUFEckNBWUFBQUR6Qy9Rd0FBQUJXR2xEUTFCSlEwTWdVSEp2Wm1sc1pRQUFlSng5a0xGTHcxQVF4cjlXcGFCMUVCMGNIREtKUTVTU0NybzR0QlZFY1FoVndlcVV2cWFwa01aSGtpSUZOLytCZ3YrQkNzNXVGb2M2T2pnSW9wUG81dVNrNEtMbGVTK0pwQ0o2aitOK2ZPKzc0emdnT1c1d2J2Y0RxRHUrVzF6S0s1dWxMU1gxakFTOUlBem04Wnl1cjByK3JqL2ovVDcwM2s3TFdiLy8vNDNCaXVreHFwK1VHY1pkSDBpb3hQcWV6eVh2RTQrNXRCUnhTN0lWOG9ua2Nzam5nV2U5V0NDK0psWll6YWdRdnhDcjVSN2Q2dUc2M1dEUkRuTDd0T2xzck1rNWxCTll4QTQ4Y05ndzBJUUNIZGsvL0xPQnY0QmRjamZoVXArRkduenF5WkVpSjVqRXkzREFNQU9WV0VPR1VwTjNqdTUzRjkxUGpiV0RKMkNoSTRTNGlMV1ZEbkEyUnlkcng5clVQREF5QkZ5MXVlRWFnZFJIbWF4V2dkZFRZTGdFak41UXo3Wlh6V3JoOXVrOE1QQW94TnNra0RvRXVpMGhQbzZFNkI1VDh3Tnc2WHdCQTZkaUU4SFlXaE1BQUVId1NVUkJWSGljN1oxNWZGVkZzdmpyM0RYN0JvUWxRQWliS0FJK1VGQnhYMUFad0hGNUlpQlBIUmNlRGk3b3FQaFRSbEZBUWNWUlVaOFBVWEhVSitMbzRLNkFBczY0b09ER0loRENrb1JBOXZWdVo2bmZIMWhObjc3bkpqY1FJSUg2Zmo3NTNDWG5kdmZwYzdwT1ZWZDFOUURETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUV6YlFEdlNEVGhhY0xsY2dJaWdhUnBZbGdWdXR4dE0wd1JOTzdndVJzU29lZ0FBTE1zQ0FCRGwwM0dhcG9rMk1DMlB4K01Cd3pBTytycktJQ0s0WEM1eHplVDNETk5tSU9Fa28ya2FlRHllRmluZjQvR0EyKzJPV1U4c1dySU54em91bDB2MHY5dnRCcmZiM2FMbGE1b0dicmNiZkQ2ZitLNmw2MkNhaGpYREZzRG44MEVrRXJGOVJ4cWFxdGsxRi9uM1RscWcvSC9XS0E0ZGJyY2JFRkg4a1NYUUVxalgwK3YxZ3E3ckxWSTJ3eHhXVkEzTVNZczdVSHcrbjAwRGJFb2JkTGxjTFdyQ0hldDRQQjdSbjNLL3RxVFdyVjR2K3N6YUlkT21rTTBud3V2MUNzSFVrcWpDamw1bE0wNyt6RUt4WlpDdnJjdmxhbkVoSlpkSHBqSlBjVEJ0RWxVUXRqUWtCRlhoR3EvR3lCdzRkRDFwMmdOZy8vVnVpWWRkckxMNGVqSnREdldtOWZ2OUFHQTNydzYyZkxVY1dldVRKOTNwZjdIYXhod1lxaVlZeTZGMUlMaGNMcUVGVWoyeUFHWU9IOXpiTFlEUDU0T0hIbm9JTzNmdURKRklwRVhOS0RsY3h6QU0wSFVkNnVycW9LU2tCUGJ1M1F1Yk5tM1NTa3BLb0xLeUVnRHNUcFNXY09BYzY1eCsrdWs0YWRJa0FBQ0lSQ0xnOVhvaEhBNkQyKzBHeTdJT1dpZ2FoZ0VWRlJYdzhNTVBhK0Z3V0h6UHpqQ21UWktZbUFqcjE2L0hjRGlNaG1FZ0lxS3U2MmhaRnBxbWlZZ29YdW4vaEs3cmlJaG9XUlphbG1YN0gzMm1WL20zOHZ0SUpJS0dZZURhdFd2eGpqdnV3THk4UEVoTlRZM1NFdWt6dlI0T3pjT3BMbnBZMEt1VDlpc2pQMXpvMkVNeEordkV1SEhqUkQrcjExS0Zyb2w2amVWcktGOXorbjc5K3ZXWWtwSWlOTVREY1Y1TU5Ld1p0Z0ErbncrbVRwMkthV2xwb0drYUpDVWxRYnQyN1dEdzRNRnd3Z2tuZ0dWWnRtQnNYZGZCNi9VS3JZOCtrNlpCbW9mSDR4SGZtYVlwaEVJa0VnR2Z6MmZUVE9UM2htSEErdlhyNGFXWFhvSVZLMVpvQlFVRkl2UkhEcytod1BCRERZV0tPR21xOUoyc0NkRjVhNW9tMnFlR0ZjbS9QWlNjY2NZWk9ISGlSUEI0UE9EeitTQTNOeGNHRHg0TUNRa0o0aGpxUzFXZ3F5RXpkSzIvL2ZaYktDZ29BRjNYSVNFaEFZcUtpdURoaHgvV2dzSGdZVDAzeGc0THc0T0VCQXFaVFhRRGV6d2VPTzY0NCtDaWl5N0NKNTk4MGliTUFFQUlRSHF0cmEyRjFOUlUyTGx6SjJ6Y3VCSFdyVnNIUlVWRlVGMWRMY3BNU2txQ3JsMjd3cm5ubmd0bm5ubW16ZU9vYVpwTmlBYURRZkQ3L2ZEenp6L0QwcVZMWWZiczJacXU2N2I0dU1NeDJOUkJUU3M0NUw2ajcwM1RqR3MxQnMyWkhnNUJUbTBEMlBlUXljdkxnK09QUHg1bnpKZ0JKNTk4c2ppR2hKMFRkTzMzN05rRE45eHdBMnpac2tVckxDeUVTQ1FDaUdpTFU2VitvUE5tZ2NpMEtXS1pOYlI2NVBiYmIwZlROREVjRHR0TUxOa01ycXlzeEI0OWVvRFA1NFBFeEVRQTJDZElhREpkRHJsd3U5MlFrWkVCTTJiTVFFVEVRQ0JnSzlNMFRadlpIUWdFY04yNmRhaWF6b2Nqam8wRWlXeldxbUVqOG1jbkI1QzYra00xOXc4bGNoMnlvOFB2OThPTEw3NklvVkRJWmc3TDB4NnlTVjFXVm9iZHVuV0RwS1FrVWE3c0NDUEhHOE8wYVdpUXlzSkZuZzg3NFlRVFlOMjZkYlpCSXd0RjB6U3hyS3dNMjdWckozNURaWGk5WHR1OEd3MGdlczNPem9hMzMzNGJxNnVyMGJJc0RJZkRVZVZibG9XR1llRG5uMytPZmZ2MkJZQWpNL2pjYmpla3A2ZkRxYWVlaXVQR2pjUFhYMzhkdDI3ZGluVjFkVUo0UkNJUkxDc3J3dzgvL0JDdnZ2cHFIRHg0TUFMWVBlaXFrRHhjYlZmbk92djE2d2ZmZmZlZHJiOXB2cEFlU0lpSTRYQVlyNzMyMnFqekFJQ29lVUphbWtmdkdhYk40QlFmUmplMEhQN3kwVWNmb1dFWXFPdTZHRGcwV0hSZHgrcnFhdXpjdWJNb1EzNXRyQzZQeHdOcGFXa3djZUpFcksrdmp4SzRobUhZQnVxeVpjc3dLeXZyc0EwME9UN3l6My8rTTc3eHhoczI0VWRFSXBFb1I1RmxXYmhqeHc1OC9QSEhNVGMzMXlZa1dpcDBxU25rNjZnNmdkeHVOenoxMUZNWURvZWpuRjB5di83Nkt4NTMzSEUyamQvcFZkV1llUVVLMHlaeDhvaktOL05ycjcwV3BRMlNTV1dhSmxaV1ZtS0hEaDFFV1RLMG9nWEFXVU1DMkNkd1R6dnROQXlIdzFHbW1teStSU0lSWExkdTNXR2JpSEs1WE5DcFV5ZFl2SGd4VmxkWEMwMkoyb0s0ejRUOCt1dXY4ZXV2dnhiSDBQK29uOWF1WFlzbm4zd3lIb21WTmVvS0ZKbHJycm5HWmlxcm5tYkxzbkRGaWhWSWdoREFua1JEZmxqd01qem1xRUVObWdYWVAzam16WnRuRTBwcW1FWkZSUVcyYTlmT2NhVURRTFN3ZFhydjlYcGgvUGp4Mk5EUVlCdVVWQ2RwWE1GZ0VLZFBuNDZIdzh6TXpzNkd6ejc3VEFnM0VoQ2hVQWkvLy81N1BQUE1NMUVPUFBiNy9mRElJNDlnVFUyTnplUkVSQ3dvS01DY25CeFI5dUVVR3JLSkxEK1loZzRkaXJXMXRlTGM1TEFhYXZmU3BVdFJuak5WelcyQTZDeERiQ0l6UngzMDVKODJiWnBOR0toVVZGUUl6WkJvem9DZ2daYWFtZ3BMbHk2MXpWblJ3SlEvLy9EREQyTCs4R0NGaXF5NUVtNjNHeElTRXVDZi8vd25tcVpwRTRhSWlNODk5eHdtSmliR1hNbzRhZEtrS0xNL0VvbmdxbFdyc0RVNUcvTHk4cUM2dXRvbUJOWHJ1M2p4WWdTd081RlkrMk9PT1VpanUvZmVlMjJEUlIwd0J5b015YnNNQUNLczV1U1RUeFphaW1xMmtVblgwTkNBRXlaTXNHa3NWQ2Vab3MzVlRtU2hxR2thVEo4K1BhcCswelJ4eVpJbG1KMmRiZnN0ZWN2cFhESXlNbURSb2tXMk5wUEF1Zm5tbTF1TlFPemV2VHRVVlZYRkRNUkdSRnkwYUpFUWhrZkMrY1BFQjErUlF3eitudi91VU1XTVVkWmxpdGt6VFJQV3JsMnJGUllXMnNKUUtGYlA1L01CSWtKU1VoS01HemRPQkh2TFFjMTBiSFBicSt1NmlMZExUMCtITVdQR2lQb2prUWhZbGdXV1pjRVhYM3dCcGFXbHRyQWJpck9qN09EVjFkV3daczBhQ0lmRDRQZjdiY3NjeDR3WjAycnkvY25uckVMWHZxR2h3ZmFkL01xMEhsZ1lIZ1ZRRURQQS9nSDR3QU1QQ0FFanI0Q2hJR2pETU9DY2M4NFJHcG5xbUdqT1lGVmpDTDFlTHd3Yk5neTdkKzh1eXZINWZPQnl1V0RQbmozdzAwOC9nYVpwUXBDclNTY1FFUklTRXVDSEgzNkFQWHYyaUhMcDNEcDI3QWpaMmRtdFlsN05NQXpidzhRSmVXVUp3Y0t3OWNIQzhBaHdJQ1pvVTJYUktnZVh5d1UrbncrV0xGbWl5Y0tGNnFObGV4NlBCNUtUazRGaUcxWE50VGxtbkZ3K0lrSWtFb0grL2ZzRE9ZUklXQ01pMU5YVlFYVjF0WVpTdG1oYXBpaG5DdytGUXJCMTYxYXRxcXJLVm9lbWFaQ2FtaXJhZmFScFNxalJFcnpXSUxpWnhtRmgyTVlob1FLd3o4U2s3RGJCWUJBYUdocUVvQUhZci8zUkFMWXNDM0p5Y2xCTllYOGc4MWxrM3RJU09aL1BKMHgzZVJPbHpNeE02Tnk1TTFMOVRyR1VWRDlsY2FGekNJVkNBQURRME5BZ2xySWRhWnpXVEt2L2s5K3pVR3k5c0RCczQ2akpGdVRVVXNGZ1VBdyswczVJR0pJUVRVdExzNVZINXZUQnBJOVNOemFTTmNmMjdkdkRzR0hEeFA5bFFVN0pLK1NFRFhRTW1jNldaVUZEUXdPVWxKUzBDbytzVS9DM0toaGxnY2xDc2ZYQ3d2QXdvUTZRbHRKcWFGRS9RTFJnRElWQ1F2akpjMjV5OGdOZDEyMUpFV1RpRVRieTRLYkJIb2xFUU5mMUtDRk43UmcyYkJoUXZLRGNmc013UkFZWUFJQ2VQWHNpQ1dzUzBvZ0lPM2Z1aElhR2hzT1dxS0V4cUsyeHJpLzFDWDFtWWRoNllXRjRpSWwxODdma25HRXM3eStsbVNLdFVON0htWTR0TEN6VVNBTWpZU05yYTAwaEQzWjUwTmZWMVVFZ0VCQnRwSGt6eTdMZzRvc3Zodjc5K3lNSkVsa1RsUHZsekRQUEJGcWlLQXZ4Ung1NVJDUEJLdmREWXl1QTVHUGtQenBPWFNQc2xKTExpYWEyY1ZYYlJkZWdLYWNMYy9oaFlkakdJZk1SWUw5emhBWlpXbG9hR0lZaGdwa3BQWlJsV2VEeGVDQWNEa05GUllYTmhIWjZiUXpTK0tndEpJQysrZVliMkxObkR4aUdJVXhnZVQ3eG80OCtndFRVVkFEWUgyeE5tcUZwbXRDaFF3YzQ1NXh6SURFeFVXaXZpQWpUcDArSGdvSUNjYnc4eDBqdElNR21DblBTak9VL09vNk9wZmxKU2ljV3ovbXpVRHM2WUdIWXh0RTBUVGdXL0g2LzBNNHlNelBCNC9HQXJQWFI4UzZYQ3d6RGdHM2J0a0Z0YmEzNFh2WTR4d3Q1c1ZVQlZGQlFvTzNldlRzcUtGeDI0bnoyMldlWWw1Y0hjcDVGRWk1VHBrekJVYU5HaVgyRUE0RUFQUC84OC9ERUUwOW9zb0NYNXpmbCtWRjVpVitzZnFQNFRBQzdJMG8xYVJ2alVHd0F4aHdaV0JpMmNVaFlBT3p6dnRJQUhqdDJiTlRhWTFub3VGd3VlT3V0dDRTNUpzOGpOa2NZMFBFa2dBRDJtWm9ORFEyd1lNRUM0ZldWdFRUTHNzRHI5Y0pKSjUwRVU2ZE94YXlzTE51ODVZSUZDL0N2Zi8ycjBEcHJhbXBneG93WjhOQkREMmxVRGdWZGsya3RhM3B5bXhwck03VkxGb2prZ0pKTi9zYncrLzE0cUtaQUdPYW9nZ2JHM1hmZjdiZ21tVGpRNVhoMGpKb2d0YkN3VUN6SGs5Y2wweHJoNnVwcXZQamlpNk9XaWNsbHh1dEFrWmVZeVdhNjIrMkc4dkp5a2FVR01YcC9FTXV5OEo1NzdrR1h5d1VEQmd6QTlldlhvNjdyWWdsZVFVRUJwcWFtUnMwbk9xMkhsdHZzbEJBMlZzWWJ0YzFxZnpiR0thZWNnclcxdFk3WmFvakhIMzhjbmZxWGhXYnJnalhETmc1cEw2UUZlVHdlT1BmY2M3RnIxNjQySVVDYWtOZnJoVWdrQXUrLy96NTg5ZFZYR3BYaHBBM0c2NjJsY3Nsa0plZUdhWnB3NG9rbmFpVWxKV0lWaGl4Z0xjc0MwelJoenB3NXNHclZLdnpsbDErZ1g3OStVRjlmRDZ0WHI0WS8vZWxQMEtkUEh5MFlESXJ6SS9PWGxzR1JWcXhxd2ZMZU1UUlBHbXRKSkgwdngwbkdpOC9uTytDVk8wenJJcjdISDlOcUlmT1d6THFzckN5NDdiYmJBQUJzMzhzZTRsQW9CRTg5OVpSdHpTekFnVzFDUk1lcjI1TlNrUFhldlh0aHhvd1o4TlJUVDRua3BpU1laTzNyakRQT2dKS1NFbmp2dmZkZzJiSmxzSHo1Y3EyK3ZsNzhuMHhtVmFEcHVnNlptWm5RcDA4ZjFIVmRyTDJtK1VEU0tPbFBUa0toYVpxWWovenl5eTgxK2Z6akZZcWtvVHJGR3JMbTE3WmdZZGpHa2IyZlBwOFBSbzRjaVNOSGpoUWJUUUhzRDc4aHdYbjU1WmZEanovK3FLbEpIR1JoSUF1ZmVDRGhJVzlrUkVMNGpUZmUwTTQvLzN3Y04yNmMwRklwN2xFVzVqNmZEejc1NUJQNDVKTlBOSnJMbzNMcGxiNmpjalJOZ3drVEp1RDk5OTh2TmxhU25UYXk0SXdsbkRaczJBQVhYSENCbUhPbGN1UDFKak5IQjJ3bXR3TEkrYUR1b0NmUHhSRWVqOGVXR1prRVhVSkNBbHh3d1FXNGNPSENxUDFZS0V5a3Vyb2FKazJhQkY5ODhZVkdqZ0paMk1udnliUnNDcldOY2tnS0NUclROT0hCQngvVWdzR2c3UnpWZWJSMjdkckI3Tm16b1ZldlhxS3NXUE5ySkhBdHk0THE2bXJZdm4wNzdOMjdGK3JyNjhIdjkwTm1aaVprWm1aQ1ZsWVdaR1ZsUVVaR2h2Z3VQVDBkYUZ2WDJ0cGFxSyt2ajhvbUUrOURnTUthNUdrR0V2SnNNak9NUkZNT0ZISW8xTlRVWUZaV0ZnRHNFeUtxSTBBdWk5N1RYOCtlUFdIMjdObFJFL2VJKy9kQjJidDNMMDZhTkFuVDA5TmIzSHhUdzFUSS9BUUFTRTlQaDJuVHB1SFBQLzhzSERpMGc1OTgvcktENTdQUFBzUE16RXl4ck05cFh0QXBjQm9Bb0cvZnZ2RGdndzlpVFUyTktKZWNTSWo3c253SEFnR2NOV3NXL3RkLy9SY09HalFJZS9ic0tjcHRMTVcvRTJQR2pJbktMSzd1aDhJT0ZJYUIrTHpKaG1GZ1pXVWx0bS9mUHVhMm1hcTJSL05mVTZaTXdmejhmQXlGUW1oWmxranhUNElRRWJHNHVCaUhEUnVHNmtCdkNSTXYxcW9QVGROZzdOaXgrTU1QUDJBb0ZFSmQxOUd5TE55MmJSdFdWRlJFN2ROaUdJWXQ2ZTBISDN5QUNRa0pqbHQxeHRwcWxJU2ozKytINTU5L1BpcTdOeUppZVhrNW5uTEtLYWhwbWtnM0pwZEI3WmZYVnpmR0ZWZGNnWUZBd0ZZUEMwT0djYUFwWVVqZlZWUlVZS2RPbmNTa1AzbE5FeElTaENCTVNrcUN0TFEwNk5ldkg4eVpNMGVFckZCNmZIWGZrK3JxYW56c3NjZHNXYUZKQ0xRa2N1QXg3ZWs4ZS9aczJ4N09obUZnZm40K0FnRGNmLy85V0Y5Zkw4Sjg1TkFiRXBLQlFBQ2ZmZmJacUxZVHF0YW1mazVOVFFXNVQzUmR4OXJhV3J6a2trdFExbUxsMzhwbHhKdTVaOXk0Y1ZIQ1VMMjJMQXpiQnV4QU9jS1FBOFRuODhFOTk5eUQ1ZVhsWWxrYUpVUnQzNzQ5ZE96WUVYSnpjNkZ2Mzc2UWxaVmxXMUtXbkp3TUFQc0djQ0FRZ0xWcjE4TDI3ZHRoM3J4NThQUFBQMnNBK3dRV2hhTVFCK0k5ZG9LQ3VRRUFoZzBiaHZmY2N3K01IajFhcE8reUxBdFdybHdKSTBlTzFOeHVOOHlhTlV2THpzN0cyMjY3RFdRUE1QNGU5SXkvTHpHY1BIa3k3TjI3RjJmUG5xMVIyQXM1U2VRNVBUbGduRDdMQ1ZXcGp6Lzc3RFA0N3J2dk5KVG1OS2tmQU96TCtRekRFSytOMGRnS0ZCWjJEQ01SYjlDMXVvTWRhUmJxZmlueVo5bkVMQzB0eFZ0dXVRVkhqeDZOdWJtNVVmV1RXWDBvdkora2FWNTY2YVc0WmN1V3FCM2lWcXhZZ2JtNXVVTDQwTExCUng1NXhHYldxNXRHV1phRmdVQUFIMzMwVVpUcmNRcVNwbmhLaXEzTXlNZ0F1YXlxcWlyOHd4LytFQ1g1blRMMXhHc2lBd0RjZU9PTkdBd0dHOTAzbVRWRGhvSDQ1d3dSOTIzUzVDUVlkRjJQV3JraFF3S0ZBcTBKSnljTW1aMHRMUlRQT3VzczNMRmpoMmhMTUJoRTB6U3hycTVPN01Jbjk0Zlg2NFdVbEJSNDhjVVhiUUpSRnFKeXY5eDMzMzJvenBuSzI0dXE1M25paVNjQzlZMXBtcmh4NDBaVUhURzBEdHBwWlVxOFp2TGt5Wk50K3lhek1HU1lHRFFsREdsT3E3UzBGSWNQSDQ0REJnekFnUU1ING9BQkEvQ3NzODdDSjU1NEF0ZXRXeWNjRExMZ2tMWEVVQ2lFaFlXRm1KT1RZOU5zYUZBZnlvUUNLU2twc0hMbFNzZnp1dlBPTzhVYWFYbEpIYjNtNXViQ1J4OTlKSTUzMGc0Ujl6aytyci8rZW94MUxxb0Q1T1dYWHhabEJnSUI3TisvdnpoV3puUWovNVpvVGw5Tm1UTEZOdWZwcENHeU1HUVlpRTh6dEN3TEt5c3JvN2JQSlBMeTh1Q2RkOTRSeDhzQ1E5WVlEY1BBSjU5OEVsTlNVc1J2MVpnK2VlMXRjd2Rqck5DZUo1NTR3bEZ6TFNnb2lHdENjc0NBQWZqdHQ5K0tjMUNuQXVpNzB0SlN2T0dHRzRTR3FKNGJDYmpNekV3b0tTa1JRbXJldkhub2RGeEw4SmUvL0VYVTQrUlJabUhJTUw5enNNS1F0SmFoUTRlS1dEMG55R3RiVjFlSEYxNTRJYXJhVHF3a3BNMUJYc3BHZ2lndkx3L0l2Q2ROaklURHRHblQ0dmJPNU9Ua3dJWU5HMndlY1hvdkM4YXFxaXE4NVpaYmJNSk5GVFQzM1hlZk1GMDNiZHFFLy9FZi94RjFmRXNKb252dXVZZUY0VkVDcjBCcDVaQTNjODJhTmRxaVJZc0FFVzFlWVZTOHd5a3BLYkJ3NFVLUVExSjhQcDlZVVNJdjBZdG5NS3BKQ0tnK1doNTM3NzMzb2h5YVlsbVdNTlBYckZrVDF6bDZ2VjRvTFMyRmE2NjVCZ29LQ2tScUxYbkpIZFdaa1pFQnMyYk5ndXV1dXc3SllTSWYwN3QzYjdqaWlpdkE3L2REWFYwZC9NLy8vSTlZZWlqM1dVc0pJazd1ZXZUQXdyQ1ZJNXVDOCtmUDF4WXNXQ0N5UnRPYVhrcUtBTEJ2cy9adTNickJLNis4Z242L0gxd3VsOWlDMCtWeWdhN3JRb0JnTThOcVZBSGNvVU1INk5ldm4xaVNGZ3FGaEFDTFJDSWk3WDlUNTZmck91aTZEai8rK0tNMmRlcFVLQzR1dHAwM1NrdmRBUGJGRU02Y09STW1UcHlJOHJwcVRkUGdubnZ1d2Y3OSt3TWl3dHExYStIMTExL1huTW80bUEydlpHSk5ON0NBWkJpRmxwZ3psRWxLU29JVksxWkVsVUZ6ZHZRYUNBVHdnUWNlUUlCb0o0cWMyaXZlYzVEbkhNa0RlL2JaWjJOUlVaRllZU0tidG1WbFpYamFhYWMxS1cxVjc3ZW1hVEJzMkRDc3FxckNZREFvY2pMSzUwZnppSHYzN3NYSmt5Y2p0ZStaWjU3QnVybzZjVnkvZnYwY3IwVkxldEpuekpqUnFLZWZ6ZVMyQTJ1R3JSeDE3aThVQ3NITW1UT2h1cnJhWnJLcWdpMHhNUkZ1dXVrbU9QUE1NNFgyUkFIT2NuNi9lSkcxU05JcXUzZnZEaGtaR2VEMysyM2JDeUFpVUVMV3BpQnpPekV4VVp6SG1qVnJ0SFBQUFZjRWM2dkxFbW5Pc24zNzlqQno1a3k0K2VhYjhmenp6OGZKa3lkRFNrb0toTU5oR0RseUpQejIyMitPdjNYcXJ3UGxRQnhSVE91RWhXRXJoekxheUFOdTFhcFYybU9QUFNiMlBnRUFZUzVUTmhvQWdFNmRPc0hreVpOdGV5TkhJaEZibnIrbVVGZXBrUERDMzFlSkpDY25pMHcxdEdxRFlnRGpFWVowYnNGZ0VIdytuekQ5MTY5ZnIwMmNPQkZLU2tyRUNoSUErOGJ5THBjTE1qTXpZZHEwYWJCNDhXSnd1OTBRREFiaDFWZGZGWWxyNmJkMEhpMnRsWEVLcjZNSEZvYXRIRTNibjR0US9qeC8vbnh0NWNxVnRwUlljc0lDL0gwNTM1VlhYZ2wzM25tbmJRVUhDYTU0NWd5ZHdta0E5bWVjbG8raDdOWUErNFRFd0lFRG15eGZ6b3hOKzZWWWxnV0dZY0Q3NzcrdjNYampqZERRMENBY1NiSmppT3J1MGFNSFpHUmtBQURBbDE5K0NiTm56OWJvUVVHYUlKVXI1MzlzQ1lFWWF5c0JwdTNCd3ZBUWc3L250cVBCNXlTQTFNRWtDeDA2bnJROUtpY2NEc01mLy9oSExSQUlpTTJSQ0RuRHRjZmpnV25UcHNIZ3dZTlJMcjg1N2FmZnlKb1ZDUll5T1Vub3lHWGZlZWVkTnUxUURZeFd0d0JRNS9SdzN3NTYyc2tubjZ6VjFkWFpjaVU2ZWROMVhZZS8vdld2VUZSVUpQcFFiVDhKYkxtZGNodWQydFlZdE4xQlkzMm5PcXRVWnc3VE9tQmgySXFJMTdzcmF6bC8rTU1mb0txcVNndzQyVlFtZ2VEMWV1R05OOTZBamgwN0FvQTlzM1Z6MjBTL3BSM3daTE9aSENzMEo5bXRXemM0OTl4elJlZ050WWxDZkdTdFZrN25UMlZTa29jdFc3YkFoQWtUaEpBRDJDZXNaQzg2elNPKzhzb3JjTVlaWnlEMUJ3bHFlVzltOWR6azc2aU5hcUxkV0RRV290UlUvemJYbTg4Y1dsZ1l0aEpVN2FFcFNDdjc1cHR2dEVXTEZvbndHWG4vWXRuazdOdTNMMHlmUGgxcEh4Skt1OThjWktjRElrSitmajRVRlJYWkJHUTRITGJ0ai96NjY2L0RjY2NkSjlwTTU2bnJPdmo5ZmlINDVFQnVUZE9FbzRmNDVKTlB0RW1USnNHbVRac0FZSjhXU09kQTUrcDJ1MkhBZ0FFd2QrNWNPT2VjYzlEcjlZcjZhRHNBcDNPU0E4bmw2eERQUGlqTjBlNVlFMnpkc0RBOERNUWo2R0taVWszOXhySXNtRHQzcnJaMTYxWWhsR2lUZGRMVUtCWFlWVmRkQlVPSERrVktUZFdjb0d1MVBZZ0ltelp0MGpadjNpeUVpV0VZWWs2UDJ0S3VYVHU0OTk1N01UMDlQU3FISU8wNVFrS05ORVlTWHFUWkVjdVdMZFBXcmwwTEFQdFRrZ0hZVjljWWhnRkRodzZGdi8vOTczREZGVmVJM0lVdWwwdG96ZFIrMGp4bDg1L0twbkxqNlI4MU1EMWVXRE5zWGJBd2JDWEVJekNkaEpkbFdWQlJVUUVqUm96UTZ1cnFBQUNFcVVyL0oxSlRVK0dERHo2SXVhTmJySGJGYWtkTlRRMHNXYklreXV5bWVVU2FtN3YyMm12aDdydnZSam52SWptRnlMU251aWljUmphZi9YNC9aR1Zsd1R2dnZJTVRKa3dRMnFDOGQ3SmhHRFp2ZG5aMk5peFlzQUJ1dnZsbWxJV3EvSjQ4OWZJbVZyUU5LUW5LbG9EbkNCbm1kelJOZ3p2dXVLUFJ3Tnl5c2pJa2oyaGpjMUJPZjhURWlSTkY2aXpFL1FIS2lDalM3bHVXaFI5Ly9ERlNUc0htNEpRSjJ1VnlpU1FTY3BwOU5lczJJdUxDaFF2eHJMUE9RcWNOM1dVdnRWeVB6K2VEMGFOSDQ3Smx5OFQ1QklOQjNMQmhBNGJEWVZ0QXRnb2QrL0RERDJQNzl1MEJ3TzRzVVpPNzB2dm1oTXM4L2ZUVFVSdklxMUEreHBaTUVNRXdiWko0aEdGcGFTbW1wNmVMNCtYZk5nVnBVcW1wcWZENDQ0L2J5bFVUclpxbWlZRkFBTysvLy82NGJEUXl0UnRyVjNaMk5uenp6VGMyd2FkQ2dtemJ0bTM0d2dzdmlCeUg2dm1SSUVwT1RvWlJvMGJoNHNXTHNhU2tCQTNEd0Vna2dxWnA0c3laTTNISWtDSDQ3cnZ2Mm9TOUNnbG4welR4ODg4L3h6Rmp4dGdTTmxEOFljK2VQZUhxcTY5R1ZRakdJeFQvOXJlL05Tb01MY3ZDMmJOblJ3bEQxaEtaWXhZbllTZ3Z6OXU3ZDI5TVlkall3RkdkQWprNU9mRE5OOS9ZRW82U1lKRHJMQ3dzeFBQT082OUpnUmhMV05ILzVIVC9SVVZGdGwzdVpJRWtaNTR4VFJOMVhjZWRPM2ZpNDQ4L2ppTkdqTUFoUTRiZ09lZWNnM2ZjY1FkKy9mWFhXRjlmTDQ2VGhmcWdRWU5zS2J5KytPSUxXejFxLzlKNTA0Tmd4WW9WMkx0M2IvRDVmT0QzKytHdXUrN0NjRGlNNFhBWTU4eVpnN0lUSng2ZWV1cXBSalZUeTdKdzFxeFpMQXdaaGdURzlkZGY3emhvNlgxWldSbW1wS1E0Yms0VWJ6MzBldm5sbDJNNEhCWnJlSjJFb3E3citNNDc3MkJxYWlvQTJMT3ZOR2ZES0hsUWp4dzVFci8rK210SFRTbldlY3VDVHY1ZU5yZDFYY2VQUHZvSWh3d1pZaFBlSG84SDJyZHZEMHVYTHNWQUlHQ3JUMDBTSzdkSm5qNmd6OTkrK3kzbTVlV0pqRHV4aEpXYW9IYmV2SG1PMjdQSzM5MXh4eDAyWWVpVWdaeGhqaGttVEpoZ0crU1VqaDV4WC82LzB0SlNFZlpDeEN1VTFPTjhQaDlNbmp3NVNnQ29nc0EwVFh6eXlTZnhZQWVuM080ZVBYckFyRm16c0xxNjJxYlpFZkpuZFU1VEZhSzZydVA2OWV2eHBwdHV3czZkT3dPQTg3eHByMTY5NExISEhoTjlTZWNxNzVkTWZTOS9wanlFbjMzMkdRNGVQTmltY2Nyem1yR21CelJOZ3llZmZGTE14OHJJZmZ5blAvMEo1WWNOZWRBWjVwaUN0dm04K3VxclJRWVdlYkRRZ0ttcXFrS0FmY3ZORGpSRnZ4eGM3UGY3WWNHQ0JWSENrT3FtdGxpV2hiZmZmanZLQTVWUWw3NDVRZlhKcjM2L0h3WU5Hb1NmZnZxcFkzMHk2aWJ5MUQ4N2R1ekFLVk9tWUU1T2pxMGY1WE9WKzlqcjljSUpKNXdBQlFVRnFDS241WmY3d0xJc25EbHpadFIrMVU3OUdtdVYwTk5QUHgyVm5Wc1ZpQmRjY0lIUWFPa2NEdVUyREF6VGFuRzVYSER6elRjMzZsd29MUzFGZVlBY3lKeVNhb1lOR2pRSXQyelo0aWdFNkwxcG1yaDE2MVljT25Rb3FpbSttb3ZjWnZwOSsvYnRZZnIwNmJodTNUb3NLQ2pBb3FJaXJLeXN4SnFhR3F5cHFjSHk4bklzS3l2RHpaczM0NlpObS9EVlYxOUZpb1ZzN0J6bCtzaTBkYnZka0pLU0FoTW1UTUIxNjliaHRtM2JzS3lzVEFqRVVDaUVaV1ZsV0ZSVWhNdVdMY096empwTE9GUmtJUldQUjVuYThlS0xMMFk5MkZUNjlPa2pmdGVjbmZlWXd3dlA0aDVpS01ENWlTZWV3THZ1dWdzQTltZGNvVmZETUtDeXNoTHk4dkkwU25RS0FDS2hRRk5RV2VyeEhvOEhycnJxS3B3L2Z6NWtabWFDcnV1T0dvbGhHTEJxMVNxNDVwcHJ0SXFLaXFpMXp2SFU3ZlY2UmZJSGRiOWhXZytjbHBZR3h4OS9QSGJyMWcwU0V4TWhNVEVSYW1wcW9LaW9DTFpzMmFMVjFOU0FZUmdpUGxGZUJ5MURnZGwwblBwL3FqOHZMdytHRFJ1RzNicDFFOWwxZHV6WUFmLzYxNyswSFR0MkFNRCtaWVQ0K3dvV2lwR1UxM2VyeU4rLy9QTExlTzIxMXdyaHFDN2owM1VkTWpNenRWQW9GSFV0WTVYUE1FY3RMcGNMM24vL2ZhRUZxazRDUk1UNitubzg5ZFJURHpvZWpiSk9FMzYvSHg1NjZDRVJqeWVIbTZnT2x1Ky8vOTQyTXB1VHJDRFc5MDJkaXp4UHA1cnE2dnljYXJiVG5KNnNFYXZKSUdSa0FhdkdHOHAxcXRwbnJEbERsOHNGcjczMm1zMVpJNXY3MXUvN1Bxc2FPOVhQRGhUbW1PT3NzODdDalJzM1JwbFA4aHhhT0J6R1YxOTlGUS9VbTZ3S0pEV2I5U2VmZkNLRXNTcUk2ZnR3T0l6NStmbllxVk9uWmdVZWt6Q1JQYTJxbWFzS29sZ0NWQlpHc2JZNVZSMCtMcGNyNmhoVmNNb1BDVmtBcTFNVDh1b1hwLzJZNWZkcGFXbnczbnZ2T2M0RjArZlBQLzhjR3hQUURIUE1NSFRvVUZ5eVpJbHRvRGlGa3BpbWljWEZ4WGpycmJkaWN5YlhuZWJwbk9qUW9RTzgvLzc3R0FxRmJDdFUxRUVjaVVSd3c0WU5PSG55Wk96ZHUvY0JuWE5qcTJTYzJ1czBSK2YwRzZmMFdySldwem9uU0FpcDJhNmRrUGQybGw5ai9kN2o4VUNYTGwxc2NZN3lLL1hucUZHalVLMlhONUZxbmZBVmFRRzhYaTljZXVtbDJMbHpad2dHZzVDWm1RbnQycldEM054Y0dESmtDSFRyMWcwU0VoS2k4Z0hLcWExb3ptcjM3dDN3M1hmZndaWXRXNkN5c2hMcTZ1ckFzaXdvTEN5RXp6Ly9YSlBuNHVLZGM2STF3TjI2ZFlQcnJyc09IM3JvSWZGN1dzOUxoTU5oOFB2OVlCZ0dyRnUzRGdvS0NxQzh2QnpxNit2Qk5FMG9MUzJGMTE1N1RhdXZyN2ZOQ3g2dDBEeWluQkNXK3Z6TU04L0V0OTU2Q3pwMzdpeldQTXZYdGFhbUJqcDA2S0JSdWpXZUgyU09hdHh1Ti9qOWZ0aTBhUk1HQWdIaE1aYk5YOWtVZFhxdm1xMjZyb3ZsWi9SKzBhSkYyS0ZEQjFGbmN5RXZwc3ZsZ3R6Y1hQajN2Lzl0MHd4bFQ3ZnFFYVcyQllOQkxDc3J3NzU5K3g2VDgxMmtjVkpmWG5ubGxXaFpsbGp0STEvVFNDU0M4K2JOdzFpZWVkWU1XeDhjK1htUW1LWUpQcDhQTm03Y0NKV1ZsVUp6SUMyQXNwK1FWaUdua0hLNzNTTFBucXc1VUV3ZHBad3lEQU8yYmRzbXZMeHkyVTFwWjJTbVJpSVI0ZkhkdFdzWERCOCtYT3ZkdXplTUd6Y09Uejc1Wk9qUW9ZUHdCSHU5WGxFMlpYYWhuSUdrSWNwN3JSek4rSHcrc1U4MTlUK2xQeHN3WUlCdzdKQ25ucTVaWFYwZGZQVFJSNDRhb2V5c1lXMng5Y0NQcDRORWpuV1QwMEhKa0FsRi95UEJLRzlzRHJCL2Jrb053VUJFOEhxOVVGOWZiNnMzM29IazFDYTVESHBOU1VtQjFOUlVNYWRGb1RLR1lVQWdFQkJDTVJnTXh0YzVSd24wRUtNSGcyVlprSmlZQ0pzM2I4YXVYYnZhMG9MaDcza1pGeTllRERmZmZMUFcwTkJ3aEZ2UE1JY1JOUnlEdnBQL1lrMmFOMmJ5eXM0Qkp5L3pnWmlxYnJjYmtwS1NvdXFSdmFqVWZ2bTk3TlFoeit5eFlDcXJqZy9pbW11dWNkd0gyelJOTEMwdHhSNDllckFwekJ4YnlLc2ZpS2E4aFhJOEhIR29NcHBRdTN3K242MWNOZUdBM0E1WmdNdW9udUZqUVJnQzJOY1NhNW9HSFR0MkJGcVBMSzkvRGdhRFdGTlRnK2VkZHg2cTE1QUZJM05NUUpvaGFVK3FOcWNLRUNjaG8rNGlSNExIYWVjMkt1ZEFCNWphSGxVd3h4THNzYlNrb3htMW56dDI3QWovK01jL01CS0pSS1VyS3k4dng5dHZ2MTNNWGNqM2c5TzFZZ0hKSEZXb2c2VTVua04xNVVOamRjaGxOWGNReGRMZzFISWIwL1RVT0x4alJTc0UySCt1NmVucE1HL2VQS3l2cjdlWnhaWmxZV1ZsSmQ1KysrMlltcG9hbCtCcnFyOFpobUVPTzJwd3VQclo2L1dDeitlRHI3LytXb1JLeVNGSnBtbmlpQkVqc0RGTm5tRVlwbFdqYW15eXRwYVltQWpEaHcvSHUrNjZDL2ZzMldQTGsyaFpGcGFYbCtQeTVjc3hPenM3YWtya1lLWXhHSVpoamhpeUZ0ZXZYejk0N0xISGNNbVNKVmhVVkJTVmg5RTBUVnl6WmcxT21EQkJiRFJGampSTzJzb3dUSnVFekdGWmt4czdkaXhHSWhHYlNXeFpGdXE2anF0WHI4Ymh3NGRqWm1ZbUFFUTd0UURzcTMyWXRnTS94cGhqR2dwY2x3UFlkVjJIaW9vS3FLeXNoRWdrQXNYRnhiQnk1VXBZdUhDaFZsVlZKVmFhVUJDNjMrK0hjRGdzUE8rUlNNUXh6eUxETUV5YndlZnpRWmN1WFdENDhPR1lsNWNuTWwrVCtSc3JwMkpqV1hZWWhtSGFERTRtTFNWZ2Rab0hsSU93Q2RuN3pEQU0wNlp3RW1ieHhHV3F2K0VFcmd6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1BekRNQXpETUF6RE1FY3ZSenpIa0xvdkxVRHpOanRxS2sxU2E5bDlMRlk3VzB2N21BT2pxZXQ2cEhmQWErNzRVSTkzYXJ0OERPMm9TTHMwSHVuelBSaU91REJVb2JXZTZnNXhoNG9qTFV4Yk11ZWRVMWx5KzUyeWFqZjMvT0laTEFkVFhuTnBxdjZXYWwrc2NscjcvWEU0QkpOVEh4M3VjZHdTdEFwaDZIYTdZZFNvVVdoWkZuZzhIdEIxWFd5OUtXZithTzd1Y3ZIUTFNM1dWSDQ2cDQyYzVETFZ6YUZpTGVvLzBQcnAvQnZiV2lCVzIrSVphT3J2MU44ZmJKcXFTQ1J5VUw5dktqTk1VL3RLTnpWWW5iWnRqZlZaM2krYlBqZFYvOEcycnluaUZZWk8yWHRpbmF0NmpLN3I4TkZISDJrMFhyMWVyOWhqdWkzUktsSjRVUm9rZ0gzWmhTT1JDTGhjTHRCMVhXeExTYlMxWGNkb3NEZFhDQjBzVklkbFdRZFZ0NU5tS1g4KzFKcEhXeks1bkFRSlBheGluVWRURC9QRG1TeldTZGc1V1JGT0R3RDVvVWlDVUo3NmFnc2NjVWxDMm9WVHB6WFdtVTJaTHkzWnZzWm9xdjVEYllZZmFUTy90ZE5TMXkvV2NRZGJmbE9hOWFIT2lYZ3dVeVpPcU5aY1crS0lhNGJ5QlZDRlgyT2QybElPbHFab2lSdmtZTXBvYVdIYzNMYW9nN1V4TS9GUWNMRG5mNmdmVm8zOVA1NTdyeW5CY1RqdjM2Ym1uSjBnazFnV2dyVG5kbE5UQUsyTkk2NFprZ0FranhRbDByUXNDMHpUUE9LYVQwc0lvNE1aTUFkNmZ2RnF6b2RhODIySzFxNjVIc3pENUhCWkw0ZVNlTzRQeXZTdHdtYnlBUkNQT1V3MEZRcWdjcVJ1eENNNUVPSXhmUTYwZlMzdFRXWU9MYzE1V0IzTXRTUmxoalRFdG13dXQybGNMcGR0SWxxZGRGWTMvQ0hjYm5mVUpqN3hiT3J1VkwvYUJxZXlEbVF5UEZhN25YRGFZTDZ4RGMzalBVYzFjN05zT3FzYklqV24zNXk4MGVwbTllcnhUdlU0WlpodWJBdFFLbC91Qi9xLzdKMlA5MXpVdHNxZjVYTFVkc2ZxeDFqbkVDL3lOZ1dOZWZ4YnUvT1JhU2J5UlhjU1JrNFhuSVNYL0ptTzkzcTlqb0l0Rms2Q1NoVVFOUEFPUmlpcTU5ZFVTSTM4T3pvZnB6Q2dwbEFITnVHVURwL2E1SGE3aFdDS1I2aTQzVzd3ZUR3eEJhdkg0eEY5S05lcENqQ3F6K2tZcDNOUWYwdnZLYktodVdGRGpaMkRYQmFkRDdWSHZTN3F0V29NK2ozVnBkWXBDM3dxbC9aMVlZNVMxRUhncERtb3gzZzhIbkhqTysxYkVjOWVGalRZWlUxRHJWK0ZidlI0Ym5oMXNNZzN1Q3dRYVVBNGFXZXg0aHpqSFJBK244LzJPNmZ0TVozK1I3OVIvOVMyeWRkRTduTlY0S3NidkFPQVRVaXEvZXIzK3gzYkpRdHRPbDZ0U3czcGFneXlLcHo2UlVVV2hDcXh0T1I0aUtVdEEvQ2VMTWNNVGhxVGt6blNtRmFrL2kvV0prRHhJQWRaazZiZ1pNYkZXNzU2SHVyZ1ZzOU5QUSsxYlU0RHR6RmltYXNKQ1FtMnRxbDdCZE5ucDkvRzBxWmxUWS9hcDJwYkFJMFBibldqSnZWQklndE5xa3VsdWVhK1hJOXFxWkMxNGRTUGNoK3BKbjY4OVR1WjkxUlhZOU1aNnYrWm80UjROUjNWWktDYnp1bEdiazdkVkI3OTNtbkFxUnBVdk1qbWoxcWVrOGFubWtyeG10UHh0TVBKZkhPYTQydEtFMm1zTGJFZUh2STFVNCtscVEyMVRmVHFKSXpraDZpcVVUc0owbGc0VFFlUTFxcWVxMnc5cUJ0SU9abjh6ZFhvWWdtL1dQM05tMWZaYWZNVEIwN2VLNWZMWlZzYWxaaVlDRU9HRE1HQkF3ZENjbkl5SkNRa1FIVjFOZno4ODgrd2NlTkdyYnk4SER3ZUQ1aW1LYUxwcWJ5bVBHSXVsd3Q2OWVvRmdVQUFrcEtTSUJnTVJqMmxBNEVBR0lZQlpXVmxVUXZiNDBFT08xSUQxSDArSCtUazVFQ2ZQbjJ3VTZkT0FBQ3dkKzllMkxadG0xWlVWQVNoVUVpY0M3VkgxM1hIc2hxRFBQN1V0NzE2OVlMZXZYdGpSa1lHSkNZbXdzNmRPNkcwdEZUYnRHa1RXSllGaUNnR29WTUVnSlBuVWw3UG1waVlLSlpsQm9OQjBWYjZyUnFCUU44bkpTVkJJQkNBbEpRVXNmb25Fb2xFSFUveGNmUTd1VTJwcWFtZzZ6b2dvaWdqbHFkVnZkZjhmci90R3BPSHRiRkVCajZmRDB6VHRMV0o3c1Y0SU1GdUdBWjR2VjdvMTY4Zm5uamlpZENqUncvd2VEd1FpVVJnKy9idHNIYnRXcTJvcUFoMFhXZFA3OUdNYkZZUlBYcjBnRGZmZkJPM2I5K090YlcxV0ZaV2hwRklCQkVSNitycXNLNnVEb3VLaXZCZi8vb1hYbmJaWmVMT2E0NW1rSkdSQWJ0Mzc4YWFtaG9zTGk3R3ZYdjNZbVZsSmU3WnN3Zkx5c3F3cEtRRVMwdExjZmZ1M2JoejUwNWN2bnc1WG5ycHBSanZKTGFxTWRENXRXL2ZIbDU2NlNYY3VYTW43dDI3RjJ0cWFyQyt2aDdyNit1eHRyWVdTMHRMc2JDd0VGOTY2U1hNeXNvQ0FPZTVyYWFRemE3TXpFeFl1SEFoYnQ2OEdjdkx5N0dpb2dJRGdRRFcxZFZoVFUwTmxwV1Y0YTVkdS9EbGwxL0d6cDA3QzNQWGFaNVExY0xrNjNibGxWZGlhV2twRmhVVllXRmhvYmcyVHVhZjJrOWp4NDdGb3FJaTNMbHpKNWFVbE9DVUtWT1E1di9VK1VtNVBmVDczTnhjS0N3c3hKMDdkMkpSVVJHU2dHN3NXc21hN1BMbHk3R3dzQkFMQ3d0eDl1elo2S1RseXU5UE8rMDAvTzY3NzdDOHZCeFhybHlKSjUxMFV0UjkyQml5QmZELy90Ly93eDA3ZG1CeGNURTJORFJnS0JUQ1lEQ0k0WEFZZzhFZ0ZoY1g0L3IxNi9HQkJ4N0E1c3lKTW0wRXVwRmxjeU01T1JuR2p4K1BvVkFJVGRORXk3TFFzaXdNaDhOWVVWR0JlL2Jzd2VycWFqUU1BeEVSRGNOQXk3SXdQejhmYzNKeW1sVi9VbElTSUNLYXBpbktJWFJkUi9sL2hHVlorTU1QUDJDUEhqM2lxa1BXTkJNVEUrSFBmLzR6MXRiV29xN3JhSm9tQmdJQjNMTm5EK2JuNTJOK2ZqN3UzYnNYZzhFZ21xWXBoUCtNR1RPd2ZmdjJCN1MvYjJabUpreWRPaFV0eThKUUtJU0lhS3R6MjdadHVIdjNiZ3lIdzZLL1E2RVEvdmQvL3pmNmZMNUdrMVdvNXI3YjdZWXhZOGFJZnJRc0N5c3JLN0YzNzk3aXQzSTU4bm4wNjljUHFxdXJiZjEvenozM29OUDVra0JTcHkrNmR1MEtWQzhpWWl5bkZLRjZpWC8rK1dkeHpVM1R4RWNmZlJRek16TUJJSHBlRldDZk1Dd3NMTVJBSUlCYnQyN0ZRWU1HTlVzWWVqd2V1T2lpaS9EWFgzOFYxd1VSTVJLSllIbDVPZTdac3dkcmEydXhycTVPZkkrSVdGWldodVBIaitjZ1VZa2p2aHp2WUNFVGgweW9wS1FrZVBEQkIvSE9PKzhVWnNiR2pSdmh3dzgvaEpLU0VpZ29LSUJJSkFJZE9uU0EzcjE3UTc5Ky9lRGlpeStHdExRMCtQVFRUOFVpODNoTldicGhLZVBPdkhuendEUk5zQ3hMbUdZcEtTblFwVXNYdU95eXk4QXdEUEI0UERCbzBDQll1SEFoamhrelJnc0VBcUJwbWlqRE1JeW9QSTh1bHd2OGZqL01uRGtUcDA2ZENycXVnOGZqZ1E4Ly9CQldyMTROMzMvL1BlemF0VXZUTkEyNmQrK09nd2NQaHVIRGg4UG8wYU1oRkFyQjlPblQ0WVFUVHNBcFU2Wm81ZVhsb2t5NW5RQVE5YjVqeDQ0d2QrNWNIRHQyck9qanQ5NTZDOWF0V3dkcjFxeUI0dUppVGRNMDZOYXRHdzRaTWdUT09PTU1HRFZxRlBqOWZwZy9mejRNSERnUUgzendRYTJzckV5WWlMS1pLeS9oUWtUUkxsbHpURTFOaFJkZmZCR3Z1KzQ2cmFTa3hOWS8xS2JNekV4NDlkVlhNVDA5WGR3VDhoeW5iR0pybWlhdU05VlA5NHFxc2RKMWtmOHZ2NmZwQi9xT3pHcTZsdE9tVFlPRWhBU2NPM2V1VmxKU1lxdVRIaEkwTFNEWHF5NXpvMWM1STR6TDVZSXJycmdDLy9hM3YwR25UcDFBMTNXb3I2K0gvL3UvLzRQZmZ2c050bTNiSnFadmV2ZnVEVDE3OW9RUkkwWkFYbDRlRkJZV1FsRlJrZTFlbHU5NURwcHVnNmczN3cwMzNJQ1ZsWlhpNlR4djNqdzgvdmpqYlpxQlBNR2RscFlHSTBlT3hERmp4bUJTVWhJQU5NL0xscEtTQXFacGlxZHVSa1lHQU95ZnRQZjcvWkNRa0FBZE9uU0FvVU9IWWlBUUVFL240dUppUFB2c3M4WFRXWTJ0VTVrMWF4WmFsaVcweklrVEoyS0hEaDJpdk9IMGw1bVpDUTg4OEFBMk5EUWdJbUk0SE1aSEhubkVVUnR3TW1VQkFLWk9uWXFCUUFBdHk4S2FtaHFjTVdNR2R1alFJZW8zSkhpeXNyTGdwcHR1RWxwNUlCREFtVE5ub3F5Qk9abWQ4dWMvL3ZHUFFyc3pEQU1OdzBEVE5JV1dwenE2TkUyREYxNTRRUnhMNTJxYUprNmJOaTFLMDJwTTQrcldyUnVnaEpOVzYvU2VZaXZYclZzbjdqMjZWb0ZBQUo5OTlsbE1TVWtCZ0dnemVjZU9IWWlJbUorZmowT0dEQkh0amFYRnkxcHNWVldWMElLcnFxcHc4T0RCb2g3WksrOXl1U0FoSVFGT1BQRkVlUFRSUjdGbno1N0NvUk1ySE90Z25XN01ZVVllWkRrNU9mRFZWMStKQWZIMjIyOWpVbEpTVkRBdFFIeWU0M2lFWWxwYUdwQkpaWm9tcHFhbVJ2MVdEdFZadEdpUk1OdXJxcXBzcG9vOFNOUzYrL1RwQTcvODhvc3czLzc2MTcraUdyeXIvcFplUC9qZ0EyRStOVFEwb094OVZiM09oS1pwa0pLU0FoVVZGVUs0dlAzMjI2aitUZzFRZDd2ZGtKaVlDRE5uemhUVEJKczJiY0wrL2ZzM2VuNnhoS0ZzOXBtbWlYMzY5TEg5enVQeHdQang0N0d5c2hJTnc4QklKQ0tFUHlMaWZmZmRkMGlGb2Z6bjgvbGd3NFlOR0FnRU1CZ000dmJ0MjlHeUxORVB6enp6VE5RRDk5eHp6OFhpNG1LMExBczNiOTZNZ3djUFJ2aytwVDVWKzlubjg4R1NKVXN3RW9tZ1lSaFlYMTl2bXdlVTUwRzlYcS80TGVVSmpjV3hMQVRiZktDUmJPYWtwS1RBYWFlZEJpNlhDMEtoRUN4WnNnVEM0YkM0d09Gd1dBUzl1dDF1U0VwS2dvU0VCTWpLeW9MazVHVEl5c3FDdExRMEVTWVRqNWxBcGhJQVFEZ2NGZ05lRHJPaC8xOTQ0WVY0M25ubkNWT3RwcVlHZHU3Y0tjcVM4OENwZFo5enpqbVlsNWNIbXFaQmZuNCtyRnk1VWlTem9IYkliWkw3NXFxcnJ0S29QUWtKQ2ZDZi8vbWZLSHZlQWNDV0ZJUE12cUZEaDJKV1ZwYnd4dDV4eHgyYWJCYktwaTBOTXRNMFFkZDFXTFpzR2V6WnN3Y1FFWHIwNkFFREJ3NUU4b2lyN1cyTUgzLzhFYVpNbVNLOC9CczJiTURqano5ZVBOZ3lNakxneGh0dmhPVGtaSEM1WFBEVVUwL0J4eDkvTEs1SGMrbzZFTWc4eHQ4OXo2RlFTRmdldDl4eUM3ejY2cXZDbTN6cnJiZkNzODgrYTlQbXcrR3dNTFhwbkdTdk4vV3BuTGlFZ3VCNzkrNHR3b3FtVDU4dVRIU2Z6d2VJS014ck9vYnU3YVNrSk1qTXpJUjI3ZHBCY25LeTdYeU9WVUVJY0JUTUdaS3crVjJUUVlCOUZ6UVNpY0MvLy8xdlRSWVdkRk8rK2VhYjZQUDVSTFlOQ2lCMnU5MVFXbG9LRHp6d2dMWnIxNjY0NmcrSHd5S3NJVEV4RVdiTm1vVTBEMFJhaGR2dGhuYnQyc0hnd1lPaGE5ZXVJdVNpb2FFQmZ2MzFWNDNhaDBxSUIzMFBzTStKa1pLU0FwWmxRWEZ4TVd6ZXZGa0lKZ0I3S2lqMW5NUGhNR3pidGczNjllc0hMcGNMemp2dlBIanp6VGZGZkpWYUY1M1BpQkVqd0xJczhQbDhVRlJVQkRUblJZS0pIZ1F1bDB1RW9uaTlYakFNQTM3ODhVZXR0TFFVdTNidENuNi9IOUxTMG15L3d6akRSckt5c3VDMTExN1RUanp4Ukp3MGFSSzQzVzZZTTJjT1huLzk5Vm80SElaNTgrYmgyV2VmRFpabHdVOC8vUVF2dnZpaTl0aGpqeUgrbmszbGNBeHVFbGgrdngvOGZyOHRZY0hOTjkrc1ZWWlc0dFNwVXlFU2ljQ0VDUlBBNS9QaGZmZmRweFVWRlVGQ1FvSUlyU0ZCSjRmaHlBOHR1bDZSU0FUNjl1MHI3Z2VYeXdYNStmbWlQU1FVWFM0WFhIbmxsWGoxMVZlRDErc0ZSQlIxVVNqWjNYZmZyVzNjdUZIOE5sWjQyckZBbXhlRzhvUnZZbUtpdUdIb0J2WDVmQkNKUkd3WDlweHp6b0hPblR1RHJ1dmlKZ0hZZC9QOTl0dHZrSktTSXI1dnlvbENUM1BTdkc2NjZTWWhoRlJuQUVGQ2NzeVlNVnB0YlMwQTJEVU1Fb3IwV1c0THhRbEdJaEZ4UTFQYlNkdVFKL3hKS0ZPOEdUMEVaSUVrTzAxSW85QTBEZExUMDIwVCtqU1lJcEdJVFpCU1dTUVVxUTAwY0gwK1g1U2podnFzcWY2dHJhMkYrdnA2bUR0M3JqWnMyREFjTUdBQVhIVFJSVEIrL0hpc3FxcUM4ZVBIaS9hZWZ2cnBHbWxOY3A4ZVN1Z2NhUDR0SEE3YkF1VXR5NEo3NzcxWHk4akl3QnR1dUFFTXc0Q3JyNzRhRUJHdnUrNDZqYTRWUGFncDVJcnVBeXFEN2duNlRIR0lkSS9SOTNJNkxVU0VidDI2d1dXWFhTYmFxenJPWnMyYWhZaW9BZXdYZ01jcWJkNU1sdWZNeXN2TE5mSTBoc05odU9TU1M5QnBqNDFmZnZrRlZxOWVEZDk4OHcwc1g3NGN0bS9mYnJ1eDNHNDM2cm9lbHpkWnp1Tm1HQVpVVmxiQ25qMTdZUGZ1M1ZCV1ZnYkZ4Y1ZBWlZtV0JZV0ZoYkJ3NFVMbzNyMjd0bTNiTnNlNVBoVmQxNFc1YjFrVzlPalJBL3IyN1l0eSswaUl5WUtRU0VoSWdPN2R1NFBYNndYVE5PSExMNyswcFdwWGhRZVZzV3JWS2xGdVZsWVdaR2RuQzYyRGhDd2REN0EvZE1UajhjREFnUU94VTZkTzRQUDVRTmQxS0M4dmp6cS9lUHFYaFBEMjdkdmh5U2VmRko3a21UTm53cUpGaTBEVE5LaXJxNFBMTHJzTVFxR1EwTERVUDVtV0hQQ2tCWnFtQ2FGUVNNenZCUUlCMFhjQUFKTW1UZExlZnZ0dGNiOU9uRGdSWG4vOWRjekp5UkhYbHN4WnVYMXlzRHlaeXk2WEMycHFhaUFVQ2dudjh0Q2hRd0VBaEFjYVlOOTFMUzB0aFZXclZzRVhYM3dCcTFhdGdpMWJ0b2hycHk0cW9MTHAvYkVzR05zazhzMmVrNU1EMzM3N0xaSjNkL255NWRpMWE5ZW9ZMm1laE15YXUrKytXMHlZNStmblk2OWV2ZUplTzV5U2tpTGlERTNUeERGanh1RG8wYU54NU1pUmVPR0ZGK0lGRjF5QTc3NzdybkI4ckZtekJvODc3cmlvQ1hGcW4vcEs3MDg5OVZUY3VuV3JjQ2JNblRzWDQwMGs4ZlRUVDR2NmEydHJNVFUxTmU2QThuQTRMQndBYytiTUVVSEVhbnlkN0xSSlNVbUJ1WFBuWWpBWVJNdXk4SmRmZnNHVFRqb0paZTAxWG0veTZ0V3JVZTZudi96bEw4SlRpNGdZQ29WdzNyeDVTTmMwTlRVVjNuNzdiZkg3KysrL1A4cnAweGdINmswRzJIYy9rWk1yR0F6aVJSZGRoQUQ3SFdNSkNRbnc5Tk5QSTkwdnVxN2oxMTkvalhWMWRXaWFKdTdZc1FQUE8rODhsUHRTcm9lbVhBRDJhZk5MbGl3Um52YlMwbExNeTh1TE9oK2Z6d2VKaVlsaVR2ZW1tMjRTOGFtSWlNT0hEMGU1M01iT2oya0QwRVh6ZUR6d3dBTVBDQTliS0JUQ1YxNTVCWjNTSmRFTjZuYTc0Yjc3N2tQRWZlRUpXN1pzd2Y3OSs4ZGRkM0p5TXRCdkVSSFQwdEtpNmpyaGhCUGcxMTkvRmQ3T3YvLzk3K0k0OVZocWszcHVtcWJCUC83eEQ0eEVJcWpyT2dZQ0Fiem1tbXRFVUhBczdycnJMbHZBN2EyMzN0cmtiMlFlZnZoaE1lQnFhMnR4MHFSSllsNVdicXU4VW1iOCtQRllYVjJOcG1saU1CakVaNTk5MWlaVTFNR3VsamQ2OUdnaDdMNzg4a3ViQU5ZMERaNTU1aG5idzR1Q21qMGVEL2g4UG5qbm5YZHNYbmY1Zk5SK1ZoOUllWGw1UWhoYWxoVWxET1cycW5nOEh2anBwNThRRWJHK3ZoNUhqQmdSMWRlWm1abncvUFBQbzY3cklxcUEycnB0MnpZOC8venpvN3pmVGhsM3ZGNHY1T2JtaW52UE1Bd3NLQ2dRRDFvWk9TR0VrekJzNnJ5WU5nYmRBQWtKQ2ZEcHA1K0tBWXk0TDlwKzFLaFIyS2RQSCtqU3BRdjA3dDBidW5idENzY2RkeHljZnZycFNEY3dEUzQ1QnFzcE1qSXlnRzdzWURBb1l2QUlldXBPbmp4WmhQem91bzYzM0hKTGxBWUFFSHR4UGczS3JWdTNvbUVZR0E2SEVSSHhuWGZld1VHREJtSDM3dDBoT3pzYnNyT3pvV3ZYcnRDM2IxOVl1SENoMEZpRHdTQXVYcndZTzNic2FHdGJVK1RsNWNISEgzK01obUVJVGUrRkYxN0FQbjM2UUxkdTNTQXJLd3M2ZHV3SVhidDJoVUdEQnVHNzc3NXIwM3pXcmwzcmFHdXBRbFJtekpneG9wOCsrK3d6bS9EMWVyM1F1M2R2K09LTEwzRHo1czNZcmwwNzBjZmtrWDNublhlRTFraHhobkxZaVpNUUprSGJzV05IUUVSeHJlSVJodkxEK0tlZmZzSklKSUtCUUFBdnZ2aGlzZnBGL2sxcWFpcTgvLzc3NHY0a1NrcEs4T0tMTDBhS1dYU3FUMDJJTVhueVpBeUh3MkwxVTBOREF6Nzg4TVBZcDA4ZjZOV3JGM1RwMGdXNmQrOE9PVGs1Y1BiWlorTmJiNzBsNmpWTkU0Y05HeVlzak1hbUZwZzJnanF3MnJkdkR5Ky8vREtHdzJHMExFdkVxRFUwTkdCaFlTRnUzcnpadGs2WmJvN1MwbEo4K09HSHhRQ0xoOS9ER01UZ1RVNU9qdm90UFptZmUrNDVSRVJ4ODk1NDQ0MG9hMzdxT2NrYUk1MWpjbkl5L08vLy9pK1dscGFLUVdRWUJtN2R1aFZYcmx5SlgzMzFGZTdZc1VNSUk5TTBzYWFtQmg5NjZDR2JHZFhZRWpPVi92Mzd3L3o1ODIxTC9BekR3UHo4ZkZ5MmJCbXVYcjBhOC9QelJVeGdPQnpHN2R1MzQzUFBQU2NFb1dxR09Ra1VPdGZSbzBlTDY3SjA2VktrOEJEQzQvRkFWbFlXZE9uU1JYd20vSDQvdlAzMjIrS2FQdlRRUTFIQ21FeEdlWDA0emN2bDVPUUlZV2haRmpyMWsycEt5dGVJSHF6aGNCalBQLzk4bEkrVmowdFBUNGRGaXhhSmV4UVJzYkN3RUMrODhNS1lmZVpVZjFKU0V0eDIyMjFpQ2FhODdMTzB0QlMzYjkrT2hZV0ZHQXFGeFA4b2dIN3AwcVdZbTVzYjgrSEF0REdjQnJUZjc0ZlUxRlFZUFhvMHJsaXhJdW9KTEdzdWlQdUNncWRObTRaRGh3NFZRYkh4a3BHUllRdTZWalZEbVlTRUJOdDhWbVZsSlY1enpUVml3TWdyWTlTZ2JabTB0RFFZTm13WXpwa3pCM2Z2M2gxMWJ0U1duVHQzNG9NUFBvakRodzlIT1o2c3VhbWJTQkNOR0RFQzU4K2ZqMlZsWlRidGdnWVk0ajVUNy83Nzc4ZFRUamxGSktOUXB5Ym9zMU91UXJmYkRXUEhqaFdDOWYzMzN4Y2F0T3lnVWR0R3h5UWxKY0Y3NzcwblZxRGNkOTk5S0Nlb2NMcGZaSUdYbTVzTHRQNDZIbUZJK0h3KzhQbDhzSGJ0V3FHaFhYTEpKV0tlVkUweTRmVjZJU3NyQzE1KytXVnhINWFVbE9EbGwxOXVXMlZEOVRrOUxLbnZFaElTWU1pUUlmamNjODloVFUyTkVPVHkrbTU2V05mVjFlR0NCUXZ3dlBQT2kxcTlwSjdqc2FZZHR2bXpsY00xNUhXYkFQdERCVFJOZzZGRGgrSnBwNTBHUFhyMGdOcmFXaWd1TG9aZmZ2a0Zmdjc1WnkwU2lkZzhzZko2NEhoU2VJMGJOdzZycTZzaElTRUIvdm5QZjJweU9pYnk0Rkk1ZnI4ZnNyT3pRZGQxOFB2OUVBd0dvYXlzVEhna0FmYmQzT1NacE8vVUhjaGtFenM1T1JtNmQrK08zYnQzQjh1eVlQdjI3ZHF1WGJzZ0VBaUk0OVhZUHRscjJOVDVxV0V4THBjTDB0UFRJUzh2RDd0MjdRcVJTQVNLaW9xMHJWdTNnbUVZNGppS0FZMVZoOXdtK1gyM2J0M2c5Tk5QUjEzWFlkZXVYZkRERHo5b1ZEOGRwNGJweUcwOTlkUlRzVXVYTHVCeXVXRExsaTN3eXkrL2FHcTlGS0lrcDgyaU9jZExMNzBVNit2cklUVTFGZDU4ODAwTjQvU3ErdjErNk55NU15UWtKRUE0SElhS2lncW9yYTJOdWZZYllKL3dQdU9NTXpBNU9SbENvUkI4Ly8zM1dtMXRyUWhmb3JybE9NUEcrdEhyOVVLdlhyM2c5Tk5QRjQ3QW9xSWkyTHAxSy96NjY2OWFhV2xwMUNidk5FN2tlNUJwb3pTbDZUaVpaaFNQRnN0WkVhODNXYTI3cWZXa3BDVTRyYStOOVpTV1RVUTVqaTVXTytTSmQvSmF5OXJWZ1NTeEpkVFVUNnFKMVZqYlZBMHhWdnRsMUcwYW5PcVZIUVN4MG9iRnlscmpWSjY2aXFneG5NeEtwK3VvSm0yVlBjWDBQOWt5b04rb1puWmo1eVgzdjJ4aXErYTJXazZzZTVWcGd6aE5NTXZlWWtLZGU2THZHak9mNGlWV212dFlYbUtuLzhzM3JteGFBZGpYVmF2dGRVb05GVXZnT0hrb20wSWVQUEtEd21tL0V2bTkwendoL1ZaZUhTSVBZdms5SFN2M3JTb3dpS2F1WDZ5NXNGZ1BzMWhseHFyWHFTNTZ3TW5PSDdWOWpUblJuQjZRc29ORFhhc3NseXUzUS9aQU81V25FbXV1a21ubHlEZU0rc1NUSjZ6Vjc1c2FLTTBSaEU0M21OTU5SZDg1RFRpbjlqalZvWllYU3pPZzM4ajk0eVJvbWtKMUZqVDIzdW1oSXM4UE9yVS8xcHljK3AwcVNKeWNTM0s4bzNwK1R2M25kSy9FK24wczFQNlBwWFU3aFhjNVhRdW5wQ0pOT2J2VSswcnRsOFkwZHFmN3N6RlBQOE13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd0RNTXdETU13RE1Nd1RDdmkvd012dE9UN1h5eDJmQUFBQUFCSlJVNUVya0pnZ2c9PSIgYWx0PSJSJmFtcDtKIEdyb29taW5nIiBjbGFzcz0ibG9nby1pbWciPgogIDwvZGl2PgoKICA8YnV0dG9uIGNsYXNzPSJvcHQiIGlkPSJib29rQnRuIj4KICAgIDxkaXYgY2xhc3M9Im9wdC1pY29uIj48c3ZnIHdpZHRoPSIzNiIgaGVpZ2h0PSIzNiIgdmlld0JveD0iMCAwIDI0IDI0IiB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciPjxyZWN0IHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgcng9IjYiIGZpbGw9InJnYmEoMjU1LDI1NSwyNTUsLjA4KSIvPjxyZWN0IHg9IjUiIHk9IjciIHdpZHRoPSIxNCIgaGVpZ2h0PSIxMyIgcng9IjEuNSIgZmlsbD0ibm9uZSIgc3Ryb2tlPSJyZ2JhKDI1NSwyNTUsMjU1LC41NSkiIHN0cm9rZS13aWR0aD0iMS41Ii8+PHBhdGggZD0iTTggNXY0TTE2IDV2NE01IDExaDE0IiBzdHJva2U9InJnYmEoMjU1LDI1NSwyNTUsLjU1KSIgc3Ryb2tlLXdpZHRoPSIxLjUiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIvPjxjaXJjbGUgY3g9IjguNSIgY3k9IjE1IiByPSIxIiBmaWxsPSJyZ2JhKDI1NSwyNTUsMjU1LC41NSkiLz48Y2lyY2xlIGN4PSIxMiIgY3k9IjE1IiByPSIxIiBmaWxsPSJyZ2JhKDI1NSwyNTUsMjU1LC41NSkiLz48Y2lyY2xlIGN4PSIxNS41IiBjeT0iMTUiIHI9IjEiIGZpbGw9InJnYmEoMjU1LDI1NSwyNTUsLjU1KSIvPjwvc3ZnPjwvZGl2PgogICAgPGRpdiBjbGFzcz0ib3B0LXRleHQiPjxkaXYgY2xhc3M9Im9wdC10aXRsZSBvcHQtdGl0bGUtYm9vayIgZGF0YS1pMThuPSJib29rX29ubGluZSI+Qm9vayBPbmxpbmU8L2Rpdj48ZGl2IGNsYXNzPSJvcHQtaGFuZGxlIG9wdC1oYW5kbGUtYm9vayIgZGF0YS1pMThuPSJib29rX2Zsb3ciPtCf0L7RgNC+0LTQsCDihpIg0KPRgdC70YPQs9CwIOKGkiDQnNCw0YHRgtC10YAg4oaSINCS0YDQtdC80Y88L2Rpdj48L2Rpdj4KICAgIDxzcGFuIGNsYXNzPSJvcHQtYXJyb3ciPuKGkjwvc3Bhbj4KICA8L2J1dHRvbj4KICA8ZGl2IGNsYXNzPSJkaXZpZGVyIj48c3BhbiBkYXRhLWkxOG49Im9yX2NvbnRhY3QiPm9yIGNvbnRhY3QgdXM8L3NwYW4+PC9kaXY+CiAgPGEgaHJlZj0iaHR0cHM6Ly93d3cuaW5zdGFncmFtLmNvbS9yal9ncm9vbWluZz9pZ3NoPU1XeG1kSE5xY1hGa2FuTnZiUT09IiB0YXJnZXQ9Il9ibGFuayIgY2xhc3M9Im9wdCI+CiAgICA8ZGl2IGNsYXNzPSJvcHQtaWNvbiI+PGltZyBzcmM9ImRhdGE6aW1hZ2UvcG5nO2Jhc2U2NCxpVkJPUncwS0dnb0FBQUFOU1VoRVVnQUFBS0FBQUFDZ0NBWUFBQUNMejJjdEFBQ1EwRWxFUVZSNG5PeTllYnh0VjFVbFBNWmNlNTl6NzMzOVM5OFRBZ1FTa0Y2a2tRUVJhVlRBMHNRcS9FcExSZEN5S2N1MmJPQWxZT2xYamRXb1dQYVdXZ29rS2xaSnF5QUpKUWdvU0FCcDA1UDJ2ZVIxdHpubjdMM1dITjhmYzYxekh4QVFGQlg5T0Q4dUwvZmVjOC9aWisrNVp6UEdtSE1DWDNoODRmR0Z4eGNlWDNoODRmR0Z4eGNlWDNoODRmR0Z4eGNlWDNoODRmR0Z4eGNlWDNoODRmR0Z4ei8xQi8raEQrRHo3TUVET01DTDhRSCtGUTd5WW55bkFPQXlYS1pQZU40bmZIOEZnU3VXMzEyTnE5bis2MlU0eUZOeHFpN0NSYm9TVitxVC8vWUxqLysvUG5nVkxrc0hjRWtuSERCQmZ5ODNveUFLc3F0d1ZUcUFTN29ET0dCL0grLzcrZnI0LzQwSEZNQ3JjWmtCbCtFeVhPWUVQOWtUblllVjd6ejY5SXQ4b3ovcm5QNmszV2YyZTA2RjYxU2xkTGE3N1J0TDNpMVpUMHRHRUtRYjZGN0VBbURNV1pKaGMyTHBMbWMrZkMrMzdyNWxjYy9OeDdtK3ZyTGEzZm1iRjd6K2Vyd0xXNTk4YkdIOFYrRFNCRnpxVitKSy83cy9JNThmajMvU0JsaTlpMTJCSzhvbkd0d1RUcnB3MTZYRFE1NTBRVG41ekVtYWZOa011UGhrN055eEw2K2VOZkYrZGNvSk9pWTRDUU5BR0F3ZGlrVVVKUXdDVWVBZ0RCQlFGS1lrRUJJaGQ4d3hZTUdDaFEwYmg3QjE1OTFjMzVySWJoeDhmTytIZE50SDVtbnpmVmR2dnZGOUp4NGJBVGdPMkJXNHhxN0F0WVgvaE1QMlB6a0RES083eGw2S3QyUS80YnA5dy80djNuMHF6bnJxUS9NcER6MHA3M2phNmRoemJ1YzhMM1U5ZWlhWWdCSEFnajdDV0Z3aWtKUkl1Y084T0FvQkZ5QVhCSkNKa2lnSlJsRGVVeUJoSUdBR3dXZ09LUnZwVGdsOXo0NFpZYVFMampqbW0rdDM0dGp0bStub2grL094MTQzYVBhVy96SDh3WWNCTEwyZ0lGNkJTOU9WdUxiZ241Z3gvbE14UUI3QUpla0tYUE54bnU1NzluNzFGNTA3N3YvU2M5UGVTMDhxYTArWVdIL21uakpGNXdrWmptd2w1MFFWRjVTTGxUTFNjMEVCNkJBRW9JTzVkMFlsRWwwSG1xazQ2QUJoRkVsNUZoM3l6QUlVMGtjQ3BUQ0JjSFFnREduU3lUb1RtTVJrM3FlSjRBUXlKb2FPaFFWempWamtZZU1ZWnZjZW5OejdqbzhNTjcvK1hkMTFyM252NW8wSDIyYzZnQU1kZ0g4eVlmb2Z0UUVld0FHN0dCL2c1Ymk2dEo5OTQ3Nm5YUHpRNGV5bm40ODl6enNOdXk0NnVkdTVtb3NqdThNTlE1RlRZK0dZYzRLWE1MUWt4OXFVM1o0VnJwNjlFOU16OXZqSy90MU11MVl4M1RYUlpPOFVhY2NVMW5ka24wU2F5UVFYUkJBQzRDclFVRFJ1RFJqV1I4eVB6TEU0TXJmRjhabG1kMjlnNjY0WkZ2Y3NWT1lGdnVXVVFHT0hOTzFrbHBUNlRpYVdYRFRwVTQ4dUdUYktnRHQwK0o1MUhYLzc3YnpqOS81Nzk3OWVoU000Qm9SWHZCeVgyOVc0MnZHUDJDditvelRBeTNCWnVncFhiUmNTSjJIWGo1Wm5mODNGNWRULzV4enRlL3hwWmZkT3lMRkFLZW80TEhKT1BoUzZMeXpEalhzbVdEbDl0OVllZGhwMm5INnk3VHh2ajZabjdFYTNzb0kwNmNPa0hQTDVTTTh1ejA1bGdTTUFqekFNQWtnRUpJaUVKRmloa0l3Mk1YRnFTSDB5OUNZblVMSnJuQTBZam0vaDJDM3JuQjJaWStPTzR6ajZrYVBZT2poRE9lNDBVSlBKUkdtbEV5WldiRFFSYWJwU09peTB3TTA4ZE52dGZzZnZYcGMrOE1vL1hMLzI3UUJBRUYrSHIwdi9XQTN4SDVVQkhzQUJ1d0pYcUJuZTkreCt4bVBQcy8zZmZ2NjQ3MmtYK1A1ejltcUNHWHlZczNncEpZM0RqQmtqTlVsY3VXQVgxcDV3cHZiZi94emJmY0dwVEx0WGxXREFSbWJlR3VGYldTcE9kd0V1eUNrQ2hNdEZVR0lVR3c0S0VDZ29HU1hJQUxwRFhoeEdVZzdKQlpJVVVZMFZRbWRJSGNGSndtUnRBa3VkRnNQSXhlWU05OTU0TCs3OTZHSGVjOTFoYkI2YU9VYjZwT3RwVS9QZUprcERsK20rWnRiakR0Njl1RmwzL3RINzV6Zjh6TytYMS93SkFEL0JFTXVuTzRlZmI0OS9GQVlZaG5jeGljc0xBUHk3L2MvOWlrZm1zLzdOR1hudEdTZHp6VGdXeURpTXlXMmN6MU11QzJHYXVPTUJKM1BuWTgvemZSZWZwWld6OW9CR2FqTlRHMWxsUGxMWllVeDBBTFNvTlFYSzNRR0FjQWdPZWRnYklZSWlwRGh2UldvbE1TVUNwQU9pQ0ZJazRtZTFxb1VnU0hKS2tBcEloNUNJTkVuczFpYXdxU0c3WS8zUUZ1Njk4WmpmZXQzZFBITFRNWFVEYk1lT1ZhRWpWRlFNblJGZE9xUmpPT2dILy9RMnUvbm4vc2ZtNy93dWdDTElyc0FWK01lU0kzNWVHNkFBQWxkWk03d2YzUEhsVDcyRTkvdWhNM0RLVit6MUtlWVlQQ2NXWDJRT3c4eWNvMDNPMjZmOWp6OGYrNy80Zkt5ZXRwY1kzTXZSdWZKV2hvclRTQXJONEVnV3dJdkxKWVRSQU1VVlphNFEzZ3lDQkJBa291b2xBTWpoVG9TRmhZRTZTSVNQSk9IaEFSbHZKYW4rcVlIaHN3Q1pnRUxJQ1plUUVwQldlL1E3TzZpamp0NnhvVHMrZUpmZGZkMFJiQjJhWTVKNllwcEFha0R1ck9la0s5MENOK2lPdDc0M2YvaW5mbnZyRmErSmMzZFZBaTczejNjSTUvUFdBSy9DWmFrVkY5Kzcrb1RIWE5vOTVNZk9MZnVmZXpwMjRsaForTkRMUzNhTzgrUFFidVBPeDUzRDA1NThJWGJkL3d4aEJMVStNQitmaXdVT0dHUU15eWxPbENXVUFwWnFaTzZvSUI3Y25SRGhndUNCQTJZSkJFR1JEZ2xPZXFSL2xDZ25ZQUJBRW9BY0FBc0p4aWwyaTlRUkltU0FLWHlqaC9HQ3RLaTdDVWhoOERCRHY5SnJzc09zbEtLRE54N2hUZSs1VzBkdW1SRUxWNytTbktTdllPcWQ5ZE43ZlIwMzZjYlgvM24zRnovOW1tUFh2cEVBWG5uQ2VmeDhmSHplR1dCY2dnTWtydlNIN1huWXZuODlQT2o3dmlpZCtjUG5hMysvVVlaeE5DL0RxSDRjMTFuMkdFLzUyb2ZvMU1jK0dDdjc5MURIWnh6dTNaSVBEck5FQTZUc0tybGFSWUdzbExDV0lzcGRpdFE5akZJdWg2SDlUSUJZb01CYW92S1Fva1lSSXF4U1VkQXFUQXVnVVVENE9sbDRTaGhrOFJOUUZBQUxlRnVpSWxURElJb0VKTkpCRWtaenVTREJFamhaNjhWRUhEdTJ4ZHV1TzZRN1BuQVllUTdmdFhNS0NTTThwYW4xL2NIdWJuMWdjZU12L1pmTlgvZ3hBUGZXTkZXZmo5N3c4OG9BVC9SNi8ybnZjLzdsWThlelgzS2g5dDF2NXFNR2xPSlN0MWdjMTJLMzRaU3Z1QmluUCtNUm5PNWF4WGpMTWZuNkltSWRDVGlCVXFBR0doZFVReE93Y0VOeFNYR3BxemRjL2w2TXdpSzhGWUJJMlFBNmtVbDVSUEJhY2paN0REUkdOZmFhU1M2aUdFRkFWdE5MVURKUU1GRWdXR3NiQUpWeUVaYXdEc0JFRjJRa0tiQjZhYURyRTd0cDUxdWJBMjk0M3lFY3ZPRzRrcExTU3BlS2ExakRCSjJseVVmS1IyLzZvSC9nUjM1MTg1V3ZCQUk5K0h3clVqNHZEREM4M21WR1hGMHVQT21jTTM5c2VOSi9mMEk1NSt0V3M3RFY1Y0dadXNYbWhvMnJvL1k4NXlFNDQybVB3Y3JxR3ZLaGRaVDFBUm9ERzJFUjRLQVhDTzVVcVRsL1ZoaWlDeHhFbEFvek80a2l5UVVYU0JtOGVxK2E1cEV1ZDBYS1pxVld1QkpCd2xGek9ob0ZxWlhIdFBDVXloR1FaWXppQm9BU0FCQ21KTURENXhHQU0rcWNGTjlLa0JnR0c1WUsxWXFhS0VRcEx1dUlsYldlRzVzTDNYamRRUnkrZmNhdTc1VW1rTEx5bEN1VEVUTjhLSC8wVmI4eWZmbDMzM2I0dHRzUDRFQjNKYTc4dkdGVS9zRU44QUFPV0t2WWZtVFBNNzd1eWVYY24zMTBPZTMwcmJJMURCT1lablBiS3NlNWRza0RlY2EvZklKV1Q5OEh2K1U0ODlHWmFCWUU3T1lJallWZUFCWkJEdEJGTHk2Vk1FQXZsVE1iQkJUQVBVd25qRldzS1dHRVMxZkVLd2VDZlNNOXZDSGlpZUdsQ3NrS1JNc2xHQ2haNUgyU3dFeUF0WUEyeUduaDRVeU5zSU9JWUpOcm9RM1FaUkhOUFl4VFFJVnowS3dtUEwzY1ZZcTRzak9objVydXVYTVROMXgzR01ObXdjcHF3b2ppSytPazdQR1Y2UWR4dzEydkszLytiYThiL3VEVkJzT0w0SGJsQ1hUZlovZ2dQc2VHK3c5cWdDM2t2bURmby9jOE0xLzRVeGZvMU84NGJiR0tkWnNQSXZwaGZwVGptWjJmK1crL25DYy83RUtPZHh6RGVNK21rcFBJZ0JZRkdMTTBkeWs3a1F1VncwR2hpQ3FDRjVkeVZBcWxTQnpEQUN1ZzNLcFpxQVJ6TEpvZzBHczlBSTh3S2FubWhnUVEzcXA5cStiRXdXYXg0ZHdjaHZDRzhBU0o4UndSTUtRb3JBME0ySnRSVHpNOEpRQ0FSbWY5TXpId0hwSmxXWmpMU0FJR3lhWHBOTEVBdXZValIzRDN4emJSbWFsTElrWU1PN0MyY2lnZDluZVY5L3preTJhLytsSUF3K2REU1A0SE04QTM0NUx1S2JnMlg3YnZ3b2Q5eC9qWTMvNmlmTmJEMXN1ODVFVFlNSExtaDdqbjJZL1U2ZC8rZE9zV1FyN2hIbmdwUUhacDdzU2l3TWNDRFE0VUYwYUhpc0lBTTRqaTh1b041WkUvbFNLbXJEQ1FVa3NBRU40cVloY2NCbGVGbDlFQ0t5SEk0SXhMejZpb3d5WURWcEVZQlMxRklISkFGWUJtQkloU2l4RFNsaVdCeU1EVUNaQWRHOExUMGxNUWNndjNSNlFvV0l4VUNjQVAxZFV1LzlNRkdqRGRNZEh4WXd0OTlBT0hPYzZjTzFaN2pGNThxZ2tOc3ZlTzczbnpMK0xYdnZtdStkRmJEdUNTN2twY216L25ydTB6ZlB4REdDQlZxOXdmMi8zVXIzOVd1ZDh2UE14UDNudW9iSTdvdXE1c0hZR2ZBcDcyL1YrdHZVOTRDUEtOaDZCNzU4VG8wR0tVNWhsWWlENFVZQ3pRS0dpVU1CYkFSUlhKUjlDTGhPSWhpeXFRaXNNZHBDTnl3M3JkNVlTN3E4VTVDWkFZK0Y3RlNOUUtWYkZWSlBCcWxZSW9XaGl3NkRWVWhyc3E5U1ZJeUNnblZURklnOWVJYXd3L0pnSzBtdWVSTG9YcnM0WS8xbHlTUzh5dzJRdTkyWTRKTGlORjd5YUVHL0N4RzlkNTdOQWNmUjh3ejBUZFlsVXJLeC9RQis1NGJYbjk4NjZadi8zYWFvVC9JSG5oMzZzQkhnRHNDa1JFKzZuZHovcnhyL2J6WDNyV3NGS09wM0VjNVN2RC9FNU12L3pCT3ZQRlg4L3B2Q0RmZkMrUUJSd2ZvSTBGdE1qQ1BOTkhRQXVIY29GR0Z6T2tMSG91UUJZOGszREJxNkdwbU9TMW5NMU9GbFNvQkdoVzBhNm5xNmxwQXVkcjhkVXIvY2VsRVJyekVxdUpuSkExZnhNc2dPc015dUwzSU4xcGlHQk9XaURVd2U2Uk1CbmRHTDludUZWbkpWZ0FnSVphTVV0dTlGcll0MGMxL01vWVduRFRadW82OEo1RE05eDl4eGFNUUdmTVJjbzd0R1AxWTd4bGZCdisvUHRlZnZ6cW43c0tWNlhMSytELzkvbjRlelBBQTRDOUJIQmRndTZWZi9IYy8vNGxkdTYvM3JYZ3NKR2NsalBYODExcHo0dWVnOU8rOGFuUVIrNkVianBDa2JDTkJYUjhBZC9NMENKRDh3SmxWYy9uOEtFSVdWSnhvamg4Qk9RV0tISjJ1S0lTZGhmZEtmTkNGSU1LQXNKVE5jQW9MaVFFVTZJbVU0NG9iZzdXeWlRc01mNk1EYkloS2o2OUxFSktQS2s2TDdoQllvVDFDTUFKSU1OQUlSQXBYdDRpM1F2R3hHcEJITjZ4MVNxTytqSWlRMWNUckdIQWtDWm5CY0FGdVl2OWhKelBNKzY4YlV0amRrOGRrWFAyTmE1aFp1djlxeGF2ZWVudnovN1BpLzhoS3VTL0Z3TThnQVAyRWx6cDJvYzlyeHUvN3FvbkRtZDl4UWJHbkZNaVpodFc5czV4eXN0ZWdCMlBmU0R5KzI0Qmo4K0pJd3Y0NWx4Y2pOQkdwbVlaV2hUNG9raVo5TkdsMGFIUnllenV4Y0VpSzFsQ0lkeFZ3N0RnRWMvZ0RqRExJbklpcUxpcVBXZ0FYRGpEQ0hWUlpVaGcwTGFSRndJQTZWRzkxb2hzOGpDUlNOK3M0aTZsRml5R21qNmFRTWpaTEFTUTFTUVF5UkI1SXFvbkRDZHNOSVVoMXlMWkNGbncwTFJXL2JnRFpLb0ZUNURiY29oUkJRSHNnRktBdXcvT05KOFhtUkZER2JWWHErTW0xbGRlVTk3OEgzNW45b3AvVnozaDM1dXk1dS9jQUEvZ2dQMEVYdUtQMkhuNnlUK2hKNzc2Q2VPWmp6dU8rWno5WkRwdUhvWWV1S0l6ZitQN01OMjkwOHBmM1FxWFpFZTNpRU16bFBrQzNCcWh6UUtmWldEaDBsaW9rU2lqNEdNUnNzQXhKRk1NMEptbFFDcUtURjBoTHZBU1ZXOFM2QjVLdmdheXFiRFdHa2FLS05YaFZYUVpja0lHTXNKZ1JHdEdPcW5nZStVRWtpOXhRVWh1VUF2cGhKdmdWYkFnd0VpcitXTU54V0FVUEZINW9vSFNuZ2pDV3IwVU9LRjNkQWx1NWhVRWorTmhSUUNRdEtROVFxMGRIcktTTC9jY0hyaTFsWVBnS1RQdHdGb1pmWno4YVg3YmYvNjV4YS84NE4rbkovdzdOY0JsMkFWMi8rSEsxNzd4VWovN3NmZjQ1cXhmWFprTTYzZVNUenFINS96Njl5bk5Nc3BINzZER0FxNXZRVWRuMUwwTFlUNVFtd08wNWREY29VV0VYR1dqWjRmR0lzOE9LeEN5ckJTWGlsQ3kwSFJRTHBNWU9abUt3K1pPRGFWZW5HQkx2SlFxZXhGOENja0lLazFYRUtVSTNDenlyaWdtY3J2RU5BSVdJZHVNN0UyS3EwMTBpV1FuN3lpWGhhVExBS0lMUFFMRDJ6SjFSQ0xRbVdDa0I3TkNyK1dHQXVRR3pBU1ppUXdQSEo0eWloMElqZ1RGRzZCUmZRbzNDMEd3RklkMjlOaUFqWTBzODlHekZVN0xTaWJMOUMvd1oxZitwL1ZmdnVMdnl3ai96Z3p3QUdBdkJmMU03TnIvY3p1ZTh1cW5qdWM4L25EWkhOSjBwUisyYnNma0dRL0c2Yi82L2JCYjcyVTVlRlE4UHFjMjVzTEdIRG8yQjQ0dEl1eHVEdEpjMU9EUVBNTUh3Yk9nd2FIUlVUTEU0cEE3SkFNOGdkbUl4WWd5RE1pTExXa3hDc1ZaZGhTM2xVSjBob0prMEFSbUxrNG5zRjJySENjbUpZUHRYNVh0N0Uwcm5RZXJJWklHN0p3Z3JhMEFJMWhLa1ZZRm1La01oV1hoc0VKcExCeVB6SUhqbzdodyt2RUJaVVBNWlJUR1VqMHhaY2lBT2lvWjZKMUlXSjZOeXVzZzBZdm9wRWtpVjFlUUpsT2tycE42TXhsVVJMcEtGQ01NZXdNQW1RRTFONDI4TUJRMlJJSUNKeEl0U2lDbmMzT3phSE56cnB3Rk44ZXFKdGt3bTE1Yi91ekF6Mjc5eGtzK29URDVPMEZxL2s0TWNGbnRuc2ExMXg3L21qYzhVMmMvOGU2OE1lU3VtNVQ1bmQ0LzdTSTcvVmYvTFhURFhkTGhEWEEyUUlkbjhJMDVzREVITmhiQXhramZ5STVacGdiUjUwVStaR2h3K2lqNDRFSVdwRVFYbkpzWlB0K2tIenVPMUdYNHFWTm94dzcyano5YjNZTk9LOWk1MTlLNUo2Rzc4R1RheVR2QmFRZExrNkRLWUxLdWg1TEZ0eDFDTDFCVC9SUE8xU2Vlci9aN2Z0TDNnc0Voenc2VVFNYTlGTWtWMzFOU2RwUkZBUmNGa0hOK3kzSE1Qbnd2OC9HWlJOZDQ5M0ZzdlBWZWpuZHRFTVdWN3hDWU9uTG5HakNkaU5PT1BoYklpbWptbmdMSFVmWFFNZ3RjeUt1VUVZQ2ltVThlTjVXR01lUDRlb1lYR1ZCODFkTllPSisreHE4OThPdGJMMzlKODRSQ1pYbyt4N2J5ZDJHQUZzZ1p1MWV2UGZ0MVR4L1BlZW9SYmMxSzE2M001bmR5K3BRSDYvUmYvMTdZRFhjakg5NEU1d3ZoK0F3NHVxRFdaOURtUUd3TXdFYUdiNDNTUEZNTHdCZUtpbmNzOUVHQWtud1V5dEYxK1BGN3dWTVR1Z3ZQWlBlc0I2bDd3UG1ZUFBuQlRIdDNDWk1FZlB5ZFN3ZG85M1UzdTBNRTNXc0s1WTJXalpmZ1VqTlZBMXdFNWlxRHFUODVvVmFnR1JWSXpOSjdOTXpPV3BteGJiU3R3bW12SDg4dmdJOEQ4OEVOSEx2Mk5teGRmMUJIWC9jeGJ0NTZGT1Z1YWJwak43cjlxM0JHMzB1RVgxdXlOTlhpZ2c2c2pBd3JvK2oxNkJmWnNiRXhocXdiMmRkODZsdlk3UCtQdi9IN1hqRjcxWCs5REpkTnJzYlZJLzRSZU1BR011c1ZLOC84dmEvbEJWOXplTnlZWXpycEZwdDNHaDkvTHMvODdSOERiN21MZnMrNnVEbVhIOThranMzQjR5TjlZd0ZzRHREbUtHMDVzQmhadGdvd0NENElHa1pKSFh3aGxrTUhWU2JIWVk4NlJ5di83SEdjUHVPeDdCOXd4dklFQ1hFKzZWbkIrMGFSNklrRWpkV0M4SEg5d2lFeXFJQmdzQnFpWEE3U3JOSVRDQ2F1Q2hMa3JtUVdsYWMxNDNFWURGNy9qUkJPYnozRWNtZXpNQUd3V3Y5RUswQWNmYk5NTU1uUzBpRFpqbTN6K3J0eDhEVWZ3c0hmK1pEbTd6K3VaTHN3MmJjVGhXQnh0THd3U25RbEZIaXdObVF0NXFVU2gwd1hNQnN6RnZNUm5jdEhaSzJXRlIxT2gvdmZ3MnVmOTlyamIzejVDL0NDL3Bmd1N4bWZ6MXp3WCtEUi9XUHdydkhscTAvNTVhL3hDNSsvWHVaRFNkYjU0b2o4b2J0NStsVXZZcnJ0Q1AzSUJyUStFNDdPZ2ZVdGFIME9iR1pxTmtKYkkzeVd4Vm1tWmdVK2QyazkweGNBMENQZmV6ZG85eUJkL25oTXYvVXIwRDNoSVZoNnMreVFPeHZ6WUdhc0xFY0ZQYXg1bzBxWjFRTDB4Q0JiZ2I5V21IcTFMTUJodHEyWHFyUWI1SktSZEZHQ2d3elZOT0NnVlVPUGgzeUpIRHRnb1FpVVJKZHE5R2VqWGVUbWdCc0FSZnVuZ3VXSmU4YkFWRjlVam52ZTlHSGQrbk4vZ1kzWDM0RysyNGQrL3k2TmVhU0xOUzlNRkJSZTBlcDlHUXlQRTJBUnJMaGpudDN6RUhYSFdFYnM1bzU4YTNjYlhydDQzZGUrYnV2YVYvOWRjTWVmTXdOc3dvS3I5MzM1dDMzbGNNNHZEUXNmWm9aZTQ1YksyUXVjOW5zdlpYOXNKTzY2Vno0ZmdjTXo2dmpNdFRFak5rWnE3dkRaSU0wek5DdmtWb2JteFgyZXdWbG41ZWhjNWZoTjBKZWZ4NTB2K1diMGo3OElRRnhWWlplbFRwNHFYTnRpWnZ3ZkJmZjJqUVU1NjlYR1FxbnFrRnVWaENLNFl3dXJKdDBkWm5KM3MvaGhOV0NnYXFhcUtyb1dxdUV1RzdmVzJOb1FvWXBlRk0xMFZRd0RSeGlhQlRXaU9EQXduS3VyR24xOVkwcnRYZ0s4REpuV0pZWUdGamo4cHV0NXc0LzlDVGJlc2VFN1RqOFZCVTVIR0tFVTJPVVNkNmRZcW9CTUFuS0lOVGdiU25Eajdpb29lVTlhUzlmamhtTy81YTk5M0hYcjcvem9pL0ZpKzF6Mm0zeE9ETEFaMzB2M1BlNUozemhlOElaVEZ0TzB3ZEZHbFRTa3UzajZLMStzbGYxN3pUOXloNUFML09nbTB0RUZ5L3BNMmh5Z3pRd3VITDQxVWd1WHp6TnRxNkRNUjFGVCtoMTNZZHg1REtzLytRMWFlZjVYQVFBMUZzQUlKcE83aDA0RUZBeFJKaVkyanlPRHgwUURCQUMzREx1UmRWVnlMSVNtN2I4ZENDRGFJaGRVR0Jab0VSM0RPQ05pR29CaWpoUXZLSy9UT2t5TnZ3UENOcmxkMUJnYVVndzRhaGNUQUZJdXRSNnArQnVLVmhtVlpvSE5ienNBNVBpZzFobktMT09HbjNvemJ2MlA3OGJPMWRPa3RRbmQ0NDBMSWpJMGJqdFg1TjBobElvdnpjZWlvU2drYWNqRldNWTl0bnZsTFhybm4vK0hZLy90RWtGenRnLzBPWGo4clNjekhRRHNNbHpsLzNMWG1TZDk1WERXVlNjdkppdnJ0cEFzMlh5OEhmdXUvQlpNempvRjVmbzdoTmtBSFY0bkQyL0NqMjFLeHhmQytnTFlXTUEzUjJCV29Oa0liSTRxVzZOUWVvNDMzaUE5dXNmdXYvaXZXSG4rVnhIakNPVUNkQWJhOXVHTEViR3FhTGdLM2NScWZKUkVlZ0ZMa1pkQ2VJYVBHY2hGTElMbjRKWlJITXBPeXc2NkUyT0Jja0Z5Z0hJcXV6d0RLSUxIQUpqSXJqSlFja0hKVHJqVHhzQWs0UUt5b094Q0tVQXBSQ25BV09MN29ZUjZ1d2hlRk85ZFhDb2V5dTZBRTlXS0YvTGpyM3NRZElFdGxxR0FuZUZCTDNrYUh2cTdYODM1Nm1Ia1l4dGdTbkF2cldKU3E1N3ErQWZSMitzS0V5TzZTQWRDOUVQWmhtL01uNEJIUGZiNXU1LzNTd1IxQUFmUzM5WnUydU52NndINVpseVNub0pyeTJ2MlBQTVZUNStmZHZuaFBCdHNPa21McmRzNS9aWXY1MG5mOVJ5TTc3a2UzTXJnMFJtMU1SUFdGOUR4QWRvWWdGbUc1a1VhUU44YW9ZV1RXNE5zWVJ6dStoRHMrVS9TMnMvOEc5anFoQm95MktkbERicThCd1Y1VlRsRm91T29TaGdLRVFEVHBJOW5mdkpuNW4zOGZPbGRERTJYaXByQ2IxZXBIbjJYSHdkTnBQcDgyNjVzNy9PMTcrUG5ud1R4ZUFHVVBTSXZnWlJTVkUxTGhwaTE1RUdqbENzWTcrcW1IZFp2T016cnZ2S1Y0cTFtM2U2ZHlPTUk3enNvMGQxRnI3eXlTeXdJRGFRZ0RTNE1JMWhVaXBuTFhiYWl2cmh0OXE4ZC92aTcvK2ZzRDM3dXFzOVJzOVBmeWdDYnB1OC83WG44OTMvbjRrSC9lVDdPeDJ6c05CNFZIbkVtVC8yNWY0M3k0ZHVGelFWeGZFYXR6NFdOQVZwZkVNZEhhS3ZsZkE0dEFOOGNvY0hKc1JQdnVvN3B4NTZEeVUrOE1Fd3RPOUZaRlNuVkUrOHVzb3J6WUlLWGdEYTZ0TFFKQjJDTGhTOCtkRGZMclllSWNndytacVJqTTVUYk4xbnUzUUR1V0hjY1cxakpHUnJuMEFnblNsd1VtTE5Bcm1Jd0VTWlhTUVNNbVNXQW1rckh5YnFxcFVra0U0VXVHSktVaUs2SDBEdFRKM2FkYVJlQVUxYlJuN3NiM1JtNzBLMU5pTDZIVGRkZzNRUnA3MFFyRHo4RHRqclpQdUdPS0xLTXFLMHJ0ZXNZN2g0dHAxNWRtUzh5dW1uSDJjMUg4WjR2KzAzZ3ppblRqalVNeWVGR2VRM0RMc3FETXFkQUZna09hQ3pnd290QzRBTVVaTzdMYStWanV0bC9YNi81MGorZXZlT2RuNHQ4c1B1Yi9xRUFNN3dsLzh1MUJ6enlxeFpuWGRrUGVUaVdjdGNKR0hjVW5QenZMcGZmZkpBNnZnbHNEdUw2Z3R5WXc5Y0hZbk1RdGpLMU9VS0xMRzA1TlFjeEh3Vk81WGQ5R04zM1BnM1RuM2doTUdRaEdaRmF3c1lsTEdkV2M2NmNCUU90U3lDZ2N0ZHhMZDcrTHROZmZOUVhiL3dnOHZwUnBvOXRnZXVqUmhRS2hoNkpRcEtRUUhRa0VvU2tHR0Znakw2MUtwcUNWUjJMSVp4TllDV0VWU3d2MVI3MndrZ2FDNEFNWUFDQUtwMGlGTE04NEVBMDIwSGNnaXVIc2N0QWVzVnl1aFdqUFdDbitwUDNZL2V6TDhES3c4N2xya3N2Z0hVUi9WUTh2REVKdDZoN3ZBcHZTYU5OTytSRjF1cjk5dktMWHYvUCthNHYvWFdralFUdW5vZ1FEVlJseEVrQXBnb2F4ZGxsUjZDa3dNd2REcE5oblJzNkgrZE9MdUVUWHZaSGVQc1RMOFlIQ3Y2V0RNbmYxQU5TMFVTa04reDQ1bHUvWW5iU2w5eUxyZEc3dnN2RFFheDl6MWRqejVjL1F1TUhiaWZHRVR3MkZ6Y0grdFpJYkM2azlSSGFjaXJnRnZoV2dTOUVxSWZmZFNmc3VXZGk3ZmYvWDJBc1lLckFTZTJ6aU5LMDZ1OWNjTG1zUzNISlgvczI1dCsrUnZrdDc0WGZ0ZzVpVmNSdU9qcDBOb0g2Sk5Db0VqV0hvM2FWZXpPTHlCcWphQ1pncVNvQTZ2OUM1WUxXSUNRU1ZZRktJTUZUQ25rWExZZ0h5VndHWnhSTGdlNUJ6aGp6NFI1aUNLZUZHVlMxalZoSm5xM0NyRG1nbVFBaFBXUVg5L3lMUjJMLzF6OEtxdzg2T1o0N0ZuaWZ4Rm8xR1F3bFVHNGFvVElVcEVuaVBXLytNTjczOU4vRDZ0NXpOS1lTb3RZSXdaUkNyRkFrdUJNQ1dWd29uWHlRVUxJRGN2UXFNTGRoQmQzS3EvMVBYdklMcy85MTRHK3JJL3diR1dDTC96KzM0N0ZYUGw4UGZIRWU1a014NjRaaEU3am9aT3gveWZOUXJyOExPTDRnNXdPME9RQmJHWmlQME1ZSWJCVm9ub0Y1a1crTzFNekJ1VEZ2emVXbkhzUE9kL3dQcGxOMlIxYVRiSG1NeTZSSkFISUIrdWlybVAzUC8rUGxaMStKOHU2N1NleUZZUzlzdWhxcTRpbzRnR0w2UWFyd3N6Y2NCTkVyQ1VUL0wwUkZNOUR5OTJnQWJoMjNJYS9pNlRCanVwTVdLcGY0R3pjQ2J2Rkxoc0N2RUlKWjlCalRRcHdZVVMrS1lEZXF5aGhESjBnUFJYUWlMUUZPK05ZY1k5NmtuelJxNTNNdXdKay8rSlZjZmZCcEtEa0ROTEcyb1BneWNmVTQ4T0t3UHVtR2wxekwydys4QTZ0bm51NTV5Q2d3bEVpYjZRSUxhdm9jdHpWQmVhWTBMMjRxSGpKYnVxOXBpa04rVjM3NStIdGYvS2J4WGU5N01XQlhmdllOVGdEK0JsWHdBY0MrSGxlWGI5MS8za08rek03OC9qNlhrcTEwVkNZbUEzWi95MVBoZHh5QjF1Zms1Z3phbkFOYkF6QWJnYzBCbUkzU2ZBQVdHVDRiaUNHRFl3YWM4c1d0V1AzRjcwVTZiVzlVYkJGMnRXMSt0V0UzakEvNSt0dXg4ZHgvaS9rMy8yZnczWTYrZnlEUzVDUXdBUm8yeFhHTEtpTllpdE5qQ2dZZGhMdE1TMldwVEY1NSt2Z0Jhek80VU9FWDFBWVFWRGlYallDTEVzY1VnREpWMnoxYnQzQWdidkdtQkFNS2RxOXFWYkNTZUZaMTFsR3FSd0pBaVhRSFNwR0dnY29MWURWaHNtOC9WcmRPNXVhdlhZK1BmTWwvMDkwL2Z5MVMxMFdxSUtjYkNBZHBIc29aQWt3R0x3WDMvOUVuWWZMRk81WHZQVWFyZ2pEQXE0N1JsV3FJTVVWSWhzUUVzWThHQTNoTVc3Sk56WFYyT212bHNkTkgvNzhDN0dKY2RWODgrZCtOQVY2TXl5Z0F6MWljK3g4ZXNsamJzYTY1WUIySDhiQ216M29NdTcxcnhOM0hZRU9Xenhid3pRVjhhd0MyUm1BclE3T1JXaFJnbnNsRm9SWVpVZzgvY2h2cytWK0t5ZE1mQnl3R0lDWGd4TUhocW9NeDhpajBTY00xZjY3TkovNHJwUDk5aTArbUR5VnRqY3d6Y0J5a05seEljb3ZPZEVwdDBxMTdjQnhGZFM0SGdSQ3BhbWswQWErMGJpV0c5Rm0xamE3RlpBbVNvVFp3VnBDT2F2MlZZWVRSdlFRUkFjWkZSM3VVNTVITmlncTVLa016M1pEcmhpQ0tqR09JeW40UmZhVnJPMC9EeXVJVTNQR2RyOVpOMy81eUlsbmNEclc3TDFnVUJNQm9ZUi9za2g3MDM1K2xoUjJGbFcwS1NGWkJjNEJKYUpncUNDZ1I2QWltNkI1QS9XM2EwR0o4dUYvNHJLL1o4YlRMdng2WGw2dHcyZDhJMHZ1cy91Z3FYSmErSGxlWG45bjFtR2Mrc1p6eTFiT3lHRVVrWDJ4SVorN2o2bE1lQnYvWVlkQmx2ckVnTmtkd2xzSFpJRzBOMER3VGkwek1NM3lScFlVN1I0R2JNd3g3Wmx4NTBiY0dqR0l0N0M2dlkveG56bFRmYy82YXQyRDJyTy9ENnNIVGFaUFRvY1dtb0V4dmhGUHJRa2MxcExCR0tqeFVlOUVvbnVGTDZVcHRNVnJxU2JoOGMxWFVvODRYOG9ZMFJuZWs0RnpxRVNKWmpaK3E4ZytSWkNHVXJTRnBiYW5FMGxQV1IzUjZ4Q3l0UU95cXg2MkdIVDRMeUhtZzZOaXgrendjLzhYcmROUHpmbGxOd0ErNFpEVzVyUUNpSllObngvNHZ1Wi90Ly9vSFluYjRIcVZFV0VoMllmV3pBQjd5aGFab2RLQUQxWE41NG1Ba0ZocTRSenZ6bC9oRERnaFkreXRjZEY4UTErZlVBSGtaTHBMT3c4cERzZm9menlDMHdOd29SOVlHZHp6bjBlQ3drQi9maEI5Zmx6YTJ3RVVtNXlNeEg2aHhwSVpCV2d6Q2ZFRnRGV0F4UXVyZ1d4L1R5amRmcW5UMnFZNlNnYjZMYmlKSG13WVFCVW5Yb2J6bDNacGYvdVBzNWc5a1RoT000d2FOVHFySTNBazRyWDdWZmtZUzBiRk9GaGhFMGhHa2lVaVVTTC9sTkJRa1pJWlJaaElacEFQSTRTMVIyK3JnOVQwRndoV3R1MUg1VWxtTXBtUWtiVDhIeWhWT0xtUzA2UkhLa0RMb0RxcUFLS0xIK3poR3VvK2hKNWdrYUdKQW44Q1VhcXU3aTNMNnVJV1ZQYWRoNCtVZnhFMy8ramVSdWdTNDF5cWlqcUNwTndjalBPdjhIN3JFZmZXNE9CdGhYdEFWZVNvRm5YdE1IWWJYR1JHRnFncUpEa0owS1F2Rm5YVGFISXY4aU83Q0J6OS83OWNjdUJKWCt0L0VDMzdHZjNBVkxqUGlTdi9QeHkvKzVrZGkzMFBuUG81TVpubmNoRjF3R2ljUFBvZCsrMkZ5R0tDTkdXdzJTTXV2RVpqVlBIQ2VVZk0rZWk3VTFvQzhhN1RwOS94ekdFQ1BLVmJSTW1FZVByQTRrQXo1SXgvRDhIVS93TW5zbkdqQUtJTk1yQzA1Y1ZHaWdhTGV6VFdYYXptY1NhZ250QnBOSlhiUkJzZDQ5WHcxNzJzd3M3Zm5pcFNIQWF0ZTJPWHJoQkFxbEFwcS95SzQ0UkxxTExtaXpTTTAyZ1NRWWpwUkdJaHE2N2s3alVDU0lXOGR4M2owRG83SDc4UjQvQTZONjRlVnNtREpnRkprRG1vK3c5cmFPVGo2UzMvQjIzN3F0U0ZlclJoTGVOeDJ0WWxTQ25ZOTVBdzcrU3Nmd0hIak9GSXlJRmppT1A1QXB0RzhJRlhicnlEMk1MQmxEZ2FPekowVmxFZmxoM3pudVhzZWRQNWx1S294azU5YkExVDFmdWZ0d2Q0djlwTi9lUGRvUG1lRzNEVml6cFduUGxTNGQwT1lEY0RXZ3RnWW9hMlJtZzNRMWdETlJtSnJCR2FaR0FxVVhjcWpBRU5aSElTZThoRG4rV2RUSlJNcGJkT25sWU9QaUZvd2Z2T1B3ZzVOeExRQ2xiRStyWVhiMWtmajhrajIyNTBzQld0OFFqR0RxRGZocWpHUlMvS3VHbUo5NHpCc3F4TFBoZ010cTZIUUhsZGdCcldybUtsZGRWL21nU0E4aXVuMk10R1dqb1lFMWRjbjVXSkswbXlMVzNZcko4ODltYWYrM0ROeDFxOWN4ak4vNGJuYys5MFB3N2gyR0huekNLMmJBRjdpRml4ejdsZzVSd2V2ZkJNMjNuY0hVNThvRjR4VWhhMmFpb2NRZU1vM1BnTEYxbUdLWVlWV3p4M2hTSWh4WHdsTnZ4cWVQSUhxWUdoQ25BN2twcy9MUlg3T2ptZm1SLzRvUVYwVkJjbm4xZ0JSdmQrUDQ1RXZlSVR2UFcvbVF6WmpLbm1ML2ZtbllIck95ZlNEeDZERkNHd013R3lnNWlPNE5STHpEQ3l5c0JpRkhJSlNqRG04aW9qTXcraSs0ZGxWbzFGQnFSTU1YMTdBUGlILzZxdWd0OTBBVHMraDhpeWd3U282cnhWeURaM2JCaFNCdHVaVTBSWlVjN3I2TjNJS1pmbGFZSXdocXMrcHIxK2FjWUVNbzZZS2llelJpVmtVY0VmMWRoQmM5WGxMekh4WncxZjBVSzBGdm5ydVp0UU9wTVN5ZFFSNk5IWEJuLzRnTDNqVjkrR1U3M3lxOW4zcjQ3VC9oVS9FT1Q5ek9lLy83aC9BNUd0T3hUQzdrNWI2YVBkekJ3eVlMbFowMjQrL0doQ1lXTHVNNjRPa0VIYXBrNTU2SVNZWDdLUnZ6cW95Tmtab0dyWTdTQ2pCUEZ4ZXZGQmh4eWhNV3BkME1hWFJTM21jTG5qZXhhYzg0SUxMY0psL050dWZQcE1uRXJqYXo5MnpaOTlqRjN1L1o5ZGlWTkVpSVdjRHRyRGppODhYRDY4RFd6TmdjeTV0TFZ5TGhUQ2ZBNHNSSEVaeUdJR2NxVHdDT1F0am5RbzUzNFNkdVV0cnozcFNnTHBtYW1xVXFPUUVNaWtmM2NEc3BiOGxzN05SeHZVNk1DREQ0bzYxYWpScUZ6Z0ZGMXR6UEptaHlKaFZEUXFHQWxRRUxHRGlySnBIMGhDemMxdk9aOGdnUnFIbWhNYmMvcDdtMVhNb1BJY3BBejZLcFFnMVZNZlBCY3JOVU1UUXBJVHpWSUhKRWZvWHA3cWtNbXhLRDBvNC8vVS9naDJQdnIvN2tPV0xFV1VvS0VPUmNzSEsvVTdXQmIvL3ZWcDU5amxhTE82RmRZWmtFSE5tdjdxWFczLzRmcXkvL2FQT2xPUk41bEtoSWpPRFNsYTNPc1crNXo0SWk5bFJkQVpIcWJtdEhDeFpTUVdHd29TQ1RsS1NhTzVJWHRnTFNuSzZNM1hlYzFPakx1RFphMCtiUGVaN0NlcGlYUHdaZThHLzFnRGZqRXNTQWYxZ1B2dWJIbVFyWjgwOUY4Sk00d0oyMG03MnArNm43dGtFaGtKc0RlWmJDMkkyZ1BNeEZNeGpkZ3lGekI3NFhYYlFSV09TNTd2UWZkVVgwM2F1QWptREJNM01HMlZBTDRBQitYZGVoKzcyUTZTdHdkd3JmaFVXMnFwVVF3TktnNEtJVWxYYm8vbHFTaFNhditvVjJiemhOallTelpFbGNNSVlOTWhvOVM2eVJJQkdzdzZwbTVCZEo3UEVaSW5zRXBnU2FFYXI0UnZ5NmptcVp3NEVPOFlkb2MxYWlPbzJlZEJkWTNjSVovN0t2MkovOGs3bXhVQW1BL3VPMWlXbWprQktVSTZHNDdQKzQ3K0FKbHZPTEFnZWFsWTZldTNBWFQvL3Btb0UxZjh5WW43YzM4Rmg3bi82UmZMVkJaakY1UEpXTkJrUTV6Z0tkS3NobWdZaHhXZGJhc29scDdNa0t5Z1B3WG5mZE5xZTArNTNHUzZ2Vk5MZjNnQjVLYTUxbkkzVmgydnZDMWRMOGdXeWdVRFdYSk9MemhFWEpZRG1yVUdZalVxTFRNeEgrRHhEWTRFV21UNW1hWFJnZEtvVXVJcUtCT2VjZk93aldvU3FKVnNMWndDU1FXTWhmdTBQa0hneXFFVkYyNVlRaUNuWVNxR0dXeW1yaGxyVzVWbU9Oc01LZ2ZLekZpemN6dnNpMFVTMFp4SXlzbFNMblFqOVJIS1h4bldVY2pjVzVXWmZqRGRwTWQ2TWViNVZpM3lyajR2YjRlTlJJQzlBa3V3bVlPb0QvVmFwWGxKcXVXbTdjYUsxUFlaamxjVXhUcDl3Rm5aODZZT2hYR0RUU2ZqeWFQTllxZ1RaR1ZRY0t4ZWVpcDNQUEo5bFdBY3N4U2NaUjZ5a2ZWei8zeC9HMWswSFlWM1gwdHRBZm9oQXVnSHNldVFaN0hjSVBndzBPT2kxZGI0K1VoamM4bnhGMTNRUWhyRytMRWpMUk9LWXR2ekNmT2F1WitCeDMwWkFWK09xenlnTWYxb3h3bFYxYU9UTDFpOSs1a080OXVCNW1XZEFxWlFNOVluZFdYdmhoemRpU05EbUdQVGFJa3REQmtkSm80Z1NZd29VYUVmazR1N3dNb2V0ZFQ1OXlpUHJsUHJsZ0o5YW1FYmxXOTcxZnBSMzNZQ0puUS8zZ1VCU2JHaERPNnZWV2tzdFhsb25iL3hheUt6eTQ5b3M1VFhycDhXY1ZHK2tXeFRMSkpDbWthcVZvd0tPc1pRUjVkU1RwQWVlRE80OUQvMzlUNmVkdXcvRm5WclA4bGt4LzlqZHdzRWpISTl1QVIrK1Raak5rYkFMaHBPcGZnZElWWVdzdzRVWWxCb0ljUnZNQm1HdWJ1L3BnZWJKZzlPek5qSXBhcDJtaDVZN2FPRGtvak8xOVFjM29lT3VvUEVnWUFxa2RjYzlyM3FuenYyK3I0SzcwMnlwNGFjZ2xleFkyYlVUTzU5MHBtYi8rekQ3bmJzSkw3TGFWRlZxWDc0TFZhQkFGYzhraUM0QTBSQzB4aDNFRVpsRU56N0t6NzM4Ti9iaFAxNTI1TExqK0F5RUNwL1dBQytyNE9LanRPZmZuQ3pYT3NjSUUyWHUvZjFQUW1jcCtmRU5zQWljRjJvczBwakJYSVFpc3pxQm5uS3dpUEoyeHh2NlBPUGlwQWw0eW1sUm1aR0l1YmJWSEhLaGttRjg3VFdSTXlVQW95bytCNlI2b3dxeGpvMlZ5b1ZJQ3c1VVdzb3NHelhTQnRwN2hmK3JHclBpUGt5cllDa2M4czBxV01BZWZUYW1UM3lXVnA3OUpQQUJEN0QrdkpQYXFXa0d2dlRlN2QrU1hlVUR0NXNmdW8yTFY3OUg4emUvVjM3ZFIxbXdFNzJkb3RKUGFHV000MURjREJMQkpIUnc5TjFxSlVJQ3cwa2czSTFtOWVaYWp2UlZCWVduTmYwSVU0NUFYTlJqaXVHYUQwSGYrNVdrTVdUUGhtaTZra2R3bUFJckQ5MkgyUi9jcWc2N0xMTG8ydWNTUXdkRFpBTlVnWGVVKzY1U1cxS0VFWTRpc0RPbWpIRjJJVTUvd0dYK0ZWOUw4TmMra3g2U1QybUFWd0dKdUxLOGFQZlpqejNUSjE4NnFvaUFsUWdacWIvZlNjTG1BQ3lxdW5jb3dPREVrT05uZ2JIR1FKTGNJUHhndTJDZyt5YnNTeTRrZDYyNHhneDJxYW5UWTVaRUY3N01QblNEZ0RYQ005am9zb3FxMUEwS3k1S01RTXpZYTU2Z0luZHNwUTNWWUFpMUJpS3FBS2tEMUttVW16bXVEZWllODBTc3ZPQjU2Si80YUtRK0xyZ0RydXdmMTgzWlpuU293VEV1cE5RamZkRTVBczdCeWxNZnoxMXp4L0IvMzQxanYvUTZESy83TUczVFpPbDBvTytJY1FEWlVSS1NGd2dUbE9PSFNSSVdveWNSTStXcUNEK0dYcUZPOHdnYlh0OWdoeDZRWkRHYm44cFp2ZTNtNXR0dTFXTGp1RloyNzJrME0walNrT1RtQnNEM1BQcitkaVM5aThqQitUVjViYlFDQm54cWlESGFyTGRBOWNlSW9FK1VrSUJoYmtNNjJWZjlZVHJ2RzY0R2Z2TXFYT1IvWFNMNEtlUDBLYmlFQXZpSTdwU3ZQeGM5TThybzBXbE43cHdvN2RtaGNtd0xHRWY0WW9qWmZUbUx1U2lLalZJbDdvSzdROFdscXBvTGlmazZlUDZwWVNVZVVBaXNwc2R5SVNYNDVnejVyUjhHc0JQeXVpZXJmaW1HTlFmU1dnSFRodHFIVllJeGpTV3l6U1VJRXZrWFF3aFFnSDRGcGF4ejV0Y0IzL2hZN2Y3TFYyTG43L3cwVnk1OU5KZ0t5ampJYzQ1eEhSM0JMZ1ZQblZLSUpWSkNTaDFrUm5WOVRQRXRUaDh6TkJSd2FsaDUybU4wMnRVdjBpbC8rWk9ZZnNlam1kTk53UHlnckorR0Zoa09Md1dwVzhQNDlwczF2L2tPc0VzeGFLYTFIWnk0K3RNQmRFU2VEOXg4emZ2VmNZZVFNd0N2azVoR3NqTmdNNk44OUY3VWt4SHUyaEU0YWMxNGRqN2lkUGhraExsa2RZSVRhOEVSNldMa3JVQ29KMDJodms2QTZGSGhJN3dpak9DOERPVUMzL2VrTTA4Ky8vNE1zZXFuelFVLzFTLzVaVEUxTTUzai9YT0VqSUtjU01FMUtwMjhpOHhPelJiQ01JRGpTT1ZNamdFVDBFV1ZFcFNRbCtpYmlLcVBnYUVCam9WODV5azFvU1ZPMEFJMG1oU2Fid0liR3lRbklFYW9UazFXalh5cEdsSURsZXVaVmdPa2F6TEY2dXJRT3MxUlZRUHNwOUw0WWZqRjFJN1gvangyL3NaUEl6M29QR2tZb0Z3aWIwb2RRVU5LeXhiTHRnTXBYQU1jb3BoQ3ZDckFZY2xnZlFkT1VvQit4WmtYbWYwRHo4WEpQLzl2ZU1wYmZ3TDhpcE14ekQ4Q1NraUpJQXVZd0hRY1BIemdxcWlQNVNoRGpnSFU4dmhJeGVIakNKcnA4TS8vRWZ5R0kwalRGVUVGMngrVW9nR2NPN2IrL05ZNDZCaVpGZTBqWmt0QTNkSUtqRW11VEZVUmx0VUtQVVU4RGp3d1BHeUZvd0t1U1pRU0hWMzREQUdHQVo3UDEvN0pOd3lQdXh4WTdteis3QXp3cXBwMy9xZnB1VTgrdC9UM0p3WlBoa1NQV1YvZC9sWHAyQmFZUjZKNlBKWU1sZ3dVaDVkUnRkaVFwRG9leUJuSklnUVVHb1QrOUxOcUtXaENUR3IwU3FDQmdQejJ3OEk0WHliR1JnZVkxUURueXFleDdkZUlBZ01XZksrVzR1bjYzTGdRVU9DTkpQTDRIdmJmOVV6c2VPZXJNSDNtSlVUTzlGS0F5VVJMQlRZQVkwdkQ2d2xTM0FTcTQ5MUVoUVpQaFpJQ2U2dnFxa2lnREdtU3lESExGOWxYSG5NaFRuL0RUMkxQTHp3UDQrUjZZRmhINmliZ3NPV1Q3aFF0ZnZNOU9QU1MzNFoxWFVtVHJxR2FoQ2gyQnB2Mk9QeUthM0RrUmEvQ2FuOGFsUmVSM3BsSUUxUElXMlJJR0EvR3BsY3ZVYzNHbk1SdGR6clpzMXZUOC9aSWVRdEpZWHdtdWFHQUt1d0VUNUppZGJlWUFDVVVVSm1CRTBwZFVJc3NSVmFndEtLRWh3NzdudzBnWFlFcm1tcjZNemZBVTNBSkNlQkMyLzNVMHppeFFWNmlOOXFsUHFGYm13RHpJYnE1aGt6bUFzc09GU2R5RmowVUZqRUZzbmhBRUVCcjR4R2dRb0JuN2lHQUFEeUFxQ3hzeVlaUU45ekJiamJXRHAvR2NxZ0N4YlU0UkdqclJLK2NiSUZIc2luU0dUUEhtLzZ1TUZtQytSeU9ENm43bFJlci85bWZnSzFPR1oxMkhXZ3BTdUpRQlRMU2IyeGZ0Q2dEcWtGR3N0NmtUNGtKaVlraDRLOUZUeDBFR0FDYkFYMEhIOFBROTd6d09UenRqMTRLUFdBZFpYRUlObGtseWh3cjNSbFlQL0E2M3Y1Vkw3V05QL3dMbFBVTjVzVUNaWnhoODgzdjAyM2YrdDl3K0htL3dSMkxzOEtpcXFvbkJsT3JwalhSZUxDNHNZYmdqa0YzeEdjenBvUlNNdnZkVTA0dTNzM2lJMUpJaEVCdGgyRElrMVd0WkZYdUxIazlodTZiTWZjcjB0QUVwUzBNWmIrdFBQcHB1NzdvTWRGRjk2a044RDZMa0tmaUxWa0FUdW5TTXdCSHBwdUI3cVhBOXF5U2h1amxIVDJtRVl3ZVE3OXp5TkRsSHRQbFZhOFIydHc2RW5MUkM2dzM4TXo5OVdJRHRkY0xGbVZjM0xVSEQ0TndSbGRNblh0UytkYWFTMFpPeDFhS0NDQ2RCYXhRVEsxU3F3ekVrcndjWjlsOWlQMXZ2UXpUWjE4Q2pEbnd4aTVWWDFiZHJRTUdyM0N5eGRoYmQ3R1VLRUJyNEFkS3ZIck5OSkVTYUNSSmwwY3pTYzI5cWlkRG5VVnBMUE1SMHlkZWpGUGUvTk00K1BRWHEzemdMcWJKS2ZKeHdkWHUvaGhmY3hDSFh2TXp3T2s3cExRQ1lBRy9hMlJYVnJEVzNVL3VZL05rdFJLS0RqZWdRQTUwNk1VUEhvMGFLUm05NW10Z0lLNjVPSkNBZnMrRUEwYkJSS3R6NGFLQmZ0bTRUb2lLSWpFRXJDSmlrUFVKK1dLZFQ0WUNqYWVYZlN0ZnBIT2UrY2Q0N3p1QVN3eTQ5ajRWMDUva0FROEE1aEQrL2Y3VEx6ckYrb3ZkUjdGMnpRUE9ibWNQelFjaWoxREprdWZ3YkNyMGtCY0ZzNkFNcUpBb2pIRkFWUmxzSHQwWVZtQjdkMGlJNWlJeGRzdUVJWVlGY3VOd3ZKWTVJbnpuR0dwTEYxa0FsdHBQM2w0L0dBRWtGNjAyVTRZS2xHNmlzVWlyZDZGNzFjc3dlZllseERBQVhRSXRSbmEwdks2NXI1aU1JWExNNUppakNhcnZhSk9lN0R0eDBwR1RucGgwdEdrUG0vU3laSEJSZWNoTDhKdmJwNXFvazg4OVFaejBLSXNSL2RtbjRMUS8raW1VQjV2RzRURFJUeVJ0cU85M1lOcWRwWlc3OTNKNng0VDk3WHV3Wm1kaXNySWZybGw0OU5vbndDWm9OVmRJelNLUldSdzhRb3dGMFlVU2VKZTFsTEFpZk5aUDVaZzMxUzJBVUVYREc1TWpTQVZnQ1lFQ3hSUk4wZWdRWFhPcGtuMGVrYWJyUzQ4SDhPU240VEtrSzNETnA0UmlQc2tEWG9wTDdFcGM2K2VYWFUrOG4wK244RHhTNnR6akluZHJpVmpNWFdNaFMxZjdieXNjMHVaRXNRVGhRNFBneSs0ZnhZZ1N5QXJrQzNCbEVocVE1Z0REeVZRNUVHQjVyb0lSWUNhdHpaRW9hRHZZSXE2ZDRONmpRZ2t3a0t3bkV5U0trRlpSRnU5Qi8zUC9BZE12ZTRJd2pzSmtFczF1amxhcXFQcXAwSUZsQjYwRCt5cWl2L04ybGplK1IvN0JEeUhQTjVnbFdYWVpPMnAxQmQzcDU2Ri94cU5rOTdzZnU5VWVBS0F4aDFjazZwZ2pxMUVkZElQWTkvQmhaSGZXWHA3Nnh6L2hoNTcwQThCdE03THJIT05DQkUwcDFkTlhKTTJKOFFTMHlaMmtoZWJmd25sWU9ISVJ4cnpZUUNrWkNXbUpGb1gyMXBmMFVMOXpGNGdCbEJsUjRub2k1c0dxYXRzaUJBYlRHVVlXclZVQ2xBZzZCWk5Zb3BlQWp1d25qVHNlL3VnL3ZmOVpCdDRLSEREY1J3dm5mUmpncFE1Y2l6Tjk1WkxreElJNXhyTkx3TVRBanVBUUU0L3BKYVovMTMwY0ZoOE1qUmxUdlNmQ09FeldFTTRtU0twVVFQTVNCT2htcWhOYllCb3Rhb3hsOXM5NlltUExRU01JYWxvV09HRHRMRVBOVXlDaFgwT2Vmd0RwaDc0Smt4ZGNUaDhHb0pzc2h4cTVuSWpLTU9DOHlqZW5ybFBabW1QMnl0OWp1ZXIxS20vNUVMQWxFcDJFRGd4Z2dnVmt3YUFSV2JOL1I5cER6a2IzdFpkaTladitPYnB6VDRxa29XVEVEcS9vQ0M4QVV1WEV2T3Voc1dCeTltbDI2bS8rc0E1OTJaVXlQemZrdEJETVM5VDh6VXVyK1hWNU5RSVdPUkRUNWVxcGlzbzh0L2FDbHVSSW9GakhEVWVnNlU5WkZaR1oxQVFkSU9rb0FNS3U2aHVTMnhLemRwWnJuN1pSTUJxTEQwUXl6eXpqZWJaLzdVRmJwejMxWGJqeDF3L2dHcnZ5UGhxWFBpa0VXMWdwVnpJZnR3MHlLSURKbFJSWm5ZZW9RRlZZRURScXUzTWFPVkR6amFCckVHRVNCamhrQlNubWVDOEQwekk3Y1RTVUl6SXdHR1dDSjQrMGtvN0F6aVMzcFlSZW9PQW1nZTVPRjFpcjdpN0o1M2ZJSG4yQitwZThpTWc1Z09mdFQwNGFXYmxOZUNreUMzeHYvSjNmNTliam42bnhXLzQ5OFBwYjJXK2RnOTdPUjllZmk3NC9VNVB1Vk91Ny9laTdrelR0ejhha3V6Lzc0WHp5dWtINXhhL0ExbU11OThWUC9qdzBMb0N1RTBwcFJDT1NCYWl4dkFwOWdzYU0vc2tQNVk1L2Z6bUc4U2F3NndFV0xJdWVLT0hVcFBNQmFSWUZveEszZkd1b1FpMzNUSEZ0cWsxS2lPNll1dllKQUpSTzJWbHhtYmp4TEU2K3VNVDhLTnV1UHBCWWgva0xqVnZ5MEFzR0JGc2dPb1UxNy9sRk92T2hBSEF4VHQzR01UK1ZBYW9tQmY5Kzc2a1BPN3V6TXh3THR5b1FnVHRUVjRtWk9vQkZLaFYvS2dROWdGMzNlanlGUWlITks2VlI2c2xVZEZ4NVJwTVZOSUtwcWtQWitNNmlnZ0t2STY4Y3hycHdPdkxKS3J0M0NKbGd5S2pjUEJpUDJqSU9aZU9PUSt4LzlhZXRtNlpJRTVPZDhNRXJQQ05RWTlYakhUbk96Y3Rmd1BrM3ZBajllOEhKOUdLa3lUNHlMVWh0QVhsRzVBVlJCaWhub0VUWG1zcWM0QUpJVTZiSitiQjdUckhaai8wNmpqeitNdVMzdlJkcDBoTmpzVmdjSjZORk82OTU5S01nSmFnVTdQbmh5OUE5NjBLTmk0TmlkUGVTMWM4MmhRNmh1dGdoV2c0TUJWYjFqbEdodWd5aWx6bDh2Z0NBQ29jQlMwZFVid2ZiUFhHdkM1QWp4Y2xvaFVXVVdvVUs3Qy8wa2lFaGcxRktJWEZqa2lzMWdhOFhPb1pra1BhWEhVOEdZUDhjdjN1ZmVlREhHZUExOWZ2VHhzbEYrenZiNGN4ZUsvUElLYVlHY3lIUTlpeVRWN3E2b0U1WE5DMmxUWUNaUjVWcXRaT1dMcGtMVm1vaFl1SDRGRUcrS3FTaThRZUl5VUlZNmdrTk0vRXE5QXdEYzRHaXBhb2JOOEhvcUVzVWhEUWxGeDlGK281L0NYdjRROXdYQTlSMVJCdWIxYUk0NmNoVitQcXU5K1A0WTc1YXV2cWQ2bFllUkUxWHFYSk04RVc3NDlVRXE2MVlVcVZZVGFIQU5zL0FzQUZYQWJ1THhIYzdqanp0bTdIMWgzOHNtL2F1UEFhSVVZSGVJTkFxMXh1Z1BFLzY2ZTlVM3JFdWxraDk2NTRsTjVUQVF4VW5LdFFiVlVmUGNBVHQrQXdPK0lBeTFNOWJ2UzhiVjFTdnZ1MHlkYWpOVHpHSHBNRm9sY0tqVXFYaHpGdkNMNW9LV3hPVFFVd2VuWWVxUHl0eTdoeW1aNTkwMGtrN0dqdjFhUTN3VWx3bUFMaG91dk5CSzdJd1BkVEtrb0IxSmcrR0k4VERxdFRhc3BSb1ZZaHE1UnVHMFppTFNENEVtSm9rcUJZVGdHckxkalFtaEhrbHlUclV5aXdwZWtRc1FqcE5rUG15d1NncWsycm9nYnhBd3pwMThpN1pEMzlQQkk5azljSllIUkFZZUo5S01mWko0NSs5RzR1blhvN1ZHNFhKNUg1QTNsQncwT1JTcHM5WW1jbGxNaFZpMkNwOXFDM3RhczA5UUY1SG4zWmhaZjVBYkg3dEQyUGpGYTlGbXZaQXlUWEJzTlo1RVBvSU02Z1VUQjk4ZHRyNXdrdHQ4TnRocVZkZGJGeDd3OWx1QUVCZS9aaU1VcDB3MUc3WWV0aGpycjdlVmVvSk40dFJtZzZRMDVTaWl6NGNRT2daYTMrTFJLb3c1am5FdGEwdDh5QVlQVGdWSTB3a1RNQVFvZHVLeG5JYWQreS9wRnowU0FHNDdENmFsajdoQnhjRnN1VzZoSXcxRzZVV0RFaU1odThTeXdGcTBsYUJzNkM1c08yZHRIUnMxanhmeTJRRm9JQU0zSXpMWTJncE51cGRWTm1kU2xwVzl4T2V6eUJuc3lESnpRbVRvdEJ6RlJTcDc1RExMY0x6THdOT1Bva1lSNkl1RGd3anJ6aGg4Ump6ZHRNZDhHZS9FSlBqSnlOTmRsSGo4Vm9ZMWVNV0VMTk1WWFUyVlFyQmV0T3A1bVhOamFNNUdkTExnQVJnTlYrSStUZjlLT2R2ZVFlczc2aWM2NmUzK21lc1JrNUpqdDNmLy9YMHZSa29KZWFNMS9PMzNEZ1N5R2lORng0NjNBaWJhRGtndkVEanVEekZjVDVaRzB6ajc5Rkh1OVlKN1poRS9WMUMyeHE3YkY1RzVIK05HR2hvVjZTbUNkRVVXdTFoUEtsTUovME1Ed1NBaTNEdzAzdEF3NVYreVNYb2R0Tk9ENEZBU1lFQkJXREpUa1Jkc0t0b2RTUlFvcTBRSWkzbVNjbThlYWNLRHNjQWl0YS9FZXVKUnRiYU9vZ2poRU50K1hPWUZyeWcwUG9SdEFLa1FpWW5MWk5XWUpaQjh6b2lxUFowSktIdkFSdG5TQ2V0SVgzWE41T2hLQWtQellyR3hqV2d3K21sWU9zN2ZvQzZKd05yYTNCdFNhbE9GOUMyUnlWRnMxSkh2emc2eTdTR1J5WmZZcExiRjZySitnSDNBY1VLKy9IK1BQcnM3OWJpL1I5RjZqb2dCMzlWWFVva3BDbkJpL3YwekZPMTlnMlBSL1pia2F6VFV0UktaMHgyOWhweVM4WDR4RVRKNkxCb1FVWEtnemhyY0JSRFJsWExuMFl0eG1RRk5TUWlRbWlVdjZocTdnaGFDcVlwVk9JRkZuUWNJZzkwR2x3SmtnVVJKVE5oSjRqSHBIMlBCWUFyY09tbnJvSVAxQUxrOFIvZGRmKzkxcDBSM1ZZMXIzV0g5U0FUYXFJYWZialI1bDlCNExCNG9sYWdGcUVLVnRzUVdYa2JtSWRiV0FvRG1yM1YvMVpNY1NIcTJqV1UramN1V2t3d29UbG9YbWRJT0pHS2tPcjdwa0pNRWpTL1hmeWFwNkk3NjJ4Z3pFSkkvWU9nUXkya2lvT3B3L0N6dnlLOTRmL0sxazREOHhicWJJb1dZZ1BVclVWT0ZFbGVmWEkxdUZibFEyenFFZU4ydzN2MEZ3djBFVXhUbnh6YnplTXZPQUJYcWRtTUx4ZFNOeXlTa2VuNXJtLzlLcmpOcVZ3WUJWOXRBMVhGUXhvYVdsTWMxVVo0ODJqTXA1d2FjM3RDaUVJcTFLOGFvV2tKaGx3OVlqTzBTTDBvMFZ3aWkwVWhvam9kSjlDS0pJbGUyUkJWblFZekhhTEwweXFKcWRMajR0TmQ4VW1WOE5JQUw2NzVHQmQyemlxNUQwU1dvcDRWQmVzckRlYWxJcElWWDFMdEhvdmNMampaSUJRRFJvZlR6VlVOaUxUSVpPTzBWWXA0cVJXSUplaXByamN6dUNWRVpLRUZGNlRrelVBSWN6RjVKZUdkMVNBbFpLaWZXYnJzT1dFV2JiWnp1OGdlZzkxb1NicjdFTXBQdlF4OWZ6N01aNkFoUUZaenNIcFVtQytyd2pyQ3AzcGRnZWEwV2d4VnVJbVZScWplNG9US0ZKQjhocjQvSGZ5ejkyTCtpeStuVFJLUTI1NGoxSDhWc3pEY2JYTFJBNUVlYzM5QW0yUktrQngxN3pxRGM2bXBTY3NKZmZzYWh5ZmVSbkVFeWVwZk44WWJBTm01R2FMcHZ0YVBVV1dyalhSd3R2YTZsaVFSV0Y2endPSUNuN1N3RnBYNE5BWUlPelJkd3dGWWFzUGY3c3NBMitQVU5OMlJTQlNydWp0R25jVWVkWlJoSlB4b0hpelVQeUZrQ1c5WElaZWdoQmpDc2ZBWENTMWhRSU9McTg0RFlzV2ZWZVY2OGU2UkRScGFBUkl0UjFhVnpLYTJLU2lhR1pMVFRjSTRFMDdaQ3ozNkVYSG5tMUdWZ3dJVS9SV2xFQVlNdi9nclNJZU9rNU1KaEFKRkx0bWdYdFJHQ2RhMElwUTVsYkZIaXpXaEJuR3JDVlRNMHE4cEFiWnZVRUl3QjByWndvcWRqY1YvK1UzNXhwenNZd1ZNcmVWcXE0cWc3T1MwNC9SeEQwVEJ2UjVUOXgzZWNzREtzemNQU0xoek93K3RzN3NLdkpRbDN1OW9HbEZid3E5Z0I2Qm5lRENndFdWaTIzZ0JORG5Xc3JVVnBwZ2N2QTNYaEpGdWovbVFSaTg0SFN2N3YvaFhIblptVldwK1hCNjROTURMNnI5ZnFwMW43WlZRV0JTcTlUQWtkSUpRakV0ZVZoQnpOWXpnZU1GQ1dJSE1xZVNHTGhPcE1IV0ZUSVcwRVV3amtFb29Ca1J0NzRGaURiOVVFd3NwUUFveGxUQTRSczZIbEdGSk5TOTBJSTFBeXJFeHFDZTVjYS94MHNlcDI3L2ZiQnhJQmZBYkE1L00zUjNzZStIb2NmSlhmeGRjT3gza0ZwbWl1SWpQVkJqdkVTbUdtY05TUnZETW1TZDhRWllseTR6anl5QnpRRThoajZxdmtZbklYYzB3U3QwZWxSc09hdXQxYjNJbGszSm8zY0xMS2hMUU92bG8rcFdQOUlLdHVuR3d3RmhZVzBWQlpBaGo1SnZSajB4RHJJNEg0ankzTHNKZ0xOamFieUt3QUtDS1MrMDFzcVJjaTR6STI5dHJRNFVtUjN3VkNXN0oyLzFhYWo1WVlKN3BYcXhYWWVHaW5GYlMvZ3VPcFFzQTRQSlBxSVJQK0NaTU1Fc1BNU0N3N0twdEJGM3M0bTRnU3lUQUxDMnNBaEZpQXhwSjRTRVlIa2xJc1pVbHBlM25JVlU0QmN1WVV6MmRSN0xmMEptNjU3SGlpS2p2RldjdGxYZ3ZLMmhobmNscEhTbHNDZWVmQnd0SlZkUTM3a2l4SGRQYVNQanh0WDhNZmV5UTBzcUt4Q0trQXFZU2MxK3NiZFB3YUY4S2NVT0lLOHdCaTBxZWtlK0doS1BkbkZiRFdjMkVhdlV2UXBLNWFDS1IyV0VYOHUrK3BoYkN3U2w2aEFKc2V5ckFIbksrWWRMQmxDdGVYMnBzaUthcUNnWXJadFpzandwcGVoMjFDTjgwa1U1emxVWUxnd2l4WndSVnI4V3RReWlzT3ZMbDJEZzA5WXZhQ20wdEkxbGxVc0l0UTZBWE0zamVyYzdXQnB3S2ZISWwvRWtoZUNhZERhaVZaYldua1UzOUc0R2lGUlVWTWtFejBxUW9UTXloRkdvNkdDZ3JLcXdLZzRCbXhKQ05CdXVEeXB3MW9pcHlEQmpBVk1Fc0pRK0RxMTVKRGRJSm82Y2EzS01pcmt5UXZ2UkpTeWxDSkY5MUFrck4xd1RRM3ZrT0dYb2c1ZkR3akNSSzlhWlNOZloyUVFHUFlZQldKU0pXVkZNcXhBekFNTktsVUM0dUpPUGkxWnUxd2lUVVFoMzJNbDk3SGZMbU9wanFsdk1BQUJCY1RaeXcvdVJUa1I1eU1vcHZRcXdmTzlvN0k4OWo0M0Jqd1VjTG4wRWE2MFRnUFJTRGtJaEVCcVVHc2ZWYTExeVhWYjRZa3hJUVJ0azhZbVFkMWFZVk9zUmF2S3M1ay9pUlJTWERxUXdYYU4vS0o5cmFmUnJnaE54Ujg3UzRrK0xUUnU1Vjk1ZGE1R0pTWXpZcU1CeS9RelMyQlVvWlhpQ0Z3cEVtV2hKaU9GMFJZNWhuTlpLUUQxbk5RK0p5S3c0eElRYlhHU0wvQ3lPb01pMFBqNXRFSmtFbUZZN1EyV2ZHbFF6anM5WkpiUUtaT21qTXlHOTZLMjNIR29HUk5NRlMyWDVkaTBKRE1ka3lDb3ZVYml5djNoZ0EzV21GeG5pTzBhT3JPOUNBYUEzY2xvbzF6eWlnQUgwSEhKMXplT2NIMEN3a2NKQnFESW4wTWJOYm04SWVzQi9Bck9Jb3BSWVlOVTl2amU1b2VHemM2TlhWVkdNSloyR28weEdJcHJza1kyVlBnT2UxRWw2aTJWcmlnNkhBaVZxS0taSWxWZkZxWEUvSkNXZXF4NE1xRVZ0UndsbmRqbDEvalFFR0NJMk9xeEZlQ21ERlNCZTZES1pDc2hCZGdTeTdSMFViMkY5WGdCVC9yUWk3b2ZlendNY3NpZWlLcWMveVNSRjZsMkpvTGJ3ZVJHeDI5TmlQVWxGVUF1NDJRbjJWMmxocEJVNlVXK2JSWUpzS1lVVktEbllMNjZla1ZsZHJrV1FSMzczS1JUeWdEbTF0QVVlT2dCTURMRDRma3BNcG9CZVpRNm5RdWdKMEdlaHk1TGtSa2drTG1JbUc0S2JOMWY0dUlKdVc5OVpjRlJIYXJYcEpRYlJld21JRDQ0YytGSmZBcTZLbFRleXJhRGNBZFpQZGNveUttOHFCeXYyR2g5Mm1KbUtYVG9ZaHhvTVlSMWhUSjBTQnErS2xHaTZpN3MxeTE4Z293Z2kwMlRrZXc1SXFLRzJ4UUNmR2k2RE8wREcyVVc1UmFrWWU2M1F2akEwdEdSTUJlemc1R1FDdStBUlJ3Z2x5ckNzQkFDc1cyNTBOb3RjMmhNcGkxTUFZektYUndSU0htaW9Yb3FDc0pFVEJxdVV0YlRWczFoczhSV3hkSXFGc2NVdk5oZGNlRHpDU2VzWFNxYXFUTVFsdElKc2hNTUNBYmNJWXRHdW5iTHBXN3pBaWRwZWJKTi91TExycm9GSWV5WDZsRmpqVjdKMklPRmk1cUNqc0ZCL2ZscmRzM0QwTTVTSGJPaVZyaW5nYVE4VVR3bTNCYUtHTVJKMEFZb0d6ZFpqQXRvN1ZGeFVjc1pOdWFaQXRWVjdiQldHb0FMY0hFb2x0VVVWWUtobm9hYzNCVk9xUVMyNkRMZ2pjcncwenFpNFFxVFZ4TWJTK1RRc1lXb0R3c1ZVckhOcVJpRmlzV1dhbE5hTnRvNzFtWkc4RlNSMHl5a2xvb2UyK0RUQWVPMlVXMlNhcmoxZWdKS2E2QXFBbHBTR0pYTTQ5VzlicmJteUtOZFdyQkZYanF3TEhaWk4rbkJsWG5jYkpJQlByK0QxV1loV0FmNXpCeG1jV1ViZW54dnFzMW5iazhyNkhkUlhLWWN3aEFNRm9FSWxkSXRoWUI0WVp1TGFyemcyVXFEb0ZueTZMQUZ3VFhGVGFNRkwyRUhNeTdxSUlsL1VEMXFzakFGWkVKNVVJS2Z4SG01a1pWQ1FGSzBZazZmaEd0YkxZdDZCYUx4QUdqeDN0NnFZVE9ncVdrQXRCcXNTU010aXlPU3B1WUlzc3ZwN3c2a1VqckN5WElMWUVIaEJ5elZCRVUwRmRRbElmYXB3d3lNYWhlMFYvS3JiWmlwN0lHR3NiZlAwQWNkM2hHdmZFNjEzMHFUeGdQRklGWHBzd1VhR2ZZc04xMms5RjByekpCbXJtb2RvOVZqbUNxaVp0cFc1emdKSEhWSGZITURTRkdJRWZkNGU0Z0pRTTZnWGtkaEpxelJYM1h6eTlaYitFVU1KMDFWNm9BZEJXd3h1clJIWjFTblVHc2RRTDF4cW1vdkRUQ2ZtWTJndGE5Y3BlczZPNFU4U2wvVzlyWmVwc2NqUkVzN1dYeDhVSzYvQ0txN21hV0VEMTNGVkRxaG1zQUtMRWJHaFlnVXIxVm9paXZuNzJTa0Y3dTNtalhtQUR4dHRKYlo3S0lrVFh0dzNDc0xrM0xRRzdxa3hpYkFWMTFQRVR5L3dSalhPdW5LVTFNd3gzVUNGTElkTlhQdEg3ZmFJQkNnQ0t4ZzRKd1hZWVlRbXhiZEc4RllTbzBnODJLQ1ZPUTNRRG9kYkdyZkFQeWlIdSt0alRvbGk5MER1b1ZEOGRBUkx1VVpVdHdYeDNhQVZFSCtQVUl2Y3h3VU1PWHQrYjBaakUwQTJaeE0xTllZaUxhb3lNSlRCMVJCRU9RR3M3aFpVSnhSR2l3VnlNOTZob25OcWRBWVFhdlU1YksxVlByN3E5dWxKQU5kVFZjMnoxL05SU3dNTVZDQVFzTWpKNTdScEFSamRkcWJkb25JNFQ0bnlvM3dHVTJWR3dkbjRaU3IwSUZnYXJRb3VJRmZkelJFMFhSUXM4TUY1T0lHbndSaWlveWxLejNEWFFKQlk0M0VPVUdad2VKSS9pMGgxTTBZa1VkR1lrVEdURkZrZ3Z3VWxuR2pJZ3NHZUJxV0JUaXgyZmJINzM0UUczcUZJUmdKQ0UyOUthSTA2NjZncU5tamNCbGNVSWcyUE5lRnNPRjJQazZ5QUplcmpGSktIN0JIZUhPTzFlNVZnQTBCa1YxRnRnaHpVOHhwVmRKa2UxSXE1WFNzbWcrVG80bjhXdjBRQTFRelFzUmZ6bi92M2hUYjJJZlh6Q2NPNE9ENXFYSU52MDJucXpSVzVZNzVyNFB0UkFMUzlCZFdCVldoWm5nUlZOQzYvaTFYRVYwSWlFTGRydU5RSnRUVTBOK2RYT3ZSNlhoZ1dKcEtaa3F5czhFVG1mTFZXRkVYTGFMZHhLdkdXdHVkeERISkNBTnpaTlhhV2dtbHp0aElyY2xxRndLYzlidmx5OTdNdlFGQ2Nrbm1SUkNBU1NNSmp1YzhIaGlRWllYNnVNVlcvVGxJNkdWdTRUQWZyV1VLVWw2RmtEdlVVOUVaazJhd0NMMTBvMVFnSFZTN0g1KzdDaXFPN2FrcjhJT3g3QUttU0UydnEzU0NqaWcxVVBVRlBPZU4rKzFncmpFQ2ZjblRoaGQxeEVCU05XcDhxN2RxSS90azVQQ2FLM2p3TzIxQ1pROGFESHhFYS9zVExYTlVKRi9IVVBxNnZwVUEwVXRXSjExSHpFcXNTczNxQ2h4UUgyblZJdHoxcE9WZTI5OFpPQStSeEVXcXFSYTRxanN1U2FXUk12Z1pGa0I4UzNqSTYxTTZUbWY2R01wQkpBYUxUR1E4V3A5SHBaV25SRlBla3RJa0FOb0E1YmE1Q2RxWTVCOGFwS3FNV1AwSHRyRHZyNHg0azRJQUZnVE16eFIvSCtiZzBJaUN1enhPRlNBTW93QVowSERFTkphVW5nQjVEYk9kbUppdWNJbmNBT3dTYkV4MlA5cktwOWI5dFpJeXBTMmpuWUNlamErelpRdW1HTThicm9CRGVCSGVFMzNkalNtT1hvbVdVeVZES3M3MkdYUEFtYUh3ZDZpM2JSbU5LSUNpNnppaXNDMXpNSEtzNW9MRHhCalNPdzFBVWJwZUtEd1k4M1NLYTlyWnBhdTZxM2ZWakFkKy9BNUVzZVZTTTNsOWRCQk9TaTlaM0tZbFM1K1pBVEU4UnE2Y3JTVkF3dVRxUEg4TWdHU2xkZ25DbExWWGtRMlZCd0FHNXRyZ3NVR1hBVWRBRTR4ejNIcG5Ha2FucXdiWEJMeXdScWtWS2xkdUUvR09zUVJ6Y1V1VHQ2WXV2VEcyQWxnKzlNSmRVVFc1QnFUMGNLY0lxZDB6cW51bTFzRUoyRHlTdFc1c1lrV1hMSjR2a0lYRkJJb3ZvQzlCNDdBZnNhcEFNUUNNZFQ1ZDdMQ2lLWkZ4dWwzb0hPWWFrSVhRRjZoM1dDNHIySnZwaDM4VHZybkJ3MjRlOTZCNkxvaTFpZ3dOamdiZzNoQmg3NUNPUmhDK3hjeGdKYVptemNLazBBRjRaZmFUNVlJY3pwblJEaitRTUxSVXNUcXNISkNpeDVkQnlFUkFSTW1ZRTFDa29rMVJIakp1M1VuYkp6endzUFVEV0xMUWRWZElOaU9IU1EvcjZiUVpzSXF2aWo0djFSNVY2aFMzUWFBMUUxdWdNWk1GOHVkWXdySE1PaUtpUWF6cmxMTHVhS2FRcDFZSktxSXhQcTRLRVFQemlzcmpCa05iNElmQ0JWNVZvb0JBbzdaSUFaQ1k1T09ocnYvd0YrS2c4SUFGaTREMGlFT2hrN05TNHMvRkhnZ3RIQjJNSnhxanV3REtqaU1Ib1lIMkJhRG14bW85RUl4QkFTTWkxVmNKRlNrUldMcjc0N1Fld3NoQWV3NnQzb01iNDlPWkpKWmg0RGpKTmdDVEVMYjgyUWJyb3BybDRBUFVnV29Cd055end3UGZlNTRDbW5pT09NU3VHQnJIUEc4WVVPcUJwUTFFdTJuWUt3R1JjcmYyMEt2YVFCVnVWbjhmdndoREhlUlNKZGlZWHNFendmWXYvY1p6Sk5Kakc0UFM0M1NRU3c0aEV2OVZjZlJab3RZbHh2YXhoaUNab0JwZVpFQWxFRXhLRHhDaDRIWTF0eFJWVW9pcTBKc0VaSWViRVVUSWFDM0NGU3BSQmJEdGdXNW15bitNMm1YWFVmZmJ0N0F1emNQaVdLWVNtMjhlazk0Tlh4ejZwNEdBQUNOMFh6dTNWQVhJZ01WRSttcC9qWlVzQVoyMDREczZ0R0NZWkFJWUQ5bXA2a2NOTEZhN1hmOGlLZVdNKzByRE9nTEhTQ0pjQTdWdEVENENaNnBlZ1lOSjNrZy96VXZTanYvci9BNFh1QnZvOUFVdFZqY3NWWXRaeUJrMDhHbi9ybDFPRjd3QlVMVDA3QXF0ZVBQaFFwUkErSU9xYXJSbWlDSjBWbEh5ck5SbHVHOWJRV2hHcUVTTzcxM0ZHV1hUNmdyRUwyRFpkRjBFcmJsNkp4cThzSytMVnZVb2NFZGxWanFDb3FaVk5mTjFHc1VKc29ETmltNTJyM2RVaGxQYkNvR0RYU2dJR3dTZGJDSmlpOTFvYXBLSndxTFJmV1Y5K3o1dXBjbXNsU3h3U0xHd0dtbUtTYVlmZWVZR2IzWVlEMWNjckVQb0JVUU1vc3drL2tNM0FnZVJRUVNhdzVZQTB4WGdXakFydGFzWnFFem9tK0VKUElBZGtKbkJTaUsyUlhTQ2djRDFETDJlMzdxcDEvVFFUMm9YeFJKMW9YM28vSmdkNkJIbUVvS1F4ZGRHaHRBanQyTi9YT3R4TUFXVW9keTIyeHNCSVZaUUNnNy9qMm1nbDVqUG51QkcvY2MrY0lqdGtKODBaSDFuTVN4K0FwOUlKTWtwSUNYb3MyQVNKNUdJMHBhblVUd1F4TWUrajRyZWlmL1JSMEQzOFlrQXRvMjc2RnJOTjh1d1FmUnVwUC8xS0plNkF5cXo3THd3Z0RSR0hsWGNIcS9RTGJyTHE5MEtEVml4MkZmVnZyRUxoZFZGRGVObDFYdUV2TkNLdVY4Z1NJTzQ1U3FGc0dFTlZ3K0NrTHZTSVRDeExjVFZKbXhucVozUUVBZi9VcDFUQTFCOXlhNGtNNXhRc3p5VkxuWUZmaXdPb0ZVUTB4UWNDRHJVcGxWd0tITlJjNnA3b0NkQkpUQVh1bnV0RHNvWE5nNGdBU0d0a2tnTVdqMGFmZWFHQlA5Mmw5YmhkZnNzZ0IyVHN3RWRRNTJWWCt1Uk01aVhsN05pSExxMTlWOTFKQ1hxcFg5YnBZSVNVeFozV1BlQXp4RGQvaWZzdWQ0bVNLWUYwa2RVMFJVd0FycnM1YkVTWXR3MjhWTURTcEdUTEVYQXVQSnVsU1ZYRm5RRmxtRW9lQjJGRXdlY21QeFRWZjRycExSNElha24xODU3dWtkMytFbU93QUZWS3YrdGdHd1ZUYllpTmFXTlh3eWVHZ1ozSWNveGFHcWtJRU5XVEZXMkdveFVZMUxHenZVeWRZYTI3V1lnUGk5bTZXMnY0YS8yMEpSU0dVS09nd29rZEdSOGVBQVI4b3h6OTlFWEoxOVkwZlBqcGIzd1JpaUdHRTFBQzNjazJTUTg5Vzg3THdBaTBrd1FBa0oxS3dIVXhTQ0RnVkVFWVNMTEhsZ0VGaWI2ZVlTSFZDYWxCcmdDWFJ1bkR5Yk5xc0RqRVFQdElBV0lxd0JsUE52d0Q0Q0o2eEIvYm0xOVB2UFN4T2VpM1Z5ZGFnSGdCbVpDbmdqLzg0ZWVhWjREMzNnS3RkZUpobHJvcFFjYktsRWFHNldYNVZNQkFXS0VrN1RxVHFxU0NZTFZzV2lFbVBjcytIZ1pmK01QQ2dCeGpISE8yaWJKUk40QmZHU0RtSFgvc3RKRXhvcWJWRnhzNjdHTTRVWWREQ2pkV1E2OHZFdW8xOVYrcGFlUjFHU2JOcVhPRlBoMFVNcGF5Z0F3RXVKeUZvV2VVMm5DNGdJQ3hkWWZ1cEFwc0lPbVlpSUtuUUJNNHM0eTdtR1FCYzg2a01zRDF1blErTEJSUVVBbVBmY2ozeWdNUnI4NmRpYVVTZ2hIVlVrZ3hxRnlieXA4aW1QVldZSmdGSW9xV293SXJuaWtmVis3T2RPdFZ1NDhoaW9RNWdWME5zY25tSENzblVSS1BtWXFvNUZrM1VyaFhZL0U3Wi8vbzFBR0NUNE5jdVkxWUZUb2diOXA4Q3ZQSlZVa25DYkFaTSs2cGlBZFZKeTN3d1psUlUzQjlTRXJ5TEFrZ3RGMDIxN3FuNUJic0t5VkRDeW9yeXh6NU0vT3R2eFBSN3Z3Y1lSOEVTbzBpbzRqa2hkT0o5eC9HMjI2RGZlNk5zZWhxZ21VS2w1R2lwdWFGb08veWllcU5ReUVUK1VtQ2R3R2wzUXVFQXNPMm5xQWFnbkJHaHpDdWVIakJMRURtVlJhbnNzQzFkdENvY1hsT0FTRmk5ZWtWMHFHTjhxZTRveStZZEs3d1JBQzc5aERGdFN3UDhxeVdvb3lQWnhnR3BKSFNGTm5XeEUySU1Xd1o2ai9EVTFVUzlFNnhyVmE4dmN6RlVEQkNka3gxcW1BTGFuQmYyb0Z4Y3ptdGdDME10RHRUK2F3TG9DalFSMEF2c1NYWU9kVUlvb0QwTXZuZWhjNFZ3RlVBZXdmUDNzN3ppRitoSERrZERqMGZYbW0yblRJREY0aGMrOUdHR3E2NkdEbS9JTmphbExzSGdTa2xnQWttSlhZRjFSZGFKN0wzaG9KSDdtb0JVUXFuZFpjaHlmQ1o2cEMrVG5yanByOGl2ZlJhNi8vcGZZMTZMSllCQktaOTRSZHdsSjMzKzBwK0dIUmN4VFJWMnI5R0lwYVhNRnVyckFxTzcxZkFZTlliSHBJWSt3WFpOV1U4c2x4ZThCbVVBc2xFcTBkVnpJdVVmSmJsUU5Ta09JcXRhZGhocXRHR0dvUzR4UWdBUUo0RW1sUVN6bzV3Zi9Jc3o4a2NKNE1ydEZQL2pEZkNLV2tmMy9kYUhObEZ1d2RUSnJpaE5SZXNMUEJVVUNUWnhzbE5RWkgwWVJjdlAyRVhWVHd2RFhGYS9CakF0dzdjamdaaEVwY1IyUU1HTHlhc0VpUURZVzlUSlBlSnJVckcvM21GZGZZOU9yQUIxZk5Wak1icXdldzFwNHhiZ2wvK2J3MHdvUmFocE0wTkpGV0hGa2pobThRbFBKbDcxT3BSakJUeDRHTDdhQTUwa0s5VWJGNmt2VkZlZ0ZQcEFzNW92VVdDWGhSUnJoNUFFS0VkYmdTWGdwdXZnMy9pMTZGL3h2OVJQSm5WZVZZVkgxQnE4Q0I4ejBIVll2T3M5OU4vNFhkbk9zNFd5V0ZKZVJDQmJiYTBsS3k2SWlrbHkyVEJXQklqSkNGdGJpYnl2Zm1nUHNuY1pxalZtQUxrQ3M5dkFkdk5tMVRSc0c0QUdZcWxQTlhpMDZDMlVHcFI2dHZVN3hLRzAyTUwxMXkrczhRdjNaWUFFNUFkZ1Y5NkpyVHNXK1ZpZzhsWUhQVW9rVWVhUmhJTTFONHZSUzB0TUxJek9XZlYrVVlOWG5IQUpRWFF3NnlqMmtDMDJLdG1GR0RCWGs3akc0Nkd1NlY3bW1FbGdyOEFCdDhPeXpCUnVyWVY1ODFqbHVCakFCNXdrdnZ4bmlacytTdllUSVdkWUpSZFJFUTh6d0x1T3pCbDg3SmVRYi93VDZINFhRUis2R2JBTTI5R0RYUjFRWWdBcTdsaXh2c0FOVTRreXMrYWoxaGs1blVMSGpzQVAzUUs5NkVmVi84WnZ4T3VVd3JhVW15SFdQakU4QnZUeUF6K095YmlMVE5GRFpHSHNMVVNpOW45RUljVEFJMk0renhLckExSENYS2VUV3NoR2hWM2wvblg4RzRqMU9UdkVvS0ZVUFJ3YjB5RlpJSEtsR215RGQrTENCUjdnOVRZU1ZZSHFTUVdxQmFGSDkwRUF5SGp4SjZWOEgvK0REOFNIVDBrZkJobk93Z0RyZ3VNdWl5Sm5oS1IyVDNCWkVOUWF2YXVBZEVBVEZmZFQyL3NVdEZZWE4xaFp6TGNWMFhDeWhZaEdoZlFXZjI4MS9BWUdKNFVCU2wzOHNYY2dFN2R6UWpORm51cUM5ZUN1T2ZUaU9oL0dTNDB3YUtjenVzTUlJU1V3RDhLREx3TGUvR2I0ajd3VTVaam90OTVHbEJtNGFzQ09DV3lhZ042Z1NhSjZFM3BLdlFFcm5iaHpCVnBOeUxNTmxkdHVSSG5FQTRFM3ZocmRTMTRDNWhGeTBsS1NzYVYrMmlhMXhneDJIZkovL0ZuYU5YOE83ajBEOGtWZ0dWWmJBVnFMS3lIVmRnU0VJVUFvYW4zTHdXbDRGR1dCTWFycVJkbk9lWk5YNW8wQmRaZVVBbHFKT0Z5YnJ0VHlsYXJuRWV2NnM2cDBWQnVPRkxMendtb0M2TUd5Z0hDN2I3MFZBSzdCTlovZUFLODVXUG5nTXI3WnU0eXVkN0YzZFZPSEphZThoRUs3aTd5UUxmeEYrRFdrZ0I2aUVnNGRHbE53djBwT2RnNFBIcGZxTTdWMVpCdFJiMG9TVkUwSkFLeXRBYXVBVmlxMkdHRTI4TFdrZUsrSkV6VWNvOUZ6bllnZXdKUkFHZWhublVKYzkzcVduNzFDZVRJRnhySE5GNm9xNnNaMEUrbzZzaFJTaHZUOVB3ejk4ZHZoTDN5eFNqb051djBvZU1NdDByM0hxTmtHc05pTXIvazZzSGtjZnRlZExCLzVhS3d2Tys4ODhOZCtFZjJiM296MGhDK0ZqVG51NU9idEFvb0RXOXZjbU1HKzAvaG5iOGY0MHArVTdYMFE1UFBnYnFQVXJneE0rSlZhQVFjZUY1aGJWT2dXR3o2TkxsWklwUForUmRkWExSM3FlUTlQZVd5bWdqWkl4UkhqM3dJM0QrVmREZDVSOURBMmxDNmwrUTBiakpJRVJUc3dvcS9VL2thMzhQZDF4NjhIZ0orL2p4bUJIeWZIT25SdHZOSmZMY1piSDRtc3ZSMjdnR3dkbG9rOEVsb1VjSTJSZm9UcU1mSzNKU29WL3pKVlhLSFJWcUZUQ290UEVDYUZuQjlwQjBSVWRBbE9SVThUcEowN3dHbUtnaWVIQ0NMQ1Q0cXl2TktFUWtSRGI2cXBDazJSd2U1cG1Jc1BPWTM0dForaVAvd0pzQ2QvaFRpTXhLUnZrQUlxcEI5UUtxTkF4amlhblg2MjhFTS9JbnovRDFIdmVDdnd6bmRCYi84VCtQVWZnU2xCcE54RVRSemRGMStxOHZndlZYcjhFNWtlOGZBVEdLd0M3MUl3MG9wQkg2Um9paGtUTEhFc2Z2MU5HTDc2ZWVqSzZjQ0t3TEUwOFZWOFRJWktuaXcxd0RXTm1Fc1dLV0FWMWxaeGRJSFpTS2F1M3VWaHRxR2M4YXFzZyt6ZWpaajlIQ0tsbW11eTZtbXBTcGxVU1VXdVowcE5oYXRXZUxUZnI3TFVRSmVtZDJoMno0ZE92dk90dUJXNEdsZC9raUxtNHd6d3Nwb2EzVE11L3Z6SXNIYlgzclYwaGtxUkpUSWxvR1J3M0JENmt3eVJnVlkySk56M3NvOGdWQ3h4bGdPOFpiQW5BbVFoUG1PWGtiYU9CS2J2bGJCSE5aaEt4M0hYYnNoU25PYysyS0FRdGFLbUowSnRYS2phSDhXMUNtRld4ZFcweEhlNysrOGkvKzNYVTcvNngvSkhQRVkyamtUZlI2OWlKRi9iVTBzQnF1OWw3dkJjcUs0WG4vQms2QWxQcHZCdkZZTWpSbkFzMWx0U21mWVF3TDdkRSs2aDNPOE1VSUxSMWVqQUpkdEIwSVlSbVBRcTl4eGgvbWZmd1A1WUwrN2JCZmtjc2hnalcxVXI5V0s3RElsbW9Dby95Z292eGFZb1cxWllSSmFtSzJMcTRuS1JiVkZWOWFsaGxPUHNLQXdwdkdWekNGQTdXQUxiVUhtRmUrcUhwQm95R09ZdWRCQlhhaFV6SmRNZE5ydnJUMjg5ZHF6dUtmaEUrL3Y0RUV4RUE4aVZ0K0Rvb2JFY1hPb0NJLytTS0pWNTNRYVZHQlBwSWlrSExTYS9XNnBYSVlHaGVrRW9aaWl4MXV3QjJqcDhkaXhPYkwxMzZ0RTBnQlRZZHhvd1hhbHlMd0E5b2I3UlpZcmNwcE1DcjR1OHNPV0lTSUFueVpQWDNMSEFWbnJwbkVIOHRxZlQzdjRXb3UvQm5PUCtZVVVmS3BpdGdMN0NSZlZkYUNwS0JzZEJxY1FlWTNRZHNUcVJwbjFNanMwWkdrZkNTM2lQbExaRFFvTURFdEE2MGpBTTBLUW5icjBOL3ZUbndENTBwMnpmcVZCWmhKaVZjUTlYOWlXdWQ0Z2I0dnR3aXBHWHQxbU1Wc0VRQXdwR3BGTk9CYnF1VVdYdHM5VVhyZ0RFc0JsOFE0eS9SWVZhR3VZWCtzUWw0MUh6UTlUWmtIRFZLYXB5dUthUVZrSUk0U01kbXlwdkJPQnZ3cE03NEpNdDhCT1RRcFluaDFmTUxHOUFsOUZOdk1nY2FlcTBxYUtEWWU3Q3FpdkdwYlhxVjBTSGFoaVJHN0lMZUNMVUs1VS9UclhabTRCbVc4dGJHNmo0RnlyMURBRDd6Z0U0Q1Jxclo2VUNBWFlPNnl0SDNBVVBXeHVLcVNSYVY0V3BDVVFIb0JjeElZb0tiUGNPOEx5RjlOM1BVdm1qbDZOMEhlZ1NTcTVOaTVYK3JPYzduRkNjS2FWRTlCTW94V0pGTnR4WDBaL0d6c0MrZDBjQ0c4WlRING9XUjFoVk9wVmhFQ1lUNVBlOVg3T25QQlg4NkszZzZhY0QyQ0piRytvMmsxUlZPQTR6TVZUaGxYK25hcnRuaEE4akFoNUtFakVYTGo0dm9uZ1RIa2dNRWpuQWRBZmc4M25VWjU1YmdsclZCYTM0OEFhNW9JNUxvOURHZGdpQW1PRjBqTmpKS2cxRDBRWkd2S2ZjL1NIZ2t4bVFUMldBdXFMK3gwZm40OXNIUUVpS2hvbXB5M3FSSFRodWxpZ3Fsb3dIcWxyRmlVNlExUzEyQVozRUtJb085WVNpTWlVQTFtTjJjWTJTWUozZjBzYmZhTHBMVEpYL2JQUlc4a2djT2dGZHNDQkt0ZUc5dlc0WWZNVW93Ylo1bVpNNlJtblhEdkRCVS9MRjN3RDg3QlZ3Q2toZHpIdU8weDZEclpza25DeHQ0U3VXNHFkdFdKZGtyT3J5QnU4NGNLTDF4YWtGSlBrNHlyb0VUQ2Nvdi9WeStKYy9BK21ZZ0ZOT0E4dU1xZ0lRV0V5VXEyTFlaZU03NkZIVXhvMHRxMWhyRzlaRVZueVZCTEZKbkg5cVJNdlN1RjRBSHROSjFCazlBN3JqdUloRTh4UEhla1R2cjFBTThMclRMNHl3VnNwZ3E4Y2hsSUJmc0lvQ3FjREF5VzFhWDl6ZUxmNGtQdisxbjVULzNaY0I0b3BybzRQbG5uc24vL2ZPUWNmUVdYSlNGbEozSVFGNVMwQ20yTE5CTUcwdWpDb25TMHVRVXZDL2RmUzZ3bWdoMEIwVGdCdkhDUUJtVk56TWJteGNzQURzWHFOT09pbk9SUmdUdERSRWhxZnRBWFlONG9IWVFlaWkyaU1oSkNwV2dJV0VDeDFBRC80MVBleFUyTzljQ1gzVDArQWZmQy9VOVFwUU1CczhRd29CdTBlYlJOdkNzalNubG1xcC9wU1ZtS2hQVmdQVUFRZExpWlhqS3ozendidVFYL2lkNEF0ZnlINTFEMjNmTG1LTXVZUlJtZ3RNb1VjTXoxZHY1Z2pCaW52akJGRUFhMTRYd3h0UlpmRVVaa2g3OXdDUkZRTllLcjFpd29sMThDUEhWUDdxWTVZd2xhbFU1RzNiRU91RE5kZFRKUlpEK2dLcHdGRnFPRjZCY3dKQW9LOGkyUkdXOTcvOE9iZmZLSWlmeUlCOFNnTWtJQjJBL2NDZEcvZXNiNDF2UXg4bGxabGdmZVdBWFJqV25lZ3JlQkdTSk1rSVN4UTZCdkVlZ0RUZEVHTFNaSFVmWENaMnJFSUgzdzJWa1dDaU16cHZyYzB4OFF4T1ZvR3pId3ZNTjRCSkY0Vk5OVFowQXRwTjBSSHFRRXlpUUhhRDBLdks2NE1pWHhxb1FTSEJWeWhPSG5vNmVPeHR3RGMvR1Rqd1hkUXRIeGE3THVBWXlqeVhHRHBldkRhWkNpR0ZCNlNZUm9OVzNhTjF6OVFUV3pMa0hybGYzMU9MQVhqWnowRmY4bmowdi9zSzRJRVhTREV2R2pYbnEzeERVT3VWdW1UanVxeng4azIwdHpTK0drYWdKbnFBY2xUVy9hVlBxamQ1cWhkY2F0T3hEQkJteDRERFIySHNnUmdKaVZibFdlM0lZTTBGdHdIdXFoR3NJWG9KdjlBMWdkQlJ4U25jeGMwL3d0VW92NFRIM0dmK2Q1OEdDQURYWEJNLy8xaldtNUVLK2w3dTVwcnNkRnB5YU9JWTEzT0F6dzFvN3NEUXh3bld0ZDRLTEhGQTYwQWtsMXRObkZZSkhyMUozTm9FVWdBdjdsVW9LVk03WHAxNVByREl3QVJONk1yZ1hPdkY2c1NLUTBxOTErL0RFTm1EN0pZaWlQaWJEbURuUWhjbEc4WUJkdG9lOEdFOStjNWZBTDd0UzhrWFBWL2xiVytFUzlLa3AzZGRIS01BWm9HbEFEa1RlVVFxQlN3Rm5uUHNBeWtaRU9CbVZQMDd2K00yREwvNDMrSFBmREw4aWg4QTF5VGQvNnlvZEx0Qm1HYWxFK2pFMWdhQkxzUUhNUXFrQUJTdGpqMWhjaks1TWNZa3M5a3ZnUGpBSmJ1ZHNrdDJ3ZjBVUDZ5OU5sVWQ3ZlZHSDk5N0U2d1V4VVoyNS9icVc2ZWpOUGtHMjFjVlFMQmhoYkVjc3FCSHhpNE9VaHJWMGRNdDNRYit2RC82YWdCNEFkNzFtYS9xQW9CTEw0WGpXdUJ0UjhjM2ZWRzI4YXdlWFhHZ213QnBFcjNmZVY0MHpwUDFPeERIVXVWUklPcE0vYUFibGJCc0hLZ3lvM2hTSXFDUldyOFgyclUzbWxpWVZDZEd0UHhLZk9TWENkZitad015bHUxS3JaUnJVU2U4UmFDU0ppQzJwa1lZcHpWUXJEWGhiQ01JeWVLNFN3WkE4TnlUZ01VQWY5Zi9oRjM3bThKSkZ4SlArR2ZTWTc5WTlvZ25HUGJ1Z3h0bDI5MUR5N3ZhQUhuWE5WUVV1dW1qMHR2ZVFieitEOEhyM2drN2NqZHcyajdob3ZPQm9SQ0xlVXpFWXUzdGkwM2RsVzFZYm9LcTAwUnFucTBpeVVJaTVvS0htQ2UwVTgzTUVrajE4TmxkNEpkZnltN25UcUJrS1hWeFhzUHVDQS9BYWZpemR5S3B4TTN0cWdWRzg1QkVSb2wrdmxvYlJqR0Nhb1JHaDVUaDJBbm5sRTdCZllxdXZ3M3pELzZYczI5NEY2K3Z1TTFuWTRDOEVuVWg1SGpkMTE0NGZjOVorK3l4bUJXbkFkMGFNR3dJSkxrNE9LcC9VRi94UFFBSjBTOVNPUTFuVkpGeHFLbzVaRTNodXltTVI1bHZmSnY4ekF0azdyRWcyaEFlVWhXOWYvQmppWlV6Z1hJdjBQVVYyUVVhdmdjTGdicFZBSSsrUFdObW1iYUU1b3Zld2s1THg5V090RHFQWEFJN092TWtFcFN2M3lKNy9iOFhYa1dXSGVlSnUzZVNGejZjZnM3RHBaVzFtRzdWR3ptTzhua0JGeU4xL1h1ZzkvdzU3TWh4OHQ3YmdEMjd3ZDI3Z2RQT2dXWVptaThJR0NxUUI5UUNKc1FFWk1EQjFhNkxSWnRNbmRoRUN6RnpzcnE1cXNaZlFoR3FuVEgzMmdoeUhmenlMMnRYZEttQlhCWkpYU0ljOHIvOElCTDJNdFpSMU5vS2tUWVA4S0RvVGpobFRTdk5pZzBXaUVMQkxvWUszSnhsTk92ZmIvUFg4M29zL2dTWGRFL0J0Zm16TWtBQXVPWlNKQUw1N3NQNGZaMkV4L1k5Y25GMDNScG9XNFE2SVI5MWxvV1Urb0R1bHEzSEZYUTJBWW9kd0pHbUIxRWZweUtaMER2U2JkZVQ0Zk9KNkFSa2syclFNN0JqdC9DWVp4UHYvUm5nckZPRWhkZG5Wa091bnJBbUUwSlpRcTV4cXB1NDErcnpseVYzK3g2MUFSMGhLaFhJVWlBSGJXMEgvUHhkNFJrMkRzTm1CNEczZjBoODQyOUNJOGdDZVludVN3alV0RWRhV1FuUnc4bHIwaGtYVUlOREMwbXpzYzQzWXZUWGdFU2RoVjJ2ZXh5UGhVQzZqWWlKem1HMnVRNW1rRXBOSzFGUSsyQ3FrN1Q2NmViSG9MTlBadi9WVDNlNFU1WlFXNzNVdHM1YVNzaDNIeUxlY3AwNm5BMzMzRTVBNkpMaUxUeUVwcUdrUkdWSW1rbUdUajVyQlJrN1VDQ2hUR0hwbzV6ckdoNjdXcmh2K3Uwek04QnJneG43czl2SHF5NDh5MTUwdnpXdWFuUlpML1E3WFlzNXFkNjF1SHZnMm9NbTBpZ2lWZDFzOVZDQmhRbWt0ZjZSV2k4YTRDT3dtcVFiLzFRQ1lOYWh5RVZZNDNuZ05ZM1VNNzROdk82WEFJMlJ5WFJhMmxJNDFqcDNxb0hZWHVzMEdjMXFmMFZyc1lwRDIvNlhxbWhEOVlJaDloRE5nc3pQRHM4aStna3dtUUI3ZGlOR01vWHJoV0oxQ21XS2Vjd3VqaFJHRVhrZWQxOFBCdDRlamd3bHlMQ3d0R0JzMktFZWgrcVlEVll2WjVIZmVMaHpDL2NHVURLU3JwRE5xMG9FMkUzcHc4MndGM3l2ZDN0MkFjTUFUbnBVeFY0MHVtVUhKa21MUDNrcmZXc0dkcjJ6TEt3eFIxSHMxK3hrbSswSXJTeWdPREpaQVpCUnRCdk9UdEFvRkRCTlA4Smo3N3Y2dEJ2K1VyZUF3TlhibzE3djQzR2ZSUWdBWEFtNExrTzY4dGJGamJkdCtKOUVFV0FPQ3RNOVpnYW5yUm9XUjF4bGNIS0NxdUFLNlowbE5BbVdXZzVZVlVCaDloS3daMHJjZEIzOXlKMTBvMUtZUkcxR1JHd1JLdGx4M2tPQmgvNEw0UEJSWWFXdnhRUzN2MXB4MFpRelZicS8xQWlhWkxIWU1zYkRSeFhObGpaRTV4NWFNU1d2alU5S2dwdkRlZ2hkV0xWOEpNWUJuTStCWVE0dTVyVEZBaGhtd0RpRzBGVGxCR2tZVUtFbHdNQTZCSjExNENZVGEwbmQxL1FsamlsR1JrWmVyYlp2TDJiTmFMblRPT2pPNXNjVnVlRjhremh0bDlLM2Ywdk1MSzh0cU9EMmRhZ3VndjRuZjRwT080Q2w4TDB4RzBIclJadVdwMWJ4dHVkVU80WUhNTTNkRlJmc0lSNDM4YSt3L2o5NUMrYlg0SkpHbkg3MkJuamk0OE9IODI5dUxzQ1VCRG5SVGFBMGpkd0VCczV2eThJa2VOK1E0VWQ5N3FFSlBLRzlFY3UyVENRU0t5c3dIaGIrOG5VVmIxdHFvQ09reFdZalVvSXUvM0ZnM0VkZ0RObFVhLzlNQWNNRXZFT3g0WS9HMnFFbldUVkd0ZWtLRGJ0Y0dqRlk1OVcwM2w4eTBjMFlVd1hhalZQZk0zcGRGRnRoNGlzNFRMT1FyaVN6SlpoV0owbVlMWEhEaVA0QkpndFdKN1lwSkd4MTRyZGFyM1VVYjl3V2RMQXBWRUpkdGV4ZlVZSDZxYlIxayt6QTk1TW43NmNYaDZlNm1pRFFWM2h4c085UTFqZUlWNzlOSFU0QmZNN1l0QlRpMDJhcmxRV3BoVkZRUGtJTUpzcDFJOVVPRkt6Q1ZWalVzK3ZlYSt2clA1OXV1Vm9BTDhXMW43TDYvWXdNa0ZlalNPQXJieW12dmZHWXJrY0hVNHhQNTNRMzQwWmZBWVo3Q3NldEFxd3lkblltbUJLc0F0SUkyVG9pTTR1Y2tPaHFsblpxeis3dHZ4MUFxUm5NMEhvVHdnNVRvcndBSjk4ZmV1NVBBVGNmQTFZcTVkKzhXZldBRFg2Snh2cjZld005a2Q0YlUxOC9jWWVnNkpKLy9Pc3N4YVVOeDR4R0o2OVJFQjBxbUZHeE9hdHd6M0o4TUtLOTFxSW9kYUxoanloVi9wbHFLaExjT0pkaTNaanNWUTIycm9Fd1F4dC9wOXFMSFRWSjdURUo4Sm1nQ2pGZGd4KzZDZVZwVHhhLzdmbmdtR0VXU3cyVzE1T2lsZEM0RFAvNzlTcDNIeVA3VlN3M1NiRXljTFhRYUhsZkRVd3RoMUdNMGN3b0tOaFh0eU5SS0lOWitqQm52M0g3YkhhYmNKbjlkZDd2cnpWQUFNRGxzRCsrRzVzZjNoeC9mcHpDdWg3S2xOSU9ZYklXVHBwVFlYN0xRRXhxTjFVS1VRRGpGcXBGaUFJVVhxcW5IYzRDbkxRRDVmYTMwcTkvUjJCVWVZeXdzcTF0UVpQaTZNdGVLRDMybTRuYjd3VFdKb0F5WWxSRlhLUWxMVmZYSTRpc1JaRERySjY4RU1ucWhNS281cWFxUmlhcXZrN3R1aE9TeTlxQ25BN0FCTnNpaXhpUUxrc2lVb2FTTzVLVWFzOUtiZEVFazFPcE1OYkpBdXBMNEpsZHFXSUxYN0lmWnJKSUVZcTNrUi9MYlFGMURJaHFnenFVWVNzVDJOSGJZQmZ1UXYrYkwyTXlGeER0bElvR3BScUNvK0oyUU9Ndi94YVM5Z00rb0k1emErM3FyTk9mQUpSSWlHdjNvbEFzY0Q5eFFNRk9ETnpCZ2dGdXEraDRJemJtYnlpMy9DSUFYSUdyLzFyais4d004T3FvVUgvbjdmbTNiempzaDdCR0V5QW1ZcktYVW9Gc0JjaEhwZUZ3QVZaamJLUlgzcmJ1OGdTN01JWks1eEVkNjhCeVE5by9JdjNoU3dHRUx1REVnNnZnU3BSNHBVRGY5TFBRT2M4Q2JyOHJqTkJxaUt5S2FLWmFqVGNRUEZIV2hXZnhDSzFMYnhXaFcxWGt5cVVuWk9Xc0szY2RFMjJEbFdqcFJIaXVoaWNhcVNTbEJGblU4dlNrUmlkdjQ1ZFVwQ2UxaFJRV21qRzJJZWlvTjFMMWlHMzRaNGdPb3RRT3VLQVNJWlE0WFlIdXV0SDlmbE9sTjc0V09QMjBTR1ZTWW9nbnQwc0FsVUxyT3BVL2VTdjRaeDlVU3Z1Qk1sWW9zVTA2Y0d4REM2akpRaHVVRmhNVXh4Z3dqNVBxYVNhOG9PdjZEMkgyNmxlTlI5NHZITEFyUHczMjkxa1pJQUZkY3duU3F6Wng4S01mODU4UmFIMzAxTERiWVp5c2dNb0FKOERzeGdGWUVkRmJlQW9TVm1lNm9NcVBsK0czVVhiS3d1bDd3UnRlQS83bHEySDlKRm9vbDViWFVCT0RHY2kwQTNyQksrSDMreHJnNWtQdzNxWHBTbmhQa3VocVUyMHpvSzdtbXlsK1Y1UDh5UDZYdkRLMk9lYjJ0eEY2VVRQdmFPR0lteWo0cWNqOVd0aXNLMkxpK2RaVk5OcXdwQUZoOHVEQnczZ1J5SkRZVmo0QURjaXZxdE1LVEZuZ21oV1hWeXdKQXRqM2xCbks3UitFdnVZSnRHdmViSGIyT2VBd3hnSEVDMGJhUXl3blFBSGc4SlAvQlRidXM4WVF0M1FWYUVLRTluMmRnM2RDU0M1d2pDamFDZGNhcEN6SENqdDhFSnQ2UTNmM1R3UEE1Ymp5MHhXK241MEJBc0NsMTBZdWVQVmJoLzl4L1QyNkU2dXdRaFdac0xJM0RvOFRJQitUNXJkazJLNzZZWmVpQVd5clpscXVaV3FKZmFRS1ozVGlIL3c0TkM3cU5kaVdMMFl3VmdXL0hKcnNCRjd3KzhDWHZSUjI5eFE4ZENTNjBLYXBKVlQxRDl1RVNBREdXcVVIaW5JaWN3T3pOanlvL3F3YWtxbDZQaXdIZFlhVU9TWTUxL2szSWI1ZyszeFZDb0JhZ0RWVWlyQ1lXd0FUdkJKY1FJVlVvcmlBWUhUVitaaExOWlRpSGhaU0lpZlRVQk1ldmxQSWg2Q2YrQkhZNy8wZWVNWnB3cGhoWFY4bkRGV3pybjZJSlpOZHd0WXJmZzk2MDd1UXVsT2xQQWgxdUZDcmFsWGxWekZWdFEycENHTXNjQXdoeGVkZVpOVHRtSU94Nzk3cFcvL25WMmEzdjEwNFlGZTNrYXlmS3dNa0lGd08reTNnM3ZjZThwZU5Ec1BFWFgxQjJpUDBhNGhjWTQ5eDg0WlJlUTVnRXNhbERyU29OZ05ucVI0RmFFYUl5T1gyN1NUbTE4bGY4NVB5MUVFbG83WlpvUVdkRmhWTWdoVVhudm5qOGg5OE8vemgzeU1jM0NuY2ZnaVlIU002QkE5cnpiMEJiYkVpU3QwRVZBdkk3ZndhTGNkdTVoT0hHTkNpbkdqTG9LTXFyMk5VMFJxR1RFQzlsRm9PaU1ZU09tbmVwVUo4bGExeHlWMHFEcUpFZTFFUlVoWlFESW9tRndBVHc1Q01oNDZEdDl4SzRpajFMNzVTOXFkL3pNbVAvUkFzWjFncFF0Y1JGcVBUVFN4V1o5Z2pPOWdsbEkvZHFmSmRQd0hhL1ZWOElTQ2FuT0tEZTl2M3dmZ1VwWjJMbWt1U0E5eEdaTzVBNFJyRVF2Y081RWU2amZLMkhYZi9GUEhaZVQ4MFUvaE1uNnNENE9PdXhNNWYrSC82OXoveURKeGRCaWwxWXQ0RU51OTJhcEpRWmtYOXZoNjd2MlFLckRzOUJab1NnM2tRaXBnVHdPcmExa1ZtajJMZzVneDl5eHZJQno4WnlJT1VKcEZMVmNLZ2NxZEJIbmdXVWtjSFhFZHVBOS8xdTdUM3Z3cTQ1NzFBT1JxVFVWY1RNRmtCVW8rbHkzSEZDankzYUlmTmdXTFhIcHVXZDdlZUd6V0ppOXpqZWhVdTk0M0FxOVl6Mm1ObFRyakhEQ09NOWFhcDZKTExnajZRUmNORmpubnp6S1NOaHVEK1RjaWtieXlBalpsc25vR3RnYjY2QTNySWhkSS9ldzdUczc4U2VNQURRVUNlTTh5NnVKRmo4bFdkdzI0bzVrZ09vampSZHhxZThWeVdOOXdJcExPQk1xaEo2b0hvcFF0NG5IVkpnWmhCeFRaZ1lJQnpYcFhQcHlOempVQldHVmE3NmZTMytudGYvbzJ6OXo1UE9HREVsWjlSN3ZjM01VRG9NaVJlamZMYmw5anp2L1pSM1M5UGs0L0s2dGxCVzNjSjgwMmpyUWpsbUd2dDRpbFdIOUFER3c3dll1bEI2L09Gc1RVQnRYQXBqREtrQkN3RzRXTWQ4RjF2aE01N0RPRVpzRzdKU05hT3VUb0hsREozUVlWSS9YYXY3NTN2cDkzOFh2cEgvaERwNEIyTzR6ZkNaN2RGMndsVEdGV3ExWE9KVjQ3S0lzckVxQkE2cUVUMVFwRHlGSTBVRWxocTBwZFNqSkIweXVydVg1Ullxc0hSWWZPUTlHc3NNWUVxaXhnQjVBSUxRUWxZd0RJdlN2TkVGYUZzTFdUc3lITWZCSngwT3NvRjV5STk3Um5RQlE4U3YraWhySndZbVhOdFB6WGk0NFN3aHJoTktYZzFuNjduOEsrK0hmcU4xNkJNdmtnY1pvaTUwdTJjSXFRUGFCdnVJaFlVeE1Ed0FuR09VUXVBZTFGMENnb3loUlZRMTAzejVvdDA4eVAreitLdW02OEErWmtXSDM4akF3UUFIWUR4U3ZpMVg1UCs5TWtQdGlkaTdnVkpxVGgxL05iNjBYdlNqenAyUDJVTi9WN0taeFp0bldUZDl4RXZGWTNvd3JMakR4UlNSMnpOb01NN29HOTVQWERPbzhBeWhBY1QwS2FxTlJJTExYaTZHMVZRUURGMWJHWWxRT1h3TGNEUnUyR2I2N0Q1SWZEd3pjU1JtOENqaDRCakdSZ0JqQTdrQXBVRWprTnR1VnlYai9OZ09NWTVNTTVnNDhEYTB3SFBCY3dWNCtpVDJFMEFkb3g1d1NzZ1Y0aHVUZWgyVURhVnBtc0EwM1pMZ3B2Z1RxMU1nYlBQSngvd1VQaVo1MEY3ZHFxNzMvMkl2ZnVsZXJvTWlKNFRGNW1Ta0F5Vk1ZdnNBdEg3QmRFZGJxR2lwWmN1MGIvcngxaGU5bXZnOUdIeWNWZzI0ZFQ5U3FyOEJwdThvUUk4RU9nWnhBRG5GakpXVUhSYTNJVk1VaDY3YnZKYmt6dC81SVZiMS8rL1YrR3lkRG11L294enY3K3hBVjUyR2RMVnY0dnlQUS9Cay83ZFUvbzNuN0hQVldaS2FRVzJPR1phdjh1UlZnVU5nSlhFdlYrMUptUkNZeDMzV3NrTk1DcFB0bDVOSTFHM0I4Q01XTjlDT2J4YitGZHZnSjM3S0xJTWtEb2d0VmFMNm84VTNXRUdVeVhYSUM5YU5pd1NnblV0a3diQzVGdG0ySExnMkwvUzdBS1F5a0NPTXlEUGhMRklaVERMQTFXeVRIVnQ0NURyOG01VXpXQVhjdnZVQTExUGRoTnFNaFVtTzBKU0h1TnV2SlpmUE9HWVlEV0lBcTMxQ0xLYzNZdFhkSjZvYWlFQ1liZFZMMVI3WXdNc2NBY3dMR0NUcWR5QTRZVS9CUCtsVjdMckh3Z3ZpM3JTWGZLbG9wV1Y1MkN6aDJhQURtS0VmQlBGQ01lcHpKb1k1SVN2cXUvZTBLKy82eGxuWC9la3E2Ni9MRjhlTFplZkVmWjM0dU96TmtBQXVPb3lwTXV2UnZtRFo2YWZlYzRUK04zWThDeXlZMjlhdjZWd1dFamRDcFdQTy90OVBYWS9ZeWR3T0lkNHo2eWQvZ29nV3dXRUd4aEJSYUZnd21JR0hkcEpmT1hQZzQvNitqakxlUlJURnk4Z1NBWUtYamZCMUoybmtTaGlHUTNjVmJ4Qkh6UkdZeC9OWXFRM0ZYdkk2MXJyMksyVXVxWHdxNTFZT3lGenJTZXZsVGZMODFnOWxxd2FVbGdMQU9VWVRjTkFacFpyUzlBc3Z5bzNWS3Z1MkhjUlo4U01aSXhXTTBmRmhPSjNiVmlxb3haWkxsamZJOTk1RDhyMy9CdndkNjhWSmhjYWh3WHFXa0g1ZG50STlYNGhCcXJBSVhLMW94SFFETUtBZ3BOUXNOT2tZdFNxSjcrbEgvdWY3Rzc1c2wvYXZQdk5Wd0hwOHMraThqM3g4VGN5UUFIRUFmRHhWMkx2TC95cjd0MFBmNURPSzBma2Fack1zK1BZVFVWS0J2WlF1ZGU1ZXRHS2RqeHV3dit2dkRNUHN1eXU3dnZubk4rOTczWDNUTTlvdEF0SkZraGpJU1FMZ1FVT01tQ0JaTGFFSmNTV29JUWtCOHBseXNSVUdST2J1QkxYU0lBcmpnUEVqbTBjRjFVQmJ3UkxFRnNoRGlZb0NCQTJGRUoyU0VBb2xOQTJXbWZUVEhlL2ZzdjkvYzdKSCtkM1g3ZFViRm9ad2FtYW1xbDUzZS9kOTk2NXY3Tjl6L2ZMSGhQYWFNdTQ5b3U4MHNlWTJvOGdhRWl5TzQwS3N4bnNIc1BPTitBdjI0VWN1OU1keEV2bm9xazZvcnVKeUthYVdienlEeGJ6VFQyU2pmZXMvZnVJTE1qbmo1ckZqbkdGZVBZSmpVTG90VlFxakxxWk1UOTU2aWRUZmE1ZVJFeHdQVWpDaTJQaXhha1VKTG9KMldpa1lFZ05kWmo2Ym1MWUZxUzBjVVcxeFBBNUppbzZkS1Z1cHpYUi81dGU4OWY0MjM4ZC9lWUloazlEWm1NSlIrc1BWemFRYlBWZTZqczJBVFJ3endnVHNzL0FGekE1UmtyZFo5RGNKaDE4b05uendUZVB2L0dtWFp6ZlhQa2Q4SDdmemI2bk5zeERUY0N2dmduNUloejR6QTM1bC9hdVlHbFpTekV4WFJRV1QxQnNabTVGcERsS1dmL3FsUFdiT3ppdWtqZzBHODNlalcwM24wUDdhUVVHQ0dLd01JRFRkOEMrUDBjKytCejQ1RHRnZlMra3RwNFNobGx4RlRldHA1NnBSenZMVlFKL2gwREJyY3kxY3cxNkpscmlDNmtua3ZaVXRkSHcwOHA4YmdTbkM5ckVHRWhUSmRzVUVVMmlLZUVpb2lsUmI0eUtGMVZ4UTh5Q2Z6OEptM1RZaUkyLy9qL0U2VkZhMEhjQTQrSjZ4clRhajRxa29jc1JWSnFFTlEzZGpYL1A1SFdYZVhudG01QTdXcEV0cDdqbjlmaTFpb1J6VEtvUU4xN2g5eldob083M1FqMzFaaGlKem84STFud3lKUThhYmE1UEIyLzV3M1RQTzV4ZCt1MjIzUjZHTHoxeXUrNThtaGQvbG56TmEzamZxODl2MzhhcXowcWhrUllaM1ZHWXJpRTZySi9wSG1mN3E1Y1luTkRDQ3NFUGcvZUFsd3BpcmRWb2RONll1MGNtQUhObDR1eGRCMytxY001Ym5IOTBLYlo4QXZRaDBvcTRsMmo5U3pLUkhvV3Z4ZW8ydjRZZ3ExRWpYOS9qQnVpSCt4V05LZk1RYTBGU20xUTM2c3pxQ09FbDBlK3pVdEFxSTZtRVlPYThRbFh3VWlrZ29oZFV6ekIxTUJGUkZjSE1UQ1IweDJwZjJFUWtsQlBEOGFLZDVVMkRnSmZpVW03NEV2NmYvOFRsdzlmZ293R3lkSko3eVhncGlLVmV1VFFHN0RCbkFvd1BUc1hyMFIxSHVVcUgrUW9lWUFNcHZxelJNUnlTWnQ5WTdJYS9iN3RmOW52cmUvL25JeTA4TnR1amNrQUg0U3Iwakl0Wit0TTNONTk3N3BueXJMTFBzaVJKcUhQb0Z2T3V1S1NCQ0ROM0RnbmJMOTRxemJiR2ZTMkxOTXdUc3pvTGRhcFdXOVgxaVJ1elNzV0I0Tm9nbzFYWU00UG1hUHlVMXlCbi9yVGJxUzhWMlhKazMwSUcrdFpZY2N6ZEtocFRvSTQ5Wkw3WjBROExIUGNVNmhzMUd4V3hVTVowcjZsWU9HeXNjTlFZNmYyWjZrQlNEWFVtSlhJcWswcmNIZS9HNTZXOHp4ZnlvZWFBODgwT0Y0K0RLVUsvcWtyellPeXczWDhmNWI5KzNPVWpINk44K1NiU09qU0x4NHRwNDducktyUkcrblpQamIzeEljeHZIbkNyNHBaR3ZldTBrWldTdmNOWkRINGd3WnpXNlBLd0hmeEJjKzg3ZjJWMHg2NC80dHoyemR6WVBSci9pVS8vVVZwZmtQejJpVHp6OWE5cmJqajVLYVN5Z3FRQjVJbkxBOThzU0EvWm40RE9ZUHZybGttTHdCaW5xZGVRUk9pQjlQMTB0RSsxTzZqTitzajRzd1k0WVRxRGc2dE9FV2lPZ2VOZlFEbjZiUGlSODVCano4S0hDOGppZGp5MTNqK1ZiSHJQM3k3LzZLdlFRRmxiYkJxWTB5UGhOd01tSUJyajBjaWx3cjhsd3BzVG9Ob1lpWVJ1U1pvUENqZmxqaHZYdGVteEIxMlBIRHhvSE5pdjl0ZWZ4Szcva3R2L3VsN2t3SXFySHVteXRFMWRrNU5uWG9xS296M3JxcnVKYXBrTGZOYlNKMjRvZDdmYTRsZHIzZFAySkd2N2piRTdROHgzYUltajNhME1HTGIvZlhEb2hsZWQ5UFVYK2kwWDVZcDBmdGhWNzBQdFVUc2diSVRpcXk3UW4zLzVTK1VEeXlhemJrWnFCNTRtcTg3cWJlWnBRTHo1S2FLbWJILzlNanJBR1JQd3B0UlBROWxJN1d0SlVUT2lqZitibjRqZ2xldVhQTVBXUnVpNmd3MEZYY0FIVzVBdHA3dHRQUmtXajNCZjNPN2FISkg4cU5PZDVWTmhZUUcwRlZJYkR0M0VKWGk3aUtSV3BJa3Q5OTVUNmgvWmRHLzBYMEFmeGpaL29ITWNpZ0RGTzVqTm5Pa1V6MEd4Nitab3RpckRGaVBZTkI2N0hkeVA3ZHZ0OW4rL0x0ei9BSDdQSHRHLyt5STZIcnNjSEVGYUZsODYya1VHNGxid2twMGlNYjN4UURmRXpTTWhUV3hTWlFieFdtMVZ6VHB4OGVBR1RjY1BkSDAxKzJqTmFUQjJZRFRxMWdtelJZYkRMOHJvcnZmbHUxLzRNUTdlK1J1NFB0eUc4N2V6eDhRQkFhNjZpblR4eFpSUFhTcnZ1ZUJaK25hbTB1SGU2a0I4ZlkvSjJtNkxtYkhpdG80a1U3WmZ1dFcxVVh3YWtQNWVyVFErbGpvVG4rK09zZUdJUGFUWW95RkdoL2NvQXhFQnl4NUp1c04wQXBQT2JJWjRBWm1nT20wY0gyQmEwZFpUWEgyQVN5dWVGTklScnJxTU5RdGlNblJOQTBlSGdnNUZ2SzNsb2xEU0lBWXlKV1p6UGpQSENqcWJRUloxVmFTYm9yTXBaYlRxak5iZFZsZEVKbVAzWWpET3dpekhyQllKUWxOUElDcG1uYk0yZGp5SkRKYWhYVWFrZFdrR0toMzRyTGpYbWJHSVFGRVB3VGZCU3dncHUwdmx0SXlVMHlyb3dZUiszY2hSbGZiRW9hOGZjbGxmbVZGQXRtTytCTXkwbE1WbW9mdUtUcHIzY3Q4RmZ6clorM21ueHV6SHlCNHpCM1FRTGtMbGF0SzEvMXcrZWVHejVVV3NNU3NtYlJvS2t6M0c2QzV6WFlvTlhsdDExTldYZjI0YnpaTGdxMWxrR0tDRGVZTXMvcTZycUVaZEdBb0g3RCtDQUc1WVFOUzBOclA3MGxZRnhLejBuVGxRSzBwblVVV1cranhkZ2J4cE1XbGFpT1ZuZDg5RnBDdlF4U1NhS2RHcGpZMGMzRUJ6VGVJTEhpeTZpc2VOQVhVUHdacFUwYTBxNGdrdEdwdDFKQ2NEbHVMblRRUFVpRUpKZUJhWVJxVG94ZDAwTnhMWUNuVXBWYld0bjFlWGhQYzlGY1NsaUlwdHdqZXJ1cFdDcXVQSE5LVGpsM3gySU10bzk4UmRSQllwYkNWd0Y0Mm02ZjBMTW55ZjNmbXJ2ejNaKzU3cm9Ia3hQT0tXeTdleXg4d0JBWGFCdmt1eGx4ckh2T3N0K3Rubi9CalBLQS9RNFRTcGhmRzl6dWgrcDFrRU04RkhzZHE0N2JKdE5NZUwrSDVEQm4xdzAzcjRJVDIrYWVNRUpJRFFFRnpaWmVNTG9IWXVBcy9wOWZIQUVXQTlPeFc5cmtzNFliYWEyTlVsNWpuclZhcGMzTWpHNjNqOGJoYkhWU3pFSzZPYlcxS3NVaFp6c3FCZGpCZzlnMnZmNHpTOGdPUUVIVWgyTVZQWHJFSjJKT1BtQ3FKQ3AxRkZ6eEowVGJDNkZqRnlxdUxaNHQ0N3ZIc0ZTV3ZmblJjYzl5cnVWSGY0M2IyNEhKa2tQVzNnZWx4TGQ2djU2czNyS3FpM0ZOa3Fqb25RbXVmWm9HM2ZMM3ZlKzdicDduOTVYZXozVnY3Zng4NGVVd2VFalZIZHoyM2pxVzk3WTdyK25KMXlVcmRDVmtwS2l6RGFqWXp2ZDljaDRpTE94TEUxMlBhelcyVDRqQUcreDVEQi9OSnFTNmErNTM3QnordHBXTzkwaDBoazZqREllNm1Mb0N3UkszVnZPd1NlNGhleXlGelhPVk9STUZIb2hOaGpmUzREbVR1ZlJSZ3ZRSzdRelJ4RFZDbXh1b0loWkRFTjJKeDRGaVFIb2hvWEo0T1gvalUwbkxFRERiUTcwdEZIMG5pTnJHaVJhRVVWamV2TEdxZG1scXBkSENncU40R2E3MkZTTnpWTVZFU2tjNHFaeWFtTG9xY01rUVlmNzU2eGZuT3NqcmJpc3BRY00zem9kS1Z0aHgvVy9lKy9mSExudjNpOG5HLytMVC9XMWxmR2x4L0xNMy9sRGUybnp6bk5qK3JXckxUSmxRV1IwYTNHK3YyZ1MvVU5PWlQ3WVBtQ0pWbThZQkVPRkNkN0lKajdxM1NsSDVIVGg1ejZRWnYwZkkxYXNaVGVPNmoza0Nweml5L1NYVGZnVmhLOTJBelVVWFVVT0RVY2g0TWJ1ZlpxK3VjdUVwcUI5UVNWUXV3ZVYza1FLeUlhVGh3L0Z5cFRidWJ1V1VXS1JGcVJCYytRQ3ZoTVJFeXhMbmJ5blZoM0lJdElFYUVFZ05tTHVuU0laNG1id3FzVFduMU9FN0ZNaFRCSFhhNmpMTDR0b2MvZUJndkp5YzdrdGpHald5Y3VLSzBpaTdFSzZtMW1WdG8wL0hCYStiUEwxMisvekFOUy9Jam12TitMUGFKSnlIZXppNittN0RxZjVrLzI4SDgrOGxmZEsyN2FMZnZhSTBuRk1KL0FscWNKVzUraVhrYjFGd1NhcDRpc1hiZk8ya2RHc0UyRXJjU1hYVmNUZzZSYzRvOUlENSt2a0hlSitXamREWm52K0dwZFFFb2VJYTNmYnF1N3c3RVpSOTJtNitzWU5sWkpCV2ZPOUY4MytwSkVXN0R1dWxUVktFRUNHVVdTZmhPUTBBKzJDdStQZzBvVXJJcDlWeUpKS1QxQ205aXBsb3FzcnJ2QklmK2xkUzFjWXFrcFdDYWdybmJXNU5NdzkxaWRheUsxMEhHSG5MNUlldEVSeUhLOHp1ek9DZE5ieDRna0dTaXkyTVlVcERYcEdEYkRqemFIL3R2bDY3ZS8wVGNBRW8rTDg5V3Yvdkd6WGVmVFhQbFo4cjg2ZzNQZitFLzFFNmZ2NUpoeXdMSTZqUXpWeC9lNHI5N3RraFlKOXV4R3NMM3VhVmxsK2ZWYmFFNXM4SDFXYWQxOG93anhpR2FoMVNLKzZXTUtKSDhCNkxsUzNDbmk1aUY1VXZ1SnZmaFBmUTZ2Snh2MTlKTmU1OFhuNlBRZUlKd2R6NGprMkFIeDdKNU0rbE0wVUEwRjE0NmV2Z012NExIS2hoZEVTbWhmV3gvcU15TFRCTVVKc3BVNnBNbkJHS1pGaEZ5VHN5eXVXYkFzdFNpUjZLMVVnY3pnVVZaWTY4UjJKRS9QM2k3cGlJSDR1RGpxakwreVJuZDc1MDZpRWZlRk5qWlFXclJ6MWNGZnlxRnJmbVp5eDhVS005dm94RDV1OXJnNklHejBDSC90T002NjVETDltM1BPOEpQS1FYS0N4QUkrMmU4eXVyWE9RQWJFOHZvYXpnalo4cEl0TFB6VUFvdzhXaXFwZHVKSzRVRWlQaHJ6Q25WNkpZSDViTW1LdS9ZTEdBWldUTFIzNUFlRlhkK29pbXVGQzRTRFo2aExZN1ZvaWI4ZHdRdXV0UjJDNFZaY3lJSmtFVHBIaWdUcUdyeWd2UU9qUmZ2bmNzK2dXWlFjenhjTlJvV3VDbUVXUllvUW9WajZuTlVEbVMzUUJaMkptK0NySlNBWno5aUtQM3M3c3BZbFRiS1lDYXRmUE9SbGYwWTFlYXZPc0hGRHhKcVFlRis4Mmc1ZWMvSGs3cDlWeUw4QmoxbXY3enZaNCs2QXNIRVN2bjRMWjczOXpmSlh6M21tN0xSRFpESkpCNURIeHVyTndSZWVGcGlQSXV3K0dQem9nQzJYTEtQTENudkxSdEpRYWtIc0Rxa0hrTVR2ZXFaSzZsWm9sc2RZeTRyMDJpc0JMTW4xQmkrK0FjRjNvS1BtZ0ZISWVIN1E2c1M4YURFWDEwS29HR1RGQytZNVZuZnBSS0pRcU5lalFpa0ViMHpOL3lndTNvbExCaTBpM29FWDhaQmJFQ3ZGNHhxTGlCZXRVcjRpRWp6bzNrdUZlRzdjMTdKN01aVlRsdEN6dHB0dlRjS2hRbXFFZlBkRTFtOVk4VEp4a2liYVpMUXRBTmFZVHIxTlN4L2gwRFdYdlB5ZW4vR3JzU3ZnWVNPYkg2azlJUTRJRzRYSnkrR2tkLzl5K2kvbm5pc3Y0RkRweXBRbURRT3hzbnF6TVZzQldZeUlTd04rQUFSaDhaVmJHSjYzQ0d2dXZsS2tIeG1MMVIzZUtEcGtEdktZczBsZzBUcUJCeFVTL1g2SWJ3cS96cWFLbU5qRE5vdVFudXZQbTlVVHNvcWhSbmgyQzBkQmkyT0YrUWxJUnhRcUVoby8zdldhMTRKbnhEc0p1dVBPZzdTbzlESVZjZnA1RnBlczR0bWRJbUltTkoyNnU0aVp1Nnk0MjB6RlR4b2daMjlGZHl6QVhnTnhiTG1oL01NcTNZMkh4R2hRRVcvYnluSlgzQnVWcnJUTjhDOXQ5Uzh1V3IzclVzZkxGWThBVnY5bzdBbHpRS2hPK0ZFS3p1TGZ2a2svOXBNL3FhL0FTMWRXVUIxSWtvRXp2czFaM3cwNkpIaFlGSHdLdGcvYVUxc1cvOWsybXBNVjltWjhBalFTb2pnZXhMWjE0U2phemlZUmdxRTZYKzlvUGZxNkFsN3lwc2Y2dkt4ZlVLb2hNVUt5U0lSZklsY1VDZlNtZ1hXZ2RSem1CZkVPbHc2UjR0QUoxclA3VnFsZmNoK0M0N1ZrS2lKVkY3eG5ZNUZLL0VBbmZYb1FtbWtUaGRVaVZuQk9Xa0pPMjRJY09SQWRGN2QxZDVZRzR1Tk0vc0tLNTl1bkFSM0RKWWxiR2dZQll5TWk0NlRwT2x0L3p6OVoyLzFyQ3Y1RWhkM045b1E2SU1DdVhlZzczNFc1b1IrL1NIN25oUmZJVzdkdmRiY0hLSWdrSFRqZFBoamQ0VkxHMEF6cjRaYkFEdUsrRHNQekZtWExQOTZDTGlnY05NaFlzTHFKekFWaVBKcko0b0VqVUFjcnhBbG1ncGozRkUveHB6cWhlKzM3NWRxRzhkcHpLL1ZVN1I4ck5lQVhpV1dqQ00xT1JyUkl0SkVLZ1VEdHFxQkJ4bU1McjM3d0dmZGMyVHRtRWJLdFFHQzNwRUlZRWJLaXB0QWhyRG1leGVYNG9YRDZGbVM1eGRjNkdCVjBJWGtaTkhRM1QzejJ4Vlh4VHExWlVCbElxYXRLZ2c1a3BvbUZiM2FkZmNHbWI3bHNiYzhmMWZGYTMrcC9RdTBKZDBDb1k3dmFZdjd6VitqbDU1M1ArNTkycW14aG4zV1dwZEVGaCt5TWJrT205K0xhSXQ2VFBEbjRYcENCc1BpU3JTdzhmd0ZwMWV5QWllYjZjMTU2SjZQTy9CMExsalBINXVGWStrTEdMRUpzMUtsQ3JoT0V2c3J1UTdTSlJGVk41SGVpdGJIczRrVmNzN2gxTHBpNDVORGs4Z3hTbFZBbGk1VWFnc1BCUEY2M0ErbnF2RFlIZlkyUnNPemdqazZTTUlwRFVZNGZ3b2xMeUZJTEk4TW14UmtLT2hEeTNaMTBYMTZudTg5TUJna1hHSWpUTkFFR2F0RENVTnNiYkh6Yk5ldWpYL2pONmFGcjYzanRjV2t5ZnkvMmZYSEEvclg5S2xRdXB2eTcwL2p4QzErVC91emM1L296R0pQTG1tdEtyZ3loMndmcnQrQjVIZEVsQWxDWEVLWmcrM0ZkVnBaZXNNVENUeTNDVWhMYm41MUpRU3V5V2F6S2I4MkhvYUJCYlJLRmk3TlJoQmh4ZXBhK2lBbU1nZGJRN0NZaXhlcmNPR3FBZWRzbW0ydFJwNHM5SmEyT1NVZFV3cTV1V1NxQUlJcUpQcGNrdTJoWHcyME4vZTdKNmNCbmpwUWtIRGxFajIvZEdoV2RwSGd2QTNVcllQZDJkRjlidzNZYjdvMzdNQ0Zta3BKNG85SGVURWw5SkNWZFg5WS90V3QxNzV1K0JIYzlIclBkaDJ2ZlR3Y0VOdG8wWjhCUnYvdDYrUS9QZXdHWGJkc0tqRHk3MGNoQ1NJeE9kOFBvdHBqVjZ5RFFNWjRFSnU2eUI5ZHRLb1B6bHhnOGY5RjFHNktyUlh3bEhFNVNjTXFwQjBpbTlnUEZiU05NQjNBQU5sb3R4SWtZWVZxODVvWFNGeWxSbElUT1hSR1JMc0t4RjlTenUrWTZQck5OT1dBaGRsMnlPRmFuTmxFTkl6T1pNMkpZa1pCaFYxR1dtOGpwWEZ4TlJGb0pzTUtxVTI2YllyZE1mYlkvY05yU3BnRHVDelJCaldkSlBHdlREdS93am8rUDEzYTlkZjNBT3dFZXpTTFJZMm5mZHdlRUtFNWU5ekdLRzd6dmhiemhGUmZLNzV6eGREK2FFVjBaYTBwdG5JYmxvRE8rSFdiN2lIYk1vRDZCaXR2RThZTWdRMkhocktFc1BHK1I1dVFCREIxV2NVWmU5eUdMR09Mem5tR3BSVWp4ZXZKc252MEN1Sk5GdkZRT2gwNmlMVk5IY1pLSkdYQlV3RzRaQ0ljVXFXMllXdmg0c1JxT2N5WFhLb3BWc0lOYWp3UVhOeFJTcXR3NUVpaXpoUGpFdmR4ZHBOemVZZmNWYkFyZXFuc0Rna2dTOTJhb3JsRnhsNEdJNUliMmM5UHBMWi91cHIvNG00Y09YZXVnVnhBTXVFL29sL3h0N0xCd1FLaDVZUTNKdjNRVVoxenlCbjdyMmMva05RdExBcXQwTnFYUmhlRHd5NGVFOGUxTzkwQUZhVFc0SkVSUzVHYitRQlFEelFrTjdZOHRlSHRtNjgxVG1tRFFHcHY0eUpDdW42d1ljMlJNWDRUa0dwNER6TERSKzZ1b2w3bFRsZG9vdGpsQXdhd2pKaTZaS0JwcTQ5cTlOcG1MdUJ1U1FnRkxyT2VGdGlET2NRVnBVakR3QzI3aklyYW5VTzRvNUhzNkdBTXA0WU00NnNvc1NGemJSU1F0aUh0UmJ6dWRrWFRoL3BUNTlIVDZvVi9Zcys5WDEyQmZiZVYvMzArOXpYYllPR0J2ZmI4UTRFTXY0ZEx6enVQZHA1OGhwNUNCTlRMbWlRVVJFdVNETUxuRG1lNkpxbHFHb0FtWEZBV0VqZDFaSlJqUmptMmtPV1BnZzZjUHBUbTVnU1VSWmppcldSaDd4ZmI1eG5RRjZIUEJ3TmtCZGJXNWQwS3ZqaWxGM0l1TEZKa0RGN3hEcEdPejgxcmY0aWtGU2JsMjFHdmtWVzFpU2FVenlncVUrL0Z5YjhiM09iNXU4ZGdnbUkwd3BIVEI4dGEwUXJzbFZxeTl1RFdTVWlaeGcwMi8vcW54K2p0MkhWejl1QUIvY1ppRTNJZmFZZWVBRUxqQ0s2Skt0dGZDc1cvL2VkNTIyazc1NWVOL2hBVkdYbXdpa0ZBZElpaWVSODUwTjB6dlFXd21yZzJpQS9xcFNZVFRpWXV2dVdtSHBLMHErdFFCNmZUV202ZUtwQjBORERUQzhOUmhBb3pkWTViclhrZGtVYm5QeDNZQnJBa0hqTHhOcTVOYXpmczAxOGExRVhKL2xlUkxGYWNrckVOMDVIUXJCVG1vWW5zTGZvOVJSa1pSY1pGZ3ZDUnBCYzRpbmdNc01XaUVkbGlwYmt4eUFJZFNlOWZNMXE0djAvOTB5WjREN3dKV3JvSjBFWE5FNVdGbmg2VUQ5cmFwY2MzdlBaMnp6MzhWLy9ya1UrWGlJNDV5WVUwSzB3QUk2OUNkUnNSSEx0MCttTjBMNVZDQURhekJwVWU5dUxoYW9KcHRCTXlDOXlodFVkSXhMZW1rQmpsTzBLTWJkRW1RVmlwcmd6Z1pwVVRMaEZuTkczUDA2ank3U0ZHZlY5Tk9CVFRVOGQ3VTNNYmdJOGNPdXRnK3QzSW9paVE3WkY3R1NGT0NMc0pWM0Z2QlUvaXVUMEZ6Z0crYlJxUmRFaG9STjFPVDdKWlVoUzAwdDgzTXZ6SHBQdktGbGZWM1gzbGdkdFBoZk9wdHRzUGFBU0Z5dzgvc0lyMzR5bWdYZlBUVi9QU0paOHF2bjNFYUZ4eHhwTURJQ2hPS1pXbTBSV2tjREM4clNIYy96UGJpZVMwNGZVU0RKTHhYelZScU1URURuNENNNDk4QXNxam9vcm91Q0xJczZHSVNhVUNIQXNNNkJpeFZ0a1hCVGFMSlBIRVl1OWpJbllNdXRtcjRxdUZqb1V4cUcwYTF3ckdRWGljdkppTkFLRDNFTks2NGF3dURCWlhCUUpCR3NLS3VVeWxxQ0cxS2QzV0YvMjNsRTE5YW4vM2JkOTI3ZGozTUs5ekhGVWIxV05saDc0QzkrUzZVc3hDNU9PN29EMTdJeTg5OGpyN3R4T1A5cFNjZVc1RnJxNVF5RXhGMzBhR0V4R3NSNzFhUnZBK21leDFmQlI5VGszMUl3VnNkbkQyQmNJNmlvZUF5VmFTNFMzRXRYYlJCMU9vK3JRV2tDaU9VNkJ5a1ZDaThWV1pIUWFzdWlwc212Q2Q0OHlDZ2tReWVGWnZoT2xPbE9GcWR2RmtFWFJCUFE0azlsTEVVTlJ6WFJFcHlkelp1R3BkUC8rMkI2WHV2ZkdEOWZ3QTRwQ3ZBRDVjSzkzdXhKNDBEOW5iVlJhU0xycUtLVThIdm5zM3p6ajVYTHQvNWRIL2x5U2R3TW9zQzYxNXNFcWhuUnpTMUFpM2c3aloxc1lOSTNvOTNCOERXb3FuZDd6eUZ1S0pBZEQ4cUtqU2NyR2RXa1JJZ0hDOUJJQlJPRjl6a0dLRXlZYUMxUXZaQ2xTQVdKenVXZTNhdVVKQlNGWm9rTkZzZ0xWWUtKQlAzRE41aGpZdUR0bVRodG5XZjNsWGttbXZXcG4vdzNydldQZ2R4YzE1OEpmSndxSEVQRjN2U09XQnYxUkhuOU8xdmhXTmVkU212Ty9rTXVmejRvK1c1Unh4WngyY1RNak12WnBJd1FuazhvU2dldklCSVdjUExBYVFjZ2p3Qkc0TjFtM3FCbGE5TWNGRVBzRURzbmtpUE0rd2ROTGhzaXpnNVNQcTlCQ0dyTkpCYWNSb2t0ZXBwRVVuREFGdGdTaHByREsybm9jK3FqY1RNQjBsN3A4WTNKK1VyWDM4Z2YvSVRoL3lQcno0d3V3bkNnNjhHT2R6enZPOWtUMW9IN08yaG9SbG8vdmhGemZOUDNGa3VPZmxwZnVGeFIrdHAyNCtxUmNPRXd0U2RLVzVWTW1lRHVMeUNEekxZRFBjTzhURXd3MjBtbEJrd1JnaWtjK3hKWmNHRFdCT25vdUlCQ2FJNFQrSkNFblRnMEdnMGR6VG9NcXdUZk9wSUZxZWo2TFFSSUNFcUpHRjlWdmpteU5mdm5IRHRqUWU2RCsyNmEvWTNSQmNRdjRoMDlkWHdaSGE4M3A3MER0aWJnM0FScWgrbDlFdDBwOEFSLy81RnpUbExPLzJTblNmWlR5d3R5Yk5PM0c3QjRsK0FkYUREbUZId0lKcXFmSHVoeDlZalpWTEZHbGFHbjRvYjhlQjAzclFHWUxXUXFPTThuMFdSWXdZK0VhTWpLRmdNMTZKYTIrZENWcGdrN2xqTDNEMnlXLy9meVAvT0pWMy9oN2VWYTc4OG5kNEtOU3BmUkxyaTZpZFhqdmZkN0FmR0FUZVpYSFVSZXN5WnlJdmZPZWRhQkJqKy9qbnR6ck4rdER0dngzSDh4STdqT0xjb1ovL0lNYlJwa2FnWDY5STV1WWJuRXUzajJrNkovWkN1eWtOWEJGYUVhQUhIZXp4Z3hRMEc3M1ZJQ0ZaZXVycFdPWVhaREc0OUpEWmV0N3RYVnVUenQ2ekxIZmZ0bjE3OWIvWnhLM0N3ditoTnA5MlRvcXA5dVBhRDZJQnpjNUNydjdVekFneC9jUWZIdnZhVm5MMmVlZkhwSjhxTzVVWE9IaHVuTmkzYmpsaWsyVEowSGZUS25BbFlkNWpWZnp1eHAxSnArekJxZ0NRU3V6SE1Wb3o5NjlqS2hDNmJUdHBXdjNwZ2phL2Vzby9WUEMwMy84T0tmZTQvM3NvKzRJSE5GM2JkK1RSN2o4Vy85Z04yMm4wcis0RjJ3TTNtSUZmc1FxNjRDYmx4Qi9yY0R3VFp4a09zZVI1c2U5bFJMRno0SW81ZWdaMHJhK3hVWmR2MlpacW1CQ0hEMWtVR1N3M3FVMXhVa3F1ejNrblpzOHJzNEFTYlRMMlVLYk1kUTcxejZPbnZyNzI5Mi8rMXU1aDlIdmJ6TFJ6S0x5SjlaZy95bWM5aVYvVExCVDhrOWtQamdBKzEzaUhQdWdrNWRRZDY3Z200dnBQOExaenlNVE1SdU9ISGFXOEVUdCtLL3pBNjNFUHQvd01RazN6MjZRTDFSUUFBQUFCSlJVNUVya0pnZ2c9PSIgYWx0PSJJbnN0YWdyYW0iIGNsYXNzPSJvcHQtaWNvbi1pbWciPjwvZGl2PgogICAgPGRpdiBjbGFzcz0ib3B0LXRleHQiPjxkaXYgY2xhc3M9Im9wdC10aXRsZSI+SW5zdGFncmFtPC9kaXY+PGRpdiBjbGFzcz0ib3B0LWhhbmRsZSI+QHJqX2dyb29taW5nPC9kaXY+PC9kaXY+CiAgICA8c3BhbiBjbGFzcz0ib3B0LWFycm93Ij7ihpI8L3NwYW4+CiAgPC9hPgogIDxhIGhyZWY9Imh0dHBzOi8vd2EubWUvMzcyNTg3MzU0NTYiIHRhcmdldD0iX2JsYW5rIiBjbGFzcz0ib3B0Ij4KICAgIDxkaXYgY2xhc3M9Im9wdC1pY29uIj48aW1nIHNyYz0iZGF0YTppbWFnZS9wbmc7YmFzZTY0LGlWQk9SdzBLR2dvQUFBQU5TVWhFVWdBQUFLQUFBQUNnQ0FZQUFBQ0x6MmN0QUFDT2lFbEVRVlI0bk95OWQ0QWxSM1UxZm01VjkzdHYwczVzRHRxVlZscmxnQkxJZ0FBSklSQTVha1dPeHNZMjhBSEdCaHVEaGZEbkJOaGdrakVZYkpJQkxXQnlNRkZFZ1FnU1FrSloydFhtT1BHOTE5MVY5L3orcUtydUhrbllZSUx4OTZOaHRETXZWbGVkdXVIY1VNQnZydDljdjdsK2MvM20rczMxbStzMzEyK3UzMXkvdVg1ei9lYjZ6ZldiNnpmWGI2NWZ3U1gvMHdQNE5iNEVCSEFSVFAzSTNqaGZxOEM3L0x0OW5RamlFaEM0aStkK2MvM21pcGRnTXl6T1FZWkxZWEhwWmd0ZWJNQmY0TVlrQktTQUZ4dGN1amw4MTJiWVgraDMvQysrL3Y4MUNSZkQ0Q3N3ZUM2SXpSY1Q1aEw5U2ZMcEhDQzc1UWtiVmkwVWV2akNnaHR4ZmUzQ21Cd0Nva2RqTzNrR01RN2VBeDcwVHRUMlBiM1JNcCsweGVoNFovK3lXK2QyM1BTZGc3TS9jVHlrQUFKc2djR2JJYmdNSHY4L2s1ai9id09RRUp3TGkzT2grQXRSNk9LMVhic1dvN01uTGo5TGwzQ2xuanExakxtNUp3Wnl2SFJ0ejA5SUQ2TjJ5b2haelk0SUlLQVZVQVFpaEFGQUNWOENDT0FBZWdVOFlWVkFkZlBxdUE4RG5URkRKWTI1T1JOOHBiaCtkcjlNdTJGbnQvdnU0UEtET3hZTlNBQm9hOHlYUUg5Rk0vVS9kdjIvQjhDTFlRQVkvQVhjSFpkdjh1RnJ6eWhHZUNKT21ieVhveitISy9NeEdjczJZc0tDSFFFeUFjVUNDbVF3b0NQZ3FDSkNBYWdnSUFGMTFvTTBDa0lrUUZHOFFJUUNNVXBERVV0TFVBVG9DSlNFcVFoVUNsTjZjTHJhcTN1cmZhYmtIclBnUHFXM0RHOHVQN3JyQ3dBVzZnRUxnQXRoQVFCYjRIOUZNL2dydmY1ZkFhRGdZdGc3Z203a09lc1BLM1lXNTNlT0dyMDNsM2Z1cHhzNlIyQkpQaUlqRnJRR3FnSlRRU25HczFMUU9WRjFBbGNSVldtZ0RsQVZxQkJpVFB3bWhMOEJnSUFoQS81RVlBUXdJQ0JSTUZyQ1pyVGRqR0p5TVZsT1pDVERJRHZNQ1lGQ0tvTFREdGhiYmN0bS9QWEY3c0hIZFZCK0ZlL2VleldTU2pZQVhuRk9CbHoyLzVSay9OOE53SXRoY05KbXdVVmJhdWt3ZHQ1aGR5dU96czdYRGZsRlBMNTdESzFaWm1oZ0pqSW9GUWFtZEtVWGxKVkJNUlM0Z1lGNlFBd3l5ZG1URE10N1UzTDQrQ291SFpsQVQwYlE3WXhJdHp1Q2tVNFhIWk1UQUR3ZFBBbWxGNjlFcFk0RFgyREJENFMrNUxBcXVXOStXdmIycHpIcit5aGRnUUplNER4QkEzUnlJdTlCT2gxYUMxVUxBNEVWQ0xUd3dMd2ZZcisvdGJPMy9JUy9kdUhUN3FPN3Z3NUVLY2lMRGM2OXhPQXl1UCtaaWYvRlhmOGJBU2pZREJOcERnV0FTV0NxK3UzREw2eU82ajNWcit5ZWhWV2RIbklDVG1Fa3E1U0FGbjJMcWkrbytvVE5zUVFqNW9qSnRUaDVhaVBYTFRrTWF5ZFh5cG9sU3puYUhVZVdkNWgzY2dFQUR5OGVuZ0lScjBxbkhpcE9GQjZrS0VrakJqQVFWVUNNQ0dnQWdaQ2VHRmFGS1gzRm9pcTRkKzZRYkQyNEMzUHowM0x6N0c3ZU9yMGI4MVZmSUo0d1BVR240MDIzQTJNelQrODd6QUVMQzkwOUJQY09yelk3cTM4ejMxLzRTUG4xL1RjQUNGTHh2c2orTnpzdi83c0F1QmtXSDRKUFU1MDlmTTA5N04ybm5xSXI3VVZtUTNlTjYxcWc4TEJxU3ZVVVYvVXRCdE1HSGx6ZW1jTHBhNC9GeWN1T3hTbnJUdUNHOFpXU2pZeUlXRUdGQWVkUVlOYjNwWFFsblMrMUVscFBoWHBQZ3FDQnFCSUFxVlNvcUlBQ1VoVUNFU09FQWlRQmdTZ0FDNE5NUk1RWVdqRXlZam9jTVYzWVRpNm95S0lZeXY3NWc3eDU3M2JjZEhDclhMUC9kaDRZekFZN014K0I3WFRWZXFPYXc3SUhZeXBDZGd4bXNiMzhqTDltNFlQKzMzZDlDa0FKQUxqNG5BeVhYUGEvRG9qL093QjRNUXhlZVRFZ2x5Z0FhNSswNnY3NWt0SGY1eWtURDNkSDV4MXhnTlZzeUZLenNsb3dLT1lGYW5qVXlDcTUvL3JUNWN6MVorR2sxWnZJTHNTandxRnlubk1Zb3VCUXZDdmhCWENpSkZRSWtxcFVvVkFnWGowOUZUUXEzaXRJalRpTFpoZ0pXRXNERlRKNHhBSUJvNzhpQWxBbFVDNEVhUUFRc0NMU3NSbDZ0c3R4MnhNakZzT2k1TzJIZHVMYVBUZksxVHUyY3Z1aC9hS2dvdE5EMXVtUkhmSG9xUldJTlFjYzVOckJEM0ZUK2RiaWcxdmZBMkErek5YL0xpRCtlZ053TXl3MmIwYXk4VWFmZU5oRGgvZGM4bkpkYWUrRjBRd21OeFFWNTUzUFVNNEN3eGxaTjdvUzU2Ky9OODVkZnphUFdyVVI3RUJtTU0vcFloWnoxUUljZlBBUlNGRW9IRDFVUER3SXBZZW5oeWRSMFlIVW9OdW9JQWhGL05lVFFjY3FBRkFrT2lFUmdRSWdPTVFRQ0JDd0hJa2JFVWo4cFVaSWtLd3dZcm1rTzRiUnJBdDR4YzdwL1hMOXJsdnh2ZHR2NU42NWc0Q3h3T2dZalRGS0F5QXptZkVHY3V2Z1Z2T2p3VDlNdlAyV2R4NEE1bXJ2ZVFzVXYrWkEvSFVGWVBCcUx3bEdkdmRocTg4MXAwMittRWVQUFh5NERqQjlxSkhjdTNJaFEzRUlvNXJyZmRiY3pWeHc5SGs0N2JDVFlMc2RUT3MwOTFZSHBIQlZBSTRFUjBHcEtId3BhbFFVd2FiejlLS2k4TW9BUi9WUWtFNDlHQ0FuQklMdUo2RWdoQXh3UzZCaUU5c3dJc0lnOXdRUU1XTGlhd0dSZ0VDQlVBZ3hZZ05Bay9BRVNZSkdySXgxT2pMUzZkSjVZdHZlM2JqNnRodHc0NTRkNkxzQzZQVEU1QjJWM0RoMjBFRUJ5RzJEVyt5MUMyOWUrcmF0LzdRSFdBQWhlQ1hrMTlsci92VUQ0R2JZeEhuMUxsaDFMNTQrOVdJY00vbzRQYXdMbGw3aFJGMnhZREYvQ0llTkxNZkRqNzRmSDNUMC9iRnlhcFhNbTc3c0t2YXg4aVVnWkNHVmNYU292R1BGQ2dXOWVEbzRlbnA0S0x3NDlmQ3E5RUx4VUhwVjhmQlVwU2dKMHBPZ0JGRkNVRUVhRmROb1lNSkE0dE1DZ01hSWtNM0RFSU1BT1ZCRVdzSlFZQ0VRdFFHUUVzQnJEQUFhZ0tSVGhiR0NpWHdFdmF5THdXQ0E2MjYvVGE3Y2VqUDJ6YzRRblM1TXQ2UE1vTktUM0hnRC9IaitPdjNPN04vb2xwM3ZBYUM0R0JrdStmVjBWSDU5QUhneERGNEpRS0NyZ2JHRGY3RHVyN083cmZqOWNtTW5NNlU2MFk1VzFURGovRTZ6M3F6QnMwOTdITTg1OXQ3UW51Vys0b0JNbDdNb3hZc0k0Y1NqOG81RERPSFVpYU5Ed1FxVk9ucDZjVlE2ZWhBcVhoVWVIbzRLRGJRS1ZRaXZLZ3FDOFRHSWlGTFRFaVlWU2lnQkUwTnFRU0xTaU5UekdzU2RRR2pTQXpCR0FCVVlRd2dNaEFtcUFndWhFUW4ySTBNK2hNVHZWbERIcENzVEl5UGlQSEg3bnQyOCt0WmJ1SFgvQVVIWE11dDJWSzJTWXpZM0J4eXk3eFRmY0Z2blh1ay9zdk1MY1k2enBGVitYYTVmRHdDMnZOdk9NdzdmN0U0ZmZSV1BIVDBlamhTdlRpRVdNM3RrdVl6Z21YZDdOQjkrL0lQaHV3WmJCenN4NytmaFFYR3NVTEtTU2gwcWVGUytvaGVGbzVkS0t6aDRWRm9pZ1UrcG9nUThGWjRhN0Q5SklBelduaWNKVWhpVFdsUmorSTBhWFl3bzlBaElVSjBRRTZnWVJLV2FzQ2h4cXNXZ0JseDhuS1JJaERNelk1RkVaSWpNTVlJNGZJYUZvU2NoSWpMVzZTRTNSbmNlT0NoWDMzZ1Q5ODNOQ3JvOWlCaGxSbVRqdWVXZUV2TEQvai8zWHIzN1QrY3h2ejk5SEg1TnBPSC9QQUQvQ1RtZWcyb0NXTzVlZnRUcnFwTW5udW9uRGV5Q2NWWXlVeFQ3VEtmZjkwODgrc0h5dE5NdkVydWtpOXVHTzduZysxVFE5SFdBUWt0V1dvcURvbEtIa2c2Vk9uZ29uSWJmblRpcHZLZFM0WUxMSVE0SzlZU0hnZ2hBWkJCcjhGUjZJamdJd1RhRHNoWC9KYUdCY2tGQ1NqTHpyQmdJaEVyQ0JBaEcvemdnVWhCanlpQU1ESVFTTUFqU0dpT2dSQ294T001UVNjWWtNcHJ3UFVZRURQeGpKOHZFWkJZN2R1M0R0VnUzY1dGUUNMb2R5YTMxMmdIUXNWYXVIOXd1MzUrK3VQcVg3ZjhDQUxnVUZoZjl6NGYzL2ljQmFLSUE0ZGdEVjU5WFhURDVEajE1eVVaWGV0Y3BMYXFxdEp6ZnkvdXNPbGxlOUZ2UDV2clY2N0YxdUVQMjZReW9oSmRLRjNRb0MzNGdsYTlRc2tJbEtxV3Y0TlFITHhZQmJGNDlTblYwOU9KVm9WUTZxQUFLVlFhd0dSVVNVZVVxd3Y4RVVGK3pLMHBBSWdnWjhTYUFrSXp3RExaZkpqWUpTQmdJS0pGN2liSXVTcmdBU0FoQUE0bE9zVFVSb3ZIVGlTUTFRUlBnR3A4SStEU1NBU1FFUkNmdndIdkYxcDI3Y2RQMlhlSWxvODB5ME5CelRETGI5OEFWYysvdnZtUDNpeGIyTHV6NWRWREovek1BYktuYzdBK1ArRE81Mi9nbGJrM1htbGwxMW1hMjdPL0dLaitHbDV6K2ROejNidmZqN3VxQTdCc2NoRnBLWDRZYytnRktkZExYSVlkK2lESkt1VW9jS25WU3FZTlRoUmNGTmRoM0pSMmNPbnBTQ0sxdHZrUzFKTWd4aU1BZ3pNUUtFTlFnVE1oMllWRFpWUHJ3ZnFFRWVzL0Mwc0tJUldhc01JSTFDRCtGSndWVVFJVEdHSUVISmFHSkZnR2RrRXlNR29SbmhJRlRoSW5TazBJREF3YmhpU0JMZ3dFcWpCTFhpUFI2WGZUN1E5eDh5eTd1bTU2RjlISVIwbkVDVG1CNzVrZnoyM2pON0xQOXYrNzZQTGpaUXJZUStKL3hsSC8xQUR6bm5BeVhYZWJHTWI3Uy9kbjZmeXp2UGZJNFZVOVRHS280ZzdtZGVQajZzL25pZXowWGRyUXJQeDdlek1wN0tEeUdMR1NvQllhK1lLR2xGRnF4OUE2bFZxamdVRkhoR0FCWWVnY0s0ZFRScVVjRkZWVlBoYUtDZ3hMaVZFRXFJVVlBZ1FOSjllanJBQXUrTDVYcjB5aEI5YWd3RkFBWWx6R09TSVlPZTlKQnpvN2tVcGdLTTM2ZXM1eG5SUzhkNW1MVVVITUQwK21pSXlPY01DTXdrZ2ZwUjRoblJZQVFNWklGQ0lHa1dqRWlZa0lhdG9qWVZrSTJSQ0FheVd4QURBeE1ORDBCd0lnTkJKR3FkUEljRTcwZWJ0OTdnTGZkdGdzVlZhVGJBYUVWUmt5ZUhYRGVmSDNoVmVVLzN2SXFDSUEvaC9tZm9HdCtsUUFVWEh5T3hTV1h1YW1UVjkxdCtLamxIOVJ6bGh6ditzUEthR2JkY0ZyR0IzMzh5VDJmaDRlYytBRGNWTjR1aDl3TXZIajAzVUFIZm9pS3BSbWk0ckFxZ2xmTFNncDFXbmhIVDJjY2lkSlg4RlJVOEtJQmtQQlFWYW80SHgwTjFlaUtHamdvWjZwWnpMZ0RzSlduYzBOemJIZWpIdFhkZ0JHTW1yc3RPWjFualo2QU1Uc21XVGFHSmRrU2RFeEcwS0JqTXVUSTRhR1kxejc2Zmg1RHY4Q3k2c3VOZzIyOFl1RXE3QzIyeTdRT2NPWEM5VmhBbnpRV05KbVo2aTVIYnJ2b01nUFZRMVZCQTFnWWlnRkVvN0lXaytKN2tkOU9mb21SaU1QbzdFVExFWkhySnRDeEdicmRMb2RsaGEzYmRzdkJRM09VWGg3RWZVNHhzRWErTm51cGUrMU56d1d3djAyQi9lcEE4YXY2bm1CSGNmUkJxNS9KQzFhOTBaODhPc1laWDVrc3Q4WEJXK1RFa1ExNC9ZUCtBbU1ySm5CMS96b0lDUS9sdEorVG9RNWs2QW90VWNwQVBVcGZTRVhISWpnY2RONUxSWWVLS3FVNktFbFBKMTRVSU9GSk9IaXFxaGhqb1FJY0hCN2lUSGxBc3FMQTNickg4OGplNFhqUXNnZmc5Q1dueUxyZWVyOHNtMnFKSGhDSWh0dC9mU25DdktiWDB3R3lkWEE3OXBjNytlbjlsK0c2K2F2TTVYTlhZbGMxUTNZNk10VmRnWkZzSEFZQ3BZTUlVWU1KZ1M4TVRqS1NlWkQ0N0pTT0tCQXllZGNTWFoxdWxrTkpXbU1sTTRZSERzemc5cDM3NEswaFZJQk1uWXptSGJsODdocjlsejBQeDlhWjIzQU9zbDlsbHMydkFvQUdBajJIeUw3NWY0NTZaWGJtNUorVkt3V3k0TDJSM0pRSGJzRmpqanhmWG5IZUgyS3ZUdVBXYWpzOEt4WmFZYWdGNXQwOGhscGk2Q29wcEVDaERwVldxTlNqcEF2T2hYTndvbExGVUpvTHpnVXJlQ0U5REN5OEdNNlgwNWdwOWpIM2xKUHRNWGpBaXZ2Slk1Yy9Bc2VPSDRlT3lSY051bElYRXYyaXlSWFl1VEJqMU1EMVVVaWhDTWtBQVFnb0llMDZSdW1pbnlYSUYzKys3Q24yNE9yWkgrTmpleitCYjAzL2dOZjBiMFE1bW1GRmI3MTBUQWRlSzZoNldyRXBuQmQyZ1JnQ0ZGSmdER2dJYU1CakpINU1VTWNFT2pZUElqTmtVVWl2azNGWWxOaTJiWjhVQXdmSmpSTHdNcFhsNW9yWlc4MWxDeGRXbjl2MS9WK2xjL0xMQmVERk1IZ2xlYUpJZnN1TGp2bWdPMy9wb3oxS0ovUGVxaVd3Ynp2KytCN1B3dVBPZkN5dUwyNWhYd2RTc3BKNVArRFFGekx3UXd4OGlXRlZzSWc4WHhrSTVmQkR6ekw4QzZjZUVJZ242ZURGdVlvS0E0cVZROVYrenZmMzREQXU1Nk9XUDRoUFdmTUVuREoxcXNsaDYvdXZXTVhGRFRHSkZOcVZRSVVrNzVPUjhnTkZ3dEtuRk9ua3Q1SlJOQVZ3cGhrbVNOWGd0QWlBVExKNm1oWmN3ZThlK2g3K2RmdDcrSWtEbjVNRGRzQmVaNjJNZHBiQWFpVktKVUkrck9RaTBmME83a2p3eWlWYWtZSWdLQVAzWTBNRWhvZ2tqd0F3eGdBQzd0NTVTUHF6UTBpZWdWNjlXWmxiM05pZjBRL3ZleG8rZCtEak1hbmhsdzdDWHlZQURYZ3hJSmRrdlpjYy9URjMzcklIbys4S3E2WlQrRGtabVp2RnE4LzlVNTYxNlF4OGIzQ05pQm90ZENoejdNdUNHNkR2Q2hRc1VQZ1NBMWV3cEdORlowbzZWdDRoY1g2VmVqajRHTCtBVk40UllrUnNqZ1BGSHN6MWQrUHUzZVA1ekRWUDV1TldQQlpUbmNsQUpVTkZxUktJRFVtc1Nad1NJcVd6SkdXWURQMDBZUnFDWlhXU2Z2cGJHWElURU9XVUNXeVRSQ3NrOEhzcHdod3BuendtdVFMQTdzRnV2bS9IKy9IT25lL0h0ZFYyR1IxWkplUFpGQ3N0QlNCeUV3R0ltdUlPUXRaSURUVVFJVGxiVENJdUFRR3NDS2lFQVNnMnc0RzljNWlmNlFPWkZVTTZMTXN5dWJsUGZuVC8wL1hUKzk2RDMwV090Nkg2SldMa2x3YkFaUFBaN29zMmZjUTllTmtqc09ETERyTnNNSmlXVmM3amRRKytSTmF2WEtmWHpOOG9QaU9HMVFBRExURG4rMWpRQVFwZm92Q0ZETFZpNlN1VVdxR0VrNUlLcng1T1Bhc1F5NlVHNGxnb0FpTTVadDBNOXM5dnhUM3k0L2o4ZGMrUng2NTdESFBrQkNDVmxpSDhaUXdNak56UnJJdllBVnIvMWxFTUJIK0FBVktVRkFaQndsZjkra2l2SUZHRFFsa00yQmpSUzBFSk9sWEFDTE1nbGFYdiszcnAxa3ZsdGR2ZUlOY1UyekUyc1FHOWZBendWU0J6YWk1UldyR05JQWtadjhIQ05pa1FJRXlRbEhGL0NZek5aSDZtajVrRDg1RE1ra0FsUzYzRjFxR1ZkKzE5a2w2MjkvMzRwek56UE9kN3Z6UVEvaklBbUNRZjhwY2U5V0djcy9MUmJqZ3NySFM2cnIrZlIraVl2TzZSbDdBem5zdTJ1VjJveEhFQkF3ejlFQU5YeUFLSEhMZ2hocjZRb2E4d1pNbFNnOXAxOEZKcGs4SGl5QkJXVXc4eHVmUTV3Tzc1VzdoSlZzc2ZIL1k4ZmRhR1p3V0hFaDZPSGlJWkRVS2lxQ0xSYTlFN2FpQVJGMDFqeWhhZ1ZBcUMzNW1KRFMrNzgvd1JBQnlEOXgyNFFSRXJGcFNZcFFvVGtoV2kyRXB3ai9paEFsRDE4UERvbWd5QTRaenJtN2ZlK2xhOGJ1dmJzTXRQYytuazBRQWh6bGNVWThWSzNCQ1JHYTlqZndDdG1HQXVCSmpTbUJnK2liZE5DbXhtMFo4Yll1N0FBcENiUUZndXlXQnVMY1JmZHZCcGVQK3U5K0Yzejh6eHRsOE9DSCt4QUF3SkJRS0I3Nzc0cUEvNEM1WS9YdWZMeXNCa2JuQ0l4K1pMekQ4ODdDOHczUmxpNzNBZmxNb0YzOGVDSDhyQUR6blFvUXkxd3NBVkxId2xCU3NNdFVUcEsxYjBjUERpVlZNdWkxUktLRVhGWk5nejJDNVpPWS9ucjNrbVhuekVpekdWVFZBQjhWckJtaHdwVzRBUXBhcFlZeVF0VFZLdkhrcXF3b1BvM2NFcFNkZUNMekZ3ZlNuY2dBNFZoQlFqQnNia0hMWGptT3FNdDdSMGRFb0JEbjBCQ2NRSERHT0FycEdzVEc4d29LZ0dMZXVnNkpoTUFIRFhZQTllZGUxZjRKMTd0cUFjbWRMSmtkVkdmU0dnSnJVYm5DYVJoR3hZWXhBU2VJSjltRVhBTTZHZmhpVEVXb055VUdIKzBBQXdRbmdGbG1lS1BaWGdBL3VmZ0kvdjN2TExja3gra1FBVWZQa2NpL3RmNW5vdjJmaDI5NkNWeitiQnFzb2t6NHZpQURibEUzekRJMThsTTlMbnptS3ZLSlVESGJMdkI3TGdCekwwaFE2MU5BT3RVTGlDdytCY1NLRUZDblYwNnFDaTRtTHNWdUVKNUJoNmgxM3oxL05CRS9mQTN4LzlOM0xpMklrQWdGSXJac1ltU2pmYVlZQW9xU1lFeUFCQW9TUkRVTUtJVFl1bnM5VzhIQ2oyeVEyRDIvalZROS9FZExrTHUveXMzRnh0NDk1aXR5eFVDNmpFRWZBUUo5SXpZMXpaVzRsamVrZHd2YXcwbWUzeXpLbTc0K3pKTXpEVldjbFZ2ZVV0N1U1VHFZTVJDeVBCeDI2bVVTTWJISndNQmVEVTFjVlEzejV3T2Y3MG1sZnh5L1BmbHBFbFI2RWptVGd0QUZpeWxuMUJMMXRKZG9FZzFxMkVQSW82aThJRTJhK0NqczA0SEpSWW1CMEVNZXJoWmFrRmRoU1E5KzEralAvOGdVLzhNaWlhWHh3QXYzeE9odnRmNWpxL3YrR05mUFNxNS9rNWxwbmFyQnpNWXEyMThnK1BmaVVLVytEQVlCb0xNc1RBRDJYb0J0clh3Z3kwRE5FTlgyQklMNFZXS0gzRmtrNkdyRUtXc25wMVVPTkplblhTTVYzdUsvZWpXcGpCeGV0ZmdEOCs2b1hSeHF0Q3hOUllDbWhxUXgxSndzU3dHMVdzMk5wOEFvQmJGbTdobHc5OEM1ZnQrenkrVzEyUEc0WTNpNWVja0JMSWpNQ09RMHdYdWMyWVNWWkhjMDMwbUN0Zm9uSjl3QTBFdmdDUUkvT0tWV1lsNzljN2s2Y3VPVVVldXVZQ25yejBGQk1IQmtjSFFKQ0pFTUUwU0phanREMXRKYUhxbU5zT0FNZ2JybnNMWDNiVFg2TS9NbzRsSTh1bGRBTWlBWTRpSUdHTXJSbHNnclJpaEFnMFVrMXNoMFFIQ0N3eksrajNDeWxtQzhBWXdrR3gyb2o1MFh4cC92SFFBOTIxKzc3K2l5YXJmekVBak9KNS9QY1BmNmw3eUpxL0tjcXlraElacGVEeW1TRmUvZmlYQzBjTkRnNm1wVlNIR1QrTFVpdjBYWW1CTDFpZ2tNSlZLTFQ1cWJSQ3FZcUtqazY5S0JqK1ZzZXVIY0hPK1J0eFFyNGU3enpoTFRoajRtNnN0SkxnV05oQTNVRkNwQlpwZ2hIQ2NPclFOWjM2L3I5MzZBZDQvNzR0dUhwd0RiODFmNlVzK0FLYUs3dmRGZEt4SXpBQXJHUVFBM2l2UUxUdmdvL0x0aThLQUJSalE1cXBBbDRJaU1KVkZlYUhCNENxUkY0Wm5yWGtOSG5FeXZONTRackh5S2F4VFdrc3JPakVpbEVESTRwUTlJUklOd1p3S1ZVZEFKSE01UGpoekxWNHpuZWZ5OHNIMTJCczhoaFFDL0ZTcTFnRU13TkozdFZXU0hDaElyY1laYTlGU0lnUWdRem1DL3FoQnpKRGVLVk1HQ3RYRHc3cHg2ZnZoNi91K1ZGTTkvK0ZnUERuQjJEY0VaT1BQT0wrNWVabFh5aVhHTldGMHNLQzNZUDc1UzhmL1VjeXNYU0N1K2IzQVZaazN2V3g0QW9VdnNEQWxTeFpTc0VTUSs4UUloc2xDbC9CcTBkRkgrSzdxaUcxUUNHZU9YZk4vSmdYTGpzUDd6anByYklrbjJEaFMrbmF6cUpoYWYxZlEwS2hxcEtid0wzTnVsbCtmTStuelR1My93dStzdkJkc0pzQjNURk1kcGNpZ3dFVVVMcElvUUVoNEJCNkRJVXJ5aGttaDdjMjdPdUVMU2dadEJ3RkZCaVRBd0o0VmN3TVpvRmlsbU11bC91UDN3TlBPZXdwZk9SaEQ1VVJPd0lBckxTaU1SWkNHSWloUWJMaktBSlJWUldQY0Q4TGJzZy8vTUhMK0xiYjM0WHU4bzFpSUhCVW9RQ1oxTTVOQWgrQW1MZ1krYzZRS2tGa1lnQUZSUUxIWFM1VXBOTm93YXJIaWp3elg1Ly9udjdwRGZjRk1VeCs5ODhMbjU4UGdKRm83aHczY1p6OHdSRmY5OGVNTHNQQlNqbGl4Ty9aSVM4KzcxazRidVBSMkRxL1F3Qnc0SWN5MUNJQTBCVVlxbU9KRXFWV012UVZTMVpTSm02UEhvNkVpNVFMUktpaVp1KytHL2pINi84UC92ckVWeEFBS25YSVRVWlZHREVoWXRXT29YbDZFYkZxQUV5WGgrU2RXOStEZCs1OUw2OHBieEdNVHNwNGJ4bXNDQUFYdkdsRk1Oa2wvcXNoU3lYNHJ0cXVQYXE1d3FUNFFwbG00SUZUdjRRa2p1cTFWMEJvUmF5RmR4NEx4VjVnTU9UcG5lUGttZXNlenlkdmZDcVdkWmNpM0Z1RkxEZ2hMVHN4MVFLSU9EcG1SZ0JrK0plYjNxUFArZUVMeFV5dUY1aU9lRGhreHRRT1NQSkdRclNHc2F3RmdFcE1aQkRHTkJzUkFrNkZmbGlselFRQm5GbVNaZmpZd1MzK2piZGVoRXMzVzF5MDVlY3VldnA1QUNqZ3hTSnlpWFplZWZ5MzNUMG56OEsrd3RtUmppbDMzeTRYbmY1Z1BQaU0rL0tXZzl1a3lvaStEcVQwSmZxdXdDQ0J6VmNzNkZDd2xOSlhET0UxRDBjbm5pRmxxbEpQUVdnSnNQL2dqL0gyNDE2SFp4NytkSGhVaEJvUll3a29ndG9LbHdIRVE2bXF5RTJHZ1MvdzFxM3Y1TnR1Znh1dTh6dEV4bGRqUEI4SHhJbnpnVjJ3Qm9rVlpHU2RKZEVrSkNqR0NEVE9kOXVEclpFV0x0Yi9BWkorWStqWGdlU3Fla1dvS1pGZ3F5cEYrdFVNMFQ4b3gva04vdmNQZjZiOHpqRy9MYVA1YUlqd0dKRU1KdHF4bWhoQUVhZzZFRjZkNlpxZWZtTGJaN0Q1bTA4VnQzeWwyR3djSXI1bUdqWEJLNW9ueXVpcEpLYVF3UVpGVXMrdzlKV0hsbFVBc0FNNUlwVjAwTUg3ZHY4OTM3M3J4YitJYU1sL0Y0QzF4enY2MHFQL29iei8xUDl4Qnd1WFpWM3JwdmZoSHV1T2tkOTl3RVc4K2RBMk9BTHpIS0JrSllXdk1QUUZDcDlzUFZkTHZkSlhBWEFCZ0ZTb09CK3lraFdDbVlQWDRqMG52VlVlditIeEtIM0IzSFNFZFNVdTZrVUpaVVFxdVdRS1FMNjQ3eXY0NCt0ZkxqOFlYbzE4YWkzSHNrbFFuU2djRURKTkVPbTU5RTlOeml5S2pDeWVxWVNuUmNSMWpINGdXcHhSR0JwcHk0aFlIZ3lKOXFPU2dESlVodGd1K29NRlZETzdjVkxuQ1B6Zmt5N0Jvdzk3aEFCZ3BSNlpzWFhBcHNtYlVoQUdsU3ZReTdyOHdxNnZ5ZVp2UEJYVDR4bEdlaXZFK1NFWnZYdUdHV053YXFJSlFUUVNIUkpEMlJJNGRpV3BDamlHL2pqT0EwdkVteGxtOHU4ekYva1BidHZ5OHpvbC96MEFSbmU4OCtUREw4VGpWbTBwVGVWTUgxYmRVTlpKemo5NjVHL0w3c0ZCRHFzU0JVcnBhNEhTdThiSjhDR3BJQ1dTT3ZXczZNV3B3aUhFZUlNUkRuaTEyTGZ2R3YzQWFXK1h4NisvVUVvdGtaa3NoVlNoSk1VWVFsV01NU2pVb1dzeUhDeG4rT2ZYdkVMZXR1Kzk0cGFzNEpMdVVnakxVRmdFVUV4TUowbjYwUUMxTFFmV0JucXEvRWl4dXZoL0NXRzNXRHdFeHM1dktWOGFVWENHV28vd0hjbGNUT2tzY2VHWnFEd2pwRUpvSVpMcmZERWpuTnVMeDR3L0VQOXd4bXV4WWVKd3REejhKSjdETjJ1SThaYStRTWQyZWNXQnErVlJYMzRzZG8xVzZJNnVvWGRGeUtVSTRwM3hucU5BaGtEamVMVGVVakcxZ1lSQ29CVDRPRm1scDZ6dTBsNVRUTHUvM25vdjdKaTc4ZWNwL1RULzlVdnU0ajFmZ1k2Y3Zud2Q3am4rOXpwaW5Ka1hBVlR5aFFVODQzNlBrK2xpbmpQRGVWbndBOHlYQS9hckFuMDNSTjhOTWZCRDlIMkp3bFVvbkdlcGpxVldVdm1Lemp0V3F2Q3FVQVhWR083YmV4WGVkZkkveU9QWFg4aFNoK3lZVGxqcVFGVFVib0Vhd0Vmd2ZYSFBGM0d2YjV5Tk4wOS9RTHJMajhaWU5nbm5obEFOYVN1RWllbjNxS2xpeGxnSUdheDJEY3NRL3FZSmNkc0lVVWJ3SWI1ZW16aEtFcHdNN0lZMmJ3QkFUZllnd1VDaXg3OGw1dEFZS2tubkNoblB4bVY4K1RIeTc4WFhjY1pYenBIMzNQd0I1Q2FITlVEd2hPTm1pSGtLQ2tWbU01UmE0aDdMVCtGbnp2c0VqK3VQbzVqWkk1bDBnVkRTbDdqcUZGR1UycUpRU1ZScHNJTkpRTVVnMXVmRk9nUWdzOEw5cGVwSjNXWDJXV3ZlQmdGdzBuL2ZsUHZaQVhncEJBTDFEMTN5ZDNyTStBYlpwMnBzSm5yb0FCNSsrcmxZdG1RSzIyZjN5WUFWRnFxaDlLc0NBMWRHeXFYRXdGVVllbDlMd01LN2FQZFJIRlFxOVNBVnh1YXlkL2ZWOHBZVFg2TlAyL2cwbEZvd014azhZckNURk1BQXhsRFZTd1lEYXpLODVzYlh5QU4vc0JrMzk5UXNtVHdLNmlzb25LZ3hFbnpETU9OaGpnbjZXTzFHQ2ZaU2FQRVN3UlpmaDVBUEUrTmRVa3ZOR0FOVGxTakpBc2hVUmFDQ1ZHT0MrRklnZ0JzME5mMVJDeU1TU2tpazZNVFJvM0pEakkwZWh1a2w0M2phVmMvQjh5NS9rUXk5SWpPWmVIVmg4WXhBR1J4a0EyRm1NaW0wTXFjdU94NWZmUEIvWUdPUllUallMMFpzR3BBd09oNXBjNlQ3aVNwRkdpNFJqUjFpSkZHSEFDWFRlVmZvZmNmT3laNTZ4RXR4RVR3Mnh6NkdQK1Axc3lFMzZ2dVJKeDIrdVhqTXNrdlZvc29MbTFXRFEzTGsrRkk4OC96SFlmdjBIczZ4a0tFdldBVnFSUXBVcUx4RHhRcVZLandVRlIwcURhU3lpM2w4Q3FCU0Q3RTlIanh3TlY1LzVDdnhndU5mZ01JWDZOaU9zQlV4aUpRRUt6amt5TGpnQi9MY3E1NlBkeDNZZ3JHcG8wUlVnTXdCRVRqQmJ3aldWMURmck9jNDFxa2hha3FRaEFtNUxVaUJmVW54VXpRQjVQaHA2YmU0aU1uZmpVdWNnck9NSUE5aXE0NENJM0ROUWVkUmFvb3VMWTU2d2tKZ2JJNzVRN2Z3UHAzVCtMNnovdFVjdnVSd2xMNWlaaTFVR1lqM2FJWW9nTXFYNk5xT1hMbi9HajNuUHg2TWhhVVRJdEtEaDJjMFBjSVVwcHVLYWlBYUVjSG1TOCsxdkdYNEZHR2tZc3FvM0ZvTStZKzd6c2FWMHo4TXUvUm5VOFUvaXdRVVhIb3hKNCtZbkhKbmpMd0dvNWJTVitOUkluZkVnKzkrTG5jdDdFZmZGeGk0SVFlK2xMNHZaUkJyT0lhKzVNQzdsRVNLMG50VXpvbFREK2M5dkpMT080cnQ0T0NCNi9EU3RYK0FGeHovQWxSYUliYzVCYUlHRU5ONGdheTBRbzRNQjR1RGVOQTNMK0M3RG41WXBxYU9rMkFXVldBVDh3OGFSK3NPVmx6VVhwZFJ6dFUyWVJJSkVXcFJWU2FwcVFqNXg0djJiNTBURUlRa1F2V0dKS3NyNUFSRU53bkpMdFJhQWlsTVkzL0ZiMDJaK0o1a3BRWEdsbTZTcjd2cmNjOHZQSUJmMlAxMWRHd3VsZmVFc1NCVUdGV3hrTksxdVpTKzVHa3JUcEovdXRkYjRRL3VKa3dZR1d1eDNwUk9KV25IZWlmR2dTVlBPVkhhaWZjME1EaW80TEc5Y1huRTByZUNFRnlhREpwZkJnQTN3MEF1MGZrTEpsN3FUcGs0UXFlZG10d2FuWm5GT2NmZkhkMlJqaHlZbjJPcFhnYSt3c0JYd2R2MURrTmZvVkFubFhxVUd0UnY1VDBjbEM1VXFOR3JoN1VkT1RpOWxlZVAzUmQvZGRwZndxR2lEYkVCVVZWQmJERUpCRHNvTnptMjkzZmdnZDk0c0h5enVsR1dUQjZId2c4aXdBeVFtclpRb3VaQjZDUVVheUJaSzlob3gybWE5RkI0a2RRa2tGUnAyUDFNWG9XR3l2VmdFOGJQQ1FzbUZCcENLSm9rWFdnbkV4T2xBejdUNzZoWjdqU3VNT0FXQ29VQ1Z3MHhNcnBlZGkvSjhMQ3ZQd3FmdVAyVDdOb2MzcFZSL2hwQ1U0bW5vR016Rks3QUU0NThtUHoxeVg5T3QvTjZaQ1lMTm9RYWllQ1NSY3dwcFVWYXB3M0lhQ05HTUZMQzMwWXlUS3VUZTB6ZHl6eHA3Zk54RVR3dS9kbk11cC91eFJmRDRGSm81OWp4NDNHUHBjK0Q4OTVReEErR3NtSEpDcm5icHFPeC9kQitlS2dNcW9KRFY2R29TZ3g5eWFKeUxMMmlWQi9TNkoyajh4clQ1aFdWRDgwd0lFYm15eGtlNVNmNXJydi9Jd0ZsYUVBQXdpaU1NV0JJc1lTams4eGt1R1hoZHB6M3pZZnkrOWlCMGQ2UktOd2dWdWZFalY1djRtQlhKejBaVXF3Q3ZaUHN0RG96RDRCS3MvR1RvSWhzU2EwKzYvV3FrL0JSVy9IcDllRzU4T0dxSVZzVUd2OEplQU5ia2xlVHBLeURLUUdjbE9RVUdYcGZzSmVOQ2xjY2hzZCs0MG40eE5aUG9wdUZGSDZCQWlZa3hTS0txVHpMVVdncGYzTHFDK1c1bTU3RGN0K3RrcUVuRWljWHllRmdCRit5RXhyd3RiZGg4MjlrSVRDazFZNm91Y2ZTVitEMDVldndCUGlmR2xjLzlRdFBna0JBZmNMYVYrdjYzamdPZVRMM0l0VkF6enp4Qk80Wkh1UkFobGpnRUFzWXl0QVhVdmlLUWZLVmNON0ZKa0NoSlFhcG9jWldWVlFobFJlV0NnNzM3WkQzbnZsT1dUZTZTdFFyck9tWTJMS25uZ1NuRHBsa3ZHMStCeS80MWlONFkrZVFMQms3VElnNUdCTTBiTkNxc1pkZlhGbEU1NEVoRFptMTZwTUlPTFp2dUxITWczb0VZT3EwZk5SYUxBUU9tblZCUzRrcUlnUkZtbHhEZ3FLdEJZN1BhUHF1SUdhb01WOEFDY3pSZ1NWQnE4WkpTV3Nzc09vSVBPNGJUNVJQYnYwME82WnJLbCtaZ0NOU1E2eGFBRU1MU3cvRnE4LytXNXc2ZG9KV3M3dG9JbEpyZXlEWkYvWE5wSEhHTzZYR0hkanFpK2cwdkhlMlVuZEtaNFc5OS9qZlFBRnMvdW5WOEg4TndNMndlRHo4MkNNUHV6K1BHWHNFcHlzSEdNdUZnUnkrZWgzR3B5WjRjREFucFhjWVZoV0dWWW5DbFJ4cUphVjNVcXFYa3A0dUZoSDVPc1NXVks5alpuTE1IN2dGTHo3ODkzaXZWV2VGRUpUTllqUk1xSUdXQVZXUm1Zd3p4UXdmODQxSHkwMW1yMHoyMXNEcGdCQVRUT2dJTkVuMk5SdExPcEt3RUlhdVZEWnllTkpRZDhDaUg5Um1ZdlFWb3FSc0xFZ0ZJdjBYdEdXVWVHRW9URTgwQWJRR3E2eXRQWktrU2ozMGlJcmEzV2EwRHFKc2p4TFV3WWdWcnRqSUN5OTdvdnpIOWkrd2E3dnFmQW5iMFBJRWdNeFlVR2xHVFVjK2NNNDdaSExvUUZhUXRoTWlVZHRRd3JjbE1YNG5MREV4LzNIM1VPQnBVTkhqekxHblptZXV2QTgrOU5ON3hmOFZBQVVuaGk4cWYydkp5N2t5Z3hsR1phREFjWWNmTG5PREJWU3FnV0lKMVdzbzFLUDA0YWZ5bnFGQmtJY3FVU21EMDRFWXI0ZVZoWVY5Y3B5c3g1K2Y5dWZpMWNPYUhLclJqQS9CQkFPanBBazAvWk8vOFNTNTBtN2wyT2g2bEg0WVJSNEZNWDh1dWh6cER1Smp3bzd0aUtmSWdnNWszZzlRMGlNem5iREJHWTRSU2VZZElWRUZoclRvS0tUcXlFZTlQdUVWU041T011d1VrTFFMMGg1b3FlWm1mTkZFaU81blNsZ09PVEFwM0pjWWtrWmxDaUZRNzJrbGgxKzVRUzc4eXVQeG5iMVhvR3M3TEVONE1XWkRxQUFRWXl4TFgvTDRaY2ViVjV6K0N1citiYlJaSnhHYXlmRUlmRkxnL21KTmFJcEgxczUvMkYxR0JNWUkxQUF3Qm5OS3Y3RUwvNER4UzBDRW84cCtiZ0J1RHRYeVk0ODc3RnpkMUR0UFo3d2lzeG43UTY1ZHNRS2pveU9ZSHd5RHVxMnFRQzZyUStHY1ZNNnpDalplY0RnWUNHYXZHcnZMSzN3OGFjTk43K1diN3Y0NlRHUWo4TW5OTUtITU5Tb3FldS9GSXNPekwvOGRmbXJ3ZFJrWjNTVE9EWUp3Q3FmSDFQNS9jdWVVQW5wQ2tJa0NNbnZ3Vm1iemgzZ3NsdUFZTjg1VmZXQnUvODF3Y0JUYStFMHRjQkZRQmthbnRyMmprNUpRbHJ4bVJnYVJBbEUwU2ZhS2xxM0h4UFlJcUNFNVZCUEVLQTBrbWJEZUFubllFTFZGRm9JV0Jxb08xblk0djJ5VlBQVHpGK0ttdVZ1UldTc2hUOXF3L3NCSVZGZGE4c1duL29IY2MrcGV4czNzaGpWNWl4RlBOd29ra2Job011b3Iyb21wOHRtRDhHSXdxNDUzSHowUEQxLzNDRndDL1dtazRIOEZRQUJBY2ZMSUgrbTRCUW9mYVVvdkc5YXV4T3h3RG4wV012QmxTSi8zSllZK3RNbHdkT0s5RCtueklhRTBSRGhJVkY3aG5ZZlFZdUhRclhqNllVL0ErZXZPUmFVVk96WmtmNWhhTVFLVjk4aHR6cmZlK0sveWp2My9ac2FYSFF0MWM0Q0IrR2pVYWV4c1pZSTRBVDNCeWdPK3k2SS9vMmIzVHI1d3hUUDVqVE0vcWRmZjgzSzk0ZXdyOFAxN2ZBVi91ZnJQS0R0M295cUhFTTBKRDZpM1RDRVJrZ2o2UDNreVNKSXlERERaY2JYSEVvQkgxZHBraXFJSU1YdExrbmRVMi9nYVhaMllLWnBjNjRUMUZDOU9LR1ZzWjAxRFlVWnhwa1NuTjRvRG8xNmU5ZVhmUlJwVnFQb0RDQ0UwRkdGSnNLZmxUZmY2YTNRTzlRa0hTT2pqaE1iVGlvQnJ1bjVFM0xINTBUZ3hFdTFBZ2FEdmdHVTU1RzRqTHdRZ0NULy9QUUJ1RHUyN3h1NjE5R1E1Y3VSQm5IY1VNUm1IQlZZc1dZTFI4VkhNRFFhQlZ2RU9aZVZRdUtCMm5mTjBEUGw4SGlwT0dicVF0aG9Da1VEcEN5d3RsK0NTMDE0ZWJvK0pHV0R0VVZicTBMVTVyOXgvRlY3eXd6L215TlJSY0s0SXVjNWhwMGFUUytyNHJVWTMxa3FYeFhBZk5nNXpmUEhzeitKMXA3OFdkNXM4UlNvVktBeFc5VmJ5WmFmOEVUNSt4cVd3Qi9lQldnSnFhd2NsQUNZQkxMWmphNkVOYk95NGNNVlhNeHBWdGU1TlM5bVNhSktBR1ZBY0hOZEltZ2ZJTUgxdkVraDFDQ2JZYmxGa0JuWHBmWUh1MkJwOGJlNEtlZEhYWDRaY01uaDZwaVVXRStiVG1Fd3FYK0hNTldmZ2FjYyttWDcvVnRpOFV3OGFrWUZDUXdNa0JDNVdxUUpBYVdCRHJ4dDRBakFXODBvY00zci83TzRyNzRYSC85ZTI0RThHNEIrY0V3Qnd6dFR6ZEUwM1Iwa1BFQ2djTnF4ZWlmbGlpSUlxdzhweDZDc01RMklCSy9Xb1NLbEMyek40WlZJMW9uSGplRlZZMjhId3dGYit3Y1puOElqeDlTaDlpY3htMFFpSk5BU1VCZ0t2SHMvOCtoOWdycGREVEJaaXVqV0RHaVpKWTV2bGNLaUNRbXpPNGZDZ0hOWHZ5VmZ1KzNuZWEvbVpMSHdsbFRveHhnQ3FJQ2lGTDNEQitnZHd5NW52UVRXOVhVU01TRHlCcTlhdXRYcXM3Y0phT3k2aWF3S1VvRFJSSlNOVmtTY2FONGI5MG5MR0RSTWoyMkNJSmNlOWxNaktoUFg0WFpIS2psL0cwS0lMaExDcVNuU1dINE0zWFA5UDh0RmJQc3ZjNUNpOWk4SGNWaFJKRER3VUw3L25TODNTY2d4YUZNSERaOVRyQ1hOTnVXZmtyRm9xdWs3R2phL1N5Q2t1cU1lUnVaaXpKMThBQXRqOG40dkJud1JBZy9NdWN5UEhqQnpHSThZdlFnR0toMlhoTURZeWlwR0pJUDBxSDZSZjRSeEs1MUF5eEhJckRaRU5LRVExYU1QVTNWWlZJUlFNaTNtdXo5YmcrU2YvSGhRS2E2MElhR28yQUNGNzJCckxsLzN3RWx4WlhDMjk4YlZTK1NxWXlkcWlyQWdKamtJUU9JYUE5MzBzSHlvK2ZiOVBjTVBZQmx2NFFybzJaMjRNTENER2hJTDBydTJpMHNvOFl1M0QrTXpWVDhIdzBGWmF5UVcrOW9rbFphM1VlcFloZ1FSQTdGb1o3UGRvSDlZcEJtbHdkVlFyZ2c0dHFkbW1hTUkvSWtsdHh3YXQ5YUl3YVc0bXk3Z1ZXNGtKUE9wS21KVXI4Znh2dlZCMkxPeEJaZzI5aHFTMmNDa3lZK0c4a3lQRzErR1B6bmdCZE4vdFJrdzNnS2lWV1FGdEFXN1JSVGFQUnkyZE9sQXJNczU3NzA3cFBUSS9hZVdwOFlTRG55am83dnFKaTg4eElHRHV0K1lKV050Wnl2bkt3MEJRT2t3dUhjT2NMekIwSmZ1dVpPRktsRDcwYXFsY3lHWU9ObzZHd0RvcFpEeERJMHBEb29OcWVoZGVmT3h6c0hwa0pid3FER3gwdXlBVVVlODljNVBKOS9aZEthKy80VTNJMXg0RlJabE0rS2lxb2dtZm1nVW93NG1Wa3NQdHVRMS9lK0pmOHJqSlkxbTRrcm50Qmltc0puVEhRcjJ5WWswR0Q4WHJUL3BiUFVvTzAySXdFOXIrZVVSRHIyVUh0ZE5rdENVUzR2MUZ4Z1phSnlNSTRHTXdMcmdTZ2FPRXRxS21qQkhCOEsvVVFsVmFueDFVU05ySXRWUGFvcTloYUp6eE1OMEpiRGY3OGRKdlhod2FlR2dveHdyL0RWUmhMaGs5bE0rOSsrOXdmVzg5ZFRnVE1oZlQzTlM3b09rTEM2VEhFSkp6VlFtWDdCSWxTaDhRTlVQUEl6bzlQWC9zcVRXZWZpWUF2dklySG9EeFIvV2VvZ1loRzlhSGRQU3hKU01ZREF1NFNMT1U2cVR5UHVYMVFUVzRCQjdSMjBYcXV4eHRLQmk0YWg1cnVVS2V0dWxwVkNpc01aQ21pb0pVaU5pZ2dQNzRCeTlET1RZQ1VST29tY0NhQmNJNUdJdWlXbnVsek5CbE1idWRqMXI1V1B6MnhxZElwWlhKVFFZVGIxYUVRUVczUUdYQ3A4cDRaeHd2Ty9iRlJ2dTdJS1lERUl5ZkhhYzQySGJKZDQzTFVRc014RzU5cksyM0FOU1VOY2FXS20rQ3JsRXlhdE5sUm1PK2F1QVN0WmErQkNJbnJ6R2h0VFlDRWNONUJBVGVsY2lXYnNEN2Jua2Z2M1RMWmNpelhMeXZVcWFBK0tnQm5QT1l6Q2ZrSldlK2dKemRRMG45YXRxWXF3dmQwMk5OaG4rTHFnbDlQN3lrY0pGbFFlZ1J2Y2NCR01XckxuT0xKdncvQmVCbVdJaHcyVU5YMzhPdjdaNmlDK29BSXlnOWVxTmQyRHpEc0hRb1NLbThsMUtWRlFPLzU1UGRSOEJyZ0lwcUNncUVoVEdTUWFkMzgxbEhQSUhMZWt2aHZFUFRqaEV4VE9aZ1llVmpXei9ETDg5ZWpzN1lXdkZWQ1JHRGtONmVRQ0dwYTFuaytvdzRMV1Z5bU10cnp2Z3JobkNsaFFraEVxWjFRck9jNmFJUlN3OG5UOTM0Skp6WU93RmxlVWhFYk9BK2ZPT0FSRThwOFlWTmFDUUFLdVgxUVdycEdNMGpUUlZDalNQU3FPSXdOOXArUGxHYmtNU3ZCNW1uamZrUnBTd0R4dE5PaUhuaW5wQ2xLL2lTYjc5QzFJTkdET29JZVFBaXJCRjRLSjU0N0lXeXhod0dYL1hqdDlUQm05cWFRUHI4Wk9oS2loazM5Q0FFNGVBd2hjaThyN2ltc3hFWEhQWmdFTUE1ZCsyTTNCVUFBUUJ6SjA0ODBpMjNGa05WQWdicU1UclJpekZkaFZQUHdPdUZUcVBleDRKcUlvVGJsTkRRRHBkUXFiMzd5cGZvdVE2ZWV2eVRJeHphUXhBUUNpTUNyNHFMdi8rWEJxTkxnTGdIeUJSZUNyZkxHZ2hCL1ZteDhOTTc4SHNibjRsanhqZUtWMGNidURCREUrekxlcnE0YUVPS0lZUUs2WmdjTHp6eUJkQzVhU0swMVlndUI1Q2NrY2FKQ0FQUWV1MGJSeVJvUjlTYkJXaldEMms5YTRja0ppQnJSRnNFb2lTdXA3RVhVMVl6eVBhbjFkWkVUQThEbks5b1IxYkk5Mlovb08rL2JndXNzWERxNm04TEJ5MFplTyt3WW5RcGZ2ZW9Kd2dPN29ZMUhkVGh0dWFHVU51dUVpMkpkUC8xN0FoZ0JmQlIySlFBbGhySWFTT1BCd0I4NWE3VHRPNE13TWZEbjNqaWlSMi9McjhRNFJ5RGpGb1o2UmhrSXprV3FpRzhldEFScXFScVNCb3dRT0MrSW1jV2lKUVlFWWhJc1NhSFg5aURCNjArRzhkTkhjUFNWOHl0aWJvcTNKQ3FpalVXSDc3cDQ3eHE0V3AyeHBlSjB4SVNWVTlVdnZVQ2hVVU1oMkU1clRCUkx1SHZibm8yQWRDS0ZVWElyaGJVVUtlcTFzcXdnU0JnakNGSmVlcVJGOG1SZGdQSzRad0tiTEFNYXJzN3B1cW5KWktnZGlTcUxVa09RcjFHOVM2cG5TdTJIcXNmWitzRjhYekU5QkNUN1p4ZTFpb25DbUNQdm5uZ0FzSm5HQkVWSjFpK2luLzlnOWVqOUU1eWF4bWxiTXkrTkxDaDF3MmVmdHBUT0Y2TTBic3FkRnB0YkF2VVRuUjd6OWJidnlhV0dLMFRnakNBWmhnb3VOWThFTWV2Vzk2MDQvblBBTGdaRmdSdVBtN1BmYkFxUDBibTFNTUtVSkUyczRBRnFvcDBKRHpqSVMrKzdqUWY3VDhnR0dtaFNTVHFtUXdhaFVXSnB4NzFwTENYSXVYYzNKZ0VPa0FoZjMvOW13UlRTd0huSVVicVRWa1hVMGMvSU9wQWlzbmcrOU84NThTWk9HckpSam9OL2Y1aVF5QmdrYmd3dGZac1BRNERBMCtQbnVuZy8yejhYV0IyajFqbU1RZ2N4eEFKSUNKRS9LRjFUVk5TMGtsVjEyeEorSktrSDV0dlRCZ1VJQ1I4Q2lTNG5zbXdXbXgyMVdWL05RUVpka0FpRlpPdWptU1dPc2VzdDFTdVdiZ1dsOTd3WVFMR2VPOEV4akEyYUJJeEZrNjlITFgwU0RuL3NITUVNL3ZGbU5SVml3bFlVYUpFc1JjNHdKaWtHcW1ZY0VoQUNEKzZ1RUFGSFZaMXA4eFI4bEFRRWsrei84OEFHUFh2MFZPUHdtZ21vV2NZQVNXN25SemVFOTZIOHpSY01JMmdDQnRXVytVUFNTQUlFVUJJaUJFalZUWEV1bXd0ejExL2Z4Sk1EYmpqVmdzSEIxcGorYzA5bC9PN0MxZlM5cGJTcXlQaW9ScTFhbXZXdWFGaGpBSDZmWG5zK29janJGWHdVUVNJdlIyUlRBUXhiY3hINlp1QVlrd0lYejE5NDVObE5WWklXY3hpVVlkQVRhcXpjWS9yRkg1dFpFSXlrTVNuQkM0SnVHRllyMFp1Qk15azNGbXdGUWl1ODdEUVdBQmNST29GUlpNeXhLSkFiYVNYQVNzdm1GeUtOMTd6endDUStyelcrbEJDV0kwQThOU1Rud0RPTGdDdzZmMVN4NmZqMk9LYkNaaW1xMG5MUklVQXFJSTJRQUdSa1V6TXlSTVBCRUM4OHM1cWVMRUI5dmdQK1lzQnd4WFplVFFHOUxUd0lRbFp1aFl1bkxFaFhnbXZkUnV5aG95cldRa2lGclZKckFRa2Flajcwemg3K1JsWTBadUM5NDZac2MzRTFZRkY0TjIzL2h0OTF3WnZTSkpHRjlOUUQ0M2dCeUR3Z0tvVGNjQVpxMDRGQURHMTVJc3JYZ3NhYllSU2VxWTFCd1lHVGgyV2RpZjVsQTFQVk00ZkNGME40bjJ4bGdaaDRxazF4eG1iOXNaMUN0UUpOWW5xMmx4cUJzOVVpNkpzb2JneDcrNlVpbGZiQUlpcFZFM0paeXVNMUREam9DZ3JNYjFKK2U2K0svSEY3VjhUYXl5OVZwRUhDR2kxeG9wWHhYMFB2dzhPRzFsUFg4ekJ0RG5Bc0FqaEFZMUVVZnVHSWpFS0lwS3dCS2poQkFrU1hHM094ZEZITDRHNVF6TGxJZ0J1aGdHSjE1eXg0blJkblIrRHdxdkFTanJiZ3BtQmM0VDNEQjNkSS9HclBqR1hhYjFyMnlGTlpEajB6eGlpUDhQTmh6OGlnTGIyeFFLajdsV1JTY2JwL3JSODZ2YlBRMGFXaXZlK3VjOFlha2hHZTdSQjZ1NWtaVFhFNGIwamNlVGtwckJjeGxCaU1LSU5jVFNqVEQ1bytzUzB6aUt4ei9MdkgvazdrcnNPcTZxSTh5OHh1U1FtRmRkdVlMaU54V0ZVWVVvNkZTWkpWejlYUHg1SXc2VGNHRHhhalJZbmExQW0wU3ExY285OFZTM3lLUFhFb3pad0NGcW9DS2dqQm0vNHpsdlQxM2pFN0c0VGU4SjRlcXdjWGNyN3I3NG5aSFlXWW13RUVsaVhDWHF3TlpiRXJhSDJ3dEtzQ29BeWJIeVczbkdWM1dEWHpOOGJCSEEvTkwySjd3QkFBSUMvZTM0NmwzUzZVbEFoQ25xcXpZS3lESjV1VUxlaFd4TWxNZ0FNSWJZay9RVWhyUnpKTUlEM3BabVVwVGh0K2VuQndoR0RsR2dxOVowQVg5cjVGZXdvZG9pMW8vRHdNY1lMa05GeG9OVGhweHJITUlCM1hKVXQ0Y3FSS1hwVkptb0hpNEdXdW1NMW9uVHhSUUZoeFZCVmNlVGtCdHh2OGg2QzRoQXQwb2xFclUyV2hHdVFTSTJkbmlSVlFtWDg1QnFGZHdBazZsaWZRZnlTSm5ZWHkzT3hpQkJuQ24yeC92NmdybG4vTHd5TlVCSDFUbVRKU254NXoxZHg2L1EyNmRnT2ZISjlFaENpMEhqTThZOFF6cytKbUF4SW5wZEs4OVdMNWkyMmlHZ3JtNWpKaFNxc1Bad1NTekx3dExFeldqZHdGd0NNQ05TSjdzUFFNV0FSYzNJOVlIT2pQcG9LM3FOeE5JSUthbml4c0d1anlvMlNRa2tqRmpxWXhnbmp4K09ZcVUzdzZtbU1MQjVNWExEM2IvMndjbVFpR3JsR3RCWWhLZDdSeUs0b0RXdHVKSFl2RllhQzFnWUtpMjdjMUQwRHcydGJ0bjE4U0FGVThEQXdlT2lhQndIREJVRHNZc2N3ZmlqcnJZUDZ2NDBkR0FhY1V1NVFVOGQzVU1tMTlraWtjL3drTlEzSTY4ZUNhWmxHd1BvRGtyYUlhRXgxQjFGQ0dOUGxuQ3pvNTdkK0VRREVlNjFuZ2RFWkFXRE9QdkkrWERtNlh0MXdMaUVycXZ3SXFFUnZwTnFSd0p3MW9qZHRNYzhRMXF2RTBCcklxTDBBQVBFVkx1cWlrTVlnMkx4RlY2L0dHS2Z5NDFFQUlFMFUveUtaTVNtVEpaVlZOWVk0dzU1Tk1WRWcxUmJXTWdJd1FML2crYXZ2bTBZWnZZcncvUjRHbVRYWVBUaUViKzYvU21Sa0hQUXhEcVlNMFlpUXB4RStQMFMwUW4xRmJYU0xsQnE2SzJTbXFRNWtJenBFa1ZqeFdLMFVuMmo1Sk5IbnJCMGtQT3J3aDZIcnVsSzVRVnd0SnJZOHBrV0ZZWVl5K1NUUTRqY3piUmVpemgyc0pSYWF6Mk5ialNHOVZwSXRIVU44S1htMDdYQklqV0pLQ09rZzFDUXZGczhrdkRjeU9tSStmTk1ubU5ZQVNJb1lzTWJRZVlmVm84dmx6R1duR3ZUblFzTWlUOVlEcjlVYkY5c1ZOZGdaY2d0dGxPZ1ZnL2dvUFBRd2UrVDQwVWV2Uk9nWVYrK29NTXViUXpScXVHcjBlRE9WSDhPRmtGc2pQcnB1VmdDTjZwZUpjR1pZek9nSzF3NUlxdENKQ0NQaklRSmxuMmV0T0RYTVhlMVloWmxUSCtwM3I5OTNEWGVXMjhUbW8xRHhpYlVQL3ErUFdGTEVWS1ZvcmNmYUQ1R2N1NFlIc1hkd1NBQURGMUlBRXRCSWhwTXBZNThnYlU5Q3VsSklWUkRQY3dPNGVuUUQxNDF0RXJnaEJDMHBtQUFRMTVmMVF0UmlOY1hmSWtjSUJHWSs3U0lrUkNicGt1eXJWSkxYK09mSlVZbDMwaVFNMUlJbkpSd3lKa1NpVmhNeFlLNnNnTzRrdnIzbmg3eHRlZ2M3dGdPdE4ya2NSbWlmeWd1T2ZnQlFERUp0ZE8wSkJlMVdHNjcxeG8vb2I2UnY0cVVJRnplTTh4V1daaHNHazdPbmdSRnZpd0NZQm5EaTFGcU9kU3c5UENnQzFkQisyQWpVb3laRVF5cFRjZ01DR0VKWm96WmpqVGwvQXNDNUNpT3lRdFpNYkFqNFN4NjdTRWk4ajFQdzNiMVhVMHczdVBRcU5SeVFvby94RHNQMEN0Ti9WSUVzNjJGWHNZMjNUOThTWmtIdmdEQlo5TnVkd0Fja056Yjg3a1BVd0d5ZnZVbDI5MjlWc1NNaDBhRlduVzJoR2U0cUd2OWhSdUtZSlNFay90S0UxK0o3R2Q4Y0VoSFNxMUNIdU5KWEpadTNMVUpyQ1ptMHZVanRLSVNQWmZKYVNWQ3lIbWJjUWZQRGZUOWlBQnlUeFlaSVJBb0EyYlRpS0VobDB4U213WVoxU1VKWW83SkF6UVdHc2RXV2xRaDhNS0RnQmVnSnpNbTlOWGVjOHlnQmcvMG5LM3Juc3BjRllnOElFWTdRSG9SZUk2QWk2eEo3RGRmUUNFK2FoTVpvZU1VRFg2cUNSK1NyY2NxeTQrQVJFOWJqT0xXeGlIRFo3VjhDc3dCTENKcHM0alR4VEpWdFlUNXFqeE1VMElBVzJObmZYaThoMGpCYUtwZXhvanl0VWJycXYwV2dWQkVqOUFDZitmWG5ZMkE4ak9Tb0V3UFNKN2V6VmVKampXSVBTMVlMaVhhOVFOMThKcjNXb0Y3Y3hEd24yNnNWaW90QURQZWtpNzRud0VPMXVhMmtMcHZjbWlBY3V4ay9kdU9uVFp5TCt1VUtBNU9GU01sOTE5NExhN0sxOE5Vd09IUDFna2t6aG5xWTdVdVNWQStTMkNFQXRRS2xsNEVUMlFWdHZBRjNrSUJGVDQ1aVJoaTJaaWVES0wwaENmcHcyazhUbFVtMlFab0VEYU9LNnBsVWdvYW9TdWxsRnIyc0o0aW45c1FKTWtJVmE2MDRWUnpVT1lQZUtEVG12aUpLL0pqRUtWR1hCQ0Q2a0FNSUJxNFJDa0dleThlMmZ5N094Q0xYSW9ZdVVudXlaRGpkY2VwQ0hibFR6d3daL3V6N0w4Zmx3Mi9CVHE0VFg1WE5QVVlKbjFSdlhaU2tDTFVlWVBBSmxNMmkxSE9FaGpPcmdkSjJZNUtuWEEreEZmdExYOVlDWGxzYXBsdXB0eDZEZ3BGZ1JOSjVzTnZCTFRPM3dwTml4WUIxdVJVZ0NxRXFwa2JHY05qRVNxQW9ZN29oUXh5cnZnL0diSWlvcEZKVlB1Tk1hd3hBS1FuSFZLc0lQOHFOQVlBbjFvTU5BTHhtU3hCdUsreXlvRXBqNU02RGFtTTBRd0ZTaGJHdW83NXRqVWxxOWZjbk5TekNrSHNxMEFwSEx6MnBoZGc0bjlHSnN6QzRlZVoyL0dEK09rcmVDd0NNVWZWNmp0UGVaNE9vZE1NS1FMVUNPbVA4L3Y0Zll1Q0dsTkJuT2I1Yld6SXZiUDYySlo4V21oQjZyZEN4T2Q1Kzgvdk4zMTcvV21UTGo0YXliQUN6bURCbW9sOFFsVnJhTk9FUmFWS2JnbGtVdnF6Tk9pS3FYcWJuRVROYjBkaUFTZHkzWUZyYmVFa01CMjNZcVBER3hXVlNsWjRPNkU3aDhuM2Z3KzNUMjJuRU1HV3J4OFBuMGpueDhxQWp6Z0VHY3pRbW5vdFMwMGhwQWU0azlwdmxUWDk2QUY2RlJNYVNrQlhaT2t3ZU1RVzVwTGJCUTVyY0pkQ2p4bGF0UW1hT2s2SVdGcWsyVDV1eXc3U2owN3F5L3M3V0RvaHpGWjR6RUtJc2NWWUFJRFFkZ2hIaHBMRmJ4Y0pnTC92bGZvcmtiTjlubUw3d1hhMGxZTHpQWUcwcG9hbzBuU1c0ZXZaNmZIdlhENUNaREUyNHpNUlZrUlNXU3pxay9wVWduSytRbTF5K3VmZEtQTytLRjlFczN5VHFIT3FzNUtUMmdoWU1Tb2dDSkpzRHBwRStpNnpPZWxLazRTNkpSWFJrbU5vRW9CalJxTmU0bFR1UkpGR2NaQVZhVkJSckowVmJkbG45aFFLeFZvWTY1Rnh4Q0swWFFJUXg0QmdBZmRqb1dxRHNtOVJsTHI2eVplZ21lelRDcE40ZzljWUlydzltb0tHbllyUnpaSmJOSHcwQXVEZ0I4T0x3MGZ2V3VXVXdabVhJNENQQmVNU2lRVktwV054REJLaHp3bElpWDNzeTY5OHRVQlZZbVMrcDMwbUFyUjBxQUxCUXpndk1hTFRXcEcxZXhFQkw4aElscVRIV2FJaW52aGdWUWNmZ0F6ZStId0RTYnE3QndOajhyQmtra0h4TkpkR3h1ZXpzNzhHVFA3dFozZGdJQkIxcUUrSUtid2tTTFRWdVdZeTF4aXRydlphTjFHejdQcnlMMzRuR0l3WWFnQzRTbUMzUWh1OXUva29jWWVoamdSZ3liRlM2RW9aR2tIWGtCd2V1Qy9mZExJTjRxRWhrVUZjc1BSeVFjVkdIR0E5dURhU2RzQnJHMUVqbGRMOEo5dzRVcFlxbjU1Z0YxMmNUQUlCckV3RGpMM3FDNmVtb3NWSUp4Q01MMno2VzRmc20zbHZ2UG8rVUNOZXcvVWtTSmpjOTZUZHhrbzJNaFAxaHhMVDZTcG8wOW1zV3RnSVpSSHdkL2ExM2Z3QUpGNmN4TDdJOXd6OWVIVEMrREIvZThSblowOTh2dWVrZ2RVUU55WmloOUlaczhZQUNLRU8yOTRIQkFUN3NreGZ3dHRGNW1PNGtQSWRoT3llSlh5ODBwSTdma29oZGp4QTkyWHF4d3p2Tkl1M1FrbDROb0pJWmxyd0NUUVhxdGJwRHk4YUw4eHhaSnNiM3BlOUdDeERCSUcwK1A2T0lJWUVodnJmanU4R2tWcDgwQ1MyUTJvL0lHYXRPd0xpWjhCNHVPdlF0V3pZT0lIS1VpWWR0bms5Q0VpU2NobjNnS2VnUlBDVmYzYnFSVmtoZzNaS2xzQVp3REZYdkdvejhHTmJnb2gzZFVFUHRXc1dXU282aXFWNkREaWU3eTVycFM5VlVOZEtBYXcvZFFOQWptcnd0M2lzdWVXMTdNUm5GNGNQclJqeUJ0ckg1S1BhNy9YekhqOThOQUtIRGZ2U05rT0pWb2ErVlJHU0cxM25IQ3ovK0JGN3BibU8yWkkwNFgwUmpYMXZxcFMzaVcxZWFrK1JjSloxZVc3bFlEREMyM2xjN0txYmhBUmM1SWkzZUw4d0JhMXV4K2E0MG5SSVRCOXNxTytuWUdNVWcwYkVjRHVmQ2s1S0FFRW9lRW9yWGRLY3dwZ2JRcEVYU0JEUUNPdWdQb3ZiQ1dsQ29MMDhBTktJa2VvQ09tV01CMUtGZlUxTXd1UnlGM0tEbTkydkRPVW1pdEV2amJxODVDTnpabHRFbWFxYXFNRExHeVh4SjJFWXhZMUlCYWlDUkNRRFQ1VDdBSnJsWmkvZTBZQUpRazlrRlFPSml0M3lKdUNiZWlVeXV3UnV2ZXp1bnEybkpUU2FKUW1laVBBQ0pIZHJpTkJMV1dvd3VXU2JRVE9oU2lSZWFIYzEwbnkxOVdIOHZFbEFhSUFuYTltQjhUV3N1YXpVV3M0Q0NGSkhtaFJFTTlXY1pMSHF1WG85bWQ5VzJXWHZNZGZaTWF5NGxseXE4SXNiYVFwVElHTVBVVDlPWUhDYkxZdjZLWWNSREdFY3J0QmhzcGtXeUpBbUdjQU5xZ3Bua0FlWUdjQnBQNXRrY0FYak4zaUFKSERjaHQ0Q1hHT2FLRlM2SlZVODdNTjE4c1BrTTZpS0x0QkNOYldBRVVDVnlaQnpOUm1xeGwvU3BNWTFOV2JGQ3lqK1ROUGxwbHdmcVFacEZqNHEwbGpScDhaVEJHUm1SM2RVZTg1YXIzaEV6UFJRaEJLSUVRK0VyRTZzYmN5Z3REUDdoM3EvQmlxb0hsSVVJc2dZNnZoNFBtcVMvdGcwRTFJbEdxb3VUcmVzb1NOeTBpZ2c2UVVxRmlmZlRVTmMxY0NORzZ0QmEzTlpKd2xFUVRraVBjNVFja25wY2JZc2xUQ3BKZ2Ntd3Y1b2xBSVRXdmJVRWxDUnhqYzB3MXBzaXZJZlVuWjdpT0VJV1hwdDZhU2FvUFhaQmJhYUFzVlZPaDhFR2pMaXJFVkJaVG5rSi9aa2oyeXcxcTVBbWxRd3RXbFZhS0UrekxYZjRkaVFEWHd5Tk1iRlBpZFFHUlMzODQ4TDVXb3h5OFgrQWtGOGFTd0daTmtEYlMwUThEMGhnS2NvaFpHb0ZYL09qTjNMbi9HN2sxdENqQW1CUVZ6YTJKTFlSZzBvOWpwNDhISzgvK3gvZ2Q5eE93d3loazVLa01FK3o4ZXBjT1RZVExDMndKUnM1ZlVkNFQxTm5tKzRoWmZScS9Ub3VXbHhOQUkyL0w0cDZJRFU2Yktub0ZsZVhiRUpCK0NOd2Q2UjZ3RnJjMXQ4bEExOHhnMDJKM1JFTkJqQ0tQTTg1T2JvTThDVkNzRXBpNW5GYzdKZ1ExcFE4MXVNSnlFL3VqMWN5Y0tLQ0NwQ0o3Q2VrWXdsNmRaQ3NTUzJMWkhJYlcxSFNCNGE5SVNGcnVaOFFqM3JIQ0lsMDBFcjcxUkdwOFp1TWh1TG0xaUxHMkcrZGVwUkNRbTFQc1A2ZTJoWWlDWnA4Rk5PY3c4dSs5UmNBREZobmZ5aGlmZWVpS3pkV1NsL2h5Y2MrRGsvZjlFejQvYmNpc3lOcFVhUHF3VjEvTDJwQUxwNkxOalBRZnZNZFZUT1F3QmsrTERVc2JNOTVjZ0xTQXRSWiswbGJwN2xKd0pjMEh0YmJQdUNUUUlaRGZoWjlQMERTWWJGc05lUWhLMkVCV1dHWEFPb2hTZE5JYzZOeEpkTVlwSG11aFlIb0tFcmRXcEZnTnhDTE9PbXlJSUhyS2JCaTB1b3lMclkwT3JobFlNYUpYU1JuNDkrTGdKZEdFYmpvTk5RMHA2Sk5xK0w2ZzFKQkRldXBSRzJJcHVkYW1LMTNJeGpiRDhTL0tmQ3VGTHQwQTk1MzgzdmxDN2Q5V1RxMlM2ZGVqREVJQnpjM1psWWNQS3dZT0hYeWhnZjhIWTdOam9PYjN5K1pkSmd5VWhhWklHa2RnRFlmR2ovTk5KTVFoVmk2LzhYQWkvWlU3VWd4VFU2elBkb2dsOVpjTjVNWGFaSjY5V1hSKzVvYWtUVFJBcE5oNElad2RXWVVXM1Z3Tkdtc1N6cTl5SENZTzBSQ1dnb3NwZVluWnlxTWpXQjl2eEpVWVhpdldBbUgrbTFHQk9DdWVZbHowZVQ0QXdnbi9WQkFOUFJIa2tpTFBLOWs1RW9EdkhxVGhKdDNWSmJxRnRtQUVJaEpXUlFBT3ZtWXdFUHFnU1M0aGdWczMxQkxNcGdFVlZPL2xtbWdBcWpDVFV6aGVWOStvUXpLdmdpRWQ4d0FpWmNBUW1zc0NISkozc09ISHZodUxKMHU2SXUrV0dSdDhDMjJmMm9icnpVdXZmTVhMQnAzUFYvMWRveHpGVjlaUzdEMmU5TEN0bDdMdGcyczdiWWE2Yk1iY1pRS0FlTmNlaGZmRTZaTFFxWmFtUGpVbnRVaVI5QjBJbzNaTEdrc1VsTXZHb2l0TzkxcjgzdjRhQ1drVHRVTWw4SGFjUUtBRDZkRmhkVkk5ZTZKUkdnS1h5UElvN0pQMlFJdE9yT1doTkhlTVJRcHE4ck11MEY0c3JFQmcyQ0xpYWxMdXlzRlRrTGJDaDhuckxhTm91M1RybE1JUTVBN0FhQTFCczlLc3ZFcFhLKzM4amxmK3lOYVkrQzlSM0lqMHF2akpRQWtON2s0ZGVhVWxjZmpZeGQ4bUwzdGU4aXlIMUtUS2dKZUFFZUVLdnk0NEQ1dXpQU3ZhckR2a2ttaGpEMzAwdTlzYkRxUDhIa05yeGo1TlRRYnZxRy9ZdFBES05FQ0J5czFNRnM5QzJvSFFFRzRCTzRJUWhFVXcyazRWOFlGWjYwTFNUQzFJdloxT3k0WHBYVDhVeVJvSk5aejBNUzMwejBrS1oxMmJFcG85LzR1RTFMVHpraVdRL3ZtV1FmTGEvQUphaHN3NWEvZEVmbHhjeGdqSUllY0syYkRyYVpRSEpKbUN0Y2twd2duY2NWYVlHdFMybHZ4T1VRR2dDMHAwTElGa3IxbUJLNHFrYTA0Q3UrNS90MW15M1VmWVc1elZGb3hKYUhld1JRRUFHUW1RK1ZMM1BlSSsrS0RELzAzY05kT3BTb01Nd1JTSHJJNDdFYkV1V2hKb0JoL2E2VGhZZ2tPb2o3OE1JNjZSVXJIejlIMDJkR0xUb1lVRzVXWTdNRTdYczFVTlZ1KzFsNUVyenNHWTV0bUJRWUk0d25Hb0FCQTZZcndoZ2lrNW9PVHRwSGF3b3JqV3Z6OXJPa0NrWlN6dUxnVFFlc1BYOXRmemMzVWtzKzBiS0RJVzdXVFRTU0ZhbG9Jakg4TERlRUx6QThPdGRDVkZLZ3ljYk9uclR3V2NLR3QyV0lSSG0zQVZzcDFMSVJwUVQxTkFCcnZPSDJMTWZET2lWbStUcC85K2VmTHRRZHVScFJ5cmRtNjQwWGtOcGZLVitZUnh6NmM3ei92WFREYmJrZm9xMnpiRXkyTEFTYUxwVkVkbzAydmFkbHpjZjRzTEV4akIyTFJ2bWlyY21sOUNOUGZyYzhtR20rNHJTWFM5MUpxOWdlcUhKVlI1TGJibXVOQWkwbjZmQUN6WmQ4QU5tbnFNS0ZKTU5YQjlDZ0FOQkhIYVN6SlBKT1VhaHZlWjh5aWVXOHlVeW56Z1pLS0h4alNYbUljOUU3MUcwbUpCWGE5dG44RUNLZWxvRllqUG9qVjBnM2pnTmpNc3pGSXVTTnJSbFlDMVFCc2YxNVNZUUdEMG5pLzhlWnFPaWhPQ0ZzRFRCS0ZDQ1oyTm1abXg0aG5mdnkzUVZVSmtublJYTFJFajFBaGFtM09TaXZ6K0pNMnk1Ynozd2R6K3phSWQySm9Rd1dXbW1ieDA4YW8xUkhSa05OSWFyamV5RUtCUVE3ZlB3alZBamJyaHJpa2ozUk5Vc0gxME5CK0xCNGVtTzQxcXVlVzNkMXM0RGlXdW53QUFsZkowczVTakdXallXS2IxU1dvTU1heVZNV2hoVVB4Q0ZGdDFoZnh2bW9NY1BGR1NXTk1FamV4bG1udmxGZzA2VFVBcmRQYlROK0hUYUlBMUFqRUlMYW5iOTZVZmsxdWY3M0swaHBVTTJuaHdWd0cxWkR4YnBNWUNGSXcycVFqblNVRU9vRHo3WUthOE9OaXZsOEtVTlEvYkUxQytqMFp5c25BRGdoV0hTS2JXb3Z2VkQvZ016L3piQnBZZUhYdDFMSjZSek9nM1FncHVjbForZ3FQT2ZGUjJQTGdmMVBac1lNNkdNQ3lDMVErOElTT3JUSEhId2ZBUnpCNG9PNGM1Ulh3QWpxRjd0dU8zNUxqc0g1NmhIN2J6ZFM1V1ZqSllWeEdPQVF3cHZlNkNNejBlQzNsR05LM2dtVWxhSmRKeGdtSU55a1JnSVFydVdGa0JVWk1EcTh4alowVVZaVjRKMUs1UXZZdDNLN0lMV25pUmtvOFlOSXlvU1FqWlYybk1zNW1NMmdqYldKcGVBellBOWdTcElqQnRZR1BzWTdicFl6N0lXZ3drVVF3cHExVXh6TGpoemZpT01DKzN2bnAzcU5xeVR2OHdZRWIyOEN0TjR1SisyL0QxSG9jTlg0MFdBNWl5a0RMNFVnU0xhaXVoaEpxdnlhZStGMDdRRkpqUEh5aEdMcHlnR3o1TWVZOU43MGZyL3ptcXlXek9Ud2RrejVwS1c2MEdZM2NabEw1aW84NS9qSHk4VWQvRk12bkNEKzNYekxUQTV3Q2tMaklMV0FrVERPT1BaNXNMVFN3VU5vRCsvaTIrNzRGbHovaEsvenVreTdEMisvekpweGhqMVMvZXl1MU9BaHJNcGc2dVRkdXdDQ2NJeU1RZ1Y1N0VQR1lpdWFPazEzUFdCbUxvQk1zVUZWWTJac0NBSGhWaVluVGJkMEVxRmRYcWtBczRnbVJ3ZE50aDlzQWdaRVVGbzJQdHFWMW5FU0plVndDb1BUUitFaWh1SGk1ZWU1aHFtSktJVVZGRTRaaVhXZWJGcWN4Z2hjQkpRMHdTRWhWSmJKY2ZyRC8rNHYwQ1VOSURFWU1uVG9zNjQ3TE1lTXJpY0VDUkd3emVZaTNFSFpXcXk2aFhvSFdwN2IrcVArT1g2c2lFQ3UrS3BHdFB4NlhmT3RWZVBmVkgwWnVjamhYdGU5Z1VhWk1sSWJJYlM3T08zbm9rZWZ6YTQvL0VzN0FFWFM3YjBXV2pVQ2FLaXRwMlZ6U2VMWlM1L2VackV1LysxYTg5cDUvSmI5ei9CTlIrUktyUjVmajJhYzlBOTk1eWxmbDM4NzdGNXdseDRuZmRodDBmbFp0bHNNWTIreUx1bG9PelU5ZE5kKzIzOU5yQTlzVk4wRjRpUy9SelRvQXdscWJ1TlNwb3prQWxsVUZuMlZCRTk3aEhtcGh3UFE5YVd3dDdqTnR2bVFaSmszUVo1Q0Fkd3pGbWQwTE0wWTFjakRSb2lJak40ZEd2YldsRHFXNTZjWFAxY1d5d2JQcWNMNC9UNmMrOU54RG1vendmTUxLQ2N2UEVCbldrWmRVZWhsdHdYaTJXVzFUcFFscC9kVEFZKzNWMTJPTGxTQ1VjRlNFV1hNNGZ1L1R6K0lWTzYrUVBPdWc4dFdpR3lQQWxIS1kvTExNWm5TK01pY3NQUktYUGZsTGVQckdKNGk3OVRwUWxkYmtFcG9rSXFqRE9pd1ZQOVlwTXR1QjMzTVRublBzNzhrTFQvODlWTDZFTmJsNHFsVGVBZXJsaVNjOEJ0OTh5aGU1NVdIdndUM3RzZkRiYnFmMnAyRnNEdU5ONjdQQnhXc0JObHFwWmI2QTBZeEtONktBTjdKeTZSRmh4NFQxRU1SMHRlUVYzakM3VFdhMEx5SlprclJ0N2JZNDN0ME9Vb1M5d0hxVENFS0JiNHk0U0lGWkFNQkpxN2dZZ050OGFTckNHTkJRUkd5TFZWNTBReWsyaVVBSEpFZWhsbjV4Sk1sZVZVTHNLTFl0N09FUEQxeFBLMEtsQmlLSmxGcmxBM2pRcHJQQmNoaU9DbTN6ZTR2VWJoeFVLT3hoL2JQSTRFL3R4Umhiem1xa2JPTHhPRm9DTnVOd3pVcDUyQWNlaVd2MlhTdTV6VkY1UnlKV0VoQlNFNlBOWkV0bWN6cjFNcHAxOEs4UC95ZDk2emx2d3JKZGZmZ0RPMkdrUStOTVVNc2VDcFhBL1RtbGxZNjZnN3Q1ajlHN3k1c3ZlRjBvemhkTEk0SDF5RzJHMEtuS1FWUjU0ZkdQNHRlZS9nWDU4QVh2d1ZrOFd2VzJiYXI5R1pwc2hFWTYyaUtCSS9qaVJrdlowQ2tMT3pnMU5VVVV1azEwOUg2cnp3QUFtblRxbUVpa0x3UGFmcnozT2xUbHRGcGp0SlY5M3piQkdodXdlVHkxUnd2MnVwT1VyaEoxR0lIY2JnMVR1U1hnTHAxb1kwby9rSUVyWWs4TDFsbk9yaTFKMmlCTFVReHQ3Y1JvcnpRdEtJUWdUSlp6b2RndjIvYmVCQ0FTN1VSb0RTLzFlU0I2N0tyanVEUmZwYjRxSUNydDhCcWF4RWNnWlNVdkJtVHQvY1h4R3FrenRqVVJBU2FxbUhEQWk4bkhzRzl5QkJlODZ5Rzg4Y0FONk5oY3ZLc0lLRnJ0YjhOeGJ5Mk5uNWx3VHBYekZaNXoxbS96VzgvNEtoNjk3SUhVWFRlTCtnV3hXVGZVa2tYcExiQkFXY3JVQXZBdkQzczc0NGtKWW8wTmV6VjhXV29hS1RCQTVaMVlWWG5zaVkvQTE1LzVCWG5QQlc4elo5bGpvZHR2RSsxUHc5bzhucUFURnp4SnFDUUlXa0dROEs4a3UxMVlEckZ4YkYxMHArdmtUQnFqcWIwZ2g3T0hHT1BBYldPOFh2cTBJZVBUc2xnS29xSHBiTWczZ2hHUm9VSm05VWN0L01IZ2t2QzJEVHZtZGxHeE5UUTJwRElkRXBBYTFOVFV3QjBHMEZhOTZXN2JxaGdnVkFXOUNmbkc5aXZTKytJTFRXaEtaQTByWDhtbXlTTnc0dWdtd1hCT2pOajZuaHRObUNSU0JCYWpjWjNPZjY1ZnFXMFZrRDRqVEVrNitkc1k4ZDdCams1aXg1U1Y4LzcxZk54eTRHYm1XUzZsY3dEcXZNVjRROGtFQ210a0lUVFdTdWxLT1hacUkvNzl3ZzlneTRQZWcrTVdWbmkvWXh1MEdJcXhWdks4aTh4MnhlL2NKbTkrMEQvaXBPWEh3Rk1oZFE4K1JJTXFqaDNoUUpsUXNRWlc2bUJCUE9YVWkzalowejRySDM3SU8vRmI3Z2p4VzdlQ3JDZ1NPM2NsTlp0V24zR1RJcHIveWVLcXZDekwxNWp4aVpWSTg1SENrMXFYWUFLM3pld1dkTWNONmNQV1lBdGRTZUxXREYrVDNWWUxJUUhxNmpoQWtBSHNLK1RhL2g2MHJsaUhBdmt4TU05NXYwTnN0QUZ0K054SWZ3VENxNjdMaUdLWFFGMkpsUVJHdlJ1WTVrTGdsZWgyK0xVZGx3dnFKRWpXRTZacUJJR3ZOQTgvNWdMQi9EekZ4TTVjVFRhd2dhK0xieEJkL1RRUkVuZTQxSDhuK0NzRDZjWm9ZdGZHZVZnYjd3cllpVWxzbndBZStKNEw4TU05MTZDWGRWQzVDaUJUdXlDa1Q0Mkp4ZFFvc1RwWkp4eEQ2NTFjZU5MamNNVnZmOE84N3F5L3hFbkZLdXEyM2F5MjM0VHExaC93MldjOGwwODY0YkVvZkFtYnV0UEUrMjk1ZGdueVlveUlpRWh1TW9UZTExNXlJM2pzU1kvRzE1L3paWG5yZVgrSHFaMTljSDQya3YxeFJScnl1OUZja1RxeHhvSUwwN2pYcWhPeGJtSWx2RG94RkRFd0Vrb1dFRFkrZ00vZCtpV2kxNk5XS2pXZHREakczUW9MSmhVdGpacjJVVVZia0lRVEVUR3pWTE5uTUhkSEFBSmJOb2Z6bXc2NVdSR0J1TG8wSUh5UTg0M3VUMStldkNGcTJIMHBycGxzZ01ZWUZ2VmVwTGNVMXh5OGpqZnV1d25XV1BFTS9iT01TR3dZR1pJUUhuSGNRekJXalZMcDBtb0htekMxSDZ0M1h0eVJtallFbTBtcUhSYzBoSGpneHdJbG12cWRTQWc5K1hJSU83NFV0NHlYdU45N0g0U1AzL3dWeWJNY3BhdnFaZy94SWdDRzJwSlFscXdBTW1PbFl6TTY5WnpvOVBEQ3M1K0w3ejc3cS9yUlI3NlBmM3pVYzNuSm1YL09mM3JnYTZsZW1kc01NYjlUYW1jK3FvL3dFNkVaaXExQVFJd1l6WXlsTlpsVTNrRWcrcHpmK20xKzhlbi9nWlg5akNncWlrYTZwTFlGSmZ5UzFrbURVU3ZPNmZGclQ2UWduUGNUZWo0UkRJZUcweHJEL1hQN3NXTityeUR2Z3FFZlg5dm1penhnc3JzMXJIY0syd2FPc0drdllsTUlSd3dLN3FpR1psdkFYQk5YcS9WeDV2RWpHU3JFVWdWR2tDSFVpTVRHMHdqQjc1Wk5pT1FJSUNRbXhOV3FuWWpHUURNMng3eGJ3QmR1KzNJOTl0QjZNR3pSekZpcUtrOWFjd0pPWEg2QzZPd01yZGo0Z1hHYkVXaHlGUVVOTFZQditQaWJOTlJOeUl4cTl1MWlWUkszb2NEN0VyYTNWR2FXanNwanRqd2FIL2p4eDlETk8rSlJ0VVJKWkJUQzZZZ2ExR1dRNWhHSVVDZ3FyZEN6WFhuVThRK1dWei82MWZqekIxNEM5VjRvb0lIUnlHTFVIOXNNcnE1M2JKNnBBeGhCTWVVMnlJYkNEZVNNZFNmSjVtTWZBODVNaTVYVTdMT2VwNVQ5a3NaTjcwSDJLM25VOFErTHVzdkVBdkF3QnhvOTRNdTNmZ2Y3QnJ0aDdValRxVFh0RWNaQk1RS3hUbDFxRzRBSW04R0lrUndpRkNOZEl6eFk3Y0Q4L0Q1Y2ZMRkpiNGhlY0VDZzdodGVnWUVIYlJhMlRCNUZxaGZDMVV1UVVvR2l6eTMxRS9GeEpsZ2hKU3hBQUZYQjJMaDg0b2JQSXZGc0ZtbmJoODlLdHNqbTR4OE56RTVEYUJ0amRuRUtVdlA3SWt1eFpaRFdubGthb3JSTUJnbGhOS0RGcFJsNjUyaXlFWER0V2o3eHcwL0ZtNy94WnVRbUI4aVkxcC91U3FCbzV4V0YrMUhBR0Jqa0pxZXFGNmNPbFN0UXVpSXVrekhKM21DZHdGUWJhR20wYkFBbmFmalNRaVVnTUxucG92Q08xeHk0QmVpTWhDelNSQTdYOXlXMWlXUmd3R0lCUjQ1dnhFbXJUaFFQaFpIa2c0ZEdsU21BZS8yKzI0ak1CdTgxQ2Voa1lrR2tQdVZiNnZWSTRBdy9vV0pTSUVLRVBueGg4dVoxTndEZ3BHdnJHdzV6R0QxaGUwdi9CZ3lxUHJQUUhSZTVDV01yMk9iZldvc1dVVkNyMndRT2pSSXBTazJCcUs4ZzQxUDQ5dmJ2WTl2MDdaTFpjSXAzRWtna3hjVE9wRTg0WlRNbVpKSys3TWRvakRRR2RzSkJzdVVTU1pwS0dldFFYQVBMSnFzWnFGUEdRK2ZCeFQrV1VEakFXakdIYlpEbmZlVlArRHNmZmE1NGhIemR5cnY0MVlSUVlyRlRiZDdIeVF4SFNSaGprWmtNZWRaRmJqb1NwaFJTbHdRbVdYY1hsOVRHZEpNL1VNTVRJWXZTR01PWi9pSDU5czRyQkNNam9mdzB6VWY2aWNJS25oUXh3TnkwUE9DSSsySFp5Q1NkY3pBaGxoQjBwRklNZzNuK2lSOTlVdER0QmZXYjl2Z2lTd0Uxc2RMNEEwa0FFWFhjT3hmUWdqUkdNU0JOSVo4SEFHelpVdDlybUxOTHdweWVlZVA4alRKZFhTODlHbEVHZDhGS0U4TnNCaEMrS1NVT2dNMGVEU2ZUcE1tbytUekMwWFp5SE5SRCtNQ1ZId0VBOGZRMGxEcGZ4OENnMG9vYnB0YmdzUnNmSWp5MEI5WmtyUjRrelZmWGwwOUdjQnhjRy9qdHVHaGJFbXB0RzBSdkNJQ294QXdCVUIzVUV0bkdvL0hQTjc4WDkvdm5jM0RWM2g4anR4a3E3Nkw0V2xRK0w4M0k0cU10Z1FrVHFNbUFxaVkxcjNYRjdWMTdVSnFrYXR5Y1NDWDFDb1Z6b1ZmZGYxejNlUlJ1RmpidmdpbUx1dTE4QklGQmhFNjJJcFhIazA5L1BJRFFyRHhhYndRQ2xacVpEUHZtOXVHRy9UY0NveU9pM2pkbVZacDMxc09OZGxSdFd6ZlBLd1FPUkRjS1NVTWpmUlYrZDNvcjduQTFXbVFMekdXQWswUFZyV0d4aExBRWpMQkp1SXpPUmxyZzVPcTNqZi9hbVl1RGtlUXdpR2psSVpOTDhhNnJQb2hoVmNHS2xacDhEMlZIdGJYMm92cytGL2xDekI2c2pkcVd4NVdVckJBMVgxaG5UZC9CTEVnMlNwTFFpd0xTbEx0OFBRV3VMSkd0MllUTHE1dHg3aitmanc5ZDlXSGtOcU1SUTY5ZUFtS2pjYXJCTm95ZlUzZCtTRmlRSnMwdEVSYnRkWTJDemlUMHhMaE5lbk1LcGdKQ2tVeUMybnovdFI4akozb0M1eGQvWUVKL2xQakdkc0NGYVo0OGVUTHZlZmc5b1BDd1lnTXFnejBidWtnWThLczNmaDI3RnJiRGRzZEkralFuZHdaMzBqeDF2VWs5MkRxTFNibzJPSkNaeWN6K2N0cjI3UTBSYS9XcjJ3QU1iNThwUDQ0RkJYT3JFS0gwSXJGY29hVjY3ekNOS1ZJaHdLS1FYVE5RZ2dMMUh0SmJ5bXNQWFlOdmJ2czZyREhxdkJPbXFoaFF4Rm80WCtIVXcwN2xnNDkrdU9pQlhiUlowNlcrMXBmTmdyWkFsOGJVdHU0bDdjcEdnZ3VhRFpJaW9lMFVwcUQ4Q0xGd3hSQjJZZzJtVjB4Zzg3OC9BOC8vMEFzd001d0xKdy81U2hRKzJMUjFxRHdZdmJHY29iSHpJdkhUZ0xTNVd0c2h4YUZyMjR0UnVpYzcwVk5wcmNWVjI2L2g1Mjc3b3NqRUtxcExIUkRRekgyeUFRRVlhOEJEMDNqMkdjOWd6M2JnblFxTVNSWFdxTTBvUUQ3eXcwOENJK1BoekxHNkZya1JJdEhHVzJ6MzFYbUljUTRkZ2N3QVhRRWNhRE1SM2VsdUwyL1pmMk9NRjlkaWRWRjJLZ0RrMXkvY1lCY3FqeXkwNkdYSEJQUEdhMHhNYU96Y2VqL0lIU2MxelI0YU96RXV2QUZGbG96aGRkLzZ4L1Q5clJMeFlISUdoZzN5VitlL1hEcERnTjVUYWhzdnF0czdrdUtwakxRdTIyeUJzRmEvU1UzWGNxa0Y2allpRUZaUHZNQVllRitJNUJuTXhpUHhwbHZlTFdmKzQ5bjQ1aTNmUUdaekdsaFVyaEpTSlpGRjBhU280OTFFWUN1a05tR0luM0RWcjBpVkdNQWlnSWF5U2dBdisreXI0RWRDTWdwTm94WVdIVlpDaFNHbzgzTmNsNjNEMDg5OG9paEFtMlZHWWdXOXhDNGplWlpoNTZGZC9QeU5YNllzbXpLK3FzTG5CWTIzV01VbU5WdFRXbWdjWVRJd0p6a0VSc2pRZ0F0bUFkOERBR3k1NkNka1JHOEo4WVRWYzZ1L2IyZktXeVdYUUNGbUFMTG9CYnRvNHdHbzZaaDJEcHpHTGR1RTRocXZPWm9OVGl0Z2NoVStkZE1YZWZsdDM1SE1abEQ2c01ralZJeXhySHlGazllY0lFODU0VW5RM2R0aGJCNmI2ZGVzZStJZG84aE5pam5GUHRQMklIREgwdExrTzlSaHFqdUFJbkdFUXNCb01NT01oMnJGYk8wUnVMbDdTQjd3M2tmeUR6LzZFdXllMjRjOHl5RktWbHBCWXllVG1xOUVLTFlKdWJlZ3lKMzJmQTI4T0xYU2FyWVRWelowUW5UZVNaNTErSm5ydjh6UDN2d3BzVXZYd3ZzS3RYZmFCa2hnSG1CTUJ0Mi9HNzk5K3BOa3Nqc083eXN4MGVBTk9XSUliZHNBZk9ISFgrYSthaTlOcDRlbVJWbmM2V0Y5RzZCSGZxZzJhRlAvR2M5UXR0Q0xpY201cUpRTzVzYmhsd0FBYjk2eWFMZTNaNFA0NEdaNzAwMDNGVGlrWDRsdWRzQjVWeUloelZod3pRWlliZEhyZ2Fad0tQMmJvQ0RCWVVBOHYyUzBJMy8xcGRlbU8yeHM4eUQ0SlZJZGZNVzVMOEVLWFFZLzZIdEpOdHppUStVQlV1djJFY0ZlRkxUNzVBYzF6UHBkU1M4dkVrUkphcmVOTTRsZ1JTMldYRlhRakV4aXVHNGRYbmZOMjNIbW04L0c2Ny8yWmc1Rm1XYzVBSVRUblJianJGWjJpS3FnRGZkNk1Vemlpa3hyNk9HL1NoVWpCZ3ZsZ3J6b295OFFybGdlRzVraE1Zck5PS01sYlVUQVlsNVdkbGJMYzgvNWZaSVVhNnlvQmh0VFlvYWZpUzFKMzNINSs0MU1UQmlVbnJYY2xSZ0lDSGVSaEVrZDY2L2o3VUVZQlZNTkF1bkdDY2drTjRlcU9idkgvUUFBc0dyeHJDK2Vwb2hPZitQQzV6QndnSTIyY1MrVzhJUU1oeVlkUEMxbUxlMmlEVmEzODJJTFhsRk9LK0hMRW1icGFuNWk2K2Y1alZzdkYyc3NYYVE0RWlLdFdLbWN3OFpsNi9VeEp6MGNtTnN2RnAzMmdqYi9wdTlMQ2JPMUxWb3ZTKzBDTjgvWHFoc05kV01RczcrNUtMTW52U25lcjZwUzZHSFhIWTZkU3hVditzcWY0TXczM1Fmdi90NmxnREhJYkM2a2w4cTdaQzAwUjkzZFNkOHZJb0lndGFKbWJUOTZxaWdWeGhnODVkTGZ4L1hEVzJGR2xrS2RKMnFia2MzbUJ3UmVJYllMdjN1WHZQanNGMkgxMkVweDZnRXhySU10SUp4NldHUHh2VnQvd0cvdC9DWmxZaG04eHJZS2d0WUdsR1krNjFtVVpBNkYxNmdJU2lVeVlaQ0FwSFNzeU0zVmJRdTM3UGtSRElBdCtBbFZjUUJ3V2NqKzYrM1k5eVU1TU54cmVqWXpJS1ZqZ0o0QUZjSUpPRUgwdHRwSXRQSEFwTnFZVm0zUjlBT0FDQXdobU9qSmE3LzgrdmplUUNiVjFXb2lxVmhCY3RpUTFXdGFkbDQ2WHI3bXZXZ2FtWmpNQVRRcXQra1l4ZFpDTlJPYVBPeWt6dEp6WUdQakpwZ1lDQzNGYXducGRKRWZzVW11TlR2TjB6LytESnozbHZ2TFo2LzdENGpZa0dJRmdYTVYxUHRZVDVnS1IrNGtmb0U0VXhvSHJBQzlkN0Jpa0ptTUwvblV4ZmpvVFIrQlBleEllSlFDSzdJb0poN21RZUJJQTBzL2Y0Z25MenRkbjN2dlo5TnBSU3NaRVZJTzBnbFQ5TEhjNG8xZmVwTlVvN0ZpUWlDZ1NsMCtTcmJxV1poNDFiQlpOWVZCUVZRS1ZCVDBRaUsxV09Pc2VuQ3YreWdBd1NzV241SjBad0FDaE1MTVhvdUQ5b0QvbHVRQ2dDb1ppSzVoVTlPZ3FEZW9Sb0MxMjBra2FSRzhwWVljRGJ1VUVJRnpGZTNrYW56MHVrL2dVOWY4aDJRMmgxUEhLSWNBaEJocjVaMWN1K3MySUJzanRRN0hKRUdaSkJzV0VhS3BzVUJTQzR0TUZ6VGh4TnBXd21LbnBiWjFBdHBxSUxieFVqK2xVaFVEbU40RTdKSEh5R1hWOWZLUTl6NEdaNy9wL3ZxKzcxMktnU3ZReVhKbTFrSzlZK1c5T0I4T1JJZ0NtN0VzZzZvQkhLb01YQis5WkRialhOWG4wejcwQjNqTnQvK09kdjFHZUZjQk5wa01DRHJJbU9EK0J2dFhqQmhnL3pUZStPaS9rL0h1Q0FCcEg5cElZd3dxZGV4YUs5ZnN2cDRmdlA0VGtLVnI0U3ZYS1BONkF6TXNYbkk2a3dCcXo2dW5vSW9hYU5RRUhyQkR5LzJsNTFVelh3QkFYTHQ0MTkwVkFJRnpnMmZhdVhtNHhRd1Z0RWFoRUhTTndDTFlnQldTbEl0aU9JRXZxck5tZ1pza2dxU09VN2s3UmVpSmJIUVNOc3VDNWRMNGljSmc4L0RnM01IQTl2ZEdxTDZtUytxRGFob1N1aldPUUpMWERFZ3psamloNlpTbTJ2bG9LY2FXa202VlFyYitYZlRDd0Z5S2daTDBaVWt6dW96MmlFMzhabm1kUE9XVHo4SUpyenVWTC9yb3krU3JOMzliakdUSWJjN00ybENQSDRkaGpJR3FyK3RqTW1QWnlYSllzZnpvTlovQnZkOThqcnpueCsrRFdYZWsrTXFqOFVxUkNIclRuSldyeURwZHVCM2I1SVgzZUNIT1BmcmVVbnBIRXgwT2c5ajNIZ0NwQWhqOHphZitqbVhIaFg3UXlWSko5bDN6ZDV6M05NOUF6UU5DZ24vZ0tNZ0Zac3hTbEpRUmE3bkQzZVJ1blA3YVhhbGZBSGNXaVZFTk03dHMxNmVMTThiM215UEdWN0JmT2VRbVExZUFQZ1FWQVNmaDNZYU44Vi9qVzFJaVFBQkpIVGFXdXJKS2pJaTZnc3ZNbU54NzQxbEFuSndvMXFBTVFaaHQwenRaaUJmWUhMRlJVbU5EU1FRR0k5QWFyQWNBS1FLRGE2S2hIbm1ITzBqcUJsL3R6NnM1NWZoN2VrNFJHbmkyN1NOSml3R29lZ0pPek1ReXlPUUtiQzM2ZVAyMS84dzNmUGVkT0dYcUtIbmNpUS9uS1VlZGd2dHR1QmZIT3hPd1dTNVdESTJ4VUlDVks3QnpkcmQ4NG9lZjRiOWQ5WDU4KytCVmdxa2x5Rlp1RkZjVlJOWUNRTExCMHYxVGFDU0htejJJRTNxYmVQR0QvMWljZDdSTlRWQ1VZSXFLaW83TjVlb2QxL0tEVjM5SXpCSHJ4RmNsYTN2ZEJJWWZyT2V1dGNIdllCZDZoaXJBU2lETExjVUlJUER3TUx4NStDRUF4SDJSNGJJNm8rQS9BU0JBZkJCMjVpSWM2dTR1UCswMzZkTjhTcXZwV2NHQ0E0WUM1Q1F5aER5N3RzR2ZDR21Rb2JSVG0rZGlBakEwOEU5Y0dNaHhTMC9HZUhjY3Fsb2YzeERPSS9Td01QTEZXNzlFbFQ0enNlS0NHNDFHK2lBcUNta25MWVJlVERYRkVpY3FwaUhYc1JaaHkzbHFicjFXUDhrZlNHY0ZwNGZyRWdLbVpZOURxcVdvQUFKVkY3S2hPMTNZdGVQaXZlTlZnKzI4NnNxL0ZmekFZRVYzRlRxa2pIYkhjZGpvV3VubFk5eGZITVR1aFIyWUtTdk05L2NJcHFiRXJ0OEFxb2ZUSVdCak1YTUNYLzJOUWVJSHg3K1EzblRCRHozLy9aZ2FYWUxLVjZIMHRkRjFWQU9CRHdOKzJjY3ZZVFZsNjBxZHV1Rmw2bFZkMS9aSUUyWk5HVEN4eFQ0cUJZbzRwK01aNEtHY01OYnNHTHJzeHd1WFZnQndMaFNYM1Jsc2R3WEFKajNyeXJuMytPUEdueXJqWWxnQ0dCRWdsL0NGQXdFNmxzamlnTk95Z1dHUkc0a1ViaTJCSXRLVzFtU2lDOU04OTlpN3d3QW8xV3VXNVhWaEZpTTNkV0QvSHNERy9nOGVnRTA3a2xKTEp4T1Jrd3FYMnJIaWhKeW04akpncDUwUVJHMTV4bXplb3lic0NGRnBvanZKMFpKMEwyZ2VCSkFpSFdKaTRwS0tjd1ZBd1BRbVlNYW1vS3JjN3l1Qjk0QTdoSnVtOXpNVXp4aEJOeWU2SFdUTGpqTHFQWHhWeHVabWtyelNlcExyelJJdGw2eVRTM1hkemZxNkM5OGhKNjQ5bms0cjVEWm5nQ2dUWFNPZWpybnR5T2V1K2h3L2VjT25ZSS9hS0w1d1NkYlYrcllHZWxOekhiNVhrYUtHVFdTa0pERnBhWHBHcUtSMHJUVzNEYjVlM2pCOWRWejlWbEM1dWU2U0ZjVVdlQkN5Y1BuQkw1bmIrdDh6ZVdhaEpLd0FJekU5MjZsZzJMYTkwTmlBUUdwZ2ZZZm5nSlM0cWdxZ0JJNWNkMHk0VFZrOGhDekw0S2o0N3I3cmdlNms4YzQxa2k1VW04V2E1YmJLYk5tYXRZUnNBU29WdHBvVzY5Z1NuZlhGZGp3THFEc2dKT01iN2NHeUNZR0ZvRzBpd0tVR1NYeE9xWEJGUmEyOENBd2s2OERrWTdBalUyTEhsNG9aWFFLeFhTT0FjVVZCOVk1TkM3ZGFnVFlObVZMcGdSTmtlUS9WOWRmcC8zM0F4Znk5czUrR3loVmlUQjZ5R2tMeURtRkNyWlNSREdWVjhBOC8vQ2VRNVZQQ0twbFJjU1hxVG1mU2FJbDJROUJZS1FOR29WQ0ZOWmVwVEtpQVpFYnNySWU1cnZ4bkFJeCt4VjFlUC9FSm5Bc0xRTE50L2JlWXdnTlpuUGh4RS9oQkZXQ29ncXBldExhbVlzMVpzQTJTOERvQjRIMkpiamJPY3pmZEp3eEVqQWdoR3F2WURBem0rM1A4M3E0cmlkNG9XTVhFdFdUajFZVTRhWElTMkNMaTJ2SEpaaUxiNGZUVzVtaUJyTDZIT2plMGtiVGhQcHRGU1NJeFpRS2xKSTMyZlNmYkYzSFJZaUYzRTFIMDRYL2VpMUtGcVM3U3RBamVaSU1tNnFzMkE4SW16MGQ2NG02K0hpLytyUmZ5eng3K2NxbDhDV3V6T0VjaG9wTk9KdkRxWVkyUmwzejBMM0J0LzBiWThTbG81UkEzVWtKaFErdlVhMWRyczJUSGg1bXNRQXc5TVNxUXNReHdWSXhiaTV2Nk8wWXUyLzF4RUlMTDd1eDgvTmNBdkN4SXdlNi83L2tJYmgzY2pwNDFVSHBrTmtoQnhKTER2a3BkLzRyMklrWHBGNW44eHQ2aXdCSndReTYxRTdKbWZKWFVBQlJBakpGVUFYREQzbHN3VUcrUVpRajlMT3FGandCTGNVOXQydmY2S0NHU3ZWSlRSNjBkM0J4QjJoS0FiTFZOYTk4REVyRGI5OWNnTXRGTDdRNVZwdlY2cWI4azBkanBUZUZIQ0JnUFdFK0lUNmRkc240dUFUdmNuMFRiTlhxL1JEN1NRN1gxZXYvc1U1NkYxejcrTmNiNVNveGtFTVJrQTQzMmlnZ3I1OURKY243aFIxL0FXNzcrQnJIckRqZStxQklyMEc3dUZ1Y0RyUG0vTkpna0FMd1BkZENGRWs2QjVYbVlyaHlLRHNEcmkzY2VBbWJ3eW5OUzN2SFBDRUNBZUNYc0lXREcvSGp1WFNZUkdrSmdpVVYwTHNJQWl0WmkxcEtRRVNBeGxKT2VBMkRGRUhOenVQK0crMkNzTnlwZXZjVEcrbUYyby8zM3Jkc3ZoM09IbUlrTis3QTU1aUJLdVpTM0tZekhpS1RuSTFqcW5kbzhIcWdFcWNGVWU1SUpFQzJRdFZWNytyZGxCallTVlpxbENlTks0R3ZGb0xYbFFMQ3g1Mm9LcS83Z1ZnK2NwTWJaUkdiaWtmVWlsTHpYaytyRzYvQ2NvNThxYjMvcW0wTk5zUmhZWXhLZlRHTXNEVXlNZUJoc1A3UkxudmkrWjdOYU01V1dwK1hNMWRxcVNWR3I5MTFiaXlITXRhT2c5SUtKREdZaWg1UktNNUZiYzlPd1gzMXU3enNoQUM2NTdDNXR2M1Q5WndBRUxnbFNNUHZXbmpkeDYyQWZ4b3hGaFJCbW1iU0JLWWNBQzBxNHVHTnJlMFVDUjFZVHdiWDBDVnR0T0pBalZ4d2Vsb2FwcnBJU2o1Y1RBTmh6WUk4aXl4dGp1NTZBOXFRSkY0V2dDTlFOd1RYQ3BabGdxUmVmU0FSckM3RHA4VFExRVZnMFRERzFSVkdVU1A2RXpZYjBJWW1nVFJSR1E0eEhxaVFDY3ZIQzFoTFZOSFV1algwZE15VU5vSVF4VmtSeVZEZGN6OTg3OWRsODZ6UGZKazRkaFlBMU51YmJKdjVmNnhRd1l3eCsrNTNQd3Y3T1BHeHZDVlVqMHB1TkxZc2RNWWI3cTlWeHN1MUJlQWpLS0dCV2RjUFlEZFZZQ0w0Ly8xNFV4VzM0WUREai9qT0kvZWNBQklndG04M0NYdXd4VjgrL3pSZ1JaRlI0QWhNVzhVQ2IwRGwwSVNWRnRpVUk2M210bjFPQlV4WFJIdTl6ek5rQVFCTkJKeEJBRlZZc1BZQXZiNzFDTURKTzlUNjUvdEtrVzBVd0tWQlhabm1rREcwUnBWaXhzTWFHNTFMSlpocFArL0NYaUlPMjMxRnZIQ1JxSXQ0VHRDV0JXeXFwU1ZGdkFKMDBnYkQxZDNwT0drMjh5SmFOQitNa0NWbm5XcXJBSzdPc0ErMFBZVysrblgveGdGZmhINS95Wm9TWU04WGFMSjRDRlQ3WGhQT3h4YWxEWml5ZTljNW44ejkyWHdhN2RKMzRxb2lBandCZm5FaVNQTnltUTBVVUhyV0pVa1RxWlRLampCbEY0WW5Kek9pUCsvM3FQMmIrSG9UZ29wK3NldE4xMXpSTSs3cG9pNEtRY2RuMWQvMWpSNStxSjQrdHgzNGxPbEVLN25lQ1RJQUJnREVDWFNKbVJFU1p3eEJCcWFlZW9DbzZsY0dwYTA0TUhvSXhDR1hCUVE5WmE5a2Y5bkh6dm1zTWx1ZHBGbHFlYTV3Z2pkSkhBZERUR0F1eE9XQUJQK3pUeng4S21uQjBXY2h6MTZoaVRZeGZKaXJHUzdEYkpJSWg2ZHdrT1dPYUlpSXJneVJWZ1NiSkltMkd4WVI0ZURMOUxZSjAzT3NpVlp4QUdNN1NhWWhmaVVpTnh3YlpQSWZiczVPcmlpVjQvN00valBOT09sOHFWOEprbHFKR0ZDRjZGTWtCRVpKZUhUcFpCLy9uMGorUmY3bnlQY2lPT2g2dUdBQmlHM29sakVHaXdkZ0VGSUxFaitJMDNwOEg0VFJJUHlHeElndG4zRm80NlNESDlZTi93ZHpjOWRnUzQyYi94ZlZmU2NBd1N4ZkJ6QUNIc20vT3ZzWVdhdEFURHcvQkVrdDBwQ25iblBhb095bWtRLzVxVDQ1Qi9Sb0RMUFJ4Mm9wVHNYVEpjb1F1bUdIUjQ1b1JBSDZ3ODhlWWRoVWw3eUVTbm90M1kxU0psZ2JXWmpCNUYxb080US9zaHIvNUZvNGRVTjVuOUc2NDE4ako0STZkMFA1QlpIa2U4Z3BjU3pVdkNyZWhrVVFCWDlFbVNqdS9CYmFrT3BONnJRdkQ0Nkl0R210UzUyVHRrZGZmbFNSeVFtajQ0cnEzaTVJMjc1RE93ZDE2RXg2KzlMN21pai81SnM0NzZYeHh2b0xOT2hCWUNmVjRKcHhNRlFxWFVER0E3K0pQdjFiZWVObnJtQjE1bkxpaWlCdUJXRFNPT29jempyMGRRNC9qcUpNZHFpaGtsblZnUnF3UkIyTFNHbnZEY0NiLzVMNVgvN1RTRC9ocEpDQVFrbFY1c1JuSUpXL1BUdW45UHM1ZWVTTDJGNHJNQ0paYVlKY0xVWTdTQS9OQ2pKbGdIMG9FWTVBOEFoQ0dPWFIyaGljZWU2U001QjB0WFNGNTFnblJ1VmorYUdGeDdiYXJVRldITkplVnR0SmhqVkdoaUJFcnlDdzluZmk1V1dCaEFNd1BzR25GaVhLM3crNkRwOTc5OFhMaWhydnh1R1dIQTRCODVyb3Y4RTgvY1RHdjJuYWx3YXJEWUxzalZLM0FsRVFoY1hjM0VtNnhUUWxwMldWQTQ4Q2c5blhxQjFOWXNoYW5TZnJGVjlkcU4zNUxPNFNaeklBZ1RjVm1PWlZlL0s3dHNyU2N3bDg5N0EzeW5Qdi9MZ1NReWp2bWtvZE9YNkdKUW9DNk1WRFNlUFhJYlFldi9zSWI4YXJQL0RteUk0K2tLOHZGZG1ndHFhVTFnaFF4U3ZmRDVoQWFTb3g2RU9nSVpHVWVHc3ZuVUdOTkx0K2RmKzF3T055R2kzNDY2UWY4dEFBRWlGZGVZaUVvT2xkTS94RVBILytrTHMvSUJUV1l5SWdaVHd4VWtGbGdsb0tjUUNZdFlTRzFDVWdvNEExTzJYUjYyUEppR2t2WGlNQ0YxYnp0MEZhZzF6WDBqa0tLdFIyb0pkVlg5SE1IZ2RsU29PQzkxcHlCZTJ3NkZSZWU4a2ljY3RRWm5PcU9KekJJNlVzWUVUemsrUE54enRIM2t6ZDg2WTE4L1dWdnhoNjVEVmk1a25aa1ZPZ2MxR3ZrTnRrQXFoMkNFMm0zM1kyUHhiOFhlZE9vSGZGQWttdFRUWlFjdHVDd0pHQ21tSEswK1lLNHlmSmNOQlA0L2J0Z0I4VHZuUG9VL09rRkwrUGh5OWFGQnFNaWt0c3M1Q1VhUTlYWVRDa3dUNkpVN2RqTS9NV24vZ1ovL3JsTFlBOC9Fcjd5eWVSb29rSmtNRDhhYWljMUpJK21TYkp4VTdvVmdWS0Jpb29OUFpGTVJFcDZMczh5WERGM1hmbkpuYS9EeFRDNDVEOTNQTnFYL05jdmFWMlh3dUlpZUh2aFlSL1VSNis4Q0xPdUFveEY0WVRiSzZsYitob0FLelBXQkcxZE4wV0JGZURXN2JqMkpkL2hDZXVQcC9OZU1wdWNKUVBuSzJRMjEvdTg1ZUh5N1prZkNpYkd4UTJHaXJsRGdvRml6RTd3Zm12dm9mYys5aHo3eUpQUDU5MDJuTHpvUGxKaUt5RElqQlVJNkx4SFprTzBjOWZNYm4zamw5K0tkMS8xQWJPajNDbFl0Z3htZER4Z3pMdlF1aXRrRnJiU2ttc2dOdVNzU2FvejJvL3QxellxTFhXUFlzT0oxaEttTmpxZ2dMV1dFQU5QRHh6YUQvU0pSeDV6QVY3K2tKZmlIdXRQSndCVTZveU5CdzFGZ2R6WUR3b2hIYXhrZ0FGZS92NC9rNys4N0crUkhYTWNYUkZQb0UvZjNNUjNJOFVUNHBtdExLWW04em5sLzNrQ0pZZ0ZMNWpLWVRhTmhHTzRKcXczQjZyTS8rdjJoK09hdVUvalF0aTd5bnI1U2RmUEJzQ0xZZkJLRXV1WEg1WTlZOFVQOWFUUkpkaGZnVjB4M0ZjSjltbUlGVHRQakJoZ0tvdTdUZ0VJUlVTb0RsUDdpV3RmOWkyc25WcWhIZ29MSTZxQU1TRXlRQVUydkhRVGQ4a09vSmcweTN2TDlmeE45OFpEVG42STNPZVlzN0JwNWNaYWlYbnZSYUVLaGJGWkZ1SXBHZ0liTFFPWEJJM3pucmtOaHhEdW50OHJIL3oyQitXZnYvMSsvR2oreDhSRUpwaFlRV3V0Q0VDdkxnWnQ3bWpLcEUzVjdLbVdHbTdSTVd5eWNXcVFTUEJtNDl1TnlRQVJlRmNRYy9QQTNCQWplUmVQMm5RQm5uZi81K1BzSSs4aEFGQTZwOVlZc1RHaGoweE1GYUhSYzNQT281UGxtS3VHZU42L1BFL2VmZFc3TkR2cWVPUExtS1NYa21xREhkcWN3eEtlYkhtOGJETURVcGRnbEJRTW9vZzljUlNtWXdKVnZUVEwrTUU5Vy9peFhSY2xBZld6UU9wbkF5QUFiQTRJeng2NDZqbll2T2F0bXFIaVFETm1GR3d0Z1FVRU5sOFpxSm9sa1FZQllLMkZuOTNQQjQrZHpVKy81TVBHTzZjbU02bFhWQXdUV1Y1KysxWHlyTGM4bmVlY2RwNCs4dVJIMjFNUFB4N3JKbFlGRVlsUW5FTlNJVERHNXBRb2RWUEtaY2g4Uzl4SGJZS0ZUb3crOUdQSWJTWUFaTUVQOEtVZlg0YjNmMzhMUG43VDU3SGdGNEFSSVpZc0UrUjVPT09PRUZXUGNPcWVCRE9pWGJKV3g1c2pFbjBEM1BBcUF4akNpSUhBd29zSFN3L016d0w5QlpncXczbnI3NE1IbmZnQVhuakdvM0hreW8wQUFPY3JVb3drcVpjMkZFRkJ3R0RJbW5aTzhpeVRiZnR1NStQLzhZbTRmUG9IWXRkdHBDOGRXanVsRnBzQWtyUnJId2JVOExXK05YUklTRFRvZTJEZ2dTTkhZRlowQkE0cUs2emdtN056L3MyN1RnUUhPMXNHeFU5OS9ld0FCSUJMTjF0Y3RBWDVrdy83ckgvMGl2TjFuM01ReWVBSjNGcUdzcnkwNHlZellnUUFSYkpPRHJmemRqNzd1Q2ZqN2IvekpwU3VaQ2ZyMUdNZ0tVcmh6SEFhTUFiTGVrdlNHRm02VW1DQ0hyRml4WWlFUUp5cWlBR0ZZaHBLQXdnOXQrTzVJNGpsaHdqVFl4Q0M4b1JLRnB0RUFzRE82VjM0MkRXZnd1VTNYc0hQYmY4eTloYXpJQmFBM0FwR2x3QzlMaUFHWWpKYWE4TzcwcEZWb28xTnFDbWxURVhwd3BkV1Fnd1dCUE96QU1jNWJuTjV3T3I3NGU1SG40YUxUbjhjamwyekNYSDBVamxIWTBSU0Q4RlFaSk1DSXBIdFp6aU0yMEpvclpXUFh2bHBQTzlkeitlT2ZFYXlsV3ZoaGdWZ0plaDl4a1l1d2M1TWRoMGFRcCtoazJxZ1c1b09GSUlRd2x6d3hNQUxsdVdRSTBjaEZZa2x4bU9vbWI1dCsxUHh3K24zL3F5cU4xMy9QUUJlRElPL2dDNWQwOXN3LzR4VlYxZW5UeTJSWGFvY01SYlREdGcyQkd6bytBWUJNU21BRlRHOUR2VFc3ZnpRMDk2RngvM1dJMUU1eDh4bUp1NWtHRWFDTDNRK2s5SlhZaVYwZjdkaWhOR09JVVFCRmFHSUJ1NHN0UHRGNEh5ekxJUDNuaEFSYXd4VWxaRFFlRVpFVkZVbGhKNWpMWTk2UUVJVUlZSUFCd2FIWk52K2JmclpHejRqdCsvWmllOGV1Z0hYSGZxUmxFb1djSUQyQVJ0VHFDQkI0b1lqTFlDcUVuZ3JrRkYyVEM3V095NGRXWVZ6RHI4M0RwdGNoUWVjZUFGT1gzTWNWaTlaVTArcFY0VlNhWTB4RUtOUUZSTTZzZFlGeGtJS0dRNFdwQ3J5TEVQZlZmeXpELzA1WHYrbE53RHJWNGp0VHNCWFJiSlJvMnBOcXgzOG91RDBSRWxOQkJzdlJaUVNGMmtsU01KNUR4UU95RVhNOFJQQnpiSjBabW1leWIvdjMrSSt0UDBpbklNc0poejhWTlRMenc5QUFQRkwzY2laazV2THA2KzcxRS9sRHJOcWtSdkIvb0xZVlFJMjlxekxGSmd3SXFOZDhLYXR2UExQdm9aVGp6d05wWFBvWkJrQWlpcVpEckQyVVZTSk1UQ2txRkJqQSsxd2dnUFNLYktFcXRLSVFXYXorbjZtRitZd05SYk9SZmErRW9pbE5TYnh1VFRCRWpJMTZ4YUZRcWpwRGF1VkJSVmRUMmdGWUdiaElPWUcwL2pobmh2dzR3TzN5UHpDSG5wWGtTcWhtRTRoTmplRVFrWW5sdUdJbFp0dzF1cVRaS296em01dkZGTWpTOXE2R3Q0NWVLRllzUXc1bzdGWm1ub1lZMG5RSk1rZGFrVUE1MHQwOHREaC9xcy8vZ2IvNUFNdjViZjJmMWZNa1J1SkNxTE9JZG9qeVVoRXRCT1NrOVFBMGtjQm16SzdmYlJPczVoV051dUFJUVZPYVk0YkRjbW1wZmV5cG12c04yYTNqcno1NWpObWlKbUlvcDlKOWFicnZ3OUFvQWFoZmRDcU4rRnA2NTdyaDc1Q0pUbEF5bzRCY0VDRnVTVjhKZExKd0JIbEVmMUpmT2RWWDhDcWlXWDB6b2ZEQ28yQmliT3N6YmpxeFUrMU5GNjlDZlZDaWp5clN6UUpBTnNQYk9kbnIvbUsrZmF0Vi9BalAveW9uSHY0Yi9IbGozMkZuTDdoRkFCZzZhdWt1aE5uQm9EUlhFeEI3UEI0N0hwRk9vcW5oNEV3eTdKVXNySEk0R3VOUWU3aTkvYkYwbGVpU2dSblFtQ3RqWlJ4WTZwcUdGU2RXME1JaFJUUDBFTVJBQTcwRCtML2Z1UXYrUS9mZUlkdytSaXp5V1hpcW1GcVJtNmlEMVRmRDlvZ0MrUnlqSGd3eFhrREtJMEpIZEVNZ0JsSERGVlFlTWo2SHV5Nm5yRHZ5ZVdaeWkwRGl3L3VlcWkvY2U0ei8xM1ZtNjZmRDRDQTRHSllYSUlSKzd2clA0c0hycmkzN25HZVZxeDRFTnNMY01ZRG1jS0tGVDg4Z1BPWDNadWYvK3RQaW5NVk01TzNRUWVoeGtSNVEwSWxWZ3dLVlpGWmkzWlhnVU1MaDNqOTdodnhpYXMveTh0ditqNHUzL1ZkOUxVdjZBa3h0ZFJnK2lBN2ZjRWZudnQ4K2VNTFhveGxveE1Bd01KWHlJd1ZRV2hmM0hSZENWZGdJY0lYUjZrckREa1NKckd6bnRTWXU0ZmFPSlBZYWhSSjI0ZUhURHBQTi8yZXVvNGx5bEZxQmhFQVJNTUxJRkVycUhySnNod0NjS0hvNHoyWHZSdC8rY1hYWUh1eG0zYkRSb0VIZlZWSzNWbFZZMXBOMmdicDZJYjBGY0h6RlhncVNBTWFvUEpBTnlPNlZzUVk4bUFKOUQxUWVNanFycGlOUGJKUVlOSTZ6R3VPZjkzK1BQMys5SnVUQVBwNUFQVFRFdEUvNlFwc2tjRmM5eVBiTDZ5V2RhN2dxV1BydUY4OTg4emk4QjV4Y3o4a0tveGFZTUZqOWFhVllidUxpQWRvb1NGZFY4Sko1MVJLcFJXRVJMZlRDZDhSQ25aazI5NnQrTncxWDhaM2J2cTJmdXFXejh2dXdReVFsY0NTU2VDd2NiRjJLVUFZZFo1MjlUcVVyc1RmZlAzditONXZmNEF2dU8vdm1XZWQrd3dzRzVzS25KcXJRQkVSTVpCd2dGNXpWRkJNWm1yK0RnVnJpYisxWWt3R20zeWFWR1hHOUZjQ3NhbXB4T0FIMENnbE5HeU13WWVXb0l0d0VZcng2dFhUbzV0MVlJM2hYREhIOTM5akMvNytNMitRNitkdUF0YXVnSjA2MHZqQk1IQlhrcEpwYWU0c2s1TXRLSW1iVEdFMWdScWlWRUV2aDR6bWdvNGg5aGJBckF0OXZTY3ptaU5HZ0VMQm5pMGxNMTErWWQvYitQM3BOMGZLNWVjQ1h6MjZuL3VLTzZGNzFOZ0QrSXkxbjZ1T0hnUDNxMEZIUkNxU3Q4Nko5UUxkdXB2dmZmWTc4S1NIUFZHcXlzSG1tVkpEbWFFbjJjbnlSY0pveC9RZTNMWm5xOWx5MVVmeGc5dXUwbS9jL2kzeG1RcTZRaXhkQm5TNnlKQ0I2cUdoa1Z1c0YyaDJ2TTF6K0dKQTdObUxJMFlPTjA4LzY0bDgycjJlZ2sycmppQ2kySERlQlN5S3dJaEpoSjVJcXZhcHRlU2QxVyt5endCUUF5c05RNGEzQnRaSU5LckNDQk5oYlpTRnoxSFZaTThpRDNOQUFOaDVjRGZlL1kzMzRWM2ZmcDljZCtnbXhjcEpzU01Ub3FVamxRS2JtbXRFQzZJaHRoT0ZFak5wSkpXd2hzUmRTbkJBNWozeWxUMll5UzZZQzkyT1BuVDNNRVJIUmczc3llTmkxRURoSzZ6S2MzeDAzL2VtUHJqOS9nY3VQbWVBU3k3N2J6a2RkN3grTVFBRUJKdVJZd3ZLN3FrVHp5eC9aLzA3dWFMcnNOOFp5YXlZaFFyWVg4Si81eVo4NWRWZnhYMVBPeHVscTZSbmN3M3RQOEkxMTUrVDIzYmN5cy9jOGxWOCs1cXY0YkxicjhDQjZxQWc4OERFbUdKOHlwZzhIRy9xNFVodFFpM1JqREsxb2QyU0xXSU1UR2JwNTJlQlE0ZWtoekUrN01oejhKZ3pIeTMzUGU3ZU9Ielpobm91bkhkeHBRS0ZZWTJSUlJDTWxFNzRqVEdVQUJHRWM2WkNTekVqUWVBQjRRak0wQXpTeExlMit6Y0JvVEVrNG1MT1YwTjg3K2J2NG4zZmVCLy8vY3JQWWIvdUU2eGFKblowVkxYMFlNVW05Y3RLRXNxQytzU0ExbFpweDN4alpyTVlFUTRjVVNyR0Q1K0NYZG9WTVlMK2JYTW9iMXRnNkxZQTJqTW5JVlpFS3U5MFZTZkR0Nlp2OEcrNjViNHcyQnM2MHYzODRHc0crNHU2TGo0bnd5V1hPWFBPMVBPekp4MytocXFuRldhUkdXdmc1K1p4K0MwOXVlYjEzOEY0YjZ4V0ZGZmQ5Q081OXJZZjRlUGYvQlIrZU9BR3ViWi9MVENSRVYwTFRDMkhkRE5ZYjBCNjhTR0pzMUdNZFoxWDBqV1JTRzJjaGFhd0dncHhvT25rOEhDQy9mdUJ2c2RrWndudWYvZzljT0daajhWWlI1NkZZMVp0YWsrc0tBbW5MdEJBaSs4MlNSY1lpVVUvalI4WUQzS0lQSlFTYWdKMWJGRFRQZlcxWjJFL2J0eDJQZDc3dlEveWE5ZC9COWZ1dnphMFFsbTVVbXluUzFhVnFHOTliODA5SW9RQlU4NGdrTGk5bE5vV1h0K1dWVE1GT2lNZFdYSENLbXBIeEJQbzN6S0xoUnRuZ0c0SXgyV25MQkVkc1dSWk9sblJ5K1g2L3UzWkcyNjdmekZYM0F6KzlJa0dQODMxaXdVZ1VJT3djOTZLUDlNbnIvbS9hcVhDMEZpZFBvaXpEbTdFTzU3OUZyTnp6M1orNUp1ZmxDdHYrcEYrKzZidkNqcjgvOG83OTFqTDdxcU9mOWJ2dC9jKzU5eDc3cDM3bkR0elp6cVZ6akNVVW13Tll0SW0ycWFnYUsycGhFd1ZSUTJhRUJMOWcyaVFCS01FalNTU3FIOG9pYURpQXcwQ0dtTjhWQ0RoVmFDQUxhL09sQXFWVHR1Wnp1dk96SDJlZS9icnQvemo5L3Z0dmUra0dBek1kS2FzNU03Yzg5cDNuMzIrWnoyL2F5MjRvUSs3ZDRrTUIycXM3d1J6cmhMdERzcHA1bDAzRGxRb053UjNYaldzcWZKaFlOdlo3eFZROEZoVkFKc2s0aEJjVWNMYWVSZ1ZESktoM3I3M0J6aTRmSWpYM0h3M2V4YjI4OEtGZ3pKSWVyRFQvSVpqQXAwR3owc2ttdGh1dEN3QVh6dnpkVDEvOGF6ZS85akh6RmUvK1VYOTBzcmpuRng5eXZ2Snd5RXlNWVhCaUtzcXo5aHB1Sld1WmJMUTVLaUpYOE0yd28zZ2syWSt0RlpPV0JzenUzOUdkNzk0dDI1dlZWS2x5TVkzMXRnNHVvTDBFMVJRKzVJaFpzS0txMXlsYzZtMVI3Y3Y2dCtlZWxWMWV2MmhXQVg3emdDeVU3NzdBQVI0dzh0UzN2TndhWDkwL3ZmbHRjdHZyUk5iNkVhVnBxZTNxWjVZVWIyNExTUXBUS2RxSm1mQXFUSFdxSnRQY1ZQaUM5L1dFQ2dlSVowUUVxU3h0Q1J0NEJkY0wybmlnT2FqRDcrMC9BLy9uaU14MVBuZzFwZ0VVcUYydGJKeFVSZ1h5cmhBc2tsZU9udFFGck9oN3A3ZEx6KzQ3K1c4ZVBGNnpkS01aTkJqMEIrU0pST2FDbUtOUVVXb2F1ZnljbXhHMjV2VStaaTZxdVQ0eG1rZU92a29UNTQ1cHF2YlcvclZjLzl0UnVNVlpXSUNNZ3U5YVRHOW5yZk90WXFySzMvT05yUTFkTW00a1hPNG84KzVHMlJvMjg5QmlIODNLNUlhMmZ2U0paM2FNeVhGWnVHMGIyWGxrUlhXanE2SXBDazZBUHZpb1pxSlJOeTRybVZ2WXUyWHh5UGV2WEpQdm5iKzQ1Y0RmSEM1QUFqQzIrNnd2UDJUVmZwVEMrK3NYcmY4WnMxZHlkZ2tGb05zbHVqVG0yaGgvSGpab2c3Wlo0SERRMkhDK1BwamRQRTZWcGRvVmx1cjZ6T3NNZk1mS1ZVK2hkYThpTGpCc2VucWg1YWRyR0RERkErYmhFRFdhRjA1b2RnUXhpTmZEU2dGcWh6cXdwR21ncG5ReFBUSUV1TnBaU3FtckN2TkpSZktMYVVxQlpzcXB1ZkxZaGxLMm9QQmxKaVFXcUVDVjlZUzZzeVJVQjhCMXFieS9HVm9sd0YyS1lkK0c1UmZwK0hCS2lKR3ZkYkxtVm1jNXJwYjlta3RVQmFsbUo3VjA1ODl4ZHJqRjBVeWk4NGthbTZlOHUzU2hYTzZsRnB6Ykd0ay8zcmwxZm1wQ3gvNWJxUmJ2alZRTHFkRSt0Wjl1LzlVN2w3OFZXZWwxRFZKVkJUV0N1SFJzWjkxWXRWUFY2cndRRG8wcWN3bHdyaW1JYlYyejdWTEFKQWRqN1YrWDZUZlI1OVFORkR2Q1NFamdZMnN6MzRWd2hZZGcvVlRUZU5rQnhHUGRoOTU0d2ZRYXR2U3FENFBhTVQ2ZGZjTnhUNVMvQjExWFVYZFJPeVZiNzVuRXZOMnRGclBuMnZRZHRwcXZ6Z0Z3b1VIL0hJWW4xWmZ6N0VpN0gvSlhsMjRmb0hSeGphdWI2aTJLM25tSTAreWZXRmJKVFdpTXhubWxtbi8rbEpyZHFlSlBMYTl3ZCtmK3VuNm0yc2Z1NXpnODIveThrcE1WRmZwWGZOL3FQZnQrWFUzSWM2dE9HRXlFUzRVY0d6TG4wVWFraFNWODVPV0R2VGh1Z0ZVdFc5MkZ1T2ZvRHQ4UUNCUVU2U1RmaEFDUUJ1Tm9IRXBodGN5Qkw2ZnRrQm9Ya2NMS0JlYTB4dkNaa0NLZDcrYS9nMXB2VkhmWHhabjlrRndEVVNiY3dIL1phTUQvc2dYVkkyTlhocWUzdlpYYTNpclhSUGNtbDEvT1JSMHUxYTJDek8zYjViOU4rOVZrMWp5VVU0MjNaUFZKMWIxeE1lZm9zWXB6Z3BMR2ZhV2FTRjNnQ3VaeXhMenlOYXF2di9zdmRYVHF3OWNidkRGUzM2NVJmZ2dodnVva3pzWDMrTHVuWCtIVyt3WnpwVTFBMk1wSFJ6YlVEWWQ5Qk9mQUZXRXNZT1pWRGs0RUNhTmtxdFNxNGxhcGpHNWpTNklKamdBdERIZm5XQWwwQko5NktpZUFtR2sxWlpld2lmZTBybjh2UzErUWxDdFFZTkZVTWFEQktDR2ZGdjBPZEgyZHRObEY4M3RqbDRSM1RFNHFZbXN3ekdjaG1wSCtKWjUwclZvVWNOR3BSUFRmWlp2MnN2MDRsQ0tyVnhkZ3FacElpYy84elFyUjFlRXFSUktWVGt3Z1QwOEtlUTF6bERKWEpMWS8xbzdsN3pyeEUrTXl2TGhLd0crK0s2dWhEUWd0SGZOSGRGWExQeU4zakF4MEpXeVF2QkZoY2UybGJPRjBMY2h5aE1oci8xdzlCZjBZWC9QMC95MzFZUFVtSTRwMWc3enVER0hyWThVTmFKRC9aU3NtSHQwbmZoVU9pWVE4TW1WYVBMWUdjdzJ4NDhSYVRTMFRXS29JYWdpc3VPbFVjTjI2ZnpSRDRWWU52UHZKMVlzR245WFc0WXlYdWRwcWJCVlNIK1E2ZExCQldhV1o4U1ZOYTVXc21HbUc4K3NjK0tCRXpKZUd5c1RLVklvOXZ1SHNLOEg2MDYwTHhXN2trUS90LzZWNUUvKzUrY0tlUFJLZ1ErdUhBQzl2STJFdDFNbFM0TWZzciswOTMzdXRzbkQxZG1xb2pDVzFNQ1RZOUdueC82RFRZSkdxSjB5ZHBBWllYL1BYemdyK01GSWpRTVhUSERIWDFUQ1JDNW96R1h6U1p0b0ptbDZOZ2lISTRLNnRaSkVhOTQ0L0VKVHVHdm1GVXJUMDBIa2VYVmlvT2E4SElweHNtT0thL1JuZldOVUo2RWMrMGVJSlRUQitiVXFXaWhzbDVJT3JDNWNOOHZjOG93ZlVwVlhaSk9wYUY3cnVTK2NrUXVQclNqOTFHZHdCb2k1ZFZyTmRBS2J0V1BlT2kwMGxVOWQrT2ZxNzA3K0FzTFdkMG91K1AvS2xRVWdOSUhKRUJiS1g5djNWL1Z0cy9mVWhTbDFyVGJTczRhdFV2UWIyNTRLMURlaGRPU2pSYlpWNkFNdjZDdkxmYzhtMkhadGlLSXhDcGJvalVVaWdQT2tUTk1tcjd1bUVHTFZMTnpvZ0Z1RER4ZzFVRk1JNllTblFwemMzOVcyaEw4U0FOUnhDWWhmRm1tMWVIUXRBTVQ0SmQvK2hzUlZGS0tDbGs0WitRRkRjd2QyTVhkZzJvbUtWS1dLNlZzVmxJdkh6c3ZLdytmVXFRcURWTWxyWkc5UHpFMitBT0FLVjV1Rk5ESG5TdHhITC81Ui9aK25mc1B6dmI2N1NlWnZSNjQ4QU1IWGpqOUZoWUs5ZCs4N3VXdnV6YnFVb21mTFNoUGp0MGlkSEN0UDVaRFhRaWFnT0VRTXBjTEkrUjZPZlQxWTZpazlVUW9WQ3RmR3hCTE5zZmpJTlhaMkdjUjN2TGx1dGFEckEwTG5JTTNOSnZpSXYzUjhVSWhkNzNRQ200Ny8yR2pUQVA1bXIxdjNwMjFLaituUDBKU080SE9qb3dxRDBabmxvY3pzMlVYYXM1UmxyY2tnRVlQUnRhOWZZT1dMWnltM1NtRTZnekZLaXRpYkpqRkxQZFh0R3NtazFtbWI2Tkd0RGZPcGpUZVZENTU1YjhkV2RIeU1LeVBQRFFDOW1HaVV6TTB6UDYvM3p2MnhlY24wWXIxWlYydzdLeE1pdXVXVTR5TTRsWHUxMWpkUjB3aVYrb245bVhIczdRbkxHUXlNYjVvdXRFM1BSUCtwdGE4UWkvVE5zeFRQeDQ2cEZ1S2RNWmhvejdveHZKMFdPWTBPSkVUZWZJdFRwYUVoZUNDMzZSTm9UMG5WUjkweHVkNTBwaW1VZm5qNTlOeFFweGFIcFAyUTMwbEVwWUxONCt0eTRkaUtGcXM1VEtZSUlwbzc3SFY5TlljblJLM2d4czdKdkZWVFkrdVBubi9JdmYvc3IwRDUxV0NSZG5ZclgwRjVMZ0hvLzM0SVRvRHIrNjlaK3JQcTd0MC9YdlV0WnJXcVFLenJHK0ZDQ1krUFlLV0VWQ0Uxa2F3Q3BTcEZ1SGJ6cVhDZ0J6T0JyT29IWjlQNGM5Q0pPTFhGV0FSZHAzOW5SMURqejVUZzUwVUtZZXQvYW91dUptaHdRUlh1NEwxd1NlOXg1MzYvZVZSQ1MyaThUMUFsR3lRTUZ5ZkpKc0w3S210MTV3dkdUMjdLNW9sTmRWdWxCNTRWMGJIQ25NVWNuc0RPcHFxRnEyc2p0WmxKZXZMRUp0Vm5MdndCOTEvNFhZUVJQM0xsZ28xdkpjODFBTDBjd2ZLUG5oQ2UzTG40bTd4aTEyL0xDNmVHOVlaek9xclExQW9KeUxsQzlQZzJYQ3o5Zk9yRU5JUUFuUHFKcmFXRDZRU1crckNVd21TWU1GcmlINE13Z0Z1bDBVN09Fd2NhUDdDTlVEcytINWNtcm5jR0lhNERuSmg3akVodUFCa2ZhOTNIMWw4TmdSSFJGeFJJakdTN01uclRQWklhNnJOanlwTmJXcHpacHQ0b0JXdGdNdkZQemdVbVJlWGdBTm5URTZrVktWM0YwUHA1YlorNytBM3pMMnR2ek0rc2ZTejRlLzdmNTFpdURnQjZhVXh5TnBVZGRxL2I4M1p6NC9CbnE0VU10MW1Wak5WSTMxZzF3RXFwUEJrMG91SlpISkhlSEJQWlpXaTZtVXRndWE4c1pHR3daakRmM2VXTENxRnFFZjAxZnlhdHp4Y2dKNTNiMHI3WWw4YmlhMExrNnRyVVRqU3pUVlRiYVR1N2RKOXhZcFMrSUltSVZLSjJ2Ulk5bDFPZkhhT2pHaXgrV0h4cWZmZGhyckRMaXIxaG9DeG1mblJnWGp0NnhzbXNUVGkralh4Ky9TK3FmejM5VzhEWjc2U0I2SExJMVFSQUw1MmlkL2J5bVh2S1Y4NytuamswZWF0TEVoaFZsZFpxeUl4ZmRyRmFvay9seXFuY2ovUklqSkEwcnA2RzVTbCtmSndWWlRhQmhWU1l6WlRKMEtnYjlVQ2MvaFQvbHc0Z2Q0QXA0Sy94MVNTbVN0aXg3Q2JtRUdNd0U4MnZnbCtFR0k1aEpLeEVVOThVTkhLd1hncXJKYXhXZmd5YUZlaFo3M3FJK3RISVJZM3NTdFVjR2lMTG1lQlVkVlE3bHhrMVE1dkl1RUsvdFBFVjdsOTlTMzFpL2NNSVhPa1V5N2NqVng4QXZSZytlRVM0NzBNMWtDUS9OdmRHdVhYNnJkeXlhMjhwQmxrckt5bkZhRStNV29HUmd4TmpPRDJHclNxWVV4UFhRbmlKUUl6YmZGTHhHMzEycFRDYitpYjZnZlY3YnB0QWdXYjVzdis5NHg4YTdlZ1FqYWE3RTRCMG4wc0FiRlNjemg5cjI4SFl3VmFGckRsMHJWVHljTHhFZlBTZmh1VXhoWG93cHFMTXA4aUJnY2hDaWlRNEhUdEhZcHhNU2tMaGpCN2RQTWtYTjk3aFBuMytMNEU4ZkttZnMwRGovNUtyRllCZWptRDVKMm9jVE1KU2NXVDNtL1NXbWRlN1E1TkwxS0FiWlVXT3FCRWpQUkd0Z1FzRm5NNlZjNFVuTTFqeEZSVFBBZlVPb3lNRUtNNHZXQWxzR0hyR2MvSW1FNWd3bmhUYU14NnNpUTFnRXEvZG9yL1lEaU9pTWIvcTJ1VXRkUWhReGc3R3FveWNzRkg2S1FObGZJN3pmZFNaQ0lsNHlxY2lmdmNhWHN0T1daR2xEUGIybEtuRTkvWGxycFlNeDFTU1VpcnkrUGE1NUpHTmQ0My80L1M3Z0JVTThKcXJUK3QxNWVvR1lKUk9rREtBL2VYcmwzOVpEMCsvZ2VWc244c0VOclRTc1FOSDRxTmtDN2xUenVmQ21Sd3UxRXBlZTZDbFJyQW1kTGxGQ1NtUUNwLzRybFdwUTZSc05HaFQ4YXlkeFBxc1dSeE9GSmsyYlNRTHprbndNZHZtNzdxeHkzNWdweUZNRU90OEJFNjEyYmNyS0lORW1PL0JVb2JNSi80Y1JrNnB0WmFlRlhaWnEyV0ZlWHo3dUg1NTg5M3BSODU4SUljbnJsWnorMnh5YlFEUWkwL1ovSXdINGdUc0dkKzMveGZkVGYzWG1ybmtWcGxOcVkzQW1xc1pPejhRdVNkK3NjM0lLV3VWY3I0UVZrcGg3SlF5ak1LMWhyQ1V1L1g3SW5VTFdsUHFnUlZ1aDhsSGpwakNhY0xtSmhLT2VXd1JEeHdpMktKdnFONjgxODZ2UFhQT244TXdWWmxQUlhkbk1KVXFXVWhDRitxdzFESzAxbVRXbVBNVjd1em9VWGx3NjYrclQ1NTVMM0MrQTd5cjB0dyttMXhMQUl5eUE0aEFtdDAyK0VsM2NQaDZEazIveWgwYzlqUXhzRm5YV2prVlgxKzFCS3RGcGJCWktadVZzRjdEZWcyYmxSK3VHU05pcGRWNjBxYjdkaEFXOFBuZW9EMGpXdHVQM2NWYWNlUWUwb0k0K29XSjlXWitNb0c1RkdaU0dGcHZnaXRWS1ZDMVR1a1prVjVpYmFYdzlGYk5pZndUOHZYeGU4clByUHdiTU9xWTJtc0dlRkd1UlFCR0VZNWdvbWtHeU1nT3Uzdm1qdWpCL3QzczZkM3VsdnZRQXdwQnhzNjVVZTBiMXl5Q1JackF3Tys1OVhOUU5pc2Z5R3c2UDR3eEpyT2IrcTZBaG9Ic2tXZllSTGNTYU4zQmhCcUIxSGN0a1Jub1crOWJUbGdZQmw4ejdRQzdWQVZYQzZqMHhXcG1qQURtYklXZUxoNHhqMjM5dTN6NHd2c0tpa2ZERmJqbU5ONmxjaTBETUlvSDRoSFlNWnR1S2IzWjNEcjdhbk5vOG5hMzI5eWxTLzFNZDZWSUFWcXFrdGZPcjZaSFVESGV4elBTYUQwblBraXAxUWNDVmUwRGhsamhNTVRaS3NIZ1NveTYvZTVlcTUwQUtPUXBiU2pieFVIZnFDTzJ6eVhZdUJKWGN2VVIvZW5pYStiNCtNSDYxUGdmT0xyK2Fmd29lRkFNOXlIWE12Q2lQQjhBMkJYREhaaElkT2pJQzh3ckYrOTBOMDdlenJUK3NIWG1SYkovUUQyVG9sVW9ZRlFDMndyT1ZTaE93amhxYUpwOFl0VGh3ZFJOdy9nMFhtUkJ4MkdQWGQ1TjJNRWxpbFhCQnJDbDR2ZE1GdzVkS1dGenZNVks5UUJmS3g3U0IwY2ZvTng2QXRpS2Z6YVV6cUl4ZjE3STh3MkFVWVMzSVh3Q3d5ZjhFSXJPWTVNWkxNc2RNNi9LWjVJYjVlRGtDNU5lY3B0YU05RDVOR0UreFZqakI0TldFVDgrc0tWMm5sZW5JV2tjR1EyQ3FBbnBhU3ZOZmtFaGFNREFwaGJuVXorNmxpTVhxMEp5OXpnWDh5KzRFOFZGczFyZVgzOSsvUmp3VE9kZGVORHRScDhQMnU3WjVQa0t3RXZGY0FSaDltV0dQMys0ZkphUGNSR3d5ZDJMTjVEWHI5QzVaSW9EZ3hsU2U2TTRibERSSHBsSk5NT3FzUVlqaWZRa2krT0twRlM4T2FXaXFDc1pBNlZ6T3RZUzBYV0RlWXE4Zm9KbjhyT3lXcTJyczUrdEh6ano2UFd3K2lTTW03UHdSUmZoVG16UWRERWtldDdLOXdvQXUrSnpJbmRnZU5ITGhIYy9YR0VhWTNtcEdHQUtzRU13bTVEMjk4eGs5ZmNsODVLeHB4Yk5zQmFiUzEyUGF5ZnF6bGVuaXRYZWhUTFA4M3dNRk9GbmswdUJGSy84NzVCd0N1RWlqZzgxQ1ovdkdmbGZCdThySTdFOC9LRUFBQUFBU1VWT1JLNUNZSUk9IiBhbHQ9IldoYXRzQXBwIiBjbGFzcz0ib3B0LWljb24taW1nIj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9Im9wdC10ZXh0Ij48ZGl2IGNsYXNzPSJvcHQtdGl0bGUiPldoYXRzQXBwPC9kaXY+PGRpdiBjbGFzcz0ib3B0LWhhbmRsZSI+KzM3MiA1ODcgMzU0NTY8L2Rpdj48L2Rpdj4KICAgIDxzcGFuIGNsYXNzPSJvcHQtYXJyb3ciPuKGkjwvc3Bhbj4KICA8L2E+CiAgPGEgaHJlZj0iaHR0cHM6Ly93d3cuZmFjZWJvb2suY29tL3NoYXJlLzFFTFA2S0M2clYvP21pYmV4dGlkPXd3WElmciIgdGFyZ2V0PSJfYmxhbmsiIGNsYXNzPSJvcHQiPgogICAgPGRpdiBjbGFzcz0ib3B0LWljb24iPjxpbWcgc3JjPSJkYXRhOmltYWdlL3BuZztiYXNlNjQsaVZCT1J3MEtHZ29BQUFBTlNVaEVVZ0FBQUtBQUFBQ2dDQVlBQUFDTHoyY3RBQUIyb0VsRVFWUjRuTzI5YWFCbDExRWUrdFhhKzV4emg3NDlEMnJOdHVSSk1qWllKZ1lUV3pJMkpBd0dZOXlDZ0JtU0FJOGtFRUo0anp3Z2NhdEpJSUVRRXZMQzhFZ2VTY0Rnb0RhRThRVmlYckJNd0FHUDJGanlwTllzdFhydXZzTTVaKys5Nm5zL3FtcnRmZHFTTEJ2SmxzRGJ2dXA3ejdEM1dyVnFWWDMxVmEyMWdNOWNuN2srYzMzbStzejFtZXN6MTJldXoxeWZ1VDV6ZmViNnpQV1o2elBYWjY3UFhKK0NTejdkRFhqNlhCVGdGZ0Z1RitDRUFEY0J1SjQ0OUNnZlArci9IanJxdjE5SGUrRVdBc0ludTdWUGwrc3pDcmh3VVlDYkUyNjhUckQvZXVLNlE4UXRJQ1FSZUFKMWhoVGNBc0h0UndVblBpQzQ3WGFhZ2g3UkorNGhUNC9yTDdrQ0hrNjRFUWszM2FLbWFJOXVtZlpkZDJoYlUxMnpmL1BzdzFkUXV6VmdGUmlOUjlVNEk4dUVGWFBLMVFSSXl3bmM2akRkMGtvU2Eyd1F6VmJHYVBUZ2FGS2QzM244N25QM25ML3QzQ00rUkFSNGd5YTg5WlprU25rMFAxazlmNnBjZjhrVWNLQndQMWdwdUdod0RoNjhjZStaK3ZJWDZkTGFidGwxN1pXU0pqZDFhWHlKcHJWYXF1VlZxVWE3S05VdTFpc0FLbE1ZTnZZN0NFZ0NVb0tRSUFGQmhpQURYUXVpbTZWYXR0ak5McURadkpBNE8xdFgrQk9ldStlOVBIZi9hVXo0b2ZtSC92UGRpKzBWNExBbUhMa2xBZEMvaUJieUw3NENIanBVQVllQU4zOU5CbnNESndEcWE3L2hjL0x5cFMrc2R6enpKZ0F2MHJVcmRuTzg3VEp5QkZSanlHZ0NRb0N1QTZoQUVraHVNM0pXRlFHcXBGQUZTUkVrUllJQUN2dFNCUkVWaUFDU0NNMGpRUkltQWRJWW9na1FnSGtLNldhUWJub0swd2RPMXUzbTNadysvQnY1N0VmdjZPNSsweDhDNlBwR0MvQzZYNjV3OUNqK29sakh2NkFLNkpidWJUL1lEWlZ1K2ZLdi85eDI1NTRiMHVxVnI4ampIWitOMVN1dTVHVGZFc2Nya0NwQjJBSElMYldsdGptaDdRVHpER1FWWkJDQ0JMZ1JFZ29nQ2xXRGh5SUVLRWdKUUlLOW1HaFdrZ0pKQ21pQ1VDRVUxRFV4SGpGVkZVV1FpTHF1cEFiWmdlZ2cwNGRWTm80ZnkxdnI3OFg1WTIrcHVvMTN0UGY4M0o4aUdpQUN2UHdOTlc1N2Vsdkd2MGdLS01DaGhNTzNFa2RFQWVCR29QN2o1MzNYUzl2dDEzNDVKM3UrZ2l2N25vbWxIU1BJQkJoTmtOaUFtbHZ0V3NHOEU4eXlvR1dDQ2pHdXNWUW4yYjU5eEV2MkxzbWVYU09zck5UY3ZuT0UxZVdhUzFXU3FrcW9rd2dCVmFXb0NsSUZ0QzNSS0RCdE1xYU5jbk9ydy9xRkRubmV5blRlOHZTWnVadzkyOGkwemRTY0FZV2dUc1J5eWhnbEpCRWlqU3BVbzZSdEE4eTNnRzVqTHUzWkI5TFcyZCtvejl6eDIvTzcvdDFiRWRieE1CT08zQ3hQUjZ2NGRGZEFBUW5jZEV1RnQvM1RMakRkeXRVM3Z6QWZlUEZyOC9KbFg4TnRCNSt0cXdlRVdTQ2NVVVE2elZrd2F3UlRUY2hKMGtpNFk4Y0V6emk0SXBmdlcrYmVBeXM0dUdlQ2Jhc1ZScE5LNnFvR0ZPaXlvczFFemtCVzg3NEtrS1NRb0VETUNST2FnQVFRQkZnbFFFV1lCQUlRbW9INXJNUFdQTXZ4TXcwZlBqNlZFeWMyY2Z6c0Z0Y3Z6TUdjZ1FURjhnZ1kxMHhTa2VDSUdBR3NJYlBqU05OakgwcnJEeDFORDczajUrWVAvK1pkQU16UzNueHp3dEdqaWljMGJIL3lycWV6QWdwdVBGemh0aU1kQU53QWpONS8xZGU5T2wzMjRtL29kajM3aS9QcTFTdW1Bek5OWUp1YnRzWldKMmhWUnFNYUIvYXQ0cmxYYlpjckxsL2w1ZnVYc1gzN1dPb0tCSVh6dHBPdE9hVHRGRTFMekR0QWxSUVJpeS9vaWtSU2toQ0VLRWxSQ0VWQVFraFNTYWdJSUNReVJhelZFTmlkSnVNa2RWMWhNa29nd0tacGNXR2p3ME1QYitIaEJ5N2c3aE5UbkR2ZkNEUURZeEtqcEtsS1N0WVZScE1FdGtqcjk2MWovZGl2MVBlOTlaZm1ELzNtVzB3eWdSVnZmc29yNHROUkFSY1VEN3V2M1Y1ZittVi9BMnZQL25idWZkNW42L0pCQ0RzVnRFMXVtaEcyY2tKTHJxNVc4c3dydDhzTG5yMkR6N3A2cCt6WU5pWUVNcHNyTnVjWjA3bWk2MGdGSVNtSkVrZ2dXZ1ZiVFFJU2RHekhUQ2dKS2lsQ3FCSVFFRXhpY0JBSXBWUUN5VjlnK1M5RUlDQUpwUUJVQ0lCS3lGR2RzRFNwSUhXQ3RDcG56ODF3NTcwWGNOZjlaM0g2MUpTNVVXQ2NnTlZSbHFycVNGbVNORUk5dnkvejdCMXYxUWYrOUdmMXZsLzhEUUF6U0FKZTkxK2Uwb3I0OUZMQVE0Y3EzSHFyUW9UWEFwTzdQLy9IdnpPdlhmTVB1ZTN5ZzVCbFNNbzVvVUhlN0FUbnN0U1RDdGRkc3gyZmMvMTJmTmF6MWpoYW1zaHNuckUrVmRtYUVXMm5VQWdvQkJVa2tpZ1ZBSmtKNUV6SkZHWlhMRldBekRBclNKajZBWUFwR2dpcEJLQllXS0lxVUNVdDhTRVFDNGpFb3hiL3NUc2tFWkJLSlFHbHFNY3p5K1BFOGFRV1NjRDZlcVAzM1hzQjk5eDFRVTZkM2JCdkxnbWtUaG1URldFMXFyQjFEdW5jZlI4WXJiL3p4K2J2LzdIL0FtQm1nZEFiMGxNeFdIbWFLT0RoaEVPM0NJNUtCakNldlBENy9tWjM4Q1YvUjdjOTQ0V1FCR0hYUWpUcDVpemhBbVQzbmhXODdFVjc4Ymt2M01tOWU4YmMyc3B5K2tMSHJibENGVW1ScUJSa3BXUWxsTXFPZ0txSXVuSXBCVXBLcG9BVVpCSUV3UXlxeTQxdTR1aktBeUpWS1lGUWdrbnNPd0JJV3FRaEFNeVBteHA2dEV5em1nSWxTWWhBUkpLcnJsbmNxaEtNUjhCa1VoTUMyVnh2NWU2UG51V2RkNTJUMmF3aDFoTFNhSnlaUmtSS28wcTNnQXNmK3RQNi9uZisyT3pELys2L0FPaHc2TmFubkRWODZpdmdqWWZyY0xmTDE3NysxZDFWZiswTnV1K3pYcHl4akpTbm5hQkplVXJnWE1iVmw2MmtWNzNzQUQvN3VsMlFsSEQ2UWljYld5MWJCYUVRczJyS05pZHBWU1RUWEdsV01JT2lHY2dremRLSkVFUm1ncEpRdHgyRzdTQ3VweUJOWndpb0NGSmxjWkVBZGk5WE5VS05wd2JNY2lhU1NDSlVpUVNNS2FsWlNZcXJvU3VuMldBQnFJUUlPS21yTkJsWG1EZUsrKzQ5eDJQM244UEcrVGt3RWNpSUhRakZlRzFTNlFiazRYZjhpWDcwTjkrUVQvN2U3MTRzMDAvMzlSUld3TU1KaDI4QmpvaE9Mbi8xdGZteWwvMFFML3Y4bS9QeTVaRG1iSk5FcTd5aENYUGxzNjlleFpkL3dVRzU3am5iWlQ0bkhqNDc1MVpyU0Y4VmFMcU10aU02RmVTc2FKbWt6V0xLbHdrRm1HazBDcFZVZ2VSTTBJd1BWSktRTUt4blFhOG8xYlRPOUE4VUVncEpJaElPVmhrR0xYU1NJa0tqQmFrUUpJUlRGcVgveCtNVUNDRWlVQ0lsSUlVYkIxQkJxWUF3QytvYW1Fd3FNaEhISDlyQW5SODlLeHZudGhSTG9GUUFNYUpNSm5YYXVoODQ4WjZmWG5yUEcyN1pCRTRZZFlOUGUySEVVMUVCQlRkOFc0MTMvV3dMb0twZjlNUGZxZnRlOVAyNjY1cDkwS2FyTUpYY29NSkRNenpqNm0wNDlPV1g0N09ldFFPYlc1bkh6elNZWlJFRjBXV2c2WlJOaHVSTzJaRnNzMGluUU50Qk1nV3FSS2RFQnFuK3QwSUVPVk9SbU1GRUJaUUpGRk5BTjJsUXMzN0djd3NabG8wVWlKanBnMXMvSWNBRU1OTTBpQkZSQ3dUdXd4Tmc2ZzRMcFEwcFVqeW1TV0tmVFlnb0dramlFYmtTZFVVdUxWVUFranh3MzNrYys5Qkp6cnRNTEZjQ3FHSzhUV1M4a25EaWp4K3NIM2piRDdUSC9zTi9NbkhmV2dFM2Y5cjR3NmVZQWxKd0dJSWpvdU9yditJNTNWVmYrWDl4end1L2lOVUtFbVlkUkN0OXVNWHUxYkY4L1ZkY3lodi95ajVabjJZOGVMTGx0S01vd1hrTGFUcGlucFZ0cDlKMG9DclFVZEYxUUtlUVRoTTZWWFlVVVJwK3k4NlV1T3BRRmFJd3VHYWZvUmt1bWowamFhd2pGWUJFdUd0OENnang3OEE4TENFUWpheU12MjRLYU1FSnhiZ2JpR05DV0VaUGhDSVVKQUhGY2FRVjV5UkpaaFNaUUVsSkNTVWd3cFhsU3JxT3VPdllHVDUwejJuaFVpVXlGckJMSFNiYjYxb2ZCbzYvNDllN2QveGYzd0U4Y0Q5dS9QMGF0NzBpNDlPQURaOUNDbmlvQ2laLzh2eHYrZmIyMGxmOXFPNytuRFcwczY2dXU5UnRac0c1aks5NitRRzgvcXN1aFZZMTducHdqbG1yNkNDWXpSVk5DNW0xR1UwRzVobnNNcVROeWdnbW1telVTYXNpYlU1UUN4ZEFrbG5OQWRJMU1Lc0loRkFJVkMyc3BmdFZBRUkxZk1laVRNZ0NnM1lzY3FXSEY0Ull3QzAwZytqS1pxNWRKS3E5d2dWSGJDMWlzYkpBUkZtWlgwWUNLUUlSdFRZbFVRcUVDWVFrMCsrVUtPT2xNZGJQVG5uc295ZlNiTm9DU3pXaG1SaXR0REpPRXpuMXZ1UFZuZi85SDdYM3ZmSG5QMTJSOGxOREFSMFViOXQyN2I3bStkLzJFL25BaS85R0h1K0RzTzFFdGRMakxhNjVhb25mOVkzUGxPYytZMDJPM1QvRHFjME1oVm04YWFlWXpjbW1wVFFkMEdSaXJrRFhHWStuU21RcU9rMmlKTG9zNkZTZzZOMXFkbXhIc2I4dEVMSFNBa0pFTlVTbExOU2RrVEY5dWxrTXZubnc2K1FmeldjNkhIUURTeUhFbUVaNmJFeC9vTHRuZjY0OVZXRVo1b2gxQmdOSElvbmRSR2lsaXluWjc4ekVlRndoMXlJbjd6MkhoKzg3UnlZQUV3R1FHb3hYSjJuclBsVEgzL0ZEN2Z1Ty9CQ0E2YWZhSlgvNkZkRE1mMWZ2KytLWDhyTys0WTM1NEV1ZmdXN1dWWmluZklFSjV6Sys1YXN1d2V0ZmV5bk9YS0RlZGJ5UldVdHBPc1cwSldZdE1XMGM4MlZpbGdWZFp5bXpOZ090bXBLcHVWeGtRTHBXcUdKUmNLWUF6RlRhOENrOHBQVm9ReWtnVW5DRmNGTkZHRlVqR1lKRVUwcGhrYWZaTmRORUFoYUpxRktTcWFkNDNReEVyWWhCQUNSejgrWmZYYTlUZ2tUOUE2Z3FJaUtnRWVXaGRBWU1KR3ltMEtaUkJVQWhNaEppYWFubSt2b2M5MzM0SkpxMmhheU1TR1lpTFNFdGpTbzg4TWR2Vy9wZi8vcHJ0L0RCaDV5dStaUW80YWRUQVFXSGFYanZHYS85YWo3MzBCdmIzUzlad3V4Q1c0MVI1UWRuY3RuYVJQN0pkMStERnp4bkRYY2NhM2h1TTJQYUtqYm5XV1lOTUdzVXMwN1J0RUNiQlcxV3pMT2dJOUIyTUp5WHphaGtRRHNGVkNWMUJEc2FCZ3VMWk5ZUUhsVFl2eFlpR2czRFFHb1VjNUVrVlFSVTBCV0NBdk9NNWw4aFZKU1l4TWdVMDJCeExqQTdWaXhCTDVXOU8wYmhCd0VSaWNaWlF0bmduaWRna2l1NDNRTXFvZ2tFUlVpb3BDb0JnS0t1S3hLUWt3K2U1Ym5UbTBSZEFVa0ZkZDJreWRvRUo5OTM3K2oyWC9xVytZbmZmTXVuQ2hkK3FoU3dwL3pqNzBOTU9DcDU5TUsvZXdUWGZNVWIycFZyczdRYkJLVG0vUmY0Rlg5MVAvN3UzN3hHbUFWM1A5eGluak12YkNrMlpvcFprMlhXQWRPR0NKZmJkWUpHZ1U2QlZoVk5Cbk1XNmFpTTZEWXJxQkNhRXJxcnBWaFdGNEpzeGdTRW96RXgzK1paanZpL0VFNDhpMEN6Rm1kcndXa3hhQ1VlUWR5VUlxV2FpbDQyNkIrRUtBSkFKbjh6YkMwaWI4Zkk4c1hzNkl1NHhhZU9pRkpjK3lGSkJFUktDVlVpNkNuRGVpUllQNytGMHcrZU40OC9rZ3hxeHRMdWNUci93VnpmODEvLzd2d2piL3habzJxRWVCS1Y4Tk5oQVl2bG03endCMytpZTk2ci9uNldQVjNLRjBSblZhclB6L0Mvdi81eWZOa3JyOEJISG1obFk1clJFYml3MVhKOXBwZzJLdE5HTVd0UmxLL0podXRhdDM2ZEt0b01ab3BrV3EyVWtjb2l5b1NzNW5xVmlNQVJsb0FMK0s4ZW5FUWtHekVvNlFLelVXRlJCZ2FWTEdFZ0pYU21LSWRybkdNOUFGWTNxTzZMZ2NvU2VnRFU4b05HQkJycWN5OGZDQUFFVXVGNkNFbkdBL25mQ21xeTdJcmRSZnh6VWJvekdndG04eXhuSGppSHJtc1Vvd3JJTFdWNUIxSStWYVdQL0xjZmFlLzQ4Zi9UYkhqcDhwT2dESi9hU3h4YU03M2t5TS9vRlYvNXY3RWV6eXRNUjNtZHNudEwrVVBmK3l4NTdyVTc4ZjQ3dDlCbXlEUUQwM25HNWp4amM2YWNkY1NzcGN4YUdNYkx4RHliSXVVTXRKWmlRNnVFS3RoWnhBcDRRWXNpU2F2MDlKb1kwV2JFTGtCeDYyWm1SK0ZoSmhJb1Jsd251cE4yUXBtOTZzSEhGcjJlaEMyTmVNU3dXWEpqNWdHMFdkaEVWTXhNOFRVNnBnTWdJaEljdUJIU2xyVVR6eUFEUkMwQkhBQXZFZ3RHMjFod2M5bEtVQ29oU0dXVlJCUWlaMCtzbzl1YUswWkkwRXhNZG5SSnVsSDZ5TkZmNkc3L2tXOXlnL0drV01KUG9RTGE3THdCVXIvMzgvN2x6K1dydnV6MTBMYXQ2cTdPcHpPdTNsN2hSNy8zT2s0bUk3bjcrSnl0U3RxY1pVd2JHdTZiWlc2MXhLeWgwU3lkUmJ4enRZSmxWU0lyMlZJa1U5QmxCU25zaXZGUlpBWElKQjNOVXBidWUySk5CNjR6Rk1SZmNsVmcvN3VROXYrQWQzMXFnOUZkaHJ0MXRheUlWTkVjb3dDdENycnMzMmRtcFliN0xaQUFhN0Y0dUlLSVFtbjRnVUF1bkRYRVRWUVNBZ2tpbHBQMmRETU5NamgzNlN5bWlDdXlab1ZJa2l6ZzlQd20ydldwb0JaQVdpSnRhNldlaktzUC9UKy8yTjN4RTYvSElWWTRLazk0SHJsK0ltLzJxTmZodzhsWG5kWHZmdW0vZVNNdmU5WE5hT2R0dllUVVBkVGdzeTRmNDRlLzUzcXV6d1gzUFRDRlFyQysyWEt6czNLcHJRYVl0WXBwQjVsM1FOT1I4dzVvMUN5Z3VWbkRmQ29Ja2hsWmpTb0RLWlFFSlFrcU0ydkQ5QkxGQXZEQ3ZoNWZtVmVVS0Vwd1hVMlJxeTA0VDN1eloyT3RodDlTRXFTUjhTRk5CdGdTdXE2Q3JRN1lVcUJWcEpTNG5ESnFGV0ZVYUdrR0V0RXFaVXVUNlc0Q1VFTXdTc0JTUWxvU1ZLTUtvOG9VVHp0RjE1bXVwUklpaGRjTnIrOVBBQ3c2cHZVSHpsOVAxbGFFU09pMnBrQWFFM2s2b3JDVjUzN2QxeTlYc2pVOUt0LzJaQ2pocDBBQkR5ZmNmb3ZjSUZLOTkvTi8raGZ5bFY5NE0yWmJYVFhSVVhkL295KzRjb0ovOXI4L0h5Zk9aUncvM3dnSlRKdWNwbzF5MWdxMkdtTFdBdE1HTW0zSVJpRnRSOHpWY3J1dEFwMFNVRWhISWlQQlNHVUcwUngwaUZYdkNaQ3BUakFqTUJyQ1NCaU5iTlpQRmRCa2hFZDRXcUdsdzlRSE4xYUltRDlXcGdvaUk4RjgzaEYzdDhDRmpGMnJTZXBsd2ZNUEpuN3U4MWRrZFFYY2Uwa3R6end3a21jZHFER3BCVlVGTUlHdFcvWUxjK0xCRGVMNHlZNG5qemZTYlhVNGV6N2pQZmZQK2RHemtPYk1GS2ZQWnlvbzZkSXg2dVZFWkJVbkl5T0tJWkpuKzl3Q3dyTTloWGMwclJjb1VDMlBrYW5nZkM2UUJEUmJkYXRMalY3enRkKzZWRzNMczZQeWQ1N29pcG9uVndFSndaZThaSVNqTXYremwvN1RmOFVyYi9vYXpHWnRWZWM2bitodzNjRUpEbi9QOC9XK2s2MmMyOGd5N1JUenVXTGFLWnBPWk40UzA4NTR2dWxjT2VzVXJRb2JoV2cybk5mQk1CL1Y4cmtOM0MwU1VKQVo0Z01BcUhFVHlPcS9leE9GQkJYS1pMaWVDaUI1K1pXRzNhQXJxQWc5WVVjb0ZJcWsxRlNMYUNMYVV4M3djTXRMdGdsZS9jSXh2dUN6dDhrclAzZUZ1M2JXV0YwMXR1OWpwTlJEb1FWSWRNTUNJZE5mRzNPaTJlcjRKeCtlNHQxL3VvbGYrY01MZU84RlJiVXRNWGRrRWlUakJnbmtnbE9CVE1BTHk4VEVZLzJOL3hDU1JqVnl6b0ttc2JSS25vNnpyRFI2OVd1L2ZVUm9lL1RtditlVzhBbmhDWjljREhqRHQ0M3dycDl0bDU3LzkvL1A5dnB2L09lNUd6V0o4NUZlVUhubXpocTNmTTluNmVsenJaemI3TkJrY25PZXBla2dzMDdadEVsbXJibGVXK0JqaFFXTmdwM1JMWkl6MFpueEF4V21qSVZHTWNGblEwUVFRTElhdHN2QjdUSnFROTBDcXNSZmd5eUVyZTd0YS9pSXBMNEVCS1JVQ3FtQTdrUkxPVDFQTjE2NXhHLzYwalc4NXBYYlpPZTJDaGdvRDBscHN5V1JSWUNxU2xJR3dDeVNNeVhKNlVaNm1RMUxWRk5WeVJiZURjYndUYjkvbGwvM3d3K2d1bnBaZEtacWR0N0M4RENKMWdDbEVWSHd1Sm9BMVpKM0Z1UzQvMWJvYkc1a3FwQ1FES1RsTG9tTTVNTS84OC96aC8vajkwY0M0YytySWsrZUJUeDBxTUxSbjIyWHJubmROK3F6di9LZlU1Ym1JcHNqM2FEc1h5Sys2MXVmaFlmT3REaHpvVUdyeEt4VnpGdHczbEdhckpnMWlubVhNTStBRlJkQTJpeG05ZFJ3WDZZWS9vT1J5QW9nVzRCb3I0V3BFaThtTUdJUEtxYWd2Vzc0dW80U04zcDFsTVNZbE9RRlRCZk1qNldsaFBaY0J6d3dsUzk1VnNWLzlCMlg4TWJQMndhNHFyYXRjU2QxTXFlZFVtSmRRVkFMUlNtU1NsaGo5MDJoT1NxVkpxQ0tndXNJWjcwNElpc2xKV2xhc0VyRW1RdmVlV3FRNVo2UWxrQzVYaHdCenlTcnhmK00wakREdzRXdWxVUVpqNFRaSkFwSlFMYzEwbXBiTjNyMjEzNWYxWjA1MGR6MmluL3pSTlFWUGprS2VPUGhHbS8rd1c2eS8xdGYxVDMzTmYrK1c3NnlTL056RlRZcFM3T01iLytXNnpHZEp6NXdha3N5QklaN0ZOT09hTE9nNlNqekREUlpqZXRUUWRNUmphclY5QkZpNUxKNWxleFl6akNmODN0UTBHaFovNENWbWxncFhrU3pDSjdDS0JHeEFvVWhtd0YyUkVyOVlDR1RDU0lUa2ZhRFczakJkdkNIZjJBL3YreGxhd0lBdWFOa0NPdEtaRlFKTkxucUlJbXlGMUhvRkdBbWlnQ29hblg0VEtWVWNHQ0s2ZFBEc25FQ1ZoV2xyZ1MxWm9GMmdFeThJc2FxV1RIUUtYVnFobGF5UXgxNDlTZ3JzNm95RDdpU0FKTVIwTUptdHdqUWJsVHQ2cjV1OUt6WC8vaHkzbmJQOUxZai85VU16U2UvSFBSSlVNRERDVy83d1c0SHVXdjlzMjc4R2QzKzNER21wMXNrcWVWc2c5ZC8wL09RNmhIdWZtZ0xMWW1teStneU9WZEIweEt0S3VlWlp1MXlRcE9CVm9rbUt6b2xYUUdwcm9SWitreEZsbkM5OXJlS0ZaZENuVlFqUmZzMG01ZmpTZEFUemdaSytGZWp6b3dMUmdrbHgxWUJrOSs5aVcvL3doWDh5RC9ZaSsyck5ick9xbEZUN2FTdk9YTVJ0ekJPMEpnUlZZMWlaL2lkelNVakNSUkVVbVlrVlA1NktKSUNnSUNKQy9VSVRLcTJjazdqR2ZSeUI2TWRlK0liQURPakxNZTc2WGVOQ2tmdnIyZDJVRmVDVmdqdGdLb1NUTTlLdC9Rc3lqVmY4blBqTSsvNlFQUG1veDhHRG4vU1ZUVHA0My9rRTdvRU55SmRSUzV0dmZTZi9Db1BmdTQxTWp2YnBsU045TlJNdi9UVlYyUHYzaVhjK2NBR3BpMWxzMUZ1elJXYkRiQTFwMnkxTk1xbEFRUC96YlB4Zm0xR2xOTERlRHhCU3hnR0pORFJKcjlaUnFJenFxeFB2d2trdzdsaUs5dVRJV2RuRUtnM0F0cW53c0s4SWxVUXRvckpuMjdKRy8vQlh2bnBIN2dFMjVjcmFUdEZYU2MxOWk2NER4T0hWUm1ZRWVtTlRnclUxV3RSUk9BSlVDU0thcnp0czhjcSt0TUEwOFU5MUJFY0dIWVQ4T29IajBCQ1gvdEVDWUl0dHhTUVRSY1hnazA5R0ZaQlJhUUt0cjBJaVpRU3A2Zlo3bmpCVG4zTzMveFZjRzBQRGwzdnZmekVyeWRXQVgyNTVQRWIvdTRiZU9XcmI4SjBjeVlWS3owM3hZdWZ2MHV1ZmNZYTdueGdFL09PV0o5bGJzMVZ0aHFScllhWU5jcFpCNWwxa0hrV21YWEFMTHZyelVTcklxMFZFYUNsSWx0K0Y1bEU1MW1Oam9JczludEdGSnNLRlRCcUJrZ2E5RU5ZU3NBeEh0emM5TENzYUI4QnFjd2xMMzEwamwvL29RUDQrbGR2NTNUYW9hTm9WWU5aTlNHcFJUQytQRmpqenFRZzlUR1AyVlh6ck5FR3g1dFIzcHJpN1NMYm9vcDk2L3d6RW1RNVFxVnBPUTkzMk5IRitENzd2ZzhVMDFOL0FPQVl4dXNna20zM1VGV0NWQW1RZ0ZRTHArc05MNzNwK3NtenYrWmZXZVhNNFU5S2w1NUFGM3lvd20xSHV1cmFiM2xkYzlsci93OXl0VWw2WWFTYkhmYXVqZm1pejk3UGV4L1lGR1ZDMTlrYW5WYVZuWkp0QmxvQ1RTYTdMTklSZlY2WFRwdlFhUkVIY2hwY1hwOUlHMHhvbUE4MGIyb2tuZ0VzcStNa2UvM3FuVEVENlNQc1kvWjBXTExBc3ZyQU9uNzlSeTdERjc5a2hmTldaWG01SnNta0NsWVJtZ3FLNGlTdkZJalNLUThIb3BDaEQ0QmhqVTRZZ0QyL3lON2llVjBza2hmUlJJUWxFSUhHQ3J1Q0xsM1pJTDBlTTRvamZOWmwyRnFCbUdnRFZ3MUhpV1VOSUR5YWN4bmwrVWc3NmZJenZ1cWI2cTU3ZTNmc3lQLzl5WlJ4UFZFV01JRzM2dXFCRit6SFZhLzZLYTVjWFZmTmVnVVZxVnJpSlM4NXlEUHJjNW42SXZDdE5zUFNiSm5UeGtxc3R1YVVhUVBNTzNMV0tlYWR1ZDdPQ3c1YXorOW1WV1lsc3dvNmRiSlk2U3ZYR0p3V2tHa0VLeFgwOG1XU0RCZGRtQWVUclFXUDlJb1JqWUV6RzFMVmd2U0JMYno1SHgvQUY3OWtoVTBIVEtvVVZvaFZTbDdLUlNsQ3RjQkc3RFVkcktvRGs0dTlETGRibzBqN2t4NFhXSGtmQlpHUjltdW9ZckNsSnVWT0loYlBGK0t5NEl4WXlJS2lXZlRLeDhqYWhSSUt2QmdpN3Vvb05Wa0k1QlVjd21hV3VxVkxWQzU3K2IvZVB2bWNhL0htci9tRUxlRVRvWUNDUTRjRUl0SmMvdGYrZ3g3NG5IMHlPOW1pRnRHTkZzOTUzbDVNSmxVNmY2RmxvOFM4elppMWlsbEh3M2tkWk40Uzg0NDY3MndiaktZVE5DcGViQUJ6dHlRemszU2FrQUZrZUFFQ0FaV0VLQXlnNWQ3NjZoYU5KSUQ0d2tZVUF4RHUxWEwxdE1MaUNBa0pnSUswWEtINzhDYit4VGR2eDFlOGFqdm1MV1ZjUFNKL2FrdmVqT2hCZ0RJcmJ2YmNDUUNMejlIbnl0d0RxaFVUTWtZbEpSdnJVT3B3NVY3YTU3TW5tZnlMT1pkUUZudDlvSmNEWis2WXp3em5JTmJ1OFc3OE40SVIrQzFGck9SYWJHa1VVa3FZYldqZTlmemwyWE8rN1A4QldUa2VmTno4OGhPZ2dJY1NqaDdOYTgvNjZtK1FaLzcxVjNNK2JRV284aXhqeCs0SkxydDhHODZjbmFITFRETXZvNXAzUU5PcXpETmwzb0V6NC9xa1VjL3ZLcnlpeGZDZVZiaEFjbkI5RVNnZ0dkL25NcVhELy9CQ3luQ3JrVElieUNhR3U0QXc3NDcwNjNHckVaRWZuT01Weng3anU3OTVENXFHR05YaEIrM2pxcUNxZ2xCeE15cjBiWWpnUmRRcFFWSWFyR2NUaXFURk1VcDluQnRCT1ZEVXhwYkp4ZTR5bEtGeWhib2x1S0dVb3YwYy9ONzNEVDRGRTRMWUdkN0orVVlBZ3VRZUhzbnVMMGtjRThZSEFiRFNwbXZ6WlRlOWZQeU0xLzk5YzhHSEhyZGUvWGtWVUlEcnVCM1lQYi9zbGYraXJTOVhOSnVKa2lCTmxtZGR2UXRiMDViVGhybzVWODZiRE52d1J6SFA1THhSek51TXBsVzBxbUNtcUxISFVGWERId3FJRWptVHFwbHFaay9JRERCYjRXWE80Yk5NSTl5VmxHaTBWSlRhQ25VV1YrU29VZFZYcFB0WWFVZGhSOVVPazFOei91UjM3Mk9Dc0s1dGRFVThxTFVWUW9LVVVLWGswRkdDcTRhSWlOcytLbWxjc0VnNDJJdEVTUUNKdzJDNVZPS3pVSHBtbDN2eXhONEd2WnpiTVVnUCtRSkllR0VNbzBEUmFzMmNkU252STRJd241bVJJSTlwWDhKa2RZdnBQM2xXNTdSVHVmOExmbWpiK1BuUEJXL1Z4K3VLLzN3S2VPUGhDamlpMHhkOHgvZDIrei8vRWs0dnFOUlY0clRCSlpkc3c5SlNKWnViclRTZFl0NHFacjZHbyttVTgxWmw3cnRQdFVxMEdkS3BzbE5sbTVXZFc4R09SS2ZxZW1pS0dadiswSU9USHU2WUVoVThUU1hWYWxsc2o0c3lPRWE0Q01LTm9ZUXdyci9WSklIM05QaWVyOXd1ejd0NlNkb09TRW1pVGlHSkt4U0FXQXNTTHRqdDhpQ1VVS3QyVGh4WU9iazRKeXdXaDBDUnhMb2hpUFZKcmhvUmF4VzdibCtzS2l3YXNwaDRRbFBLOG1IMnE3QlNDYm9XbThGRjdTNzMxWWpLdmN4VklzS21XY1haT3ZPQnY3SThmOTRYL1V1STBGM3h4NzMrSEFwNE9PRzJXL0o0NytjK094LzRndS9VTG1YUlJwakpVVjNKd1FQYnVMSFpzYU40aFllbDB3em5pVmpFQzdZRVdvVjBDbTFwbFMxWnJlU3QwMkFFYlA1bFc0bUk0UThMaGpOeDBZc3RMWEhSZXpMNjRvdGlIclM0cDdoaytFYWV0N0s3eXZqdWI5Z05FcWhTM0FsREE5VHpIK29XMFFHVlEwZ1RzdSthcWlBU2t2MnRBMjN5Wm9nSEtLb3BXQnQ3VG9sQ0VjRnljYzBBVUNWRUJXQzRaamRuanRXY2RocjJFbUNLMXBadURYbXBDSWVpYWlPZUdZRzNBVndVdTV4WWFUTnY5Y0NOWHo3ZTk0clh1aXV1SGwxL1hEWWY3d09QZWgwR0FLRSs2elgvakR1ZnY0SjJrNmhTd3J6RC9uM2JBSVhNTzBYYlpXazZvczJVTmlzYlQ3RzFDclJLR1ZnNnlSbmlOWHlpU29sZ2d3d1BZc1VwWnN4NmdVWW9VYnBVZkpnRVd0S0N1ME9jWllZUE5OaCtwQnFMOE1FRzMzampDdmJ1ckpuVmdvTDRqUHF1R1FpWFQxS2gwaTg3Z2xGQWFyeWVBa0JTUUlUcVFRaTgvakRhNGVHeldiN2trOFVZb2I3Y01LR3NwaXR4QitETDQ2cmU5WllZcXhUbEwvNklXQWJJRW9LaGJpWUZHWnBBNlRFZ1lETWgrVHlXRlB0Zjl3cmViaWFkWEVxNS9BdVBBRmpENFZ1TFgzKzA2NU5Ud0VPSEtoeTVoVXVYdnVLbHV1dUcxekhuTElLS3JXSXlTYkpqeDVqVGVXUDdzR1JobThsV3liWlR0TmwvV3BVdUUxa1Z1WFBYcWlRTjFSc0d6QVkzVkxOdGI2QzJlbHdJVHowNTFndHJRc0tTNTdCS1VpaEJsUUtpREI5S2ovMzgzNFhJbDhna3BWRisvWmZ1TWlFbEh4Q1BWQ094RHdpWVBHUkZzbXlGbFF6YUZocUNpRnFoVnN3cW9vWWdGN1RGNFprcUJObmw0QlNNRXlLaGI2UnR6MUhXcEFCQTIxSnM2VEJqcGIyN1MrTUlpcjh3NUdyVkdnV0hzbTlFa1EyZGVIVXNvemx3c2xmbkR2QmtOQ0puQUpJNG43YmRKVGMrZjN6TjMvNE9IQkhGb2NjT1NDNTZrNCtwcmVXNjdsWUtoUG1xbTQ3b3Rpc0Y3YVlsN0RPd2M5ZXFCUmtkWVpaUHBWVkttNEV1QTYxWDczWktaZ1Z5dGdIUEduU0xNbnZSVUFaTFpVcy9XQnpzSDFEMHBoU01GcjhzQ0lFR2p4RWZESjlITHdNWnVDR2lxa0NleS9pcjE0N3gyYzllUnM1R2ZaVG5ET2dORlZDaUZDZjUvUVNRSk03L0dYRnVDOFcxbUxvb0ZJaDRkQ2gwcDQ4a0s5bDFpcTZGZXdkSzd0UzJCODYyd1VHMlpRandoWHAybFRUeEFDeHc4Rjd3ZTMwQmVOUkNSSWd6VkFOdnN4ak9FL2FmdTFoWENwL1VWbG1XTk8rNzRYdFhWdlpkWXRzRlAzcEFjdEViajJQbDA2RkRGWTZJcmx6ejEyN1NmUzkrQmRxbUExZ3hLMGRMTlVhVG1qT3I2ek1DT1Z0Wldac1JPVjEzdjVhL3pVeStjaTJ3WGdKVlNpQXhTRnNLU3dNNVhLTGg4dmFvTFRCUStkc2pTSEhuVW54WHNQb1lLaXVrRnVCQ2gxZTlhRVhxRktzcU1IVGFvWWswZ0dQKzJUeTBFRkRib3NqK1U3ZzlFUXRCZk4xVHVSSGRQblhaaW5icUtrbGRDZXNxb2E2VDFDT2dyZ1JWSlJqVlNVWlZRbFVKNjBxd05FbW9rbUJ0emJkQ1FESjNHWjZYMHY4OXdCZ3UzQjdWTWtMZ0FTenBrV2NQYjRZbFBIRE9NNUxkZ1hwRmtqUWJtYnRmdUxPOS9EWGZDWUM0OGRFOTdTZWVpcnZ1VmdLQzl1QVhIdGFseXlwc2JYU1FTcEFWeTJzamRsbXQzSWtDcG1SVks5NkgxdXYzZk5oOERUWVpaUnZGTUJBa1ZVcHVQVUJ3NGExWVhoeGMvWHZ4dVZBd1JndEVGZ3RKK3B2SElHUUJvQ3AvN2ZQWDRuM0tnaWtwZzFUc1FRS29IaHdNSXA4ZXVNT216QUx0NXByWmRjWXRwbG9BUUk3ZFA4Zld1VTVPYlNqUGJ3TE5QTnRXSTYyU2hLUktVZFdWSkJJVkUzYnNFUHpwaDJmQThzZ1dLb2w2ZjBxYjNjcWxtTTNXWWxFQms2ZmwyUGZQQ0drcGZVMFVaQzhXaktERlYwMFZzVVBzbmhITDVLNG1WeWk3bnZ2M3NIekZUK0syV3g0QzhJZ1ZNNStZQWg2NnRjSVIwY2x6dnZLbWJzZjFMMmZiZENBcklDTlZDWFV0eUsydDk0ZEF0TXRKSmRGWFFVTFZkaG93RWxsc0s0cWtVQ2FGYmVibnV6VkdRYm1qZjQwb2xzYkNHNDlsczA1VE1XaGd2c2dGcVNVNE5NS000cnNIQ28weU1DSUNiblk0c0ExODVoVmpBQlJaS0QrT0piS21hRkV5cmRiTXdHMFlHR3F4d0NPTWtFbkdpNXdGcWh6VlNSNThjTTZmK1pYVDhxNzNYdUQvT05aaTFrRlF3WEtOUVhua1hHNEFnZFdlSlFHUWdVa2xPTEFFN2FMaTBIUGJMSUp4dkt4d3BRc1FEUGpHQ29odDRrSnBJMFdudlVtQTBFL1hpVW9IbUtVUk43NnFSSWFOMld4RHNlZHpkMHl1ZXUxM3pUOG8vd2lIYmszbEFNZFBXZ0d2TzBRQTFGMHYvaTVkdVRSaHZ0RWhWVFUwbzE0YVI0eFFTdWlzRzBaYlVZU1pZS1lXTTZBVkRDTzdTTndndWcyenhlQUxBREFvaFpoMXNUK0Z5U3pJMFVHV1FNUjVzSktad0xEVWVXQ3NJRmFxbWJjb24zdlpFdmF0MVd3elVWY0xmaWNzc29RK2l4dFZWVVdLRW1jenFFVkJVNEdQZmkvdlhRUElqLzdrZy95M2J6NlBrd0N4cXdhdXFnVXB1YWYyVUZ2RWxLZXNPZktXUkJ6aFc3d1dKQ2VNb29HTDNHdDVMWG9TZlRmbEs3SUUrc04yQ0lodGFkT0hRZ09RV1FoMWQ4WGxjMW15VnF6M1hQMU5hMnRyUDdwKzlHdE85dzNvcjA4Z0NqNmNjRVIwMjFVdmY1N3VmdTZYc05NT2tHUllLN0ZLRmJ0T05SUHNiTVdhZGlxMHNubVJUUFBNVnJGc3V3elFBd3pHenJRTDNrRFJaeXRDRm9VU1lCbnBJWW9YaUJYVGUzaXNSV2crVXVycHBjaVFGR2tEdEcyWk1RTjJyRll4TFBnWXJPMy9LV2dxQ2hCU0d0Q1JwcGhRWWRrVEVLNzM1Z1p3ZmwzbHk3N2pYdnlUVzgvajVOVXJVbDI3aW1wSGJSeFBTM0FPNkp6UU9VUm4vbThMMFRtaFUwSm5DcDBEYkVpMjhNUmJ6Qlh4QUN1NkYrTXUxcXVGVUlyd3dJUjl2WmgzMWFwMmZCQmtDQ2tXbEdnSURFdEhwUkowczFhM1hYK2cyL3ZsZndPZ0p5NFdyOGV2Z001c3p3Kzg0dS9veXRVVGRGdm0yRlNSS3FQVlk5UHZUbFZ5MXBTcFVLdUtJMG14YlZUQ0ZFa3AvdlM1SEVSdWVDc3U5ZzFSTG1wVG1VQ2tORjNRYm5hZHZJaFp2ekRuUXVidTFtS1F3aXFJQ0xyTWZWZU1ZZ1FvZ28vQkxTWmpNVjlxQnlWQm9SSzVlek4zRkNlL2ZXVGhxL0VFSFlEWC9NTjc4WHNmVkl5djN5YlNLdkxVSWx3UEdnUXBpVmNrT0o4cy9VK0MyVlZ4aTJpd3cxbm5vYldLUUNLVUw0Zzg4V0pUU2M3anBXSXhpNXhZdm03eWlnTFpoZG5YeTVmeGViY3doQ0IzVlpZVjVOMHYvQmJndWpIZWVzdkhsR285WGdVVUhEMmtBTmE2cGF0Zncxd1RWS1BzYlhFOXFKazU2SUpNRWhrZXpub1ZsRUpzTTF6WEkxKzV5d0hlSUgySnBIY2dyS0F2OXU0Sk0zK3Q1T09DbW1XVXc3aG1SLzRTZzNJakFIRE9wMWcvUXprS0FFMG5MNzFtWk1SWmJLa1Nud3BrMTZ0OFNiNEJWcWRuemxrWi9peUowT3Vqa0R1aXJvUS8vZ3NQNDYxL3RvN1J0VFdhQzYxUHFZSDVLUVBzbGp6Q2NPUGVCaE9DTElXa3hoR0V5K2pkU01tYm1HQTlySmZ5ZDVIZFVEZTh0S2pJTGZmOXA3SnZJRXBHSng0SGlIRkVtaUZrWXRlMFdMdm1oZVBMcnYwaVU5QWJGMkRmNDFQQVE0Y1NJRng5M3F1L3BOcDErUlZvcDUwUm40bndBZ2t2aDZjWEJZanh4a0pWSmxVbTQ0UE5CVm1pTEFJTkZuMk1KRm81UlpVc2ZZZHZPVkdFaVlFY3l1VFZ4YjhsQnMxbnArK1c3NjRvMWo1Z2NCY0E1TUVkbFFPYUFhTVl5U21nRUNrRGxnY1JGSm1xOUtWZllZMnpBcU94NEw0VERYNzRUV2RSWGJ2TWJwcFJyQndHUCtFU3kydGx4NWplNGxFOFB6c1VnbWx5QVFyaUdGcWtmTFBJRWh4UTNQR3NvVXhMejh4cXF0V0c5Wk15OHMyQkZkR0xNUFFiQVBJY3VuUVFzdWVGWHdlQU9QeldCWS95K0JUd3Vsc0pJT1Vkei8xYk90bFA1TVpOT1d5T3NCRG45TDMwM0VEWmEyYlkxQTFZR0xveVdTOGlIOFZ3V2ovMmZkOFdySmhmSEVRaVpTTVhEZmNxYmgxY0NkSC9YclMwdDc1aGNTZTE0Y3pVaTdUWGtYaXNLMkh3bDBnR0x1a29JVDRhVmN4VU02aHYvdi9PeTRXT0t1T3FSd29jS0U3ZjdZSFBEZmVYb21wd1VKeWZITk80dGFjdE9uVWw4NmdncUJrTWNSd2cydmVvM0RFOENVTFdqaXFHdytIdExjc2JodU1oaUFJUG4wWXBrOVRsZzErMGlnUDcvU0RKY3FmSG80QUpSMFNYbG5CNVhyMzJaZHJNNEFBQ1VDMUM4Z09HYkxta3JRQUhMTWlnc214OEZsYkIvdXNtM1NuTjNzTFpkeU5seHFJbmZWV1NTVFBBZEVRdTdHOWVIQ1k0U0o0T3RHaW9nd0hiWElCVlZjUWl1aWplSXY4ZW0xT1NhYUhWMTVReEF5aGlTbXdSdHdEUTMzajdPbVhIeUZ6SFFpVk9yK3Q5eXdaNTdUNnF4VkFTcmlRQ1ZNUFpTcURRVCtpelBiM2g4b3hINFpYNmpETmNxUmtkbFNKWUJjQVVXdGRUWVVQTENZaFhMSUpRSTJtNmFlYmFNL2ZsSzc3NFMreW1mWHJ1NHl2Z29WdXRtOC80NWxkMTI1NjlnbVp1WkJzN0FTd3BIN1Y0R25WMkxpZzdqaWlMUlBsVVNjMHJRQU12QmhXVmlPV0RROXhuRGkvd1d6eklYYW4wcysvaTdMd0NaV3VFcUdaUzBPdm1neU5pY2RNTFpuWGdsOFEyWm9zN3h4QzZPYUNJUkFqY1U5dytIOTM4dUxVWHBnUzBXZVg4ZVNVbkZZeXoxTUdkL2ZjZ3BBU0lwZkw5RkhBVkR4bEx0TDhQOTEwalVoRmdJWlZqTXFzTXVoRkdRQWJJTnRvVWc5aTNVVHg2S1N2dFFubkZrWUkvd3lhUHUwSUl1c3hjYldQZXR2dDFBSWhEaDRwNmZYd0ZQUEVCQVFTNmN2QkxVZThFMkxsRkVyWHpBb1QwM0hrL3U3eG5uaElsSVZEeG91RmVsaGdDYWdzdVhKTzlZOUdmM3I2bG50ZUt2Z0xGRlpUSkdnTURLWll6cWJzcXNFL3F1ZEJqOGE1L05wVXpCM3VMT2JDZGhld1lncG5Zd3M4K1d6WWF0UEliV2pCeXo1bFc3cnd3RjR5c2xtTHc3ZWl2RlBkVzZLZm9mMWc2TENabUNKU3Q4YzFkOTgyVUFSYUl6d2VwM0F2T2xmN2lEeUtlWi9kbGdLMWhjTUxDZUxuSmp4bllENkFwUnFKbXlQWm5mZDRLY05BWEx0bmN4V05mZ3R1T2REdkFuVnk3OHFWc1p3QUxOTExxekhCeDZ2WEVhb1l2WEpGM1J5ZzYyS01rQUIwdkdtRUp6SVlpcE5MblVCWjRNV1hCTmIzSkRaZFFySWg5ZU9BZWt1YzVVZlJJaWt5bHlIcXdZZzE5SHdZdk1ZWlAvTGhNWlBYWlpXQUN0aktTSEg3eDVObk1DMXUyeXBGQkJRWFBIWk9zdCtURCtyeStNVVhwUW55REIzaHNFZHhYd1lmaGh1MWZqNUdLUC9iWkdBbGpiMEM1ZDU5TWg4K3RRdVVNeEdiM1ZwUUpqcklXaGdRVDJrN3o4dVY3ZWNVWHZkZ2FhMjc0NHlpZ2ZXajZuSy83SWwyNTRpQnltd0ZXbHNsRjMxYXIvSEM3WXVGWjZBMmozRW5CNGw2TE14dU01dERVUTFCY1VocTRoamdaVUNNQ2cwbVdBMFZtQ01yQmR3bHh3cTFJNzI0R1l6MVFVcStoTWp1dUJYU1ZHL2lZbUhVVDIzVVZ5ZFpia0lNdGJjWFNKTVdrNTRZV1RaWmlnV0hUVUtaMTZjc3dNMFovcjFodWx4RkRFbTd4N0VnbHR3QytlSHJRY3BkUjlOQ2VGZmxSK3NSbmxNUVUvUXlyNlIvU2dieUtuZzM3MDMvZnVWL1JCaGp0WmJkNjNaY0RBRzY4N25GWVFQOVEzbjNORjNGeUNhR05XdDFBU0NGb2o2QkZCbTZqMzRuSnJaa08wallLK0pKSlYwb09ZQWY3WEdPeHJzYktBYkJrWHhGVy83eFEyZ2o2UWdtcElSd3VWTllOUVk5UEJGZmpQaXFYb2IvclZjSUFWTlJlaWZQRUhEaWdVbW95MEJFL2NGQ2pYSUVMZm4xd2hjazFVQjNCZ1ZtWXZ0TmxJaS84YnZWWjlQNUdNMk55bXhFSVE2Qk91cUxjSkxic04xbVo0RFFQWksybVhKN1Q2cVBna0pqMDkwS3dEZEdXREtxSVRQYStFRUNOMjh4ZlBaWUNDbTc3d1E3QUVsWXUveXZzT2dFMXVTWHlvTGJxRlNlRVpoTXROQ0xrMG90Wnkrd0plOU5Ic1FvdFp5VUVMeHZ3UkZ6TEN2QnlIRjN5NHA1OUhnUno2T1hpVWFoS1QrNnlqN2pkQzRrVFJrRXc5YlRYUlpjdmoyVDhYeWpRVkpoazJ1WlpJa0tXU21yWTJxbGVEc01HTHR6YlpCT1FwT0RoSW9pK3pRTUJ1MVhzRzVzZ3ZnOFJQL1l4d1FNeUtCWlg2ckI0SG13b3hIRytOYUlZRytreDFtQWVSSDFUd1JiQkNXajIvRTFETE8rL1lZeTFaMXBsek9ISFVrQzc4OUtsMSs5alduNHV0QVZpcDI0enUxSndRUURvS0VndnRNZ0NHUE5HcHFFQWZaYkdlN1FVRkFoYkZUSnNESHM5SnRuYjd1RXNHeGduRFB3OE5TWkcyTGVCRHlRSzZVb0NTQklhbUNRU0F2RHYrTlBMR0FOV0FDMVVMMzFXc096WmE0a2lEY2JVRGVyZ1ZrUERZYzB4V1lXRlhHaXpEeXk5anlud2IrZ0gzQ3I1aDIyck1EcGZ1MGlkQnlpTS94VklFbDBiVnNZTVh1N3hIa3R2U21rdDRheW5vK0NDSWVIMS82Qm1jdWxnUGI3Mk5TOEVBQnk2L2pFVThORFJCQUROenB0ZXh2SEJNWExuOUlzR1p3Y3c5OWlobUtxb0pvMzMzZW9VdDFuY3J2am1pUDU3cE45aUJnMEdxTkNJYmc3RWxUMTRXdnBuakdHUUVwa1ZLRElBckZpSTFvdTgzZXFhT0UzZzlOVHlndjBJakZmTWtWdERvYVdGeE90amZBbXVKQ1Q0YkJFUG9LUnNkeEZpNjJWb3lzVUErLzY2bHRFUExCWnd4dVJpMi8xRnB3U1FGS3JSVDdDaDJaVHdPQ2phWDRLeHVJVXV1bStRbnVNZHBFYjl3d3RMRVh2SEVMTDJiUXNoN1Z4WmJVZUQ3YThFQUp6NHdHTW80TEhmczNUQTB1NFhzTm9sMEp4alJZcTNod3VLaFNLUXZrRTlkR2R4bDMxam5YSWh5cDRhcW9ENlFwRFFrS0swTVRqK3pGSkdIWjhCSEhUN09IbzdJbFdoNWZPUk1SZ0lTajBqV1BCUkVZTU1vWVNQcFF3VHQxN2kzaGN0V0N3U3FvNStwR21uWElZUzlYbkEzdEpMcnpUaEhvZVppN0p3R2IwTEYvUnp2N1RUKzFBQ0ZpMjZiV1N0MG1zbnlqU0REaFJuQVhjTUprSlVQWk1SbUF4ZUQ4T3FmUi85WmRzaTJaYVlxUkk2UG5nRkFPQ21XL1RSRmZETC8yL2phclpmOG16dlZGcVlnUTdNK2dERW53UXdKbXNSRm1CNXh6THJlNSs3WU1aaUVQdE5Lb29XQUZEZjJ0VHVlekhuNXdGRTc2NkcySWU5VUl1MUhsZ0llMFlxbXVhUjlJSVpYb1JmNFNTUmtwYnQxd1FPT215ZkEycm9aT2dadmRPQjdjcGRRbDYrQ05VVWIvaEVIMmdSQ0dJRlBJYlNSVG5rSmxwWFZEZWUyZytiS1kyQUV2dFVPZ1BhdTVPWWNIMnI0NGxrc2FnQ0xOQmMwYWVleHVxeHJIMitFblRnMHRvTkFQYmhpRHlxQWdwK3NGSUF5MmlibC9nc1NmNFFManlrSjdSQ1VJRzNCalBRTVVjTVpDaUVVSHBMNU9Bb2xYcS9Qa0lySWlDY1p1aWxYL0FTaCtvKzZJbjA3OXZ3eVNLK2lSOS92NVJZQlVnYXpnVC9TdWtKb1pxb3FyNHFncVVDMm9iRzlyR0o5b2luZFljdzBONG9sc2YrazBwN0JvMlVzR0tEQmtub1IwelI4RkJhREVXZjVuUzVoOHVGSUNxY0E5Ykl4OXlUQmJLd2ZBL0Y1SnIzNlB1d3lCOUkrUlVDK2w1SXFrcXBsdzlNc0dQTlpQU0kxMkZ6WDllOGJBK3IxVjNvcHFab1d2Z2hkNmx4a0M1UjhFRlp4QnV1VE9FbjhhSjN2eTdUeU5xVUZ5SElzWXd5QXBtaCtXQlkxb2pJRmwrTEc3c2Q3dCtUSHJzb01NQTE4ZDMrQnhSZmwydGU2YUpkN1h1enBKSHNzUlZKQ1dHOFloWGR4WG9teWRjTWxjZzkyakpnbldCK0lFWU5oZk1NQ3huN01SUnJHZmZCQUQ4V2RScGtWT0puVU1NVFNoODdPNVVQRnRBODhCTEJUZ1FKWGVUVzIxQ05lN21iRDEwUUVGbTlGamdUbzkzZ2xWLzlPWSttZ0lKRHR3c0FqTWM3YmtoTGU1ZVFXMTFRQnZ1ZEM3OExDbmt3MUlJZUkrcEZ2dyt3VmcvRXZlR0szazM3SUpVQUpiQ0hEMlQ4Z0wwUXltRy9nVkVMVm1TUjc4ZFl6R2diV1hBZEVNNHVSaUc2YSt2ZUxOd2dQTm9OdzFMd1hleG01VityYXJNU1F6cXY5TFBncTBGZU55UmFiS3EvR21sREVBUEk0UzBOSlMwWWpRaE95OFRyeHNCbjRVTDk1RkJaaWY2UDhvc2dFZ3ZsMlFNbEhYand3WXoyY2ZNY05MTnl0SVkyei84SzhNZ0tTQnpiNWFSQ2ZqYkd1d1NNekdWWURwVWlPUzVxNWFEV0ROQkJiVnFZY0pjellnL2pJUVpjdUZjOHlyNiswRmtKYlJ0MGZLancwWmJvdlBGNy9veVlDQ0VZb0xjWWpMck1jR1hESXBIaVlJYkpDc0pyRWhRUUtkcmdIeFFPdm9BQ3dDU3c3RkN4aHBZNDRBY0hiWXlsWFhBbEtwL3R2WjMyalJxMG9aL0p5UkNreVNOU25pNDBHZHd2SHJyd08yRGtOSUdlQkJ0bzdNRGFtaWpkQXJ0MFl3QkZtYVJHUGQ2eEhYaTBSVW5iRGhJQWRQbmdjdVIxU3k5alFnMnNOR0tnNWVLT3F4OEpYanByMzAvUkhpZEtZN2hKS1d0Ym9RUXE2WE9lL2FOUXRFNzZrU01rVlhDUGFUZjBYVzdkTnZ0SVNUelIybEI2SVY0Qm53Q3AraTZVSGc3K1pYRkN0RVVmQ1V5U29GQXZpZ2xJRmpPdmdFRFVkWkkwQWpRSkZ2UFVneWNwbklSMFYxY3N0QURCQm9ER0ZTdjcxWEN4MXFYTXdRRTJGNWRWUHo5RDl0S1AwYUF0OFRxQ0RBMGpJWVAwS0h2RElwN1pNdGtQK2t6RVlpVXFqWmNDZ2VYOWx6MjZBdTYvM3BxeGROMmw0QWpRTGF2Y1pZZ2ZkbkpwakY5WlcrR0tGYUc0d0I1dXh3blpERTJWNlpZZElXV0xBaW1wRUtZa2JEdHhMQXljV2Yrd1pPaHIwUWlQZkJXNm9mMW5FNlJYT3ZaV0k0bEowT2FIRk5kRlpSNkw0RnlEYm1ZRFNrSlNHcG9ZQUZEYnhpMzVlS2ZrL2JLK2VnUUFrV1FkTUFOTWc3Y1ozZG1PbUV4dFlVaE1iaHRiWXhpTU1yUmRZcktkYUdjbDUxRUdoVVhGWEVyQUtnVmRHaWlZSzFEeEFzQmdsM1NmTzFTSDh6YU9xdjFuQ1NtVmtNRm9GTXNEK0U3d1V1WS9PTUN4akI4aWF1Wm9TL3FZTTZoSldCRlNiM3MyZ05WSFVFQUszbHlaSTJKNmdYTU1UdnBMYUppRk5RSlRIQ1U5YlJEenhoWEY4VkIwQ3U1NndxcjExU0FEd0NwVzZ5ZURBc3NpVk1KWGFwbGtxR2F5RktnM0ZKOTl5UmlyOEVPa3F6QVlZQzBNVzRxRWhKa0NtZ1dKQXFsQXFZQVJhdFFqUVg0ZVpXVVpnOW03a0FzMjFVcHVrdHk5dUxwUVlzdFM5dzV1ZjR1LzJyYzZ4amU4YkEzMU9MSHJiTTIzencrVFhSVXpVQW1ON2R6S3p0SE9SWWhLbDVObTVYaFU0OTBYV3J6Ny9pbGt0UUpMa1JPZHJhQlhDQlhYSFA5QkdZUitJc2VnR082Z2Z5OFV2WlRGMGI0UWk1Ujg4bUlRcy9kZXFmd3JNVjZTQkZCQWtRNEEyUFlJQ3VpdUFhaUZ6VUVPOFZjMGdqSEQwa0NUSk55Q0taOVVnVVZzVHpMNitKT0FER2ZyWUFGcVFLdFlvK3BlMmswOGJUVXRBMnNRRUtscUlKOXM4QVhYci9JdC8rcFo1Y3ZzZllFRzErQVBZNGR5L2c4cTlQUlZWVWJIRW1waXh6a01KWXYrM3BTRjdxc0pkK2hOdzNaR2hmVzFsMDM0OHovNnpMZ2xzWWpCMlgvell3YUVqL0MzQU9BUC9lY1Q4dTUvdjhsNmgyMTdNbXhvc2Y3bFc4VU5FMEtEUjFCNEtaMHNOS1BVL3hVY0hlTUpsRzBRRUlheGh3V2w0RFVNeTBWWWtCUWlJMVZMZFFhcVIxMllmaDJRUHVpTHpnTkcybzNVaGlyMkp3djhVSGFhOTk3YVJKRVNhRVRsaFZUUlFhQWl2SW9ycEJaNHpXWVBLM0hnTUlpMEZnWUNJWVNkMjRRakl0bEdrcFNFc2gwUVNEQ0pIY3luc09MMXFvZTBKUkdjYlpNTnBEcEpHQUs1cUxDQzR0dmxCaEFVS3BYR0x3NGpBaUtDRXRBblJWWkkxMlg0UkpLcXA3N2RBSVdOTVo5Y2lRelZwK2dIWVZ0NmpFY2lKKy9iQk1heG8zM01QYmRTeFNxVnUwZ1pHL2JQUSt6REJhQjRMR3VDUXhTQnJVZlJRU3Y4SGdWK0ZUelJGM29VVk9udHN2b1hxeGVYVVRXWnZIanBrVnd3Z0lUYkFSR01FdGlXVVN4R2hhSzlwcWZBVUVDQVVTU2lERDhFTktiV2xDMTJSb2hzUUZqQ1B0UHRiZkMvMFhkY2ZDejdYdHZITTlIT2MvSXNhT2lyK0FGYlZyNUtzcEpFRlR0b3lHdUlJMDRoUWFRS0V1dDQvRHp6aTYxUkVaSk5QeEVJSlNYN25RajM2M0YzNEFzVndwQWNKdVBLcXF3OCtJOEhXRnRZMWxFNkdLVFQ4Z3VNT0FHcEVpbVY0TU9uR3NHb3N0bFdTdFhjVS9TbFFsakFnVXFCd01kUUFpLzJ4QUlLSmtTdmxFT002UGNZWm5wNlRsZEs4VWZoSnJNWWU1UEJxa2JTRGlKVlZlKzRidHNqRTlFRTlnQmpxZXE2VDhzT2FJRElIZEpMZFlweUJrQU9kWEhUalNDV3VmaVFraGVOWnl3U1NLVno5dDRBTEE4NkhrQ3ZZOStTb2RxRUwwYUtCZDVSMlRDSVYreDlNY1d5T2lybitTNjJ1VVZsaWxkSUZ1eEx2OFRNcUJrZkRRSzJlRG9LVlJ5d3hIcVN3YzBKaWIwREdkejJ3Z2Y4M2dCUnBTUWIwNHgzZkhnS0xOZlVjdmlic0dRbDRvR2hOSU9QdUdTR3VXOTNvV2JXd3phaWhMbnh5bUM4eXRzaFpFSC92U0FxZmUyb2l6MmlNcVM2eWl2MXprZkpoQkN6L2Z0WHdGVDNabGVzUXdzUDlFcUkyS2k1SjdPa3h3dmx5MmFLaCszdkpSLy9oZFVjTXNyeWUrRzVZWW5SNjV2SzhBNUY5L3dYVndCZlBKUW8wRUNVVXA3ckJ0Q3QyUEJzajk3NUZLSDAveWt2QyttNnJUSGdoUThUdTZORnl2Q1ZxSDBrdVNnRUh5V2daT0lHZzQ2Qk9IM1FCZENOTE0yR0FwTUVPNndhZ2tXOVRuM1QyYmZQZXBBdzVISUxoMGVnbEZVNXhMSjEzTEZtUjRZd295OThBQVlCSjZQVi9UUVBOYkFraEtSVWFjTGFveFlqYkk0UDFtWGxXY3dDRXYxdUE3RW5ocytPeU5MMUhmR0l1TmpzL3MzUzRPSmxCMjdXWjlGUTZWemJ2UTlTckthVXR4aUUwb0x4Y3pGYm9ZQ1FxVkJIZmN0RGpoYlBlaE1rZHJNdkgwWGY0Z0hhSVNpazl1ZTZsVDdRNDkvZTV2UWJqc2REaHZKbS83SUFhZmhaZjI1dmpZTlJPbmE2d1ZRVXFNVE1Yc2d5eE1hRlc3Z29nKytqeVRMeUFIM1MzSnBUMUtaZ0JTbmp0WmkyOCtFZGRLWXdJQ3dhQUFlNVJnMmJybFJNeTQrZ2dMY0lBQ3hwbDFoTnFuNzYwWE8vWlJKNUtaWFpWREFMaGp0TEdwVmduL1ZqYkFlWkZCZFErZGR6bUpUSXNRNEVHUGpFWnlYY1VRNnNvMUxxMUN0QTJHZzZiakdxemdkUmlvakxxQ2NLcmF4ME1INmVXaHRZUXd6dVh6SzlDUW45RnIweFVVdFpEaExFUTB6cDI4QmVKOG80MGlMMHlLYVF3eVFNeTNQWkR6L2Y4OUV0dGpORlpUTDNMT0RRbXJrdEdxN0ZLUlhudEFQZnk2MkgvM29xVDkzRHNTeDBkK0dVeEMrS040eHhIejZMZzdHR0F0cjZZUG9wRVRLckg4VUNDbWF6OHlvY2JDVmZTTzRZK0dBcVdkeEV0SDBnZ0FLM1BGd3YrcmNBZUdQaWxmemxFQUI3UjN4Y2VuQWR6dGIvR1c1RDVZMUtLVUVSeHpSUWJJTkM5cjQzMnB3QXBCVFF4eFhIVi9mMHptTkJRSU13Q0hIOFFnUjl2bFl3dkVhUlRBSWRYQUxGb3Z2TlpTREZTTUxKVUtZb2lMTjQrQlAzemdVcE1jVkdSV0FrQlJoT0hWd1VKY3E2RHVtRkY0b3psT2VnYlc0eEI4WkhJa1hJOHZsUWhjQ1hjWkpLTUJ2UmpwS1dJNUQ1U0FXcFIrek5hcDZSVlFkQmhUY3VMSnB2OHc4ZDRnZGVoUDFDL1Z3WVd0cFU5S2VNYkNuTDhpNFBoRmRvbVVGcWFlQXRVQ1hrUVZjQ3FCSys2TVdkRjZXbkNHVndHNlBDVU1LQ2tOMndJNE43OTY5TFNWVlROYlpuaXdOSXZFKytPNEt2SEdURU5ld1ZzN1NqTCtQcXlWRWREdkpGVjdQWkFFdUNmaHhnOHJQZUU3RzJvTFMzUkFrRW1JeXhpSW1OT044c3VpbGx3dmR3MUx0ZmhxaGZveU1oeU9FUStiZ0ZKMW5NV1Zrck9IdFVDNGlUSjZmUXR1MVJUMGhrb083RnVvbVVNdTBJQmhscGpxaWV3RURoWEJuRGhRZmlDQUVOSXZ3Q0g0ZWdOc3FIeXN0RW5DNXE2S09ZRHBocnRFM3hmQ0wzdHNiYlkwdzFpOHNWWDlEdC9QdEZRKzl6WTJBdCs0cUpRWnRodGhBaGNET0c0ZUllVWVTV0lRaVpscGs0S0UwMUMxbzUrbjdIM1ROZ3FSSXRiTXZBNmczckloZ1BHQ2dPQVBSTW1mUkNDdzhXUUtIQUhmcVk5MzBZdXRyaDVnd3NYakVVM0NlRS9XcEZrNWtDbm5zRUJXVDRnNW5JUnFibGJrMDB6UDFuRnNxbVNsMTh6M3lYZWpEMHVLVGdPamYzRXZqQ0c2ZUViWFJjeXBsNnR4SFZJNzBySDB3S08ra0dRRThyd3ZnSU5RckVsa0pGK2RTQ21nQmVXMlhERkhPZzRNVkZTMldHRkhSbTNWZGpwb2czTGRqdy9ZUTlDak96SzI0SWJPZnlpLzBpUUNJbElNVjg5YnIrb2YveCt6Q0pvRkhpamxNdFVJdnRDRXBHSUdkZ3llVEZYZ0hVWkt2bGIvWXowUCtPclVJaUJUZFVzbGdrWnVPaC9ldGx5dlF3S2RCckdBNldtaTlVVktZRWlMWTV6YzV2UG9JQ21yeHZBRlR5UE50cDJURmM0YjRaeHRnNk1ZQVRpT0NrLzB4ZjZsUTZXNTRseGJvV2ZEZDRWdndlZnhJc2doTFFUN0h4WmcwOWxUWEV6blZNcnNJcWZxaE1BZmZGSHZrQk1xNjZGSlhZL29HRDVSRkZYekh3czZHNHh1S1UrZ3N4RzlrNzQ3NjR4SW9KTDVaMytkY2lFWGN4QzdwUFJER0pBSGtyWS9OY0Mwd1NtUWtVU3VEaUt6b2J2dzQ4RXRDekZNV3hEY1lueHBPUWZwZUZZbGd1bWtRczN5Z0tVNkNUVzNRUk1Ja1FGVVRiTG04OWZQNVJhWmgzQVdSdTUvNkFNdnhtbkxLRU4wU0FsejQ2dHAvZ2owSjlPSkJ5SDh2MUpHT3ZaSTlnNHVOVFlaNkdBTTM4VGZhejF5NG05YTFNejdZTVVJa2kwb1ZwWXgraVV4a0FLTzZ6ZTZUZXR6MjZXbDdTd3VtWk1oWXpnSjRSZGlHSzBJcjNQMVpSaHFqRHN4OFhqNDI0c1FjQUhqL2ZZYk9qbmVFUU5YNG1xOUQwWHE3REJVS0ZuSFlqSFRJT0JTMEtPS0J6VU5ERVlJd1kyR2NJbDNxdlZHWWNZT2hIZXdRaUF1WjVWNjIvODVFc1lCbmFEUEFDY3RmTFhzcnNvUkU2bWdZVnZGSTZiRDFKd1g0WHZGWXNtN3ZRNFZrVmhVajBXL1ZSay92VmtFNjhpRWpUQ2tESkFRL2lMWmdmdEswenpQYWlaRFlHZFE5bVBFd3h5OTRJNGxPdW1PNWVPUFRrc1JUMTgva1hXVE9LUjB1RHFzMmUzQmI3L1JHTlZTall3cnNYOFpicWJYekhQVk5zeklBcTBVNnNUb1YwampVejBvc3IwdDBoSVVFNVlzdnQrQ0RtRGxXUFgveUsvUWZMaDRESUJoWHJpcmpuNENnd0V3QWkvUWozMExuUktheUk3T0pMb05uR1JzWVBHWXNGMzEwUy9heGhWcFNqNXduRXBvZzkxdlBjQkQzUThNOE15L0hMNTdYM25PcERHcDAxdlFzc0dkOFZ4RkZlTGh0cGRXR2dFSnFHc2x6REt5SDdKb0tJV2duMlEyM2ZTeEZXRG1pVUVMSXJOQTNhaUtTVTNHMUc0T0IzOTUxaUJXS24xOUNPVWpSbjNROWEvRjRCVE1rYktJTi9mU1FsanFVRGNPN2VUYUFqVTJ6S1hwYkpxaUxuUGhkTXVsYTdwUzRZMENPVk1xOWQ0TVhYa1lYWHRiVW45RFhoZ2w2QzZOUHlROGhVVnZmMTNvNWhSRldGR1JUZEFOQThzZ1c4MlJhbFMzUHVkdk1pcVZjUWdSVkhxbXZYY0ZiRlA4VWErdnZGRXNaTjJJZnpKWmh3QWNMckFVc3NnMkx0Z1lGRlJhOFhwaFVERU9pRzBSTzZGcUtHaFZMeFU0eFFiSklwcUlZTHBzc3hKY3VJOUpJUlozTGdjRjBTa3lQS1VxNkx3akxhUTEzQ0lXaWhTUEpOdW5vUGhZZzFreXBFeGFkOThSUitQNi9LQklDelp6cGdaR3RUZXZtRStmZVNPQ3kwUGVURmoyRXlocW00MEsyb2RvcE9DWk1Sbk1PMTRDR3RZakJZN21HNHh2LzJKUndTU0RsQlpIUUhnRWVLZ2dHYytFa0JnT3JDaDA5Sm5xSGZRc1lEaFdHUW9Nb1M1YUkwZ21INmVqTWUvNGgxcmhRYURheThPOEJpVGN1dEhCQUdGUUR2ZFBoQUFIbm9RUUkzV3NHTnhIWTA2SHZpL0tEOXg2eGpDcEozUU9UeFlyVG1JMndzak4yclh6V25pSVVuWWlBUXRsaEowVzlGcWJTVW9NR3hCY1FoQkFaNzN6blBHbGxabDBTVnpJLzlyM3RteEhqa2N6WUdmdERTNGVibElmYzBVT1J3MVNWMkdlcHJUUHp3ZHVoM2NyeDQ2N3A0WnF4OGpKUmRuN0VvRFdPcXJGMnBBdVluN29JeFlJOStwVG8vbUhRVFlPeXI1K1pVeUxMdFJtbkx3Szlad3kzWnpUeUFVZVY5NjNHSndIekdxTnVCb0FFNHVIZTVwS2g0ai9HVlVoY3hCMDhTanNFT093cjh0eEJhK2tUU0lINTYxdDJaeHZpUTMycmhiOEFDRlplaGlwRXRVWmRWM0w0WlRzczFsMTJORm5JOXhUM1pGOFF5RldYU0FIMmlOb2tRQ3R6eG9YVmdPZm1oeW9VRVdzUjZCY1lNYllCM2tVa1dHUWxlbEUzU3dhTVpWbElHdlplRjhReExiYy96M3dON21uUlRUSFhONE96a2FlL3VJMXkzM2FRQVVQR0JkNk05MFVKUzVZMHp4VktpM3doNm9TRjlGVVhKYkpURzkzUmhRZTRsTHprVVJFVFA0Wm9kdThBSjZFaFUwUTdlOFBjclIxVlp6U2hUS1RrcjR5UUMyb21qOXBheS95cGQ4VHhucDJZMy9MalZLQ3FJRVlyK29nUWkvZHVwY0l6cUI0UVlZWkljQXhwOTBkbHBvVlQ2T2NuZWxpNFR1Vk8wbmRwQlA1bkltZWd5bUpYb3NySnJGUm1VazJjYk9YOXVMcGdJQjZjSWlGYzJBN0VMR1BQQUNwV2NidUM2N0ZqUHhxM2tkSXJWOHVubjhpNEkyM2xhOXF6NUl2NXpheEtlVVFqbmo0MWFTRldTYmgwSitqN2dVWS9xT2tKQVVCOTczLzNUbldlbm1EeG5oTHdseHF3VjNJQ3lKZ091eTZXMnJ4VEtCQzF0empJeThtWGxHb3FyZEN6Qk11dEtsT1l1Tzk3ajRIc29jaE1abTE2TVIwT25XYXF6THBKVTJLaFNvTmZyaTVqL0tldCtTemFqcHczaDZNYm10NWR3U1NycDI4VVR1NHovU203YXhuV0tOZzJFZ09FamhtM3VmVFRFMWc4QVBIYW1reE5kQmFrRTJxTC91R0J3bm0vSXB4Z09sS1NCRlEvM0U5dzQ5ZDRZbENrWGhpWEd1OVFYU3Q5REJ4ZkZwTHN5RjlldlFGYUlWSUtVa0ZJU3pNNmdQZmVlZHdPUGRWYWNDTTZUMDhSMEg2UytIaEJiaEpROTh1MWRMc052bEFSbVg0NER6MW1sd2RDRmpPbGRrYUpvcHJ6R0lDbjdaWUs5TjRxYlJreHY0R1JjNGV5VWVPQ0JLZHVPa3V4WUM0YjU5K1VybnIrd2djbmVGS2ZDbVVSRktOUk1YSEp3SXFOUk9ITVp3S00rd1BDR0NFV1lWSWJPMi93OGJVdHMxd2xKSWp4elh2a3IvL09NVkxWSUJXRmRXeGN6N2JCcWFsL0xYdGNXOEpBcXlRR2pxbkJsVXVFdGYzSU9ITlZJRE9wOG9MS2xFbElIdjN1TGh4VVVRNTZiRXRwcGw0YXNoL2xRaFBBOEJTRkFiRmhPd0pabFltQ0UzR2hZOGtCUTFZUWtNRldROXNJNWJQeloyY2RTUU9MbC82VEdiVWZtMU9YYklQbDZDQldFTGVnZ0NiSE1ndSszRUlESHlNYytnUEpHVjdHTXN4ZEcxSHRvRC9sN3ZPR2RaN0padjZDZ0lXVDdYVnNBYXpYZStaRVpudkdhZHdsbjhLWEZFc3MrcFFReHB0eXdPcjBFc2JVYzlqSkJxVVQwMUFiLytOYy9oeTk2L2s1MFdhV3V5b3dwTFVkSU9QcVlZcE9Wdm1iUHEvVkZCTXdaU0JYazlydTMrRzNmODFGZ3p3UTJ3YVNubGlMZ3NJa25aUzFPa1B5MUFDbEo3SUtFZzh2UURySmd0ZnBOTjFtczM4ZWsxVVJjNFZDV1VWZ0JjSS8xL01VQUdYWnpKVlFsU09WZTg4dG5ZM2g2RHdrS3RBT29JbWxFVWViRXFnYTMzZzdnQkE3ZCt1aUxrckR4a01uNndvZE95cTdudVY3RmN4MmlKK2VGVklsU2RDY21nTENDc1pCYVBHZHFhb05GNm1BSXMyaWJMOEpQaTlZWWQvWEVoSzlCQ1l6c2JvQ3JDZTNhR20yMWtVc2pBZjNLdllESmNUNkg5Qnhta0I0VEFUOEtFYW5DTlE5MGJxQjVvTUFYalZDaXlNSDhnZ3hWMUQ5dFRGRGlhRVVrWFRsbWRmbXE1RXc3Y3JXNE9BQ0p5ZGRpdUlOWHdIaDBwekNEQW9Gb0c5OHJ0amhjS3dGNmdnQ1JCdXNWTUJURDBHTWdXUmtzclhSejdtNDAwbTVsM2EvQ2EwQ2xQSHRJZXZlUmo1OVcxRUdZS1ZWZDdDKzM3bjBZQUhIczk5S2pLK0M3emlvQXBQbEQvMU5ucHhYMVNnVnRXV1l0eE5CK1VzZHBET05jQkYrc29TY21BSWd2U2cvZnAyNmZYQnVsUDl0c2lFTUNvdEhOWUNiTFNqblhWWkEyUStQSTRaQ1BQWUpsbkVyeVhLSTh3QkVTSWEwQVhiREV2ZVAvbU10U0xBSUpEaEVnbFpXVThpZjMzSllPTGs1S0FaMHJNQ2UwMVY1NWlncVRsbDBTRHlEZ2JHVkJIajF1QytVS2RiZFlvWGVUR3B4NG4rMjBmK2hIMWlJdEpBaUtRUUFzMzR0UWwxNEtQVFdXWW80czdQRVRTaDdyazRXMEJkZzFrUkloS1VrK0Q1bHYvVjdvMkdQUU1FY1ZFT1FIZitQZE1qMnhqbXJpTkdqcFNMSW5hYmpTMHN3b2cvRFNMQmE4VVpUSjcrRzdVSm55cXQyem41R0R5UVFNSW1BVVNraWkwd2psUktUUm1NU1VWSHo1cTMrZUltUmxJa1JpbndBTVRLa29hYWt3NWhkZk1aZXMyU3pyaThzYUsvUW96TzQyd0EraFQ4VzRlbTVZUWpESjFYZXdUN0NnQjZ3aDQwSW9FVUVpK2YxOUxBb3ZWL1MzeUFveEZrVXJZMHk1NEhWNmVtWDRvYjVLdDJ3SUd2Y09CUTRmVEFFenBSb1p0UzZqSkxPSGtCLzZyKysxRHg3bFl5Z2dpTU9hQUp5cjlNTGJVaHFCbXJXWWU1Q1c4dkZRdnh5ZkVQV0FHcTB5ZDYyT0Y4cjRGQXBHeW00SDViM2lLbnJCUnBzS3hoalF1NEdCSW9mSndEV1JSaGtJTVVsL1NHRTBjY0V5QTY1RkM4aDcyQVRwSmR5aml0UlR5TkczcUhweG0yb1JRZXhsTG5EOEhJMEpmenBRbXBpMHZaWHMyMHR2WjlSaDB0Mm81dWlwMDJIaHNZb0RsSDdTaHB1RmpWdHNibFRTcmZHNFFsakZXQTZTQW00UWdtTHJ4MCtRSGY5VkkwcFdyVVJFdXd0M3RHanZqUTQ5SmhHTnQ5NlNBR2phK01CN0VxZGNPTDBlUGx1NjFwZUUrWXh6UnRkKzk0NFl3eWJGcEpYT0VZaWNjVmk0WWtZR0F1LzNsYmJQMisyNTBHazY3MVJTY294NzlacERHdWtXRHpmTEIxdG9WWEsvRjB0Qkh1MVA2NTNRVGdaMXN5RUFrQkJidDNGUWhXb0VUQTd0R2Z5RUcvT0JWbHJaazdWYUMxWFNmN0NmckZFZWg0RjNNRTdRcTVkQ2JvRUppMkxaanBIOWJYdU5Fd1dnN0lzUzhwQWZsTEltaElQbkZ6QXdtTjFXeUVKSmxkZXprOVhXUSs4RHNPNTdrSDhjQmJTekhKRE9mK2oveGZRQlFUV3BvbXpUZ0dobDUwaG9YalJmL1VibWtWQXZJSEZCK2NLaVFjTHE5SXRrQm5GSjcyc0RFMW0vaHBwUTZnQWNmSllKd2tEZUpRcVAvL2FXQUFsSWFtRkZYK3AxY1JxdXRJUmV1bW1sODJxRnBMNHlwS1JEUWJHaitncVVaR0JaRGF2RU1DS0R3Q1VzVTBUSEpWOExtNlNGR2NrZVlNUjMzV0Qzd3lDbDVENEFCZVAwVExyOEdBcGFLT1VJeUR6d2RYbEliMGhLMlpjL0tQYVd0dWYwazE4VjBJNnBHZ21TQ05OSXBEMG5NcjMzTndBQVI0L0NKZjlZMXhFQ2xPbVo5OTJPcmVOM01xMVk5Q1RKUUt3b29ObXNZUEdocm54OXgwTnhZdWlIUnNSL2kyUXRvM2lWTHVDaG0rek5FL3N4WFp4OThUaVhaNy9UZkIrbExXNzJpSjZhaUN3eEMveGE5UDZMWC9OL1JaSE1Bc1pxSWRONVZ6V2hxZ1lHSE9SVVlzeGs4SGZJSnpBV1E2UjlINkowM2hJQUE5a3hmbVhmaVpEU1FHbUt5SVhPVGx3YzhxRXNleXBJQXo2ZTVUTTlCaXdXTWZvUWtBaUFab0laVW8vTjFxWXF5ZGI5Ni9WRC8rMlA3THRIWStvLzVrWGNlRXNGWUQxdEhmdGQwVXpmU1kzRjFBTkExOWpzR0dEaGZ0QUpNTW9vdlpGbGk0aUJhelgzTGYySmxoN0crNEU5dGlTMFlEN0JvTUxBVWoxeHI1QkxFR2hBY2NVV0JNV0F1c0tXRkpXUERzVU5BaTBIcjhPUkErRUhyWHIvRWl6WG5CQ0ZyK0xaWU5zcnVrb0piakFqVUtObzRHYWZMS1VJUjQxWmlMNXA2YS8vNU42cjRDTFowZkYzZ1NzREMwK3ZYdExzM2luKzF2aGViN1VpeFdhcE5OZXk3QlpXZTFrYXZCcGFaeHVqN0M1YU94R0k0YitrdWFwcTRjYWQ3ejRQM0kzRERQTGc0eW9nY052dE5nM092ZWNvcHZjSzBsS2tkVmhtb21ZZ056Wkk3Q2VLZDFTS0VNSis5Zi9GSXI2THVyWmlydnBCS29udXdZeGYySE9hWGd2TVBpM2tDZUNCa09MOTNwcVd3c21BaThyWUpjdFNhQXZUSGNVUStPTDFzS2ZXS2pyWW9OditDRUswTjZtTTVSVGFVNVJEM0Rrb29NYkZvTFJZR2NEUHhoM0lhRENySXJBYjF1WEYyNnI5NTNueGZRZktIb2lEUXd0bkU5VEljSWNxTVQ2RnZsUUNMWkJiU2wwekpRaWtKbldLZHVQaFh3WUFITG01V055UHI0QTRxZ0JsZnZvRDc1U3RlejRpOVZJU1VpRVZJQlhOdXRHc29KVHBJNldUdlJrUEt4anNmYURFZ1lBS2h1aEp6bmd2M0lJYmtlSVFGaXM2QnE3V3NXVU0vTEJhdnBRU2xBRWMxaDBQYmpkWWV6TzQ0dE1mSTd5THZ4N01VUG1lNzZVam5vMG9PQThoajZGRlk5OS9vUENDM2k2VURJZHJUVm5jWDBKbnc4T1JYZ2NDV3BoUzJ1bWxIclFOMGE0alhKdU1iaFM4R3dYdmhYVjFZcnpJVWQwS1pnQlpaTFFNa3BRMHJtWGpubWwzOG85KzM1NXhYWkhSNDFEQTRvWTMwdGFkUjVPdUM5T0V2dE1vRUJnZ2R3THRCSko2TXBxRG53WE9LZnE4YUZUOFh2NkJRaWYwMXJCOGtYMVJaWkVwQjlrUnd0TkYwbU1lNysrQ1JzVTllbVBTMjdFaDFsa1VCNEo4VUlxamk2aWU0U0sxVXl6cHdBSVdhQnhNZWR4NG1LcDBSWE9MTTh6ckRzbm9pR0packpVL085Q0NFWXlGY0M3ZDlqSFM0VTRXWVJUWVQvS2hoUVE4T3ZZdjZFQjI1V1gvTjNkQUdsRkdFNHNaMGpqcHhwMjNBYWMvYU83M1NBSGlqMGNCZ2R2OEMyZmYrcXRvSHU2UUtnQktTL0Q3UXpVVHpieFh0Q0hnaHVNSHpTZ3VWQm5IUHZUWW93UXJORTdLbGhJT3NJcWR0bU5ZTVpTRzhOT2ZQVkliNExtZ0NNckpuRzRaK3JMMFhuRDlBRVgyd0RlaldLam9RVVJKQWpPa3liaVdGR3RDckhMU1l3NnJTYkF5UWNBalQvWFRxZ216TUYzdkZpV3diRWFQelliTEhETEwzanoydHhRc3BpcEFOdnlZczIvcEcrVll6dGN5UjNRdDVSVE1CZHhZRkRKZTh4bWpLTS90cVM5M3p4bTlHeUtRTTVCYmtkR1NKQ29oSTAzZFNWVG4zL0dmQUFCSGJsclF1Y2VuZ0lEaU1GTjcvdFI3dVhuOG5USmFybUhSc09XQVFTdEU2T2JXc1pJMGRrMTAvUnNNT0F1R01aMWd3WUNnWGxSZUh3b3pXSUdsWHVCVnZoOEQ2alBiallhU2lEVWtkSldLNEVTak1jVUsrV29abHJYMVVYN3ZIMlBmcEw0dWlZd1NmMUU0QUU1bDNXOTVoSnVvd2FvVUR0byt0TXJPM3BrcEJrby9pN2RJZzJ4RktERVFwNmdacVEvRDBYRThhOXlvckxkUlFNUVB0ZlpCaXQ0Vmd4R1ROVEMwUDJzWS9QUUlXQW9QbVZ0QXdEUmVBalZyTlo2TTA0WGI3NTZkZmVkdm0yRy9iZUdzK01lcmdNRHRSd1ZBeHNuM3ZERjFteUpwNHVtZ2lpNTN3NFB6TFJhTVU0WnNZRVBDd2trSVd5d3FLMzV3dUYxRXlFNkNXUm5TS2FIQVhHd29VVEkxS0k5bXFaenZLWngrbHB0d2srdGZZUm85dllDTGVDTmFTWWl2RjVFNGRDMzhFZWg4aUFVMklvd2d4UFdNZnV0QkZaQy9yaTZid1llUjQxUDB5ZWhINW9heTJQd2xzbXBmTU9CTmpUeDdaQzJHcWJ6WXBxbklBR1VPRjlFcHNlQXRoa0ZNWCt6UWYwZGJJamRNNDJXa1ZGUFNLRmZTaUt4LzlPY0JiT0IxdjF5RlFzVDErQlh3Nk0wS0VlakozL29GMmJqalhobXRqa1RWZ3BIb2RFcUNkZ2EwTFgxeHRqY3hUTFkvc2c4c1kvcTdLeEdVL1V6Sy9ubnNCOFBnZjBoWWVtR0hBTXJ2WVE1NlR0Sk9IRXJGZ3hwdkhRb3R0dGxtdkRmc3VBeE5wYjBBTCs0MTlXY0N5cGxYQU1DU1MwNDBZNVA4R2FaRUpkTmNoc0xsZzNqZEU4YkRDSG5vRmRDcnZEK3d2MzlKdFRHVWJKajk2SjlWU3JDR2JZaklISEMyM0NSbWdVemZ2dmpwOXhxek5yU3RBSVJNdHBubHJDYzFMM3hrcXp2eDdsOEFCRGg2ODBYRzRoTlJRSUI0K1J0cUFCZlMraC8vZE1VTmdWUVpsVVJhM0hFYmdmbjY0R3RxSURnaXZoNEh5b0JYNnBVazJQK1Nqd3plTUFZd0V1RCtYdFFsMHA4Rmp3UzkxaDRTaDJ5TDl0WENoSE5iRVUzU0lsNE85WTlScmgvYTNBdkNWOVdSVWxsK1VuUlEyakRZV3hCVlFyZ3FNeWZCTzVZOGRtQ3g0dUk4RUtEWXRxOXEySTJRQlF5b2diMjA1K1FBUDNGVUYrOFZ1RGJ1WXdvVFBCOTZ6SWZlZWtZUUdBclAwcTY0cjQ5QmhtSFBEc2d6eUdnaWtrYWtVcXQ2bkxCMTdMZm44enMvNm5VRkpmajRaQlRRZ2hGU211Ty84Wi9rd2dkT3kyaTFGZ1dSUmk1Y3QxRHRISmh2d25mS0w3MXlITmE3UFRQcEEzVGwzSlZSQXdNWDdPOFhBVVdxYnlnMDE4cnNBMW0rVTB5SVdZT0lsS05RWnlnU2ExUFE1dUlIT1VUVEMxMHNOaGxpUDVtd0lqb3NYQkN4QW1leUx4L3hKSS8ySXVHaUJYY2pQN0JnQTVxSnJqd0RXMXpXMzRRZmRDaFRtaEd5WHVCT3ZlalZiOHJlTFBiUEczSTNCUVBLWUl6UTcrRURHOE91QVFCV3l6c2diSUhSc25EOUx0WGpiL3ZYZ0N4d2Y4UHJFMU5BUUNFM0oyemhPRSsvNjZlRmJXSTF5cWhIcm15dU9DTEFiQVBRenQzdDRFaUFFSnpERWp0ZkJJWVpJMzd1RlhVNFFJN2VKSlRQWGZWQVZzUEl0M2M5RVJYakl2ZlVhelF6ZW90cnBiRDJoa3JnOUJLcWdERDRMc1UyMkxpcTczNFFhUzRBWmYxeFA4b0FrL1FwUWhRMDRTcFFsTXFRaVhzWGdlOU13QkswOUlITDBNMktLNkV6QXNVN2hkemhFMVV2VXZUZVhSZVplQUNTZFNESHdSZ1V1T3RlclpzalRiWkpTalZKWktuR1NVNzh3VnRtNng5OE8vQzZDamk2RUh6RTlZa3FJSUNqQkNudEE3Zjk2M1QrM1NlbFdxbUVJS29KakJib2JLeHlCOHpXZ1ZTc1lERWxMZzI2VitJQ2VSMmpYUUExdytWR2FJdCtPZ09sbG5CeHZ4ajdmRmxKRjE4cGExazhDSko0amhUREVuZ0toWFlaQWlXN2RYL21tb1M3UzVJa01pT0cvWWRmS3dkOXk0QW54RVh0QldMWCt1RjZEUFVKMHUvN1p6K2w4TnY3dzM1K2VDWE00TDVGWVIxbkw1eElHdjk2VThMVlp2dFNqMGdHR1JGWFVEbzh5SzM1aTZVMVVsdWlYa3Zwd2gxb2p2L3h2N0QrbkhoRTZ3ZDhVZ29JUW01T3dJVXpvN04vL0crVHpoSWxkWGI0UnVYOFVtZFV6Tlo1d1d3RFNKVU9iRVZBbDBHdW1ESHpYZkRaRWJqandkTHA0cjdUUUJBOVQxZytFd0xYUVhBTmxxTUg3RHVSVWtLNWtjQW1rU2xoaE5tTElRbENCLzF2U2FLYWZHaXRPRlhva1RMTkJyclhwUlRNVzhBWGpOY3MxZ3hnUjEvR2FEak13eWo3eWIwN0xYbmljSzlxK0ZCVnkxcnN5RGJGczhwSkJ6Skk4eEU5RjNnUjlnNytGRG1lTFRZMk9iQ21nQjNRelpDV3RrR3FHc3JVVnBJVHpyM3p0M08rODYxbS9XN3JubGdGeEhYRVlhYXRCOS96YjNIKy9mZkthUHRJcUpycVpmYUxoeWhJaWRnOFpTWmE0a3dpd3JmNGNtWHdNZGJJQTBPRCtDK1lSRnlZUlMyRytDaXNnZCs3RDU3RHd2Ykt1UEQzSUtxMGljR29RM1hiNHRodDBZT1YzZjdvUjliUWxiRlVvbENRMUd3SEFLSFF0NGdUSzJEUnNvaGthTWpMRXpqc0gzclNYTnlsWHZ6WnZnOW0zYnp3eHF1aXJUMExaZk9ESUtVdlNDQStadkl5TEN6TC9lUE5vdkFLZEhOYmNiVzBBOGd0MG1TMXJyZnVhUFQwSC94ams5M1J4MVNtVDBZQkJUaWlCaXJQWEVnbi91ZjNWOTA1UWIyaVREVXhXaVU2bjBVQzYrREdDY05HQmV1SWxFQWlmRWs1Uk1OUGNieFlFQWdoREFjQXcwRWN1SjBnVE9PZUNzU0o3eEU1WHJ6Uk9ZTndLYmVVMEtyNEtkRnd5YXdSbVJyN1M0akhWQ3diRlFpb3FkaFdSN2VGcExhWFM1MkNGbjNwRzBIN1ZpaEt5SDlZQlJPYTJHTmFHYlI0bUMvdWxVaW92YlVERmxKNlJibWlvSUg5NUJwNklzMkMzQUp0ZzdTeUV5a0ppQ3JYU1NzNSs0NmZiTGZ1ZVMvd2hocHhBTzZqWEora0JRU3NTT0ZRMVo3Kzc3K0lVMjk5aThoa0JIYVV5VEpRMVVSdVhRYkpJdUt0TTRJMEd1UXJBZEEzb1lLbnFHTGh1Z0FMYnFxdjFBaFhIVU5ocjVlZEUrSnpnQXZRQTVZQlhpcjhXeWgzK1NrWWh4SmJlNmJpZzB2Z0dhNCtDZzJxMUN0SGJBcmtybDJqcDJIekFsRmw5aHgzc1dUdU9lSTAwSENld3o0dGxHQ0ZoeGllM2F5eXdDcTRaUjlVSkxFc1Z5anlJTUNjRm1aaU9ablRsWmErMDBLQlVSNEZ0NXVVOFFxcXlTclF6Vmt0YlU4NDg3OU9idHo3NWg4QkRpL2tmQi90K21RVU1DNWFVYUVnM2YvRzcwZ1gzcldGYWsxRU0yUmx1OHU5dFFVWUtTazJ6d0RUODBTcXZRTTBnYWovQkVDbjE2R1Z4VFYwbzhYZzYrejlUT2NVRWRSQk1XN2xOWkNnSm1pMmJTc0tGMVpjRHdZMGcybEN6cVZ1encrY0REVXZodGhQVVBJK1FsSXlpcXM0d0FTSWVWcGJrUmtGaHJFYXROUkI1cUVzUXJrQzU0bXQ3d2llRUFaVHNrZXc5SHh5YjdWY2FhQUZEMnRrUjZMZlNzOEJoM1ZqZWE2Tmc1UWQyaU1ydzhpL085OEl0YUNqblFwQVNhczdDTFpndlpycjJVT3BldWovL1g0QUR3TzNDeDZCOTd2NCt2TW9JT3dCdjF3MXpmcUgwN20zL2RPSzZ3blZKSXNrWUhrN21MdkFEN2ExMS9vSm9Oc0NVdDBMMVNwV1dKUng2SEtIbUM3Y0N3Y2ZJQWF1cVNqY3dKMVJDcWNZT05KZTUwRFI0N3RtM2ZvUTBrRmNlTTk0VVF6ZktXSzMxU2ptS3U4TGJKdG9MdlN4NUlnTEN6TnNkbzlkYWNTOUxTT1RnWXMydHlnWVVDRGhGUWlITTI0eHZXTXlZQlpNRHowNEdlVGhBdzhQQzRnTFJpeU1nMmNmQWJBVmRJMmduYUZhMmNOVWo0Q3NlVlF2MWUzSnQvK1BqYlB2L1k4NGRPaFJhWmVMcnordkFnSzRXWEhvMXFwNTRQZCtWRTcrd1I5aHRHTkVwY3BvRlpoc0I5VVBPNHdoT252Y1psRGx4TGc0RUFwNmdxWHN2S2RWWkVGSXR1dFc3K0RpUFJZOEdZTlFaci9QWml1bTdBbHVvSGRwcGVBMUd0c3ZJUFdYK3hvcGd2RHp4bFhWbkhyeW5KWHB0cW1lQ0gzbmhmRGdQVVZTZnVJdFJYRzc0UVlSNzlQY2ExbjFBUFNGMmc1SndndUFVYmZYTXd2RFlLZFhPcGU3RHFKd29wY1BVVFkxSnhVcG1ZSG9HcUtiVWlhcmxPVWRncTZqalBlUVo5NjVqbnQrNlRzaEtlUG8wY0VESC90NkFoUVF4TkVQRUNUYnUvN29iNlZ6N3o0ajQ5MlFQR2RhM2tuVUs1VGNHaGNyTkhMNjFMMDJjNE84SHE0L0tETnp3UGxwOEhVRnRBellmUjhjSzJyQW9sVUx1UllsYzRWM0RpelNWZzRIWXJsdlliNEFRQytxMm1PUm1tM1RsTXpUSmNTcWVzREthSWFwWG84RXNnMThXUmVxZzM0S2dDakVLQWFZS0RRTnZaL0kxczh5TXdvWUdFeWt3Smowd0llRHllamZDMWdURHd6aXVYaWFuamFUVkZtSHVpblF6czJhcis0WHlRMVJMZVUwZTZEdVR2enU5emJZdUIzODZxQTdIdGYxUkNnZ2dDT1dJY0g3UDFROS9OdHZxUExKV2taTExUaFBzbTJ2TUZXZ3RzYmRKU0YwRHB5OHl4UGlxZEN6eUFwZnQrQkJCWG9pRms0c0IzMVFoQjkvUjFXTHYyWUNaWEVwL1ZFUnZqZWJNZ0x6c0NCMDNqRDBKVEs3M3NuaVlzc2VRb1JRWTIyY3VxTzMxQWRJaURLMk9yWHZrcFpaaURJeGlXaWRnejVGN3BjQlNWaEt5TXdyRENaZkNjNkFPRElnZ2dXTlVoWkt5YkZyWUVESGc5QmM4Q1k3LzE1Z2JRTGFRa2FWb0VwQW5oUGRUSkFicHRVRHRBUlduYXVxR3ZINGI3NjVPLzBIUDRNYkQ5ZVAxL1hHOVFRcElBQWN6Ymp4Y0QwLytkYWZyQi8rblZ0VHRUUUJSMDBsZ3JUdFFIL01xaW9nTmRITmlGUEg2TllpWm5VL1VBQ0xTekZqd0Y3NVF0RXd3RTVSbE9CYUd3UFVGektJRHlyY1phTzRuNEo5Q0dpcEJ6TW1jdUZNRDN1cVJIdGtzTkJORTIwRkk4b21sQ0hoZ2hGbGlGL2oxL0p2Ny9yS292SUJMdXZOcVc5cVZJS0hSYmNhVVROOFd1Zmh1U3MrZ1RsUWVzMDkzSW1KQVZ2amsxYTNTVFVhRTdrUnppNEltaG5UNmo1VWsxV2h0amt0N2F6bDVPL2V1L1RRcjMwclNDbUZ5NS9BOVFRcUlJRGJqbVNRVXQzenhyOG5KLy9IdldscDEwU1ZIVVpMSW11WDJQSk5jVUJiallCMlJwdzhabFFOb2xUTUdXZ1MvY253UTNvRi92c1FYY2VMR0E1RzVDM2R0TVVBdVNQdVNjbWVtblA5WW5oY29wUXpjNWppTnlWelV3ZVUvV3VUZUNXRDJ1OXVTeVVGbkl5SWhqM0gzWHZCSVMvcFdZZDRIZEYyNlZmTk1TWWZmYitYZU4ydHBFZGlZZlY3U09OZFZPZGkyVTlDbzcrQXRrVzlZNmVreVlqTXJYRHJ2S0taVWxaMklhM3NoSGFOWXJ3UDFhay96TTJ4WC96bTg1Qno1Z0Vmdit1TjY0bFZRSUNRbTlNRzVCUWYrTlhYeUptM1gwaWp0UXJkTEtmSktyRnRIOUhPZkFaMmdxb1N6QzhBSis4RTZoR2N1clYvQzNWUUxCTmNtdjJBRklFT1pyWHAyY0FTQk41akFQZlVseUVKeXV1S09OQlFTclJoaVF3NktRaEI3T0hNV0o5dnh5TFpYZnRUcnVHR3lVdi9vaVRmNTRQNE5xNjlGU3ZFZVZqbUNJalk5MHNMcXpFRWw3M2xSd1JuZEM1UUUyS3J0K0lkZlBzTjVzQy8vV1FYQW0wcm9pM0h1M2V6R28rWnVnNTY3Z1F3dXlBWUxhTmEydzlwdHpTTlZuWFUzbHZ4K08vOTdaelAvLzVqRlJ0OHZPdUpWa0JZUTk1UXQxdjN2RWVQLzk3WHBObmRxQ1k3TXZJbVpYVVBzTEtMYURadE5ISUhWR05pZWhZNDhSRkJQYUpsUkhKZmVGb0dJWWNRK3pSY3JGME5xeGR1dm5DQ2lwNzNBajBmREVTMDF3Y2dnS3F5NERLN3pQbEwwUjRDb2xFTWFBdDlYRUdkakhialNVVXNGQTY2em5RdmlVOHE1Ly9DdW9leWxRMGZBL01PK29PQ1k3V2ZiSzYwdlNYVG9zUkdNNm50WEZId24vOG83Q0JLZFV1cndHekdxa3BZT3JCZnF0RW9KVlYwcHg0U05odVVlb1I2KzZVaXpJSnFvclZzamZqUXIvL0U3TndmL21mZ3hrOFk5dzJ2SjBFQkFlQklCOXhZNTdQLzgzZlN3Nzk5Uzhyclk2bjNkc2h6WU1mbGd1WGRRRHUxRFgxVUUrb2xZUE1NOGRDSGpDTkUzVWV6REk4WjBWcjh3REZOV0RPUFhndEZNY0JhVU5oQ2o5d0Q3SDU1cU5zNlRiQUNFSmFOM3p5TFVtSVV1QzRWZ2toZ0pjU3hYdEhYSWNUaEVJN25vaVRmQTFMTHkramcyYjF5bVpVQzJMZlR1eEprZlFqbFk2eStDeXdVQ2x6c2J4UWM5QlZGZGg5VllqYkRlRzFGVmk3YkQwa1ZSSVR6NDNkRFordEFxcEYyWEFtcEVnaTA5WGhwaEFkLzYzZG5ELzdXOTFyUWNkc25yWHpBazZhQUFIQmJ4ZzNmTnBvOStGdi9UTzcvNWYrWU1GOWl2ZFloTjhEdVp3RExPNGgyYmx4WnpvSjZBc3pPQVE5OHdBUlZqeTFpRytUQUVRdHRBNWpINzFIVEZxdnMra1hUanZtQ0V3d1FIdTQ0ckF0TXFRYjFwUDZmQUpEd0ZXNEZ4cGxpc1hoNlFmTGtqVmhWZzBCczExMlVESGJzaGxHdW9nY0Erb0NqNXpMVmVjUGhCMHVHZ3NQdkdlNGJMaURTZ2pQUTQyWDA3d3NFYlpzd24yUDF3RjZzN05zRFprVXR4UHplTzZuVERZRUlxNTFYU0JwTndKeTdlckpubEU3Ky9udW45LzdTelVCcWNOdVJXTjczU1Y5UG9nS0NlTmZQZGpoMGF6Vjc0TmYrZHZYZ0x4NGQxenFHVksxMFU4ck9LeUNUYlVBelE5bWhxaDRCN1FaeDMzc0U4L09DeVJLc3RDdE1tdk5aOUxMK0VnbEgzcmRZaGVBRUNUQzVWU1R5SUFNZ1JOa3B5elpZQ3NxbkxPNzFyL1hGSy81dkxqb1BBRXoyT1JYYWdkZ0NKcS9qTk9WVzN5elNOTGhGdjgySWw2NXhvRGltUEQzajNQZUxKWDBYUzA4amtyZitKaFJ1MDVYVWFDM0RoRmx0eVdUY2IyTUxJeEhzZis2Vk10bStBdVNNS3JmWVBIWTc4dXk4Z0lwcTkxV294c3RNZWRxTWxuYlY2ZVRiN3RyODhKdStDcEFMOE4xdy9yeEs4bVFxSUFBUVIyOVdIS1pNNy8ydlg4OTdmK1YzcXRIU1NLUnVvWTNJbnFzZ3k5c1UzUmFSSkVFN29Cb0JJSEhmbndMbjdnUEdLd0FnMEp3d3hHaUY1b3FCR2xqQ1lqRjE4Y2RjamhSTUZDVmYvcjRnQTlyMVc3d0JvQ1JqNjZUWEVYdThPVllsQ2ZIMTdraE9Bbmw0eXRoRzJrMjJFQkpyWXZvMlkwQ25TQW1wN1dHUlFneUZHdmhrNy9jd0NFUDBYd1dxV3ZaMDZTY3MwYmJBNWhhMjc5dXVsMXgvRmNWMlFwWTgyK1NGajd3ZmViNEZDQ1R0ZlNabGFadWduZWMwWGh1blUzOXdFaC82cWE4R0hyNGI0Q2RFTmovVzlXUXJJQUFRUndRQXUvYStuMzlOT3Y2YmIwMlQ3V05KcXgyNkR0aHpUY0x5YnFEWnNIM09vSURVZ3RFcWNPSVk4T0FkUUJvQm80a0RacjhXS29MRExjWmdEaUNhT2FlK3NMbHNTeHRsVFlnSU8weGJlY0tRdVNqN1NNSlhtdnN1dGJiWkFsMU5ncFpKUFVzaU9pQVM1U0xkR2VnVHZEZ2d6dmdvQmFmb1UzZGhpZ2ZBZEVEYkJBVHhXZVZZTURZMUI0SHBMRldhdWY4NWwySG4xUWVrbVRkZ3F0Q2NPNlhySDdsZG1KV294cWoyWHN0cXZDS2NUelV0NzZpcmMzODh4ZDF2L0pvdG5IcVBCUjJQWFdMMWlWeVB2a2YwRTNzcGNFc0Mwcnk5KzJlK29tTCtiN2owdFYrQW5CcnB0a2F5NXhyTDZxOGZKOGFyZ2lDU1J5dUM5UlBBN0R4d3lYT0ExZDFFTzdYYjlja3hBS0RSQzhrUGhGQlh5Z0UxRlVBY1FNR1ZnTGswTDJheDRZNnN2dDNXaklxRkR5SnhJRlcvL0lndzBBZUI5SEN5dkJkN0V6akRKbUMvTEFCbFlSVDlwdlRGN2ZSR0Z1VmkvTlZYQ0lVeWk3dGp3Tncwc0pqdnJpQ1l0Y1IwTG11WDdNQ2VaMTRpdVdrNW44NEZWWTJOajN3WVd5Y2ZFb3lTWUx5TnN1c1pWbVhSZFYyOXZLK3V6dnpCT1I3NzFaczM1L2Y4UG5Db0FvNCthblh6SjNOOXFoUVFYaHVXSUdrOTMvUHZ2M2lFMlcvSXBhOTdaY2ZsdVRSYm83VHJLbVNwaGVmdkV4bXYrQmhrWURJeHJIYmZlNEU5VndCN3J4RXcwelpES3FCZUNvbGFUbHNibWcxRjJFSGJCSjEyYjZiZUI1QVVKQ0V6MUpPMkhTa0pMUHY3bFpXVUV1TXNOTTFQNkplT2VGV05MWnUxSUNhREhZQXVaMEJhU3U3S0NRL0RuR3VoVThwNlhsKzdJYTZnWVZmTG9pNE04V0pQdjVDbVJFcGlmWWJ4WkNSN1gzQVZKdHVYT1YvZkFxb0ptQnVjZmQrNzBHNWVBQ29SV2RxTGF2ZlZFRFlVMVZ3dDc2cmw1Ty9jUHovMmExL1ZkY2ZlNmNyM2hGbSt1QzVXd0VFVy8wbTVGTlFLU0ZzSDd2bUZWNS9JK2JlcXkxLzNoVXhyY3pZYm83VHJDbWcxSXM4ZUE2cXhaVXMwRzEwelhoS2N2Wi9ZUEVzY2VEYXdzaE5HYXJ1YktxMlBBUk56WldsZzdVQ1U3TXB3V1dZYXhMdXFBK3RtRnN1WklQR1JMOTZ2QkNjQ1dvSTRKVXRvS0lVV0pkc0dsaGxaZ2FTRWRObklRUVhLc3RXSTZrdmhaN2hkTjJNYUJhZTlPUzRUTEt5Zks2WWtVWFlVYnMyWWttTFAxWHV4NCtCT05MTU8wL09iVWkwdGNmT2g0emovMFR1VXVVdW9LOGl1SzFDdkhhRGttUkRqcmw1YXF1V2hYN3U3dmZPbnY2UURQdWhjM3hOcStlSzZXQUdmVE9XTEt3T2E3a2VhNHY1Zit0S3h6djg5THIvNUc3aTB1OFhzVkVwcmx5U09Wa1JQZmhEUUtWQXY5d0I2dEN6b0d1Sys5d0RiOWdQN3J3WEdFNnZRVURwRjR1NDNQSFJvaXUxR1pScG5qREJLUms0ZFYzbnBrUVM0eTdUVDVRVkkyVEtHaW9FbDlKSXNjZUtIWWlmakFBQ2RqQ2xRc2ZqM1dOMEc5SUVJTUZBOEdTaldnTWNqZ1RpTHlza2REVmVjeTdoeHF3RzZCdHYzNzVTOVYrOWhHaVZzWFpoQ1pJUTBHdlBzSGJkajYvajl4QWlDOFFUVjNtZWhXbDRqbTAyd1htdkhOVWE4NzQyM2I5N3pwaThGNUI3ZzVmVmpMU3I2ODE2ZlFoZThjQ21nQ1dUVGlIempwRDExcjF4NjZBZnl0bWNEMDFNNUxhMVd2UFJGNU1NZkZzeE9BK00xQUFuUVROUVZ3QnBZZnhqWU9BUHNmd2F3NnpLcmNNbHRKUDN0c3NGbXNTUldmZU51THFGc0xTWVJ5SkFRamMwcExIVlBrbjZHcjZoRlBzWk9LcURKQ2FKc0dSTW1oM0dFdVhPaGRsbWlHS1VjcXA1TE5VMG9IMUQ0djhoRFI0UStTTDMxN2hrRngwWmZteGJvT3E3dFhwSGRWMTNPMGFSQ3U5WElmRTVXb3lWc1BYeEN6bjdrdzZMVFRXS2NJT1B0a3ZaZXc1UnFZYlBGdExRblYrMkZVZnZSWDM5TDgvQ2JYZy9JaVNkYitZQlBud0lDc0pwaTROWnFmdkxtZjV4bXh6ODZ2dnpyZmh5N1AzOVhiczYybGVRNlgvcFp3UGw3d1hOM0FsSUQxWks1SUdSZ05MR0JPMzQ3Y09ZZTRNQnpCS3Q3QVczdGVDaHpjWXMrc3dRZmJwUEMyaGlCUzhrak1KTktJZ05zRkNtQlRKS2daRG1XVWYya2VLRUdEV3hFTkl3SDFBaGwzT3hsVldRbUNOUTI2eE5QeVJXYUpQQmpVQ2x3Vit5V0dRNDFJeW11c0hKOW9XRFdBVjBucXp1M1llK1ZPem5aTnVGc2F5N3I1elBTZU14Mi9RSTI3bnFmekU2Zk5PTThFa25iOTdQYWRZVks3Z0R0MnNuS0FVbWJIeDNONzNyVFR6Vm5iL3RPcy9GTVQ3YnlBWjllQlFRQUFqZG40TlpLMTIvK1Q3TTdmdlRkSzlmOHpWdmxrbGM5cCszcURyT3pGYlpmS1dsNUYzanFnOExaSmpCZVJqa3pEUUpNdGdGdEk3ajNQY1RhSG1EZnRZTGxIVUE3aCsxWE10d01DYjR0UmVUc1VuRnFaaWMxZ1RsQ0ZGRUZVMHAyckZaWnR0UHJIbUJ3THdtZ3lFUVVOMm80VzFyVXF4YlcyS0dmcGVMWXN6TGVGS2dIUkhTbEhQNGVRWWtydGRBcWkxckY4dlpsN0w3MEFKWjNUTWlzYkJzRlVvVjI4enkyUHZRQnprNCtiSmk0VWlBdEkrMTdEdXZsRmFaMmk2alg4bWk4UE9HWlA5aHM3anY2RDJZWGJ2OUo4eEx5U1ZXMmZETFhwMXNCL2JvNUF6ZU1nSGUvYit2T2YvZXl5ZlNobjZvdSs2clg2Y29sMEsySE0rcHhKUWRmVEp5L0J6eDNqOW11MFVUQVpPeCtQUUtxTWJCMURyanJUNGp0bHdqMlBnTlkybVlGRDEwRGsyZGxMcmxrMnJRUFJpd2tEZzZHOVBPVkMvT290dUl0b0p2bGhxMDBVRU0vd0lMV0JCQkRaa29pc2RCMkVkbXFSaEF5cVBrTFM0MzRlOUg5MnJ3Uk5CM0drNUhzdUh3TjIvYXVJbFVWbTRac1pvM016NTdsNWdQM3NUbDdVcEJTd2lnUjJrSzJYU3BwNzdWTXpFQzN4YlM4RjFWZW43VEhmdlYvZFEvK3duZDF3SjhBdDFhK2pPcFRvbnpBVTBZQkFlQmRMWENvZ3Z6S3lmbUR2M0pvdkg3L2Q2WExYdnZQMDU2WExPZlptWmJ0UnNLT3ExTmEyUStlL1NpNWNWcFFWZWFLbzdTOW1naHFHRDY4OERDeHRnL1lmWVZGekFvaU44bDNpRUt4Zm5aRnRZenBua2dxNUlabUVwV0lyOGtJMndrZ0NjWGNzZWRXeldnRllWZFlPV1NhZnlaRXFNcWs3b2FEeml4YkJEc1c5V1djZzFwQnUzRzIwcitsL1R1NSs4QjJxWk5ndnJtRjZabFQyRHArUnViblR0dkdVS09LV0tvTkY2WmxwUDNYTWEzc2t0UnVBRkozNCtYOUkxejRNOHp1L2MwZmE4Kzk5UWNBYVF6djNmeWt1OXlMcjZlUUFnTEEwUXhDY09qVzFCeTkrU2Z3d2ZmL3diWm5mZXVQNVgydmZFWER2Vm1iYzIycVVvMzl6NFB1MklLZU93Wk16NWdscVpiRFRRbnFpWTNheGtuZy9FUEEwcHBnMTJYRTJpWEVaTVYzOWZjeXBTUUFWWkVCVkJTa29JL0JUS2JrNnp3eUVrVlZDcnlERnloWVNiY3Y5RjVZUG1YS2FpQ1I1cDg3cHFSZ3pvSHAraWk0YkFlanZRc1dUeGRLTGFtZVFPcUtkVjBoc2NXWk80OHhYemlGOXV4cDRXd0tqRWFDeVlneVNtQ3pKV2dTWk1kVnJIWmRKUWtkTVR1WHE4bDJyVkk5cWsvOTdoM1R1OTcwWGUzOG9iZVl5NzM1Q1NlWUgrLzFGRk5BQUFCeDlPYnN4T2U3Tno3eUUxKzB0UDZSN3h2dGVjVXQzUDJpTVp1TmVXN09WekpaVHRXQno0RnVuUUxQM2dYTUx4RDFHRWdUSUhKejR5WHppdDFjY1B4MndjbVBFR3NIQk5zUEFzczdpY215dWI5MmFtc2lOQVBhUlhJaHVSWkpMQXlMOVNCZU1XL0ZXbDV1WlJGQlNSQktmMklZU1FXcEtLdmFMR0xXZmpQSUlLT044TEhGV2hWUVRvM3ZHbURqTExGMURzM21CY3hubTFKMnpSOVBJTnUyQzdVaG1pbVlSa3hybDR2c3VBSlZQUUs3TFVVYTUvSHEvaEUyNzZxYWg5LzJIelllL0tYdkEzREtQSTU0VmNTbjUzb3FLcUJmUnpOd09JRzM2RXprbjlYSC8vdi9TRmYrN1g5VFhmS0ZuNHR0bDZMZE90OUozcWhrWlJka2RZOXc4NVR3M0wzQS9Cd2hJOEZvS2R5YUlGV0tlaTFCTTNEaEllTE1BMEMxSk5pMkMxamJSMHkyQWFObGtmRllXU1ZKSU9hempMYk4wRlMyMlpCa0RBNTk1WWY0Y3VaQ21ERGNwWjliSFNubUxxdTBYWWVjS1dWaGVRREhWQnZobm9TUURtaGJZSGJCZGhhYmJ3aG1XMEM3QmRWV0lNbDJHeHVOaUdwaWs2ZWJnVzFEMUN0TU81NHBzdU1nVWowRzJrMGdNOWVUblhYaUxPWGpiL216Zk8rdmZWL1RmUFMzVExHLytrbkpiSHlpbDN6OGp6d0ZyaHNQMTdqdFNBZGdVcTk5M3Q5UGw3N3llNnRkTDlyYnBUWG83SHdIemlyVUt3SVprVnVuaE9mdkJXY1hERkRWWTl0R1dIM3JEenRLd1lLQXJnVzBCU29TNHpXcHQrOW5kMmJLdC8vT044dm52ZWlnekJSb1dqQXJwSmtyWnZOc2hkTWVWcXZ6aXduUzAzSzJYTXlBWkVWT3hvS3FHbUhuZHVBRHh6Yng0ci8rcTBpWDdoRnRPZ0tWTURmQTlBTFFiQURUZFdLK0lkREd5czRTQmF5QXFnSVNhSW9EYXpOYlFTYXd0QU5wKzJXc3R1MHhpNW0zVkZRMGpYZFVrOVNrOXVRN3p6U24zL0V2dTdQLy9kOEIyTUNoV3lzY3ZibVVWbnk2cjZlSEF0cFZRU1I3V3UzeXliNHYvTitxL1MvN085aHh3NTZPQW0wMnMwQUY5VVRBSk4xOENtdzhBRzZkc0FLR1ZCSFZKTmtDYTkvQ3k3ZkJCeEtST3l0QW5jN3d2T2ZzNVBPdk9jam52ZmdxZWVWTHI4THk4aTdadG5NTkIvZXRjT2RLR1Rpbm5PMXdoY3AvZEVCdnI4OFZEenkwanZWem02aDBoajk2ejBQeTNkLy9POFRTQ0d5YkFXZnAxRXBWSzFJbGlOeHlXVE1OSStHMXNYL1RLTW5hUWFUVlN5RkwyNUhZVXJpcEF0RTAycEhHRmFwODRRUG44OG0zdlhIcitPLzhHSUM3TGVKNmFsaTk0ZlYwVWtDL0RsV1FONGNpWGoyKzdEWC9NTzI0NFcvSmpoZXRxbFRVWmwyUkcyRWFKVlFUcUxiZzFtbHk0emd3TzJ2Wmttb0V5TWlVVWp3UkxGN0pKQUEyNXNCMEJsUU4wSFZBR3NudXkzZmdXVmRzdys3dEs2aGtCSkdhcVFKbFBKR1VBRUZpMTh4RWthRk5DN0xEcWZVcFAzTDNlWnc5MVJDem1Sbks3U3RnNVpzMFNRS1FpSlI4UW5oWUhEV0N1VFhDV3JNZ0pXQ3lFN0x0TXFTMWZVaXBZdXJtb3RwbXFaZDFOS3BIZGFxUkwzendBczc4OGM4M0Q5ejZVdzF3aDNYb2RaVnRKdlhVc0hyRDYybW9nQUFBQVE0bDRGYzhxNC9ybHk0NTlBK3c2NGEvSVRzL2F6WExHSGwyb1VPZWlpUkpySmJBcWhhMk04cldHZXJzYk1MV1dhQ2IyaFloYVV5a2tRY0ppU2xWU0NuWjdpNEtZVzdCTGhQelZ0QmtTL3VKRXl6bzJOY2FPcThqdFlVam96R3hNZ1pHSTBtcElyUVYxUTU5elVjVlJRYXdMWGc3TzNHSzlnQ01WcEpNZGdLcmV5SExPeUQxTW9TWmFPY1FNS2Rxa3FyUmFrSjdGbWw2N0l5ZWU5K2J1Z2ZlOUpORjhRNjlyc0xScDZiaXhmVjBWY0M0RW5CSWdEY0h1ZmZjMFlIWGZsdmFmY09oeWRwVmx6ZlZBZVE4emN5Ynl0eUlwRG9oVGFBWWlhaUM4M1BDNldsaWZnRnM1d0kyS0trd2lDa25FeUdWU0VwSWtoaEh1VXF3ZDVLTkdBY0ZrandsWnlrK2tscXFvZ2tDWFdRL2pBclMxcXA5bUlCNkNSZ3RNWTIzQWNzN1JaYldLT01WRVZTZ05oQnRWSEpXVm1PcHgydFZ4VGxrODE3a3Jmdi9zRDN6dnA5dnovN083d0U0QmdDTzgwcU81YWw4UGQwVk1LNEUzQ3FXVVFFQUhOaTE1NFd2MFQxLzlldXc3WnFYYy9WcXRMa0NjdTZJUnZPOEU2WlVTUm9McWtwVVkrT2RyY1IyRTJ4bllMTk9kQnRHN01iQ29GZ3NMbDVpUXdFa1M2eU1zM3hkUWpuc3owckRnalUwMmtURTNQOW9XVkN2UUNiYmdja095R1FiNjFRenBRcGtLOGdOeVVZRjBGU3RWcW11VTVJa3FUa0QzYmpuaEs3LzJXL2wwMy93YzdQWlEzOVlwUEEwVXJ5NG51NEtXQXFiL0RKRmxLL05VU2UzZ3BVWDRjREx2aTd0dU80bXJGeHhBMWFlaVl3bHFDWGlzK3BjdGN1UWhJUlVWNEtLS2pVZ0NZUUt0RFdhUTF0QU80SE9pUzRMdEhPWE9ZZmJQeFd4dzVRRDE0a2tzcG9JcWhHbG1naFNSYWxHZ21wQ1NaWFZ1VkxOQ3VxY3RiUktGVVUxZ1ZUamxGSlZWVkpEWncrQjA3dk9ZdmJnYlRqLzBWczN6OTcyTmdBUEFIQWx2em41QWRCUEc4V0w2K211Z0k5MkdVYmtyWXArYjVjUlJ0ZThZSG5YTlRkejlWazNvRDc0c21yMTByRXVIWVJpRExJRnRhRjJiWWJPSGZoREtFa00wMVZBcXZ4VWllUTdSUSt5WmtncUVtc3U2UmsxTHhTRVFsUXRsMHlsNU00VGRZbTJFM1dkVWoxSm95UytOdXNzTUh1WTBPbnR1djZCUDJuUDNmV3J6ZnJiL3hUQWZhVjdONzZodHJPY24xcFI3U2Q2L1VWVndPR1ZjT1BoaE50K3NMc0lpMTg5WHZ2ODUzTnk1VitYYlZmdGtORzJWNDFHNDB2UzZpWElrMzBnYXFoMjBKekJiSGlOdHZaUllSVmJodWY4VkNRQlFCR2txT2VyazFkYkpTQWxrVFFTUWxOS0lpS215SklTMERhUTdqUTRQMEZ0ejM2SXN6TnZ6NXYzdkw4NTgvYi9CbXpjQTJCcXpSWGcwQzlYT0hvVVQxZHI5MGpYWHdZRmpFdUF3NElia1hEYkxibWNPdGxmZTdZQisyVHZEVjlNMmZuTWF0ZXo5NmJSbnBkbTVmNk1lbHlOZDFZeTJwNmtXZ0prQ1pDcXIrc3JoYzRKeWZiL1FHeTJKSGtPNWN4Mmh1MDJGVHJ0Tk04M0lPbjliRS9mcFp0M25hNmFyZHVhMDI5Ny94dzRCV0Jqb2NtOXBmc0xvM1RENnkrVEFnNHZzWjlEZ2h1dkUreS9uamo2dFRrS0F3ZlhOZ0RMeThDb1hudk9tbTUvNW1XcGE2NG04eVZTalNhUWFrUkZuV1FrbWFLKythNklhS2M1ZDZwdGxqeHRrT1Zod2V3WXVvZFBibTdlZnhiQURNRHBqMjBTWE9IZUN1QW1CWTU4VElQK29sMS9XUlh3a1M1WHloc1RibmlPWU50QjRtMy8xQ3RFdkZidkNYbUt3MFROZ3B0dXFiRHhrT0JkWjlVdDNGOTRoYnY0K3Y4QnZtOGtBcGI3dFRrQUFBQUFTVVZPUks1Q1lJST0iIGFsdD0iRmFjZWJvb2siIGNsYXNzPSJvcHQtaWNvbi1pbWciPjwvZGl2PgogICAgPGRpdiBjbGFzcz0ib3B0LXRleHQiPjxkaXYgY2xhc3M9Im9wdC10aXRsZSI+RmFjZWJvb2s8L2Rpdj48ZGl2IGNsYXNzPSJvcHQtaGFuZGxlIj5SJmFtcDtKIEdyb29taW5nPC9kaXY+PC9kaXY+CiAgICA8c3BhbiBjbGFzcz0ib3B0LWFycm93Ij7ihpI8L3NwYW4+CiAgPC9hPgogIDxidXR0b24gY2xhc3M9Im9wdCIgb25jbGljaz0id2luZG93LmxvY2F0aW9uLmhyZWY9J3RlbDorMzcyNTg3MzU0NTYnIj4KICAgIDxkaXYgY2xhc3M9Im9wdC1pY29uIj48c3ZnIHdpZHRoPSIzMiIgaGVpZ2h0PSIzMiIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJub25lIiBzdHJva2U9InJnYmEoMjU1LDI1NSwyNTUsLjQ1KSIgc3Ryb2tlLXdpZHRoPSIxLjYiPjxwYXRoIGQ9Ik0yMiAxNi45MnYzYTIgMiAwIDAxLTIuMTggMiAxOS43OSAxOS43OSAwIDAxLTguNjMtMy4wN0ExOS41IDE5LjUgMCAwMTMuMDcgOS44MmExOS43OSAxOS43OSAwIDAxLTMuMDctOC42N0EyIDIgMCAwMTIgMWgzYTIgMiAwIDAxMiAxLjcyYy4xMjcuOTYuMzYxIDEuOTAzLjcgMi44MWEyIDIgMCAwMS0uNDUgMi4xMUw2LjkxIDguOTFhMTYgMTYgMCAwMDYgNmwxLjI3LTEuMjdhMiAyIDAgMDEyLjExLS40NWMuOTA3LjMzOSAxLjg1LjU3MyAyLjgxLjdBMiAyIDAgMDEyMiAxNi45MnoiLz48L3N2Zz48L2Rpdj4KICAgIDxkaXYgY2xhc3M9Im9wdC10ZXh0Ij48ZGl2IGNsYXNzPSJvcHQtdGl0bGUiIGRhdGEtaTE4bj0iY2FsbF91cyI+Q2FsbCBVczwvZGl2PjxkaXYgY2xhc3M9Im9wdC1oYW5kbGUiPiszNzIgNTg3IDM1NDU2PC9kaXY+PC9kaXY+CiAgICA8c3BhbiBjbGFzcz0ib3B0LWFycm93Ij7ihpI8L3NwYW4+CiAgPC9idXR0b24+CiAgPGRpdiBjbGFzcz0iaG9tZS1mb290Ij4KICAgIDxzcGFuPlRhbGxpbm48L3NwYW4+PGRpdiBjbGFzcz0iZmRvdCI+PC9kaXY+PHNwYW4+RXN0b25pYTwvc3Bhbj48ZGl2IGNsYXNzPSJmZG90Ij48L2Rpdj48c3Bhbj5BbGx2ZWVsYWV2YSA0PC9zcGFuPgogIDwvZGl2Pgo8L2Rpdj4KPC9kaXY+Cgo8IS0tIEJPT0tJTkcgLS0+CjxkaXYgY2xhc3M9InNjcmVlbiIgaWQ9ImJvb2tTY3JlZW4iPgo8ZGl2IGNsYXNzPSJjb24iPgogIDxidXR0b24gY2xhc3M9ImJhY2stYnRuIiBpZD0iYmFja0J0biIgZGF0YS1pMThuPSJiYWNrIj7ihpAg0J3QsNC30LDQtDwvYnV0dG9uPgogIDxkaXYgY2xhc3M9ImxvZ28tcmoiPlImYW1wO0o8L2Rpdj4KICA8ZGl2IGNsYXNzPSJsb2dvLXN1YiIgZGF0YS1pMThuPSJsb2dvX3N1YiI+R3Jvb21pbmcgwrcg0KLQsNC70LvQuNC9PC9kaXY+CiAgPGRpdiBjbGFzcz0icHJvZ3Jlc3MiPgogICAgPGRpdiBjbGFzcz0icHMgYWN0aXZlIiBpZD0icHMxIj48ZGl2IGNsYXNzPSJwZG90Ij48L2Rpdj48c3BhbiBkYXRhLWkxOG49InBzX3NlcnZpY2UiPtCj0YHQu9GD0LPQsDwvc3Bhbj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9InBsIiBpZD0icGwxIj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9InBzIiBpZD0icHMyIj48ZGl2IGNsYXNzPSJwZG90Ij48L2Rpdj48c3BhbiBkYXRhLWkxOG49InBzX3BldCI+0J/QuNGC0L7QvNC10YY8L3NwYW4+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJwbCIgaWQ9InBsMiI+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJwcyIgaWQ9InBzMyI+PGRpdiBjbGFzcz0icGRvdCI+PC9kaXY+PHNwYW4gZGF0YS1pMThuPSJwc19tYXN0ZXIiPtCc0LDRgdGC0LXRgDwvc3Bhbj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9InBsIiBpZD0icGwzIj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9InBzIiBpZD0icHM0Ij48ZGl2IGNsYXNzPSJwZG90Ij48L2Rpdj48c3BhbiBkYXRhLWkxOG49InBzX2RhdGUiPtCU0LDRgtCwPC9zcGFuPjwvZGl2PgogICAgPGRpdiBjbGFzcz0icGwiIGlkPSJwbDQiPjwvZGl2PgogICAgPGRpdiBjbGFzcz0icHMiIGlkPSJwczUiPjxkaXYgY2xhc3M9InBkb3QiPjwvZGl2PjxzcGFuIGRhdGEtaTE4bj0icHNfZGV0YWlscyI+0JTQsNC90L3Ri9C1PC9zcGFuPjwvZGl2PgogIDwvZGl2PgoKICA8IS0tIFN0ZXAgMSAtLT4KICA8ZGl2IGNsYXNzPSJzdGVwIHNob3ciIGlkPSJiazEiPgogICAgPGRpdiBjbGFzcz0ic2xibCIgZGF0YS1pMThuPSJzdGVwMV9sYmwiPjAxIMK3INCf0L7RgNC+0LTQsDwvZGl2PgogICAgPGRpdiBjbGFzcz0iYndyYXAiPgogICAgICA8ZGl2IGNsYXNzPSJzYm94Ij4KICAgICAgICA8c3BhbiBjbGFzcz0ic2kiPvCflI08L3NwYW4+CiAgICAgICAgPGlucHV0IGlkPSJiSW5wdXQiIHR5cGU9InRleHQiIHBsYWNlaG9sZGVyPSLQndCw0YfQvdC40YLQtSDQstCy0L7QtNC40YLRjCDQv9C+0YDQvtC00YMuLi4iIGRhdGEtaTE4bi1waD0iYnJlZWRfcGgiIGF1dG9jb21wbGV0ZT0ib2ZmIj4KICAgICAgICA8YnV0dG9uIGNsYXNzPSJjbHIiIGlkPSJjbHJCdG4iPuKclTwvYnV0dG9uPgogICAgICA8L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iZHJvcCIgaWQ9ImJEcm9wIj48L2Rpdj4KICAgIDwvZGl2PgogICAgPGRpdiBjbGFzcz0ic2JhZGdlIiBpZD0ic0JhZGdlIj48L2Rpdj4KICAgIDxkaXYgaWQ9InN2Y1NlYyIgc3R5bGU9ImRpc3BsYXk6bm9uZTttYXJnaW4tdG9wOjE2cHgiPgogICAgICA8ZGl2IGNsYXNzPSJzbGJsIiBpZD0ic3RlcDJMYmxFbCIgZGF0YS1pMThuPSJzdGVwMl9sYmwiPjAyIMK3INCj0YHQu9GD0LPQsDwvZGl2PgogICAgICA8ZGl2IGlkPSJzdmNMaXN0Ij48L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2PgoKICA8IS0tIFN0ZXAgMiAtLT4KICA8ZGl2IGNsYXNzPSJzdGVwIiBpZD0iYmsyIj4KICAgIDxkaXYgY2xhc3M9InNsYmwiIGRhdGEtaTE4bj0ic3RlcDNfbGJsIj7QmtCw0Log0LTQsNCy0L3QviDQstGLINC/0L7RgdC10YnQsNC70Lgg0LPRgNGD0LzQuNC90LM/PC9kaXY+CiAgICA8YnV0dG9uIGNsYXNzPSJnYnRuIiBkYXRhLXZhbD0i0J/QtdGA0LLRi9C5INGA0LDQtyIgZGF0YS1pMThuPSJnMSI+0J/QtdGA0LLRi9C5INGA0LDQtzwvYnV0dG9uPgogICAgPGJ1dHRvbiBjbGFzcz0iZ2J0biIgZGF0YS12YWw9ItCe0YIgMSDQtNC+IDMg0LzQtdGB0Y/RhtC10LIiIGRhdGEtaTE4bj0iZzIiPtCe0YIgMSDQtNC+IDMg0LzQtdGB0Y/RhtC10LI8L2J1dHRvbj4KICAgIDxidXR0b24gY2xhc3M9ImdidG4iIGRhdGEtdmFsPSLQntGCIDMg0LTQviA2INC80LXRgdGP0YbQtdCyIiBkYXRhLWkxOG49ImczIj7QntGCIDMg0LTQviA2INC80LXRgdGP0YbQtdCyPC9idXR0b24+CiAgICA8YnV0dG9uIGNsYXNzPSJnYnRuIiBkYXRhLXZhbD0i0JHQvtC70LXQtSA2INC80LXRgdGP0YbQtdCyIiBkYXRhLWkxOG49Imc0Ij7QkdC+0LvQtdC1IDYg0LzQtdGB0Y/RhtC10LI8L2J1dHRvbj4KICA8L2Rpdj4KCiAgPCEtLSBTdGVwIDMgLS0+CiAgPGRpdiBjbGFzcz0ic3RlcCIgaWQ9ImJrMyI+CiAgICA8ZGl2IGNsYXNzPSJzbGJsIiBkYXRhLWkxOG49InN0ZXAyX21hc3RlciI+0JLRi9Cx0LXRgNC40YLQtSDQvNCw0YHRgtC10YDQsDwvZGl2PgogICAgPGRpdiBjbGFzcz0ibWFzdGVycyI+CiAgICAgIDxkaXYgY2xhc3M9Im1idG4iIGRhdGEtbWFzdGVyPSLQotCw0YLRjNGP0L3QsCI+PGRpdiBjbGFzcz0ibW5hbWUiPtCi0LDRgtGM0Y/QvdCwPC9kaXY+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9Im1idG4iIGRhdGEtbWFzdGVyPSLQkNC70LXQutGB0LDQvdC00YDQsCI+PGRpdiBjbGFzcz0ibW5hbWUiPtCQ0LvQtdC60YHQsNC90LTRgNCwPC9kaXY+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9Im1idG4iIGRhdGEtbWFzdGVyPSLQmtGB0LXQvdC40Y8iPjxkaXYgY2xhc3M9Im1uYW1lIj7QmtGB0LXQvdC40Y88L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0ibWJ0biIgZGF0YS1tYXN0ZXI9ItCQ0L3QvdCwIj48ZGl2IGNsYXNzPSJtbmFtZSI+0JDQvdC90LA8L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0ibWJ0biIgZGF0YS1tYXN0ZXI9ItCQ0LvQuNGB0LAiPjxkaXYgY2xhc3M9Im1uYW1lIj7QkNC70LjRgdCwPC9kaXY+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9Im1idG4iIGRhdGEtbWFzdGVyPSLQmtGA0LjRgdGC0LjQvdCwIj48ZGl2IGNsYXNzPSJtbmFtZSI+0JrRgNC40YHRgtC40L3QsDwvZGl2PjwvZGl2PgogICAgPC9kaXY+CiAgPC9kaXY+CgogIDwhLS0gU3RlcCA0IC0tPgogIDxkaXYgY2xhc3M9InN0ZXAiIGlkPSJiazQiPgogICAgPGRpdiBjbGFzcz0ic2xibCIgZGF0YS1pMThuPSJzdGVwNF9sYmwiPtCS0YvQsdC10YDQuNGC0LUg0LTQsNGC0YM8L2Rpdj4KICAgIDxkaXYgY2xhc3M9ImNhbC1oIj4KICAgICAgPGJ1dHRvbiBjbGFzcz0iY2FsLW4iIGlkPSJwcmV2TSI+JiM4MjQ5OzwvYnV0dG9uPgogICAgICA8ZGl2IGNsYXNzPSJjYWwtbSIgaWQ9ImNhbE0iPjwvZGl2PgogICAgICA8YnV0dG9uIGNsYXNzPSJjYWwtbiIgaWQ9Im5leHRNIj4mIzgyNTA7PC9idXR0b24+CiAgICA8L2Rpdj4KICAgIDxkaXYgY2xhc3M9ImNhbC13cmFwIiBpZD0iY2FsV3JhcCI+CiAgICAgIDxkaXYgY2xhc3M9ImNnIiBpZD0iY2FsRyI+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImNhbC1sb2FkaW5nIiBpZD0iY2FsTG9hZGluZyI+CiAgICAgICAgPGRpdiBjbGFzcz0iY2FsLXNwaW5uZXIiPjwvZGl2PgogICAgICAgIDxkaXYgY2xhc3M9ImNhbC1sb2FkaW5nLXRleHQiIGRhdGEtaTE4bj0iY2FsX2xvYWRpbmciPtCX0LDQs9GA0YPQttCw0LXQvCDRgdCy0L7QsdC+0LTQvdGL0LUg0LTQvdC4Li4uPC9kaXY+CiAgICAgIDwvZGl2PgogICAgPC9kaXY+CiAgICA8ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7Z2FwOjIwcHg7YWxpZ24taXRlbXM6Y2VudGVyO21hcmdpbi10b3A6MTJweDtwYWRkaW5nLXRvcDoxMnB4O2JvcmRlci10b3A6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjEpO2ZsZXgtd3JhcDp3cmFwOyI+PGRpdiBzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6OHB4OyI+PGRpdiBzdHlsZT0id2lkdGg6MTZweDtoZWlnaHQ6MTZweDtib3JkZXItcmFkaXVzOjUwJTtiYWNrZ3JvdW5kOnJnYmEoOTAsMTgwLDkwLC4xNSk7Ym9yZGVyOjFweCBzb2xpZCAjNWFiNDVhO2ZsZXgtc2hyaW5rOjA7Ij48L2Rpdj48c3BhbiBzdHlsZT0iZm9udC1zaXplOjFyZW07Y29sb3I6I2ZmZmZmZjtsZXR0ZXItc3BhY2luZzouMDNlbTsiIGRhdGEtaTE4bj0iY2FsX2F2YWlsIj7QldGB0YLRjCDRgdCy0L7QsdC+0LTQvdC+0LUg0LLRgNC10LzRjzwvc3Bhbj48L2Rpdj48ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo4cHg7Ij48ZGl2IHN0eWxlPSJ3aWR0aDoxNnB4O2hlaWdodDoxNnB4O2JvcmRlci1yYWRpdXM6NTAlO2JhY2tncm91bmQ6cmdiYSgyNTUsMjU1LDI1NSwuMDQpO2JvcmRlcjoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMSk7ZmxleC1zaHJpbms6MDsiPjwvZGl2PjxzcGFuIHN0eWxlPSJmb250LXNpemU6MXJlbTtjb2xvcjojZmZmZmZmO2xldHRlci1zcGFjaW5nOi4wM2VtOyIgZGF0YS1pMThuPSJjYWxfbm9uZSI+0KHQstC+0LHQvtC00L3QvtCz0L4g0LLRgNC10LzQtdC90Lgg0L3QtdGCPC9zcGFuPjwvZGl2PjwvZGl2PgogICAgPGRpdiBpZD0idGltZVNlYyIgc3R5bGU9ImRpc3BsYXk6bm9uZTttYXJnaW4tdG9wOjE2cHgiPgogICAgICA8ZGl2IGNsYXNzPSJzbGJsIiBkYXRhLWkxOG49InN0ZXA0X3RpbWUiPtCS0YvQsdC10YDQuNGC0LUg0LLRgNC10LzRjzwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJ0ZyIgaWQ9InRpbWVHIj48L2Rpdj4KICAgIDwvZGl2PgogICAgPGRpdiBzdHlsZT0ibWFyZ2luLXRvcDoyMHB4O3BhZGRpbmctdG9wOjE2cHg7Ym9yZGVyLXRvcDoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDgpO3RleHQtYWxpZ246Y2VudGVyIj4KICAgICAgPGJ1dHRvbiBpZD0iY2FsbGJhY2tCdG4iIGNsYXNzPSJjYmstYnRuIj7QndC1INC90LDRiNC70Lgg0YPQtNC+0LHQvdC+0LUg0LLRgNC10LzRjz8g4oaSPC9idXR0b24+CiAgICA8L2Rpdj4KICA8L2Rpdj4KCiAgPCEtLSBTdGVwIDUgLS0+CiAgPGRpdiBjbGFzcz0ic3RlcCIgaWQ9ImJrNSI+CiAgICA8ZGl2IGNsYXNzPSJzbGJsIiBkYXRhLWkxOG49InN0ZXA1X2xibCI+0JLQsNGI0Lgg0LTQsNC90L3Ri9C1PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJmZyI+PGxhYmVsIGNsYXNzPSJmbCIgZGF0YS1pMThuPSJsYmxfbmFtZSI+0JjQvNGPPC9sYWJlbD48aW5wdXQgY2xhc3M9ImZpIiBpZD0iY05hbWUiIHR5cGU9InRleHQiIHBsYWNlaG9sZGVyPSLQktCw0YjQtSDQuNC80Y8iIGRhdGEtaTE4bi1waD0icGhfbmFtZSI+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJmZyI+PGxhYmVsIGNsYXNzPSJmbCIgZGF0YS1pMThuPSJsYmxfcGhvbmUiPtCi0LXQu9C10YTQvtC9PC9sYWJlbD48aW5wdXQgY2xhc3M9ImZpIiBpZD0iY1Bob25lIiB0eXBlPSJ0ZWwiIHBsYWNlaG9sZGVyPSIrMzcyIC4uLiI+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJmZyI+PGxhYmVsIGNsYXNzPSJmbCIgZGF0YS1pMThuPSJsYmxfZW1haWwiPkVtYWlsPC9sYWJlbD48aW5wdXQgY2xhc3M9ImZpIiBpZD0iY0VtYWlsIiB0eXBlPSJlbWFpbCIgcGxhY2Vob2xkZXI9ImVtYWlsQGV4YW1wbGUuY29tIj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9ImZnIj48bGFiZWwgY2xhc3M9ImZsIiBkYXRhLWkxOG49ImxibF9wZXQiPtCa0LvQuNGH0LrQsCDQv9C40YLQvtC80YbQsDwvbGFiZWw+PGlucHV0IGNsYXNzPSJmaSIgaWQ9ImNQZXQiIHR5cGU9InRleHQiIHBsYWNlaG9sZGVyPSLQndC10L7QsdGP0LfQsNGC0LXQu9GM0L3QviIgZGF0YS1pMThuLXBoPSJwaF9vcHRpb25hbCI+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzdW0iIGlkPSJzdW1CbG9jayI+PC9kaXY+CiAgICA8YnV0dG9uIGNsYXNzPSJjYnRuIiBpZD0iY29uZmlybUJ0biIgZGF0YS1pMThuPSJjb25maXJtX2J0biI+0J/QvtC00YLQstC10YDQtNC40YLRjCDQt9Cw0L/QuNGB0Yw8L2J1dHRvbj4KICA8L2Rpdj4KCiAgPCEtLSBTdWNjZXNzIC0tPgogIDxkaXYgY2xhc3M9InNibG9jayIgaWQ9InN1Y0Jsb2NrIj4KICAgIDxkaXYgY2xhc3M9InNpMiI+8J+QvjwvZGl2PgogICAgPGRpdiBjbGFzcz0ic3QiIGRhdGEtaTE4bj0ic3VjY2Vzc190aXRsZSI+0JfQsNC/0LjRgdGMINC/0YDQuNC90Y/RgtCwITwvZGl2PgogICAgPGRpdiBjbGFzcz0ic3MiIGRhdGEtaTE4bj0ic3VjY2Vzc19zdWIiPtCc0Ysg0YHQstGP0LbQtdC80YHRjyDRgSDQstCw0LzQuCDQtNC70Y8g0L/QvtC00YLQstC10YDQttC00LXQvdC40Y8uPGJyPtCh0L/QsNGB0LjQsdC+LCDRh9GC0L4g0LLRi9Cx0YDQsNC70LggUiZKIEdyb29taW5nITwvZGl2PgogICAgPGJ1dHRvbiBjbGFzcz0iaGJ0biIgaWQ9ImhvbWVCdG4iIGRhdGEtaTE4bj0idG9faG9tZSI+4oaQINCd0LAg0LPQu9Cw0LLQvdGD0Y48L2J1dHRvbj4KICA8L2Rpdj4KPC9kaXY+CjwvZGl2PgoKPGRpdiBpZD0iY2JrTW9kYWwiIHN0eWxlPSJkaXNwbGF5Om5vbmU7cG9zaXRpb246Zml4ZWQ7aW5zZXQ6MDtiYWNrZ3JvdW5kOnJnYmEoMCwwLDAsLjc1KTt6LWluZGV4OjMwMDthbGlnbi1pdGVtczpjZW50ZXI7anVzdGlmeS1jb250ZW50OmNlbnRlcjtwYWRkaW5nOjIwcHgiPgogIDxkaXYgc3R5bGU9ImJhY2tncm91bmQ6IzBhMGEwYTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjEyKTtib3JkZXItdG9wOjFweCBzb2xpZCAjZmZmZmZmO3BhZGRpbmc6MjhweCAyNHB4O3dpZHRoOjEwMCU7bWF4LXdpZHRoOjM2MHB4Ij4KICAgIDxkaXYgc3R5bGU9ImZvbnQtc2l6ZTowLjgzOHJlbTtsZXR0ZXItc3BhY2luZzouMmVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtjb2xvcjojZmZmZmZmO21hcmdpbi1ib3R0b206MTZweDtmb250LXdlaWdodDo2MDA7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWYiPtCe0LHRgNCw0YLQvdGL0Lkg0LfQstC+0L3QvtC6PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJmZyI+PGxhYmVsIGNsYXNzPSJmbCI+0JjQvNGPPC9sYWJlbD48aW5wdXQgY2xhc3M9ImZpIiBpZD0iY2JrTmFtZSIgdHlwZT0idGV4dCIgcGxhY2Vob2xkZXI9ItCS0LDRiNC1INC40LzRjyI+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJmZyI+CiAgICAgIDxsYWJlbCBjbGFzcz0iZmwiPtCi0LXQu9C10YTQvtC9PC9sYWJlbD4KICAgICAgPGRpdiBzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOnN0cmV0Y2g7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMTUpIj4KICAgICAgICA8c3BhbiBzdHlsZT0icGFkZGluZzoxMHB4IDEwcHggMTBweCAwO2NvbG9yOiNmZmZmZmY7Zm9udC1zaXplOjEuMzYzcmVtO2JvcmRlci1yaWdodDoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMSk7bWFyZ2luLXJpZ2h0OjEwcHg7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWYiPiszNzI8L3NwYW4+CiAgICAgICAgPGlucHV0IGlkPSJjYmtQaG9uZSIgdHlwZT0idGVsIiBwbGFjZWhvbGRlcj0iWFhYWFhYWFgiIHN0eWxlPSJmbGV4OjE7YmFja2dyb3VuZDp0cmFuc3BhcmVudDtib3JkZXI6bm9uZTtvdXRsaW5lOm5vbmU7Zm9udC1mYW1pbHk6J1BsYXlmYWlyIERpc3BsYXknLHNlcmlmO2ZvbnQtc2l6ZToxLjQzOHJlbTtjb2xvcjojZmZmZmZmO3BhZGRpbmc6MTBweCAwIj4KICAgICAgPC9kaXY+CiAgICA8L2Rpdj4KICAgIDxkaXYgaWQ9ImNia1N1Y2Nlc3MiIHN0eWxlPSJkaXNwbGF5Om5vbmU7dGV4dC1hbGlnbjpjZW50ZXI7cGFkZGluZzoyMHB4IDAiPgogICAgICA8ZGl2IHN0eWxlPSJmb250LXNpemU6Mi44NzVyZW07bWFyZ2luLWJvdHRvbToxMHB4O29wYWNpdHk6LjUiPuKckzwvZGl2PgogICAgICA8ZGl2IHN0eWxlPSJmb250LWZhbWlseTonUGxheWZhaXIgRGlzcGxheScsc2VyaWY7Zm9udC1zaXplOjEuODc1cmVtO2NvbG9yOiNmZmZmZmY7bWFyZ2luLWJvdHRvbTo2cHgiPtCX0LDRj9Cy0LrQsCDQv9GA0LjQvdGP0YLQsCE8L2Rpdj4KICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjEuMDM3cmVtO2NvbG9yOiNmZmZmZmY7bGluZS1oZWlnaHQ6MS42O2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmIj7QnNGLINC/0LXRgNC10LfQstC+0L3QuNC8INCy0LDQvCDQsiDQsdC70LjQttCw0LnRiNC10LUg0LLRgNC10LzRjzwvZGl2PgogICAgPC9kaXY+CiAgICA8YnV0dG9uIGlkPSJjYmtTdWJtaXQiIGNsYXNzPSJjYnRuIiBzdHlsZT0ibWFyZ2luLXRvcDoxNHB4Ij7QntGC0L/RgNCw0LLQuNGC0Yw8L2J1dHRvbj4KICAgIDxidXR0b24gaWQ9ImNia0Nsb3NlIiBzdHlsZT0iZGlzcGxheTpibG9jazt3aWR0aDoxMDAlO21hcmdpbi10b3A6OHB4O2JhY2tncm91bmQ6bm9uZTtib3JkZXI6bm9uZTtjb2xvcjojZmZmZmZmO2ZvbnQtc2l6ZTowLjgzOHJlbTtsZXR0ZXItc3BhY2luZzouMTJlbTtjdXJzb3I6cG9pbnRlcjtwYWRkaW5nOjhweDtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZiI+0J7RgtC80LXQvdCwPC9idXR0b24+CiAgPC9kaXY+CjwvZGl2PgoKPHNjcmlwdD4KdmFyIERBVEEgPSBbeyJicmVlZCI6ItCQ0LLRgdGC0YDQsNC70LjQudGB0LrQsNGPINC+0LLRh9Cw0YDQutCwIDE14oCTMjUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjYwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo4MCwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjg1fSwiYnJlZWRfZW4iOiJBdXN0cmFsaWFuIFNoZXBoZXJkIDE14oCTMjUga2ciLCJicmVlZF9ldCI6IkF1c3RyYWFsaWEgbGFtYmFrb2VyIDE14oCTMjUga2cifSx7ImJyZWVkIjoi0JDQstGB0YLRgNCw0LvQuNC50YHQutCw0Y8g0L7QstGH0LDRgNC60LAgMjXigJMzNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjU1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NzAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjk1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTAwfSwiYnJlZWRfZW4iOiJBdXN0cmFsaWFuIFNoZXBoZXJkIDI14oCTMzUga2ciLCJicmVlZF9ldCI6IkF1c3RyYWFsaWEgbGFtYmFrb2VyIDI14oCTMzUga2cifSx7ImJyZWVkIjoi0JDQutC40YLQsC3QuNC90YMgMjDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjYwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NzUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo5MH0sImJyZWVkX2VuIjoiQWtpdGEgSW51IDIw4oCTNDAga2ciLCJicmVlZF9ldCI6IkFraXRhIEludSAyMOKAkzQwIGtnIn0seyJicmVlZCI6ItCQ0LrQuNGC0LAt0LjQvdGDINCx0L7Qu9C10LUgNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo4NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjk1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTE1fSwiYnJlZWRfZW4iOiJBa2l0YSBJbnUgb3ZlciA0MCBrZyIsImJyZWVkX2V0IjoiQWtpdGEgSW51IMO8bGUgNDAga2cifSx7ImJyZWVkIjoi0JDQutC40YLQsC3QuNC90YMg0YTQu9Cw0YTRhNC4IDIw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjc1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6OTB9LCJicmVlZF9lbiI6IkFraXRhIEludSBmbHVmZnkgMjDigJM0MCBrZyIsImJyZWVkX2V0IjoiQWtpdGEgSW51IHBlaG1la2FydmFsaW5lIDIw4oCTNDAga2cifSx7ImJyZWVkIjoi0JDQutC40YLQsC3QuNC90YMg0YTQu9Cw0YTRhNC4INCx0L7Qu9C10LUgNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo4NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjk1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTE1fSwiYnJlZWRfZW4iOiJBa2l0YSBJbnUgZmx1ZmZ5IG92ZXIgNDAga2ciLCJicmVlZF9ldCI6IkFraXRhIEludSBwZWhtZWthcnZhbGluZSDDvGxlIDQwIGtnIn0seyJicmVlZCI6ItCQ0LvQsNCx0LDQuSA0MOKAkzYwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6ODUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo5NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjExNX0sImJyZWVkX2VuIjoiQ2VudHJhbCBBc2lhbiBTaGVwaGVyZCA0MOKAkzYwIGtnIiwiYnJlZWRfZXQiOiJLZXNrLUFhc2lhIGxhbWJha29lciA0MOKAkzYwIGtnIn0seyJicmVlZCI6ItCQ0LvQsNCx0LDQuSDQsdC+0LvQtdC1IDYwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MTAwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6MTE1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTMwfSwiYnJlZWRfZW4iOiJDZW50cmFsIEFzaWFuIFNoZXBoZXJkIG92ZXIgNjAga2ciLCJicmVlZF9ldCI6Iktlc2stQWFzaWEgbGFtYmFrb2VyIMO8bGUgNjAga2cifSx7ImJyZWVkIjoi0JDQu9GP0YHQutC40L3RgdC60LjQuSDQvNCw0LvQsNC80YPRgiAyMOKAkzQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NjAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo3NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjkwfSwiYnJlZWRfZW4iOiJBbGFza2FuIE1hbGFtdXRlIDIw4oCTNDAga2ciLCJicmVlZF9ldCI6IkFsYXNrYSBtYWxhbXV1dCAyMOKAkzQwIGtnIn0seyJicmVlZCI6ItCQ0LvRj9GB0LrQuNC90YHQutC40Lkg0LzQsNC70LDQvNGD0YIg0LHQvtC70LXQtSA0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjg1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6OTUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMTV9LCJicmVlZF9lbiI6IkFsYXNrYW4gTWFsYW11dGUgb3ZlciA0MCBrZyIsImJyZWVkX2V0IjoiQWxhc2thIG1hbGFtdXV0IMO8bGUgNDAga2cifSx7ImJyZWVkIjoi0JDQu9GP0YHQutC40L3RgdC60LjQuSDQvNCw0LvQsNC80YPRgiDRhNC70LDRhNGE0LggMjDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjYwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NzUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo5MH0sImJyZWVkX2VuIjoiQWxhc2thbiBNYWxhbXV0ZSBmbHVmZnkgMjDigJM0MCBrZyIsImJyZWVkX2V0IjoiQWxhc2thIG1hbGFtdXV0IHBlaG1la2FydmFsaW5lIDIw4oCTNDAga2cifSx7ImJyZWVkIjoi0JDQu9GP0YHQutC40L3RgdC60LjQuSDQvNCw0LvQsNC80YPRgiDRhNC70LDRhNGE0Lgg0LHQvtC70LXQtSA0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjg1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6OTUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMTV9LCJicmVlZF9lbiI6IkFsYXNrYW4gTWFsYW11dGUgZmx1ZmZ5IG92ZXIgNDAga2ciLCJicmVlZF9ldCI6IkFsYXNrYSBtYWxhbXV1dCBwZWhtZWthcnZhbGluZSDDvGxlIDQwIGtnIn0seyJicmVlZCI6ItCQ0LzQtdGA0LjQutCw0L3RgdC60LDRjyDQsNC60LjRgtCwIDIw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjc1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6OTB9LCJicmVlZF9lbiI6IkFtZXJpY2FuIEFraXRhIDIw4oCTNDAga2ciLCJicmVlZF9ldCI6IkFtZWVyaWthIEFraXRhIDIw4oCTNDAga2cifSx7ImJyZWVkIjoi0JDQvNC10YDQuNC60LDQvdGB0LrQsNGPINCw0LrQuNGC0LAg0LHQvtC70LXQtSA0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjg1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6OTUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMTV9LCJicmVlZF9lbiI6IkFtZXJpY2FuIEFraXRhIG92ZXIgNDAga2ciLCJicmVlZF9ldCI6IkFtZWVyaWthIEFraXRhIMO8bGUgNDAga2cifSx7ImJyZWVkIjoi0JDQvNC10YDQuNC60LDQvdGB0LrQsNGPINCw0LrQuNGC0LAg0YTQu9Cw0YTRhNC4IDIw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjc1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6OTB9LCJicmVlZF9lbiI6IkFtZXJpY2FuIEFraXRhIGZsdWZmeSAyMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJBbWVlcmlrYSBBa2l0YSBwZWhtZWthcnZhbGluZSAyMOKAkzQwIGtnIn0seyJicmVlZCI6ItCQ0LzQtdGA0LjQutCw0L3RgdC60LDRjyDQsNC60LjRgtCwINGE0LvQsNGE0YTQuCDQsdC+0LvQtdC1IDQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6ODUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo5NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjExNX0sImJyZWVkX2VuIjoiQW1lcmljYW4gQWtpdGEgZmx1ZmZ5IG92ZXIgNDAga2ciLCJicmVlZF9ldCI6IkFtZWVyaWthIEFraXRhIHBlaG1la2FydmFsaW5lIMO8bGUgNDAga2cifSx7ImJyZWVkIjoi0JDQvNC10YDQuNC60LDQvdGB0LrQuNC5INC60L7QutC10YAt0YHQv9Cw0L3QuNC10LvRjCAxMOKAkzE1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo1NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NzB9LCJicmVlZF9lbiI6IkFtZXJpY2FuIENvY2tlciBTcGFuaWVsIDEw4oCTMTUga2ciLCJicmVlZF9ldCI6IkFtZWVyaWthIGtva2Vyc3BhbmplbCAxMOKAkzE1IGtnIn0seyJicmVlZCI6ItCQ0LzQtdGA0LjQutCw0L3RgdC60LjQuSDQutC+0LrQtdGALdGB0L/QsNC90LjQtdC70YwgMTXigJMyMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQ1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NjUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjgwfSwiYnJlZWRfZW4iOiJBbWVyaWNhbiBDb2NrZXIgU3BhbmllbCAxNeKAkzIwIGtnIiwiYnJlZWRfZXQiOiJBbWVlcmlrYSBrb2tlcnNwYW5qZWwgMTXigJMyMCBrZyJ9LHsiYnJlZWQiOiLQkNC80LXRgNC40LrQsNC90YHQutC40Lkg0YHRgtCw0YTRhNC+0YDQtNGI0LjRgNGB0LrQuNC5INGC0LXRgNGM0LXRgCAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiQW1lcmljYW4gU3RhZmZvcmRzaGlyZSBUZXJyaWVyIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6IkFtZWVyaWthIFN0YWZmb3Jkc2hpcmUgdGVyamVyIDIw4oCTMzAga2cifSx7ImJyZWVkIjoi0JDQvNC10YDQuNC60LDQvdGB0LrQuNC5INGB0YLQsNGE0YTQvtGA0LTRiNC40YDRgdC60LjQuSDRgtC10YDRjNC10YAgMzDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjU1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NjV9LCJicmVlZF9lbiI6IkFtZXJpY2FuIFN0YWZmb3Jkc2hpcmUgVGVycmllciAzMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJBbWVlcmlrYSBTdGFmZm9yZHNoaXJlIHRlcmplciAzMOKAkzQwIGtnIn0seyJicmVlZCI6ItCQ0L3Qs9C70LjQudGB0LrQuNC5INCx0YPQu9GM0LTQvtCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQ1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo3MH0sImJyZWVkX2VuIjoiRW5nbGlzaCBCdWxsZG9nIiwiYnJlZWRfZXQiOiJJbmdsaXNlIGJ1bGRvZyJ9LHsiYnJlZWQiOiLQkNC90LPQu9C40LnRgdC60LjQuSDQutC+0LrQtdGALdGB0L/QsNC90LjQtdC70YwgMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjcwfSwiYnJlZWRfZW4iOiJFbmdsaXNoIENvY2tlciBTcGFuaWVsIDEw4oCTMTUga2ciLCJicmVlZF9ldCI6IkluZ2xpc2Uga29rZXJzcGFuamVsIDEw4oCTMTUga2cifSx7ImJyZWVkIjoi0JDQvdCz0LvQuNC50YHQutC40Lkg0LrQvtC60LXRgC3RgdC/0LDQvdC40LXQu9GMIDE14oCTMjAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjY1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo4MH0sImJyZWVkX2VuIjoiRW5nbGlzaCBDb2NrZXIgU3BhbmllbCAxNeKAkzIwIGtnIiwiYnJlZWRfZXQiOiJJbmdsaXNlIGtva2Vyc3BhbmplbCAxNeKAkzIwIGtnIn0seyJicmVlZCI6ItCQ0YTQs9Cw0L0gMjDigJMzMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjUwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NzAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjkwfSwiYnJlZWRfZW4iOiJBZmdoYW4gSG91bmQgMjDigJMzMCBrZyIsImJyZWVkX2V0IjoiQWZnYW5pc3Rhbmkga29lciAyMOKAkzMwIGtnIn0seyJicmVlZCI6ItCQ0YTQs9Cw0L0gMzDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjYwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6ODAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjEwMH0sImJyZWVkX2VuIjoiQWZnaGFuIEhvdW5kIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IkFmZ2FuaXN0YW5pIGtvZXIgMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQkdCw0YHRgdC10YIt0YXQsNGD0L3QtCAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiQmFzc2V0IEhvdW5kIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6IkJhc3NldGhvdW5kIDIw4oCTMzAga2cifSx7ImJyZWVkIjoi0JHQsNGB0YHQtdGCLdGF0LDRg9C90LQgMzDigJMzNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjU1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NjV9LCJicmVlZF9lbiI6IkJhc3NldCBIb3VuZCAzMOKAkzM1IGtnIiwiYnJlZWRfZXQiOiJCYXNzZXRob3VuZCAzMOKAkzM1IGtnIn0seyJicmVlZCI6ItCR0LXRgNC90YHQutC40Lkg0LfQtdC90L3QtdC90YXRg9C90LQgMzDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjcwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6ODUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjExMCwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjEwMH0sImJyZWVkX2VuIjoiQmVybmVzZSBNb3VudGFpbiBEb2cgMzDigJM0MCBrZyIsImJyZWVkX2V0IjoiQmVybmkgbcOkZ2lrb2VyIDMw4oCTNDAga2cifSx7ImJyZWVkIjoi0JHQtdGA0L3RgdC60LjQuSDQt9C10L3QvdC10L3RhdGD0L3QtCDQsdC+0LvQtdC1IDQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6ODUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo5NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTMwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTE1fSwiYnJlZWRfZW4iOiJCZXJuZXNlIE1vdW50YWluIERvZyBvdmVyIDQwIGtnIiwiYnJlZWRfZXQiOiJCZXJuaSBtw6RnaWtvZXIgw7xsZSA0MCBrZyJ9LHsiYnJlZWQiOiLQkdC40LLQtdGALdC50L7RgNC6INCx0L7Qu9C10LUgMyw1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NjB9LCJicmVlZF9lbiI6IkJpZXdlciBZb3Jrc2hpcmUgVGVycmllciBvdmVyIDMsNSBrZyIsImJyZWVkX2V0IjoiQmlld2VyIFlvcmtzaGlyZSBUZXJyaWVyIMO8bGUgMyw1IGtnIn0seyJicmVlZCI6ItCR0LjQstC10YAt0LnQvtGA0Log0LTQviAzLDUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiQmlld2VyIFlvcmtzaGlyZSBUZXJyaWVyIHVwIHRvIDMsNSBrZyIsImJyZWVkX2V0IjoiQmlld2VyIFlvcmtzaGlyZSBUZXJyaWVyIGt1bmkgMyw1IGtnIn0seyJicmVlZCI6ItCR0LjQs9C70YwgMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo2MH0sImJyZWVkX2VuIjoiQmVhZ2xlIDEw4oCTMTUga2ciLCJicmVlZF9ldCI6IkJpaWdlbCAxMOKAkzE1IGtnIn0seyJicmVlZCI6ItCR0LjQs9C70YwgMTXigJMyMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo2NX0sImJyZWVkX2VuIjoiQmVhZ2xlIDE14oCTMjAga2ciLCJicmVlZF9ldCI6IkJpaWdlbCAxNeKAkzIwIGtnIn0seyJicmVlZCI6ItCR0LjRiNC+0L0t0YTRgNC40LfQtSA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MH0sImJyZWVkX2VuIjoiQmljaG9uIEZyaXPDqSA14oCTMTAga2ciLCJicmVlZF9ldCI6IkJpxaFvbiBGcmlzw6kgNeKAkzEwIGtnIn0seyJicmVlZCI6ItCR0LjRiNC+0L0t0YTRgNC40LfQtSDQtNC+IDUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiQmljaG9uIEZyaXPDqSB1cCB0byA1IGtnIiwiYnJlZWRfZXQiOiJCacWhb24gRnJpc8OpIGt1bmkgNSBrZyJ9LHsiYnJlZWQiOiLQkdC+0LrRgdC10YAgMjDigJMzMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTB9LCJicmVlZF9lbiI6IkJveGVyIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6IkJva3NlciAyMOKAkzMwIGtnIn0seyJicmVlZCI6ItCR0L7QutGB0LXRgCAzMOKAkzQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NTAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo2MH0sImJyZWVkX2VuIjoiQm94ZXIgMzDigJM0MCBrZyIsImJyZWVkX2V0IjoiQm9rc2VyIDMw4oCTNDAga2cifSx7ImJyZWVkIjoi0JHQvtGA0LTQtdGALdC60L7Qu9C70LggMTXigJMyMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjUwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NzAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjkwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6ODB9LCJicmVlZF9lbiI6IkJvcmRlciBDb2xsaWUgMTXigJMyMCBrZyIsImJyZWVkX2V0IjoiQm9yZGVya29sbCAxNeKAkzIwIGtnIn0seyJicmVlZCI6ItCR0L7RgNC00LXRgC3QutC+0LvQu9C4IDIw4oCTMjUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjc1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMDAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo5MH0sImJyZWVkX2VuIjoiQm9yZGVyIENvbGxpZSAyMOKAkzI1IGtnIiwiYnJlZWRfZXQiOiJCb3JkZXJrb2xsIDIw4oCTMjUga2cifSx7ImJyZWVkIjoi0JHQvtGB0YLQvtC9LdGC0LXRgNGM0LXRgCAxMOKAkzE1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0NX0sImJyZWVkX2VuIjoiQm9zdG9uIFRlcnJpZXIgMTDigJMxNSBrZyIsImJyZWVkX2V0IjoiQm9zdG9uaSB0ZXJqZXIgMTDigJMxNSBrZyJ9LHsiYnJlZWQiOiLQkdC+0YHRgtC+0L0t0YLQtdGA0YzQtdGAIDXigJMxMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjMwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDB9LCJicmVlZF9lbiI6IkJvc3RvbiBUZXJyaWVyIDXigJMxMCBrZyIsImJyZWVkX2V0IjoiQm9zdG9uaSB0ZXJqZXIgNeKAkzEwIGtnIn0seyJicmVlZCI6ItCR0YDQsNCx0LDQvdGB0L7QvSIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NTB9LCJicmVlZF9lbiI6IkdyaWZmb24gQnJ1eGVsbG9pcyIsImJyZWVkX2V0IjoiQnLDvHNzZWxpIGdyaWZvbiJ9LHsiYnJlZWQiOiLQkdGD0LvRjNGC0LXRgNGM0LXRgCAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo1MH0sImJyZWVkX2VuIjoiQnVsbCBUZXJyaWVyIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6IkJ1bGx0ZXJqZXIgMjDigJMzMCBrZyJ9LHsiYnJlZWQiOiLQktC10LvRjNGILdC60L7RgNCz0LggMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo2NX0sImJyZWVkX2VuIjoiV2Vsc2ggQ29yZ2kgMTDigJMxNSBrZyIsImJyZWVkX2V0IjoiV2FsZXNpIGtvcmdpIDEw4oCTMTUga2cifSx7ImJyZWVkIjoi0JLQtdC70YzRiC3QutC+0YDQs9C4IDE14oCTMjAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjY1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NzV9LCJicmVlZF9lbiI6IldlbHNoIENvcmdpIDE14oCTMjAga2ciLCJicmVlZF9ldCI6IldhbGVzaSBrb3JnaSAxNeKAkzIwIGtnIn0seyJicmVlZCI6ItCS0LXRgdGCLdGF0LDQudC70LXQvdC0LdCy0LDQudGCLdGC0LXRgNGM0LXRgCIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MCwi0KLRgNC40LzQvNC40L3QsyI6NjV9LCJicmVlZF9lbiI6Ildlc3QgSGlnaGxhbmQgV2hpdGUgVGVycmllciIsImJyZWVkX2V0IjoiTMOkw6RuZS3FoG90aW1hYSB2YWxnZSB0ZXJqZXIifSx7ImJyZWVkIjoi0JLQvtGB0YLQvtGH0L3QvtGB0LjQsdC40YDRgdC60LDRjyDQu9Cw0LnQutCwIDE44oCTMjUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjYwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6ODV9LCJicmVlZF9lbiI6IkVhc3QgU2liZXJpYW4gTGFpa2EgMTjigJMyNSBrZyIsImJyZWVkX2V0IjoiSWRhLVNpYmVyaSBsYWlrYSAxOOKAkzI1IGtnIn0seyJicmVlZCI6ItCS0L7RgdGC0L7Rh9C90L7RgdC40LHQuNGA0YHQutCw0Y8g0LvQsNC50LrQsCDQsdC+0LvQtdC1IDI1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NTUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo3MCwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjEwMH0sImJyZWVkX2VuIjoiRWFzdCBTaWJlcmlhbiBMYWlrYSBvdmVyIDI1IGtnIiwiYnJlZWRfZXQiOiJJZGEtU2liZXJpIGxhaWthIMO8bGUgMjUga2cifSx7ImJyZWVkIjoi0JPQvtC70LTQtdC9LdGA0LXRgtGA0LjQstC10YAgMjDigJMzMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjYwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NzUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjEwMCwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjkwfSwiYnJlZWRfZW4iOiJHb2xkZW4gUmV0cmlldmVyIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6Ikt1bGRuZSByZXRyaWl2ZXIgMjDigJMzMCBrZyJ9LHsiYnJlZWQiOiLQk9C+0LvQtNC10L0t0YDQtdGC0YDQuNCy0LXRgCAzMOKAkzQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo4NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTEwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTAwfSwiYnJlZWRfZW4iOiJHb2xkZW4gUmV0cmlldmVyIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6Ikt1bGRuZSByZXRyaWl2ZXIgMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQk9GA0LjRhNGE0L7QvSIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MCwi0KLRgNC40LzQvNC40L3QsyI6NjV9LCJicmVlZF9lbiI6IkdyaWZmb24iLCJicmVlZF9ldCI6IkdyaWZvbiJ9LHsiYnJlZWQiOiLQlNCw0LvQvNCw0YLQuNC9Iiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQ1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo3MH0sImJyZWVkX2VuIjoiRGFsbWF0aWFuIiwiYnJlZWRfZXQiOiJEYWxtYWF0c2lhIGtvZXIifSx7ImJyZWVkIjoi0JTQttC10Lot0YDQsNGB0YHQtdC7LdGC0LXRgNGM0LXRgCDQs9C70LDQtNC60L7RiNC10YDRgdGC0L3Ri9C5Iiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjMwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo1MH0sImJyZWVkX2VuIjoiSmFjayBSdXNzZWxsIFRlcnJpZXIgc21vb3RoIiwiYnJlZWRfZXQiOiJKYWNrIFJ1c3NlbGxpIHRlcmplciBsw7xoaWthcnZhbGluZSJ9LHsiYnJlZWQiOiLQlNC20LXQui3RgNCw0YHRgdC10Lst0YLQtdGA0YzQtdGAINC20LXRgdGC0LrQvtGI0LXRgNGB0YLQvdGL0LkiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NjAsItCi0YDQuNC80LzQuNC90LMiOjY1fSwiYnJlZWRfZW4iOiJKYWNrIFJ1c3NlbGwgVGVycmllciB3aXJlLWhhaXJlZCIsImJyZWVkX2V0IjoiSmFjayBSdXNzZWxsaSB0ZXJqZXIga2FydWthcnZhbGluZSJ9LHsiYnJlZWQiOiLQlNC+0LHQtdGA0LzQsNC9IDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo1NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjY1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6ODB9LCJicmVlZF9lbiI6IkRvYmVybWFubiAzMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJEb2Jlcm1hbm4gMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQlNC+0LHQtdGA0LzQsNC9INCx0L7Qu9C10LUgNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo3MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjgwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6OTV9LCJicmVlZF9lbiI6IkRvYmVybWFubiBvdmVyIDQwIGtnIiwiYnJlZWRfZXQiOiJEb2Jlcm1hbm4gw7xsZSA0MCBrZyJ9LHsiYnJlZWQiOiLQl9Cw0L/QsNC00L3QvtGB0LjQsdC40YDRgdC60LDRjyDQu9Cw0LnQutCwIDE44oCTMjUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjYwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6ODV9LCJicmVlZF9lbiI6Ildlc3QgU2liZXJpYW4gTGFpa2EgMTjigJMyNSBrZyIsImJyZWVkX2V0IjoiTMOkw6RuZS1TaWJlcmkgbGFpa2EgMTjigJMyNSBrZyJ9LHsiYnJlZWQiOiLQl9C+0LvQvtGC0LjRgdGC0YvQuSDRgNC10YLRgNC40LLQtdGAIDIw4oCTMzAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo1MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjY1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo5MCwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjkwfSwiYnJlZWRfZW4iOiJHb2xkZW4gUmV0cmlldmVyIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6Ikt1bGRuZSByZXRyaWl2ZXIgMjDigJMzMCBrZyJ9LHsiYnJlZWQiOiLQl9C+0LvQvtGC0LjRgdGC0YvQuSDRgNC10YLRgNC40LLQtdGAIDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjgwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMTAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMTB9LCJicmVlZF9lbiI6IkdvbGRlbiBSZXRyaWV2ZXIgMzDigJM0MCBrZyIsImJyZWVkX2V0IjoiS3VsZG5lIHJldHJpaXZlciAzMOKAkzQwIGtnIn0seyJicmVlZCI6ItCY0YDQu9Cw0L3QtNGB0LrQuNC5INC80Y/Qs9C60L7RiNC10YDRgdGC0L3Ri9C5INC/0YjQtdC90LjRh9C90YvQuSDRgtC10YDRjNC10YAiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo2NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6ODB9LCJicmVlZF9lbiI6IklyaXNoIFNvZnQgQ29hdGVkIFdoZWF0ZW4gVGVycmllciIsImJyZWVkX2V0IjoiSWlyaSBwZWhtZWthcnZhbmUgbmlzdXbDpHJ2aSB0ZXJqZXIifSx7ImJyZWVkIjoi0JjRgNC70LDQvdC00YHQutC40Lkg0YLQtdGA0YzQtdGAIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjcwLCLQotGA0LjQvNC80LjQvdCzIjo3NX0sImJyZWVkX2VuIjoiSXJpc2ggVGVycmllciIsImJyZWVkX2V0IjoiSWlyaSB0ZXJqZXIifSx7ImJyZWVkIjoi0JjRgdC/0LDQvdGB0LrQuNC5INCz0LDQu9GM0LPQviAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo1NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjcwfSwiYnJlZWRfZW4iOiJTcGFuaXNoIEdhbGdvIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6Ikhpc3BhYW5pYSBnYWxnbyAyMOKAkzMwIGtnIn0seyJicmVlZCI6ItCY0YHQv9Cw0L3RgdC60LjQuSDQs9Cw0LvRjNCz0L4gMzDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjU1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NjUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo4MH0sImJyZWVkX2VuIjoiU3BhbmlzaCBHYWxnbyAzMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJIaXNwYWFuaWEgZ2FsZ28gMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQmdC+0YDQutGI0LjRgNGB0LrQuNC5INGC0LXRgNGM0LXRgCDQsdC+0LvQtdC1IDMsNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwfSwiYnJlZWRfZW4iOiJZb3Jrc2hpcmUgVGVycmllciBvdmVyIDMsNSBrZyIsImJyZWVkX2V0IjoiWW9ya3NoaXJlIHRlcmplciDDvGxlIDMsNSBrZyJ9LHsiYnJlZWQiOiLQmdC+0YDQutGI0LjRgNGB0LrQuNC5INGC0LXRgNGM0LXRgCDQtNC+IDMsNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjMwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjU1fSwiYnJlZWRfZW4iOiJZb3Jrc2hpcmUgVGVycmllciB1cCB0byAzLDUga2ciLCJicmVlZF9ldCI6IllvcmtzaGlyZSB0ZXJqZXIga3VuaSAzLDUga2cifSx7ImJyZWVkIjoi0JrQsNCy0LDQu9C10YAt0LrQuNC90LMt0YfQsNGA0LvRjNC3LdGB0L/QsNC90LjQtdC70YwgMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjcwfSwiYnJlZWRfZW4iOiJDYXZhbGllciBLaW5nIENoYXJsZXMgU3BhbmllbCAxMOKAkzE1IGtnIiwiYnJlZWRfZXQiOiJDYXZhbGllciBLaW5nIENoYXJsZXMgU3BhbmllbCAxMOKAkzE1IGtnIn0seyJicmVlZCI6ItCa0LDQstCw0LvQtdGALdC60LjQvdCzLdGH0LDRgNC70YzQty3RgdC/0LDQvdC40LXQu9GMIDXigJMxMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwfSwiYnJlZWRfZW4iOiJDYXZhbGllciBLaW5nIENoYXJsZXMgU3BhbmllbCA14oCTMTAga2ciLCJicmVlZF9ldCI6IkNhdmFsaWVyIEtpbmcgQ2hhcmxlcyBTcGFuaWVsIDXigJMxMCBrZyJ9LHsiYnJlZWQiOiLQmtCw0L3QtS3QutC+0YDRgdC+IDQw4oCTNjAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo3MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjg1fSwiYnJlZWRfZW4iOiJDYW5lIENvcnNvIDQw4oCTNjAga2ciLCJicmVlZF9ldCI6IkNhbmUgQ29yc28gNDDigJM2MCBrZyJ9LHsiYnJlZWQiOiLQmtCw0L3QtS3QutC+0YDRgdC+INCx0L7Qu9C10LUgNjAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo5MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjEwNX0sImJyZWVkX2VuIjoiQ2FuZSBDb3JzbyBvdmVyIDYwIGtnIiwiYnJlZWRfZXQiOiJDYW5lIENvcnNvIMO8bGUgNjAga2cifSx7ImJyZWVkIjoi0JrQsNGA0LXQu9C+LdGE0LjQvdGB0LrQsNGPINC70LDQudC60LAg0LTQviAxMyDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo2NX0sImJyZWVkX2VuIjoiS2FyZWxpYW4tRmlubmlzaCBMYWlrYSB1cCB0byAxMyBrZyIsImJyZWVkX2V0IjoiS2FyamFsYS1Tb29tZSBsYWlrYSBrdW5pIDEzIGtnIn0seyJicmVlZCI6ItCa0LjRgtCw0LnRgdC60LDRjyDRhdC+0YXQu9Cw0YLQsNGPINCz0L7Qu9Cw0Y8gNeKAkzEwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzIsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0Miwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NjB9LCJicmVlZF9lbiI6IkNoaW5lc2UgQ3Jlc3RlZCBoYWlybGVzcyA14oCTMTAga2ciLCJicmVlZF9ldCI6IkhpaW5hIGhhcmpha29lciBrYXJ2YXR1IDXigJMxMCBrZyJ9LHsiYnJlZWQiOiLQmtC40YLQsNC50YHQutCw0Y8g0YXQvtGF0LvQsNGC0LDRjyDQs9C+0LvQsNGPINC00L4gNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjI4LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6MzUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjU1fSwiYnJlZWRfZW4iOiJDaGluZXNlIENyZXN0ZWQgaGFpcmxlc3MgdXAgdG8gNSBrZyIsImJyZWVkX2V0IjoiSGlpbmEgaGFyamFrb2VyIGthcnZhdHUga3VuaSA1IGtnIn0seyJicmVlZCI6ItCa0LjRgtCw0LnRgdC60LDRjyDRhdC+0YXQu9Cw0YLQsNGPINC/0YPRhdC+0LLQsNGPIDXigJMxMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwfSwiYnJlZWRfZW4iOiJDaGluZXNlIENyZXN0ZWQgcG93ZGVycHVmZiA14oCTMTAga2ciLCJicmVlZF9ldCI6IkhpaW5hIGhhcmpha29lciBQb3dkZXJwdWZmIDXigJMxMCBrZyJ9LHsiYnJlZWQiOiLQmtC40YLQsNC50YHQutCw0Y8g0YXQvtGF0LvQsNGC0LDRjyDQv9GD0YXQvtCy0LDRjyDQtNC+IDUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiQ2hpbmVzZSBDcmVzdGVkIHBvd2RlcnB1ZmYgdXAgdG8gNSBrZyIsImJyZWVkX2V0IjoiSGlpbmEgaGFyamFrb2VyIFBvd2RlcnB1ZmYga3VuaSA1IGtnIn0seyJicmVlZCI6ItCa0L7QutCw0L/RgyA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjU1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo3NX0sImJyZWVkX2VuIjoiQ29ja2Fwb28gNeKAkzEwIGtnIiwiYnJlZWRfZXQiOiJDb2NrYXBvbyA14oCTMTAga2cifSx7ImJyZWVkIjoi0JrQvtC60LDQv9GDINC00L4gNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjY1fSwiYnJlZWRfZW4iOiJDb2NrYXBvbyB1cCB0byA1IGtnIiwiYnJlZWRfZXQiOiJDb2NrYXBvbyBrdW5pIDUga2cifSx7ImJyZWVkIjoi0JrQvtC70LvQuCAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NjAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo3NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTAwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6OTB9LCJicmVlZF9lbiI6IkNvbGxpZSAyMOKAkzMwIGtnIiwiYnJlZWRfZXQiOiJLb2xsIDIw4oCTMzAga2cifSx7ImJyZWVkIjoi0JrQvtC70LvQuCAzMOKAkzQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo4NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTEwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTAwfSwiYnJlZWRfZW4iOiJDb2xsaWUgMzDigJM0MCBrZyIsImJyZWVkX2V0IjoiS29sbCAzMOKAkzQwIGtnIn0seyJicmVlZCI6ItCa0L7QvNC+0L3QtNC+0YAgMzDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjcwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6ODUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjExMH0sImJyZWVkX2VuIjoiS29tb25kb3IgMzDigJM0MCBrZyIsImJyZWVkX2V0IjoiS29tb25kb3IgMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQmtC+0LzQvtC90LTQvtGAINCx0L7Qu9C10LUgNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo4NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjk1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMzB9LCJicmVlZF9lbiI6IktvbW9uZG9yIG92ZXIgNDAga2ciLCJicmVlZF9ldCI6IktvbW9uZG9yIMO8bGUgNDAga2cifSx7ImJyZWVkIjoi0JvQsNCx0YDQsNC00L7RgCDQs9C70LDQtNC60L7RiNC10YDRgdGC0L3Ri9C5IDIw4oCTMzAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjU1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NzB9LCJicmVlZF9lbiI6IkxhYnJhZG9yIFJldHJpZXZlciBzbW9vdGggMjDigJMzMCBrZyIsImJyZWVkX2V0IjoiTGFicmFkb3JpIHJldHJpaXZlciBsw7xoaWthcnZhbGluZSAyMOKAkzMwIGtnIn0seyJicmVlZCI6ItCb0LDQsdGA0LDQtNC+0YAg0LPQu9Cw0LTQutC+0YjQtdGA0YHRgtC90YvQuSAzMOKAkzQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NTUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo2NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjgwfSwiYnJlZWRfZW4iOiJMYWJyYWRvciBSZXRyaWV2ZXIgc21vb3RoIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IkxhYnJhZG9yaSByZXRyaWl2ZXIgbMO8aGlrYXJ2YWxpbmUgMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQm9Cw0LHRgNCw0LTQvtGAINCz0LvQsNC00LrQvtGI0LXRgNGB0YLQvdGL0Lkg0LHQvtC70LXQtSA0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjcwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6ODAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo5NX0sImJyZWVkX2VuIjoiTGFicmFkb3IgUmV0cmlldmVyIHNtb290aCBvdmVyIDQwIGtnIiwiYnJlZWRfZXQiOiJMYWJyYWRvcmkgcmV0cmlpdmVyIGzDvGhpa2FydmFsaW5lIMO8bGUgNDAga2cifSx7ImJyZWVkIjoi0JvQsNCx0YDQsNC00L7RgCDQtNC70LjQvdC90L7RiNC10YDRgdGC0L3Ri9C5IDIw4oCTMzAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjc1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMDAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo5MH0sImJyZWVkX2VuIjoiTGFicmFkb3IgUmV0cmlldmVyIGxvbmctY29hdGVkIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6IkxhYnJhZG9yaSByZXRyaWl2ZXIgcGlra2FydmFsaW5lIDIw4oCTMzAga2cifSx7ImJyZWVkIjoi0JvQsNCx0YDQsNC00L7RgCDQtNC70LjQvdC90L7RiNC10YDRgdGC0L3Ri9C5IDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo3MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjg1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMTAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMDB9LCJicmVlZF9lbiI6IkxhYnJhZG9yIFJldHJpZXZlciBsb25nLWNvYXRlZCAzMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJMYWJyYWRvcmkgcmV0cmlpdmVyIHBpa2thcnZhbGluZSAzMOKAkzQwIGtnIn0seyJicmVlZCI6ItCb0LDQsdGA0LDQtNC+0YAg0LTQu9C40L3QvdC+0YjQtdGA0YHRgtC90YvQuSDQsdC+0LvQtdC1IDQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6ODUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo5NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTMwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTE1fSwiYnJlZWRfZW4iOiJMYWJyYWRvciBSZXRyaWV2ZXIgbG9uZy1jb2F0ZWQgb3ZlciA0MCBrZyIsImJyZWVkX2V0IjoiTGFicmFkb3JpIHJldHJpaXZlciBwaWtrYXJ2YWxpbmUgw7xsZSA0MCBrZyJ9LHsiYnJlZWQiOiLQm9Cw0LHRgNCw0LTRg9C00LXQu9GMIDEw4oCTMjAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjU1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo3MH0sImJyZWVkX2VuIjoiTGFicmFkb29kbGUgMTDigJMyMCBrZyIsImJyZWVkX2V0IjoiTGFicmFkb29kbGUgMTDigJMyMCBrZyJ9LHsiYnJlZWQiOiLQm9Cw0LHRgNCw0LTRg9C00LXQu9GMIDIw4oCTMzAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo1MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjcwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo5MH0sImJyZWVkX2VuIjoiTGFicmFkb29kbGUgMjDigJMzMCBrZyIsImJyZWVkX2V0IjoiTGFicmFkb29kbGUgMjDigJMzMCBrZyJ9LHsiYnJlZWQiOiLQm9Cw0LHRgNCw0LTRg9C00LXQu9GMIDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjgwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMDB9LCJicmVlZF9lbiI6IkxhYnJhZG9vZGxlIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IkxhYnJhZG9vZGxlIDMw4oCTNDAga2cifSx7ImJyZWVkIjoi0JvQtdCy0YDQtdGC0LrQsCA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwfSwiYnJlZWRfZW4iOiJJdGFsaWFuIEdyZXlob3VuZCA14oCTMTAga2ciLCJicmVlZF9ldCI6Ikl0YWFsaWEgdmluZGtvZXIgNeKAkzEwIGtnIn0seyJicmVlZCI6ItCb0LXQstGA0LXRgtC60LAg0LTQviA1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MjUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0IjozNX0sImJyZWVkX2VuIjoiSXRhbGlhbiBHcmV5aG91bmQgdXAgdG8gNSBrZyIsImJyZWVkX2V0IjoiSXRhYWxpYSB2aW5ka29lciBrdW5pIDUga2cifSx7ImJyZWVkIjoi0JvRhdCw0YHRgdC60LjQuSDQsNC/0YHQviA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjU1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo3NX0sImJyZWVkX2VuIjoiTGhhc2EgQXBzbyA14oCTMTAga2ciLCJicmVlZF9ldCI6IkxoYXNhIEFwc28gNeKAkzEwIGtnIn0seyJicmVlZCI6ItCb0YXQsNGB0YHQutC40Lkg0LDQv9GB0L4g0LTQviA1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NjV9LCJicmVlZF9lbiI6IkxoYXNhIEFwc28gdXAgdG8gNSBrZyIsImJyZWVkX2V0IjoiTGhhc2EgQXBzbyBrdW5pIDUga2cifSx7ImJyZWVkIjoi0JzQsNC70YzRgtC10LfQtSIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiTWFsdGVzZSIsImJyZWVkX2V0IjoiTWFsdGEgYm9sb25lZXMifSx7ImJyZWVkIjoi0JzQsNC70YzRgtC40LnRgdC60LDRjyDQsdC+0LvQvtC90LrQsCA14oCTOCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjY1fSwiYnJlZWRfZW4iOiJNYWx0ZXNlIEJvbG9nbmVzZSA14oCTOCBrZyIsImJyZWVkX2V0IjoiTWFsdGEgYm9sb25lZXMgNeKAkzgga2cifSx7ImJyZWVkIjoi0JzQsNC70YzRgtC40LnRgdC60LDRjyDQsdC+0LvQvtC90LrQsCDQtNC+IDUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MH0sImJyZWVkX2VuIjoiTWFsdGVzZSBCb2xvZ25lc2UgdXAgdG8gNSBrZyIsImJyZWVkX2V0IjoiTWFsdGEgYm9sb25lZXMga3VuaSA1IGtnIn0seyJicmVlZCI6ItCc0LDQu9GM0YLQuNC/0YMgMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjcwfSwiYnJlZWRfZW4iOiJNYWx0aXBvbyAxMOKAkzE1IGtnIiwiYnJlZWRfZXQiOiJNYWx0aXB1dSAxMOKAkzE1IGtnIn0seyJicmVlZCI6ItCc0LDQu9GM0YLQuNC/0YMgNeKAkzEwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NjB9LCJicmVlZF9lbiI6Ik1hbHRpcG9vIDXigJMxMCBrZyIsImJyZWVkX2V0IjoiTWFsdGlwdXUgNeKAkzEwIGtnIn0seyJicmVlZCI6ItCc0LDQu9GM0YLQuNC/0YMg0LTQviA1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NTV9LCJicmVlZF9lbiI6Ik1hbHRpcG9vIHVwIHRvIDUga2ciLCJicmVlZF9ldCI6Ik1hbHRpcHV1IGt1bmkgNSBrZyJ9LHsiYnJlZWQiOiLQnNC10YLQuNGBINC60YDRg9C/0L3Ri9C5IDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjc1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMDAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMDB9LCJicmVlZF9lbiI6Ik1peGVkIGJyZWVkIGxhcmdlIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IlNlZ2F2ZXJkIHN1dXIgMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQnNC10YLQuNGBINC60YDRg9C/0L3Ri9C5INCx0L7Qu9C10LUgNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo3NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjkwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMjAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMjB9LCJicmVlZF9lbiI6Ik1peGVkIGJyZWVkIGxhcmdlIG92ZXIgNDAga2ciLCJicmVlZF9ldCI6IlNlZ2F2ZXJkIHN1dXIgw7xsZSA0MCBrZyJ9LHsiYnJlZWQiOiLQnNC10YLQuNGBINC80LXQu9C60LjQuSA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiTWl4ZWQgYnJlZWQgc21hbGwgNeKAkzEwIGtnIiwiYnJlZWRfZXQiOiJTZWdhdmVyZCB2w6Rpa2UgNeKAkzEwIGtnIn0seyJicmVlZCI6ItCc0LXRgtC40YEg0LzQtdC70LrQuNC5INC00L4gNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjI1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6MzUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjUwfSwiYnJlZWRfZW4iOiJNaXhlZCBicmVlZCBzbWFsbCB1cCB0byA1IGtnIiwiYnJlZWRfZXQiOiJTZWdhdmVyZCB2w6Rpa2Uga3VuaSA1IGtnIn0seyJicmVlZCI6ItCc0LXRgtC40YEg0YHRgNC10LTQvdC40LkgMTDigJMyMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjcwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NzB9LCJicmVlZF9lbiI6Ik1peGVkIGJyZWVkIG1lZGl1bSAxMOKAkzIwIGtnIiwiYnJlZWRfZXQiOiJTZWdhdmVyZCBrZXNrbWluZSAxMOKAkzIwIGtnIn0seyJicmVlZCI6ItCc0LXRgtC40YEg0YHRgNC10LTQvdC40LkgMjDigJMzMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjUwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NjAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjg1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6ODV9LCJicmVlZF9lbiI6Ik1peGVkIGJyZWVkIG1lZGl1bSAyMOKAkzMwIGtnIiwiYnJlZWRfZXQiOiJTZWdhdmVyZCBrZXNrbWluZSAyMOKAkzMwIGtnIn0seyJicmVlZCI6ItCc0LjRgtGC0LXQu9GM0YjQvdCw0YPRhtC10YAgMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjcwLCLQotGA0LjQvNC80LjQvdCzIjo3NX0sImJyZWVkX2VuIjoiU3RhbmRhcmQgU2NobmF1emVyIDEw4oCTMTUga2ciLCJicmVlZF9ldCI6IlN0YW5kYXJkxaFuYXV0c2VyIDEw4oCTMTUga2cifSx7ImJyZWVkIjoi0JzQuNGC0YLQtdC70YzRiNC90LDRg9GG0LXRgCAxNeKAkzIwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo2NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6ODAsItCi0YDQuNC80LzQuNC90LMiOjg1fSwiYnJlZWRfZW4iOiJTdGFuZGFyZCBTY2huYXV6ZXIgMTXigJMyMCBrZyIsImJyZWVkX2V0IjoiU3RhbmRhcmTFoW5hdXRzZXIgMTXigJMyMCBrZyJ9LHsiYnJlZWQiOiLQnNC+0L/RgSIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NTB9LCJicmVlZF9lbiI6IlB1ZyIsImJyZWVkX2V0IjoiTW9wcyJ9LHsiYnJlZWQiOiLQndC10LLRgdC60LDRjyDQvtGA0YXQuNC00LXRjyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiTmV2YSBPcmNoaWQiLCJicmVlZF9ldCI6Ik5lZXZhIG9yaGlkZWUifSx7ImJyZWVkIjoi0J3QtdC80LXRhtC60LDRjyDQvtCy0YfQsNGA0LrQsCAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NjAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo3NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjkwfSwiYnJlZWRfZW4iOiJHZXJtYW4gU2hlcGhlcmQgMjDigJMzMCBrZyIsImJyZWVkX2V0IjoiU2Frc2EgbGFtYmFrb2VyIDIw4oCTMzAga2cifSx7ImJyZWVkIjoi0J3QtdC80LXRhtC60LDRjyDQvtCy0YfQsNGA0LrQsCAzMOKAkzQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo4NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjEwMH0sImJyZWVkX2VuIjoiR2VybWFuIFNoZXBoZXJkIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IlNha3NhIGxhbWJha29lciAzMOKAkzQwIGtnIn0seyJicmVlZCI6ItCd0LXQvNC10YbQutCw0Y8g0L7QstGH0LDRgNC60LAg0LHQvtC70LXQtSA0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjg1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6OTUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMTV9LCJicmVlZF9lbiI6Ikdlcm1hbiBTaGVwaGVyZCBvdmVyIDQwIGtnIiwiYnJlZWRfZXQiOiJTYWtzYSBsYW1iYWtvZXIgw7xsZSA0MCBrZyJ9LHsiYnJlZWQiOiLQqNCy0LXQudGG0LDRgNGB0LrQsNGPINC+0LLRh9Cw0YDQutCwIDIw4oCTMzAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjc1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6OTB9LCJicmVlZF9lbiI6IlN3aXNzIFNoZXBoZXJkIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6IsWgdmVpdHNpIGxhbWJha29lciAyMOKAkzMwIGtnIn0seyJicmVlZCI6ItCo0LLQtdC50YbQsNGA0YHQutCw0Y8g0L7QstGH0LDRgNC60LAgMzDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjcwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6ODUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMDB9LCJicmVlZF9lbiI6IlN3aXNzIFNoZXBoZXJkIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IsWgdmVpdHNpIGxhbWJha29lciAzMOKAkzQwIGtnIn0seyJicmVlZCI6ItCo0LLQtdC50YbQsNGA0YHQutCw0Y8g0L7QstGH0LDRgNC60LAg0LHQvtC70LXQtSA0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjg1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6OTUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMTV9LCJicmVlZF9lbiI6IlN3aXNzIFNoZXBoZXJkIG92ZXIgNDAga2ciLCJicmVlZF9ldCI6IsWgdmVpdHNpIGxhbWJha29lciDDvGxlIDQwIGtnIn0seyJicmVlZCI6ItCd0L7RgNCy0LjRhy3RgtC10YDRjNC10YAiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NjAsItCi0YDQuNC80LzQuNC90LMiOjY1fSwiYnJlZWRfZW4iOiJOb3J3aWNoIFRlcnJpZXIiLCJicmVlZF9ldCI6Ik5vcndpdMWhaSB0ZXJqZXIifSx7ImJyZWVkIjoi0J3QvtGA0YTQvtC70Lot0YLQtdGA0YzQtdGAIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwLCLQotGA0LjQvNC80LjQvdCzIjo2NX0sImJyZWVkX2VuIjoiTm9yZm9sayBUZXJyaWVyIiwiYnJlZWRfZXQiOiJOb3Jmb2xraSB0ZXJqZXIifSx7ImJyZWVkIjoi0J3RjNGO0YTQsNGD0L3QtNC70LXQvdC0IDQw4oCTNjAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo4NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjk1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMzAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMTV9LCJicmVlZF9lbiI6Ik5ld2ZvdW5kbGFuZCA0MOKAkzYwIGtnIiwiYnJlZWRfZXQiOiJOZXdmb3VuZGxhbmRpIGtvZXIgNDDigJM2MCBrZyJ9LHsiYnJlZWQiOiLQndGM0Y7RhNCw0YPQvdC00LvQtdC90LQg0LHQvtC70LXQtSA2MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjEwMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjExNSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTUwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTMwfSwiYnJlZWRfZW4iOiJOZXdmb3VuZGxhbmQgb3ZlciA2MCBrZyIsImJyZWVkX2V0IjoiTmV3Zm91bmRsYW5kaSBrb2VyIMO8bGUgNjAga2cifSx7ImJyZWVkIjoi0J/QsNC/0LjQudC+0L0iLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NTV9LCJicmVlZF9lbiI6IlBhcGlsbG9uIiwiYnJlZWRfZXQiOiJQYXBpbGxvbiJ9LHsiYnJlZWQiOiLQn9C10LrQuNC90LXRgSA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MH0sImJyZWVkX2VuIjoiUGVraW5nZXNlIDXigJMxMCBrZyIsImJyZWVkX2V0IjoiUGVraW5lc2kga29lciA14oCTMTAga2cifSx7ImJyZWVkIjoi0J/QtdC60LjQvdC10YEg0LTQviA1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NTV9LCJicmVlZF9lbiI6IlBla2luZ2VzZSB1cCB0byA1IGtnIiwiYnJlZWRfZXQiOiJQZWtpbmVzaSBrb2VyIGt1bmkgNSBrZyJ9LHsiYnJlZWQiOiLQn9GD0LTQtdC70Ywg0LHQvtC70YzRiNC+0LkgMjDigJMzMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjUwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NzAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjkwfSwiYnJlZWRfZW4iOiJTdGFuZGFyZCBQb29kbGUgMjDigJMzMCBrZyIsImJyZWVkX2V0IjoiU3RhbmRhcmRwdXVkZWwgMjDigJMzMCBrZyJ9LHsiYnJlZWQiOiLQn9GD0LTQtdC70Ywg0LHQvtC70YzRiNC+0LkgMzDigJM0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjYwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6ODAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjEwMH0sImJyZWVkX2VuIjoiU3RhbmRhcmQgUG9vZGxlIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IlN0YW5kYXJkcHV1ZGVsIDMw4oCTNDAga2cifSx7ImJyZWVkIjoi0J/Rg9C00LXQu9GMINC60LDRgNC70LjQutC+0LLRi9C5IDXigJMxMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwfSwiYnJlZWRfZW4iOiJNaW5pYXR1cmUgUG9vZGxlIDXigJMxMCBrZyIsImJyZWVkX2V0IjoiS8Okw6RidXNwdXVkZWwgNeKAkzEwIGtnIn0seyJicmVlZCI6ItCf0YPQtNC10LvRjCDQvNCw0LvRi9C5IDEw4oCTMTUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjU1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo3MH0sImJyZWVkX2VuIjoiU21hbGwgUG9vZGxlIDEw4oCTMTUga2ciLCJicmVlZF9ldCI6IlbDpGlrZSBwdXVkZWwgMTDigJMxNSBrZyJ9LHsiYnJlZWQiOiLQn9GD0LTQtdC70Ywg0LzQsNC70YvQuSAxNeKAkzIwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo2NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6ODB9LCJicmVlZF9lbiI6IlNtYWxsIFBvb2RsZSAxNeKAkzIwIGtnIiwiYnJlZWRfZXQiOiJWw6Rpa2UgcHV1ZGVsIDE14oCTMjAga2cifSx7ImJyZWVkIjoi0J/Rg9C00LXQu9GMINGC0L7QuSDQtNC+IDUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiVG95IFBvb2RsZSB1cCB0byA1IGtnIiwiYnJlZWRfZXQiOiJNw6RuZ3Vhc2phIHB1dWRlbCBrdW5pIDUga2cifSx7ImJyZWVkIjoi0KDQuNC30LXQvdGI0L3QsNGD0YbQtdGAIDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjgwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMDAsItCi0YDQuNC80LzQuNC90LMiOjExMH0sImJyZWVkX2VuIjoiR2lhbnQgU2NobmF1emVyIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IlN1dXLFoW5hdXRzZXIgMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQoNC40LfQtdC90YjQvdCw0YPRhtC10YAg0LHQvtC70LXQtSA0MCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjc1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6OTUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjEyMCwi0KLRgNC40LzQvNC40L3QsyI6MTI1fSwiYnJlZWRfZW4iOiJHaWFudCBTY2huYXV6ZXIgb3ZlciA0MCBrZyIsImJyZWVkX2V0IjoiU3V1csWhbmF1dHNlciDDvGxlIDQwIGtnIn0seyJicmVlZCI6ItCg0YPRgdGB0LrQsNGPINGG0LLQtdGC0L3QsNGPINCx0L7Qu9C+0L3QutCwIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwfSwiYnJlZWRfZW4iOiJSdXNzaWFuIENvbG9yZWQgTGFwZG9nIiwiYnJlZWRfZXQiOiJWZW5lIHbDpHJ2aWxpbmUgc8O8bGVrb2VyIn0seyJicmVlZCI6ItCg0YPRgdGB0LrQuNC5INC+0YXQvtGC0L3QuNGH0LjQuSDRgdC/0LDQvdC40LXQu9GMIDEw4oCTMTUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjU1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo3MH0sImJyZWVkX2VuIjoiUnVzc2lhbiBTcGFuaWVsIDEw4oCTMTUga2ciLCJicmVlZF9ldCI6IlZlbmUgamFoaXNwYW5qZWwgMTDigJMxNSBrZyJ9LHsiYnJlZWQiOiLQoNGD0YHRgdC60LjQuSDQvtGF0L7RgtC90LjRh9C40Lkg0YHQv9Cw0L3QuNC10LvRjCAxNeKAkzIwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo2NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6ODB9LCJicmVlZF9lbiI6IlJ1c3NpYW4gU3BhbmllbCAxNeKAkzIwIGtnIiwiYnJlZWRfZXQiOiJWZW5lIGphaGlzcGFuamVsIDE14oCTMjAga2cifSx7ImJyZWVkIjoi0KDRg9GB0YHQutC40Lkg0YLQvtC5INCz0LvQsNC00LrQvtGI0LXRgNGB0YLQvdGL0LkiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MjUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0IjozNX0sImJyZWVkX2VuIjoiUnVzc2lhbiBUb3kgc21vb3RoIiwiYnJlZWRfZXQiOiJWZW5lIFRveSBsw7xoaWthcnZhbGluZSJ9LHsiYnJlZWQiOiLQoNGD0YHRgdC60LjQuSDRgtC+0Lkg0LTQu9C40L3QvdC+0YjQtdGA0YHRgtC90YvQuSIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiUnVzc2lhbiBUb3kgbG9uZy1jb2F0ZWQiLCJicmVlZF9ldCI6IlZlbmUgVG95IHBpa2thcnZhbGluZSJ9LHsiYnJlZWQiOiLQoNGD0YHRgdC60LjQuSDRh9C10YDQvdGL0Lkg0YLQtdGA0YzQtdGAIDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo2MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjgwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMDB9LCJicmVlZF9lbiI6IkJsYWNrIFJ1c3NpYW4gVGVycmllciAzMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJNdXN0IFZlbmUgdGVyamVyIDMw4oCTNDAga2cifSx7ImJyZWVkIjoi0KDRg9GB0YHQutC40Lkg0YfQtdGA0L3Ri9C5INGC0LXRgNGM0LXRgCDQsdC+0LvQtdC1IDQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NzUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo5NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTIwfSwiYnJlZWRfZW4iOiJCbGFjayBSdXNzaWFuIFRlcnJpZXIgb3ZlciA0MCBrZyIsImJyZWVkX2V0IjoiTXVzdCBWZW5lIHRlcmplciDDvGxlIDQwIGtnIn0seyJicmVlZCI6ItCg0YPRgdGB0LrQvi3QtdCy0YDQvtC/0LXQudGB0LrQsNGPINC70LDQudC60LAgMjDigJMyOCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjUwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NjUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo5MH0sImJyZWVkX2VuIjoiUnVzc2lhbi1FdXJvcGVhbiBMYWlrYSAyMOKAkzI4IGtnIiwiYnJlZWRfZXQiOiJWZW5lLUV1cm9vcGEgbGFpa2EgMjDigJMyOCBrZyJ9LHsiYnJlZWQiOiLQodCw0LzQvtC10LQgMjDigJMzMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjYwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NzUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo5MH0sImJyZWVkX2VuIjoiU2Ftb3llZCAyMOKAkzMwIGtnIiwiYnJlZWRfZXQiOiJTYW1vamVlZCAyMOKAkzMwIGtnIn0seyJicmVlZCI6ItCh0LDQvNC+0LXQtCAzMOKAkzQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo4NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjEwMH0sImJyZWVkX2VuIjoiU2Ftb3llZCAzMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJTYW1vamVlZCAzMOKAkzQwIGtnIn0seyJicmVlZCI6ItCh0LXRgtGC0LXRgCDQsNC90LPQu9C40LnRgdC60LjQuSAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NTAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo3MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6OTB9LCJicmVlZF9lbiI6IkVuZ2xpc2ggU2V0dGVyIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6IkluZ2xpc2Ugc2V0dGVyIDIw4oCTMzAga2cifSx7ImJyZWVkIjoi0KHQtdGC0YLQtdGAINCz0L7RgNC00L7QvSAzMOKAkzQwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NjAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo4MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTAwfSwiYnJlZWRfZW4iOiJHb3Jkb24gU2V0dGVyIDMw4oCTNDAga2ciLCJicmVlZF9ldCI6IkdvcmRvbmkgc2V0dGVyIDMw4oCTNDAga2cifSx7ImJyZWVkIjoi0KHQtdGC0YLQtdGAINC40YDQu9Cw0L3QtNGB0LrQuNC5IDIw4oCTMzAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo1MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjcwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo5MH0sImJyZWVkX2VuIjoiSXJpc2ggU2V0dGVyIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6Iklpcmkgc2V0dGVyIDIw4oCTMzAga2cifSx7ImJyZWVkIjoi0KHQuNCx0LAt0LjQvdGDIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQ1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NjAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo3MH0sImJyZWVkX2VuIjoiU2hpYmEgSW51IiwiYnJlZWRfZXQiOiJTaGliYSBJbnUifSx7ImJyZWVkIjoi0KHQuNC70LjRhdC10Lwt0YLQtdGA0YzQtdGAIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwLCLQotGA0LjQvNC80LjQvdCzIjo2NX0sImJyZWVkX2VuIjoiU2VhbHloYW0gVGVycmllciIsImJyZWVkX2V0IjoiU2VhbHloYW1pIHRlcmplciJ9LHsiYnJlZWQiOiLQodC60L7RgtGHLdGC0LXRgNGM0LXRgCIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MCwi0KLRgNC40LzQvNC40L3QsyI6NjV9LCJicmVlZF9lbiI6IlNjb3R0aXNoIFRlcnJpZXIiLCJicmVlZF9ldCI6IsWgb3RpIHRlcmplciJ9LHsiYnJlZWQiOiLQotCw0LrRgdCwINCz0LvQsNC00LrQvtGI0LXRgNGB0YLQvdCw0Y8g0LrQsNGA0LvQuNC60L7QstCw0Y8gNeKAkzEwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0MCwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjUwfSwiYnJlZWRfZW4iOiJEYWNoc2h1bmQgc21vb3RoIG1pbmlhdHVyZSA14oCTMTAga2ciLCJicmVlZF9ldCI6IlRha3Npa29lciBsw7xoaWthcnZhbGluZSBrw6TDpGJ1cyA14oCTMTAga2cifSx7ImJyZWVkIjoi0KLQsNC60YHQsCDQs9C70LDQtNC60L7RiNC10YDRgdGC0L3QsNGPINC60YDQvtC70LjRh9GM0Y8g0LTQviA1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MjUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0IjozNSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjQ1fSwiYnJlZWRfZW4iOiJEYWNoc2h1bmQgc21vb3RoIHJhYmJpdCB1cCB0byA1IGtnIiwiYnJlZWRfZXQiOiJUYWtzaWtvZXIgbMO8aGlrYXJ2YWxpbmUga8O8w7xsaWsga3VuaSA1IGtnIn0seyJicmVlZCI6ItCi0LDQutGB0LAg0LPQu9Cw0LTQutC+0YjQtdGA0YHRgtC90LDRjyDRgdGC0LDQvdC00LDRgNGC0L3QsNGPIDEw4oCTMTUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NjB9LCJicmVlZF9lbiI6IkRhY2hzaHVuZCBzbW9vdGggc3RhbmRhcmQgMTDigJMxNSBrZyIsImJyZWVkX2V0IjoiVGFrc2lrb2VyIGzDvGhpa2FydmFsaW5lIHN0YW5kYXJkIDEw4oCTMTUga2cifSx7ImJyZWVkIjoi0KLQsNC60YHQsCDQtNC70LjQvdC90L7RiNC10YDRgdGC0L3QsNGPINC60LDRgNC70LjQutC+0LLQsNGPIDXigJMxMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwfSwiYnJlZWRfZW4iOiJEYWNoc2h1bmQgbG9uZy1jb2F0ZWQgbWluaWF0dXJlIDXigJMxMCBrZyIsImJyZWVkX2V0IjoiVGFrc2lrb2VyIHBpa2thcnZhbGluZSBrw6TDpGJ1cyA14oCTMTAga2cifSx7ImJyZWVkIjoi0KLQsNC60YHQsCDQtNC70LjQvdC90L7RiNC10YDRgdGC0L3QsNGPINC60YDQvtC70LjRh9GM0Y8g0LTQviA1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NTV9LCJicmVlZF9lbiI6IkRhY2hzaHVuZCBsb25nLWNvYXRlZCByYWJiaXQgdXAgdG8gNSBrZyIsImJyZWVkX2V0IjoiVGFrc2lrb2VyIHBpa2thcnZhbGluZSBrw7zDvGxpayBrdW5pIDUga2cifSx7ImJyZWVkIjoi0KLQsNC60YHQsCDQtNC70LjQvdC90L7RiNC10YDRgdGC0L3QsNGPINGB0YLQsNC90LTQsNGA0YLQvdCw0Y8gMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjcwfSwiYnJlZWRfZW4iOiJEYWNoc2h1bmQgbG9uZy1jb2F0ZWQgc3RhbmRhcmQgMTDigJMxNSBrZyIsImJyZWVkX2V0IjoiVGFrc2lrb2VyIHBpa2thcnZhbGluZSBzdGFuZGFyZCAxMOKAkzE1IGtnIn0seyJicmVlZCI6ItCi0LDQutGB0LAg0LbQtdGB0YLQutC+0YjQtdGA0YHRgtC90LDRjyDQutCw0YDQu9C40LrQvtCy0LDRjyA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MCwi0KLRgNC40LzQvNC40L3QsyI6NjV9LCJicmVlZF9lbiI6IkRhY2hzaHVuZCB3aXJlLWhhaXJlZCBtaW5pYXR1cmUgNeKAkzEwIGtnIiwiYnJlZWRfZXQiOiJUYWtzaWtvZXIga2FydWthcnZhbGluZSBrw6TDpGJ1cyA14oCTMTAga2cifSx7ImJyZWVkIjoi0KLQsNC60YHQsCDQttC10YHRgtC60L7RiNC10YDRgdGC0L3QsNGPINC60YDQvtC70LjRh9GM0Y8g0LTQviA1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NTUsItCi0YDQuNC80LzQuNC90LMiOjU1fSwiYnJlZWRfZW4iOiJEYWNoc2h1bmQgd2lyZS1oYWlyZWQgcmFiYml0IHVwIHRvIDUga2ciLCJicmVlZF9ldCI6IlRha3Npa29lciBrYXJ1a2FydmFsaW5lIGvDvMO8bGlrIGt1bmkgNSBrZyJ9LHsiYnJlZWQiOiLQotCw0LrRgdCwINC20LXRgdGC0LrQvtGI0LXRgNGB0YLQvdCw0Y8g0YHRgtCw0L3QtNCw0YDRgtC90LDRjyAxMOKAkzE1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo1NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NzAsItCi0YDQuNC80LzQuNC90LMiOjc1fSwiYnJlZWRfZW4iOiJEYWNoc2h1bmQgd2lyZS1oYWlyZWQgc3RhbmRhcmQgMTDigJMxNSBrZyIsImJyZWVkX2V0IjoiVGFrc2lrb2VyIGthcnVrYXJ2YWxpbmUgc3RhbmRhcmQgMTDigJMxNSBrZyJ9LHsiYnJlZWQiOiLQo9C40L/Qv9C10YIgMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDV9LCJicmVlZF9lbiI6IldoaXBwZXQgMTDigJMxNSBrZyIsImJyZWVkX2V0IjoiV2hpcHBldCAxMOKAkzE1IGtnIn0seyJicmVlZCI6ItCj0LjQv9C/0LXRgiAxNeKAkzIwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NDAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo1MH0sImJyZWVkX2VuIjoiV2hpcHBldCAxNeKAkzIwIGtnIiwiYnJlZWRfZXQiOiJXaGlwcGV0IDE14oCTMjAga2cifSx7ImJyZWVkIjoi0KTQuNC90YHQutC40Lkg0LvQsNC/0YXRg9C90LQgMTXigJMyMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjUwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NjUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjg1fSwiYnJlZWRfZW4iOiJGaW5uaXNoIExhcHBodW5kIDE14oCTMjAga2ciLCJicmVlZF9ldCI6IlNvb21lIGxhbWJha29lciAxNeKAkzIwIGtnIn0seyJicmVlZCI6ItCk0LjQvdGB0LrQuNC5INC70LDQv9GF0YPQvdC0IDIw4oCTMjQg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo1NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjcwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo5MH0sImJyZWVkX2VuIjoiRmlubmlzaCBMYXBwaHVuZCAyMOKAkzI0IGtnIiwiYnJlZWRfZXQiOiJTb29tZSBsYW1iYWtvZXIgMjDigJMyNCBrZyJ9LHsiYnJlZWQiOiLQpNC+0LrRgdGC0LXRgNGM0LXRgCDQttC10YHRgtC60L7RiNC10YDRgdGC0L3Ri9C5IDEw4oCTMTUg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjU1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo3MCwi0KLRgNC40LzQvNC40L3QsyI6NzV9LCJicmVlZF9lbiI6IldpcmUgRm94IFRlcnJpZXIgMTDigJMxNSBrZyIsImJyZWVkX2V0IjoiS2FydWthcnZhbGluZSBmb3h0ZXJqZXIgMTDigJMxNSBrZyJ9LHsiYnJlZWQiOiLQpNC+0LrRgdGC0LXRgNGM0LXRgCDQttC10YHRgtC60L7RiNC10YDRgdGC0L3Ri9C5IDXigJMxMCDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwLCLQotGA0LjQvNC80LjQvdCzIjo2NX0sImJyZWVkX2VuIjoiV2lyZSBGb3ggVGVycmllciA14oCTMTAga2ciLCJicmVlZF9ldCI6IkthcnVrYXJ2YWxpbmUgZm94dGVyamVyIDXigJMxMCBrZyJ9LHsiYnJlZWQiOiLQpNGA0LDQvdGG0YPQt9GB0LrQuNC5INCx0YPQu9GM0LTQvtCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjo2MH0sImJyZWVkX2VuIjoiRnJlbmNoIEJ1bGxkb2ciLCJicmVlZF9ldCI6IlByYW50c3VzZSBidWxkb2cifSx7ImJyZWVkIjoi0KXQsNGB0LrQuCAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NjAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo3NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjkwfSwiYnJlZWRfZW4iOiJTaWJlcmlhbiBIdXNreSAyMOKAkzMwIGtnIiwiYnJlZWRfZXQiOiJTaWJlcmkgaHVza3kgMjDigJMzMCBrZyJ9LHsiYnJlZWQiOiLQpdCw0YHQutC4IDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo3MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjg1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6MTAwfSwiYnJlZWRfZW4iOiJTaWJlcmlhbiBIdXNreSAzMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJTaWJlcmkgaHVza3kgMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQptCy0LXRgNCz0YjQvdCw0YPRhtC10YAgMTDigJMxNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjcwLCLQotGA0LjQvNC80LjQvdCzIjo3NX0sImJyZWVkX2VuIjoiTWluaWF0dXJlIFNjaG5hdXplciAxMOKAkzE1IGtnIiwiYnJlZWRfZXQiOiJLw6TDpGJ1c8WhbmF1dHNlciAxMOKAkzE1IGtnIn0seyJicmVlZCI6ItCm0LLQtdGA0LPRiNC90LDRg9GG0LXRgCA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MCwi0KLRgNC40LzQvNC40L3QsyI6NjV9LCJicmVlZF9lbiI6Ik1pbmlhdHVyZSBTY2huYXV6ZXIgNeKAkzEwIGtnIiwiYnJlZWRfZXQiOiJLw6TDpGJ1c8WhbmF1dHNlciA14oCTMTAga2cifSx7ImJyZWVkIjoi0KfQsNGDLdGH0LDRgyAyMOKAkzMwINC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6NjAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo3NSwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6MTAwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6OTB9LCJicmVlZF9lbiI6IkNob3cgQ2hvdyAyMOKAkzMwIGtnIiwiYnJlZWRfZXQiOiJDaG93IENob3cgMjDigJMzMCBrZyJ9LHsiYnJlZWQiOiLQp9Cw0YMt0YfQsNGDIDMw4oCTNDAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo3MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjg1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0IjoxMTAsItCt0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwIjoxMDB9LCJicmVlZF9lbiI6IkNob3cgQ2hvdyAzMOKAkzQwIGtnIiwiYnJlZWRfZXQiOiJDaG93IENob3cgMzDigJM0MCBrZyJ9LHsiYnJlZWQiOiLQp9C40YXRg9Cw0YXRg9CwINCz0LvQsNC00LrQvtGI0LXRgNGB0YLQvdGL0LkiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MjUsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0IjozNX0sImJyZWVkX2VuIjoiQ2hpaHVhaHVhIHNtb290aCIsImJyZWVkX2V0IjoiVMWhaWh1YWh1YSBsw7xoaWthcnZhbGluZSJ9LHsiYnJlZWQiOiLQp9C40YXRg9Cw0YXRg9CwINC00LvQuNC90L3QvtGI0LXRgNGB0YLQvdGL0LkiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NTV9LCJicmVlZF9lbiI6IkNoaWh1YWh1YSBsb25nLWNvYXRlZCIsImJyZWVkX2V0IjoiVMWhaWh1YWh1YSBwaWtrYXJ2YWxpbmUifSx7ImJyZWVkIjoi0KjQsNGA0L/QtdC5IDE14oCTMjAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjUwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NjV9LCJicmVlZF9lbiI6IlNoYXIgUGVpIDE14oCTMjAga2ciLCJicmVlZF9ldCI6IsWgYXItUGVpIDE14oCTMjAga2cifSx7ImJyZWVkIjoi0KjQsNGA0L/QtdC5IDIw4oCTMzAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0NSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjU1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NzB9LCJicmVlZF9lbiI6IlNoYXIgUGVpIDIw4oCTMzAga2ciLCJicmVlZF9ldCI6IsWgYXItUGVpIDIw4oCTMzAga2cifSx7ImJyZWVkIjoi0KjQtdC70YLQuCIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjUwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjYwfSwiYnJlZWRfZW4iOiJTaGV0bGFuZCBTaGVlcGRvZyIsImJyZWVkX2V0IjoixaBldGxhbmRpIGxhbWJha29lciJ9LHsiYnJlZWQiOiLQqNC4LdGC0YbRgyA14oCTMTAg0LrQsyIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozNSwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQ1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2MH0sImJyZWVkX2VuIjoiU2hpaCBUenUgNeKAkzEwIGtnIiwiYnJlZWRfZXQiOiJTaGloIFR6dSA14oCTMTAga2cifSx7ImJyZWVkIjoi0KjQuC3RgtGG0YMg0LTQviA1INC60LMiLCJzZXJ2aWNlcyI6eyLQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCI6MzAsItCT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0Ijo0MCwi0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCI6NTV9LCJicmVlZF9lbiI6IlNoaWggVHp1IHVwIHRvIDUga2ciLCJicmVlZF9ldCI6IlNoaWggVHp1IGt1bmkgNSBrZyJ9LHsiYnJlZWQiOiLQqNC90LDRg9GG0LXRgCDQvNC40L3QuNCw0YLRjtGA0L3Ri9C5INC00L4gNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCi0YDQuNC80LzQuNC90LMiOjY1fSwiYnJlZWRfZW4iOiJNaW5pYXR1cmUgU2NobmF1emVyIHVwIHRvIDUga2ciLCJicmVlZF9ldCI6IkvDpMOkYnVzxaFuYXV0c2VyIGt1bmkgNSBrZyJ9LHsiYnJlZWQiOiLQqNC/0LjRhiDQvdC10LzQtdGG0LrQuNC5IC8g0L/QvtC80LXRgNCw0L3RgdC60LjQuSDQsdC+0LvQtdC1IDMsNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTAsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjY1LCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NjB9LCJicmVlZF9lbiI6Ikdlcm1hbiBTcGl0eiAvIFBvbWVyYW5pYW4gb3ZlciAzLDUga2ciLCJicmVlZF9ldCI6IlNha3NhIHNwaXRzIC8gUG9tZXJhbmlhbiDDvGxlIDMsNSBrZyJ9LHsiYnJlZWQiOiLQqNC/0LjRhiDQvdC10LzQtdGG0LrQuNC5IC8g0L/QvtC80LXRgNCw0L3RgdC60LjQuSDQtNC+IDMsNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjM1LCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwLCLQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCI6NTV9LCJicmVlZF9lbiI6Ikdlcm1hbiBTcGl0eiAvIFBvbWVyYW5pYW4gdXAgdG8gMyw1IGtnIiwiYnJlZWRfZXQiOiJTYWtzYSBzcGl0cyAvIFBvbWVyYW5pYW4ga3VuaSAzLDUga2cifSx7ImJyZWVkIjoi0KjQv9C40YYg0Y/Qv9C+0L3RgdC60LjQuSIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0Ijo0MCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjUwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo2NSwi0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAiOjYwfSwiYnJlZWRfZW4iOiJKYXBhbmVzZSBTcGl0eiIsImJyZWVkX2V0IjoiSmFhcGFuaSBzcGl0cyJ9LHsiYnJlZWQiOiLQqdC10L3QutC4Iiwic2VydmljZXMiOnsi0JLRgdGPINC/0YDQvtCz0YDQsNC80LzQsCI6NTV9LCJicmVlZF9lbiI6IlB1cHBpZXMiLCJicmVlZF9ldCI6Ikt1dHNpa2FkIn0seyJicmVlZCI6ItCt0YHRgtC+0L3RgdC60LDRjyDQs9C+0L3Rh9Cw0Y8gMTXigJMyNSDQutCzIiwic2VydmljZXMiOnsi0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQiOjQwLCLQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCI6NTB9LCJicmVlZF9lbiI6IkVzdG9uaWFuIEhvdW5kIDE14oCTMjUga2ciLCJicmVlZF9ldCI6IkVlc3RpIGhhZ2lqYXMgMTXigJMyNSBrZyJ9LHsiYnJlZWQiOiLQr9C/0L7QvdGB0LrQuNC5INGF0LjQvSIsInNlcnZpY2VzIjp7ItCR0LDQt9C+0LLRi9C5INGD0YXQvtC0IjozMCwi0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQiOjQwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo1NX0sImJyZWVkX2VuIjoiSmFwYW5lc2UgQ2hpbiIsImJyZWVkX2V0IjoiSmFhcGFuaSBDaGluIn0seyJicmVlZCI6ItCa0L7RiNC60LAg0LrQvtGA0L7RgtC60L7RiNC10YDRgdGC0L3QsNGPIiwic2VydmljZXMiOnsi0JLRi9GH0LXRgSI6NDUsItCa0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQiOjYwfSwiYnJlZWRfZW4iOiJDYXQgc2hvcnQtaGFpcmVkIiwiYnJlZWRfZXQiOiJLYXNzIGzDvGhpa2FydmFsaW5lIn0seyJicmVlZCI6ItCa0L7RiNC60LAg0LTQu9C40L3QvdC+0YjQtdGA0YHRgtC90LDRjyIsInNlcnZpY2VzIjp7ItCS0YvRh9C10YEiOjU1LCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo4MH0sImJyZWVkX2VuIjoiQ2F0IGxvbmctaGFpcmVkIiwiYnJlZWRfZXQiOiJLYXNzIHBpa2thcnZhbGluZSJ9LHsiYnJlZWQiOiLQmtC+0YjQutCwINCc0LXQudC9LdC60YPQvSIsInNlcnZpY2VzIjp7ItCS0YvRh9GR0YEiOjYwLCLQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Ijo5MH0sImJyZWVkX2VuIjoiQ2F0IE1haW5lIENvb24iLCJicmVlZF9ldCI6Ikthc3MgTWFpbmUgQ29vbiJ9XTsKdmFyIFJBSUxXQVkgPSAiaHR0cHM6Ly9yamdyb29taW5nLnVwLnJhaWx3YXkuYXBwL2Jvb2siOwp2YXIgR09PR0xFX1NDUklQVCA9ICJodHRwczovL3NjcmlwdC5nb29nbGUuY29tL21hY3Jvcy9zL0FLZnljYnlUU1otZUpNZGVwLUQwTHItbngwX1Y0SEJXZ0lJY3RuUlQycmpTRHZCeWJqNUNZSTNOSzJNcWNBd19jZmN6Z1JFaWZnL2V4ZWMiOwp2YXIgRkFMTEJBQ0tfVElNRVMgPSBbJzEwOjAwJywnMTA6MzAnLCcxMTowMCcsJzExOjMwJywnMTI6MDAnLCcxMjozMCcsJzEzOjAwJywnMTM6MzAnLCcxNDowMCcsJzE0OjMwJywnMTU6MDAnLCcxNTozMCcsJzE2OjAwJywnMTY6MzAnLCcxNzowMCcsJzE3OjMwJywnMTg6MDAnXTsKdmFyIGJvb2tpbmcgPSB7YnJlZWQ6JycsYnJlZWREaXNwbGF5OicnLHNlcnZpY2U6JycscHJpY2U6MCxtYXN0ZXI6JycsZ3Jvb21IaXN0b3J5OicnLGRhdGU6JycsdGltZTonJyxsYW5nOidydSd9Owp2YXIgc2VsQnJlZWQgPSBudWxsOwp2YXIgY1kgPSBuZXcgRGF0ZSgpLmdldEZ1bGxZZWFyKCk7CnZhciBjTSA9IG5ldyBEYXRlKCkuZ2V0TW9udGgoKTsKdmFyIHN0ZXAgPSAxOwp2YXIgTU9OVEhTID0gWyfQr9C90LLQsNGA0YwnLCfQpNC10LLRgNCw0LvRjCcsJ9Cc0LDRgNGCJywn0JDQv9GA0LXQu9GMJywn0JzQsNC5Jywn0JjRjtC90YwnLCfQmNGO0LvRjCcsJ9CQ0LLQs9GD0YHRgicsJ9Ch0LXQvdGC0Y/QsdGA0YwnLCfQntC60YLRj9Cx0YDRjCcsJ9Cd0L7Rj9Cx0YDRjCcsJ9CU0LXQutCw0LHRgNGMJ107CgpmdW5jdGlvbiBzaG93U2NyZWVuKGlkKSB7CiAgZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLnNjcmVlbicpLmZvckVhY2goZnVuY3Rpb24ocyl7cy5jbGFzc0xpc3QucmVtb3ZlKCdhY3RpdmUnKTt9KTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZChpZCkuY2xhc3NMaXN0LmFkZCgnYWN0aXZlJyk7CiAgd2luZG93LnNjcm9sbFRvKDAsMCk7Cn0KCmZ1bmN0aW9uIGdvU3RlcChuKSB7CiAgWydiazEnLCdiazInLCdiazMnLCdiazQnLCdiazUnXS5mb3JFYWNoKGZ1bmN0aW9uKGlkLGkpewogICAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoaWQpLmNsYXNzTmFtZSA9ICdzdGVwJyArIChpKzE9PT1uPycgc2hvdyc6JycpOwogIH0pOwogIGZvcih2YXIgaT0xO2k8PTU7aSsrKXsKICAgIHZhciBwcz1kb2N1bWVudC5nZXRFbGVtZW50QnlJZCgncHMnK2kpOwogICAgdmFyIHBsPWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdwbCcraSk7CiAgICBpZihpPG4pe3BzLmNsYXNzTmFtZT0ncHMgZG9uZSc7aWYocGwpcGwuY2xhc3NOYW1lPSdwbCBkb25lJzt9CiAgICBlbHNlIGlmKGk9PT1uKXtwcy5jbGFzc05hbWU9J3BzIGFjdGl2ZSc7aWYocGwpcGwuY2xhc3NOYW1lPSdwbCc7fQogICAgZWxzZXtwcy5jbGFzc05hbWU9J3BzJztpZihwbClwbC5jbGFzc05hbWU9J3BsJzt9CiAgfQogIHN0ZXA9bjsgd2luZG93LnNjcm9sbFRvKDAsMCk7CiAgaWYobj09PTMpIGZpbHRlck1hc3RlcnMoKTsKfQoKZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2Jvb2tCdG4nKS5vbmNsaWNrID0gZnVuY3Rpb24oKXsKICBzaG93U2NyZWVuKCdib29rU2NyZWVuJyk7IGdvU3RlcCgxKTsgYnVpbGRDYWwoKTsKfTsKZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2JhY2tCdG4nKS5vbmNsaWNrID0gZnVuY3Rpb24oKXsKICBpZihzdGVwPjEpe2dvU3RlcChzdGVwLTEpO31lbHNle3Nob3dTY3JlZW4oJ2hvbWVTY3JlZW4nKTt9Cn07CmRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdob21lQnRuJykub25jbGljayA9IGZ1bmN0aW9uKCl7CiAgc2hvd1NjcmVlbignaG9tZVNjcmVlbicpOyByZXNldEFsbCgpOwp9OwoKLy8gQnJlZWQgc2VhcmNoCnZhciBpbnAgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnYklucHV0Jyk7CnZhciBkcm9wID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2JEcm9wJyk7CnZhciBjbHIgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2xyQnRuJyk7CnZhciBiYWRnZSA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzQmFkZ2UnKTsKCmlucC5hZGRFdmVudExpc3RlbmVyKCdpbnB1dCcsIGZ1bmN0aW9uKCl7CiAgdmFyIHEgPSBpbnAudmFsdWUudHJpbSgpOwogIGNsci5jbGFzc0xpc3QudG9nZ2xlKCdzaG93JywgcS5sZW5ndGg+MCk7CiAgaWYoIXEpe2Ryb3AuY2xhc3NMaXN0LnJlbW92ZSgnb3BlbicpO2Ryb3AuaW5uZXJIVE1MPScnO3JldHVybjt9CiAgdmFyIHNmPUxBTkc9PT0nZW4nPydicmVlZF9lbic6TEFORz09PSdldCc/J2JyZWVkX2V0JzonYnJlZWQnOwogIHZhciByZXM9REFUQS5maWx0ZXIoZnVuY3Rpb24oYil7cmV0dXJuKGJbc2ZdfHxiLmJyZWVkKS50b0xvd2VyQ2FzZSgpLmluZGV4T2YocS50b0xvd2VyQ2FzZSgpKSE9PS0xO30pLnNsaWNlKDAsMzUpOwogIGRyb3AuaW5uZXJIVE1MPScnOwogIHZhciBfbnI9TEFORz09PSdlbic/J0JyZWVkIG5vdCBmb3VuZCc6TEFORz09PSdldCc/J1TDtXVndSBlaSBsZWl0dWQnOifQn9C+0YDQvtC00LAg0L3QtSDQvdCw0LnQtNC10L3QsCc7CiAgdmFyIF9udD1MQU5HPT09J2VuJz8iQ2FuJ3QgZmluZCB5b3VyIGJyZWVkPyI6TEFORz09PSdldCc/J0VpIGxlaWEgb21hIHTDtXVndT8nOifQndC1INC90LDRiNC70Lgg0YHQstC+0Y4g0L/QvtGA0L7QtNGDPyc7CiAgdmFyIF9ucz1MQU5HPT09J2VuJz8nQ29udGFjdCB1cyDigJQgd2Ugd2lsbCBoZWxwIHlvdSBjaG9vc2UgYSBzZXJ2aWNlJzpMQU5HPT09J2V0Jz8nVsO1dGtlIG1laWVnYSDDvGhlbmR1c3Qg4oCUIGFpdGFtZSB0ZWVudXNlIHZhbGlkYSc6J9Ch0LLRj9C20LjRgtC10YHRjCDRgSDQvdCw0LzQuCDQu9GO0LHRi9C8INGD0LTQvtCx0L3Ri9C8INGB0L/QvtGB0L7QsdC+0Lwg4oCUINC80Ysg0L/QvtC80L7QttC10Lwg0L/QvtC00L7QsdGA0LDRgtGMINGD0YHQu9GD0LPRgyc7CiAgaWYoIXJlcy5sZW5ndGgpe2Ryb3AuaW5uZXJIVE1MPSc8ZGl2IGNsYXNzPSJub3JlcyI+JytfbnIrJzwvZGl2PjxkaXYgY2xhc3M9Im5vLWJyZWVkLWJhbm5lciIgb25jbGljaz0ic2hvd1NjcmVlbihcJ2hvbWVTY3JlZW5cJykiPjxkaXYgY2xhc3M9Im5vLWJyZWVkLWJhbm5lci1pY29uIj7wn5C+PC9kaXY+PGRpdiBjbGFzcz0ibm8tYnJlZWQtYmFubmVyLXRleHQiPjxkaXYgY2xhc3M9Im5vLWJyZWVkLWJhbm5lci10aXRsZSI+JytfbnQrJzwvZGl2PjxkaXYgY2xhc3M9Im5vLWJyZWVkLWJhbm5lci1zdWIiPicrX25zKyc8L2Rpdj48L2Rpdj48ZGl2IGNsYXNzPSJuby1icmVlZC1iYW5uZXItYXJyb3ciPuKGkjwvZGl2PjwvZGl2Pic7fQogIGVsc2V7CiAgICByZXMuZm9yRWFjaChmdW5jdGlvbihiKXsKICAgICAgdmFyIGQ9ZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnZGl2Jyk7IGQuY2xhc3NOYW1lPSdkaXRlbSc7CiAgICAgIHZhciBibmFtZT1iW3NmXXx8Yi5icmVlZDsKICAgICAgdmFyIGlkeD1ibmFtZS50b0xvd2VyQ2FzZSgpLmluZGV4T2YocS50b0xvd2VyQ2FzZSgpKTsKICAgICAgZC5pbm5lckhUTUw9Ym5hbWUuc3Vic3RyaW5nKDAsaWR4KSsnPG1hcms+JytibmFtZS5zdWJzdHJpbmcoaWR4LGlkeCtxLmxlbmd0aCkrJzwvbWFyaz4nK2JuYW1lLnN1YnN0cmluZyhpZHgrcS5sZW5ndGgpOwogICAgICBkLm9uY2xpY2s9ZnVuY3Rpb24oKXtzZWxlY3RCcmVlZChiKTt9OwogICAgICBkcm9wLmFwcGVuZENoaWxkKGQpOwogICAgfSk7CiAgfQogIGRyb3AuY2xhc3NMaXN0LmFkZCgnb3BlbicpOwp9KTsKCmRvY3VtZW50LmFkZEV2ZW50TGlzdGVuZXIoJ2NsaWNrJyxmdW5jdGlvbihlKXsKICBpZighZS50YXJnZXQuY2xvc2VzdCgnLmJ3cmFwJykpZHJvcC5jbGFzc0xpc3QucmVtb3ZlKCdvcGVuJyk7Cn0pOwpjbHIub25jbGljayA9IHJlc2V0QnJlZWQ7CgpmdW5jdGlvbiBzZWxlY3RCcmVlZChiKXsKICBzZWxCcmVlZD1iOyBib29raW5nLmJyZWVkPWIuYnJlZWQ7CiAgaW5wLnZhbHVlPScnOyBjbHIuY2xhc3NMaXN0LnJlbW92ZSgnc2hvdycpOwogIGRyb3AuY2xhc3NMaXN0LnJlbW92ZSgnb3BlbicpOyBkcm9wLmlubmVySFRNTD0nJzsKICBiYWRnZS5pbm5lckhUTUw9Jyc7CiAgdmFyIGJGaWVsZD1MQU5HPT09J2VuJz8nYnJlZWRfZW4nOkxBTkc9PT0nZXQnPydicmVlZF9ldCc6J2JyZWVkJzsKICB2YXIgZGlzcEJyZWVkPWJbYkZpZWxkXXx8Yi5icmVlZDsKICBib29raW5nLmJyZWVkRGlzcGxheT1kaXNwQnJlZWQ7CiAgdmFyIGJuPWRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ3NwYW4nKTtibi5jbGFzc05hbWU9J2JuYW1lJztibi50ZXh0Q29udGVudD1kaXNwQnJlZWQ7CiAgdmFyIGNoZ1R4dD1MQU5HPT09J2VuJz8nQ2hhbmdlJzpMQU5HPT09J2V0Jz8nTXV1ZGEnOifQmNC30LzQtdC90LjRgtGMJzsKICB2YXIgYmM9ZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnc3BhbicpO2JjLmNsYXNzTmFtZT0nYmNoZyc7YmMudGV4dENvbnRlbnQ9Y2hnVHh0OwogIGJjLm9uY2xpY2s9cmVzZXRCcmVlZDsKICBiYWRnZS5hcHBlbmRDaGlsZChibik7YmFkZ2UuYXBwZW5kQ2hpbGQoYmMpOwogIGJhZGdlLmNsYXNzTGlzdC5hZGQoJ3Nob3cnKTsKICByZW5kZXJTdmNzKGIpOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdmNTZWMnKS5zdHlsZS5kaXNwbGF5PSdibG9jayc7CiAgICAvLyBBZGQgaW1wb3J0YW50IG5vdGUgaWYgbm90IGV4aXN0cwogICAgaWYoIWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdmNOb3RlJykpewogICAgICB2YXIgbm90ZT1kb2N1bWVudC5jcmVhdGVFbGVtZW50KCdkaXYnKTsKICAgICAgbm90ZS5pZD0nc3ZjTm90ZSc7CiAgICAgIG5vdGUuc3R5bGUuY3NzVGV4dD0nYm9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4wOCk7cGFkZGluZzoxNHB4IDE2cHg7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wMik7bWFyZ2luLXRvcDoxMnB4Oyc7CiAgICAgIHZhciBub3RlVGl0bGU9TEFORz09PSdlbic/J1BsZWFzZSBub3RlJzpMQU5HPT09J2V0Jz8nUGFuZ2UgdMOkaGVsZSc6J9CS0LDQttC90L4g0LfQvdCw0YLRjCc7CiAgICAgIHZhciBub3RlQm9keT1MQU5HPT09J2VuJz8nRmluYWwgcHJpY2UgZGVwZW5kcyBvbiBjb2F0IGNvbmRpdGlvbiBhbmQgcGV0IGJlaGF2aW91ci48YnI+RGVtYXR0aW5nIGZyb20gNSDigqwuPGJyPkFnZ3Jlc3NpdmUgYmVoYXZpb3VyIHN1cmNoYXJnZSBtYXkgYXBwbHk6ICs1MCUuJzpMQU5HPT09J2V0Jz8nTMO1cGxpayBoaW5kIHPDtWx0dWIga2FydmFzdGlrdSBzZWlzdW5kaXN0IGphIGxlbW1pa2xvb21hIGvDpGl0dW1pc2VzdC48YnI+S29sdHN1bml0ZSBsYWh0aWhhcnV0YW1pbmUgYWxhdGVzIDUg4oKsLjxicj5BZ3Jlc3NpaXZzZSBrw6RpdHVtaXNlIGtvcnJhbCB2w7VpYiBsaXNhbmR1ZGEgNTAlIGp1dXJkZWhpbmRsdXMuJzon0J7QutC+0L3Rh9Cw0YLQtdC70YzQvdCw0Y8g0YHRgtC+0LjQvNC+0YHRgtGMINC30LDQstC40YHQuNGCINC+0YIg0YHQvtGB0YLQvtGP0L3QuNGPINGI0LXRgNGB0YLQuCDQuCDQv9C+0LLQtdC00LXQvdC40Y8g0L/QuNGC0L7QvNGG0LAuPGJyPtCg0LDQt9Cx0L7RgCDQutC+0LvRgtGD0L3QvtCyIOKAlCDQvtGCIDUg4oKsLjxicj7Qn9GA0Lgg0LDQs9GA0LXRgdGB0LjQstC90L7QvCDQv9C+0LLQtdC00LXQvdC40Lgg0LzQvtC20LXRgiDQv9GA0LjQvNC10L3Rj9GC0YzRgdGPINC00L7Qv9C70LDRgtCwIDUwJS4nOwogICAgICBub3RlLmlubmVySFRNTD0nPGRpdiBzdHlsZT0iZm9udC1zaXplOjAuODM4cmVtO2xldHRlci1zcGFjaW5nOi4xNWVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtjb2xvcjojZmZmZmZmO21hcmdpbi1ib3R0b206OHB4O2ZvbnQtd2VpZ2h0OjYwMDtmb250LWZhbWlseTpcJ01vbnRzZXJyYXRcJyxzYW5zLXNlcmlmIj4nK25vdGVUaXRsZSsnPC9kaXY+PGRpdiBzdHlsZT0iZm9udC1zaXplOjEuMDI1cmVtO2NvbG9yOiNmZmZmZmY7bGluZS1oZWlnaHQ6MS44O2ZvbnQtZmFtaWx5OlwnTW9udHNlcnJhdFwnLHNhbnMtc2VyaWYiPicrbm90ZUJvZHkrJzwvZGl2Pic7CiAgICAgIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdmNTZWMnKS5hcHBlbmRDaGlsZChub3RlKTsKICAgIH0KICBmaWx0ZXJNYXN0ZXJzKCk7Cn0KCmZ1bmN0aW9uIHJlc2V0QnJlZWQoKXsKICBzZWxCcmVlZD1udWxsO2Jvb2tpbmcuYnJlZWQ9Jyc7Ym9va2luZy5zZXJ2aWNlPScnO2Jvb2tpbmcucHJpY2U9MDsKICBpbnAudmFsdWU9Jyc7Y2xyLmNsYXNzTGlzdC5yZW1vdmUoJ3Nob3cnKTsKICBiYWRnZS5jbGFzc0xpc3QucmVtb3ZlKCdzaG93Jyk7YmFkZ2UuaW5uZXJIVE1MPScnOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdmNTZWMnKS5zdHlsZS5kaXNwbGF5PSdub25lJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnc3ZjTGlzdCcpLmlubmVySFRNTD0nJzsKfQoKCnZhciBTVkNfVFJBTlNMQVRJT05TID0gewogICfQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCc6ICAgICAge2VuOidCYXNpYyBncm9vbScsICAgICAgZXQ6J1DDtWhpaG9vbGR1cyd9LAogICfQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCc6e2VuOidIeWdpZW5lIGdyb29tJywgICAgZXQ6J0jDvGdpZWVuaWhvb2xkdXMnfSwKICAn0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCc6ICB7ZW46J0Z1bGwgZ3Jvb20nLCAgICAgICAgZXQ6J1TDpGllbGlrIGhvb2xkdXMnfSwKICAn0KLRgNC40LzQvNC40L3Qsyc6ICAgICAgICAgIHtlbjonVHJpbW1pbmcnLCAgICAgICAgICBldDonVHJpbW1lcmltaW5lJ30sCiAgJ9Ct0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwJzogICB7ZW46J0V4cHJlc3Mgc2hlZCcsICAgICAgZXQ6J0tpaXJrYXJ2YXZhaGV0dXMnfSwKICAn0JLRi9GH0LXRgSc6ICAgICAgICAgICAgIHtlbjonQnJ1c2gtb3V0JywgICAgICAgICBldDonSGFyamFtaW5lJ30sCiAgJ9CS0YHRjyDQv9GA0L7Qs9GA0LDQvNC80LAnOiAgICAge2VuOidGdWxsIHByb2dyYW0nLCAgICAgIGV0OidLb2d1IHByb2dyYW1tJ30KfTsKdmFyIFNWQ19UQUdMSU5FX0kxOE49ewogIHJ1Onsn0JLRi9GH0LXRgSc6J9Ch0YLQvtC40LzQvtGB0YLRjCDQt9Cw0LLQuNGB0LjRgiDQvtGCINGB0L7RgdGC0L7Rj9C90LjRjyDRiNC10YDRgdGC0Lgg0Lgg0L7QsdGK0ZHQvNCwINGA0LDQsdC+0YInLCfQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCc6J9Cf0L7QtNGF0L7QtNC40YIg0LTQu9GPINC/0L7QtNC00LXRgNC20LDQvdC40Y8g0YfQuNGB0YLQvtGC0Ysg0LzQtdC20LTRgyDQv9GA0L7RhtC10LTRg9GA0LDQvNC4Jywn0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQnOifQlNC70Y8g0LrQvtC80YTQvtGA0YLQsCDQuCDQsNC60LrRg9GA0LDRgtC90L7RgdGC0Lgg0L/QuNGC0L7QvNGG0LAnLCfQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0Jzon0J/QvtC70L3Ri9C5INGD0YXQvtC0INGB0L4g0YHRgtGA0LjQttC60L7QuScsJ9Ct0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwJzon0J/QvtC80L7Qs9Cw0LXRgiDRg9C80LXQvdGM0YjQuNGC0Ywg0LrQvtC70LjRh9C10YHRgtCy0L4g0LvQuNC90Y/RjtGJ0LXQuSDRiNC10YDRgdGC0LgnLCfQotGA0LjQvNC80LjQvdCzJzon0JTQu9GPINC20LXRgdGC0LrQvtGI0LXRgNGB0YLQvdGL0YUg0L/QvtGA0L7QtCd9LAogIGVuOnsn0JLRi9GH0LXRgSc6J1ByaWNlIGRlcGVuZHMgb24gY29hdCBjb25kaXRpb24gYW5kIHZvbHVtZSBvZiB3b3JrJywn0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQnOidJZGVhbCBmb3IgbWFpbnRhaW5pbmcgY2xlYW5saW5lc3MgYmV0d2VlbiBmdWxsIGdyb29tcycsJ9CT0LjQs9C40LXQvdC40YfQtdGB0LrQuNC5INGD0YXQvtC0JzonRm9yIHlvdXIgcGV0XCdzIGNvbWZvcnQgYW5kIG5lYXRuZXNzJywn0JrQvtC80L/Qu9C10LrRgdC90YvQuSDRg9GF0L7QtCc6J0Z1bGwgZ3Jvb21pbmcgd2l0aCBoYWlyY3V0Jywn0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAnOidTaWduaWZpY2FudGx5IHJlZHVjZXMgc2hlZGRpbmcnLCfQotGA0LjQvNC80LjQvdCzJzonRm9yIHdpcmUtaGFpcmVkIGJyZWVkcyd9LAogIGV0Onsn0JLRi9GH0LXRgSc6J0hpbmQgc8O1bHR1YiBrYXJ2YXN0aWt1IHNlaXN1bmRpc3QgamEgdMO2w7ZtYWh1c3QnLCfQkdCw0LfQvtCy0YvQuSDRg9GF0L7QtCc6J1NvYmliIHB1aHR1c2UgaG9pZG1pc2VrcyBwcm90c2VkdXVyaWRlIHZhaGVsJywn0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQnOidMZW1taWtsb29tYSBtdWdhdnVzZWtzIGphIGtvcnJhc2hvaXVrcycsJ9Ca0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQnOidUw6RpZWxpayBob29sZHVzIGtvb3MgbMO1aWt1c2VnYScsJ9Ct0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwJzonVsOkaGVuZGFiIG9sdWxpc2VsdCBrYXJ2YWRlIGxhbmdlbWlzdCcsJ9Ci0YDQuNC80LzQuNC90LMnOidUcmFhdGthcnZhbGlzdGVsZSB0w7V1Z3VkZWxlJ30KfTsKdmFyIFNWQ19ERVNDX0kxOE49ewogIHJ1Onsn0JLRi9GH0LXRgSc6J9Cn0LjRgdGC0LrQsCDQs9C70LDQtywg0YPRiNC10LksINC/0L7QtNGB0YLRgNC40LPQsNC90LjQtSDQutC+0LPRgtC10LksINCy0YvRh9GR0YEgKNC00LvRjyDQutC+0YjQtdC6KScsJ9CR0LDQt9C+0LLRi9C5INGD0YXQvtC0Jzon0JzRi9GC0YzRkSDQv9GA0L7RhNC10YHRgdC40L7QvdCw0LvRjNC90YvQvNC4INGB0YDQtdC00YHRgtCy0LDQvNC4LCDQtNC10LvQuNC60LDRgtC90LDRjyDRgdGD0YjQutCwJywn0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQnOifQodGC0YDQuNC20LrQsCDQutC+0LPRgtC10LksINGH0LjRgdGC0LrQsCDRg9GI0LXQuSDQuCDQs9C70LDQtywg0LrRg9C/0LDQvdC40LUsINGB0YPRiNC60LAsINGD0YXQvtC0INC30LAg0LvQsNC/0LrQsNC80Lgg0Lgg0YfRg9Cy0YHRgtCy0LjRgtC10LvRjNC90YvQvNC4INC30L7QvdCw0LzQuCcsJ9Ca0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQnOifQodGC0YDQuNC20LrQsCDQutC+0LPRgtC10LksINGH0LjRgdGC0LrQsCDRg9GI0LXQuSDQuCDQs9C70LDQtywg0LrRg9C/0LDQvdC40LUsINGB0YPRiNC60LAsINGD0YXQvtC0INC30LAg0LvQsNC/0LrQsNC80Lgg0Lgg0YfRg9Cy0YHRgtCy0LjRgtC10LvRjNC90YvQvNC4INC30L7QvdCw0LzQuCwg0LzQvtC00LXQu9GM0L3QsNGPINGB0YLRgNC40LbQutCwJywn0K3QutGB0L/RgNC10YHRgS3Qu9C40L3RjNC60LAnOifQnNGL0YLRjNGRLCDRgdGD0YjQutCwLCDRg9GF0L7QtCDQt9CwINGI0LXRgNGB0YLRjNGOLCDQvNCw0YHQutCwLCDQv9C+0LTRgdGC0YDQuNCz0LDQvdC40LUg0LrQvtCz0YLQtdC5LCDRh9C40YHRgtC60LAg0YPRiNC10Lkg0Lgg0LPQu9Cw0LcsINGD0YXQvtC0INC30LAg0LvQsNC/0LDQvNC4INC4INC30L7QvdCw0LzQuCDRgtGA0LXQsdGD0Y7RidC40LzQuCDQvtGB0L7QsdC+0LPQviDQstC90LjQvNCw0L3QuNGPJywn0KLRgNC40LzQvNC40L3Qsyc6J9CS0YvRidC40L/Ri9Cy0LDQvdC40LUg0YHRgtCw0YDQvtCz0L4g0YHQu9C+0Y8g0YjQtdGA0YHRgtC4LCDQvNGL0YLRjNGRLCDRgdGD0YjQutCwLCDRgdGC0YDQuNC20LrQsCDQutC+0LPRgtC10LksINGH0LjRgdGC0LrQsCDRg9GI0LXQuSDQuCDQs9C70LDQtywg0L7RhNC+0YDQvNC70LXQvdC40LUg0YjQtdGA0YHRgtC4Jywn0JLRgdGPINC/0YDQvtCz0YDQsNC80LzQsCc6J9Cf0JXQoNCS0KvQmSDQktCY0JfQmNCiICgyMC0zMCDQvNC40L0pIOKAlCAyMCDigqxcbuKAoiDQt9C90LDQutC+0LzRgdGC0LLQviDRgdC+INGB0YLQvtC70L7QvCDQuCDQuNC90YHRgtGA0YPQvNC10L3RgtCw0LzQuFxu4oCiINC70ZHQs9C60L7QtSDQstGL0YfRkdGB0YvQstCw0L3QuNC1XG7igKIg0LfQstGD0LrQuCDRhNC10L3QsCDQuCDQu9C10LPQutCw0Y8g0L/RgNC+0LTRg9Cy0LrQsFxu4oCiINC+0YHQstC10LbQtdC90LjQtSDQs9C70LDQt9C+0Log0Lgg0YPRiNC10LpcbuKAoiDQutC+0LPQvtGC0LrQuFxu4oCiINCy0LrRg9GB0L3Rj9GI0LrQuCDQuCDRgdC/0L7QutC+0LnQvdCw0Y8g0LDQtNCw0L/RgtCw0YbQuNGPXG5cbtCS0KLQntCg0J7QmSDQktCY0JfQmNCiICg0MC02MCDQvNC40L0pIOKAlCAzNSDigqxcbuKAoiDQv9C10YDQstC+0LUg0LrRg9C/0LDQvdC40LUg0Lgg0YHRg9GI0LrQsFxu4oCiINCy0YvRh9GR0YHRi9Cy0LDQvdC40LVcbuKAoiDQs9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtFxu4oCiINC90LXQsdC+0LvRjNGI0LDRjyDRgdGC0YDQuNC20LrQsCAvINC60L7RgNGA0LXQutGG0LjRjyDRiNC10YDRgdGC0LggKNC/0YDQuCDQvdC10L7QsdGF0L7QtNC40LzQvtGB0YLQuClcbuKAoiDQt9Cw0LrRgNC10L/Qu9C10L3QuNC1INC/0L7Qu9C+0LbQuNGC0LXQu9GM0L3QvtCz0L4g0L7Qv9GL0YLQsCd9LAogIGVuOnsn0JLRi9GH0LXRgSc6J0V5ZSBhbmQgZWFyIGNsZWFuaW5nLCBuYWlsIHRyaW1taW5nLCBicnVzaGluZyAoZm9yIGNhdHMpJywn0JHQsNC30L7QstGL0Lkg0YPRhdC+0LQnOidXYXNoaW5nIHdpdGggcHJvZmVzc2lvbmFsIHByb2R1Y3RzLCBnZW50bGUgZHJ5aW5nJywn0JPQuNCz0LjQtdC90LjRh9C10YHQutC40Lkg0YPRhdC+0LQnOidOYWlsIHRyaW1taW5nLCBlYXIgYW5kIGV5ZSBjbGVhbmluZywgYmF0aGluZywgZHJ5aW5nLCBwYXcgYW5kIHNlbnNpdGl2ZSBhcmVhIGNhcmUnLCfQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0JzonTmFpbCB0cmltbWluZywgZWFyIGFuZCBleWUgY2xlYW5pbmcsIGJhdGhpbmcsIGRyeWluZywgcGF3IGFuZCBzZW5zaXRpdmUgYXJlYSBjYXJlLCBzdHlsaW5nIGhhaXJjdXQnLCfQrdC60YHQv9GA0LXRgdGBLdC70LjQvdGM0LrQsCc6J1dhc2hpbmcsIGRyeWluZywgY29hdCBjYXJlLCBtYXNrLCBuYWlsIHRyaW1taW5nLCBlYXIgYW5kIGV5ZSBjbGVhbmluZywgcGF3IGFuZCBzcGVjaWFsIGFyZWEgY2FyZScsJ9Ci0YDQuNC80LzQuNC90LMnOidSZW1vdmluZyBvbGQgY29hdCBsYXllciwgd2FzaGluZywgZHJ5aW5nLCBuYWlsIHRyaW1taW5nLCBlYXIgYW5kIGV5ZSBjbGVhbmluZywgY29hdCBzdHlsaW5nJywn0JLRgdGPINC/0YDQvtCz0YDQsNC80LzQsCc6J0ZJUlNUIFZJU0lUICgyMC0zMCBtaW4pIOKAlCDigqwyMFxu4oCiIGdldHRpbmcgdXNlZCB0byB0aGUgdGFibGUgYW5kIHRvb2xzXG7igKIgZ2VudGxlIGJydXNoaW5nXG7igKIgZHJ5ZXIgc291bmRzIGFuZCBsaWdodCBhaXJmbG93XG7igKIgZXllIGFuZCBlYXIgcmVmcmVzaFxu4oCiIG5haWwgdHJpbVxu4oCiIHRyZWF0cyBhbmQgY2FsbSBhZGFwdGF0aW9uXG5cblNFQ09ORCBWSVNJVCAoNDAtNjAgbWluKSDigJQg4oKsMzVcbuKAoiBmaXJzdCBiYXRoIGFuZCBkcnlpbmdcbuKAoiBicnVzaGluZ1xu4oCiIGh5Z2llbmUgY2FyZVxu4oCiIGxpZ2h0IHRyaW0gLyBjb2F0IGFkanVzdG1lbnQgKGlmIG5lZWRlZClcbuKAoiByZWluZm9yY2luZyB0aGUgcG9zaXRpdmUgZXhwZXJpZW5jZSd9LAogIGV0Onsn0JLRi9GH0LXRgSc6J1NpbG1hZGUgamEga8O1cnZhZGUgcHVoYXN0YW1pbmUsIGvDvMO8bnRlIGzDtWlrYW1pbmUsIGhhcmphbWluZSAoa2Fzc2lkZWxlKScsJ9CR0LDQt9C+0LLRi9C5INGD0YXQvtC0JzonUGVzZW1pbmUgcHJvZmVzc2lvbmFhbHNldGUgdmFoZW5kaXRlZ2EsIMO1cm4ga3VpdmF0YW1pbmUnLCfQk9C40LPQuNC10L3QuNGH0LXRgdC60LjQuSDRg9GF0L7QtCc6J0vDvMO8bnRlIGzDtWlrYW1pbmUsIGvDtXJ2YWRlIGphIHNpbG1hZGUgcHVoYXN0YW1pbmUsIHBlc2VtaW5lLCBrdWl2YXRhbWluZSwga8OkcHBhZGUgamEgdHVuZGxpa2UgcGlpcmtvbmRhZGUgaG9vbGR1cycsJ9Ca0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQnOidLw7zDvG50ZSBsw7Vpa2FtaW5lLCBrw7VydmFkZSBqYSBzaWxtYWRlIHB1aGFzdGFtaW5lLCBwZXNlbWluZSwga3VpdmF0YW1pbmUsIGvDpHBwYWRlIGphIHR1bmRsaWtlIHBpaXJrb25kYWRlIGhvb2xkdXMsIG1vZGVsbMO1aWt1cycsJ9Ct0LrRgdC/0YDQtdGB0YEt0LvQuNC90YzQutCwJzonUGVzZW1pbmUsIGt1aXZhdGFtaW5lLCBrYXJ2YXN0aWt1IGhvb2xkdXMsIG1hc2ssIGvDvMO8bnRlIGzDtWlrYW1pbmUsIGvDtXJ2YWRlIGphIHNpbG1hZGUgcHVoYXN0YW1pbmUsIGvDpHBwYWRlIGphIGVyaWxpc3RlIHBpaXJrb25kYWRlIGhvb2xkdXMnLCfQotGA0LjQvNC80LjQvdCzJzonVmFuYSBrYXJ2YWtpaGkgZWVtYWxkYW1pbmUsIHBlc2VtaW5lLCBrdWl2YXRhbWluZSwga8O8w7xudGUgbMO1aWthbWluZSwga8O1cnZhZGUgamEgc2lsbWFkZSBwdWhhc3RhbWluZSwga2FydmFzdGlrdSBrdWp1bmRhbWluZScsJ9CS0YHRjyDQv9GA0L7Qs9GA0LDQvNC80LAnOidFU0lNRU5FIEvDnExBU1RVUyAoMjAtMzAgbWluKSDigJQgMjAg4oKsXG7igKIgdHV0dnVtaW5lIGxhdWFnYSBqYSB0w7bDtnJpaXN0YWRlZ2FcbuKAoiBrZXJnZSBoYXJqYW1pbmVcbuKAoiBmw7bDtm5paGVsaWQgamEga2VyZ2Ugw7VodXZvb2xcbuKAoiBzaWxtYWRlIGphIGvDtXJ2YWRlIHbDpHJza2VuZHVzXG7igKIga8O8w7xudGUgbMO1aWthbWluZVxu4oCiIG1haXVzZWQgamEgcmFodWxpayBrb2hhbmVtaW5lXG5cblRFSU5FIEvDnExBU1RVUyAoNDAtNjAgbWluKSDigJQgMzUg4oKsXG7igKIgZXNpbWVuZSB2YW5uaXRhbWluZSBqYSBrdWl2YXRhbWluZVxu4oCiIGhhcmphbWluZVxu4oCiIGjDvGdpZWVuaWhvb2xkdXNcbuKAoiBrZXJnZSBsw7Vpa3VzIC8ga2FydmEga29ycmlnZWVyaW1pbmUgKHZhamFkdXNlbClcbuKAoiBwb3NpdGlpdnNlIGtvZ2VtdXNlIGtpbm5pc3RhbWluZSd9Cn07CnZhciBTVkNfREVTQ19DQVRfQ09NUExFWD17CiAgcnU6J9Cc0YvRgtGM0ZEsINGB0YPRiNC60LAsINCy0YvRh9GR0YHRi9Cy0LDQvdC40LUsINGB0YLRgNC40LbQutCwINC60L7Qs9GC0LXQuSwg0LAg0YLQsNC60LbQtSDQvtCx0YDQsNCx0L7RgtC60LAg0LPQu9Cw0Lcg0Lgg0YPRiNC10LonLAogIGVuOidXYXNoaW5nLCBkcnlpbmcsIGJydXNoaW5nLCBuYWlsIHRyaW1taW5nLCBhbmQgZXllIGFuZCBlYXIgY2FyZScsCiAgZXQ6J1Blc2VtaW5lLCBrdWl2YXRhbWluZSwgaGFyamFtaW5lLCBrw7zDvG50ZSBsw7Vpa2FtaW5lIG5pbmcgc2lsbWFkZSBqYSBrw7VydmFkZSBob29sZHVzJwp9OwpmdW5jdGlvbiBnZXRTdmNUYWcobmFtZSl7cmV0dXJuKFNWQ19UQUdMSU5FX0kxOE5bTEFOR10mJlNWQ19UQUdMSU5FX0kxOE5bTEFOR11bbmFtZV0pfHxTVkNfVEFHTElORV9JMThOLnJ1W25hbWVdfHwnJzt9CmZ1bmN0aW9uIGdldFN2Y0Rlc2MobmFtZSl7CiAgaWYobmFtZT09PSfQmtC+0LzQv9C70LXQutGB0L3Ri9C5INGD0YXQvtC0JyAmJiBib29raW5nLmJyZWVkICYmIGJvb2tpbmcuYnJlZWQuaW5kZXhPZign0JrQvtGI0LrQsCcpPT09MCl7CiAgICB2YXIgZD1TVkNfREVTQ19DQVRfQ09NUExFWFtMQU5HXXx8U1ZDX0RFU0NfQ0FUX0NPTVBMRVgucnU7CiAgICByZXR1cm4gZDsKICB9CiAgcmV0dXJuKFNWQ19ERVNDX0kxOE5bTEFOR10mJlNWQ19ERVNDX0kxOE5bTEFOR11bbmFtZV0pfHxTVkNfREVTQ19JMThOLnJ1W25hbWVdfHwnJzsKfQoKZnVuY3Rpb24gcmVuZGVyU3ZjcyhiKXsKICB2YXIgbGJsRWw9ZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ3N0ZXAyTGJsRWwnKTsKICBpZihsYmxFbCl7CiAgICB2YXIgYmFzZUxibD0oVFtMQU5HXSYmVFtMQU5HXS5zdGVwMl9sYmwpfHwnMDIgwrcg0KPRgdC70YPQs9CwJzsKICAgIGxibEVsLnRleHRDb250ZW50PShiLmJyZWVkPT09J9Cp0LXQvdC60LgnKT8oYmFzZUxibCsnIFB1cHB5IFN0YXInKTpiYXNlTGJsOwogIH0KICB2YXIgbGlzdD1kb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnc3ZjTGlzdCcpO2xpc3QuaW5uZXJIVE1MPScnOwogIE9iamVjdC5lbnRyaWVzKGIuc2VydmljZXMpLmZvckVhY2goZnVuY3Rpb24oa3YpewogICAgdmFyIG5hbWU9a3ZbMF0scHJpY2U9a3ZbMV07CgogICAgdmFyIGRpc3BsYXlOYW1lPShMQU5HIT09J3J1JyYmU1ZDX1RSQU5TTEFUSU9OU1tuYW1lXSk/U1ZDX1RSQU5TTEFUSU9OU1tuYW1lXVtMQU5HXTpuYW1lOwogICAgdmFyIGJ0bj1kb2N1bWVudC5jcmVhdGVFbGVtZW50KCdidXR0b24nKTtidG4uY2xhc3NOYW1lPSdzdmJ0bic7CiAgICB2YXIgcm93PWRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2RpdicpO3Jvdy5jbGFzc05hbWU9J3N2YnRuLXJvdyc7CiAgICB2YXIgbnM9ZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnc3BhbicpO25zLmNsYXNzTmFtZT0nc3ZidG4tbmFtZSc7bnMudGV4dENvbnRlbnQ9ZGlzcGxheU5hbWU7CiAgICB2YXIgcHM9ZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnc3BhbicpO3BzLmNsYXNzTmFtZT0nc3ZidG4tcHJpY2UnO3BzLnRleHRDb250ZW50PXByaWNlKycg4oKsJzsKICAgIHJvdy5hcHBlbmRDaGlsZChucyk7cm93LmFwcGVuZENoaWxkKHBzKTsKICAgIGJ0bi5hcHBlbmRDaGlsZChyb3cpOwogICAgdmFyIGRlc2M9Z2V0U3ZjRGVzYyhuYW1lKTsKICAgIGlmKGRlc2Mpe3ZhciBkcz1kb2N1bWVudC5jcmVhdGVFbGVtZW50KCdzcGFuJyk7ZHMuY2xhc3NOYW1lPSdzdmJ0bi1kZXNjJztkcy50ZXh0Q29udGVudD1kZXNjO2J0bi5hcHBlbmRDaGlsZChkcyk7fQogICAgdmFyIHRhZz1nZXRTdmNUYWcobmFtZSk7CiAgICBpZih0YWcpe3ZhciB0cz1kb2N1bWVudC5jcmVhdGVFbGVtZW50KCdzcGFuJyk7dHMuY2xhc3NOYW1lPSdzdmJ0bi10YWcnO3RzLnRleHRDb250ZW50PXRhZztidG4uYXBwZW5kQ2hpbGQodHMpO30KICAgIGJ0bi5vbmNsaWNrPWZ1bmN0aW9uKCl7CiAgICAgIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5zdmJ0bicpLmZvckVhY2goZnVuY3Rpb24oYil7Yi5jbGFzc0xpc3QucmVtb3ZlKCdhY3RpdmUnKTt9KTsKICAgICAgYnRuLmNsYXNzTGlzdC5hZGQoJ2FjdGl2ZScpOwogICAgICBib29raW5nLnNlcnZpY2U9bmFtZTtib29raW5nLnByaWNlPXByaWNlOwogICAgICBmaWx0ZXJNYXN0ZXJzKCk7CiAgICAgIHNldFRpbWVvdXQoZnVuY3Rpb24oKXtnb1N0ZXAoMik7fSwzMDApOwogICAgfTsKICAgIGxpc3QuYXBwZW5kQ2hpbGQoYnRuKTsKICB9KTsKfQoKLy8gTWFzdGVycwpmdW5jdGlvbiBzb3J0TWFzdGVyc0J5QXZhaWxhYmlsaXR5KCl7CiAgdmFyIG5vdyA9IG5ldyBEYXRlKCk7CiAgdmFyIG1vbnRoID0gbm93LmdldE1vbnRoKCkrMSwgeWVhciA9IG5vdy5nZXRGdWxsWWVhcigpOwogIHZhciBiYXNlT3JkZXIgPSBbJ9Ci0LDRgtGM0Y/QvdCwJywn0JDQu9C10LrRgdCw0L3QtNGA0LAnLCfQmtGB0LXQvdC40Y8nLCfQkNC90L3QsCcsJ9CQ0LvQuNGB0LAnLCfQmtGA0LjRgdGC0LjQvdCwJ107CiAgdmFyIHZpc2libGVCdG5zID0gQXJyYXkucHJvdG90eXBlLmZpbHRlci5jYWxsKGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5tYnRuJyksIGZ1bmN0aW9uKGIpeyByZXR1cm4gYi5zdHlsZS5kaXNwbGF5ICE9PSAnbm9uZSc7IH0pOwogIGlmKCF2aXNpYmxlQnRucy5sZW5ndGgpIHJldHVybjsKICBQcm9taXNlLmFsbCh2aXNpYmxlQnRucy5tYXAoZnVuY3Rpb24oYnRuKXsKICAgIHZhciBtYXN0ZXIgPSBidG4uZ2V0QXR0cmlidXRlKCdkYXRhLW1hc3RlcicpOwogICAgcmV0dXJuIGZldGNoKHdpbmRvdy5sb2NhdGlvbi5vcmlnaW4gKyAnL2FwaS9hdmFpbGFibGVfZGF5cz9tb250aD0nICsgbW9udGggKyAnJnllYXI9JyArIHllYXIgKyAnJm1hc3Rlcj0nICsgZW5jb2RlVVJJQ29tcG9uZW50KG1hc3RlcikpCiAgICAgIC50aGVuKGZ1bmN0aW9uKHIpeyByZXR1cm4gci5qc29uKCk7IH0pCiAgICAgIC50aGVuKGZ1bmN0aW9uKGQpeyByZXR1cm4ge2J0bjogYnRuLCBtYXN0ZXI6IG1hc3RlciwgY291bnQ6IChkLmF2YWlsYWJsZXx8W10pLmxlbmd0aH07IH0pCiAgICAgIC5jYXRjaChmdW5jdGlvbigpeyByZXR1cm4ge2J0bjogYnRuLCBtYXN0ZXI6IG1hc3RlciwgY291bnQ6IC0xfTsgfSk7CiAgfSkpLnRoZW4oZnVuY3Rpb24ocmVzdWx0cyl7CiAgICByZXN1bHRzLnNvcnQoZnVuY3Rpb24oYSxiKXsKICAgICAgaWYoYi5jb3VudCAhPT0gYS5jb3VudCkgcmV0dXJuIGIuY291bnQgLSBhLmNvdW50OwogICAgICByZXR1cm4gYmFzZU9yZGVyLmluZGV4T2YoYS5tYXN0ZXIpIC0gYmFzZU9yZGVyLmluZGV4T2YoYi5tYXN0ZXIpOwogICAgfSk7CiAgICByZXN1bHRzLmZvckVhY2goZnVuY3Rpb24ociwgaSl7IHIuYnRuLnN0eWxlLm9yZGVyID0gaTsgfSk7CiAgfSk7Cn0KCmZ1bmN0aW9uIGZpbHRlck1hc3RlcnMoKXsKICB2YXIgaXNDYXQgPSBib29raW5nLmJyZWVkICYmIGJvb2tpbmcuYnJlZWQuaW5kZXhPZign0JrQvtGI0LrQsCcpID09PSAwOwogIHZhciBicmVlZCA9IGJvb2tpbmcuYnJlZWQgfHwgJyc7CiAgdmFyIGlzQ2F0Q29tcGxleCA9IGlzQ2F0ICYmIGJvb2tpbmcuc2VydmljZSA9PT0gJ9Ca0L7QvNC/0LvQtdC60YHQvdGL0Lkg0YPRhdC+0LQnOwogIHZhciBhbm5hRXhjbHVkZSA9IFsn0JzQsNC70YzRgtC40L/RgycsJ9Cf0YPQtNC10LvRjCcsJ9CZ0L7RgNC6Jywn0JHQuNGI0L7QvScsJ9CR0L7Qu9C+0L3QutCwJywn0JzQsNC70YzRgtC40LnRgdC60LDRjyddOwogIHZhciBpc0FubmFCcmVlZCA9IGJyZWVkICYmICFhbm5hRXhjbHVkZS5zb21lKGZ1bmN0aW9uKGIpeyByZXR1cm4gYnJlZWQuaW5kZXhPZihiKSAhPT0gLTE7IH0pOwogIHZhciBhbGV4YW5kcmFFeGNsdWRlID0gWyfQpNC+0LrRgdGC0LXRgNGM0LXRgCcsJ9Cm0LLQtdGA0LPRiNC90LDRg9GG0LXRgCddOwogIHZhciBpc0FsZXhhbmRyYUJyZWVkID0gIWFsZXhhbmRyYUV4Y2x1ZGUuc29tZShmdW5jdGlvbihiKXsgcmV0dXJuIGJyZWVkLmluZGV4T2YoYikgIT09IC0xOyB9KTsKICB2YXIga3NlbmlhRXhjbHVkZSA9IFsn0J/Rg9C00LXQu9GMJywn0JzQsNC70YzRgtC40L/RgycsJ9CZ0L7RgNC6Jywn0JHQvtC70L7QvdC60LAnXTsKICB2YXIgaXNLc2VuaWFCcmVlZCA9ICFrc2VuaWFFeGNsdWRlLnNvbWUoZnVuY3Rpb24oYil7IHJldHVybiBicmVlZC5pbmRleE9mKGIpICE9PSAtMTsgfSk7CiAgZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLm1idG4nKS5mb3JFYWNoKGZ1bmN0aW9uKGJ0bil7CiAgICB2YXIgbWFzdGVyID0gYnRuLmdldEF0dHJpYnV0ZSgnZGF0YS1tYXN0ZXInKTsKICAgIHZhciBpc1RyaW1taW5nID0gYm9va2luZy5zZXJ2aWNlID09PSAn0KLRgNC40LzQvNC40L3Qsyc7CiAgICBpZihpc0NhdENvbXBsZXgpewogICAgICBidG4uc3R5bGUuZGlzcGxheSA9IChtYXN0ZXIgPT09ICfQotCw0YLRjNGP0L3QsCcgfHwgbWFzdGVyID09PSAn0JrRgdC10L3QuNGPJykgPyAnJyA6ICdub25lJzsKICAgICAgcmV0dXJuOwogICAgfQogICAgaWYobWFzdGVyID09PSAn0JDQu9C40YHQsCcpewogICAgICBidG4uc3R5bGUuZGlzcGxheSA9IGlzQ2F0ID8gJycgOiAnbm9uZSc7CiAgICB9IGVsc2UgaWYobWFzdGVyID09PSAn0JDQvdC90LAnKXsKICAgICAgYnRuLnN0eWxlLmRpc3BsYXkgPSAoaXNBbm5hQnJlZWQgJiYgIWlzVHJpbW1pbmcpID8gJycgOiAnbm9uZSc7CiAgICB9IGVsc2UgaWYobWFzdGVyID09PSAn0JDQu9C10LrRgdCw0L3QtNGA0LAnKXsKICAgICAgYnRuLnN0eWxlLmRpc3BsYXkgPSAoaXNBbGV4YW5kcmFCcmVlZCAmJiAhaXNUcmltbWluZyAmJiAhaXNDYXQpID8gJycgOiAnbm9uZSc7CiAgICB9IGVsc2UgaWYobWFzdGVyID09PSAn0JrRgdC10L3QuNGPJyl7CiAgICAgIGJ0bi5zdHlsZS5kaXNwbGF5ID0gaXNLc2VuaWFCcmVlZCA/ICcnIDogJ25vbmUnOwogICAgfSBlbHNlIGlmKGlzVHJpbW1pbmcpewogICAgICBidG4uc3R5bGUuZGlzcGxheSA9ICdub25lJzsKICAgIH0gZWxzZSB7CiAgICAgIGJ0bi5zdHlsZS5kaXNwbGF5ID0gJyc7CiAgICB9CiAgfSk7CiAgc29ydE1hc3RlcnNCeUF2YWlsYWJpbGl0eSgpOwp9Cgpkb2N1bWVudC5xdWVyeVNlbGVjdG9yQWxsKCcubWJ0bicpLmZvckVhY2goZnVuY3Rpb24oYnRuKXsKICBidG4ub25jbGljaz1mdW5jdGlvbigpewogICAgZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLm1idG4nKS5mb3JFYWNoKGZ1bmN0aW9uKGIpe2IuY2xhc3NMaXN0LnJlbW92ZSgnYWN0aXZlJyk7fSk7CiAgICBidG4uY2xhc3NMaXN0LmFkZCgnYWN0aXZlJyk7CiAgICBib29raW5nLm1hc3Rlcj1idG4uZ2V0QXR0cmlidXRlKCdkYXRhLW1hc3RlcicpOwogICAgc2V0VGltZW91dChmdW5jdGlvbigpe2dvU3RlcCg0KTtidWlsZENhbCgpO30sMzAwKTsKICB9Owp9KTsKCi8vIEdyb29tIGhpc3RvcnkKZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLmdidG4nKS5mb3JFYWNoKGZ1bmN0aW9uKGJ0bil7CiAgYnRuLm9uY2xpY2s9ZnVuY3Rpb24oKXsKICAgIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5nYnRuJykuZm9yRWFjaChmdW5jdGlvbihiKXtiLmNsYXNzTGlzdC5yZW1vdmUoJ2FjdGl2ZScpO30pOwogICAgYnRuLmNsYXNzTGlzdC5hZGQoJ2FjdGl2ZScpOwogICAgYm9va2luZy5ncm9vbUhpc3Rvcnk9YnRuLmdldEF0dHJpYnV0ZSgnZGF0YS12YWwnKTsKICAgIGZpbHRlck1hc3RlcnMoKTsKICAgIHNldFRpbWVvdXQoZnVuY3Rpb24oKXtnb1N0ZXAoMyk7fSwzMDApOwogIH07Cn0pOwoKLy8gQ2FsZW5kYXIKZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ3ByZXZNJykub25jbGljaz1mdW5jdGlvbigpe2NNLS07aWYoY008MCl7Y009MTE7Y1ktLTt9YnVpbGRDYWwoKTt9Owpkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbmV4dE0nKS5vbmNsaWNrPWZ1bmN0aW9uKCl7Y00rKztpZihjTT4xMSl7Y009MDtjWSsrO31idWlsZENhbCgpO307Cgp2YXIgYXZhaWxhYmxlRGF5cyA9IFtdOwoKZnVuY3Rpb24gbG9hZEF2YWlsYWJsZURheXMoKSB7CiAgdmFyIG1hc3RlciA9IGJvb2tpbmcubWFzdGVyOwogIGlmICghbWFzdGVyKSByZXR1cm47CiAgYXZhaWxhYmxlRGF5cyA9IFtdOwogIHZhciBsb2FkaW5nRWwgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2FsTG9hZGluZycpOwogIGlmKGxvYWRpbmdFbCkgbG9hZGluZ0VsLmNsYXNzTGlzdC5hZGQoJ3Nob3cnKTsKICBmZXRjaCh3aW5kb3cubG9jYXRpb24ub3JpZ2luICsgJy9hcGkvYXZhaWxhYmxlX2RheXM/bW9udGg9JyArIChjTSsxKSArICcmeWVhcj0nICsgY1kgKyAnJm1hc3Rlcj0nICsgZW5jb2RlVVJJQ29tcG9uZW50KG1hc3RlcikpCiAgICAudGhlbihmdW5jdGlvbihyKXsgcmV0dXJuIHIuanNvbigpOyB9KQogICAgLnRoZW4oZnVuY3Rpb24oZGF0YSl7CiAgICAgIGF2YWlsYWJsZURheXMgPSBkYXRhLmF2YWlsYWJsZSB8fCBbXTsKICAgICAgbWFya0F2YWlsYWJsZURheXMoKTsKICAgIH0pCiAgICAuY2F0Y2goZnVuY3Rpb24oKXsgYXZhaWxhYmxlRGF5cyA9IFtdOyB9KQogICAgLmZpbmFsbHkoZnVuY3Rpb24oKXsKICAgICAgaWYobG9hZGluZ0VsKSBsb2FkaW5nRWwuY2xhc3NMaXN0LnJlbW92ZSgnc2hvdycpOwogICAgfSk7Cn0KCmZ1bmN0aW9uIG1hcmtBdmFpbGFibGVEYXlzKCkgewogIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5jZCcpLmZvckVhY2goZnVuY3Rpb24oYyl7aWYoIWMuY2xhc3NMaXN0LmNvbnRhaW5zKCdkaXMnKSljLmNsYXNzTGlzdC5yZW1vdmUoJ3NlbCcpO30pOwogIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5jZDpub3QoLmRpcyk6bm90KC5jZG4pOm5vdCgucGFkKScpLmZvckVhY2goZnVuY3Rpb24oZWwpIHsKICAgIHZhciBkYXkgPSBlbC50ZXh0Q29udGVudC50cmltKCk7CiAgICBpZiAoIWRheSB8fCBpc05hTihwYXJzZUludChkYXkpKSkgcmV0dXJuOwogICAgdmFyIGRhdGVTdHIgPSBTdHJpbmcocGFyc2VJbnQoZGF5KSkucGFkU3RhcnQoMiwnMCcpICsgJy4nICsgU3RyaW5nKGNNKzEpLnBhZFN0YXJ0KDIsJzAnKSArICcuJyArIGNZOwogICAgaWYgKGF2YWlsYWJsZURheXMuaW5kZXhPZihkYXRlU3RyKSAhPT0gLTEpIHsKICAgICAgZWwuY2xhc3NMaXN0LmFkZCgnYXZhaWwnKTsKICAgICAgZWwuY2xhc3NMaXN0LnJlbW92ZSgnYnVzeScpOwogICAgfSBlbHNlIHsKICAgICAgZWwuY2xhc3NMaXN0LmFkZCgnYnVzeScpOwogICAgICBlbC5jbGFzc0xpc3QucmVtb3ZlKCdhdmFpbCcpOwogICAgfQogIH0pOwp9CgpmdW5jdGlvbiBidWlsZENhbCgpewogIGxvYWRBdmFpbGFibGVEYXlzKCk7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2NhbE0nKS50ZXh0Q29udGVudD1NT05USFNbY01dKycgJytjWTsKICBib29raW5nLmRhdGU9Jyc7IGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5jZCcpLmZvckVhY2goZnVuY3Rpb24oYyl7Yy5jbGFzc0xpc3QucmVtb3ZlKCdzZWwnKTtjLmNsYXNzTGlzdC5yZW1vdmUoJ2F2YWlsJyk7Yy5jbGFzc0xpc3QucmVtb3ZlKCdidXN5Jyk7fSk7CiAgdmFyIGc9ZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2NhbEcnKTtnLmlubmVySFRNTD0nJzsKICBbJ9Cf0L0nLCfQktGCJywn0KHRgCcsJ9Cn0YInLCfQn9GCJywn0KHQsScsJ9CS0YEnXS5mb3JFYWNoKGZ1bmN0aW9uKGQpewogICAgdmFyIGVsPWRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2RpdicpO2VsLmNsYXNzTmFtZT0nY2RuJztlbC50ZXh0Q29udGVudD1kO2cuYXBwZW5kQ2hpbGQoZWwpOwogIH0pOwogIHZhciBmaXJzdD1uZXcgRGF0ZShjWSxjTSwxKS5nZXREYXkoKTsKICB2YXIgZGF5cz1uZXcgRGF0ZShjWSxjTSsxLDApLmdldERhdGUoKTsKICB2YXIgc3RhcnQ9Zmlyc3Q9PT0wPzY6Zmlyc3QtMTsKICB2YXIgdG9kYXk9bmV3IERhdGUoKTsKICBmb3IodmFyIGk9MDtpPHN0YXJ0O2krKyl7dmFyIGVsPWRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2RpdicpO2VsLmNsYXNzTmFtZT0nY2QgcGFkJztnLmFwcGVuZENoaWxkKGVsKTt9CiAgZm9yKHZhciBkYXk9MTtkYXk8PWRheXM7ZGF5KyspewogICAgdmFyIGVsPWRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2RpdicpO2VsLmNsYXNzTmFtZT0nY2QnOwogICAgdmFyIGRhdGU9bmV3IERhdGUoY1ksY00sZGF5KTsKICAgIHZhciBpc1Bhc3Q9ZGF0ZTxuZXcgRGF0ZSh0b2RheS5nZXRGdWxsWWVhcigpLHRvZGF5LmdldE1vbnRoKCksdG9kYXkuZ2V0RGF0ZSgpKTsKICAgIHZhciBpbm5lcj1kb2N1bWVudC5jcmVhdGVFbGVtZW50KCdkaXYnKTtpbm5lci5jbGFzc05hbWU9J2NkLWlubmVyJztpbm5lci50ZXh0Q29udGVudD1kYXk7ZWwuYXBwZW5kQ2hpbGQoaW5uZXIpOwogICAgaWYoaXNQYXN0KXtlbC5jbGFzc0xpc3QuYWRkKCdkaXMnKTt9CiAgICBlbHNlewogICAgICBpZihkYXRlLnRvRGF0ZVN0cmluZygpPT09dG9kYXkudG9EYXRlU3RyaW5nKCkpZWwuY2xhc3NMaXN0LmFkZCgndG9kJyk7CiAgICAgIChmdW5jdGlvbihkLCBlbFJlZil7CiAgICAgICAgZWxSZWYub25jbGljaz1mdW5jdGlvbigpewogICAgICAgICAgZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLmNkJykuZm9yRWFjaChmdW5jdGlvbihjKXtjLmNsYXNzTGlzdC5yZW1vdmUoJ3NlbCcpO30pOwogICAgICAgICAgZWxSZWYuY2xhc3NMaXN0LmFkZCgnc2VsJyk7CiAgICAgICAgICBib29raW5nLmRhdGU9U3RyaW5nKGQpLnBhZFN0YXJ0KDIsJzAnKSsnLicrU3RyaW5nKGNNKzEpLnBhZFN0YXJ0KDIsJzAnKSsnLicrY1k7CiAgICAgICAgICBzaG93VGltZXMoKTsKICAgICAgICB9OwogICAgICB9KShkYXksIGVsKTsKICAgIH0KICAgIGcuYXBwZW5kQ2hpbGQoZWwpOwogIH0KICAvLyBmaWxsIHRyYWlsaW5nIGNlbGxzIHRvIGNvbXBsZXRlIGxhc3QgZ3JpZCByb3cKICB2YXIgdG90YWwgPSBzdGFydCArIGRheXM7CiAgdmFyIHRyYWlsID0gKDcgLSAodG90YWwgJSA3KSkgJSA3OwogIGZvcih2YXIgdD0wO3Q8dHJhaWw7dCsrKXt2YXIgZXA9ZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnZGl2Jyk7ZXAuY2xhc3NOYW1lPSdjZCBwYWQnO2cuYXBwZW5kQ2hpbGQoZXApO30KfQoKZnVuY3Rpb24gc2hvd1RpbWVzKCl7CiAgdmFyIHRnPWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCd0aW1lRycpOwogIHRnLmlubmVySFRNTD0nPGRpdiBjbGFzcz0ibG9hZGluZy1zbG90cyI+4o+zINCX0LDQs9GA0YPQttCw0LXQvCDRgNCw0YHQv9C40YHQsNC90LjQtS4uLjwvZGl2Pic7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ3RpbWVTZWMnKS5zdHlsZS5kaXNwbGF5PSdibG9jayc7CgogIHZhciB1cmwgPSB3aW5kb3cubG9jYXRpb24ub3JpZ2luICsgIi9hcGkvc2xvdHMiICsgJz9hY3Rpb249c2xvdHMmZGF0ZT0nICsgZW5jb2RlVVJJQ29tcG9uZW50KGJvb2tpbmcuZGF0ZSkgKyAnJm1hc3Rlcj0nICsgZW5jb2RlVVJJQ29tcG9uZW50KGJvb2tpbmcubWFzdGVyKTsKCiAgZmV0Y2godXJsKQogICAgLnRoZW4oZnVuY3Rpb24ocil7cmV0dXJuIHIuanNvbigpO30pCiAgICAudGhlbihmdW5jdGlvbihkYXRhKXsKICAgICAgdmFyIHNsb3RzID0gKGRhdGEuc2xvdHMgJiYgZGF0YS5zbG90cy5sZW5ndGggPiAwKSA/IGRhdGEuc2xvdHMgOiBbXTsKICAgICAgcmVuZGVyVGltZVNsb3RzKHNsb3RzKTsKICAgIH0pCiAgICAuY2F0Y2goZnVuY3Rpb24oKXsKICAgICAgcmVuZGVyVGltZVNsb3RzKFtdKTsKICAgIH0pOwp9CgpmdW5jdGlvbiByZW5kZXJUaW1lU2xvdHMoc2xvdHMpewogIHZhciB0Zz1kb2N1bWVudC5nZXRFbGVtZW50QnlJZCgndGltZUcnKTt0Zy5pbm5lckhUTUw9Jyc7CiAgaWYoc2xvdHMubGVuZ3RoPT09MCl7CiAgICB0Zy5pbm5lckhUTUw9JzxkaXYgY2xhc3M9Im5vLXNsb3RzLXBhbmVsIj48ZGl2IGNsYXNzPSJuby1zbG90cy1pY29uIj7wn5C+PC9kaXY+PGRpdiBjbGFzcz0ibm8tc2xvdHMtdGl0bGUiPtCd0LXRgiDQtNC+0YHRgtGD0L/QvdGL0YUg0YHQu9C+0YLQvtCyPGJyPtC90LAg0Y3RgtGDINC00LDRgtGDPC9kaXY+PGRpdiBjbGFzcz0ibm8tc2xvdHMtZGl2aWRlciI+PC9kaXY+PGRpdiBjbGFzcz0ibm8tc2xvdHMtY3RhIiBvbmNsaWNrPSJzaG93U2NyZWVuKFwnaG9tZVNjcmVlblwnKSI+PGRpdiBjbGFzcz0ibm8tc2xvdHMtY3RhLXRpdGxlIj7QndC1INC90LDRiNC70Lgg0L/QvtC00YXQvtC00Y/RidC10LUg0LLRgNC10LzRjz88L2Rpdj48ZGl2IGNsYXNzPSJuby1zbG90cy1jdGEtc3ViIj7QodCy0Y/QttC40YLQtdGB0Ywg0YEg0L3QsNC80Lgg0LvRjtCx0YvQvCDRg9C00L7QsdC90YvQvCDRgdC/0L7RgdC+0LHQvtC8IOKAlDxicj7QvNGLINC/0L7QtNCx0LXRgNGR0Lwg0YPQtNC+0LHQvdC+0LUg0LLRgNC10LzRjzwvZGl2PjxkaXYgY2xhc3M9Im5vLXNsb3RzLWN0YS1hcnJvdyI+0KHQstGP0LfQsNGC0YzRgdGPIOKGkjwvZGl2PjwvZGl2PjwvZGl2Pic7CiAgICByZXR1cm47CiAgfQogIHNsb3RzLmZvckVhY2goZnVuY3Rpb24odCl7CiAgICB2YXIgYnRuPWRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2J1dHRvbicpO2J0bi5jbGFzc05hbWU9J3RidG4nO2J0bi50ZXh0Q29udGVudD10OwogICAgYnRuLm9uY2xpY2s9ZnVuY3Rpb24oKXsKICAgICAgZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLnRidG4nKS5mb3JFYWNoKGZ1bmN0aW9uKGIpe2IuY2xhc3NMaXN0LnJlbW92ZSgnYWN0aXZlJyk7fSk7CiAgICAgIGJ0bi5jbGFzc0xpc3QuYWRkKCdhY3RpdmUnKTtib29raW5nLnRpbWU9dDsKICAgICAgc2V0VGltZW91dChmdW5jdGlvbigpe2dvU3RlcCg1KTtidWlsZFN1bSgpO30sMzAwKTsKICAgIH07CiAgICB0Zy5hcHBlbmRDaGlsZChidG4pOwogIH0pOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCd0aW1lU2VjJykuc2Nyb2xsSW50b1ZpZXcoe2JlaGF2aW9yOidzbW9vdGgnLGJsb2NrOiduZWFyZXN0J30pOwp9CgpmdW5jdGlvbiBidWlsZFN1bSgpewogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdW1CbG9jaycpLmlubmVySFRNTD0KICAgICc8ZGl2IGNsYXNzPSJzciI+PHNwYW4gY2xhc3M9InNsIj4nK1RbTEFOR10uc3VtX2JyZWVkKyc8L3NwYW4+PHNwYW4gY2xhc3M9InN2Ij4nKyhib29raW5nLmJyZWVkRGlzcGxheXx8Ym9va2luZy5icmVlZCkrJzwvc3Bhbj48L2Rpdj4nKwogICAgJzxkaXYgY2xhc3M9InNyIj48c3BhbiBjbGFzcz0ic2wiPicrVFtMQU5HXS5zdW1fc2VydmljZSsnPC9zcGFuPjxzcGFuIGNsYXNzPSJzdiI+JysoKExBTkchPT0ncnUnJiZTVkNfVFJBTlNMQVRJT05TW2Jvb2tpbmcuc2VydmljZV0pP1NWQ19UUkFOU0xBVElPTlNbYm9va2luZy5zZXJ2aWNlXVtMQU5HXTpib29raW5nLnNlcnZpY2UpKyc8L3NwYW4+PC9kaXY+JysKICAgICc8ZGl2IGNsYXNzPSJzciI+PHNwYW4gY2xhc3M9InNsIj4nK1RbTEFOR10uc3VtX21hc3RlcisnPC9zcGFuPjxzcGFuIGNsYXNzPSJzdiI+Jytib29raW5nLm1hc3RlcisnPC9zcGFuPjwvZGl2PicrCiAgICAnPGRpdiBjbGFzcz0ic3IiPjxzcGFuIGNsYXNzPSJzbCI+JytUW0xBTkddLnN1bV9ncm9vbSsnPC9zcGFuPjxzcGFuIGNsYXNzPSJzdiI+Jytib29raW5nLmdyb29tSGlzdG9yeSsnPC9zcGFuPjwvZGl2PicrCiAgICAnPGRpdiBjbGFzcz0ic3IiPjxzcGFuIGNsYXNzPSJzbCI+JytUW0xBTkddLnN1bV9kYXRlKyc8L3NwYW4+PHNwYW4gY2xhc3M9InN2Ij4nK2Jvb2tpbmcuZGF0ZSsnPC9zcGFuPjwvZGl2PicrCiAgICAnPGRpdiBjbGFzcz0ic3IiPjxzcGFuIGNsYXNzPSJzbCI+JytUW0xBTkddLnN1bV90aW1lKyc8L3NwYW4+PHNwYW4gY2xhc3M9InN2Ij4nK2Jvb2tpbmcudGltZSsnPC9zcGFuPjwvZGl2PicrCiAgICAnPGRpdiBjbGFzcz0ic3IiPjxzcGFuIGNsYXNzPSJzbCI+JytUW0xBTkddLnN1bV9wcmljZSsnPC9zcGFuPjxzcGFuIGNsYXNzPSJzcCI+Jytib29raW5nLnByaWNlKycg4oKsPC9zcGFuPjwvZGl2Pic7Cn0KCmRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjb25maXJtQnRuJykub25jbGljayA9IGZ1bmN0aW9uKCl7CiAgdmFyIG5hbWU9ZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2NOYW1lJykudmFsdWU7CiAgdmFyIHBob25lPWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjUGhvbmUnKS52YWx1ZTsKICBpZighbmFtZXx8IXBob25lKXthbGVydChUW0xBTkddLmFsZXJ0X2ZpbGwpO3JldHVybjt9CiAgaWYoIS9eXCtcZHsxMCx9JC8udGVzdChwaG9uZS50cmltKCkpKXthbGVydChUW0xBTkddLmFsZXJ0X3Bob25lKTtyZXR1cm47fQogIGJvb2tpbmcubmFtZT1uYW1lOyBib29raW5nLnBob25lPXBob25lOyBib29raW5nLmVtYWlsPWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjRW1haWwnKS52YWx1ZTsgYm9va2luZy5wZXQ9ZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2NQZXQnKS52YWx1ZTsgYm9va2luZy5sYW5nPUxBTkc7CiAgYm9va2luZy5kdXJhdGlvbiA9IGJvb2tpbmcuYnJlZWQgPT09ICfQqdC10L3QutC4JyA/IDYwIDogKGJvb2tpbmcuYnJlZWQgJiYgYm9va2luZy5icmVlZC5pbmRleE9mKCfQmtC+0YjQutCwJykgPT09IDAgPyAxMjAgOiAxODApOwogIHZhciBidG49ZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2NvbmZpcm1CdG4nKTsKICBidG4udGV4dENvbnRlbnQ9VFtMQU5HXS5zZW5kaW5nOyBidG4uZGlzYWJsZWQ9dHJ1ZTsKICBmZXRjaChSQUlMV0FZLCB7CiAgICBtZXRob2Q6J1BPU1QnLAogICAgaGVhZGVyczp7J0NvbnRlbnQtVHlwZSc6J2FwcGxpY2F0aW9uL2pzb24nfSwKICAgIGJvZHk6SlNPTi5zdHJpbmdpZnkoYm9va2luZykKICB9KS50aGVuKGZ1bmN0aW9uKCl7c2hvd1N1Y2Nlc3MoKTt9KS5jYXRjaChmdW5jdGlvbigpe3Nob3dTdWNjZXNzKCk7fSk7Cn07CgpmdW5jdGlvbiBzaG93U3VjY2VzcygpewogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdiazUnKS5jbGFzc05hbWU9J3N0ZXAnOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdWNCbG9jaycpLmNsYXNzTGlzdC5hZGQoJ3Nob3cnKTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgncHJvZ3Jlc3MnKS5zdHlsZS5kaXNwbGF5PSdub25lJzsKfQoKZnVuY3Rpb24gcmVzZXRBbGwoKXsKICBib29raW5nPXticmVlZDonJyxicmVlZERpc3BsYXk6Jycsc2VydmljZTonJyxwcmljZTowLG1hc3RlcjonJyxncm9vbUhpc3Rvcnk6JycsZGF0ZTonJyx0aW1lOicnLGxhbmc6J3J1J307CiAgc2VsQnJlZWQ9bnVsbDsgaW5wLnZhbHVlPScnOyBjbHIuY2xhc3NMaXN0LnJlbW92ZSgnc2hvdycpOwogIGJhZGdlLmNsYXNzTGlzdC5yZW1vdmUoJ3Nob3cnKTsgYmFkZ2UuaW5uZXJIVE1MPScnOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdmNTZWMnKS5zdHlsZS5kaXNwbGF5PSdub25lJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgndGltZVNlYycpLnN0eWxlLmRpc3BsYXk9J25vbmUnOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdWNCbG9jaycpLmNsYXNzTGlzdC5yZW1vdmUoJ3Nob3cnKTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgncHJvZ3Jlc3MnKS5zdHlsZS5kaXNwbGF5PSdmbGV4JzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY05hbWUnKS52YWx1ZT0nJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY1Bob25lJykudmFsdWU9Jyc7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2NFbWFpbCcpLnZhbHVlPScnOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjUGV0JykudmFsdWU9Jyc7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2NvbmZpcm1CdG4nKS50ZXh0Q29udGVudD1UW0xBTkddLmNvbmZpcm1fYnRuOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjb25maXJtQnRuJykuZGlzYWJsZWQ9ZmFsc2U7CiAgZ29TdGVwKDEpOwp9Cgp2YXIgTEFORyA9IGxvY2FsU3RvcmFnZS5nZXRJdGVtKCdyamxhbmcnKSB8fCAncnUnOwp2YXIgVCA9IHsKICBydTp7CiAgICBsb2dvX3RhZzon0J/RgNC10LzQuNCw0LvRjNC90YvQuSDQs9GA0YPQvNC40L3Qsy08YnI+0YHQsNC70L7QvSDQsiDQotCw0LvQu9C40L3QtScsCiAgICBjaG9vc2VfaG93OidDaG9vc2UgaG93IHRvIGNvbm5lY3QnLAogICAgYm9va19vbmxpbmU6J9Ce0L3Qu9Cw0LnQvSDQsdGA0L7QvdC40YDQvtCy0LDQvdC40LUnLAogICAgYm9va19mbG93OifQn9C+0YDQvtC00LAg4oaSINCj0YHQu9GD0LPQsCDihpIg0JzQsNGB0YLQtdGAIOKGkiDQktGA0LXQvNGPJywKICAgIG9yX2NvbnRhY3Q6J9C40LvQuCDRgdCy0Y/QttC40YLQtdGB0Ywg0YEg0L3QsNC80LgnLAogICAgY2FsbF91czon0J/QvtC30LLQvtC90LjRgtC1INC90LDQvCcsCiAgICBiYWNrOifihpAg0J3QsNC30LDQtCcsCiAgICBsb2dvX3N1YjonR3Jvb21pbmcgwrcg0KLQsNC70LvQuNC9JywKICAgIHBzX3NlcnZpY2U6J9Cj0YHQu9GD0LPQsCcscHNfbWFzdGVyOifQnNCw0YHRgtC10YAnLHBzX3BldDon0J/QuNGC0L7QvNC10YYnLHBzX2RhdGU6J9CU0LDRgtCwJyxwc19kZXRhaWxzOifQlNCw0L3QvdGL0LUnLAogICAgc3RlcDFfbGJsOicwMSDCtyDQn9C+0YDQvtC00LAnLAogICAgYnJlZWRfcGg6J9Cd0LDRh9C90LjRgtC1INCy0LLQvtC00LjRgtGMINC/0L7RgNC+0LTRgy4uLicsCiAgICBzdGVwMl9sYmw6JzAyIMK3INCj0YHQu9GD0LPQsCcsCiAgICBzdGVwMl9tYXN0ZXI6J9CS0YvQsdC10YDQuNGC0LUg0LzQsNGB0YLQtdGA0LAnLAogICAgc3RlcDNfbGJsOifQmtCw0Log0LTQsNCy0L3QviDQstGLINC/0L7RgdC10YnQsNC70Lgg0LPRgNGD0LzQuNC90LM/JywKICAgIGcxOifQn9C10YDQstGL0Lkg0YDQsNC3JyxnMjon0J7RgiAxINC00L4gMyDQvNC10YHRj9GG0LXQsicsZzM6J9Ce0YIgMyDQtNC+IDYg0LzQtdGB0Y/RhtC10LInLGc0OifQkdC+0LvQtdC1IDYg0LzQtdGB0Y/RhtC10LInLAogICAgc3RlcDRfbGJsOifQktGL0LHQtdGA0LjRgtC1INC00LDRgtGDJywKICAgIGNhbF9hdmFpbDon0JXRgdGC0Ywg0YHQstC+0LHQvtC00L3QvtC1INCy0YDQtdC80Y8nLGNhbF9ub25lOifQodCy0L7QsdC+0LTQvdC+0LPQviDQstGA0LXQvNC10L3QuCDQvdC10YInLAogICAgc3RlcDRfdGltZTon0JLRi9Cx0LXRgNC40YLQtSDQstGA0LXQvNGPJywKICAgIHN0ZXA1X2xibDon0JLQsNGI0Lgg0LTQsNC90L3Ri9C1JywKICAgIGxibF9uYW1lOifQmNC80Y8nLHBoX25hbWU6J9CS0LDRiNC1INC40LzRjycsCiAgICBsYmxfcGhvbmU6J9Ci0LXQu9C10YTQvtC9JyxsYmxfZW1haWw6J0VtYWlsJywKICAgIGxibF9wZXQ6J9Ca0LvQuNGH0LrQsCDQv9C40YLQvtC80YbQsCcscGhfb3B0aW9uYWw6J9Cd0LXQvtCx0Y/Qt9Cw0YLQtdC70YzQvdC+JywKICAgIGNvbmZpcm1fYnRuOifQn9C+0LTRgtCy0LXRgNC00LjRgtGMINC30LDQv9C40YHRjCcsCiAgICBzdWNjZXNzX3RpdGxlOifQl9Cw0L/QuNGB0Ywg0L/RgNC40L3Rj9GC0LAhJywKICAgIHN1Y2Nlc3Nfc3ViOifQnNGLINGB0LLRj9C20LXQvNGB0Y8g0YEg0LLQsNC80Lgg0LTQu9GPINC/0L7QtNGC0LLQtdGA0LbQtNC10L3QuNGPLjxicj7QodC/0LDRgdC40LHQviwg0YfRgtC+INCy0YvQsdGA0LDQu9C4IFImYW1wO0ogR3Jvb21pbmchJywKICAgIHRvX2hvbWU6J+KGkCDQndCwINCz0LvQsNCy0L3Rg9GOJywKICAgIGFsZXJ0X2ZpbGw6J9CS0LLQtdC00LjRgtC1INC40LzRjyDQuCDRgtC10LvQtdGE0L7QvScsYWxlcnRfcGhvbmU6J9CS0LLQtdC00LjRgtC1INC90L7QvNC10YAg0LIg0YTQvtGA0LzQsNGC0LUgKzM3MjEyMzQ1Njc4JywKICAgIHNlbmRpbmc6J9Ce0YLQv9GA0LDQstC70Y/QtdC8Li4uJywKICAgIHN1bV9icmVlZDon0J/QvtGA0L7QtNCwJyxzdW1fc2VydmljZTon0KPRgdC70YPQs9CwJyxzdW1fbWFzdGVyOifQnNCw0YHRgtC10YAnLHN1bV9ncm9vbTon0J/QvtGB0LvQtdC00L3QuNC5INCz0YDRg9C8JyxzdW1fZGF0ZTon0JTQsNGC0LAnLHN1bV90aW1lOifQktGA0LXQvNGPJyxzdW1fcHJpY2U6J9Ch0YLQvtC40LzQvtGB0YLRjCcsCiAgICBtb250aHM6WyfQr9C90LLQsNGA0YwnLCfQpNC10LLRgNCw0LvRjCcsJ9Cc0LDRgNGCJywn0JDQv9GA0LXQu9GMJywn0JzQsNC5Jywn0JjRjtC90YwnLCfQmNGO0LvRjCcsJ9CQ0LLQs9GD0YHRgicsJ9Ch0LXQvdGC0Y/QsdGA0YwnLCfQntC60YLRj9Cx0YDRjCcsJ9Cd0L7Rj9Cx0YDRjCcsJ9CU0LXQutCw0LHRgNGMJ10KICB9LAogIGVuOnsKICAgIGxvZ29fdGFnOidQcmVtaXVtIGdyb29taW5nPGJyPnNhbG9uIGluIFRhbGxpbm4nLAogICAgY2hvb3NlX2hvdzonQ2hvb3NlIGhvdyB0byBjb25uZWN0JywKICAgIGJvb2tfb25saW5lOidCb29rIE9ubGluZScsCiAgICBib29rX2Zsb3c6J0JyZWVkIOKGkiBTZXJ2aWNlIOKGkiBNYXN0ZXIg4oaSIFRpbWUnLAogICAgb3JfY29udGFjdDonb3IgY29udGFjdCB1cycsCiAgICBjYWxsX3VzOidDYWxsIFVzJywKICAgIGJhY2s6J+KGkCBCYWNrJywKICAgIGxvZ29fc3ViOidHcm9vbWluZyDCtyBUYWxsaW5uJywKICAgIHBzX3NlcnZpY2U6J1NlcnZpY2UnLHBzX21hc3RlcjonTWFzdGVyJyxwc19wZXQ6J1BldCcscHNfZGF0ZTonRGF0ZScscHNfZGV0YWlsczonRGV0YWlscycsCiAgICBzdGVwMV9sYmw6JzAxIMK3IERvZyBicmVlZCcsCiAgICBicmVlZF9waDonU3RhcnQgdHlwaW5nIGJyZWVkLi4uJywKICAgIHN0ZXAyX2xibDonMDIgwrcgU2VydmljZScsCiAgICBzdGVwMl9tYXN0ZXI6J0Nob29zZSBtYXN0ZXInLAogICAgc3RlcDNfbGJsOidIb3cgbG9uZyBhZ28gd2FzIHlvdXIgbGFzdCBncm9vbWluZz8nLAogICAgZzE6J0ZpcnN0IHRpbWUnLGcyOicx4oCTMyBtb250aHMgYWdvJyxnMzonM+KAkzYgbW9udGhzIGFnbycsZzQ6J092ZXIgNiBtb250aHMnLAogICAgc3RlcDRfbGJsOidDaG9vc2UgZGF0ZScsCiAgICBjYWxfYXZhaWw6J0F2YWlsYWJsZScsY2FsX25vbmU6J05vdCBhdmFpbGFibGUnLAogICAgc3RlcDRfdGltZTonQ2hvb3NlIHRpbWUnLAogICAgc3RlcDVfbGJsOidZb3VyIGRldGFpbHMnLAogICAgbGJsX25hbWU6J05hbWUnLHBoX25hbWU6J1lvdXIgbmFtZScsCiAgICBsYmxfcGhvbmU6J1Bob25lJyxsYmxfZW1haWw6J0VtYWlsJywKICAgIGxibF9wZXQ6IlBldCdzIG5hbWUiLHBoX29wdGlvbmFsOidPcHRpb25hbCcsCiAgICBjb25maXJtX2J0bjonQ29uZmlybSBib29raW5nJywKICAgIHN1Y2Nlc3NfdGl0bGU6J0Jvb2tpbmcgY29uZmlybWVkIScsCiAgICBzdWNjZXNzX3N1YjonV2Ugd2lsbCBjb250YWN0IHlvdSB0byBjb25maXJtLjxicj5UaGFuayB5b3UgZm9yIGNob29zaW5nIFImYW1wO0ogR3Jvb21pbmchJywKICAgIHRvX2hvbWU6J+KGkCBIb21lJywKICAgIGFsZXJ0X2ZpbGw6J1BsZWFzZSBlbnRlciBuYW1lIGFuZCBwaG9uZScsYWxlcnRfcGhvbmU6J0VudGVyIHBob25lIG51bWJlciBpbiBmb3JtYXQgKzM3MjEyMzQ1Njc4JywKICAgIHNlbmRpbmc6J1NlbmRpbmcuLi4nLAogICAgc3VtX2JyZWVkOidCcmVlZCcsc3VtX3NlcnZpY2U6J1NlcnZpY2UnLHN1bV9tYXN0ZXI6J01hc3Rlcicsc3VtX2dyb29tOidMYXN0IGdyb29taW5nJyxzdW1fZGF0ZTonRGF0ZScsc3VtX3RpbWU6J1RpbWUnLHN1bV9wcmljZTonUHJpY2UnLAogICAgbW9udGhzOlsnSmFudWFyeScsJ0ZlYnJ1YXJ5JywnTWFyY2gnLCdBcHJpbCcsJ01heScsJ0p1bmUnLCdKdWx5JywnQXVndXN0JywnU2VwdGVtYmVyJywnT2N0b2JlcicsJ05vdmVtYmVyJywnRGVjZW1iZXInXQogIH0sCiAgZXQ6ewogICAgbG9nb190YWc6J0VzbWFrbGFzc2lsaW5lIGhvb2xkdXN0ZWVudXM8YnI+VGFsbGlubmFzJywKICAgIGNob29zZV9ob3c6J1ZhbGkgw7xoZW5kdXN2aWlzJywKICAgIGJvb2tfb25saW5lOidCcm9uZWVyaSB2ZWViaXMnLAogICAgYm9va19mbG93OidUw7V1ZyDihpIgVGVlbnVzIOKGkiBNZWlzdGVyIOKGkiBBZWcnLAogICAgb3JfY29udGFjdDondsO1aSB2w7V0YSDDvGhlbmR1c3QnLAogICAgY2FsbF91czonSGVsaXN0YSBtZWlsZScsCiAgICBiYWNrOifihpAgVGFnYXNpJywKICAgIGxvZ29fc3ViOidHcm9vbWluZyDCtyBUYWxsaW5uJywKICAgIHBzX3NlcnZpY2U6J1RlZW51cycscHNfbWFzdGVyOidNZWlzdGVyJyxwc19wZXQ6J0xlbW1pa2xvb20nLHBzX2RhdGU6J0t1dXDDpGV2Jyxwc19kZXRhaWxzOidBbmRtZWQnLAogICAgc3RlcDFfbGJsOicwMSDCtyBLb2VyYSB0w7V1ZycsCiAgICBicmVlZF9waDonQWx1c3RhZ2UgdMO1dSBzaXNlc3RhbWlzdC4uLicsCiAgICBzdGVwMl9sYmw6JzAyIMK3IFRlZW51cycsCiAgICBzdGVwMl9tYXN0ZXI6J1ZhbGkgbWVpc3RlcicsCiAgICBzdGVwM19sYmw6J01pbGxhbCBrw6Rpc2l0ZSB2aWltYXRpIGdyb29taW5ndXM/JywKICAgIGcxOidFc2ltZXN0IGtvcmRhJyxnMjonMeKAkzMga3V1ZCB0YWdhc2knLGczOicz4oCTNiBrdXVkIHRhZ2FzaScsZzQ6J8OcbGUgNiBrdXUnLAogICAgc3RlcDRfbGJsOidWYWxpIGt1dXDDpGV2JywKICAgIGNhbF9hdmFpbDonVmFidSBhZWd1IG9uJyxjYWxfbm9uZTonVmFidSBhZWd1IHBvbGUnLAogICAgc3RlcDRfdGltZTonVmFsaSBrZWxsYWFlZycsCiAgICBzdGVwNV9sYmw6J1RlaWUgYW5kbWVkJywKICAgIGxibF9uYW1lOidOaW1pJyxwaF9uYW1lOidUZWllIG5pbWknLAogICAgbGJsX3Bob25lOidUZWxlZm9uJyxsYmxfZW1haWw6J0VtYWlsJywKICAgIGxibF9wZXQ6J0xlbW1pa2xvb21hIG5pbWknLHBoX29wdGlvbmFsOidWYWxpa3VsaW5lJywKICAgIGNvbmZpcm1fYnRuOidLaW5uaXRhIGJyb25lZXJpbmcnLAogICAgc3VjY2Vzc190aXRsZTonQnJvbmVlcmluZyBraW5uaXRhdHVkIScsCiAgICBzdWNjZXNzX3N1YjonVsO1dGFtZSB0ZWllZ2Egw7xoZW5kdXN0IGtpbm5pdGFtaXNla3MuPGJyPlTDpG5hbWUsIGV0IHZhbGlzaXRlIFImYW1wO0ogR3Jvb21pbmchJywKICAgIHRvX2hvbWU6J+KGkCBBdmFsZWhlbGUnLAogICAgYWxlcnRfZmlsbDonUGFsdW4gc2lzZXN0YWdlIG5pbWkgamEgdGVsZWZvbicsYWxlcnRfcGhvbmU6J1Npc2VzdGFnZSB0ZWxlZm9uaW51bWJlciB2b3JtaW5ndXMgKzM3MjEyMzQ1Njc4JywKICAgIHNlbmRpbmc6J1NhYWRhbi4uLicsCiAgICBzdW1fYnJlZWQ6J1TDtXVnJyxzdW1fc2VydmljZTonVGVlbnVzJyxzdW1fbWFzdGVyOidNZWlzdGVyJyxzdW1fZ3Jvb206J1ZpaW1hbmUgZ3Jvb21pbmcnLHN1bV9kYXRlOidLdXVww6Rldicsc3VtX3RpbWU6J0tlbGxhYWVnJyxzdW1fcHJpY2U6J0hpbmQnLAogICAgbW9udGhzOlsnSmFhbnVhcicsJ1ZlZWJydWFyJywnTcOkcnRzJywnQXByaWxsJywnTWFpJywnSnV1bmknLCdKdXVsaScsJ0F1Z3VzdCcsJ1NlcHRlbWJlcicsJ09rdG9vYmVyJywnTm92ZW1iZXInLCdEZXRzZW1iZXInXQogIH0KfTsKCmZ1bmN0aW9uIHNldExhbmcobCl7CiAgTEFORz1sOwogIGxvY2FsU3RvcmFnZS5zZXRJdGVtKCdyamxhbmcnLGwpOwogIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5sYW5nLWJ0bicpLmZvckVhY2goZnVuY3Rpb24oYil7CiAgICBiLmNsYXNzTGlzdC50b2dnbGUoJ2FjdGl2ZScsIGIudGV4dENvbnRlbnQudG9Mb3dlckNhc2UoKT09PWwpOwogIH0pOwogIHZhciB0cj1UW2xdOwogIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJ1tkYXRhLWkxOG5dJykuZm9yRWFjaChmdW5jdGlvbihlbCl7CiAgICB2YXIgaz1lbC5nZXRBdHRyaWJ1dGUoJ2RhdGEtaTE4bicpOwogICAgaWYodHJba10hPT11bmRlZmluZWQpIGVsLmlubmVySFRNTD10cltrXTsKICB9KTsKICBkb2N1bWVudC5xdWVyeVNlbGVjdG9yQWxsKCdbZGF0YS1pMThuLXBoXScpLmZvckVhY2goZnVuY3Rpb24oZWwpewogICAgdmFyIGs9ZWwuZ2V0QXR0cmlidXRlKCdkYXRhLWkxOG4tcGgnKTsKICAgIGlmKHRyW2tdIT09dW5kZWZpbmVkKSBlbC5wbGFjZWhvbGRlcj10cltrXTsKICB9KTsKICBNT05USFM9dHIubW9udGhzOwogIGJ1aWxkQ2FsKCk7CiAgLy8gUmUtcmVuZGVyIGJhZGdlIGFuZCBzZXJ2aWNlcyBpZiBicmVlZCBhbHJlYWR5IHNlbGVjdGVkCiAgaWYoc2VsQnJlZWQpewogICAgdmFyIGJmPWw9PT0nZW4nPydicmVlZF9lbic6bD09PSdldCc/J2JyZWVkX2V0JzonYnJlZWQnOwogICAgdmFyIGRiPXNlbEJyZWVkW2JmXXx8c2VsQnJlZWQuYnJlZWQ7CiAgICBib29raW5nLmJyZWVkRGlzcGxheT1kYjsKICAgIHZhciBibkVsPWRvY3VtZW50LnF1ZXJ5U2VsZWN0b3IoJyNzQmFkZ2UgLmJuYW1lJyk7CiAgICBpZihibkVsKSBibkVsLnRleHRDb250ZW50PWRiOwogICAgdmFyIGJjRWw9ZG9jdW1lbnQucXVlcnlTZWxlY3RvcignI3NCYWRnZSAuYmNoZycpOwogICAgaWYoYmNFbCkgYmNFbC50ZXh0Q29udGVudD1sPT09J2VuJz8nQ2hhbmdlJzpsPT09J2V0Jz8nTXV1ZGEnOifQmNC30LzQtdC90LjRgtGMJzsKICAgIGlmKGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzdmNTZWMnKS5zdHlsZS5kaXNwbGF5IT09J25vbmUnKSByZW5kZXJTdmNzKHNlbEJyZWVkKTsKICAgIHZhciBzbj1kb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnc3ZjTm90ZScpOwogICAgaWYoc24pewogICAgICB2YXIgbnQ9bD09PSdlbic/J1BsZWFzZSBub3RlJzpsPT09J2V0Jz8nUGFuZ2UgdMOkaGVsZSc6J9CS0LDQttC90L4g0LfQvdCw0YLRjCc7CiAgICAgIHZhciBuYj1sPT09J2VuJz8nRmluYWwgcHJpY2UgZGVwZW5kcyBvbiBjb2F0IGNvbmRpdGlvbiBhbmQgcGV0IGJlaGF2aW91ci48YnI+RGVtYXR0aW5nIGZyb20gNSDigqwuPGJyPkFnZ3Jlc3NpdmUgYmVoYXZpb3VyIHN1cmNoYXJnZSBtYXkgYXBwbHk6ICs1MCUuJzpsPT09J2V0Jz8nTMO1cGxpayBoaW5kIHPDtWx0dWIga2FydmFzdGlrdSBzZWlzdW5kaXN0IGphIGxlbW1pa2xvb21hIGvDpGl0dW1pc2VzdC48YnI+S29sdHN1bml0ZSBsYWh0aWhhcnV0YW1pbmUgYWxhdGVzIDUg4oKsLjxicj5BZ3Jlc3NpaXZzZSBrw6RpdHVtaXNlIGtvcnJhbCB2w7VpYiBsaXNhbmR1ZGEgNTAlIGp1dXJkZWhpbmRsdXMuJzon0J7QutC+0L3Rh9Cw0YLQtdC70YzQvdCw0Y8g0YHRgtC+0LjQvNC+0YHRgtGMINC30LDQstC40YHQuNGCINC+0YIg0YHQvtGB0YLQvtGP0L3QuNGPINGI0LXRgNGB0YLQuCDQuCDQv9C+0LLQtdC00LXQvdC40Y8g0L/QuNGC0L7QvNGG0LAuPGJyPtCg0LDQt9Cx0L7RgCDQutC+0LvRgtGD0L3QvtCyIOKAlCDQvtGCIDUg4oKsLjxicj7Qn9GA0Lgg0LDQs9GA0LXRgdGB0LjQstC90L7QvCDQv9C+0LLQtdC00LXQvdC40Lgg0LzQvtC20LXRgiDQv9GA0LjQvNC10L3Rj9GC0YzRgdGPINC00L7Qv9C70LDRgtCwIDUwJS4nOwogICAgICBzbi5pbm5lckhUTUw9JzxkaXYgc3R5bGU9ImZvbnQtc2l6ZTowLjgzOHJlbTtsZXR0ZXItc3BhY2luZzouMTVlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Y29sb3I6I2ZmZmZmZjttYXJnaW4tYm90dG9tOjhweDtmb250LXdlaWdodDo2MDA7Zm9udC1mYW1pbHk6XCdNb250c2VycmF0XCcsc2Fucy1zZXJpZiI+JytudCsnPC9kaXY+PGRpdiBzdHlsZT0iZm9udC1zaXplOjEuMDI1cmVtO2NvbG9yOiNmZmZmZmY7bGluZS1oZWlnaHQ6MS44O2ZvbnQtZmFtaWx5OlwnTW9udHNlcnJhdFwnLHNhbnMtc2VyaWYiPicrbmIrJzwvZGl2Pic7CiAgICB9CiAgfQp9CgovLyBBcHBseSBzYXZlZCBsYW5ndWFnZSBvbiBsb2FkCihmdW5jdGlvbigpeyBzZXRMYW5nKExBTkcpOyB9KSgpOwoKLy8gQ2FsbGJhY2sgZm9ybQpkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2FsbGJhY2tCdG4nKS5vbmNsaWNrID0gZnVuY3Rpb24oKXsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2JrTW9kYWwnKS5zdHlsZS5kaXNwbGF5ID0gJ2ZsZXgnOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjYmtOYW1lJykudmFsdWUgPSAnJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2JrUGhvbmUnKS52YWx1ZSA9ICcnOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjYmtTdWNjZXNzJykuc3R5bGUuZGlzcGxheSA9ICdub25lJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2JrU3VibWl0Jykuc3R5bGUuZGlzcGxheSA9ICdibG9jayc7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2Nia0Nsb3NlJykudGV4dENvbnRlbnQgPSAn0J7RgtC80LXQvdCwJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2JrU3VibWl0JykudGV4dENvbnRlbnQgPSAn0J7RgtC/0YDQsNCy0LjRgtGMJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2JrU3VibWl0JykuZGlzYWJsZWQgPSBmYWxzZTsKfTsKZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2Nia0Nsb3NlJykub25jbGljayA9IGZ1bmN0aW9uKCl7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2Nia01vZGFsJykuc3R5bGUuZGlzcGxheSA9ICdub25lJzsKfTsKZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2Nia1N1Ym1pdCcpLm9uY2xpY2sgPSBmdW5jdGlvbigpewogIHZhciBuYW1lID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2Nia05hbWUnKS52YWx1ZS50cmltKCk7CiAgdmFyIHBob25lID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2Nia1Bob25lJykudmFsdWUudHJpbSgpLnJlcGxhY2UoL1xEL2csJycpOwogIGlmKCFuYW1lIHx8ICFwaG9uZSl7YWxlcnQoJ9CS0LLQtdC00LjRgtC1INC40LzRjyDQuCDRgtC10LvQtdGE0L7QvScpO3JldHVybjt9CiAgdmFyIGJ0biA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjYmtTdWJtaXQnKTsKICBidG4udGV4dENvbnRlbnQgPSAn0J7RgtC/0YDQsNCy0LvRj9C10LwuLi4nOyBidG4uZGlzYWJsZWQgPSB0cnVlOwogIGZldGNoKCcvYXBpL2NhbGxiYWNrJyx7CiAgICBtZXRob2Q6J1BPU1QnLAogICAgaGVhZGVyczp7J0NvbnRlbnQtVHlwZSc6J2FwcGxpY2F0aW9uL2pzb24nfSwKICAgIGJvZHk6SlNPTi5zdHJpbmdpZnkoe25hbWU6bmFtZSwgcGhvbmU6JyszNzInK3Bob25lfSkKICB9KS50aGVuKGZ1bmN0aW9uKCl7CiAgICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2JrU3VjY2VzcycpLnN0eWxlLmRpc3BsYXkgPSAnYmxvY2snOwogICAgYnRuLnN0eWxlLmRpc3BsYXkgPSAnbm9uZSc7CiAgICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2JrQ2xvc2UnKS50ZXh0Q29udGVudCA9ICfihpAg0JfQsNC60YDRi9GC0YwnOwogICAgc2V0VGltZW91dChmdW5jdGlvbigpe2RvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdjYmtNb2RhbCcpLnN0eWxlLmRpc3BsYXk9J25vbmUnO30sMzAwMCk7CiAgfSkuY2F0Y2goZnVuY3Rpb24oKXsKICAgIGJ0bi50ZXh0Q29udGVudCA9ICfQntGC0L/RgNCw0LLQuNGC0YwnOyBidG4uZGlzYWJsZWQgPSBmYWxzZTsKICAgIGFsZXJ0KCfQntGI0LjQsdC60LAuINCf0L7Qv9GA0L7QsdGD0LnRgtC1INC10YnRkSDRgNCw0LcuJyk7CiAgfSk7Cn07Cgo8L3NjcmlwdD4KPC9ib2R5Pgo8L2h0bWw+Cg=="



@app.route("/api/available_days")
def api_available_days():
    month = request.args.get("month", "")
    year = request.args.get("year", "")
    master = request.args.get("master", "")
    gs_params = {"action": "available_days", "month": month, "year": year, "master": master}
    print(f"[api/available_days] GET {_gs_url(gs_params)}", flush=True)
    if not GOOGLE_SCRIPT:
        print("[api/available_days] GOOGLE_SCRIPT not configured", flush=True)
        resp = jsonify({"available": [], "error": "GOOGLE_SCRIPT not configured"})
        resp.headers["Access-Control-Allow-Origin"] = "*"
        return resp
    try:
        r = requests.get(GOOGLE_SCRIPT, params=gs_params, timeout=25)
        print(f"[api/available_days] → {r.status_code}: {r.text[:300]}", flush=True)
        resp = jsonify(r.json())
    except Exception as e:
        print(f"[api/available_days] error: {e}", flush=True)
        resp = jsonify({"available": [], "error": str(e)})
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp

@app.route("/api/slots")
def api_slots():
    date = request.args.get("date", "")
    master = request.args.get("master", "")
    gs_params = {"action": "slots", "date": date, "master": master}
    print(f"[api/slots] GET {_gs_url(gs_params)}", flush=True)
    if not GOOGLE_SCRIPT:
        print("[api/slots] GOOGLE_SCRIPT not configured", flush=True)
        resp = jsonify({"slots": [], "error": "GOOGLE_SCRIPT not configured"})
        resp.headers["Access-Control-Allow-Origin"] = "*"
        return resp
    try:
        r = requests.get(GOOGLE_SCRIPT, params=gs_params, timeout=25)
        print(f"[api/slots] → {r.status_code}: {r.text[:300]}", flush=True)
        resp = jsonify(r.json())
    except Exception as e:
        print(f"[api/slots] error: {e}", flush=True)
        resp = jsonify({"slots": [], "error": str(e)})
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp
@app.route("/app")
def booking_app():
    import base64
    html = base64.b64decode(BOOKING_HTML_B64).decode("utf-8")
    html = html.replace("https://dynamic-cooperation-production-dd95.up.railway.app/book", "/book")
    html = html.replace(
        '<div class="fg"><label class="fl">Кличка питомца</label>',
        '<div class="fg"><label class="fl">Email</label><input class="fi" id="cEmail" type="email" placeholder="email@example.com"></div><div class="fg"><label class="fl">Кличка питомца</label>'
    )
    html = html.replace(
        "booking.pet=document.getElementById('cPet').value;",
        "booking.pet=document.getElementById('cPet').value; booking.email=document.getElementById('cEmail').value;"
    )
    # ── Phone validation: must start with + and have at least 10 digits ──────
    html = html.replace(
        "if(!name||!phone){alert(T[LANG].alert_fill);return;}",
        "if(!name||!phone){alert(T[LANG].alert_fill);return;}"
        "if(!/^\\+\\d{10,}$/.test(phone.trim())){alert(T[LANG].alert_phone);return;}"
    )
    html = html.replace(
        "alert_fill:'Введите имя и телефон',",
        "alert_fill:'Введите имя и телефон',alert_phone:'Введите номер в формате +37212345678',"
    )
    html = html.replace(
        "alert_fill:'Please enter name and phone',",
        "alert_fill:'Please enter name and phone',alert_phone:'Enter phone number in format +37212345678',"
    )
    html = html.replace(
        "alert_fill:'Palun sisestage nimi ja telefon',",
        "alert_fill:'Palun sisestage nimi ja telefon',alert_phone:'Sisestage telefoninumber vormingus +37212345678',"
    )
    anna_filter_script = """
<script>
(function(){
  var ANNA_KEYS = ["француз","такса","джек","рассел","самоед","шелти","корги","лабрадор","чихуа"];
  function annaCanGroom(breed){
    if(!breed) return true;
    var b = String(breed).toLowerCase();
    if(b.indexOf("такса") !== -1){
      if(b.indexOf("жёстк") !== -1 || b.indexOf("жестк") !== -1) return false;
      if(b.indexOf("кроличь") !== -1) return false;
      return true;
    }
    if(b.indexOf("джек") !== -1 || b.indexOf("рассел") !== -1){
      if(b.indexOf("жёстк") !== -1 || b.indexOf("жестк") !== -1) return false;
      if(b.indexOf("брокен") !== -1) return false;
      return true;
    }
    for(var i=0; i<ANNA_KEYS.length; i++){
      if(b.indexOf(ANNA_KEYS[i]) !== -1) return true;
    }
    return false;
  }
  function updateAnnaVisibility(){
    var annaBtn = document.querySelector('.mbtn[data-master="Анна"]');
    if(!annaBtn) return;
    var b = (window.booking && window.booking.breed) ? window.booking.breed : "";
    if(annaCanGroom(b)){
      annaBtn.style.display = "";
    } else {
      annaBtn.style.display = "none";
      if(window.booking && window.booking.master === "Анна"){
        window.booking.master = "";
        annaBtn.classList.remove("active");
      }
    }
  }
  function install(){
    if(!window.booking){ setTimeout(install, 100); return; }
    updateAnnaVisibility();
    try{
      var _breed = window.booking.breed;
      Object.defineProperty(window.booking, "breed", {
        get: function(){ return _breed; },
        set: function(v){ _breed = v; updateAnnaVisibility(); },
        configurable: true
      });
    }catch(e){
      var last = window.booking.breed;
      setInterval(function(){
        if(window.booking.breed !== last){
          last = window.booking.breed;
          updateAnnaVisibility();
        }
      }, 300);
    }
  }
  if(document.readyState === "loading"){
    document.addEventListener("DOMContentLoaded", install);
  } else {
    install();
  }
})();
</script>
"""
    html = html.replace("</body>", anna_filter_script + "</body>")
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}


# ── WEBHOOKS ───────────────────────────────────────────────────────────────
@app.route("/webhook", methods=["GET"])
def verify():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    if mode == "subscribe" and token in (VERIFY_TOKEN, INSTAGRAM_VERIFY_TOKEN):
        return challenge, 200
    return "Forbidden", 403

@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.json
    try:
        obj = data.get("object", "")
        entry = data.get("entry", [{}])[0]
        if obj == "instagram" or "messaging" in entry:
            for msg_event in entry.get("messaging", []):
                sender_id = msg_event.get("sender", {}).get("id", "")
                recipient_id = msg_event.get("recipient", {}).get("id", "")
                if sender_id == recipient_id: continue
                msg = msg_event.get("message", {})
                if msg.get("is_echo"): continue
                text = msg.get("text", "")
                if sender_id and text:
                    handle_message(sender_id, text, "instagram")
            return "ok", 200
        value = entry.get("changes", [{}])[0].get("value", {})
        messages = value.get("messages", [])
        if not messages: return "ok", 200
        msg = messages[0]
        phone = msg["from"]
        text = msg.get("text", {}).get("body", "")
        if text:
            handle_message(phone, text, "whatsapp")
    except Exception as e:
        print("Error:", str(e))
    return "ok", 200

# ── ADMIN API ──────────────────────────────────────────────────────────────
@app.route("/admin/api/toggle", methods=["POST"])
def api_toggle():
    global jarvis_enabled, manual_mode, pause_until
    data = request.get_json()
    jarvis_enabled = bool(data.get("enabled", True))
    manual_mode = False; pause_until = None
    return jsonify({"ok": True, "message": "Jarvis " + ("включён ✅" if jarvis_enabled else "выключен ❌")})

@app.route("/admin/api/toggle-instagram", methods=["POST"])
def api_toggle_instagram():
    global instagram_enabled
    instagram_enabled = bool(request.get_json().get("enabled", True))
    return jsonify({"ok": True, "message": "Instagram " + ("включён ✅" if instagram_enabled else "выключен ❌")})

@app.route("/admin/api/messages")
def api_messages():
    result = []
    for phone, info in clients.items():
        last = info.get("last_seen")
        result.append({"phone": phone, "ts": last.isoformat() if isinstance(last, datetime) else "", "last_text": info.get("last_text", ""), "channel": info.get("channel", "whatsapp")})
    return jsonify(result)

@app.route("/admin/api/message-log")
def api_message_log():
    return jsonify({"log": list(reversed(_msg_log[-20:])), "counts": _channel_counts})

@app.route("/admin/api/send", methods=["POST"])
def api_send():
    data = request.get_json() or {}
    phone = data.get("phone", "").strip()
    text = data.get("text", "").strip()
    if not phone or not text:
        return jsonify({"ok": False, "error": "phone and text required"}), 400
    try:
        send_whatsapp(phone, text)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/cancel-booking", methods=["POST"])
def api_cancel_booking():
    data = request.get_json() or {}
    phone = data.get("phone", "").strip()
    if not phone:
        return jsonify({"success": False, "error": "phone required"}), 400
    return jsonify(_gs_cancel(phone))

@app.route("/api/reschedule-booking", methods=["POST"])
def api_reschedule_booking():
    data = request.get_json() or {}
    phone    = data.get("phone", "").strip()
    new_date = data.get("newDate", "").strip()
    new_time = data.get("newTime", "").strip()
    if not all([phone, new_date, new_time]):
        return jsonify({"success": False, "error": "phone, newDate, newTime required"}), 400
    return jsonify(_gs_reschedule(phone, new_date, new_time))

# ── ElevenLabs voice agent endpoints ─────────────────────────────────────────

@app.route("/api/elevenlabs/book", methods=["POST"])
def elevenlabs_book():
    """Accept booking from ElevenLabs voice agent and forward to Google Script.
    Required JSON fields: breed, service, date (DD.MM.YYYY), time (HH:MM), name, phone.
    Optional: pet, master, price, lang (ru/en/et), breedDisplay.
    """
    data = request.get_json() or {}
    missing = [f for f in ("breed", "service", "date", "time", "name", "phone")
               if not data.get(f)]
    if missing:
        return jsonify({"success": False,
                        "error": f"Missing required fields: {', '.join(missing)}"}), 400

    payload = {
        "breed":        data.get("breed", ""),
        "breedDisplay": data.get("breedDisplay") or data.get("breed", ""),
        "service":      data.get("service", ""),
        "price":        data.get("price", 0),
        "date":         data.get("date", ""),
        "time":         data.get("time", ""),
        "name":         data.get("name", ""),
        "pet":          data.get("pet", ""),
        "master":       data.get("master", ""),
        "phone":        data.get("phone", ""),
        "lang":         data.get("lang", "ru"),
        "source":       "elevenlabs",
    }
    print(f"[elevenlabs/book] {payload}", flush=True)

    if not GOOGLE_SCRIPT:
        return jsonify({"success": False, "error": "GOOGLE_SCRIPT not configured"}), 500
    try:
        r = requests.get(GOOGLE_SCRIPT, params=payload, timeout=15)
        print(f"[elevenlabs/book] GS → {r.status_code}: {r.text[:300]}", flush=True)
        return jsonify({"success": True, "booking": payload})
    except Exception as e:
        print(f"[elevenlabs/book] error: {e}", flush=True)
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/elevenlabs/slots")
def elevenlabs_slots():
    """Return available time slots for a given date.
    Query params: date=DD.MM.YYYY (required), master= (optional).
    Returns: {success, date, slots: ["10:00", "11:00", ...]}
    """
    date   = request.args.get("date", "").strip()
    master = request.args.get("master", "").strip()

    if not date:
        return jsonify({"success": False, "slots": [],
                        "error": "date parameter required (DD.MM.YYYY)"}), 400
    if not GOOGLE_SCRIPT:
        return jsonify({"success": False, "slots": [],
                        "error": "GOOGLE_SCRIPT not configured"}), 500
    try:
        r = requests.get(GOOGLE_SCRIPT,
                         params={"action": "slots", "date": date, "master": master},
                         timeout=25)
        slots = [str(s) for s in r.json().get("slots", []) if str(s).endswith(":00")]
        print(f"[elevenlabs/slots] {date} master={master!r} → {slots}", flush=True)
        return jsonify({"success": True, "date": date, "slots": slots})
    except Exception as e:
        print(f"[elevenlabs/slots] error: {e}", flush=True)
        return jsonify({"success": False, "slots": [], "error": str(e)}), 500


@app.route("/api/elevenlabs/available_days")
def elevenlabs_available_days():
    """Return available days for the current (or specified) month.
    Query params: month= (1-12), year= (YYYY), master= (optional).
    Returns: {success, month, year, available_days: ["01.06.2026", ...]}
    """
    import datetime as _dt
    _now   = _dt.date.today()
    month  = request.args.get("month",  str(_now.month)).strip()
    year   = request.args.get("year",   str(_now.year)).strip()
    master = request.args.get("master", "").strip()

    if not GOOGLE_SCRIPT:
        return jsonify({"success": False, "available_days": [],
                        "error": "GOOGLE_SCRIPT not configured"}), 500
    try:
        r = requests.get(GOOGLE_SCRIPT,
                         params={"action": "available_days",
                                 "month": month, "year": year, "master": master},
                         timeout=25)
        available = r.json().get("available", [])
        print(f"[elevenlabs/available_days] {month}/{year} master={master!r} → {available}",
              flush=True)
        return jsonify({"success": True, "month": month, "year": year,
                        "available_days": available})
    except Exception as e:
        print(f"[elevenlabs/available_days] error: {e}", flush=True)
        return jsonify({"success": False, "available_days": [], "error": str(e)}), 500


@app.route("/api/elevenlabs/notify", methods=["POST"])
def elevenlabs_notify():
    """Send SMS to admin when voice agent captures a new lead.
    Required JSON fields: breed, service, phone.
    """
    data = request.get_json() or {}
    breed   = data.get("breed", "").strip()
    service = data.get("service", "").strip()
    phone   = data.get("phone", "").strip()

    missing = [f for f in ("breed", "service", "phone") if not data.get(f)]
    if missing:
        return jsonify({"success": False,
                        "error": f"Missing required fields: {', '.join(missing)}"}), 400

    print(f"[elevenlabs/notify] breed={breed!r} service={service!r} phone={phone!r}", flush=True)

    twilio_sid   = os.environ.get("TWILIO_ACCOUNT_SID")
    twilio_token = os.environ.get("TWILIO_AUTH_TOKEN")
    twilio_from  = os.environ.get("TWILIO_PHONE", "+37266922128")
    admin_phone  = "+37266922128"

    if not (twilio_sid and twilio_token):
        return jsonify({"success": False, "error": "Twilio not configured"}), 500

    sms_body = f"Новый клиент с голосового агента: {breed}, {service}, тел: {phone}"
    sms_url  = f"https://api.twilio.com/2010-04-01/Accounts/{twilio_sid}/Messages.json"
    try:
        resp = requests.post(
            sms_url,
            auth=(twilio_sid, twilio_token),
            data={"From": twilio_from, "To": admin_phone, "Body": sms_body},
            timeout=10
        )
        print(f"[elevenlabs/notify] SMS → {resp.status_code}: {resp.text[:200]}", flush=True)
        if resp.status_code == 201:
            return jsonify({"success": True, "message": "Admin notified"})
        return jsonify({"success": False, "error": f"SMS error {resp.status_code}"}), 500
    except Exception as e:
        print(f"[elevenlabs/notify] error: {e}", flush=True)
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/cron/reminders")
def cron_reminders():
    """Send SMS reminders for tomorrow's bookings.
    Railway cron: 0 10 * * *
    Command:      curl https://rjgrooming.up.railway.app/cron/reminders
    Requires GAS action=bookings (see gas_cancel_reschedule.gs).
    """
    import datetime as _dt

    secret = os.environ.get("CRON_SECRET", "")
    if secret and request.args.get("secret") != secret:
        return "Unauthorized", 401

    if not GOOGLE_SCRIPT:
        return "GOOGLE_SCRIPT not configured", 500

    twilio_sid   = os.environ.get("TWILIO_ACCOUNT_SID")
    twilio_token = os.environ.get("TWILIO_AUTH_TOKEN")
    twilio_from  = os.environ.get("TWILIO_PHONE", "+37266922128")

    tomorrow  = _dt.date.today() + _dt.timedelta(days=1)
    date_gs   = tomorrow.strftime("%d.%m.%Y")   # DD.MM.YYYY for GAS

    # ── Fetch tomorrow's bookings from Google Script ──────────────────────
    try:
        r = requests.get(GOOGLE_SCRIPT,
                         params={"action": "bookings", "date": date_gs},
                         timeout=30)
        data = r.json()
    except Exception as e:
        print(f"[cron/reminders] GAS error: {e}", flush=True)
        return f"GAS error: {e}", 500

    bookings = data.get("bookings", [])
    print(f"[cron/reminders] {date_gs}: {len(bookings)} bookings", flush=True)

    sent, failed, skipped = [], [], []

    for b in bookings:
        phone  = re.sub(r'[\s\-()]', '', str(b.get("phone", "")).strip())
        time_  = b.get("time", "")
        master = b.get("master", "")
        lang   = b.get("lang", "ru")

        if not phone:
            skipped.append(f"no phone — {b.get('title', '?')}")
            continue
        if not phone.startswith("+"):
            phone = "+" + phone

        master_part = f"Мастер: {master}. " if master else ""

        if lang == "en":
            body = (f"R&J Grooming reminder: your appointment is tomorrow "
                    f"{date_gs} at {time_}. Groomer: {master}. "
                    "Address: Allveelaeva 4, Noblessner. See you! 🐾")
        elif lang == "et":
            body = (f"R&J Grooming meeldetuletus: teie broneering on homme "
                    f"{date_gs} kell {time_}. Meister: {master}. "
                    "Aadress: Allveelaeva 4, Noblessner. Ootame teid! 🐾")
        else:
            body = (f"Напоминаем о вашей записи в R&J Grooming завтра "
                    f"{date_gs} в {time_}. {master_part}"
                    "Адрес: Allveelaeva 4, Noblessner. Ждём вас! 🐾")

        if twilio_sid and twilio_token:
            try:
                resp = requests.post(
                    f"https://api.twilio.com/2010-04-01/Accounts/{twilio_sid}/Messages.json",
                    auth=(twilio_sid, twilio_token),
                    data={"From": twilio_from, "To": phone, "Body": body},
                    timeout=10,
                )
                if resp.status_code == 201:
                    sent.append(phone)
                    print(f"[cron/reminders] ✓ {phone}", flush=True)
                else:
                    failed.append(f"{phone}: {resp.status_code} {resp.text[:80]}")
                    print(f"[cron/reminders] ✗ {phone}: {resp.status_code}", flush=True)
            except Exception as e:
                failed.append(f"{phone}: {e}")
                print(f"[cron/reminders] ✗ {phone}: {e}", flush=True)
        else:
            skipped.append(f"Twilio not configured — would send to {phone}")
            print(f"[cron/reminders] Twilio not set — skipping {phone}", flush=True)

    summary = (f"Reminders {date_gs}: {len(bookings)} bookings | "
               f"sent={len(sent)} failed={len(failed)} skipped={len(skipped)}")
    print(f"[cron/reminders] {summary}", flush=True)
    lines = ([summary]
             + [f"✓ {p}" for p in sent]
             + [f"✗ {e}" for e in failed]
             + [f"- {s}" for s in skipped])
    return "\n".join(lines), 200

@app.route("/cron/review-request")
def cron_review_request():
    """Отправляет клиентам, посетившим салон ВЧЕРА, письмо + WhatsApp со ссылкой на отзыв.
    Railway cron: 0 11 * * *
    Command:      curl https://rjgrooming.up.railway.app/cron/review-request
    """
    import datetime as _dt

    secret = os.environ.get("CRON_SECRET", "")
    if secret and request.args.get("secret") != secret:
        return "Unauthorized", 401

    if not GOOGLE_SCRIPT:
        return "GOOGLE_SCRIPT not configured", 500

    review_link = os.environ.get("GOOGLE_REVIEW_LINK", "")
    resend_key = os.environ.get("RESEND_API_KEY")

    # ── Тестовый режим: ?test_email=you@example.com шлёт одно письмо сразу ──
    test_email = request.args.get("test_email", "").strip()
    if test_email:
        if not review_link:
            return "GOOGLE_REVIEW_LINK not set", 500
        if not resend_key:
            return "RESEND_API_KEY not set", 500
        try:
            rr = requests.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                json={
                    "from": "R&J Grooming <booking@rjgrooming.salon>",
                    "to": [test_email],
                    "subject": "[ТЕСТ] Спасибо, что были у нас! 🐾",
                    "html": (
                        "<div style='background:#0a0a09;padding:32px 24px;font-family:Arial,sans-serif;color:#f2ede2'>"
                        "<img src='https://rjgrooming.up.railway.app/assets/logo.png' alt='R&amp;J Grooming' style='height:60px;margin-bottom:14px;display:block'>"
                        "<p style='color:#cfc9ba'>Здравствуйте, Константин! Спасибо, что доверили нам уход за питомцем.</p>"
                        "<p style='color:#cfc9ba'>Будем очень благодарны, если оставите короткий отзыв — это помогает другим владельцам питомцев нас найти.</p>"
                        f"<p style='margin:24px 0'><a href='{review_link}' style='background:#e6e1d5;color:#0a0a09;text-decoration:none;padding:12px 22px;border-radius:8px;font-weight:bold;display:inline-block'>Оставить отзыв</a></p>"
                        "<p style='color:#8a8578;font-size:12px'>Это тестовое письмо, отправлено вручную для проверки.</p>"
                        "</div>"
                    )
                },
                timeout=10
            )
            if rr.status_code < 300:
                return f"Test email sent to {test_email}", 200
            else:
                return f"Resend error {rr.status_code}: {rr.text[:300]}", 500
        except Exception as e:
            return f"Error: {e}", 500

    # ── Тестовый режим: ?test_sms=+79114204546 шлёт одно SMS сразу ──
    test_sms = request.args.get("test_sms", "").strip()
    if test_sms:
        if not review_link:
            return "GOOGLE_REVIEW_LINK not set", 500
        twilio_sid = os.environ.get("TWILIO_ACCOUNT_SID")
        twilio_token = os.environ.get("TWILIO_AUTH_TOKEN")
        twilio_from = os.environ.get("TWILIO_PHONE", "+37266922128")
        if not (twilio_sid and twilio_token):
            return "TWILIO not configured", 500
        sms_to = test_sms if test_sms.startswith("+") else "+" + test_sms
        if sms_to.startswith("+7"):
            return "Twilio не доставляет SMS в РФ (+7) с января 2023 — тест на этот номер невозможен.", 200
        sms_text = f"[ТЕСТ] Спасибо, что были у нас в R&J Grooming! 🐾 Будем рады короткому отзыву: {review_link}"
        try:
            sr = requests.post(
                f"https://api.twilio.com/2010-04-01/Accounts/{twilio_sid}/Messages.json",
                auth=(twilio_sid, twilio_token),
                data={"From": twilio_from, "To": sms_to, "Body": sms_text},
                timeout=10,
            )
            if sr.status_code == 201:
                return f"Test SMS sent to {sms_to}", 200
            else:
                return f"Twilio error {sr.status_code}: {sr.text[:300]}", 500
        except Exception as e:
            return f"Error: {e}", 500

    yesterday = _dt.date.today() - _dt.timedelta(days=1)
    date_gs = yesterday.strftime("%d.%m.%Y")

    try:
        r = requests.get(GOOGLE_SCRIPT, params={"action": "bookings", "date": date_gs}, timeout=30)
        data = r.json()
    except Exception as e:
        print(f"[cron/review-request] GAS error: {e}", flush=True)
        return f"GAS error: {e}", 500

    bookings = data.get("bookings", [])
    print(f"[cron/review-request] {date_gs}: {len(bookings)} bookings", flush=True)

    twilio_sid = os.environ.get("TWILIO_ACCOUNT_SID")
    twilio_token = os.environ.get("TWILIO_AUTH_TOKEN")
    twilio_from = os.environ.get("TWILIO_PHONE", "+37266922128")
    email_sent, email_failed, email_skipped = [], [], []
    sms_sent, sms_failed, sms_skipped = [], [], []
    wa_sent, wa_failed, wa_skipped = [], [], []

    _REVIEW_MSG = {
        "ru": {
            "subject": "Спасибо, что были у нас! 🐾",
            "greet": "Здравствуйте, {name}! Спасибо, что доверили нам уход за питомцем.",
            "ask": "Будем очень благодарны, если оставите короткий отзыв — это помогает другим владельцам питомцев нас найти.",
            "btn": "Оставить отзыв",
            "sms": "Спасибо, что были у нас в R&J Grooming! 🐾 Будем рады короткому отзыву: {link}",
        },
        "en": {
            "subject": "Thank you for visiting us! 🐾",
            "greet": "Hi {name}! Thank you for trusting us with your pet's grooming.",
            "ask": "We'd really appreciate a short review — it helps other pet owners find us.",
            "btn": "Leave a review",
            "sms": "Thank you for visiting R&J Grooming! 🐾 We'd love a quick review: {link}",
        },
        "et": {
            "subject": "Aitäh, et meid külastasite! 🐾",
            "greet": "Tere, {name}! Aitäh, et usaldasite oma lemmiklooma hoolduse meile.",
            "ask": "Oleksime väga tänulikud lühikese arvustuse eest — see aitab teistel lemmikloomaomanikel meid leida.",
            "btn": "Jäta arvustus",
            "sms": "Aitäh, et külastasite R&J Grooming'ut! 🐾 Ootame lühikest arvustust: {link}",
        },
    }

    for b in bookings:
        name = b.get("clientName") or b.get("name") or ""
        phone = re.sub(r'[\s\-()]', '', str(b.get("phone", "")).strip())
        email = (b.get("clientEmail") or b.get("email") or "").strip()
        lang = (b.get("lang") or "ru").strip().lower()
        if lang not in _REVIEW_MSG:
            lang = "ru"
        t = _REVIEW_MSG[lang]

        if not review_link:
            email_skipped.append("GOOGLE_REVIEW_LINK not set")
            sms_skipped.append("GOOGLE_REVIEW_LINK not set")
            wa_skipped.append("GOOGLE_REVIEW_LINK not set")
            continue

        # ── Email через Resend ──────────────────────────────
        if email and "@" in email and resend_key:
            try:
                rr = requests.post(
                    "https://api.resend.com/emails",
                    headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                    json={
                        "from": "R&J Grooming <booking@rjgrooming.salon>",
                        "to": [email],
                        "subject": t["subject"],
                        "html": (
                            "<div style='background:#0a0a09;padding:32px 24px;font-family:Arial,sans-serif;color:#f2ede2'>"
                            "<img src='https://rjgrooming.up.railway.app/assets/logo.png' alt='R&amp;J Grooming' style='height:60px;margin-bottom:14px;display:block'>"
                            f"<p style='color:#cfc9ba'>{t['greet'].format(name=name)}</p>"
                            f"<p style='color:#cfc9ba'>{t['ask']}</p>"
                            f"<p style='margin:24px 0'><a href='{review_link}' style='background:#e6e1d5;color:#0a0a09;text-decoration:none;padding:12px 22px;border-radius:8px;font-weight:bold;display:inline-block'>{t['btn']}</a></p>"
                            "</div>"
                        )
                    },
                    timeout=10
                )
                if rr.status_code < 300:
                    email_sent.append(email)
                else:
                    email_failed.append(f"{email}: {rr.status_code} {rr.text[:80]}")
            except Exception as e:
                email_failed.append(f"{email}: {e}")
        else:
            email_skipped.append(email or phone or "no email/key")

        # ── SMS через Twilio ─────────────────────────────────
        sms_phone_norm = phone if phone.startswith("+") else "+" + phone
        if sms_phone_norm.startswith("+7"):
            sms_skipped.append(f"{sms_phone_norm}: Twilio не доставляет SMS в РФ с 2023")
        elif phone and twilio_sid and twilio_token:
            sms_text = t["sms"].format(link=review_link)
            try:
                sr = requests.post(
                    f"https://api.twilio.com/2010-04-01/Accounts/{twilio_sid}/Messages.json",
                    auth=(twilio_sid, twilio_token),
                    data={"From": twilio_from, "To": sms_phone_norm, "Body": sms_text},
                    timeout=10,
                )
                if sr.status_code == 201:
                    sms_sent.append(sms_phone_norm)
                else:
                    sms_failed.append(f"{sms_phone_norm}: {sr.status_code} {sr.text[:80]}")
            except Exception as e:
                sms_failed.append(f"{phone}: {e}")
        else:
            sms_skipped.append(phone or "no phone/twilio")

        # ── WhatsApp через Meta Business API ────────────────
        # Временно отключено: свободный текст на след. день после визита требует
        # одобренного message template в Meta Business Manager (freeform вне 24ч
        # окна будет отклонён). Включим, когда шаблон review_request_ru одобрят.
        if False and phone and WHATSAPP_TOKEN and WHATSAPP_PHONE_ID:
            if not phone.startswith("+"):
                phone = "+" + phone
            wa_text = (f"Здравствуйте, {name}! Спасибо, что были у нас в R&J Grooming 🐾 "
                       f"Будем рады короткому отзыву: {review_link}")
            try:
                wr = requests.post(
                    f"https://graph.facebook.com/v18.0/{WHATSAPP_PHONE_ID}/messages",
                    headers={"Authorization": f"Bearer {WHATSAPP_TOKEN}", "Content-Type": "application/json"},
                    json={"messaging_product": "whatsapp", "to": phone.lstrip("+"), "type": "text", "text": {"body": wa_text}},
                    timeout=10
                )
                if wr.status_code < 300:
                    wa_sent.append(phone)
                else:
                    wa_failed.append(f"{phone}: {wr.status_code} {wr.text[:150]}")
            except Exception as e:
                wa_failed.append(f"{phone}: {e}")
        else:
            wa_skipped.append("WhatsApp отключён (ждём одобрения шаблона)")

    summary = (f"Review requests {date_gs}: {len(bookings)} bookings | "
               f"email sent={len(email_sent)} failed={len(email_failed)} skipped={len(email_skipped)} | "
               f"sms sent={len(sms_sent)} failed={len(sms_failed)} skipped={len(sms_skipped)} | "
               f"wa sent={len(wa_sent)} failed={len(wa_failed)} skipped={len(wa_skipped)}")
    print(f"[cron/review-request] {summary}", flush=True)
    lines = ([summary]
             + [f"✓ email {e}" for e in email_sent]
             + [f"✗ email {e}" for e in email_failed]
             + [f"✓ sms {p}" for p in sms_sent]
             + [f"✗ sms {p}" for p in sms_failed]
             + [f"✓ wa {p}" for p in wa_sent]
             + [f"✗ wa {p}" for p in wa_failed])
    return "\n".join(lines), 200

@app.route("/admin/whatsapp")
def admin_whatsapp():
    global jarvis_enabled, instagram_enabled
    html = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>R&J Admin</title>
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{background:#1c1c18;color:#c8c2b8;font-family:'Montserrat',sans-serif;min-height:100vh;padding:20px 14px;font-size:16px;line-height:1.5}
h1{color:#c9a84c;font-size:1.3rem;font-weight:700;margin-bottom:20px;letter-spacing:.04em}
h2{color:#c9a84c;font-size:.95rem;font-weight:700;margin-bottom:14px;letter-spacing:.03em;text-transform:uppercase}
.card{background:#26261f;border-radius:14px;padding:18px 16px;margin-bottom:16px}
.bot-row{display:flex;align-items:center;gap:10px;margin-bottom:12px}
.bot-row:last-of-type{margin-bottom:0}
.bot-label{font-size:.9rem;min-width:80px}
.badge{display:inline-block;padding:5px 12px;border-radius:20px;font-size:.75rem;font-weight:700;min-width:52px;text-align:center}
.on{background:#2a4a2a;color:#6fcf6f}
.off{background:#4a2a2a;color:#cf6f6f}
.btn{display:block;width:100%;min-height:48px;border:none;border-radius:10px;font-family:inherit;font-size:1rem;font-weight:700;cursor:pointer;transition:.15s;letter-spacing:.02em;padding:0 16px}
.btn-gold{background:#c9a84c;color:#1c1c18}
.btn-gold:active{background:#b89640}
.btn-row{margin-top:14px}
#msg{font-size:.85rem;color:#c9a84c;margin-top:10px;min-height:1.2em;text-align:center}
.client-item{padding:12px 0;border-bottom:1px solid #2e2e26;display:flex;flex-direction:column;gap:3px}
.client-item:last-child{border-bottom:none}
.c-phone{font-size:1rem;font-weight:600;color:#e8e0d0}
.c-text{font-size:.85rem;color:#a09880;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.c-meta{display:flex;gap:8px;align-items:center}
.c-time{font-size:.75rem;color:#666}
.c-ch{font-size:.7rem;color:#888;text-transform:uppercase;background:#2e2e26;padding:2px 7px;border-radius:10px}
#clients-empty{font-size:.85rem;color:#666;text-align:center;padding:12px 0}
select{width:100%;background:#1c1c18;border:1px solid #3a3a30;border-radius:10px;color:#c8c2b8;font-family:inherit;font-size:1rem;padding:14px 12px;margin-bottom:12px;appearance:none;-webkit-appearance:none}
select:focus{outline:none;border-color:#c9a84c}
textarea{width:100%;background:#1c1c18;border:1px solid #3a3a30;border-radius:10px;color:#c8c2b8;font-family:inherit;font-size:1rem;padding:14px 12px;margin-bottom:12px;resize:vertical;min-height:100px;line-height:1.5}
textarea:focus{outline:none;border-color:#c9a84c}
#sendMsg{font-size:.85rem;color:#c9a84c;margin-top:10px;min-height:1.2em;text-align:center}
.counter-row{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:4px}
.counter-item{background:#1c1c18;border-radius:10px;padding:10px 12px;flex:1;min-width:72px;text-align:center}
.counter-num{font-size:1.5rem;font-weight:700;color:#c9a84c;line-height:1.1}
.counter-lbl{font-size:.65rem;color:#888;text-transform:uppercase;margin-top:3px}
.msg-item{padding:8px 0;border-bottom:1px solid #2e2e26}
.msg-item:last-child{border-bottom:none}
.msg-head{display:flex;gap:6px;align-items:center;margin-bottom:2px}
.msg-dir{font-size:.7rem;font-weight:700}
.msg-in .msg-dir{color:#6fcf6f}
.msg-out .msg-dir{color:#c9a84c}
.msg-ph{font-size:.73rem;color:#a09880;flex:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.msg-t{font-size:.68rem;color:#555;white-space:nowrap}
.msg-body{font-size:.8rem;color:#c8c2b8;line-height:1.35;overflow:hidden;text-overflow:ellipsis;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical}
</style>
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700&display=swap" rel="stylesheet">
</head>
<body>
<h1>R&J Grooming — Admin</h1>

<div class="card">
  <h2>Боты</h2>
  <div class="bot-row">
    <span class="bot-label">Jarvis</span>
    <span class="badge" id="jarvisBadge"></span>
  </div>
  <div class="btn-row"><button class="btn btn-gold" id="jarvisBtn" onclick="toggleJarvis()"></button></div>
  <div class="bot-row" style="margin-top:14px">
    <span class="bot-label">Instagram</span>
    <span class="badge" id="igBadge"></span>
  </div>
  <div class="btn-row"><button class="btn btn-gold" id="igBtn" onclick="toggleIg()"></button></div>
  <div class="bot-row" style="margin-top:14px">
    <span class="bot-label">Jarvis WA</span>
    <span class="badge" id="waJarvisBadge"></span>
  </div>
  <div class="btn-row"><button class="btn btn-gold" id="waJarvisBtn" onclick="toggleWaJarvis()"></button></div>
  <div id="msg"></div>
</div>

<div class="card">
  <h2>Клиенты</h2>
  <div id="clientsList"><div id="clients-empty">Нет данных</div></div>
</div>

<div class="card">
  <h2>Статистика сообщений</h2>
  <div class="counter-row">
    <div class="counter-item"><div class="counter-num" id="cntWa">0</div><div class="counter-lbl">WhatsApp</div></div>
    <div class="counter-item"><div class="counter-num" id="cntIg">0</div><div class="counter-lbl">Instagram</div></div>
    <div class="counter-item"><div class="counter-num" id="cntFb">0</div><div class="counter-lbl">Facebook</div></div>
  </div>
</div>

<div class="card">
  <h2>Последние сообщения</h2>
  <div id="msgLog"><div style="font-size:.85rem;color:#666;text-align:center;padding:12px 0">Нет данных</div></div>
</div>

<div class="card">
  <h2>Отправить сообщение</h2>
  <select id="selPhone"><option value="">— выбрать клиента —</option></select>
  <textarea id="txtMsg" placeholder="Текст сообщения..."></textarea>
  <button class="btn btn-gold" onclick="sendManual()">Отправить в WhatsApp</button>
  <div id="sendMsg"></div>
</div>

<script>
var jarvisOn = """ + ("true" if jarvis_enabled else "false") + """;
var igOn = """ + ("true" if instagram_enabled else "false") + """;
var waJarvisOn = true;

function setBadge(id, on){
  var el=document.getElementById(id);
  el.className='badge '+(on?'on':'off');
  el.textContent=on?'ВКЛ':'ВЫКЛ';
}
function updateBadges(){
  setBadge('jarvisBadge', jarvisOn);
  document.getElementById('jarvisBtn').textContent=jarvisOn?'Выключить Jarvis':'Включить Jarvis';
  setBadge('igBadge', igOn);
  document.getElementById('igBtn').textContent=igOn?'Выключить Instagram':'Включить Instagram';
  setBadge('waJarvisBadge', waJarvisOn);
  document.getElementById('waJarvisBtn').textContent=waJarvisOn?'Выключить Jarvis WA':'Включить Jarvis WA';
}
updateBadges();

function toggleJarvis(){
  fetch('/admin/api/toggle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({enabled:!jarvisOn})})
    .then(r=>r.json()).then(d=>{jarvisOn=!jarvisOn;updateBadges();document.getElementById('msg').textContent=d.message||'';});
}
function toggleIg(){
  fetch('/admin/api/toggle-instagram',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({enabled:!igOn})})
    .then(r=>r.json()).then(d=>{igOn=!igOn;updateBadges();document.getElementById('msg').textContent=d.message||'';});
}
function toggleWaJarvis(){
  var msgEl=document.getElementById('msg');
  msgEl.textContent='...';
  fetch('http://mydeal.railway.internal/toggle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({enabled:!waJarvisOn})})
    .then(r=>r.json()).then(d=>{
      waJarvisOn=!waJarvisOn;updateBadges();
      msgEl.textContent='Jarvis WA '+(waJarvisOn?'включён ✅':'выключен ❌');
    }).catch(function(){msgEl.textContent='Ошибка: недоступен Jarvis WA сервис';});
}

function loadClients(){
  fetch('/admin/api/messages').then(r=>r.json()).then(data=>{
    var list=document.getElementById('clientsList');
    var sel=document.getElementById('selPhone');
    var cur=sel.value;
    sel.innerHTML='<option value="">— выбрать клиента —</option>';
    data.sort(function(a,b){return b.ts.localeCompare(a.ts);});
    if(!data.length){list.innerHTML='<div id="clients-empty">Нет клиентов</div>';return;}
    list.innerHTML='';
    data.forEach(function(c){
      var ts=c.ts?new Date(c.ts).toLocaleString('ru-RU',{day:'2-digit',month:'2-digit',hour:'2-digit',minute:'2-digit'}):'';
      var div=document.createElement('div');
      div.className='client-item';
      div.innerHTML='<div class="c-phone">'+esc(c.phone)+'</div>'
        +'<div class="c-text">'+esc(c.last_text||'—')+'</div>'
        +'<div class="c-meta"><span class="c-ch">'+esc(c.channel||'wa')+'</span><span class="c-time">'+ts+'</span></div>';
      list.appendChild(div);
      var opt=document.createElement('option');
      opt.value=c.phone; opt.textContent=c.phone+(c.last_text?' — '+c.last_text.substring(0,28):'');
      sel.appendChild(opt);
    });
    if(cur) sel.value=cur;
  });
}

function esc(s){var d=document.createElement('div');d.appendChild(document.createTextNode(String(s)));return d.innerHTML;}

function sendManual(){
  var phone=document.getElementById('selPhone').value;
  var text=document.getElementById('txtMsg').value.trim();
  var msgEl=document.getElementById('sendMsg');
  if(!phone){msgEl.textContent='Выберите клиента';return;}
  if(!text){msgEl.textContent='Введите текст';return;}
  msgEl.textContent='Отправка...';
  fetch('/admin/api/send',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({phone:phone,text:text})})
    .then(r=>r.json()).then(d=>{
      if(d.ok){msgEl.textContent='✓ Отправлено';document.getElementById('txtMsg').value='';}
      else{msgEl.textContent='Ошибка: '+(d.error||'неизвестно');}
    });
}

loadClients();
setInterval(loadClients, 15000);

function loadMsgLog(){
  fetch('/admin/api/message-log').then(r=>r.json()).then(data=>{
    document.getElementById('cntWa').textContent=data.counts.whatsapp||0;
    document.getElementById('cntIg').textContent=data.counts.instagram||0;
    document.getElementById('cntFb').textContent=data.counts.facebook||0;
    var log=document.getElementById('msgLog');
    if(!data.log||!data.log.length){log.innerHTML='<div style="font-size:.85rem;color:#666;text-align:center;padding:12px 0">Нет сообщений</div>';return;}
    log.innerHTML='';
    data.log.forEach(function(m){
      var ts=m.ts?new Date(m.ts).toLocaleString('ru-RU',{day:'2-digit',month:'2-digit',hour:'2-digit',minute:'2-digit'}):'';
      var div=document.createElement('div');
      div.className='msg-item msg-'+(m.direction==='in'?'in':'out');
      div.innerHTML='<div class="msg-head"><span class="msg-dir">'+(m.direction==='in'?'←':'→')+'</span>'
        +'<span class="c-ch">'+esc(m.channel||'wa')+'</span>'
        +'<span class="msg-ph">'+esc(m.phone||'')+'</span>'
        +'<span class="msg-t">'+ts+'</span></div>'
        +'<div class="msg-body">'+esc(m.text||'')+'</div>';
      log.appendChild(div);
    });
  }).catch(function(){});
}
loadMsgLog();
setInterval(loadMsgLog, 15000);
</script>
</body>
</html>"""
    return html

# ── PUBLIC ─────────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return "MyDeal Jarvis rabotaet!"

@app.route("/privacy")
def privacy():
    return "<h1>Privacy Policy</h1><p>R&J Grooming, Tallinn, Estonia</p>"

@app.route("/terms")
def terms():
    return "<h1>Terms of Service</h1><p>R&J Grooming, Tallinn, Estonia</p>"


# ── STATS PAGE ─────────────────────────────────────────────────────────────
STATS_HTML_B64 = "PCFET0NUWVBFIGh0bWw+CjxodG1sIGxhbmc9InJ1Ij4KPGhlYWQ+CjxtZXRhIGNoYXJzZXQ9IlVURi04Ij4KPG1ldGEgbmFtZT0idmlld3BvcnQiIGNvbnRlbnQ9IndpZHRoPWRldmljZS13aWR0aCxpbml0aWFsLXNjYWxlPTEiPgo8bWV0YSBuYW1lPSJ0aGVtZS1jb2xvciIgY29udGVudD0iIzFjMWMxOCI+Cjx0aXRsZT5SJkogU3RhdHM8L3RpdGxlPgo8bGluayBocmVmPSJodHRwczovL2ZvbnRzLmdvb2dsZWFwaXMuY29tL2NzczI/ZmFtaWx5PUNvcm1vcmFudCtHYXJhbW9uZDppdGFsLHdnaHRAMCw2MDA7MSw0MDAmZmFtaWx5PU1vbnRzZXJyYXQ6d2dodEAzMDA7NDAwOzUwMDs2MDAmZGlzcGxheT1zd2FwIiByZWw9InN0eWxlc2hlZXQiPgo8c3R5bGU+Cip7Ym94LXNpemluZzpib3JkZXItYm94O21hcmdpbjowO3BhZGRpbmc6MH0KaHRtbCxib2R5e21pbi1oZWlnaHQ6MTAwdmg7YmFja2dyb3VuZDojMWMxYzE4O2NvbG9yOiNjOGMyYjg7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWY7Zm9udC13ZWlnaHQ6MzAwfQouY29ue3dpZHRoOjEwMCU7bWF4LXdpZHRoOjUwMHB4O3BhZGRpbmc6MCAyMnB4O21hcmdpbjowIGF1dG99Ci5wYWdle3BhZGRpbmc6MzJweCAwIDYwcHh9Ci5sb2dvLXJqe2ZvbnQtZmFtaWx5OidDb3Jtb3JhbnQgR2FyYW1vbmQnLHNlcmlmO2ZvbnQtc2l6ZToxLjhyZW07Zm9udC13ZWlnaHQ6NjAwO2NvbG9yOiNlOGUwZDB9Ci5iYWNrLWxpbmt7ZGlzcGxheTppbmxpbmUtYmxvY2s7Zm9udC1zaXplOjAuNzRyZW07Y29sb3I6I2M5YTA1YTt0ZXh0LWRlY29yYXRpb246bm9uZTttYXJnaW4tYm90dG9tOjE0cHh9Ci5sb2dvLXN1Yntmb250LXNpemU6LjQ0cmVtO2xldHRlci1zcGFjaW5nOi40ZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2NvbG9yOiM4YThhNTU7bWFyZ2luLXRvcDoycHg7cGFkZGluZy1ib3R0b206MTZweDtib3JkZXItYm90dG9tOjFweCBzb2xpZCByZ2JhKDIwMSwxNjgsNzYsLjIpO21hcmdpbi1ib3R0b206MjBweH0KLnRhYnMtbWFpbntkaXNwbGF5OmZsZXg7Z2FwOjA7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyMDEsMTY4LDc2LC4yKTttYXJnaW4tYm90dG9tOjB9Ci50bWJ7cGFkZGluZzoxMnB4IDE4cHg7Zm9udC1zaXplOi41NnJlbTtsZXR0ZXItc3BhY2luZzouMThlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Y29sb3I6IzU1NTU1MDtjdXJzb3I6cG9pbnRlcjtib3JkZXItYm90dG9tOjJweCBzb2xpZCB0cmFuc3BhcmVudDtiYWNrZ3JvdW5kOm5vbmU7Ym9yZGVyLXRvcDpub25lO2JvcmRlci1sZWZ0Om5vbmU7Ym9yZGVyLXJpZ2h0Om5vbmU7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWY7dHJhbnNpdGlvbjphbGwgLjJzfQoudG1iLmFjdGl2ZXtjb2xvcjojYzlhODRjO2JvcmRlci1ib3R0b20tY29sb3I6I2M5YTg0Y30KLnRtYjpob3Zlcntjb2xvcjojYzhjMmI4fQoucGFuZWx7ZGlzcGxheTpub25lO3BhZGRpbmc6MjBweCAwfQoucGFuZWwuYWN0aXZle2Rpc3BsYXk6YmxvY2t9Ci5zbGJse2ZvbnQtc2l6ZTouNTJyZW07bGV0dGVyLXNwYWNpbmc6LjJlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Y29sb3I6I2M5YTg0YzttYXJnaW4tYm90dG9tOjEwcHg7Zm9udC13ZWlnaHQ6NTAwfQoubWV0cmljc3tkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOnJlcGVhdCgzLDFmcik7Z2FwOjZweDttYXJnaW4tYm90dG9tOjIycHh9Ci5tZXRyaWN7YmFja2dyb3VuZDojMTQxNDEwO2JvcmRlcjoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDcpO3BhZGRpbmc6MTJweCAxNHB4fQoubWV0cmljLWxhYmVse2ZvbnQtc2l6ZTouNTRyZW07bGV0dGVyLXNwYWNpbmc6LjFlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7Y29sb3I6IzU1NTU1MDttYXJnaW4tYm90dG9tOjVweH0KLm1ldHJpYy12YWx7Zm9udC1mYW1pbHk6J0Nvcm1vcmFudCBHYXJhbW9uZCcsc2VyaWY7Zm9udC1zaXplOjEuMzVyZW07Y29sb3I6I2M5YTg0Yztmb250LXdlaWdodDo2MDB9Ci5tZXRyaWMtc3Vie2ZvbnQtc2l6ZTouNThyZW07Y29sb3I6IzQ0NDQ0MDttYXJnaW4tdG9wOjJweH0KLmRpc2NvdW50LXJvd3tkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDoxMHB4O2JhY2tncm91bmQ6IzE0MTQxMDtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjAxLDE2OCw3NiwuMik7cGFkZGluZzoxMnB4IDE0cHg7bWFyZ2luLWJvdHRvbToxOHB4fQouZGlzY291bnQtbGFiZWx7Zm9udC1zaXplOi41OHJlbTtsZXR0ZXItc3BhY2luZzouMWVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtjb2xvcjojOGE4YTU1O2ZsZXg6MX0KLmRpc2NvdW50LWlucHV0e2JhY2tncm91bmQ6dHJhbnNwYXJlbnQ7Ym9yZGVyOm5vbmU7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyMDEsMTY4LDc2LC40KTtjb2xvcjojYzlhODRjO2ZvbnQtZmFtaWx5OidDb3Jtb3JhbnQgR2FyYW1vbmQnLHNlcmlmO2ZvbnQtc2l6ZToxLjJyZW07Zm9udC13ZWlnaHQ6NjAwO3dpZHRoOjgwcHg7dGV4dC1hbGlnbjpyaWdodDtvdXRsaW5lOm5vbmU7cGFkZGluZzoycHggNHB4fQouZGlzY291bnQtaW5wdXQ6Zm9jdXN7Ym9yZGVyLWJvdHRvbS1jb2xvcjojYzlhODRjfQouZGlzY291bnQtZXVye2NvbG9yOiM4YThhNTU7Zm9udC1zaXplOi43NXJlbTttYXJnaW4tbGVmdDoycHh9Ci5wZXJpb2Qtcm93e2Rpc3BsYXk6ZmxleDtnYXA6OHB4O21hcmdpbi1ib3R0b206MThweDthbGlnbi1pdGVtczpjZW50ZXJ9Ci5wZXJpb2Qtc2VsZWN0e2JhY2tncm91bmQ6IzE0MTQxMDtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjEpO2NvbG9yOiNjOGMyYjg7Zm9udC1mYW1pbHk6J01vbnRzZXJyYXQnLHNhbnMtc2VyaWY7Zm9udC1zaXplOi43MnJlbTtwYWRkaW5nOjhweCAxMnB4O291dGxpbmU6bm9uZTtmbGV4OjF9Ci5wZXJpb2Qtc2VsZWN0OmZvY3Vze2JvcmRlci1jb2xvcjojYzlhODRjfQoucmVmcmVzaC1idG57YmFja2dyb3VuZDpyZ2JhKDIwMSwxNjgsNzYsLjEyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjAxLDE2OCw3NiwuMyk7Y29sb3I6I2M5YTg0YztwYWRkaW5nOjhweCAxNHB4O2N1cnNvcjpwb2ludGVyO2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmO2ZvbnQtc2l6ZTouNThyZW07bGV0dGVyLXNwYWNpbmc6LjE0ZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO3doaXRlLXNwYWNlOm5vd3JhcH0KLnJlZnJlc2gtYnRuOmhvdmVye2JhY2tncm91bmQ6cmdiYSgyMDEsMTY4LDc2LC4yMil9Ci5tYXN0ZXItY2FyZHtiYWNrZ3JvdW5kOiMxNDE0MTA7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4wNyk7Ym9yZGVyLWxlZnQ6M3B4IHNvbGlkIHJnYmEoMjAxLDE2OCw3NiwuMyk7bWFyZ2luLWJvdHRvbTo2cHg7b3ZlcmZsb3c6aGlkZGVufQoubWFzdGVyLWhlYWR7ZGlzcGxheTpmbGV4O2p1c3RpZnktY29udGVudDpzcGFjZS1iZXR3ZWVuO2FsaWduLWl0ZW1zOmNlbnRlcjtwYWRkaW5nOjEzcHggMTVweDtjdXJzb3I6cG9pbnRlcjt0cmFuc2l0aW9uOmJhY2tncm91bmQgLjJzfQoubWFzdGVyLWhlYWQ6aG92ZXJ7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wMil9Ci5tYXN0ZXItaGVhZC5vcGVue2JvcmRlci1sZWZ0LWNvbG9yOiNjOWE4NGN9Ci5tbmFtZXtmb250LXNpemU6Ljg4cmVtO2ZvbnQtd2VpZ2h0OjUwMDtjb2xvcjojZThlMGQwfQoubWNvdW50e2ZvbnQtc2l6ZTouNnJlbTtjb2xvcjojNTU1NTUwO21hcmdpbi10b3A6MnB4fQoubWVhcm5pbmdze2Rpc3BsYXk6ZmxleDtnYXA6MTRweDthbGlnbi1pdGVtczpjZW50ZXJ9Ci5lYXJuLWl0ZW17dGV4dC1hbGlnbjpyaWdodH0KLmVhcm4tbGFiZWx7Zm9udC1zaXplOi41cmVtO2NvbG9yOiM1NTU1NTA7bGV0dGVyLXNwYWNpbmc6LjFlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2V9Ci5lYXJuLXZhbHtmb250LXNpemU6Ljg4cmVtO2NvbG9yOiNjOWE4NGM7Zm9udC13ZWlnaHQ6NTAwfQouZWFybi12YWwuc2Fsb257Y29sb3I6IzhhOGE1NX0KLmNoZXZyb257Y29sb3I6IzhhOGE1NTtmb250LXNpemU6Ljc1cmVtO3RyYW5zaXRpb246dHJhbnNmb3JtIC4yNXM7bWFyZ2luLWxlZnQ6OHB4fQoubWFzdGVyLWhlYWQub3BlbiAuY2hldnJvbnt0cmFuc2Zvcm06cm90YXRlKDE4MGRlZyl9Ci5tYXN0ZXItYm9keXtkaXNwbGF5Om5vbmU7Ym9yZGVyLXRvcDoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDYpfQoubWFzdGVyLWJvZHkub3BlbntkaXNwbGF5OmJsb2NrfQoudmlzaXRzLWhlYWRlcntkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOjcycHggMWZyIDg1cHggNjBweDtnYXA6NnB4O3BhZGRpbmc6OHB4IDE1cHg7Zm9udC1zaXplOi41cmVtO2NvbG9yOiM0NDQ0NDA7bGV0dGVyLXNwYWNpbmc6LjFlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wMil9Ci52aXNpdC1yb3d7ZGlzcGxheTpncmlkO2dyaWQtdGVtcGxhdGUtY29sdW1uczo3MnB4IDFmciA4NXB4IDYwcHg7Z2FwOjZweDtwYWRkaW5nOjlweCAxNXB4O2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjA0KTtmb250LXNpemU6LjcycmVtO2FsaWduLWl0ZW1zOnN0YXJ0fQoudmlzaXQtcm93Omxhc3QtY2hpbGR7Ym9yZGVyLWJvdHRvbTpub25lfQoudmlzaXQtZGF0ZXtjb2xvcjojNTU1NTUwfQoudmlzaXQtY2xpZW50e2NvbG9yOiNjOGMyYjg7Zm9udC1zaXplOi43NXJlbX0KLnZpc2l0LXBldHtjb2xvcjojNTU1NTUwO2ZvbnQtc2l6ZTouNjJyZW07bWFyZ2luLXRvcDoxcHh9Ci52aXNpdC1zdmN7Y29sb3I6IzY2NjY2MDtmb250LXNpemU6LjY1cmVtfQoudmlzaXQtcHJpY2V7Y29sb3I6I2M5YTg0Yzt0ZXh0LWFsaWduOnJpZ2h0O2ZvbnQtZmFtaWx5OidDb3Jtb3JhbnQgR2FyYW1vbmQnLHNlcmlmO2ZvbnQtc2l6ZTouOTVyZW07Zm9udC13ZWlnaHQ6NjAwfQoubm8tdmlzaXRze3BhZGRpbmc6MTZweCAxNXB4O2ZvbnQtc2l6ZTouNzJyZW07Y29sb3I6IzQ0NDQ0MDtmb250LXN0eWxlOml0YWxpY30KLnN1Yi10YWJze2Rpc3BsYXk6ZmxleDtnYXA6NXB4O21hcmdpbi1ib3R0b206MThweH0KLnN0YntwYWRkaW5nOjdweCAxNHB4O2ZvbnQtc2l6ZTouNTRyZW07bGV0dGVyLXNwYWNpbmc6LjE0ZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2NvbG9yOiM1NTU1NTA7Y3Vyc29yOnBvaW50ZXI7YmFja2dyb3VuZDojMTQxNDEwO2JvcmRlcjoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDcpO2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmO3RyYW5zaXRpb246YWxsIC4yc30KLnN0Yi5hY3RpdmV7Y29sb3I6I2M5YTg0Yztib3JkZXItY29sb3I6cmdiYSgyMDEsMTY4LDc2LC40KTtiYWNrZ3JvdW5kOnJnYmEoMjAxLDE2OCw3NiwuMDYpfQouc3RiOmhvdmVye2NvbG9yOiNjOGMyYjh9Ci5zdWItcGFuZWx7ZGlzcGxheTpub25lfQouc3ViLXBhbmVsLmFjdGl2ZXtkaXNwbGF5OmJsb2NrfQouZml7d2lkdGg6MTAwJTtiYWNrZ3JvdW5kOiMxNDE0MTA7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4xKTtjb2xvcjojYzhjMmI4O2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmO2ZvbnQtc2l6ZTouODJyZW07cGFkZGluZzoxMXB4IDE0cHg7b3V0bGluZTpub25lO21hcmdpbi1ib3R0b206OHB4fQouZmk6Zm9jdXN7Ym9yZGVyLWNvbG9yOiNjOWE4NGN9CnNlbGVjdC5maXthcHBlYXJhbmNlOm5vbmU7LXdlYmtpdC1hcHBlYXJhbmNlOm5vbmV9Ci5maS1yb3d7ZGlzcGxheTpncmlkO2dyaWQtdGVtcGxhdGUtY29sdW1uczoxZnIgMWZyO2dhcDo4cHh9Ci5jYnRue2Rpc3BsYXk6YmxvY2s7d2lkdGg6MTAwJTtiYWNrZ3JvdW5kOiM0YTRhMmU7Y29sb3I6I2M4YzJiODtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZjtmb250LXNpemU6LjZyZW07Zm9udC13ZWlnaHQ6NjAwO2xldHRlci1zcGFjaW5nOi4yMmVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtwYWRkaW5nOjEzcHg7Ym9yZGVyOm5vbmU7Y3Vyc29yOnBvaW50ZXI7bWFyZ2luLXRvcDo0cHg7dHJhbnNpdGlvbjpiYWNrZ3JvdW5kIC4yc30KLmNidG46aG92ZXJ7YmFja2dyb3VuZDojNmI2YjQyfQouY2J0bi5naG9zdHtiYWNrZ3JvdW5kOm5vbmU7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDIwMSwxNjgsNzYsLjMpO2NvbG9yOiNjOWE4NGN9Ci5jYnRuLmdob3N0OmhvdmVye2JhY2tncm91bmQ6cmdiYSgyMDEsMTY4LDc2LC4wOCl9Ci5saXN0LWl0ZW17ZGlzcGxheTpmbGV4O2p1c3RpZnktY29udGVudDpzcGFjZS1iZXR3ZWVuO2FsaWduLWl0ZW1zOmNlbnRlcjtwYWRkaW5nOjExcHggMTRweDtiYWNrZ3JvdW5kOiMxNDE0MTA7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4wNyk7bWFyZ2luLWJvdHRvbTo1cHh9Ci5saS1uYW1le2ZvbnQtc2l6ZTouODJyZW07Y29sb3I6I2M4YzJiOH0KLmxpLXN1Yntmb250LXNpemU6LjZyZW07Y29sb3I6IzU1NTU1MDttYXJnaW4tdG9wOjJweH0KLmRlbC1idG57YmFja2dyb3VuZDpub25lO2JvcmRlcjpub25lO2NvbG9yOiM0NDQ0NDA7Y3Vyc29yOnBvaW50ZXI7Zm9udC1zaXplOi44cmVtO3BhZGRpbmc6NHB4IDhweDt0cmFuc2l0aW9uOmNvbG9yIC4yc30KLmRlbC1idG46aG92ZXJ7Y29sb3I6I2MwNTA1MH0KLmJyZWVkLWNhcmR7YmFja2dyb3VuZDojMTQxNDEwO2JvcmRlcjoxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDcpO2JvcmRlci1sZWZ0OjNweCBzb2xpZCByZ2JhKDIwMSwxNjgsNzYsLjIpO21hcmdpbi1ib3R0b206NnB4fQouYnJlZWQtaGVhZHtkaXNwbGF5OmZsZXg7anVzdGlmeS1jb250ZW50OnNwYWNlLWJldHdlZW47YWxpZ24taXRlbXM6Y2VudGVyO3BhZGRpbmc6MTFweCAxNHB4O2N1cnNvcjpwb2ludGVyfQouYnJlZWQtaGVhZDpob3ZlcntiYWNrZ3JvdW5kOnJnYmEoMjU1LDI1NSwyNTUsLjAyKX0KLmJyZWVkLW5hbWV7Zm9udC1zaXplOi44MnJlbTtjb2xvcjojZThlMGQwfQouYnJlZWQtY291bnR7Zm9udC1zaXplOi42cmVtO2NvbG9yOiM1NTU1NTB9Ci5icmVlZC1ib2R5e2Rpc3BsYXk6bm9uZTtib3JkZXItdG9wOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4wNik7cGFkZGluZzoxMHB4IDE0cHh9Ci5icmVlZC1ib2R5Lm9wZW57ZGlzcGxheTpibG9ja30KLnN2Yy1yb3d7ZGlzcGxheTpmbGV4O2p1c3RpZnktY29udGVudDpzcGFjZS1iZXR3ZWVuO2FsaWduLWl0ZW1zOmNlbnRlcjtwYWRkaW5nOjZweCAwO2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjA0KTtmb250LXNpemU6Ljc1cmVtfQouc3ZjLXJvdzpsYXN0LWNoaWxke2JvcmRlci1ib3R0b206bm9uZX0KLnN2Yy1uYW1le2NvbG9yOiNjOGMyYjh9Ci5zdmMtcHJpY2V7Y29sb3I6I2M5YTg0Yztmb250LWZhbWlseTonQ29ybW9yYW50IEdhcmFtb25kJyxzZXJpZjtmb250LXNpemU6Ljk1cmVtO2ZvbnQtd2VpZ2h0OjYwMH0KLmFkZC1zdmMtcm93e2Rpc3BsYXk6Z3JpZDtncmlkLXRlbXBsYXRlLWNvbHVtbnM6MWZyIDcwcHggMzRweDtnYXA6NnB4O21hcmdpbi10b3A6MTBweH0KLmFkZC1zdmMtcm93IC5maXttYXJnaW4tYm90dG9tOjA7Zm9udC1zaXplOi43NXJlbTtwYWRkaW5nOjhweCAxMHB4fQouaWNvbi1idG57d2lkdGg6MzRweDtoZWlnaHQ6MzRweDtiYWNrZ3JvdW5kOnJnYmEoMjAxLDE2OCw3NiwuMTIpO2JvcmRlcjoxcHggc29saWQgcmdiYSgyMDEsMTY4LDc2LC4zKTtjb2xvcjojYzlhODRjO2N1cnNvcjpwb2ludGVyO2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7anVzdGlmeS1jb250ZW50OmNlbnRlcjtmb250LXNpemU6MS4xcmVtO2ZvbnQtd2VpZ2h0OjMwMH0KLmljb24tYnRuOmhvdmVye2JhY2tncm91bmQ6cmdiYSgyMDEsMTY4LDc2LC4yNCl9Ci5jbGllbnQtY2FyZHtiYWNrZ3JvdW5kOiMxNDE0MTA7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4wNyk7cGFkZGluZzoxNHB4O21hcmdpbi1ib3R0b206NnB4fQouY2wtcm93e2Rpc3BsYXk6ZmxleDtqdXN0aWZ5LWNvbnRlbnQ6c3BhY2UtYmV0d2VlbjthbGlnbi1pdGVtczpmbGV4LXN0YXJ0fQouY2wtYXZhdGFye3dpZHRoOjM2cHg7aGVpZ2h0OjM2cHg7Ym9yZGVyLXJhZGl1czo1MCU7YmFja2dyb3VuZDpyZ2JhKDIwMSwxNjgsNzYsLjEyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjAxLDE2OCw3NiwuMjUpO2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7anVzdGlmeS1jb250ZW50OmNlbnRlcjtmb250LWZhbWlseTonQ29ybW9yYW50IEdhcmFtb25kJyxzZXJpZjtmb250LXNpemU6Ljg4cmVtO2ZvbnQtd2VpZ2h0OjYwMDtjb2xvcjojYzlhODRjO2ZsZXgtc2hyaW5rOjA7bWFyZ2luLXJpZ2h0OjEwcHh9Ci5jbC1uYW1le2ZvbnQtc2l6ZTouODVyZW07Zm9udC13ZWlnaHQ6NTAwO2NvbG9yOiNlOGUwZDB9Ci5jbC1kZXRhaWx7Zm9udC1zaXplOi42MnJlbTtjb2xvcjojNTU1NTUwO21hcmdpbi10b3A6MnB4fQouY2wtc3RhdHN7ZGlzcGxheTpmbGV4O2dhcDoxNHB4O21hcmdpbi10b3A6MTBweDtwYWRkaW5nLXRvcDoxMHB4O2JvcmRlci10b3A6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjA1KX0KLmNzdC12YWx7Zm9udC1mYW1pbHk6J0Nvcm1vcmFudCBHYXJhbW9uZCcsc2VyaWY7Zm9udC1zaXplOjEuMXJlbTtjb2xvcjojYzlhODRjO2ZvbnQtd2VpZ2h0OjYwMH0KLmNzdC1sYWJlbHtmb250LXNpemU6LjUycmVtO2NvbG9yOiM1NTU1NTA7dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2xldHRlci1zcGFjaW5nOi4xZW19Ci5jbC1sYXN0e2ZvbnQtc2l6ZTouNnJlbTtjb2xvcjojNTU1NTUwO21hcmdpbi10b3A6OHB4fQouY2wtbGFzdCBzcGFue2NvbG9yOiM4YThhNTV9Ci5jbC1iYWRnZXtmb250LXNpemU6LjU4cmVtO2NvbG9yOiNjOWE4NGM7YmFja2dyb3VuZDpyZ2JhKDIwMSwxNjgsNzYsLjEpO2JvcmRlcjoxcHggc29saWQgcmdiYSgyMDEsMTY4LDc2LC4yKTtwYWRkaW5nOjNweCA4cHg7d2hpdGUtc3BhY2U6bm93cmFwfQoubG9hZGluZ3t0ZXh0LWFsaWduOmNlbnRlcjtwYWRkaW5nOjQwcHggMDtjb2xvcjojNDQ0NDQwO2ZvbnQtc2l6ZTouNzVyZW07bGV0dGVyLXNwYWNpbmc6LjEyZW19Ci5ncm93dGgtdGFic3tkaXNwbGF5OmZsZXg7Z2FwOjhweDttYXJnaW4tYm90dG9tOjE0cHh9Ci5ndGFie2JhY2tncm91bmQ6bm9uZTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjAxLDE2OCw3NiwuMjUpO2NvbG9yOiM4YTg1Nzg7cGFkZGluZzo3cHggMTZweDtmb250LXNpemU6LjY1cmVtO2xldHRlci1zcGFjaW5nOi4xZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2N1cnNvcjpwb2ludGVyO2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmO3RyYW5zaXRpb246YWxsIC4yc30KLmd0YWIuYWN0aXZle2NvbG9yOiNjOWE4NGM7Ym9yZGVyLWNvbG9yOiNjOWE4NGM7YmFja2dyb3VuZDpyZ2JhKDIwMSwxNjgsNzYsLjA4KX0KLmdyb3d0aC1zdmd7d2lkdGg6MTAwJTtoZWlnaHQ6YXV0bztkaXNwbGF5OmJsb2NrfQouZ3Jvd3RoLWJhci1sYWJlbHtmb250LXNpemU6OHB4O2ZpbGw6IzhhODU3ODtmb250LWZhbWlseTonTW9udHNlcnJhdCcsc2Fucy1zZXJpZn0KLmdyb3d0aC1iYXItdmFse2ZvbnQtc2l6ZTo5cHg7ZmlsbDojYzlhODRjO2ZvbnQtZmFtaWx5OidNb250c2VycmF0JyxzYW5zLXNlcmlmO2ZvbnQtd2VpZ2h0OjYwMH0KLnNlY3R7bWFyZ2luLWJvdHRvbToyMnB4fQouZm9ybS1ibG9ja3tiYWNrZ3JvdW5kOiMxNDE0MTA7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSwyNTUsMjU1LC4wNyk7cGFkZGluZzoxNnB4O21hcmdpbi1ib3R0b206MTZweH0KLnRhZ3tmb250LXNpemU6LjU4cmVtO2NvbG9yOiM4YThhNTU7YmFja2dyb3VuZDpyZ2JhKDI1NSwyNTUsMjU1LC4wNSk7cGFkZGluZzoycHggOHB4O21hcmdpbi1yaWdodDo0cHg7bWFyZ2luLWJvdHRvbTo0cHg7ZGlzcGxheTppbmxpbmUtYmxvY2t9CkBrZXlmcmFtZXMgZnV7ZnJvbXtvcGFjaXR5OjA7dHJhbnNmb3JtOnRyYW5zbGF0ZVkoOHB4KX10b3tvcGFjaXR5OjE7dHJhbnNmb3JtOnRyYW5zbGF0ZVkoMCl9fQouYW5pbXthbmltYXRpb246ZnUgLjNzIGVhc2UgYm90aH0KPC9zdHlsZT4KPC9oZWFkPgo8Ym9keT4KPGRpdiBjbGFzcz0iY29uIj4KPGRpdiBjbGFzcz0icGFnZSI+Cgo8YSBocmVmPSIvYWRtaW4/cGFzcz1hbnphMTk4NSIgY2xhc3M9ImJhY2stbGluayI+4oaQINCQ0LTQvNC40L0t0L/QsNC90LXQu9GMPC9hPgo8ZGl2IGNsYXNzPSJsb2dvLXJqIj5SJmFtcDtKPC9kaXY+CjxkaXYgY2xhc3M9ImxvZ28tc3ViIj5Hcm9vbWluZyAmbWlkZG90OyDQodGC0LDRgtC40YHRgtC40LrQsCAmbWlkZG90OyDQotCw0LvQu9C40L08L2Rpdj4KCjxkaXYgY2xhc3M9InRhYnMtbWFpbiI+CiAgPGJ1dHRvbiBjbGFzcz0idG1iIGFjdGl2ZSIgb25jbGljaz0ic3dpdGNoTWFpbignc3RhdHMnLHRoaXMpIj7QodGC0LDRgtC40YHRgtC40LrQsDwvYnV0dG9uPgogIDxidXR0b24gY2xhc3M9InRtYiIgb25jbGljaz0ic3dpdGNoTWFpbignbWdtdCcsdGhpcykiPtCj0L/RgNCw0LLQu9C10L3QuNC1PC9idXR0b24+CjwvZGl2PgoKPCEtLSDilZDilZDilZAg0KHQotCQ0KLQmNCh0KLQmNCa0JAg4pWQ4pWQ4pWQIC0tPgo8ZGl2IGNsYXNzPSJwYW5lbCBhY3RpdmUiIGlkPSJwYW5lbC1zdGF0cyI+CgogIDxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDtnYXA6OHB4O21hcmdpbi1ib3R0b206MTJweDttYXJnaW4tdG9wOjIwcHg7YWxpZ24taXRlbXM6Y2VudGVyIj4KICAgIDxzZWxlY3QgY2xhc3M9InBlcmlvZC1zZWxlY3QiIGlkPSJwZXJpb2RTZWxlY3QiIG9uY2hhbmdlPSJsb2FkU3RhdHMoKSI+CiAgICAgIDxvcHRpb24gdmFsdWU9Im1vbnRoIj7QrdGC0L7RgiDQvNC10YHRj9GGPC9vcHRpb24+CiAgICAgIDxvcHRpb24gdmFsdWU9Imxhc3RfbW9udGgiPtCf0YDQvtGI0LvRi9C5INC80LXRgdGP0YY8L29wdGlvbj4KICAgICAgPG9wdGlvbiB2YWx1ZT0iM21vbnRocyI+MyDQvNC10YHRj9GG0LA8L29wdGlvbj4KICAgICAgPG9wdGlvbiB2YWx1ZT0iYWxsIj7QktGB0ZEg0LLRgNC10LzRjzwvb3B0aW9uPgogICAgPC9zZWxlY3Q+CiAgICA8YnV0dG9uIGNsYXNzPSJyZWZyZXNoLWJ0biIgb25jbGljaz0ibG9hZFN0YXRzKCkiPiYjODYzNTsg0J7QsdC90L7QstC40YLRjDwvYnV0dG9uPgogIDwvZGl2PgoKICA8ZGl2IGNsYXNzPSJkaXNjb3VudC1yb3ciPgogICAgPGRpdiBjbGFzcz0iZGlzY291bnQtbGFiZWwiPtCh0LrQuNC00LrQsCDRgdCw0LvQvtC90LAgKNCy0YvRh9C40YLQsNC10YLRgdGPINGC0L7Qu9GM0LrQviDQuNC3INC00L7Qu9C4INGB0LDQu9C+0L3QsCk8L2Rpdj4KICAgIDxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjRweCI+CiAgICAgIDxpbnB1dCBjbGFzcz0iZGlzY291bnQtaW5wdXQiIGlkPSJkaXNjb3VudElucHV0IiB0eXBlPSJudW1iZXIiIG1pbj0iMCIgdmFsdWU9IjAiIG9uaW5wdXQ9InJlY2FsYygpIj4KICAgICAgPHNwYW4gY2xhc3M9ImRpc2NvdW50LWV1ciI+4oKsPC9zcGFuPgogICAgPC9kaXY+CiAgPC9kaXY+CgogIDxkaXYgY2xhc3M9Im1ldHJpY3MiIGlkPSJtZXRyaWNzQmxvY2siPgogICAgPGRpdiBjbGFzcz0ibWV0cmljIj48ZGl2IGNsYXNzPSJtZXRyaWMtbGFiZWwiPtCS0YvRgNGD0YfQutCwPC9kaXY+PGRpdiBjbGFzcz0ibWV0cmljLXZhbCIgaWQ9Im1Ub3RhbCI+4oCUPC9kaXY+PGRpdiBjbGFzcz0ibWV0cmljLXN1YiI+0YHRg9C80LzQsCDRg9GB0LvRg9CzPC9kaXY+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJtZXRyaWMiPjxkaXYgY2xhc3M9Im1ldHJpYy1sYWJlbCI+0JzQsNGB0YLQtdGA0LDQvDwvZGl2PjxkaXYgY2xhc3M9Im1ldHJpYy12YWwiIGlkPSJtTWFzdGVycyI+4oCUPC9kaXY+PGRpdiBjbGFzcz0ibWV0cmljLXN1YiI+0L/QviDRgdGC0LDQstC60LUg0LzQsNGB0YLQtdGA0LA8L2Rpdj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9Im1ldHJpYyI+PGRpdiBjbGFzcz0ibWV0cmljLWxhYmVsIj7QodCw0LvQvtC90YM8L2Rpdj48ZGl2IGNsYXNzPSJtZXRyaWMtdmFsIiBpZD0ibVNhbG9uIj7igJQ8L2Rpdj48ZGl2IGNsYXNzPSJtZXRyaWMtc3ViIj7QvtGB0YLQsNGC0L7QuiDiiJIg0YHQutC40LTQutCwPC9kaXY+PC9kaXY+CiAgPC9kaXY+CiAgPGRpdiBjbGFzcz0ibWV0cmljcyIgc3R5bGU9Im1hcmdpbi1ib3R0b206MCI+CiAgICA8ZGl2IGNsYXNzPSJtZXRyaWMiPjxkaXYgY2xhc3M9Im1ldHJpYy1sYWJlbCI+0JfQsNC/0LjRgdC10Lk8L2Rpdj48ZGl2IGNsYXNzPSJtZXRyaWMtdmFsIiBpZD0ibUNvdW50Ij7igJQ8L2Rpdj48ZGl2IGNsYXNzPSJtZXRyaWMtc3ViIj7Qt9CwINC/0LXRgNC40L7QtDwvZGl2PjwvZGl2PgogICAgPGRpdiBjbGFzcz0ibWV0cmljIj48ZGl2IGNsYXNzPSJtZXRyaWMtbGFiZWwiPtCa0LvQuNC10L3RgtC+0LI8L2Rpdj48ZGl2IGNsYXNzPSJtZXRyaWMtdmFsIiBpZD0ibUNsaWVudHMiPuKAlDwvZGl2PjxkaXYgY2xhc3M9Im1ldHJpYy1zdWIiPtGD0L3QuNC60LDQu9GM0L3Ri9GFPC9kaXY+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJtZXRyaWMiPjxkaXYgY2xhc3M9Im1ldHJpYy1sYWJlbCI+0KHRgC4g0YfQtdC6PC9kaXY+PGRpdiBjbGFzcz0ibWV0cmljLXZhbCIgaWQ9Im1BdmciPuKAlDwvZGl2PjxkaXYgY2xhc3M9Im1ldHJpYy1zdWIiPtC30LAg0YPRgdC70YPQs9GDPC9kaXY+PC9kaXY+CiAgPC9kaXY+CiAgPGRpdiBjbGFzcz0ibWV0cmljcyIgc3R5bGU9Im1hcmdpbi1ib3R0b206MDtncmlkLXRlbXBsYXRlLWNvbHVtbnM6MWZyIj4KICAgIDxkaXYgY2xhc3M9Im1ldHJpYyI+PGRpdiBjbGFzcz0ibWV0cmljLWxhYmVsIj7QndC+0LLRi9GFINC60LvQuNC10L3RgtC+0LI8L2Rpdj48ZGl2IGNsYXNzPSJtZXRyaWMtdmFsIiBpZD0ibU5ld0NsaWVudHMiPuKAlDwvZGl2PjxkaXYgY2xhc3M9Im1ldHJpYy1zdWIiPtC/0LXRgNCy0YvQuSDQstC40LfQuNGCINCyINGN0YLQvtC8INC/0LXRgNC40L7QtNC1PC9kaXY+PC9kaXY+CiAgPC9kaXY+CgogIDxkaXYgc3R5bGU9Im1hcmdpbi10b3A6MjJweCI+CiAgICA8ZGl2IGNsYXNzPSJzbGJsIj7Qn9GA0LjRgNC+0YHRgiDQutC70LjQtdC90YLQvtCyPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJncm93dGgtdGFicyI+CiAgICAgIDxidXR0b24gY2xhc3M9Imd0YWIgYWN0aXZlIiBpZD0iZ3RhYkRheSIgb25jbGljaz0ic3dpdGNoR3Jvd3RoVmlldygnZGF5JykiPtCf0L4g0LTQvdGP0Lw8L2J1dHRvbj4KICAgICAgPGJ1dHRvbiBjbGFzcz0iZ3RhYiIgaWQ9Imd0YWJNb250aCIgb25jbGljaz0ic3dpdGNoR3Jvd3RoVmlldygnbW9udGgnKSI+0J/QviDQvNC10YHRj9GG0LDQvDwvYnV0dG9uPgogICAgPC9kaXY+CiAgICA8ZGl2IGlkPSJncm93dGhDaGFydCI+PGRpdiBjbGFzcz0ibG9hZGluZyI+0JfQsNCz0YDRg9C30LrQsCDQtNCw0L3QvdGL0YUuLi48L2Rpdj48L2Rpdj4KICA8L2Rpdj4KCiAgPGRpdiBzdHlsZT0ibWFyZ2luLXRvcDoyMnB4Ij4KICAgIDxkaXYgY2xhc3M9InNsYmwiPtCf0L4g0LzQsNGB0YLQtdGA0LDQvDwvZGl2PgogICAgPGRpdiBpZD0ibWFzdGVyc0xpc3QiPjxkaXYgY2xhc3M9ImxvYWRpbmciPtCX0LDQs9GA0YPQt9C60LAg0LTQsNC90L3Ri9GFLi4uPC9kaXY+PC9kaXY+CiAgPC9kaXY+CjwvZGl2PgoKPCEtLSDilZDilZDilZAg0KPQn9Cg0JDQktCb0JXQndCY0JUg4pWQ4pWQ4pWQIC0tPgo8ZGl2IGNsYXNzPSJwYW5lbCIgaWQ9InBhbmVsLW1nbXQiPgogIDxkaXYgc3R5bGU9Im1hcmdpbi10b3A6MjBweCIgY2xhc3M9InN1Yi10YWJzIj4KICAgIDxidXR0b24gY2xhc3M9InN0YiBhY3RpdmUiIG9uY2xpY2s9InN3aXRjaFN1YignbWFzdGVycycsdGhpcykiPtCc0LDRgdGC0LXRgNCwPC9idXR0b24+CiAgICA8YnV0dG9uIGNsYXNzPSJzdGIiIG9uY2xpY2s9InN3aXRjaFN1YignYnJlZWRzJyx0aGlzKSI+0J/QvtGA0L7QtNGLPC9idXR0b24+CiAgICA8YnV0dG9uIGNsYXNzPSJzdGIiIG9uY2xpY2s9InN3aXRjaFN1YignY2xpZW50cycsdGhpcykiPtCa0LvQuNC10L3RgtGLPC9idXR0b24+CiAgPC9kaXY+CgogIDwhLS0g0JzQsNGB0YLQtdGA0LAgLS0+CiAgPGRpdiBjbGFzcz0ic3ViLXBhbmVsIGFjdGl2ZSIgaWQ9InN1Yi1tYXN0ZXJzIj4KICAgIDxkaXYgY2xhc3M9InNlY3QiPgogICAgICA8ZGl2IGNsYXNzPSJzbGJsIj7QlNC+0LHQsNCy0LjRgtGMINC80LDRgdGC0LXRgNCwPC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImZvcm0tYmxvY2siPgogICAgICAgIDxkaXYgY2xhc3M9ImZpLXJvdyI+CiAgICAgICAgICA8aW5wdXQgY2xhc3M9ImZpIiBpZD0ibmV3TWFzdGVyTmFtZSIgcGxhY2Vob2xkZXI9ItCY0LzRjyDQvNCw0YHRgtC10YDQsCI+CiAgICAgICAgICA8aW5wdXQgY2xhc3M9ImZpIiBpZD0ibmV3TWFzdGVyUGhvbmUiIHBsYWNlaG9sZGVyPSLQotC10LvQtdGE0L7QvSAo0L3QtdC+0LHRj9C3LikiPgogICAgICAgIDwvZGl2PgogICAgICAgIDxidXR0b24gY2xhc3M9ImNidG4iIG9uY2xpY2s9ImFkZE1hc3RlcigpIj4rINCU0L7QsdCw0LLQuNGC0Yw8L2J1dHRvbj4KICAgICAgPC9kaXY+CiAgICA8L2Rpdj4KICAgIDxkaXYgY2xhc3M9InNlY3QiPgogICAgICA8ZGl2IGNsYXNzPSJzbGJsIj7QnNCw0YHRgtC10YDQsCDRgdCw0LvQvtC90LA8L2Rpdj4KICAgICAgPGRpdiBpZD0ibWFzdGVyTGlzdFVJIj48L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2PgoKICA8IS0tINCf0L7RgNC+0LTRiyDQuCDRg9GB0LvRg9Cz0LggLS0+CiAgPGRpdiBjbGFzcz0ic3ViLXBhbmVsIiBpZD0ic3ViLWJyZWVkcyI+CiAgICA8ZGl2IGNsYXNzPSJzZWN0Ij4KICAgICAgPGRpdiBjbGFzcz0ic2xibCI+0JTQvtCx0LDQstC40YLRjCDQv9C+0YDQvtC00YM8L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iZm9ybS1ibG9jayI+CiAgICAgICAgPGlucHV0IGNsYXNzPSJmaSIgaWQ9Im5ld0JyZWVkTmFtZSIgcGxhY2Vob2xkZXI9ItCd0LDQt9Cy0LDQvdC40LUg0L/QvtGA0L7QtNGLICjQvdCw0L/RgC4g0KXQsNGB0LrQuCAyMOKAkzMwINC60LMpIj4KICAgICAgICA8YnV0dG9uIGNsYXNzPSJjYnRuIiBvbmNsaWNrPSJhZGRCcmVlZCgpIj4rINCU0L7QsdCw0LLQuNGC0Ywg0L/QvtGA0L7QtNGDPC9idXR0b24+CiAgICAgIDwvZGl2PgogICAgPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzZWN0Ij4KICAgICAgPGRpdiBjbGFzcz0ic2xibCI+0J/QvtGA0L7QtNGLINC4INGD0YHQu9GD0LPQuDwvZGl2PgogICAgICA8ZGl2IGlkPSJicmVlZExpc3RVSSI+PC9kaXY+CiAgICA8L2Rpdj4KICA8L2Rpdj4KCiAgPCEtLSDQmtC70LjQtdC90YLRiyAtLT4KICA8ZGl2IGNsYXNzPSJzdWItcGFuZWwiIGlkPSJzdWItY2xpZW50cyI+CiAgICA8ZGl2IGNsYXNzPSJzZWN0Ij4KICAgICAgPGRpdiBjbGFzcz0ic2xibCI+0J3QvtCy0LDRjyDQutCw0YDRgtC+0YfQutCwINC60LvQuNC10L3RgtCwPC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImZvcm0tYmxvY2siPgogICAgICAgIDxkaXYgY2xhc3M9ImZpLXJvdyI+CiAgICAgICAgICA8aW5wdXQgY2xhc3M9ImZpIiBpZD0ibmV3Q2xpZW50TmFtZSIgcGxhY2Vob2xkZXI9ItCY0LzRjyDQutC70LjQtdC90YLQsCI+CiAgICAgICAgICA8aW5wdXQgY2xhc3M9ImZpIiBpZD0ibmV3Q2xpZW50UGhvbmUiIHBsYWNlaG9sZGVyPSIrMzcyIC4uLiI+CiAgICAgICAgPC9kaXY+CiAgICAgICAgPGRpdiBjbGFzcz0iZmktcm93Ij4KICAgICAgICAgIDxpbnB1dCBjbGFzcz0iZmkiIGlkPSJuZXdDbGllbnRQZXQiIHBsYWNlaG9sZGVyPSLQmtC70LjRh9C60LAg0L/QuNGC0L7QvNGG0LAiPgogICAgICAgICAgPHNlbGVjdCBjbGFzcz0iZmkiIGlkPSJuZXdDbGllbnRCcmVlZCI+PG9wdGlvbiB2YWx1ZT0iIj7Qn9C+0YDQvtC00LAuLi48L29wdGlvbj48L3NlbGVjdD4KICAgICAgICA8L2Rpdj4KICAgICAgICA8aW5wdXQgY2xhc3M9ImZpIiBpZD0ibmV3Q2xpZW50Tm90ZSIgcGxhY2Vob2xkZXI9ItCa0L7QvNC80LXQvdGC0LDRgNC40LkgKNCw0LvQu9C10YDQs9C40LgsINC+0YHQvtCx0LXQvdC90L7RgdGC0LguLi4pIj4KICAgICAgICA8YnV0dG9uIGNsYXNzPSJjYnRuIiBvbmNsaWNrPSJhZGRDbGllbnQoKSI+KyDQodC+0LfQtNCw0YLRjCDQutCw0YDRgtC+0YfQutGDPC9idXR0b24+CiAgICAgIDwvZGl2PgogICAgPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzZWN0Ij4KICAgICAgPGRpdiBjbGFzcz0ic2xibCI+0JHQsNC30LAg0LrQu9C40LXQvdGC0L7QsjwvZGl2PgogICAgICA8ZGl2IGlkPSJjbGllbnRMaXN0VUkiPjwvZGl2PgogICAgPC9kaXY+CiAgPC9kaXY+CjwvZGl2PgoKPC9kaXY+CjwvZGl2PgoKPHNjcmlwdD4KdmFyIEdPT0dMRV9TQ1JJUFQgPSB3aW5kb3cubG9jYXRpb24ub3JpZ2luICsgJy9hcGkvc3RhdHMnOwp2YXIgZGIgPSB7CiAgbWFzdGVyczogSlNPTi5wYXJzZShsb2NhbFN0b3JhZ2UuZ2V0SXRlbSgncmpfbWFzdGVycycpIHx8ICdbItCi0LDRgtGM0Y/QvdCwIiwi0JDQu9C40YHQsCIsItCa0YDQuNGB0YLQuNC90LAiLCLQkNC90L3QsCJdJyksCiAgYnJlZWRzOiAgSlNPTi5wYXJzZShsb2NhbFN0b3JhZ2UuZ2V0SXRlbSgncmpfYnJlZWRzJykgIHx8ICdbXScpLAogIGNsaWVudHM6IEpTT04ucGFyc2UobG9jYWxTdG9yYWdlLmdldEl0ZW0oJ3JqX2NsaWVudHMnKSB8fCAnW10nKSwKICBkaXNjb3VudDogcGFyc2VGbG9hdChsb2NhbFN0b3JhZ2UuZ2V0SXRlbSgncmpfZGlzY291bnQnKSB8fCAnMCcpCn07CnZhciBhbGxCb29raW5ncyA9IFtdOwp2YXIgc3RhdHNMb2FkZWQgPSBmYWxzZTsKCmZ1bmN0aW9uIHNhdmUoKSB7CiAgbG9jYWxTdG9yYWdlLnNldEl0ZW0oJ3JqX21hc3RlcnMnLCAgSlNPTi5zdHJpbmdpZnkoZGIubWFzdGVycykpOwogIGxvY2FsU3RvcmFnZS5zZXRJdGVtKCdyal9icmVlZHMnLCAgIEpTT04uc3RyaW5naWZ5KGRiLmJyZWVkcykpOwogIGxvY2FsU3RvcmFnZS5zZXRJdGVtKCdyal9jbGllbnRzJywgIEpTT04uc3RyaW5naWZ5KGRiLmNsaWVudHMpKTsKICBsb2NhbFN0b3JhZ2Uuc2V0SXRlbSgncmpfZGlzY291bnQnLCBkYi5kaXNjb3VudCk7Cn0KCi8vIOKUgOKUgCDQndCQ0JLQmNCT0JDQptCY0K8g4pSA4pSACmZ1bmN0aW9uIHN3aXRjaE1haW4obmFtZSwgYnRuKSB7CiAgZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLnRtYicpLmZvckVhY2goZnVuY3Rpb24odCl7dC5jbGFzc0xpc3QucmVtb3ZlKCdhY3RpdmUnKX0pOwogIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5wYW5lbCcpLmZvckVhY2goZnVuY3Rpb24ocCl7cC5jbGFzc0xpc3QucmVtb3ZlKCdhY3RpdmUnKX0pOwogIGJ0bi5jbGFzc0xpc3QuYWRkKCdhY3RpdmUnKTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgncGFuZWwtJytuYW1lKS5jbGFzc0xpc3QuYWRkKCdhY3RpdmUnKTsKICBpZiAobmFtZSA9PT0gJ3N0YXRzJyAmJiAhc3RhdHNMb2FkZWQpIGxvYWRTdGF0cygpOwp9CmZ1bmN0aW9uIHN3aXRjaFN1YihuYW1lLCBidG4pIHsKICBkb2N1bWVudC5xdWVyeVNlbGVjdG9yQWxsKCcuc3RiJykuZm9yRWFjaChmdW5jdGlvbih0KXt0LmNsYXNzTGlzdC5yZW1vdmUoJ2FjdGl2ZScpfSk7CiAgZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLnN1Yi1wYW5lbCcpLmZvckVhY2goZnVuY3Rpb24ocCl7cC5jbGFzc0xpc3QucmVtb3ZlKCdhY3RpdmUnKX0pOwogIGJ0bi5jbGFzc0xpc3QuYWRkKCdhY3RpdmUnKTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnc3ViLScrbmFtZSkuY2xhc3NMaXN0LmFkZCgnYWN0aXZlJyk7Cn0KCi8vIOKUgOKUgCDQodCi0JDQotCY0KHQotCY0JrQkCDilIDilIAKZnVuY3Rpb24gZ2V0UGVyaW9kUGFyYW1zKCkgewogIHZhciB2YWwgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgncGVyaW9kU2VsZWN0JykudmFsdWU7CiAgdmFyIG5vdyA9IG5ldyBEYXRlKCk7CiAgdmFyIGZyb20sIHRvOwogIGlmICh2YWwgPT09ICdtb250aCcpIHsKICAgIGZyb20gPSBuZXcgRGF0ZShub3cuZ2V0RnVsbFllYXIoKSwgbm93LmdldE1vbnRoKCksIDEpOwogICAgdG8gICA9IG5ldyBEYXRlKG5vdy5nZXRGdWxsWWVhcigpLCBub3cuZ2V0TW9udGgoKSsxLCAwKTsKICB9IGVsc2UgaWYgKHZhbCA9PT0gJ2xhc3RfbW9udGgnKSB7CiAgICBmcm9tID0gbmV3IERhdGUobm93LmdldEZ1bGxZZWFyKCksIG5vdy5nZXRNb250aCgpLTEsIDEpOwogICAgdG8gICA9IG5ldyBEYXRlKG5vdy5nZXRGdWxsWWVhcigpLCBub3cuZ2V0TW9udGgoKSwgMCk7CiAgfSBlbHNlIGlmICh2YWwgPT09ICczbW9udGhzJykgewogICAgZnJvbSA9IG5ldyBEYXRlKG5vdy5nZXRGdWxsWWVhcigpLCBub3cuZ2V0TW9udGgoKS0yLCAxKTsKICAgIHRvICAgPSBuZXcgRGF0ZShub3cuZ2V0RnVsbFllYXIoKSwgbm93LmdldE1vbnRoKCkrMSwgMCk7CiAgfSBlbHNlIHsKICAgIGZyb20gPSBuZXcgRGF0ZSgyMDI0LCAwLCAxKTsKICAgIHRvICAgPSBuZXcgRGF0ZShub3cuZ2V0RnVsbFllYXIoKSwgbm93LmdldE1vbnRoKCkrMSwgMCk7CiAgfQogIHZhciBmbXQgPSBmdW5jdGlvbihkKSB7IHJldHVybiBTdHJpbmcoZC5nZXREYXRlKCkpLnBhZFN0YXJ0KDIsJzAnKSsnLicrU3RyaW5nKGQuZ2V0TW9udGgoKSsxKS5wYWRTdGFydCgyLCcwJykrJy4nK2QuZ2V0RnVsbFllYXIoKTsgfTsKICByZXR1cm4ge2Zyb206IGZtdChmcm9tKSwgdG86IGZtdCh0byl9Owp9Cgp2YXIgYWxsVGltZUJvb2tpbmdzID0gW107CnZhciBncm93dGhWaWV3ID0gJ2RheSc7CgpmdW5jdGlvbiBsb2FkU3RhdHMoKSB7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ21hc3RlcnNMaXN0JykuaW5uZXJIVE1MID0gJzxkaXYgY2xhc3M9ImxvYWRpbmciPtCX0LDQs9GA0YPQt9C60LAg0LTQsNC90L3Ri9GFINC40Lcg0LrQsNC70LXQvdC00LDRgNGPLi4uPC9kaXY+JzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZ3Jvd3RoQ2hhcnQnKS5pbm5lckhUTUwgPSAnPGRpdiBjbGFzcz0ibG9hZGluZyI+0JfQsNCz0YDRg9C30LrQsCDQtNCw0L3QvdGL0YUuLi48L2Rpdj4nOwogIHZhciBwID0gZ2V0UGVyaW9kUGFyYW1zKCk7CiAgUHJvbWlzZS5hbGwoWwogICAgZmV0Y2god2luZG93LmxvY2F0aW9uLm9yaWdpbiArICcvYXBpL3N0YXRzP2Zyb209JyArIHAuZnJvbSArICcmdG89JyArIHAudG8pLnRoZW4oZnVuY3Rpb24ocil7cmV0dXJuIHIuanNvbigpO30pLAogICAgZmV0Y2god2luZG93LmxvY2F0aW9uLm9yaWdpbiArICcvYXBpL3N0YXRzP2Zyb209MDEuMDEuMjAyMCZ0bz0nICsgcC50bykudGhlbihmdW5jdGlvbihyKXtyZXR1cm4gci5qc29uKCk7fSkKICBdKS50aGVuKGZ1bmN0aW9uKHJlc3VsdHMpewogICAgdmFyIHBlcmlvZERhdGEgPSByZXN1bHRzWzBdLCBhbGxEYXRhID0gcmVzdWx0c1sxXTsKICAgIGlmIChwZXJpb2REYXRhLnN1Y2Nlc3MpIHsKICAgICAgYWxsQm9va2luZ3MgPSBwZXJpb2REYXRhLmJvb2tpbmdzIHx8IFtdOwogICAgICBzdGF0c0xvYWRlZCA9IHRydWU7CiAgICB9IGVsc2UgewogICAgICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbWFzdGVyc0xpc3QnKS5pbm5lckhUTUwgPSAnPGRpdiBjbGFzcz0ibG9hZGluZyI+0J7RiNC40LHQutCwOiAnICsgKHBlcmlvZERhdGEuZXJyb3J8fCfQvdC10YIg0LTQsNC90L3Ri9GFJykgKyAnPC9kaXY+JzsKICAgIH0KICAgIGlmIChhbGxEYXRhLnN1Y2Nlc3MpIHsKICAgICAgYWxsVGltZUJvb2tpbmdzID0gYWxsRGF0YS5ib29raW5ncyB8fCBbXTsKICAgIH0KICAgIHJlY2FsYygpOwogICAgcmVuZGVyR3Jvd3RoQ2hhcnQoKTsKICB9KS5jYXRjaChmdW5jdGlvbihlKXsKICAgIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdtYXN0ZXJzTGlzdCcpLmlubmVySFRNTCA9ICc8ZGl2IGNsYXNzPSJsb2FkaW5nIj7QndC10YIg0YHQvtC10LTQuNC90LXQvdC40Y8g0YEg0LrQsNC70LXQvdC00LDRgNGR0Lw8L2Rpdj4nOwogICAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2dyb3d0aENoYXJ0JykuaW5uZXJIVE1MID0gJzxkaXYgY2xhc3M9ImxvYWRpbmciPtCd0LXRgiDRgdC+0LXQtNC40L3QtdC90LjRjyDRgSDQutCw0LvQtdC90LTQsNGA0ZHQvDwvZGl2Pic7CiAgfSk7Cn0KCmZ1bmN0aW9uIHBhcnNlRE1ZKHMpewogIGlmKCFzKSByZXR1cm4gbnVsbDsKICB2YXIgcGFydHMgPSBzLnNwbGl0KCcuJyk7CiAgaWYocGFydHMubGVuZ3RoIT09MykgcmV0dXJuIG51bGw7CiAgdmFyIGQgPSBuZXcgRGF0ZShwYXJzZUludChwYXJ0c1syXSwxMCksIHBhcnNlSW50KHBhcnRzWzFdLDEwKS0xLCBwYXJzZUludChwYXJ0c1swXSwxMCkpOwogIHJldHVybiBpc05hTihkLmdldFRpbWUoKSkgPyBudWxsIDogZDsKfQoKZnVuY3Rpb24gZ2V0Q2xpZW50Rmlyc3RWaXNpdE1hcCgpewogIHZhciBtYXAgPSB7fTsKICBhbGxUaW1lQm9va2luZ3MuZm9yRWFjaChmdW5jdGlvbihiKXsKICAgIHZhciBrZXkgPSBiLmNsaWVudFBob25lIHx8IGIuY2xpZW50TmFtZTsKICAgIGlmKCFrZXkpIHJldHVybjsKICAgIHZhciBkID0gcGFyc2VETVkoYi5kYXRlKTsKICAgIGlmKCFkKSByZXR1cm47CiAgICBpZighbWFwW2tleV0gfHwgZCA8IG1hcFtrZXldKSBtYXBba2V5XSA9IGQ7CiAgfSk7CiAgcmV0dXJuIG1hcDsKfQoKZnVuY3Rpb24gc3dpdGNoR3Jvd3RoVmlldyh2aWV3KXsKICBncm93dGhWaWV3ID0gdmlldzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZ3RhYkRheScpLmNsYXNzTGlzdC50b2dnbGUoJ2FjdGl2ZScsIHZpZXc9PT0nZGF5Jyk7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2d0YWJNb250aCcpLmNsYXNzTGlzdC50b2dnbGUoJ2FjdGl2ZScsIHZpZXc9PT0nbW9udGgnKTsKICByZW5kZXJHcm93dGhDaGFydCgpOwp9CgpmdW5jdGlvbiByZW5kZXJHcm93dGhDaGFydCgpewogIHZhciBjb250YWluZXIgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZ3Jvd3RoQ2hhcnQnKTsKICBpZighYWxsVGltZUJvb2tpbmdzLmxlbmd0aCl7IGNvbnRhaW5lci5pbm5lckhUTUwgPSAnPGRpdiBjbGFzcz0ibG9hZGluZyI+0J3QtdGCINC00LDQvdC90YvRhTwvZGl2Pic7IHJldHVybjsgfQogIHZhciBmaXJzdFZpc2l0cyA9IGdldENsaWVudEZpcnN0VmlzaXRNYXAoKTsKICB2YXIgbGFiZWxzID0gW107CgogIGlmKGdyb3d0aFZpZXcgPT09ICdkYXknKXsKICAgIHZhciBwID0gZ2V0UGVyaW9kUGFyYW1zKCk7CiAgICB2YXIgZnJvbUQgPSBwYXJzZURNWShwLmZyb20pLCB0b0QgPSBwYXJzZURNWShwLnRvKTsKICAgIHZhciBidWNrZXRzID0ge307CiAgICBPYmplY3Qua2V5cyhmaXJzdFZpc2l0cykuZm9yRWFjaChmdW5jdGlvbihrKXsKICAgICAgdmFyIGQgPSBmaXJzdFZpc2l0c1trXTsKICAgICAgaWYoZCA8IGZyb21EIHx8IGQgPiB0b0QpIHJldHVybjsKICAgICAgdmFyIGtleSA9IGQuZ2V0RnVsbFllYXIoKSsnLScrKGQuZ2V0TW9udGgoKSsxKSsnLScrZC5nZXREYXRlKCk7CiAgICAgIGJ1Y2tldHNba2V5XSA9IChidWNrZXRzW2tleV18fDApKzE7CiAgICB9KTsKICAgIHZhciBjdXIgPSBuZXcgRGF0ZShmcm9tRCk7CiAgICB3aGlsZShjdXIgPD0gdG9EKXsKICAgICAgdmFyIGtleSA9IGN1ci5nZXRGdWxsWWVhcigpKyctJysoY3VyLmdldE1vbnRoKCkrMSkrJy0nK2N1ci5nZXREYXRlKCk7CiAgICAgIGxhYmVscy5wdXNoKHtrZXk6IFN0cmluZyhjdXIuZ2V0RGF0ZSgpKS5wYWRTdGFydCgyLCcwJykrJy4nK1N0cmluZyhjdXIuZ2V0TW9udGgoKSsxKS5wYWRTdGFydCgyLCcwJyksIHZhbDogYnVja2V0c1trZXldfHwwfSk7CiAgICAgIGN1ci5zZXREYXRlKGN1ci5nZXREYXRlKCkrMSk7CiAgICB9CiAgfSBlbHNlIHsKICAgIHZhciBtb250aE5hbWVzID0gWyfQr9C90LInLCfQpNC10LInLCfQnNCw0YAnLCfQkNC/0YAnLCfQnNCw0LknLCfQmNGO0L0nLCfQmNGO0LsnLCfQkNCy0LMnLCfQodC10L0nLCfQntC60YInLCfQndC+0Y8nLCfQlNC10LonXTsKICAgIHZhciBidWNrZXRzID0ge307CiAgICBPYmplY3Qua2V5cyhmaXJzdFZpc2l0cykuZm9yRWFjaChmdW5jdGlvbihrKXsKICAgICAgdmFyIGQgPSBmaXJzdFZpc2l0c1trXTsKICAgICAgdmFyIGtleSA9IGQuZ2V0RnVsbFllYXIoKSsnLScrU3RyaW5nKGQuZ2V0TW9udGgoKSsxKS5wYWRTdGFydCgyLCcwJyk7CiAgICAgIGJ1Y2tldHNba2V5XSA9IChidWNrZXRzW2tleV18fDApKzE7CiAgICB9KTsKICAgIHZhciBrZXlzID0gT2JqZWN0LmtleXMoYnVja2V0cykuc29ydCgpOwogICAgbGFiZWxzID0ga2V5cy5tYXAoZnVuY3Rpb24oayl7CiAgICAgIHZhciBwYXJ0cyA9IGsuc3BsaXQoJy0nKTsKICAgICAgcmV0dXJuIHtrZXk6IG1vbnRoTmFtZXNbcGFyc2VJbnQocGFydHNbMV0sMTApLTFdICsgJyAnICsgcGFydHNbMF0uc2xpY2UoMiksIHZhbDogYnVja2V0c1trXX07CiAgICB9KTsKICB9CiAgcmVuZGVyR3Jvd3RoU1ZHKGNvbnRhaW5lciwgbGFiZWxzKTsKfQoKZnVuY3Rpb24gcmVuZGVyR3Jvd3RoU1ZHKGNvbnRhaW5lciwgbGFiZWxzKXsKICBpZighbGFiZWxzLmxlbmd0aCl7IGNvbnRhaW5lci5pbm5lckhUTUwgPSAnPGRpdiBjbGFzcz0ibG9hZGluZyI+0J3QtdGCINC00LDQvdC90YvRhTwvZGl2Pic7IHJldHVybjsgfQogIHZhciBtYXhWYWwgPSBNYXRoLm1heC5hcHBseShudWxsLCBsYWJlbHMubWFwKGZ1bmN0aW9uKGwpe3JldHVybiBsLnZhbDt9KS5jb25jYXQoWzFdKSk7CiAgdmFyIHcgPSAzNDAsIGggPSAxNDA7CiAgdmFyIGJhckdhcCA9IDM7CiAgdmFyIGJhclcgPSBNYXRoLm1heCgyLCAodyAtIChsYWJlbHMubGVuZ3RoLTEpKmJhckdhcCkgLyBsYWJlbHMubGVuZ3RoKTsKICB2YXIgc3ZnID0gJzxzdmcgY2xhc3M9Imdyb3d0aC1zdmciIHZpZXdCb3g9IjAgMCAnICsgdyArICcgJyArIChoKzIwKSArICciIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+JzsKICBsYWJlbHMuZm9yRWFjaChmdW5jdGlvbihsLCBpKXsKICAgIHZhciBiYXJIID0gbWF4VmFsID4gMCA/IChsLnZhbCAvIG1heFZhbCkgKiBoIDogMDsKICAgIHZhciB4ID0gaSAqIChiYXJXICsgYmFyR2FwKTsKICAgIHZhciB5ID0gaCAtIGJhckg7CiAgICBzdmcgKz0gJzxyZWN0IHg9IicreC50b0ZpeGVkKDEpKyciIHk9IicreS50b0ZpeGVkKDEpKyciIHdpZHRoPSInK2JhclcudG9GaXhlZCgxKSsnIiBoZWlnaHQ9IicrYmFySC50b0ZpeGVkKDEpKyciIGZpbGw9InJnYmEoMjAxLDE2OCw3NiwnKyhsLnZhbD4wPzAuNzU6MC4xMikrJykiPjwvcmVjdD4nOwogICAgaWYobGFiZWxzLmxlbmd0aCA8PSAyMCAmJiBsLnZhbCA+IDApewogICAgICBzdmcgKz0gJzx0ZXh0IHg9IicrKHgrYmFyVy8yKS50b0ZpeGVkKDEpKyciIHk9IicrTWF0aC5tYXgoOSx5LTQpLnRvRml4ZWQoMSkrJyIgdGV4dC1hbmNob3I9Im1pZGRsZSIgY2xhc3M9Imdyb3d0aC1iYXItdmFsIj4nK2wudmFsKyc8L3RleHQ+JzsKICAgIH0KICB9KTsKICB2YXIgbGFiZWxTdGVwID0gTWF0aC5tYXgoMSwgTWF0aC5jZWlsKGxhYmVscy5sZW5ndGggLyA4KSk7CiAgbGFiZWxzLmZvckVhY2goZnVuY3Rpb24obCwgaSl7CiAgICBpZihpICUgbGFiZWxTdGVwICE9PSAwICYmIGkgIT09IGxhYmVscy5sZW5ndGgtMSkgcmV0dXJuOwogICAgdmFyIHggPSBpICogKGJhclcgKyBiYXJHYXApICsgYmFyVy8yOwogICAgc3ZnICs9ICc8dGV4dCB4PSInK3gudG9GaXhlZCgxKSsnIiB5PSInKyhoKzE0KSsnIiB0ZXh0LWFuY2hvcj0ibWlkZGxlIiBjbGFzcz0iZ3Jvd3RoLWJhci1sYWJlbCI+JytsLmtleSsnPC90ZXh0Pic7CiAgfSk7CiAgc3ZnICs9ICc8L3N2Zz4nOwogIGNvbnRhaW5lci5pbm5lckhUTUwgPSBzdmc7Cn0KCmZ1bmN0aW9uIGdldE1hc3RlclJhdGlvKG5hbWUpIHsKICB2YXIgbWFwID0gewogICAgJ9CQ0LvQuNGB0LAnOiAwLjQwLAogICAgJ9Ca0YHQtdC90LjRjyc6IDAuNDUsCiAgICAn0JDQvdC90LAnOiAxLjAwLAogICAgJ9CQ0LvQtdC60YHQsNC90LTRgNCwJzogMC41MCwKICAgICfQotCw0YLRjNGP0L3QsCc6IDAuNTAsCiAgICAn0JrRgNC40YHRgtC40L3QsCc6IDAuNTAKICB9OwogIHJldHVybiBtYXAuaGFzT3duUHJvcGVydHkobmFtZSkgPyBtYXBbbmFtZV0gOiAwLjUwOwp9CgpmdW5jdGlvbiByZWNhbGMoKSB7CiAgdmFyIGRpc2NvdW50ID0gcGFyc2VGbG9hdChkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZGlzY291bnRJbnB1dCcpLnZhbHVlKSB8fCAwOwogIGRiLmRpc2NvdW50ID0gZGlzY291bnQ7CiAgc2F2ZSgpOwoKICB2YXIgdG90YWwgPSAwLCBjb3VudCA9IDA7CiAgdmFyIG1hc3Rlck1hcCA9IHt9OwogIHZhciBwaG9uZXMgPSB7fTsKCiAgYWxsQm9va2luZ3MuZm9yRWFjaChmdW5jdGlvbihiKSB7CiAgICB2YXIgcHJpY2UgPSBwYXJzZUZsb2F0KGIucHJpY2UpIHx8IDA7CiAgICB0b3RhbCArPSBwcmljZTsKICAgIGNvdW50Kys7CiAgICBpZiAoYi5jbGllbnRQaG9uZSkgcGhvbmVzW2IuY2xpZW50UGhvbmVdID0gdHJ1ZTsKICAgIHZhciBtID0gYi5tYXN0ZXI7CiAgICBpZiAoIW1hc3Rlck1hcFttXSkgbWFzdGVyTWFwW21dID0ge2Jvb2tpbmdzOltdLCB0b3RhbDowfTsKICAgIG1hc3Rlck1hcFttXS5ib29raW5ncy5wdXNoKGIpOwogICAgbWFzdGVyTWFwW21dLnRvdGFsICs9IHByaWNlOwogIH0pOwoKICB2YXIgbWFzdGVyVG90YWwgPSAwLCBzYWxvbkJlZm9yZURpc2NvdW50ID0gMDsKICBPYmplY3Qua2V5cyhtYXN0ZXJNYXApLmZvckVhY2goZnVuY3Rpb24obSkgewogICAgdmFyIHJhdGlvID0gZ2V0TWFzdGVyUmF0aW8obSk7CiAgICBtYXN0ZXJUb3RhbCArPSBtYXN0ZXJNYXBbbV0udG90YWwgKiByYXRpbzsKICAgIHNhbG9uQmVmb3JlRGlzY291bnQgKz0gbWFzdGVyTWFwW21dLnRvdGFsICogKDEgLSByYXRpbyk7CiAgfSk7CiAgbWFzdGVyVG90YWwgPSBNYXRoLnJvdW5kKG1hc3RlclRvdGFsKTsKICB2YXIgc2Fsb25Ub3RhbCA9IE1hdGgucm91bmQoc2Fsb25CZWZvcmVEaXNjb3VudCAtIGRpc2NvdW50KTsKICB2YXIgYXZnID0gY291bnQgPiAwID8gTWF0aC5yb3VuZCh0b3RhbCAvIGNvdW50KSA6IDA7CiAgdmFyIHVuaXF1ZUNsaWVudHMgPSBPYmplY3Qua2V5cyhwaG9uZXMpLmxlbmd0aDsKCiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ21Ub3RhbCcpLnRleHRDb250ZW50ICAgPSB0b3RhbCArICcg4oKsJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbU1hc3RlcnMnKS50ZXh0Q29udGVudCA9IG1hc3RlclRvdGFsICsgJyDigqwnOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdtU2Fsb24nKS50ZXh0Q29udGVudCAgID0gc2Fsb25Ub3RhbCArICcg4oKsJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbUNvdW50JykudGV4dENvbnRlbnQgICA9IGNvdW50OwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdtQ2xpZW50cycpLnRleHRDb250ZW50ID0gdW5pcXVlQ2xpZW50czsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbUF2ZycpLnRleHRDb250ZW50ICAgICA9IGF2ZyArICcg4oKsJzsKCiAgdmFyIG5ld0NsaWVudHNFbCA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdtTmV3Q2xpZW50cycpOwogIGlmIChuZXdDbGllbnRzRWwpIHsKICAgIGlmIChhbGxUaW1lQm9va2luZ3MubGVuZ3RoKSB7CiAgICAgIHZhciBwID0gZ2V0UGVyaW9kUGFyYW1zKCk7CiAgICAgIHZhciBmcm9tRCA9IHBhcnNlRE1ZKHAuZnJvbSksIHRvRCA9IHBhcnNlRE1ZKHAudG8pOwogICAgICB2YXIgZmlyc3RWaXNpdHMgPSBnZXRDbGllbnRGaXJzdFZpc2l0TWFwKCk7CiAgICAgIHZhciBuZXdDb3VudCA9IDA7CiAgICAgIE9iamVjdC5rZXlzKGZpcnN0VmlzaXRzKS5mb3JFYWNoKGZ1bmN0aW9uKGspewogICAgICAgIHZhciBkID0gZmlyc3RWaXNpdHNba107CiAgICAgICAgaWYgKGQgPj0gZnJvbUQgJiYgZCA8PSB0b0QpIG5ld0NvdW50Kys7CiAgICAgIH0pOwogICAgICBuZXdDbGllbnRzRWwudGV4dENvbnRlbnQgPSBuZXdDb3VudDsKICAgIH0gZWxzZSB7CiAgICAgIG5ld0NsaWVudHNFbC50ZXh0Q29udGVudCA9ICfigJQnOwogICAgfQogIH0KCiAgcmVuZGVyTWFzdGVycyhtYXN0ZXJNYXAsIGRpc2NvdW50LCB0b3RhbCk7Cn0KCmZ1bmN0aW9uIHJlbmRlck1hc3RlcnMobWFzdGVyTWFwLCBkaXNjb3VudCwgdG90YWxBbGwpIHsKICB2YXIgZWwgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbWFzdGVyc0xpc3QnKTsKICBpZiAoT2JqZWN0LmtleXMobWFzdGVyTWFwKS5sZW5ndGggPT09IDApIHsKICAgIGVsLmlubmVySFRNTCA9ICc8ZGl2IGNsYXNzPSJsb2FkaW5nIj7QndC10YIg0LfQsNC/0LjRgdC10Lkg0LfQsCDQstGL0LHRgNCw0L3QvdGL0Lkg0L/QtdGA0LjQvtC0PC9kaXY+JzsKICAgIHJldHVybjsKICB9CiAgdmFyIGh0bWwgPSAnJzsKICB2YXIgbWFzdGVycyA9IE9iamVjdC5rZXlzKG1hc3Rlck1hcCkuc29ydChmdW5jdGlvbihhLGIpeyByZXR1cm4gbWFzdGVyTWFwW2JdLnRvdGFsIC0gbWFzdGVyTWFwW2FdLnRvdGFsOyB9KTsKICBtYXN0ZXJzLmZvckVhY2goZnVuY3Rpb24obmFtZSkgewogICAgdmFyIGQgPSBtYXN0ZXJNYXBbbmFtZV07CiAgICB2YXIgcmF0aW8gPSBnZXRNYXN0ZXJSYXRpbyhuYW1lKTsKICAgIHZhciBtYXN0ZXJFYXJuID0gTWF0aC5yb3VuZChkLnRvdGFsICogcmF0aW8pOwogICAgdmFyIHNhbG9uU2hhcmUgPSBNYXRoLnJvdW5kKGQudG90YWwgKiAoMSAtIHJhdGlvKSk7CiAgICB2YXIgcmF0aW8gPSB0b3RhbEFsbCA+IDAgPyBkLnRvdGFsIC8gdG90YWxBbGwgOiAwOwogICAgdmFyIHNhbG9uRGlzY291bnQgPSBNYXRoLnJvdW5kKGRpc2NvdW50ICogcmF0aW8pOwogICAgdmFyIHNhbG9uRWFybiA9IHNhbG9uU2hhcmUgLSBzYWxvbkRpc2NvdW50OwogICAgdmFyIGlkID0gJ21jXycgKyBuYW1lLnJlcGxhY2UoL1xzL2csJ18nKTsKICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9Im1hc3Rlci1jYXJkIGFuaW0iPic7CiAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJtYXN0ZXItaGVhZCIgaWQ9Im1oXycraWQrJyIgb25jbGljaz0idG9nZ2xlTWFzdGVyKFwnJyArIGlkICsgJ1wnKSI+JzsKICAgIGh0bWwgKz0gJzxkaXY+PGRpdiBjbGFzcz0ibW5hbWUiPicgKyBuYW1lICsgJzwvZGl2Pic7CiAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJtY291bnQiPicgKyBkLmJvb2tpbmdzLmxlbmd0aCArICcg0LfQsNC/0LjRgdC10LkgwrcgJyArIGQudG90YWwgKyAnIOKCrCDQvtCx0YnQsNGPINGB0YPQvNC80LA8L2Rpdj48L2Rpdj4nOwogICAgaHRtbCArPSAnPGRpdiBjbGFzcz0ibWVhcm5pbmdzIj4nOwogICAgaHRtbCArPSAnPGRpdiBjbGFzcz0iZWFybi1pdGVtIj48ZGl2IGNsYXNzPSJlYXJuLWxhYmVsIj7QnNCw0YHRgtC10YA8L2Rpdj48ZGl2IGNsYXNzPSJlYXJuLXZhbCI+JyArIG1hc3RlckVhcm4gKyAnIOKCrDwvZGl2PjwvZGl2Pic7CiAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJlYXJuLWl0ZW0iPjxkaXYgY2xhc3M9ImVhcm4tbGFiZWwiPtCh0LDQu9C+0L08L2Rpdj48ZGl2IGNsYXNzPSJlYXJuLXZhbCBzYWxvbiI+JyArIHNhbG9uRWFybiArICcg4oKsPC9kaXY+PC9kaXY+JzsKICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9ImNoZXZyb24iPuKWvDwvZGl2PjwvZGl2PjwvZGl2Pic7CiAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJtYXN0ZXItYm9keSIgaWQ9Im1iXycraWQrJyI+JzsKICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9InZpc2l0cy1oZWFkZXIiPjxzcGFuPtCU0LDRgtCwPC9zcGFuPjxzcGFuPtCa0LvQuNC10L3RgiAvINCf0LjRgtC+0LzQtdGGPC9zcGFuPjxzcGFuPtCj0YHQu9GD0LPQsDwvc3Bhbj48c3BhbiBzdHlsZT0idGV4dC1hbGlnbjpyaWdodCI+0KbQtdC90LA8L3NwYW4+PC9kaXY+JzsKICAgIHZhciBzb3J0ZWQgPSBkLmJvb2tpbmdzLnNsaWNlKCkuc29ydChmdW5jdGlvbihhLGIpeyByZXR1cm4gYi5kYXRlLmxvY2FsZUNvbXBhcmUoYS5kYXRlKTsgfSk7CiAgICBzb3J0ZWQuZm9yRWFjaChmdW5jdGlvbihiKSB7CiAgICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9InZpc2l0LXJvdyI+JzsKICAgICAgaHRtbCArPSAnPGRpdiBjbGFzcz0idmlzaXQtZGF0ZSI+JyArIGIuZGF0ZSArICc8L2Rpdj4nOwogICAgICBodG1sICs9ICc8ZGl2PjxkaXYgY2xhc3M9InZpc2l0LWNsaWVudCI+JyArIChiLmNsaWVudE5hbWV8fCfigJQnKSArICc8L2Rpdj4nOwogICAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJ2aXNpdC1wZXQiPicgKyAoYi5wZXROYW1lID8gYi5wZXROYW1lICsgJyDCtyAnIDogJycpICsgYi5icmVlZCArICc8L2Rpdj48L2Rpdj4nOwogICAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJ2aXNpdC1zdmMiPicgKyBiLnNlcnZpY2UgKyAnPC9kaXY+JzsKICAgICAgaHRtbCArPSAnPGRpdiBjbGFzcz0idmlzaXQtcHJpY2UiPicgKyAoYi5wcmljZXx8MCkgKyAnIOKCrDwvZGl2PjwvZGl2Pic7CiAgICB9KTsKICAgIGh0bWwgKz0gJzwvZGl2PjwvZGl2Pic7CiAgfSk7CiAgZWwuaW5uZXJIVE1MID0gaHRtbDsKfQoKZnVuY3Rpb24gdG9nZ2xlTWFzdGVyKGlkKSB7CiAgdmFyIGhlYWQgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbWhfJytpZCk7CiAgdmFyIGJvZHkgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbWJfJytpZCk7CiAgaGVhZC5jbGFzc0xpc3QudG9nZ2xlKCdvcGVuJyk7CiAgYm9keS5jbGFzc0xpc3QudG9nZ2xlKCdvcGVuJyk7Cn0KCi8vIOKUgOKUgCDQo9Cf0KDQkNCS0JvQldCd0JjQlTog0JzQkNCh0KLQldCg0JAg4pSA4pSACmZ1bmN0aW9uIHJlbmRlck1hc3Rlckxpc3QoKSB7CiAgdmFyIGVsID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ21hc3Rlckxpc3RVSScpOwogIGlmICghZWwpIHJldHVybjsKICB2YXIgaHRtbCA9ICcnOwogIGRiLm1hc3RlcnMuZm9yRWFjaChmdW5jdGlvbihuYW1lLCBpKSB7CiAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJsaXN0LWl0ZW0iPjxkaXY+PGRpdiBjbGFzcz0ibGktbmFtZSI+JytuYW1lKyc8L2Rpdj48L2Rpdj4nOwogICAgaHRtbCArPSAnPGJ1dHRvbiBjbGFzcz0iZGVsLWJ0biIgb25jbGljaz0iZGVsTWFzdGVyKCcraSsnKSI+4pyVPC9idXR0b24+PC9kaXY+JzsKICB9KTsKICBlbC5pbm5lckhUTUwgPSBodG1sIHx8ICc8ZGl2IHN0eWxlPSJjb2xvcjojNDQ0NDQwO2ZvbnQtc2l6ZTouNzVyZW07cGFkZGluZzoxMnB4IDAiPtCd0LXRgiDQvNCw0YHRgtC10YDQvtCyPC9kaXY+JzsKfQpmdW5jdGlvbiBhZGRNYXN0ZXIoKSB7CiAgdmFyIG5hbWUgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbmV3TWFzdGVyTmFtZScpLnZhbHVlLnRyaW0oKTsKICBpZiAoIW5hbWUpIHJldHVybjsKICBkYi5tYXN0ZXJzLnB1c2gobmFtZSk7CiAgc2F2ZSgpOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCduZXdNYXN0ZXJOYW1lJykudmFsdWUgPSAnJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbmV3TWFzdGVyUGhvbmUnKS52YWx1ZSA9ICcnOwogIHJlbmRlck1hc3Rlckxpc3QoKTsKfQpmdW5jdGlvbiBkZWxNYXN0ZXIoaSkgewogIGlmICghY29uZmlybSgn0KPQtNCw0LvQuNGC0Ywg0LzQsNGB0YLQtdGA0LA/JykpIHJldHVybjsKICBkYi5tYXN0ZXJzLnNwbGljZShpLCAxKTsKICBzYXZlKCk7CiAgcmVuZGVyTWFzdGVyTGlzdCgpOwp9CgovLyDilIDilIAg0KPQn9Cg0JDQktCb0JXQndCY0JU6INCf0J7QoNCe0JTQqyDilIDilIAKZnVuY3Rpb24gcmVuZGVyQnJlZWRMaXN0KCkgewogIHZhciBlbCA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdicmVlZExpc3RVSScpOwogIGlmICghZWwpIHJldHVybjsKICB2YXIgaHRtbCA9ICcnOwogIGRiLmJyZWVkcy5mb3JFYWNoKGZ1bmN0aW9uKGJyZWVkLCBiaSkgewogICAgdmFyIGJpZCA9ICdicl8nK2JpOwogICAgaHRtbCArPSAnPGRpdiBjbGFzcz0iYnJlZWQtY2FyZCI+JzsKICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9ImJyZWVkLWhlYWQiIG9uY2xpY2s9InRvZ2dsZUJyZWVkKFwnJyArIGJpZCArICdcJykiPic7CiAgICBodG1sICs9ICc8ZGl2PjxkaXYgY2xhc3M9ImJyZWVkLW5hbWUiPicrYnJlZWQubmFtZSsnPC9kaXY+JzsKICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9ImJyZWVkLWNvdW50Ij4nKyhicmVlZC5zZXJ2aWNlc3x8W10pLmxlbmd0aCsnINGD0YHQu9GD0LM8L2Rpdj48L2Rpdj4nOwogICAgaHRtbCArPSAnPGRpdiBzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6MTBweCI+JzsKICAgIGh0bWwgKz0gJzxidXR0b24gY2xhc3M9ImRlbC1idG4iIG9uY2xpY2s9ImV2ZW50LnN0b3BQcm9wYWdhdGlvbigpO2RlbEJyZWVkKCcrYmkrJykiPuKclTwvYnV0dG9uPic7CiAgICBodG1sICs9ICc8c3BhbiBzdHlsZT0iY29sb3I6IzhhOGE1NTtmb250LXNpemU6Ljc1cmVtIj7ilrw8L3NwYW4+PC9kaXY+PC9kaXY+JzsKICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9ImJyZWVkLWJvZHkiIGlkPSInK2JpZCsnIj4nOwogICAgKGJyZWVkLnNlcnZpY2VzfHxbXSkuZm9yRWFjaChmdW5jdGlvbihzdmMsIHNpKSB7CiAgICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9InN2Yy1yb3ciPjxzcGFuIGNsYXNzPSJzdmMtbmFtZSI+JytzdmMubmFtZSsnPC9zcGFuPic7CiAgICAgIGh0bWwgKz0gJzxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjEwcHgiPic7CiAgICAgIGh0bWwgKz0gJzxzcGFuIGNsYXNzPSJzdmMtcHJpY2UiPicrc3ZjLnByaWNlKycg4oKsPC9zcGFuPic7CiAgICAgIGh0bWwgKz0gJzxidXR0b24gY2xhc3M9ImRlbC1idG4iIG9uY2xpY2s9ImRlbFN2YygnK2JpKycsJytzaSsnKSI+4pyVPC9idXR0b24+PC9kaXY+PC9kaXY+JzsKICAgIH0pOwogICAgaHRtbCArPSAnPGRpdiBjbGFzcz0iYWRkLXN2Yy1yb3ciPic7CiAgICBodG1sICs9ICc8aW5wdXQgY2xhc3M9ImZpIiBpZD0ic25fJytiaSsnIiBwbGFjZWhvbGRlcj0i0KPRgdC70YPQs9CwIj4nOwogICAgaHRtbCArPSAnPGlucHV0IGNsYXNzPSJmaSIgaWQ9InNwXycrYmkrJyIgcGxhY2Vob2xkZXI9ItCm0LXQvdCwIOKCrCI+JzsKICAgIGh0bWwgKz0gJzxidXR0b24gY2xhc3M9Imljb24tYnRuIiBvbmNsaWNrPSJhZGRTdmMoJytiaSsnKSI+KzwvYnV0dG9uPjwvZGl2Pic7CiAgICBodG1sICs9ICc8L2Rpdj48L2Rpdj4nOwogIH0pOwogIGVsLmlubmVySFRNTCA9IGh0bWwgfHwgJzxkaXYgc3R5bGU9ImNvbG9yOiM0NDQ0NDA7Zm9udC1zaXplOi43NXJlbTtwYWRkaW5nOjEycHggMCI+0J3QtdGCINC/0L7RgNC+0LQ8L2Rpdj4nOwogIHJlbmRlckJyZWVkU2VsZWN0KCk7Cn0KZnVuY3Rpb24gdG9nZ2xlQnJlZWQoaWQpIHsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZChpZCkuY2xhc3NMaXN0LnRvZ2dsZSgnb3BlbicpOwp9CmZ1bmN0aW9uIGFkZEJyZWVkKCkgewogIHZhciBuYW1lID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ25ld0JyZWVkTmFtZScpLnZhbHVlLnRyaW0oKTsKICBpZiAoIW5hbWUpIHJldHVybjsKICBkYi5icmVlZHMucHVzaCh7bmFtZTogbmFtZSwgc2VydmljZXM6IFtdfSk7CiAgc2F2ZSgpOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCduZXdCcmVlZE5hbWUnKS52YWx1ZSA9ICcnOwogIHJlbmRlckJyZWVkTGlzdCgpOwp9CmZ1bmN0aW9uIGRlbEJyZWVkKGkpIHsKICBpZiAoIWNvbmZpcm0oJ9Cj0LTQsNC70LjRgtGMINC/0L7RgNC+0LTRgyDQuCDQstGB0LUg0LXRkSDRg9GB0LvRg9Cz0Lg/JykpIHJldHVybjsKICBkYi5icmVlZHMuc3BsaWNlKGksIDEpOwogIHNhdmUoKTsKICByZW5kZXJCcmVlZExpc3QoKTsKfQpmdW5jdGlvbiBhZGRTdmMoYmkpIHsKICB2YXIgbmFtZSAgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnc25fJytiaSkudmFsdWUudHJpbSgpOwogIHZhciBwcmljZSA9IHBhcnNlRmxvYXQoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ3NwXycrYmkpLnZhbHVlKSB8fCAwOwogIGlmICghbmFtZSB8fCAhcHJpY2UpIHJldHVybjsKICBkYi5icmVlZHNbYmldLnNlcnZpY2VzLnB1c2goe25hbWU6IG5hbWUsIHByaWNlOiBwcmljZX0pOwogIHNhdmUoKTsKICByZW5kZXJCcmVlZExpc3QoKTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnYnJfJytiaSkuY2xhc3NMaXN0LmFkZCgnb3BlbicpOwp9CmZ1bmN0aW9uIGRlbFN2YyhiaSwgc2kpIHsKICBkYi5icmVlZHNbYmldLnNlcnZpY2VzLnNwbGljZShzaSwgMSk7CiAgc2F2ZSgpOwogIHJlbmRlckJyZWVkTGlzdCgpOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdicl8nK2JpKS5jbGFzc0xpc3QuYWRkKCdvcGVuJyk7Cn0KZnVuY3Rpb24gcmVuZGVyQnJlZWRTZWxlY3QoKSB7CiAgdmFyIHNlbCA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCduZXdDbGllbnRCcmVlZCcpOwogIGlmICghc2VsKSByZXR1cm47CiAgc2VsLmlubmVySFRNTCA9ICc8b3B0aW9uIHZhbHVlPSIiPtCf0L7RgNC+0LTQsC4uLjwvb3B0aW9uPic7CiAgZGIuYnJlZWRzLmZvckVhY2goZnVuY3Rpb24oYikgewogICAgc2VsLmlubmVySFRNTCArPSAnPG9wdGlvbj4nK2IubmFtZSsnPC9vcHRpb24+JzsKICB9KTsKfQoKLy8g4pSA4pSAINCj0J/QoNCQ0JLQm9CV0J3QmNCVOiDQmtCb0JjQldCd0KLQqyDilIDilIAKZnVuY3Rpb24gcmVuZGVyQ2xpZW50TGlzdCgpIHsKICB2YXIgZWwgPSBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY2xpZW50TGlzdFVJJyk7CiAgaWYgKCFlbCkgcmV0dXJuOwogIHZhciBtZXJnZWQgPSB7fTsKICBkYi5jbGllbnRzLmZvckVhY2goZnVuY3Rpb24oYykgewogICAgbWVyZ2VkW2MucGhvbmV8fGMubmFtZV0gPSBjOwogIH0pOwogIGFsbEJvb2tpbmdzLmZvckVhY2goZnVuY3Rpb24oYikgewogICAgaWYgKCFiLmNsaWVudFBob25lICYmICFiLmNsaWVudE5hbWUpIHJldHVybjsKICAgIHZhciBrZXkgPSBiLmNsaWVudFBob25lIHx8IGIuY2xpZW50TmFtZTsKICAgIGlmICghbWVyZ2VkW2tleV0pIHsKICAgICAgbWVyZ2VkW2tleV0gPSB7bmFtZTogYi5jbGllbnROYW1lLCBwaG9uZTogYi5jbGllbnRQaG9uZSwgcGV0OiBiLnBldE5hbWUsIGJyZWVkOiBiLmJyZWVkLCBub3RlOiAnJywgdmlzaXRzOiAwLCB0b3RhbDogMH07CiAgICB9CiAgICBtZXJnZWRba2V5XS52aXNpdHMgPSAobWVyZ2VkW2tleV0udmlzaXRzfHwwKSArIDE7CiAgICBtZXJnZWRba2V5XS50b3RhbCAgPSAobWVyZ2VkW2tleV0udG90YWx8fDApICsgKHBhcnNlRmxvYXQoYi5wcmljZSl8fDApOwogICAgbWVyZ2VkW2tleV0ubGFzdERhdGUgICA9IGIuZGF0ZTsKICAgIG1lcmdlZFtrZXldLmxhc3RNYXN0ZXIgPSBiLm1hc3RlcjsKICB9KTsKICB2YXIgYXJyID0gT2JqZWN0LnZhbHVlcyhtZXJnZWQpOwogIGFyci5zb3J0KGZ1bmN0aW9uKGEsYil7IHJldHVybiAoYi52aXNpdHN8fDApLShhLnZpc2l0c3x8MCk7IH0pOwogIHZhciBodG1sID0gJyc7CiAgYXJyLmZvckVhY2goZnVuY3Rpb24oYywgaSkgewogICAgdmFyIGluaXRpYWxzID0gKGMubmFtZXx8Jz8nKS5zcGxpdCgnICcpLm1hcChmdW5jdGlvbih3KXtyZXR1cm4gd1swXXx8Jyc7fSkuam9pbignJykuc3Vic3RyaW5nKDAsMikudG9VcHBlckNhc2UoKTsKICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9ImNsaWVudC1jYXJkIGFuaW0iPic7CiAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJjbC1yb3ciPic7CiAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJjbC1hdmF0YXIiPicraW5pdGlhbHMrJzwvZGl2Pic7CiAgICBodG1sICs9ICc8ZGl2IHN0eWxlPSJmbGV4OjEiPjxkaXYgY2xhc3M9ImNsLW5hbWUiPicrKGMubmFtZXx8J+KAlCcpKyc8L2Rpdj4nOwogICAgaHRtbCArPSAnPGRpdiBjbGFzcz0iY2wtZGV0YWlsIj4nKyhjLnBob25lfHwnJykrKGMucGV0PycgwrcgJytjLnBldDonJykrKGMuYnJlZWQ/JyDCtyAnK2MuYnJlZWQ6JycpKyc8L2Rpdj48L2Rpdj4nOwogICAgaWYgKCFjLmZyb21DYWxlbmRhcikgewogICAgICBodG1sICs9ICc8YnV0dG9uIGNsYXNzPSJkZWwtYnRuIiBvbmNsaWNrPSJkZWxDbGllbnQoJytpKycpIj7inJU8L2J1dHRvbj4nOwogICAgfQogICAgaHRtbCArPSAnPC9kaXY+JzsKICAgIGlmICgoYy52aXNpdHN8fDApID4gMCkgewogICAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJjbC1zdGF0cyI+JzsKICAgICAgaHRtbCArPSAnPGRpdiBjbGFzcz0iY3N0YXQiPjxkaXYgY2xhc3M9ImNzdC12YWwiPicrKGMudmlzaXRzfHwwKSsnPC9kaXY+PGRpdiBjbGFzcz0iY3N0LWxhYmVsIj7QstC40LfQuNGC0L7QsjwvZGl2PjwvZGl2Pic7CiAgICAgIGh0bWwgKz0gJzxkaXYgY2xhc3M9ImNzdGF0Ij48ZGl2IGNsYXNzPSJjc3QtdmFsIj4nKyhjLnRvdGFsfHwwKSsnIOKCrDwvZGl2PjxkaXYgY2xhc3M9ImNzdC1sYWJlbCI+0L/QvtGC0YDQsNGH0LXQvdC+PC9kaXY+PC9kaXY+JzsKICAgICAgaHRtbCArPSAnPC9kaXY+JzsKICAgICAgaWYgKGMubGFzdERhdGUpIHsKICAgICAgICBodG1sICs9ICc8ZGl2IGNsYXNzPSJjbC1sYXN0Ij7Qn9C+0YHQu9C10LTQvdC40Lkg0LLQuNC30LjRgjogPHNwYW4+JytjLmxhc3REYXRlKyhjLmxhc3RNYXN0ZXI/JyDCtyAnK2MubGFzdE1hc3RlcjonJykrJzwvc3Bhbj48L2Rpdj4nOwogICAgICB9CiAgICB9CiAgICBpZiAoYy5ub3RlKSBodG1sICs9ICc8ZGl2IHN0eWxlPSJmb250LXNpemU6LjYycmVtO2NvbG9yOiM1NTU1NTA7bWFyZ2luLXRvcDo2cHgiPicrYy5ub3RlKyc8L2Rpdj4nOwogICAgaHRtbCArPSAnPC9kaXY+JzsKICB9KTsKICBlbC5pbm5lckhUTUwgPSBodG1sIHx8ICc8ZGl2IHN0eWxlPSJjb2xvcjojNDQ0NDQwO2ZvbnQtc2l6ZTouNzVyZW07cGFkZGluZzoxMnB4IDAiPtCd0LXRgiDQutC70LjQtdC90YLQvtCyPC9kaXY+JzsKfQpmdW5jdGlvbiBhZGRDbGllbnQoKSB7CiAgdmFyIG5hbWUgID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ25ld0NsaWVudE5hbWUnKS52YWx1ZS50cmltKCk7CiAgdmFyIHBob25lID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ25ld0NsaWVudFBob25lJykudmFsdWUudHJpbSgpOwogIHZhciBwZXQgICA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCduZXdDbGllbnRQZXQnKS52YWx1ZS50cmltKCk7CiAgdmFyIGJyZWVkID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ25ld0NsaWVudEJyZWVkJykudmFsdWU7CiAgdmFyIG5vdGUgID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ25ld0NsaWVudE5vdGUnKS52YWx1ZS50cmltKCk7CiAgaWYgKCFuYW1lKSB7IGFsZXJ0KCfQktCy0LXQtNC40YLQtSDQuNC80Y8g0LrQu9C40LXQvdGC0LAnKTsgcmV0dXJuOyB9CiAgZGIuY2xpZW50cy5wdXNoKHtuYW1lOm5hbWUsIHBob25lOnBob25lLCBwZXQ6cGV0LCBicmVlZDpicmVlZCwgbm90ZTpub3RlLCB2aXNpdHM6MCwgdG90YWw6MH0pOwogIHNhdmUoKTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbmV3Q2xpZW50TmFtZScpLnZhbHVlID0gJyc7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ25ld0NsaWVudFBob25lJykudmFsdWUgPSAnJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbmV3Q2xpZW50UGV0JykudmFsdWUgPSAnJzsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbmV3Q2xpZW50Tm90ZScpLnZhbHVlID0gJyc7CiAgcmVuZGVyQ2xpZW50TGlzdCgpOwp9CmZ1bmN0aW9uIGRlbENsaWVudChpKSB7CiAgaWYgKCFjb25maXJtKCfQo9C00LDQu9C40YLRjCDQutCw0YDRgtC+0YfQutGDINC60LvQuNC10L3RgtCwPycpKSByZXR1cm47CiAgZGIuY2xpZW50cy5zcGxpY2UoaSwgMSk7CiAgc2F2ZSgpOwogIHJlbmRlckNsaWVudExpc3QoKTsKfQoKLy8g4pSA4pSAIElOSVQg4pSA4pSACmRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdkaXNjb3VudElucHV0JykudmFsdWUgPSBkYi5kaXNjb3VudDsKcmVuZGVyTWFzdGVyTGlzdCgpOwpyZW5kZXJCcmVlZExpc3QoKTsKcmVuZGVyQ2xpZW50TGlzdCgpOwpsb2FkU3RhdHMoKTsKPC9zY3JpcHQ+CjwvYm9keT4KPC9odG1sPg=="

@app.route("/stats")
def stats_page():
    import base64
    html = base64.b64decode(STATS_HTML_B64).decode("utf-8")
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/api/stats")
def api_stats():
    from_date = request.args.get("from", "")
    to_date   = request.args.get("to", "")
    try:
        params = {"action": "stats"}
        if from_date: params["from"] = from_date
        if to_date:   params["to"]   = to_date
        r = requests.get(GOOGLE_SCRIPT, params=params, timeout=20)
        resp = jsonify(r.json())
    except Exception as e:
        resp = jsonify({"success": False, "error": str(e), "bookings": []})
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp

# ── /test-chat ────────────────────────────────────────────────────────────────
_TEST_CHAT_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Анна — тест чата</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
     background:#0b141a;height:100vh;display:flex;flex-direction:column;overflow:hidden}
.header{background:#202c33;padding:10px 16px;display:flex;align-items:center;
        gap:12px;border-bottom:1px solid #111b21;flex-shrink:0}
.avatar{width:40px;height:40px;border-radius:50%;background:#00a884;
        display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0}
.hinfo{flex:1}
.hname{color:#e9edef;font-size:15px;font-weight:600}
.hstatus{color:#8696a0;font-size:12px;margin-top:1px}
.reset{background:none;border:1px solid #8696a0;color:#8696a0;padding:5px 12px;
       border-radius:6px;cursor:pointer;font-size:12px;white-space:nowrap}
.reset:hover{border-color:#e9edef;color:#e9edef}
.messages{flex:1;overflow-y:auto;padding:16px 5vw;display:flex;flex-direction:column;gap:2px}
.messages::-webkit-scrollbar{width:5px}
.messages::-webkit-scrollbar-thumb{background:#374045;border-radius:3px}
.bubble{max-width:72%;padding:8px 12px 4px;border-radius:8px;
        font-size:14px;line-height:1.5;word-break:break-word;margin:1px 0}
.bubble .ts{font-size:11px;color:#8696a0;text-align:right;margin-top:3px}
.jarvis{background:#202c33;color:#e9edef;align-self:flex-start;border-top-left-radius:0}
.client{background:#005c4b;color:#e9edef;align-self:flex-end;border-top-right-radius:0}
.typing-row{display:none;align-self:flex-start;padding:2px 0}
.dots{display:flex;gap:5px;padding:10px 14px;background:#202c33;border-radius:8px;
      border-top-left-radius:0}
.dots span{width:8px;height:8px;border-radius:50%;background:#8696a0;
           animation:blink 1.2s infinite}
.dots span:nth-child(2){animation-delay:.2s}
.dots span:nth-child(3){animation-delay:.4s}
@keyframes blink{0%,60%,100%{opacity:.3}30%{opacity:1}}
.input-bar{background:#202c33;padding:10px 12px;display:flex;align-items:center;
           gap:8px;flex-shrink:0}
.input-bar input{flex:1;background:#2a3942;border:none;border-radius:24px;
                 padding:10px 16px;color:#e9edef;font-size:15px;outline:none}
.input-bar input::placeholder{color:#8696a0}
.send{width:44px;height:44px;border-radius:50%;background:#00a884;border:none;
      cursor:pointer;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.send:disabled{background:#3a4a52;cursor:default}
.send svg{fill:#fff}
</style>
</head>
<body>
<div class="header">
  <div class="avatar">🐾</div>
  <div class="hinfo">
    <div class="hname">Анна — R&amp;J Grooming</div>
    <div class="hstatus" id="st">онлайн</div>
  </div>
  <button class="reset" onclick="resetChat()">Сбросить</button>
</div>
<div class="messages" id="msgs"></div>
<div class="typing-row" id="typing"><div class="dots"><span></span><span></span><span></span></div></div>
<div class="input-bar">
  <input id="inp" type="text" placeholder="Сообщение" onkeydown="if(event.key==='Enter')send()">
  <button class="send" id="sb" onclick="send()">
    <svg viewBox="0 0 24 24" width="22" height="22"><path d="M2.01 21L23 12 2.01 3 2 10l15 2-15 2z"/></svg>
  </button>
</div>
<script>
const msgs=document.getElementById('msgs'),inp=document.getElementById('inp'),
      typing=document.getElementById('typing'),sb=document.getElementById('sb'),
      st=document.getElementById('st');
function ts(){return new Date().toLocaleTimeString('ru',{hour:'2-digit',minute:'2-digit'})}
function addBubble(text,role){
  const d=document.createElement('div');
  d.className='bubble '+role;
  const html=text.replace(/\\*\\*(.*?)\\*\\*/g,'<b>$1</b>').replace(/\\n/g,'<br>');
  d.innerHTML=html+'<div class="ts">'+ts()+'</div>';
  msgs.appendChild(d);
  msgs.scrollTop=msgs.scrollHeight;
}
async function send(){
  const text=inp.value.trim();if(!text)return;
  inp.value='';sb.disabled=true;
  addBubble(text,'client');
  typing.style.display='flex';msgs.scrollTop=msgs.scrollHeight;
  st.textContent='печатает…';
  try{
    const r=await fetch('/test-chat/send',{method:'POST',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({text})});
    const d=await r.json();
    typing.style.display='none';st.textContent='онлайн';
    if(d.reply)addBubble(d.reply,'jarvis');
    if(d.second_reply){
      await new Promise(r=>setTimeout(r,900));
      typing.style.display='flex';msgs.scrollTop=msgs.scrollHeight;
      await new Promise(r=>setTimeout(r,700));
      typing.style.display='none';
      addBubble(d.second_reply,'jarvis');
    }
    if(d.booked){st.textContent='запись принята ✓';setTimeout(()=>st.textContent='онлайн',4000);}
  }catch(e){
    typing.style.display='none';st.textContent='ошибка';
    addBubble('Ошибка соединения. Попробуйте ещё раз.','jarvis');
  }
  sb.disabled=false;inp.focus();
}
async function resetChat(){
  await fetch('/test-chat/reset',{method:'POST'});
  msgs.innerHTML='';st.textContent='онлайн';
}
inp.focus();
</script>
</body>
</html>"""

@app.route("/test-chat")
def test_chat():
    return _TEST_CHAT_HTML, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/test-chat/send", methods=["POST"])
def test_chat_send():
    try:
        return _test_chat_send()
    except Exception:
        import traceback
        print(f"[test-chat] UNHANDLED ERROR:\n{traceback.format_exc()}", flush=True)
        return jsonify({"reply": "Произошла ошибка сервера. Попробуйте ещё раз 🤍",
                        "state": {}, "booked": False}), 200

def _test_chat_send():
    data = request.get_json() or {}
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"error": "empty"}), 400

    sid = session.get("chat_sid")
    if not sid or sid not in _chat_sessions:
        sid = str(uuid.uuid4())
        session["chat_sid"] = sid
        _chat_sessions[sid] = {"history": [], "state": _blank_state(), "avail": {}}

    sess = _chat_sessions[sid]
    history, state = sess["history"], sess["state"]
    avail = sess.setdefault("avail", {})

    history.append({"role": "user", "content": text})
    if len(history) > MAX_HISTORY:
        history[:] = history[-MAX_HISTORY:]

    # Save flags before extraction to detect transitions
    prev_service_confirmed = bool(state.get("service_confirmed"))
    prev_breed  = state.get("breed")
    prev_phone  = state.get("clientPhone")

    # Extract state BEFORE generating reply so bot knows about slot availability
    new_state = _extract_state(history, state)

    booking_intent = bool(new_state.get("booking_intent"))

    # Shared flags for step guards
    _services_shown = sess.get("services_shown", False)
    _breed_known    = bool(new_state.get("breed"))
    _weight_known   = bool(new_state.get("weight"))
    _had_greeting   = any(m["role"] == "assistant" for m in history)

    # ── Subtype clarification: breed with varieties → ask before weight/services ──
    if (_breed_known and not _services_shown
            and not new_state.get("service_confirmed")
            and not booking_intent
            and _had_greeting):
        _clarify_q = _check_breed_needs_clarification(new_state.get("breed"))
        if _clarify_q:
            last_asst = next((m["content"] for m in reversed(history)
                              if m["role"] == "assistant"), "")
            if "уточните" not in last_asst.lower():
                history.append({"role": "assistant", "content": _clarify_q})
                sess["state"] = new_state
                print(f"[test-chat] subtype-clarify: breed={new_state.get('breed')!r}", flush=True)
                return jsonify({"reply": _clarify_q, "state": new_state, "booked": False})

    # ── Step 3 (strict): breed known, weight not yet → inject weight question ──
    # Ensures Claude never skips asking for weight before showing services.
    if (_breed_known and not _weight_known and not _services_shown
            and not new_state.get("service_confirmed")
            and not booking_intent
            and _had_greeting):
        last_asst = next((m["content"] for m in reversed(history)
                          if m["role"] == "assistant"), "")
        _weight_already_asked = any(w in last_asst.lower()
                                    for w in ["вес", "весит", "килограмм", "сколько кг"])
        if not _weight_already_asked:
            reply = (_breed_compliment(new_state["breed"])
                     + "\n\nПодскажите, пожалуйста, примерный вес питомца?")
            history.append({"role": "assistant", "content": reply})
            sess["state"] = new_state
            print(f"[test-chat] step3-weight: breed={new_state['breed']!r}", flush=True)
            return jsonify({"reply": reply, "state": new_state, "booked": False})

    # ── Step 4 (strict): breed + weight both known → show services ──────────────
    if (_breed_known and _weight_known and not _services_shown
            and not new_state.get("service_confirmed")
            and not booking_intent
            and _had_greeting):
        services = _get_all_services_for_breed(new_state["breed"])
        sess["services_shown"] = True  # mark regardless so we don't retry on empty
        if services:
            lines = [f"• {svc} — {_SVC_DESCRIPTIONS.get(svc, '')}" for svc, _ in services]
            comp = next((svc for svc, _ in services if 'комплексный' in svc.lower()), None)
            gig  = next((svc for svc, _ in services if 'гигиенический' in svc.lower()), None)
            recommend = comp or gig or services[-1][0]
            compliment = _breed_compliment(new_state["breed"])
            msg = (compliment + "\n\nДля вашей породы доступны следующие услуги:\n"
                   + "\n".join(lines))
            if comp:
                msg += (f"\n\nКомплексный уход — самый полный вариант: питомец выходит "
                        f"полностью ухоженным — шерсть, когти, уши, глаза и стрижка 🤍 "
                        f"Многие клиенты выбирают именно его.\n\n"
                        f"Я бы посоветовала {recommend}. Вам подходит?")
            else:
                msg += (f"\n\nЯ бы посоветовала {recommend} — для вашей породы это "
                        "наиболее полный вариант ухода. Вам подходит?")
            history.append({"role": "assistant", "content": msg})
            sess["state"] = new_state
            print(f"[test-chat] auto-services: breed={new_state['breed']!r} prev={prev_breed!r} "
                  f"services={[s for s,_ in services]} recommend={recommend!r}", flush=True)
            return jsonify({"reply": msg, "state": new_state, "booked": False})

    # ── Phone validation: must be +372XXXXXXX format ─────────────────────────
    _new_phone = new_state.get("clientPhone")
    if _new_phone and _new_phone != prev_phone:
        _clean = re.sub(r'[\s\-()]', '', str(_new_phone))
        if not re.match(r'^\+372\d{6,8}$', _clean):
            new_state["clientPhone"] = None  # reset invalid phone
            reply = ("Пожалуйста, укажите номер телефона в формате +372XXXXXXX "
                     "(например, +37253112233) 🤍")
            history.append({"role": "assistant", "content": reply})
            sess["state"] = new_state
            print(f"[test-chat] phone-invalid: {_new_phone!r}", flush=True)
            return jsonify({"reply": reply, "state": new_state, "booked": False})

    # ── Step 6 (strict): service just confirmed → inject price from Python ─────
    just_confirmed_service = (not prev_service_confirmed) and bool(new_state.get("service_confirmed"))
    if just_confirmed_service:
        breed   = new_state.get("breed")
        service = new_state.get("service")
        price   = _lookup_price(breed, service)
        service_label = service or "Услуга"
        if price:
            disclaimer = _DISCLAIMER_SMOOTH if _is_smooth_coat(breed) else _DISCLAIMER_OTHER
            reply = f"{service_label} — от {price}€.\n\n{disclaimer}\n\nХотите записаться?"
        else:
            disclaimer = _DISCLAIMER_SMOOTH if _is_smooth_coat(breed) else _DISCLAIMER_OTHER
            reply = f"{service_label} отлично подойдёт вашему питомцу 🤍\n\n{disclaimer}\n\nХотите записаться?"
        history.append({"role": "assistant", "content": reply})
        new_state["booking_intent"] = False  # reset: next "да" must be answer to "Хотите записаться?"
        new_state["price_shown"]    = True   # persist in state so it survives across requests
        sess["state"]       = new_state
        sess["price_shown"] = True           # also in session for same-request reads
        print(f"[test-chat] price-inject: breed={breed!r} service={service!r} price={price}", flush=True)
        return jsonify({"reply": reply, "state": new_state, "booked": False})

    # ── Schedule trigger ──────────────────────────────────────────────────────
    _no_date_yet     = not new_state.get("date")
    _schedule_cached = "full_schedule" in avail
    # Read price_shown from state (persists via extraction merge) OR session fallback
    _price_shown     = bool(new_state.get("price_shown")) or sess.get("price_shown", False)
    _needs_schedule  = booking_intent and _price_shown and _no_date_yet and not _schedule_cached

    print(
        f"[schedule-trigger] service={new_state.get('service')!r} "
        f"booking_intent={booking_intent} price_shown={_price_shown} "
        f"(state={new_state.get('price_shown')} sess={sess.get('price_shown', False)}) "
        f"date={new_state.get('date')!r} cached={_schedule_cached} → needs_schedule={_needs_schedule}",
        flush=True,
    )

    if _needs_schedule:
        schedule = _build_schedule_for_days()
        if schedule:
            avail["full_schedule"] = schedule
            reply = f"{schedule}\n\nКакое время вам удобно? 🤍"
        else:
            reply = "Расписание сейчас недоступно. Назовите удобную дату — я проверю наличие слотов 🤍"
        history.append({"role": "assistant", "content": reply})
        sess["state"] = new_state
        return jsonify({"reply": reply, "state": new_state, "booked": False})

    # ── Fetch slots for a specific date when client names one ─────────────────
    curr_date = new_state.get("date")
    curr_time = new_state.get("time")
    if curr_date and curr_date != avail.get("cached_date"):
        date_iso = _parse_date_to_iso(curr_date)
        if date_iso:
            slots = _fetch_slots_for_date(date_iso)
            avail.update({
                "cached_date": curr_date,
                "date_label": curr_date,
                "slots": slots,
                "requested_time": curr_time,
            })
    elif curr_time and curr_time != avail.get("requested_time"):
        avail["requested_time"] = curr_time

    # ── Normal single-reply flow ──────────────────────────────────────────────
    try:
        response = client_ai.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=400,
            system=TEST_CHAT_SYSTEM_PROMPT + _state_context(new_state) + _avail_context(avail),
            messages=history,
        )
        reply = _process_action_markers(response.content[0].text.strip())
        reply = _add_price_disclaimer(reply, new_state.get("breed"))
    except Exception as e:
        print(f"[test-chat] claude error: {e}", flush=True)
        reply = "Произошла ошибка. Пожалуйста, попробуйте ещё раз 🤍"
    history.append({"role": "assistant", "content": reply})

    sess["state"] = new_state

    # ── State logging ─────────────────────────────────────────────────────────
    _required = ("breed", "service", "date", "time", "ownerName", "petName")
    _missing = [k for k in _required if not new_state.get(k)]
    if "date" not in _missing and new_state.get("date"):
        if _parse_date_to_iso(new_state["date"]) is None:
            _missing.append("date")
            print(f"[test-chat] date {new_state['date']!r} not parseable", flush=True)
    print(
        f"[test-chat] state: confirmed={new_state.get('confirmed')} missing={_missing} | "
        f"breed={new_state.get('breed')!r} service={new_state.get('service')!r} "
        f"date={new_state.get('date')!r} time={new_state.get('time')!r} "
        f"owner={new_state.get('ownerName')!r} pet={new_state.get('petName')!r}",
        flush=True,
    )

    # ── Booking ───────────────────────────────────────────────────────────────
    booked = new_state.get("confirmed") and not _missing
    if booked:
        booking_date = _to_booking_date(new_state["date"])
        price = _extract_price_from_history(history)
        payload = {
            "breed":        new_state["breed"],
            "breedDisplay": " ".join(filter(None, [new_state["breed"], new_state.get("weight")])),
            "service":      new_state["service"],
            "price":        price,
            "date":         booking_date,
            "time":         new_state["time"],
            "name":         new_state["ownerName"],
            "pet":          new_state["petName"] or "",
            "master":       new_state.get("master") or "",
            "phone":        new_state.get("clientPhone") or "test-chat",
            "lang":         "ru",
            "source":       "test-chat",
        }
        # Validate all required fields before sending
        _required_booking = {"breed": payload["breed"], "service": payload["service"],
                             "date": payload["date"], "time": payload["time"],
                             "name": payload["name"]}
        _empty = [k for k, v in _required_booking.items() if not v]
        if _empty:
            print(f"[test-chat] ⚠️ booking missing required fields: {_empty}", flush=True)
        if not price:
            print("[test-chat] ⚠️ price=0 — Anna may not have mentioned price in conversation", flush=True)
        print(f"[test-chat] ✅ booking payload: {payload}", flush=True)
        if GOOGLE_SCRIPT:
            try:
                gs_resp = requests.get(GOOGLE_SCRIPT, params=payload, timeout=15)
                print(f"[test-chat] GS → {gs_resp.status_code}: {gs_resp.text[:300]}", flush=True)
            except Exception as e:
                print(f"[test-chat] GS call failed: {e}", flush=True)
        else:
            print("[test-chat] ⚠️ GOOGLE_SCRIPT not set", flush=True)
        del _chat_sessions[sid]
        session.pop("chat_sid", None)

    return jsonify({"reply": reply, "state": new_state, "booked": bool(booked)})

@app.route("/test-chat/reset", methods=["POST"])
def test_chat_reset():
    sid = session.pop("chat_sid", None)
    if sid:
        _chat_sessions.pop(sid, None)
    return jsonify({"ok": True})

@app.route("/api/callback", methods=["POST"])
def api_callback():
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    phone = data.get("phone", "").strip()
    msg = f"Заявка: {name}, тел {phone}"
    twilio_sid   = os.environ.get("TWILIO_ACCOUNT_SID")
    twilio_token = os.environ.get("TWILIO_AUTH_TOKEN")
    twilio_from  = os.environ.get("TWILIO_ADMIN_PHONE", "+37266922128")
    admin_to     = "+37258243141"
    print(f"[callback] name={name!r} phone={phone!r} sid_set={bool(twilio_sid)} from={twilio_from}", flush=True)

    # Telegram (не зависит от статуса email-домена)
    try:
        _send_reminder_telegram(
            f"📞 <b>Заявка на обратный звонок</b>\n\n"
            f"👤 Имя: {name or '—'}\n"
            f"📱 Телефон: {phone or '—'}\n\n"
            f"Клиент не нашёл удобное время в виджете бронирования."
        )
    except Exception as e:
        print(f"[callback] Telegram error: {e}", flush=True)

    # SMS через Twilio
    if twilio_sid and twilio_token:
        try:
            sms_url = f"https://api.twilio.com/2010-04-01/Accounts/{twilio_sid}/Messages.json"
            resp = requests.post(
                sms_url,
                auth=(twilio_sid, twilio_token),
                data={"From": twilio_from, "To": admin_to, "Body": msg},
                timeout=10
            )
            print(f"[callback] Twilio SMS response: {resp.status_code} {resp.text[:200]}", flush=True)
        except Exception as e:
            print(f"[callback] Twilio error: {e}", flush=True)
    else:
        print("[callback] Twilio creds not set — skipping", flush=True)

    # Email через Resend
    resend_key = os.environ.get("RESEND_API_KEY")
    if resend_key:
        try:
            email_resp = requests.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                json={
                    "from": "booking@rjgrooming.salon",
                    "to": ["myrnj1@gmail.com"],
                    "subject": "Новая заявка на обратный звонок",
                    "html": f"<h2>Заявка с сайта</h2><p><b>Имя:</b> {name}</p><p><b>Телефон:</b> {phone}</p>"
                },
                timeout=10
            )
            print(f"[callback] Resend response: {email_resp.status_code}", flush=True)
        except Exception as e:
            print(f"[callback] Resend error: {e}", flush=True)

    return jsonify({"ok": True}), 200

def _build_client_export_workbook():
    """Собирает Excel-книгу (2 листа: Клиенты, Напоминания) и возвращает
    (bytes, filename). Используется и веб-роутом, и еженедельной рассылкой."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
    bookings = _fetch_full_history_bookings()

    def parse_date(s):
        try:
            return datetime.strptime(s, "%d.%m.%Y")
        except Exception:
            return datetime.max

    bookings_sorted = sorted(bookings, key=lambda b: parse_date(b.get("date", "")))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    headers = ["Дата", "Время", "Клиент", "Телефон", "Email", "Питомец", "Порода", "Услуга", "Мастер", "Цена (EUR)"]
    ws.append(headers)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for b in bookings_sorted:
        ws.append([
            b.get("date", ""),
            b.get("time", ""),
            b.get("clientName", ""),
            b.get("clientPhone", ""),
            b.get("clientEmail", ""),
            b.get("petName", ""),
            b.get("breed", ""),
            b.get("service", ""),
            b.get("master", ""),
            b.get("price", "")
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"

    # ── Лист 2: Напоминания (последний визит на клиента, начиная с 20 мая) ──
    cutoff = datetime(2026, 5, 20)
    by_phone = {}
    for b in bookings:
        phone = (b.get("clientPhone") or "").strip()
        if not phone:
            continue
        d = parse_date(b.get("date", ""))
        if d == datetime.max or d < cutoff:
            continue
        if phone not in by_phone or d > by_phone[phone]["_d"]:
            by_phone[phone] = {
                "_d": d,
                "date": b.get("date", ""),
                "name": b.get("clientName", ""),
                "pet": b.get("petName", ""),
                "breed": b.get("breed", ""),
                "master": b.get("master", ""),
                "email": b.get("clientEmail", ""),
            }

    ws2 = wb.create_sheet("Напоминания")
    headers2 = ["Клиент", "Телефон", "Email", "Питомец", "Порода", "Последний визит", "Мастер", "Напоминание #1 (+35 дн.)", "Напоминание #2 (+42 дн.)"]
    ws2.append(headers2)
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    rows2 = sorted(by_phone.items(), key=lambda kv: kv[1]["_d"])
    for phone, info in rows2:
        r1 = info["_d"] + timedelta(days=35)
        r2 = info["_d"] + timedelta(days=42)
        ws2.append([
            info["name"], phone, info["email"], info["pet"], info["breed"],
            info["date"], info["master"],
            r1.strftime("%d.%m.%Y"), r2.strftime("%d.%m.%Y")
        ])

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    for col in ws2.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
    ws2.freeze_panes = "A2"

    # ── Лист 3: Дубли (по формату номера и по совпадению имени) ──
    by_raw_phone = {}
    for b in bookings:
        raw_phone = (b.get("clientPhone") or "").strip()
        if not raw_phone:
            continue
        norm = _normalize_phone(raw_phone)
        entry = by_raw_phone.setdefault(raw_phone, {"norm": norm, "names": set(), "pets": set(), "count": 0})
        if b.get("clientName"):
            entry["names"].add(b.get("clientName"))
        if b.get("petName"):
            entry["pets"].add(b.get("petName"))
        entry["count"] += 1

    by_norm = {}
    for raw_phone, info in by_raw_phone.items():
        by_norm.setdefault(info["norm"], []).append({"raw": raw_phone, **info})

    by_name = {}
    for raw_phone, info in by_raw_phone.items():
        for nm in info["names"]:
            nk = nm.strip().lower()
            if nk:
                by_name.setdefault(nk, set()).add(info["norm"])

    ws3 = wb.create_sheet("Дубли")
    ws3.append(["Тип дубля", "Телефон / Имя", "Детали"])
    for col_idx in range(1, 4):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for norm, variants in by_norm.items():
        if len(variants) > 1:
            for v in variants:
                ws3.append([
                    "Один номер, разный формат",
                    norm,
                    f"как записано: {v['raw']} | имена: {', '.join(sorted(v['names']))} | питомцы: {', '.join(sorted(v['pets']))} | броней: {v['count']}"
                ])

    for name_key, phones in by_name.items():
        if len(phones) > 1:
            ws3.append([
                "Одно имя, разные номера",
                name_key,
                f"телефоны: {', '.join(sorted(phones))}"
            ])

    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    for col in ws3.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws3.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)
    ws3.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"RJ_Grooming_clients_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename

@app.route("/admin/client")
def admin_client_detail():
    import urllib.parse as _urlp

    phone = (request.args.get("phone") or "").strip()
    if not phone:
        return "Не указан телефон клиента.", 400

    error = None
    name, email, instagram = "", "", ""
    visits = []
    pets = {}
    total = 0.0

    try:
        today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
        bookings = _fetch_full_history_bookings()

        def parse_date(s):
            try:
                return datetime.strptime(s, "%d.%m.%Y").date()
            except Exception:
                return None

        for b in bookings:
            if (b.get("clientPhone") or "").strip() != phone:
                continue
            if b.get("clientName"):
                name = b.get("clientName")
            if b.get("clientEmail"):
                email = b.get("clientEmail")
            if b.get("clientInstagram"):
                instagram = b.get("clientInstagram")
            pet = b.get("petName", "")
            breed = b.get("breed", "")
            if pet:
                if breed or pet not in pets:
                    pets[pet] = breed

            try:
                price = float(b.get("price") or 0)
            except Exception:
                price = 0
            total += price
            visits.append({
                "date": parse_date(b.get("date", "")),
                "date_str": b.get("date", ""),
                "service": b.get("service", ""),
                "master": b.get("master", ""),
                "price": price,
                "pet": pet,
                "breed": breed
            })
        visits.sort(key=lambda v: v["date"] or datetime.min.date(), reverse=True)
    except Exception as e:
        error = str(e)

    saved_data = _load_client_data(phone)
    if saved_data.get("email"):
        email = saved_data["email"]
    if saved_data.get("instagram"):
        instagram = saved_data["instagram"]
    if saved_data.get("name"):
        name = saved_data["name"]
    comment = saved_data.get("comment", "")
    display_phone = _normalize_phone(saved_data.get("phone_override") or phone)

    wa_digits = re.sub(r"[^\d]", "", display_phone)
    pets_str = ", ".join(f"{p} ({b})" if b else p for p, b in pets.items()) or "—"
    last_visit_str = visits[0]["date_str"] if visits else "—"
    first_visit_str = visits[-1]["date_str"] if visits else "—"

    email_html = f'<a class="contact-val link" href="mailto:{email}">{email}</a>' if email else '<span class="contact-val empty">не указан</span>'
    ig_html = f'<a class="contact-val link" href="https://instagram.com/{instagram.lstrip("@")}" target="_blank" rel="noopener">@{instagram.lstrip("@")}</a>' if instagram else '<span class="contact-val empty">не указан</span>'

    def pet_photo_card(pet_name, breed):
        key = _pet_photo_key(phone, pet_name)
        photo_url = _pet_photo_url(key)
        return f"""
        <div class="pet-photo-card">
          <div class="pet-photo-name">{pet_name}{f' · {breed}' if breed else ''}</div>
          <img src="{photo_url}" class="pet-photo-img" id="img-{key}"
               onerror="this.outerHTML='<div class=&quot;pet-photo-placeholder&quot; id=&quot;img-{key}&quot;>Нет фото анкеты</div>'">
          <label class="pet-photo-btn">
            📷 Загрузить / обновить фото анкеты
            <input type="file" accept="image/*" capture="environment" style="display:none" onchange="uploadPetPhoto(this, '{key}', '{_urlp.quote(phone)}', '{_urlp.quote(pet_name)}')">
          </label>
        </div>"""

    pet_photos_html = "".join(pet_photo_card(p, b) for p, b in pets.items()) if pets else '<div class="empty">Питомцы не найдены</div>'

    all_memberships = _load_memberships()
    client_memberships = [m for m in all_memberships.values() if (m.get("client_phone") or "").strip() == phone]
    client_memberships.sort(key=lambda m: m.get("id", ""), reverse=True)

    def membership_row_html(m):
        pct = round((m["used_visits"] / m["total_visits"]) * 100) if m["total_visits"] else 0
        status_label = "Завершён" if m["status"] == "completed" else "Активен"
        status_color = "#8a8578" if m["status"] == "completed" else "#4ade80"
        return f"""
        <a class="membership-row" href="/membership/{m['id']}" target="_blank">
          <div class="membership-row-top">
            <span class="membership-row-id">{m['id']}</span>
            <span class="membership-row-status" style="color:{status_color};border-color:{status_color}55">{status_label}</span>
          </div>
          <div class="membership-row-meta">{m.get('plan_name','')} · {m.get('pet_name','')} · до {m.get('expiry_date','—')}</div>
          <div class="membership-row-progress-bar"><div class="membership-row-progress-fill" style="width:{pct}%"></div></div>
          <div class="membership-row-count">{m['used_visits']}/{m['total_visits']} посещений</div>
        </a>"""

    memberships_html = "".join(membership_row_html(m) for m in client_memberships) if client_memberships else '<div class="empty">Абонементов нет</div>'
    first_pet_name = next(iter(pets.keys()), "")
    first_pet_type = pets.get(first_pet_name, "") if first_pet_name else ""
    new_membership_link = (
        f"/admin/memberships?client_name={_urlp.quote(name)}&client_phone={_urlp.quote(phone)}"
        f"&pet_name={_urlp.quote(first_pet_name)}&pet_type={_urlp.quote(first_pet_type)}"
        f"&client_email={_urlp.quote(email)}"
    )

    def visit_row(v):
        return f"""
        <div class="vrow">
          <div class="vrow-date">{v['date_str']}</div>
          <div class="vrow-mid">
            <span class="vrow-service">{v['service'] or '—'}</span>
            <span class="vrow-sub">{v['pet'] or '—'} · {v['master'] or '—'}</span>
          </div>
          <div class="vrow-price">{v['price']:.0f}€</div>
        </div>"""

    visits_html = "".join(visit_row(v) for v in visits) if visits else '<div class="empty">Визитов не найдено</div>'
    error_html = f'<div class="empty" style="color:#e0824a">Ошибка загрузки данных: {error}</div>' if error else ""

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{name or 'Клиент'} — R&J Grooming</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,500;0,600;0,700;1,500&family=Montserrat:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a09;color:#f2ede2;font-family:'Montserrat',sans-serif;padding:36px 20px 80px}}
  .wrap{{max-width:640px;margin:0 auto}}
  .back-link{{display:inline-block;font-size:0.74rem;color:rgba(201,160,90,.75);text-decoration:none;margin-bottom:20px}}
  .eyebrow{{font-size:0.68rem;letter-spacing:.3em;text-transform:uppercase;color:rgba(242,237,226,.4);margin-bottom:10px}}
  h1{{font-family:'Playfair Display',serif;font-weight:700;font-size:2.4rem;margin-bottom:4px}}
  .pets-sub{{font-size:0.85rem;color:rgba(242,237,226,.55);margin-bottom:28px}}

  .actions{{display:flex;gap:10px;margin-bottom:28px}}
  .abtn{{flex:1;text-align:center;padding:14px;border-radius:12px;text-decoration:none;font-size:0.85rem;font-weight:600}}
  .abtn.call{{background:rgba(201,160,90,.12);border:1px solid rgba(201,160,90,.4);color:#c9a05a}}
  .abtn.wa{{background:rgba(74,222,128,.1);border:1px solid rgba(74,222,128,.4);color:#4ade80}}

  .card{{background:#141310;border:1px solid rgba(255,255,255,.07);border-radius:14px;padding:20px;margin-bottom:16px}}
  .card-title{{font-size:0.66rem;letter-spacing:.15em;text-transform:uppercase;color:#c9a05a;margin-bottom:14px}}
  .card-title-row{{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px}}
  .card-title-row .card-title{{margin-bottom:0}}
  .edit-btn{{background:none;border:1px solid rgba(201,160,90,.35);color:#c9a05a;font-size:0.68rem;padding:5px 12px;border-radius:20px;cursor:pointer;font-family:'Montserrat',sans-serif}}
  .edit-field{{margin-bottom:14px}}
  .edit-field label{{display:block;font-size:0.72rem;color:rgba(242,237,226,.5);margin-bottom:6px}}
  .edit-field input{{width:100%;background:#0e0d0b;border:1px solid rgba(201,160,90,.3);border-radius:8px;padding:10px 12px;color:#f2ede2;font-family:'Montserrat',sans-serif;font-size:0.88rem}}
  .edit-field input:focus{{outline:none;border-color:#c9a05a}}
  .edit-field textarea{{width:100%;background:#0e0d0b;border:1px solid rgba(201,160,90,.3);border-radius:8px;padding:10px 12px;color:#f2ede2;font-family:'Montserrat',sans-serif;font-size:0.88rem;min-height:80px;resize:vertical}}
  .edit-field textarea:focus{{outline:none;border-color:#c9a05a}}
  .edit-actions{{display:flex;gap:8px;margin-top:4px}}
  .edit-save{{flex:1;background:#c9a05a;color:#0a0a09;border:none;border-radius:8px;padding:11px;font-weight:600;font-size:0.85rem;cursor:pointer}}
  .edit-cancel{{flex:1;background:none;border:1px solid rgba(255,255,255,.15);color:rgba(242,237,226,.6);border-radius:8px;padding:11px;font-size:0.85rem;cursor:pointer}}
  .contact-line{{display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid rgba(255,255,255,.05)}}
  .contact-line:last-child{{border-bottom:none}}
  .contact-label{{font-size:0.78rem;color:rgba(242,237,226,.5)}}
  .contact-val{{font-size:0.85rem;color:#f2ede2;text-decoration:none}}
  .contact-val.link{{color:#c9a05a}}
  .contact-val.empty{{color:rgba(242,237,226,.3);font-style:italic}}

  .stats-mini{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:16px}}
  .stat{{background:#141310;border:1px solid rgba(201,160,90,.18);border-radius:12px;padding:16px 10px;text-align:center}}
  .stat .n{{font-family:'Playfair Display',serif;font-size:1.5rem;font-weight:600;color:#c9a05a}}
  .stat .l{{font-size:0.6rem;letter-spacing:.06em;text-transform:uppercase;color:rgba(242,237,226,.5);margin-top:3px}}

  .list-label{{font-size:0.68rem;letter-spacing:.2em;text-transform:uppercase;color:#c9a05a;margin:24px 0 12px}}
  .vrow{{display:flex;align-items:center;gap:14px;background:#131210;border:1px solid rgba(255,255,255,.06);border-radius:10px;padding:12px 16px;margin-bottom:8px}}
  .vrow-date{{font-size:0.72rem;color:rgba(242,237,226,.45);width:78px;flex-shrink:0}}
  .vrow-mid{{flex:1;display:flex;flex-direction:column}}
  .vrow-service{{font-size:0.85rem;font-weight:500}}
  .vrow-sub{{font-size:0.72rem;color:rgba(242,237,226,.5);margin-top:2px}}
  .vrow-price{{font-size:0.85rem;font-weight:600;color:#c9a05a}}
  .empty{{text-align:center;padding:30px 0;color:rgba(242,237,226,.4);font-size:0.85rem}}

  .pet-photos-grid{{display:flex;flex-direction:column;gap:12px;margin-bottom:8px}}
  .pet-photo-card{{background:#141310;border:1px solid rgba(255,255,255,.07);border-radius:14px;padding:16px}}
  .pet-photo-name{{font-family:'Playfair Display',serif;font-size:1rem;font-weight:600;margin-bottom:10px}}
  .pet-photo-img{{width:100%;max-height:340px;object-fit:contain;border-radius:10px;background:#0a0a09;display:block;margin-bottom:10px}}
  .pet-photo-placeholder{{width:100%;height:80px;border:1px dashed rgba(242,237,226,.2);border-radius:10px;display:flex;align-items:center;justify-content:center;color:rgba(242,237,226,.35);font-size:0.78rem;margin-bottom:10px}}
  .pet-photo-btn{{display:block;text-align:center;background:rgba(201,160,90,.1);border:1px solid rgba(201,160,90,.4);color:#c9a05a;border-radius:10px;padding:11px;font-size:0.82rem;font-weight:600;cursor:pointer}}
  .membership-row{{display:block;background:#141310;border:1px solid rgba(255,255,255,.07);border-radius:12px;padding:14px 16px;margin-bottom:8px;text-decoration:none;color:inherit}}
  .membership-row-top{{display:flex;justify-content:space-between;align-items:center;margin-bottom:6px}}
  .membership-row-id{{font-family:'Playfair Display',serif;font-size:1rem;font-weight:600;color:#c9a05a}}
  .membership-row-status{{font-size:0.6rem;text-transform:uppercase;letter-spacing:.05em;padding:3px 9px;border-radius:20px;border:1px solid}}
  .membership-row-meta{{font-size:0.74rem;color:rgba(242,237,226,.55);margin-bottom:8px}}
  .membership-row-progress-bar{{height:5px;background:rgba(255,255,255,.08);border-radius:4px;overflow:hidden;margin-bottom:6px}}
  .membership-row-progress-fill{{height:100%;background:#c9a05a}}
  .membership-row-count{{font-size:0.72rem;color:rgba(242,237,226,.6)}}
  .new-membership-link{{display:block;text-align:center;background:rgba(201,160,90,.08);border:1px dashed rgba(201,160,90,.4);color:#c9a05a;border-radius:10px;padding:11px;font-size:0.8rem;text-decoration:none;margin-bottom:24px}}
  .upload-toast{{position:fixed;bottom:24px;left:50%;transform:translateX(-50%) translateY(20px);background:#151310;border:1px solid rgba(201,160,90,.4);color:#f2ede2;padding:10px 20px;border-radius:20px;font-size:0.8rem;opacity:0;pointer-events:none;transition:all .25s;z-index:999}}
  .upload-toast.show{{opacity:1;transform:translateX(-50%) translateY(0)}}
</style>
</head>
<body>
<div class="wrap">
  <a href="/admin/clients?pass=anza1985" class="back-link">← Все клиенты</a>
  <div class="eyebrow">R&J Grooming · Клиент</div>
  <h1>{name or '—'}</h1>
  <div class="pets-sub">{pets_str}</div>

  <div class="actions">
    <a class="abtn call" href="tel:{display_phone}">📞 Позвонить</a>
    <a class="abtn wa" href="https://wa.me/{wa_digits}" target="_blank" rel="noopener">💬 WhatsApp</a>
  </div>

  <div class="stats-mini">
    <div class="stat"><div class="n">{len(visits)}</div><div class="l">визитов</div></div>
    <div class="stat"><div class="n">{total:.0f}€</div><div class="l">всего</div></div>
    <div class="stat"><div class="n" style="font-size:1.05rem">{last_visit_str}</div><div class="l">последний</div></div>
  </div>

  <div class="card">
    <div class="card-title-row">
      <div class="card-title">Контакты</div>
      <button class="edit-btn" onclick="toggleEditContacts()" id="editBtnLabel">✏️ Изменить</button>
    </div>
    <div id="contactsView">
      <div class="contact-line"><span class="contact-label">Имя</span><span class="contact-val">{name or '—'}</span></div>
      <div class="contact-line"><span class="contact-label">Телефон</span><a class="contact-val link" href="tel:{display_phone}">{display_phone}</a></div>
      <div class="contact-line"><span class="contact-label">Email</span>{email_html}</div>
      <div class="contact-line"><span class="contact-label">Instagram</span>{ig_html}</div>
      <div class="contact-line"><span class="contact-label">Первый визит</span><span class="contact-val">{first_visit_str}</span></div>
      <div class="contact-line" style="border-bottom:none;flex-direction:column;align-items:flex-start;gap:4px">
        <span class="contact-label">Комментарий</span>
        <span class="contact-val" style="white-space:pre-wrap">{comment or 'нет комментария'}</span>
      </div>
    </div>
    <div id="contactsEdit" style="display:none">
      <div class="edit-field">
        <label>Имя</label>
        <input type="text" id="editName" value="{name}" placeholder="Имя клиента">
      </div>
      <div class="edit-field">
        <label>Телефон</label>
        <input type="text" id="editPhone" value="{display_phone}" placeholder="+372...">
      </div>
      <div class="edit-field">
        <label>Email</label>
        <input type="email" id="editEmail" value="{email}" placeholder="client@example.com">
      </div>
      <div class="edit-field">
        <label>Instagram (без @)</label>
        <input type="text" id="editInstagram" value="{instagram}" placeholder="username">
      </div>
      <div class="edit-field">
        <label>Комментарий</label>
        <textarea id="editComment" placeholder="Заметки о клиенте...">{comment}</textarea>
      </div>
      <div class="edit-actions">
        <button class="edit-save" onclick="saveContacts('{_urlp.quote(phone)}')">Сохранить</button>
        <button class="edit-cancel" onclick="toggleEditContacts()">Отмена</button>
      </div>
    </div>
  </div>

  <div class="list-label">Абонементы</div>
  <div class="memberships-list">{memberships_html}</div>
  <a class="new-membership-link" href="{new_membership_link}">+ Новый абонемент</a>

  <div class="list-label">Анкеты питомцев</div>
  <div class="pet-photos-grid">{pet_photos_html}</div>

  <div class="list-label">История визитов ({len(visits)})</div>
  {error_html}
  {visits_html}
</div>
<div id="uploadToast" class="upload-toast"></div>
<script>
function toggleEditContacts(){{
  var view = document.getElementById('contactsView');
  var edit = document.getElementById('contactsEdit');
  var isEditing = edit.style.display !== 'none';
  view.style.display = isEditing ? '' : 'none';
  edit.style.display = isEditing ? 'none' : '';
}}
function saveContacts(phoneEncoded){{
  var name = document.getElementById('editName').value.trim();
  var phoneOverride = document.getElementById('editPhone').value.trim();
  var email = document.getElementById('editEmail').value.trim();
  var instagram = document.getElementById('editInstagram').value.trim();
  var comment = document.getElementById('editComment').value.trim();
  var toast = document.getElementById('uploadToast');
  toast.textContent = 'Сохраняю...';
  toast.classList.add('show');
  fetch('/api/save-client-data', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{
      phone: decodeURIComponent(phoneEncoded),
      name: name, phone_override: phoneOverride,
      email: email, instagram: instagram, comment: comment
    }})
  }})
  .then(function(r){{ return r.json(); }})
  .then(function(data){{
    if(data.success){{
      toast.textContent = 'Сохранено ✓';
      setTimeout(function(){{ location.reload(); }}, 600);
    }} else {{
      toast.textContent = 'Ошибка: ' + (data.error || 'не удалось сохранить');
      setTimeout(function(){{ toast.classList.remove('show'); }}, 2500);
    }}
  }})
  .catch(function(){{
    toast.textContent = 'Ошибка сохранения';
    setTimeout(function(){{ toast.classList.remove('show'); }}, 2500);
  }});
}}
function uploadPetPhoto(input, key, phone, pet){{
  var file = input.files[0];
  if(!file) return;
  var toast = document.getElementById('uploadToast');
  toast.textContent = 'Загружаю...';
  toast.classList.add('show');
  var fd = new FormData();
  fd.append('phone', decodeURIComponent(phone));
  fd.append('pet', decodeURIComponent(pet));
  fd.append('photo', file);
  fetch('/api/upload-pet-photo', {{method: 'POST', body: fd}})
    .then(function(r){{ return r.json(); }})
    .then(function(data){{
      if(data.success){{
        toast.textContent = 'Фото сохранено ✓';
        var el = document.getElementById('img-' + key);
        if(el){{
          var img = document.createElement('img');
          img.src = data.url + '?v=' + Date.now();
          img.className = 'pet-photo-img';
          img.id = 'img-' + key;
          el.replaceWith(img);
        }}
      }} else {{
        toast.textContent = 'Ошибка: ' + (data.error || 'не удалось загрузить');
      }}
      setTimeout(function(){{ toast.classList.remove('show'); }}, 2500);
    }})
    .catch(function(err){{
      toast.textContent = 'Ошибка загрузки';
      setTimeout(function(){{ toast.classList.remove('show'); }}, 2500);
    }});
}}
</script>
</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/api/find-duplicates")
def api_find_duplicates():
    try:
        today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
        bookings = _fetch_full_history_bookings()
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

    by_raw_phone = {}
    for b in bookings:
        raw_phone = (b.get("clientPhone") or "").strip()
        if not raw_phone:
            continue
        norm = _normalize_phone(raw_phone)
        entry = by_raw_phone.setdefault(raw_phone, {"norm": norm, "names": set(), "pets": set(), "count": 0})
        if b.get("clientName"):
            entry["names"].add(b.get("clientName"))
        if b.get("petName"):
            entry["pets"].add(b.get("petName"))
        entry["count"] += 1

    # Группировка по нормализованному телефону — ловим разные форматы одного номера
    by_norm = {}
    for raw_phone, info in by_raw_phone.items():
        by_norm.setdefault(info["norm"], []).append({"raw": raw_phone, **info})

    phone_dupes = []
    for norm, variants in by_norm.items():
        if len(variants) > 1:
            phone_dupes.append({
                "normalized_phone": norm,
                "variants": [
                    {"raw_phone": v["raw"], "names": sorted(v["names"]), "pets": sorted(v["pets"]), "bookings_count": v["count"]}
                    for v in variants
                ]
            })

    # Совпадение имени при разных телефонах — возможный дубль под другим номером
    by_name = {}
    for raw_phone, info in by_raw_phone.items():
        for name in info["names"]:
            name_key = name.strip().lower()
            if not name_key:
                continue
            by_name.setdefault(name_key, set()).add(info["norm"])

    name_dupes = []
    for name_key, phones in by_name.items():
        if len(phones) > 1:
            name_dupes.append({"name": name_key, "phones": sorted(phones)})

    return jsonify({
        "success": True,
        "phone_format_duplicates": phone_dupes,
        "same_name_different_phone": name_dupes,
        "phone_dupes_count": len(phone_dupes),
        "name_dupes_count": len(name_dupes)
    })

@app.route("/admin/memberships")
def admin_memberships_page():
    import urllib.parse as _urlp
    from datetime import datetime, timedelta
    default_purchase = datetime.now().strftime("%d.%m.%Y")
    default_expiry = (datetime.now() + timedelta(days=182)).strftime("%d.%m.%Y")
    prefill_name = request.args.get("client_name", "")
    prefill_phone = request.args.get("client_phone", "")
    prefill_pet = request.args.get("pet_name", "")
    prefill_pet_type = request.args.get("pet_type", "")
    prefill_email = request.args.get("client_email", "")
    memberships = _load_memberships()
    items = sorted(memberships.values(), key=lambda m: m.get("id", ""), reverse=True)

    def card_html(m):
        pct = round((m["used_visits"] / m["total_visits"]) * 100) if m["total_visits"] else 0
        status_label = "Завершён" if m["status"] == "completed" else "Активен"
        status_color = "#8a8578" if m["status"] == "completed" else "#4ade80"
        actions = ""
        if m["status"] != "completed" and m["used_visits"] < m["total_visits"]:
            actions += f'<button class="mem-btn mem-btn-mark" onclick="markVisit(\'{m["id"]}\')">Отметить визит</button>'
        if m["used_visits"] > 0:
            actions += f'<button class="mem-btn mem-btn-undo" onclick="undoVisit(\'{m["id"]}\')">Отменить отметку</button>'

        pricing_html = ""
        single_price = m.get("single_visit_price") or 0
        total_price = m.get("total_price") or 0
        total_visits = m.get("total_visits") or 0
        discount_pct = m.get("discount_percent") or 0
        if single_price and total_price and total_visits:
            per_visit = m.get("per_visit_price") or (total_price / total_visits)
            without_total = single_price * total_visits
            total_saving = without_total - total_price
            service_line = f'{m.get("service_type","")} · ' if m.get("service_type") else ""
            discount_line = f' · скидка {discount_pct:.0f}%' if discount_pct else ""
            pricing_html = f"""
          <div class="mem-pricing">
            <div class="mem-pricing-row">{service_line}разово {single_price:.0f}€ → по абонементу {per_visit:.1f}€/визит{discount_line}</div>
            <div class="mem-pricing-row">Без абонемента: {without_total:.0f}€ · С абонементом: {total_price:.0f}€</div>
            <div class="mem-pricing-row mem-pricing-highlight">Выгода клиента: {total_saving:.0f}€</div>
          </div>"""

        wa_digits = _normalize_phone(m.get("client_phone", "")).lstrip("+")
        wa_msg = _urlp.quote(f"Здравствуйте, {m.get('client_name','')}! Ваш абонемент {m['id']} ({m.get('plan_name','')}) готов: https://rjgrooming.up.railway.app/membership/{m['id']}")
        has_email = bool((m.get("client_email") or "").strip())

        history = m.get("visit_history") or []
        if history:
            hist_rows = "".join(
                f'<div class="mem-hist-row"><span class="mem-hist-date">{h.get("date","")}</span><span class="mem-hist-note">{h.get("note","уход")}</span></div>'
                for h in reversed(history)
            )
            history_block = f'<div class="mem-hist"><div class="mem-hist-label">История посещений</div>{hist_rows}</div>'
        else:
            history_block = ''

        return f"""
        <div class="mem-card" id="card-{m['id']}">
          <div class="mem-top">
            <div>
              <div class="mem-id">{m['id']}</div>
              <div class="mem-client">{m['client_name']} · {m['pet_name']}</div>
            </div>
            <div class="mem-status" style="color:{status_color};border-color:{status_color}55">{status_label}</div>
          </div>
          <div class="mem-progress-row">
            <div class="mem-progress-bar"><div class="mem-progress-fill" style="width:{pct}%"></div></div>
            <span class="mem-progress-text">{m['used_visits']}/{m['total_visits']}</span>
          </div>
          <div class="mem-meta">{m.get('plan_name','')} · до {m.get('expiry_date','—')}</div>
          {pricing_html}
          {history_block}
          <div class="mem-actions">
            {actions}
            <a class="mem-btn mem-btn-link" href="/membership/{m['id']}" target="_blank">Открыть карточку</a>
            <a class="mem-btn mem-btn-wa" href="https://wa.me/{wa_digits}?text={wa_msg}" target="_blank" rel="noopener">WhatsApp</a>
            <button class="mem-btn mem-btn-mail" onclick="sendMembershipEmail('{m['id']}')" {'disabled title="У клиента нет email"' if not has_email else ''}>Отправить на почту</button>
            <button class="mem-btn mem-btn-delete" onclick="deleteMembership('{m['id']}')">Удалить</button>
          </div>
        </div>"""

    cards_html = "".join(card_html(m) for m in items) if items else '<div class="empty">Абонементов пока нет</div>'

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>R&J Grooming — Абонементы</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,500;0,600;0,700;1,500&family=Montserrat:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a09;color:#f2ede2;font-family:'Montserrat',sans-serif;padding:36px 20px 80px}}
  .wrap{{max-width:640px;margin:0 auto}}
  .back-link{{display:inline-block;font-size:0.74rem;color:rgba(201,160,90,.75);text-decoration:none;margin-bottom:16px}}
  .eyebrow{{font-size:0.68rem;letter-spacing:.3em;text-transform:uppercase;color:rgba(242,237,226,.4);margin-bottom:10px}}
  h1{{font-family:'Playfair Display',serif;font-weight:600;font-size:2.1rem;margin-bottom:24px}}
  .form-card{{background:#141310;border:1px solid rgba(201,160,90,.25);border-radius:14px;padding:20px;margin-bottom:28px}}
  .form-title{{font-size:0.66rem;letter-spacing:.15em;text-transform:uppercase;color:#c9a05a;margin-bottom:14px}}
  .form-row{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px}}
  .form-field{{margin-bottom:10px}}
  .form-field label{{display:block;font-size:0.72rem;color:rgba(242,237,226,.5);margin-bottom:5px}}
  .form-field input{{width:100%;background:#0e0d0b;border:1px solid rgba(201,160,90,.3);border-radius:8px;padding:10px 12px;color:#f2ede2;font-family:'Montserrat',sans-serif;font-size:0.85rem}}
  .form-field input:focus{{outline:none;border-color:#c9a05a}}
  .form-field select{{width:100%;background:#0e0d0b;border:1px solid rgba(201,160,90,.3);border-radius:8px;padding:10px 12px;color:#f2ede2;font-family:'Montserrat',sans-serif;font-size:0.85rem}}
  .form-field input[readonly]{{color:rgba(242,237,226,.5);cursor:not-allowed}}
  .breed-search-wrap{{position:relative}}
  .breed-drop{{display:none;position:absolute;top:100%;left:0;right:0;background:#141310;border:1px solid rgba(201,160,90,.35);border-radius:8px;margin-top:4px;max-height:220px;overflow-y:auto;z-index:20}}
  .breed-drop.open{{display:block}}
  .breed-drop-item{{padding:9px 12px;font-size:0.82rem;color:#f2ede2;cursor:pointer;border-bottom:1px solid rgba(255,255,255,.05)}}
  .breed-drop-item:last-child{{border-bottom:none}}
  .breed-drop-item:hover{{background:rgba(201,160,90,.1)}}
  .breed-drop-item mark{{background:none;color:#c9a05a;font-weight:600}}
  .breed-drop-empty{{padding:12px;font-size:0.78rem;color:rgba(242,237,226,.4);text-align:center}}
  .form-field select:focus{{outline:none;border-color:#c9a05a}}
  .savings-preview{{background:rgba(74,222,128,.06);border:1px solid rgba(74,222,128,.25);border-radius:10px;padding:12px 14px;margin-bottom:14px}}
  .savings-row{{display:flex;justify-content:space-between;font-size:0.78rem;color:rgba(242,237,226,.7);padding:4px 0}}
  .savings-row span:last-child{{font-weight:600;color:#4ade80}}
  .savings-total{{border-top:1px solid rgba(74,222,128,.2);margin-top:4px;padding-top:8px}}
  .savings-total span{{font-size:0.85rem}}
  .create-btn{{width:100%;background:#c9a05a;color:#0a0a09;border:none;border-radius:8px;padding:12px;font-weight:600;font-size:0.88rem;cursor:pointer;margin-top:6px}}
  .list-label{{font-size:0.68rem;letter-spacing:.2em;text-transform:uppercase;color:#c9a05a;margin-bottom:14px}}
  .mem-card{{background:#131210;border:1px solid rgba(255,255,255,.06);border-radius:12px;padding:16px 18px;margin-bottom:10px}}
  .mem-top{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:12px}}
  .mem-id{{font-family:'Playfair Display',serif;font-size:1.1rem;font-weight:600;color:#c9a05a}}
  .mem-client{{font-size:0.85rem;color:rgba(242,237,226,.75);margin-top:2px}}
  .mem-status{{font-size:0.62rem;text-transform:uppercase;letter-spacing:.05em;padding:4px 10px;border-radius:20px;border:1px solid;white-space:nowrap}}
  .mem-progress-row{{display:flex;align-items:center;gap:10px;margin-bottom:8px}}
  .mem-progress-bar{{flex:1;height:6px;background:rgba(255,255,255,.08);border-radius:4px;overflow:hidden}}
  .mem-progress-fill{{height:100%;background:#c9a05a}}
  .mem-progress-text{{font-size:0.75rem;color:rgba(242,237,226,.7);font-weight:600}}
  .mem-meta{{font-size:0.72rem;color:rgba(242,237,226,.5);margin-bottom:12px}}
  .mem-pricing{{background:rgba(74,222,128,.05);border:1px solid rgba(74,222,128,.2);border-radius:8px;padding:8px 12px;margin-bottom:12px}}
  .mem-pricing-row{{font-size:0.72rem;color:rgba(242,237,226,.6)}}
  .mem-pricing-highlight{{color:#4ade80;font-weight:600;margin-top:2px}}
  .mem-hist{{margin-top:10px;padding-top:10px;border-top:1px solid rgba(255,255,255,.06)}}
  .mem-hist-label{{font-size:0.6rem;letter-spacing:.1em;text-transform:uppercase;color:rgba(242,237,226,.4);margin-bottom:6px}}
  .mem-hist-row{{display:flex;justify-content:space-between;font-size:0.75rem;padding:4px 0;color:rgba(242,237,226,.75)}}
  .mem-hist-date{{color:rgba(242,237,226,.5)}}
  .mem-actions{{display:flex;gap:8px;flex-wrap:wrap}}
  .mem-btn{{font-size:0.72rem;padding:8px 12px;border-radius:8px;border:1px solid;cursor:pointer;text-decoration:none;font-family:'Montserrat',sans-serif}}
  .mem-btn-mark{{background:rgba(74,222,128,.1);border-color:rgba(74,222,128,.4);color:#4ade80}}
  .mem-btn-undo{{background:none;border-color:rgba(224,82,74,.4);color:#e0524a}}
  .mem-btn-link{{background:rgba(201,160,90,.1);border-color:rgba(201,160,90,.4);color:#c9a05a}}
  .mem-btn-wa{{background:rgba(74,222,128,.08);border-color:rgba(74,222,128,.3);color:#4ade80}}
  .mem-btn-mail{{background:rgba(120,170,230,.1);border-color:rgba(120,170,230,.35);color:#78aae6}}
  .mem-btn-mail:disabled{{opacity:.35;cursor:not-allowed}}
  .mem-btn-delete{{background:none;border-color:rgba(224,82,74,.25);color:rgba(224,82,74,.6)}}
  .empty{{text-align:center;padding:40px 0;color:rgba(242,237,226,.4);font-size:0.85rem}}
  .toast{{position:fixed;bottom:24px;left:50%;transform:translateX(-50%) translateY(20px);background:#151310;border:1px solid rgba(201,160,90,.4);color:#f2ede2;padding:10px 20px;border-radius:20px;font-size:0.8rem;opacity:0;pointer-events:none;transition:all .25s;z-index:999}}
  .toast.show{{opacity:1;transform:translateX(-50%) translateY(0)}}
</style>
</head>
<body>
<div class="wrap">
  <a href="/admin" class="back-link">← Админ-панель</a>
  <div class="eyebrow">R&J Grooming · Абонементы</div>
  <h1>Абонементы</h1>

  <div class="form-card">
    <div class="form-title">Новый абонемент</div>
    <div class="form-row">
      <div class="form-field"><label>Имя клиента</label><input type="text" id="fClientName" value="{prefill_name}"></div>
      <div class="form-field"><label>Телефон</label><input type="text" id="fClientPhone" value="{prefill_phone}"></div>
    </div>
    <div class="form-row">
      <div class="form-field"><label>Кличка питомца</label><input type="text" id="fPetName" value="{prefill_pet}"></div>
      <div class="form-field"><label>Email</label><input type="email" id="fClientEmail" value="{prefill_email}" placeholder="client@example.com"></div>
    </div>
    <div class="form-row">
      <div class="form-field">
        <label>Порода</label>
        <div class="breed-search-wrap">
          <input type="text" id="fBreedSearch" placeholder="Начните вводить породу..." autocomplete="off" oninput="onBreedSearchInput()">
          <div class="breed-drop" id="breedDrop"></div>
        </div>
      </div>
      <div class="form-field">
        <label>Тип услуги</label>
        <select id="fServiceType" onchange="onServiceChange()">
          <option value="">— сначала выбери породу —</option>
        </select>
      </div>
    </div>
    <div class="form-row">
      <div class="form-field"><label>Разовая цена, €</label><input type="number" id="fSingleVisitPrice" min="0" step="0.5" placeholder="—" oninput="updateSavingsPreview()"></div>
      <div class="form-field"></div>
    </div>
    <div class="form-row">
      <div class="form-field"><label>Название абонемента</label><input type="text" id="fPlanName" placeholder="сформируется автоматически" readonly></div>
      <div class="form-field"><label>Кол-во посещений</label><input type="number" id="fTotalVisits" min="1" value="5" oninput="updatePlanName();updateSavingsPreview()"></div>
    </div>
    <div class="form-field">
      <label>Выгода в процентах от обычного посещения, %</label>
      <input type="number" id="fDiscountPercent" min="0" max="100" step="1" placeholder="0" oninput="updateSavingsPreview()">
    </div>
    <div class="form-row">
      <div class="form-field"><label>Дата покупки</label><input type="text" id="fPurchaseDate" placeholder="24.08.2026" value="{default_purchase}"></div>
      <div class="form-field"><label>Действителен до</label><input type="text" id="fExpiryDate" placeholder="24.02.2027" value="{default_expiry}"></div>
    </div>
    <div class="savings-preview" id="savingsPreview" style="display:none">
      <div class="savings-row"><span>Цена за визит по абонементу</span><span id="spPerVisit">—</span></div>
      <div class="savings-row"><span>Стоимость абонемента</span><span id="spTotalPrice">—</span></div>
      <div class="savings-row"><span>Без абонемента за визиты</span><span id="spWithoutTotal">—</span></div>
      <div class="savings-row savings-total"><span>Общая выгода клиента</span><span id="spTotalSaving">—</span></div>
    </div>
    <button class="create-btn" onclick="createMembership()">Создать абонемент</button>
  </div>

  <div class="list-label" id="memListLabel">Все абонементы ({len(items)})</div>
  {cards_html}
</div>
<div id="toast" class="toast"></div>
<script>
function showToast(msg){{
  var t = document.getElementById('toast');
  t.textContent = msg;
  t.classList.add('show');
  setTimeout(function(){{ t.classList.remove('show'); }}, 2500);
}}

var breedData = [];
var selectedBreed = null;
var PREFILL_BREED = {json.dumps(prefill_pet_type)};

function loadBreeds(){{
  fetch('/api/breed-prices').then(function(r){{ return r.json(); }}).then(function(res){{
    if(res.success){{
      breedData = res.breeds;
      if(PREFILL_BREED){{
        var match = breedData.find(function(b){{ return b.breed === PREFILL_BREED; }});
        if(!match){{
          match = breedData.find(function(b){{ return b.breed.indexOf(PREFILL_BREED) === 0 || PREFILL_BREED.indexOf(b.breed) === 0; }});
        }}
        if(match){{ pickBreed(match); }}
        else {{ document.getElementById('fBreedSearch').value = PREFILL_BREED; }}
      }}
    }}
  }}).catch(function(){{}});
}}
loadBreeds();

function onBreedSearchInput(){{
  var q = document.getElementById('fBreedSearch').value.trim();
  var drop = document.getElementById('breedDrop');
  if(!q){{ drop.classList.remove('open'); drop.innerHTML=''; return; }}
  var qLower = q.toLowerCase();
  var res = breedData.filter(function(b){{ return b.breed.toLowerCase().indexOf(qLower) !== -1; }}).slice(0, 30);
  drop.innerHTML = '';
  if(!res.length){{
    drop.innerHTML = '<div class="breed-drop-empty">Порода не найдена</div>';
  }} else {{
    res.forEach(function(b){{
      var idx = b.breed.toLowerCase().indexOf(qLower);
      var item = document.createElement('div');
      item.className = 'breed-drop-item';
      item.innerHTML = b.breed.substring(0, idx) + '<mark>' + b.breed.substring(idx, idx + q.length) + '</mark>' + b.breed.substring(idx + q.length);
      item.onclick = function(){{ pickBreed(b); }};
      drop.appendChild(item);
    }});
  }}
  drop.classList.add('open');
}}

function pickBreed(b){{
  selectedBreed = b;
  document.getElementById('fBreedSearch').value = b.breed;
  document.getElementById('breedDrop').classList.remove('open');
  document.getElementById('breedDrop').innerHTML = '';

  var serviceSel = document.getElementById('fServiceType');
  serviceSel.innerHTML = '<option value="">— выбрать услугу —</option>' +
    Object.keys(b.services).map(function(s){{ return '<option value="' + s + '">' + s + ' — ' + b.services[s] + '€</option>'; }}).join('');
  document.getElementById('fSingleVisitPrice').value = '';
  updateSavingsPreview();
}}

document.addEventListener('click', function(e){{
  if(!e.target.closest('.breed-search-wrap')) document.getElementById('breedDrop').classList.remove('open');
}});

function onServiceChange(){{
  var serviceName = document.getElementById('fServiceType').value;
  if(selectedBreed && selectedBreed.services[serviceName] !== undefined){{
    document.getElementById('fSingleVisitPrice').value = selectedBreed.services[serviceName];
  }}
  updatePlanName();
  updateSavingsPreview();
}}

function updatePlanName(){{
  var visits = document.getElementById('fTotalVisits').value;
  var service = document.getElementById('fServiceType').value;
  document.getElementById('fPlanName').value = (visits && service) ? (visits + ' посещений — ' + service) : '';
}}

function createMembership(){{
  var data = {{
    client_name: document.getElementById('fClientName').value.trim(),
    client_phone: document.getElementById('fClientPhone').value.trim(),
    client_email: document.getElementById('fClientEmail').value.trim(),
    pet_name: document.getElementById('fPetName').value.trim(),
    pet_type: selectedBreed ? selectedBreed.breed : document.getElementById('fBreedSearch').value.trim(),
    plan_name: document.getElementById('fPlanName').value.trim(),
    total_visits: document.getElementById('fTotalVisits').value,
    purchase_date: document.getElementById('fPurchaseDate').value.trim(),
    expiry_date: document.getElementById('fExpiryDate').value.trim(),
    service_type: document.getElementById('fServiceType').value,
    single_visit_price: document.getElementById('fSingleVisitPrice').value,
    discount_percent: document.getElementById('fDiscountPercent').value
  }};
  if(!data.client_name || !data.pet_name || !data.total_visits){{
    showToast('Заполни имя клиента, кличку и число посещений');
    return;
  }}
  fetch('/api/membership/create', {{
    method: 'POST', headers: {{'Content-Type':'application/json'}}, body: JSON.stringify(data)
  }}).then(function(r){{ return r.json(); }}).then(function(res){{
    if(res.success){{
      showToast('Создан: ' + res.id);
      setTimeout(function(){{ location.reload(); }}, 700);
    }} else {{
      showToast('Ошибка: ' + (res.error||'не удалось создать'));
    }}
  }}).catch(function(){{ showToast('Ошибка сети'); }});
}}

function updateSavingsPreview(){{
  var totalVisits = parseFloat(document.getElementById('fTotalVisits').value) || 0;
  var singlePrice = parseFloat(document.getElementById('fSingleVisitPrice').value) || 0;
  var discountPct = parseFloat(document.getElementById('fDiscountPercent').value) || 0;
  var block = document.getElementById('savingsPreview');
  if(!totalVisits || !singlePrice){{
    block.style.display = 'none';
    return;
  }}
  var perVisit = singlePrice * (1 - discountPct / 100);
  var totalPrice = perVisit * totalVisits;
  var withoutTotal = singlePrice * totalVisits;
  var totalSaving = withoutTotal - totalPrice;
  document.getElementById('spPerVisit').textContent = perVisit.toFixed(2) + ' €';
  document.getElementById('spTotalPrice').textContent = totalPrice.toFixed(2) + ' €';
  document.getElementById('spWithoutTotal').textContent = totalVisits + ' × ' + singlePrice.toFixed(2) + ' € = ' + withoutTotal.toFixed(2) + ' €';
  document.getElementById('spTotalSaving').textContent = totalSaving.toFixed(2) + ' €';
  block.style.display = 'block';
}}

function markVisit(id){{
  fetch('/api/membership/mark-visit', {{
    method:'POST', headers:{{'Content-Type':'application/json'}}, body: JSON.stringify({{id:id}})
  }}).then(function(r){{return r.json();}}).then(function(res){{
    if(res.success){{ showToast('Визит отмечен ✓'); setTimeout(function(){{location.reload();}}, 600); }}
    else{{ showToast('Ошибка: ' + (res.error||'')); }}
  }}).catch(function(){{ showToast('Ошибка сети'); }});
}}

function undoVisit(id){{
  fetch('/api/membership/undo-visit', {{
    method:'POST', headers:{{'Content-Type':'application/json'}}, body: JSON.stringify({{id:id}})
  }}).then(function(r){{return r.json();}}).then(function(res){{
    if(res.success){{ showToast('Отметка отменена'); setTimeout(function(){{location.reload();}}, 600); }}
    else{{ showToast('Ошибка: ' + (res.error||'')); }}
  }}).catch(function(){{ showToast('Ошибка сети'); }});
}}

function sendMembershipEmail(id){{
  var btn = event && event.target;
  if(btn){{ btn.disabled = true; var oldText = btn.textContent; btn.textContent = 'Отправка...'; }}
  fetch('/api/membership/send-email', {{
    method:'POST', headers:{{'Content-Type':'application/json'}}, body: JSON.stringify({{id:id}})
  }}).then(function(r){{return r.json();}}).then(function(res){{
    if(res.success){{ showToast('Письмо отправлено'); }}
    else {{ showToast('Ошибка: ' + (res.error||'')); }}
  }}).catch(function(){{
    showToast('Ошибка сети');
  }}).finally(function(){{
    if(btn){{ btn.disabled = false; btn.textContent = oldText; }}
  }});
}}

function deleteMembership(id){{
  if(!confirm('Удалить абонемент ' + id + '? Это действие необратимо.')) return;
  var card = document.getElementById('card-' + id);
  if(card){{ card.style.opacity = '0.4'; card.style.pointerEvents = 'none'; }}
  fetch('/api/membership/delete', {{
    method:'POST', headers:{{'Content-Type':'application/json'}}, body: JSON.stringify({{id:id}})
  }}).then(function(r){{return r.json();}}).then(function(res){{
    if(res.success){{
      showToast('Удалён');
      if(card){{ card.remove(); }}
      var label = document.getElementById('memListLabel');
      if(label){{
        var n = document.querySelectorAll('.mem-card').length;
        label.textContent = 'Все абонементы (' + n + ')';
      }}
    }} else {{
      if(card){{ card.style.opacity = '1'; card.style.pointerEvents = 'auto'; }}
      showToast('Ошибка: ' + (res.error||''));
    }}
  }}).catch(function(){{
    if(card){{ card.style.opacity = '1'; card.style.pointerEvents = 'auto'; }}
    showToast('Ошибка сети');
  }});
}}
</script>
</body>
</html>"""
    return html, 200, {
        "Content-Type": "text/html; charset=utf-8",
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache"
    }

@app.route("/admin/clients")
def admin_clients_page():
    import urllib.parse as _urlp

    error = None
    clients = []
    try:
        today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
        bookings = _fetch_full_history_bookings()

        def parse_date(s):
            try:
                return datetime.strptime(s, "%d.%m.%Y").date()
            except Exception:
                return None

        by_phone = {}
        for b in bookings:
            phone = (b.get("clientPhone") or "").strip()
            if not phone:
                continue
            d = parse_date(b.get("date", ""))
            entry = by_phone.setdefault(phone, {
                "name": b.get("clientName", ""), "phone": phone, "pets": {},
                "visits": 0, "total": 0.0, "last_date": None, "last_master": ""
            })
            if b.get("clientName"):
                entry["name"] = b.get("clientName")
            pet = b.get("petName", "")
            breed = b.get("breed", "")
            if pet:
                entry["pets"][pet] = breed
            entry["visits"] += 1
            try:
                entry["total"] += float(b.get("price") or 0)
            except Exception:
                pass
            if d and (not entry["last_date"] or d > entry["last_date"]):
                entry["last_date"] = d
                entry["last_master"] = b.get("master", "")

        all_client_data = _load_all_client_data()
        for phone, entry in by_phone.items():
            saved = all_client_data.get(_client_data_key(phone), {})
            if saved.get("name"):
                entry["name"] = saved["name"]
            entry["display_phone"] = _normalize_phone(saved.get("phone_override") or phone)

        clients = sorted(by_phone.values(), key=lambda c: c["last_date"] or datetime.min.date(), reverse=True)
    except Exception as e:
        error = str(e)

    def client_html(c):
        pets_str = ", ".join(f"{p} ({b})" if b else p for p, b in c["pets"].items()) or "—"
        last_str = c["last_date"].strftime("%d.%m.%Y") if c["last_date"] else "—"
        phone_encoded = _urlp.quote(c["phone"] or "")
        return f"""
        <a class="row" href="/admin/client?phone={phone_encoded}&pass=anza1985">
          <div class="row-top">
            <div class="who">
              <span class="name">{c['name'] or '—'}</span>
              <span class="pet">{pets_str}</span>
            </div>
            <div class="visits-badge">{c['visits']} {'визит' if c['visits']==1 else 'визита' if 2<=c['visits']<=4 else 'визитов'}</div>
          </div>
          <div class="row-bottom">
            <span>Последний визит: {last_str} · {c['last_master'] or '—'}</span>
            <span class="days">{c['total']:.0f}€</span>
          </div>
          <div class="row-bottom" style="margin-top:6px">
            <span class="contacts">{c.get('display_phone', c['phone'])}</span>
            <span class="row-arrow">→</span>
          </div>
        </a>"""

    rows_html = "".join(client_html(c) for c in clients) if clients else '<div class="empty">Клиентов не найдено</div>'
    error_html = f'<div class="empty" style="color:#e0824a">Ошибка загрузки данных: {error}</div>' if error else ""
    total_revenue = sum(c["total"] for c in clients)
    total_pets = sum(len(c["pets"]) for c in clients)
    returning_clients = sum(1 for c in clients if c["visits"] >= 2)

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>R&J Grooming — Клиенты</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,500;0,600;0,700;1,500&family=Montserrat:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a09;color:#f2ede2;font-family:'Montserrat',sans-serif;padding:36px 20px 80px}}
  .wrap{{max-width:720px;margin:0 auto}}
  .eyebrow{{font-size:0.68rem;letter-spacing:.3em;text-transform:uppercase;color:rgba(242,237,226,.4);margin-bottom:10px}}
  .back-link{{display:inline-block;font-size:0.74rem;color:rgba(201,160,90,.75);text-decoration:none;margin-bottom:16px}}
  h1{{font-family:'Playfair Display',serif;font-weight:600;font-size:2.1rem;margin-bottom:6px}}
  .sub{{font-size:0.78rem;color:rgba(242,237,226,.5);margin-bottom:20px}}
  .quick-links{{display:flex;gap:10px;margin-bottom:24px}}
  .search-link{{flex:1;display:block;text-align:center;background:rgba(201,160,90,.1);border:1px solid rgba(201,160,90,.4);color:#c9a05a;border-radius:10px;padding:12px;font-size:0.85rem;font-weight:600;text-decoration:none}}
  .stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:32px}}
  .stat{{background:#151310;border:1px solid rgba(201,160,90,.18);border-radius:12px;padding:18px 14px}}
  .stat .n{{font-family:'Playfair Display',serif;font-size:1.9rem;font-weight:600;color:#c9a05a}}
  .stat .l{{font-size:0.66rem;letter-spacing:.08em;text-transform:uppercase;color:rgba(242,237,226,.5);margin-top:4px}}
  .list-label{{font-size:0.68rem;letter-spacing:.2em;text-transform:uppercase;color:#c9a05a;margin-bottom:14px}}
  .row{{display:block;background:#131210;border:1px solid rgba(255,255,255,.06);border-radius:12px;padding:16px 18px;margin-bottom:10px;text-decoration:none;color:inherit;transition:border-color .15s}}
  .row:active{{border-color:rgba(201,160,90,.4)}}
  .row-top{{display:flex;justify-content:space-between;align-items:flex-start;gap:10px;margin-bottom:10px}}
  .who{{display:flex;flex-direction:column}}
  .name{{font-family:'Playfair Display',serif;font-size:1.15rem;font-weight:600}}
  .pet{{font-size:0.76rem;color:rgba(242,237,226,.5);margin-top:2px}}
  .visits-badge{{font-size:0.62rem;letter-spacing:.04em;color:#c9a05a;border:1px solid rgba(201,160,90,.35);background:rgba(201,160,90,.08);padding:5px 10px;border-radius:20px;white-space:nowrap;flex-shrink:0}}
  .row-bottom{{display:flex;justify-content:space-between;align-items:center;font-size:0.74rem;color:rgba(242,237,226,.55)}}
  .row-bottom .days{{font-weight:600;color:#f2ede2}}
  .contacts{{color:rgba(242,237,226,.6)}}
  .row-arrow{{color:rgba(201,160,90,.6)}}
  .wa{{color:#4ade80;text-decoration:none;font-size:0.72rem;font-weight:600;border:1px solid rgba(74,222,128,.35);border-radius:20px;padding:3px 10px}}
  .empty{{text-align:center;padding:40px 0;color:rgba(242,237,226,.4);font-size:0.85rem}}
</style>
</head>
<body>
<div class="wrap">
  <a href="/admin?pass=anza1985" class="back-link">← Админ-панель</a>
  <div class="eyebrow">R&J Grooming · Клиенты</div>
  <h1>Все клиенты</h1>
  <div class="sub">Полная база — вся история бронирований</div>

  <div class="quick-links">
    <a href="/admin/search?pass=anza1985" class="search-link">🔍 Поиск клиента</a>
    <a href="/admin/export-clients?pass=anza1985" class="search-link">⬇️ Excel-выгрузка</a>
  </div>

  <div class="stats" style="grid-template-columns:repeat(2,1fr)">
    <div class="stat"><div class="n">{len(clients)}</div><div class="l">клиентов</div></div>
    <div class="stat"><div class="n">{total_pets}</div><div class="l">питомцев</div></div>
    <div class="stat"><div class="n">{returning_clients}</div><div class="l">повторных (2+ визита)</div></div>
    <div class="stat"><div class="n">{total_revenue:.0f}€</div><div class="l">выручка всего</div></div>
  </div>

  <div class="list-label">Список ({len(clients)})</div>
  {error_html}
  {rows_html}
</div>
</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/admin/dns-records")
def admin_dns_records():
    resend_key = request.args.get("key") or os.environ.get("RESEND_API_KEY")
    if not resend_key:
        return "RESEND_API_KEY не настроен на Railway.", 500

    try:
        r = requests.get(
            "https://api.resend.com/domains",
            headers={"Authorization": f"Bearer {resend_key}"},
            timeout=15
        )
        domains = r.json().get("data", [])
        domain = next((d for d in domains if "rjgrooming" in d.get("name", "")), None)
        if not domain:
            return f"Домен не найден. Ответ API: {r.text}", 404

        r2 = requests.get(
            f"https://api.resend.com/domains/{domain['id']}",
            headers={"Authorization": f"Bearer {resend_key}"},
            timeout=15
        )
        detail = r2.json()
    except Exception as e:
        return f"Ошибка запроса к Resend API: {e}", 500

    records = detail.get("records", [])
    rows_html = ""
    for rec in records:
        rows_html += f"""
        <div class="card">
          <div class="row"><b>Type:</b> {rec.get('type','')}</div>
          <div class="row"><b>Name:</b> <span class="copyable">{rec.get('name','')}</span></div>
          <div class="row"><b>Value:</b></div>
          <textarea readonly onclick="this.select()">{rec.get('value','')}</textarea>
          <div class="row"><b>Priority:</b> {rec.get('priority','—')}</div>
          <div class="row"><b>Status:</b> {rec.get('status','')}</div>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>DNS Records — {domain.get('name','')}</title>
<style>
  body{{background:#0a0a09;color:#f2ede2;font-family:-apple-system,sans-serif;padding:20px;margin:0}}
  h1{{font-size:1.3rem}}
  .status{{color:#e0824a;font-weight:600;margin-bottom:20px}}
  .card{{background:#151310;border:1px solid rgba(201,160,90,.25);border-radius:12px;padding:16px;margin-bottom:14px}}
  .row{{font-size:0.85rem;margin-bottom:6px;color:rgba(242,237,226,.8)}}
  textarea{{width:100%;background:#0a0a09;color:#c9a05a;border:1px solid rgba(201,160,90,.3);border-radius:8px;padding:10px;font-family:monospace;font-size:0.78rem;min-height:70px;margin-bottom:8px;word-break:break-all}}
</style>
</head>
<body>
  <h1>{domain.get('name','')}</h1>
  <div class="status">Статус: {domain.get('status','')}</div>
  {rows_html}
  <p style="font-size:0.75rem;color:rgba(242,237,226,.5)">Нажми на поле Value, выбери всё (Select All) и скопируй — там полное значение без обрезки.</p>
</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/api/membership/create", methods=["POST"])
def api_membership_create():
    body = request.get_json(force=True) or {}
    client_name = (body.get("client_name") or "").strip()
    client_phone = (body.get("client_phone") or "").strip()
    client_email = (body.get("client_email") or "").strip()
    pet_name = (body.get("pet_name") or "").strip()
    pet_type = (body.get("pet_type") or "").strip()
    plan_name = (body.get("plan_name") or "").strip()
    total_visits = body.get("total_visits")
    purchase_date = (body.get("purchase_date") or "").strip()
    expiry_date = (body.get("expiry_date") or "").strip()
    service_type = (body.get("service_type") or "").strip()
    single_visit_price = body.get("single_visit_price")
    discount_percent = body.get("discount_percent")

    if not client_name or not pet_name or not total_visits:
        return jsonify({"success": False, "error": "client_name, pet_name и total_visits обязательны"}), 400
    try:
        total_visits = int(total_visits)
    except Exception:
        return jsonify({"success": False, "error": "total_visits должно быть числом"}), 400

    try:
        single_visit_price = float(single_visit_price) if single_visit_price not in (None, "") else 0.0
    except Exception:
        single_visit_price = 0.0
    try:
        discount_percent = float(discount_percent) if discount_percent not in (None, "") else 0.0
    except Exception:
        discount_percent = 0.0
    discount_percent = max(0.0, min(100.0, discount_percent))

    per_visit_price = round(single_visit_price * (1 - discount_percent / 100), 2)
    total_price = round(per_visit_price * total_visits, 2)

    memberships = _load_memberships()
    mid = _next_membership_id(memberships)
    memberships[mid] = {
        "id": mid,
        "client_name": client_name,
        "client_phone": client_phone,
        "client_email": client_email,
        "pet_name": pet_name,
        "pet_type": pet_type,
        "plan_name": plan_name or f"{total_visits} посещений",
        "service_type": service_type,
        "total_visits": total_visits,
        "used_visits": 0,
        "purchase_date": purchase_date,
        "expiry_date": expiry_date,
        "single_visit_price": single_visit_price,
        "discount_percent": discount_percent,
        "per_visit_price": per_visit_price,
        "total_price": total_price,
        "visit_history": [],
        "status": "active",
        "created_at": datetime.now(_REMINDER_TZ).isoformat() if _REMINDER_TZ else datetime.utcnow().isoformat()
    }
    wallet_serial, wallet_share_url, wallet_google_url = _wallet_create_pass(memberships[mid])
    if wallet_serial:
        memberships[mid]["wallet_serial"] = wallet_serial
        memberships[mid]["wallet_share_url"] = wallet_share_url
        memberships[mid]["wallet_google_url"] = wallet_google_url
    ok = _save_memberships(memberships)
    return jsonify({"success": ok, "id": mid, "wallet_share_url": wallet_share_url})

@app.route("/api/membership/mark-visit", methods=["POST"])
def api_membership_mark_visit():
    body = request.get_json(force=True) or {}
    mid = (body.get("id") or "").strip()
    note = (body.get("note") or "уход").strip()
    memberships = _load_memberships()
    m = memberships.get(mid)
    if not m:
        return jsonify({"success": False, "error": "Абонемент не найден"}), 404
    if m["used_visits"] >= m["total_visits"]:
        return jsonify({"success": False, "error": "Все посещения уже использованы"}), 400

    today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
    m["used_visits"] += 1
    m["visit_history"].append({"date": today.strftime("%d.%m.%Y"), "note": note})
    if m["used_visits"] >= m["total_visits"]:
        m["status"] = "completed"
    _wallet_update_pass(m)
    ok = _save_memberships(memberships)
    return jsonify({"success": ok, "used_visits": m["used_visits"], "status": m["status"]})

@app.route("/api/membership/undo-visit", methods=["POST"])
def api_membership_undo_visit():
    body = request.get_json(force=True) or {}
    mid = (body.get("id") or "").strip()
    memberships = _load_memberships()
    m = memberships.get(mid)
    if not m:
        return jsonify({"success": False, "error": "Абонемент не найден"}), 404
    if m["used_visits"] <= 0:
        return jsonify({"success": False, "error": "Нет отметок для отмены"}), 400

    m["used_visits"] -= 1
    if m["visit_history"]:
        m["visit_history"].pop()
    m["status"] = "active"
    _wallet_update_pass(m)
    ok = _save_memberships(memberships)
    return jsonify({"success": ok, "used_visits": m["used_visits"], "status": m["status"]})

@app.route("/api/membership/delete", methods=["POST"])
def api_membership_delete():
    body = request.get_json(force=True) or {}
    mid = (body.get("id") or "").strip()
    memberships = _load_memberships()
    if mid not in memberships:
        return jsonify({"success": False, "error": "Абонемент не найден"}), 404
    del memberships[mid]
    ok = _save_memberships(memberships)
    return jsonify({"success": ok})

@app.route("/api/membership/send-email", methods=["POST"])
def api_membership_send_email():
    body = request.get_json(force=True) or {}
    mid = (body.get("id") or "").strip()
    memberships = _load_memberships()
    m = memberships.get(mid)
    if not m:
        return jsonify({"success": False, "error": "Абонемент не найден"}), 404
    to_email = (body.get("email") or m.get("client_email") or "").strip()
    if not to_email or "@" not in to_email:
        return jsonify({"success": False, "error": "У клиента не указан email"}), 400
    resend_key = os.environ.get("RESEND_API_KEY")
    if not resend_key:
        return jsonify({"success": False, "error": "RESEND_API_KEY не настроен"}), 500
    card_url = f"https://rjgrooming.up.railway.app/membership/{mid}"
    try:
        r = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
            json={
                "from": "R&J Grooming <booking@rjgrooming.salon>",
                "to": [to_email],
                "subject": f"Ваш абонемент {mid} — R&J Grooming",
                "html": (
                    "<div style='background:#0a0a09;padding:32px 24px;font-family:Arial,sans-serif;color:#f2ede2'>"
                    "<img src='https://rjgrooming.up.railway.app/assets/logo.png' alt='R&amp;J Grooming' style='height:60px;margin-bottom:14px;display:block'>"
                    f"<p style='color:#cfc9ba'>Здравствуйте, {m.get('client_name','')}!</p>"
                    f"<p style='color:#cfc9ba'>Ваш абонемент <b>{m.get('plan_name','')}</b> для {m.get('pet_name','')} готов. "
                    "Откройте карточку по кнопке ниже — там всегда видно, сколько посещений использовано и сколько осталось.</p>"
                    f"<p style='margin:24px 0'><a href='{card_url}' style='background:#e6e1d5;color:#0a0a09;text-decoration:none;padding:12px 22px;border-radius:8px;font-weight:bold;display:inline-block'>Открыть абонемент</a></p>"
                    f"<p style='color:#8a8578;font-size:13px'>{card_url}</p>"
                    "</div>"
                )
            },
            timeout=10
        )
        if r.status_code >= 300:
            return jsonify({"success": False, "error": f"Resend error {r.status_code}: {r.text[:200]}"}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    return jsonify({"success": True})

@app.route("/membership/<mid>")
def public_membership_card(mid):
    try:
        return _render_membership_card(mid)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"MEMBERSHIP CARD ERROR: {tb}", flush=True)
        return f"<pre>{tb}</pre>", 500

def _render_membership_card(mid):
    import urllib.parse as _urlp_q
    memberships = _load_memberships()
    m = memberships.get(mid)
    if not m:
        return "Абонемент не найден. Проверьте ссылку.", 404

    total = m.get("total_visits", 0)
    used = m.get("used_visits", 0)
    remaining = max(0, total - used)
    is_completed = m.get("status") == "completed" or used >= total

    def circle_html(i):
        filled = i <= used
        cls = "visit-circle filled" if filled else "visit-circle"
        mark = "\u2713" if filled else str(i)
        return f'<div class="{cls}">{mark}</div>'

    circles_html = "".join(circle_html(i) for i in range(1, total + 1))

    history = m.get("visit_history", [])
    if history:
        history_html = "".join(
            f'<div class="hist-row"><span class="hist-date">{h.get("date","")}</span><span class="hist-note">{h.get("note","уход")} \u2713</span></div>'
            for h in reversed(history)
        )
    else:
        history_html = '<div class="hist-empty">Визитов пока не было</div>'

    completed_banner = ""
    if is_completed:
        completed_banner = """
        <div class="completed-banner">
          <div class="completed-title">Абонемент завершён \U0001F90D</div>
          <a class="new-membership-btn" href="/app">Приобрести новый абонемент</a>
        </div>"""

    wallet_button_html = ""
    if m.get("wallet_share_url"):
        wallet_button_html = f'<a class="wallet-btn" href="{m["wallet_share_url"]}" target="_blank" rel="noopener">\U0001F4F1 Добавить в Wallet</a>'

    icon_person = '<svg width="18" height="18" viewBox="0 0 24 24" fill="none"><circle cx="12" cy="8" r="4" stroke="rgba(230,225,215,.55)" stroke-width="1.5"/><path d="M4 20c0-4.4 3.6-7 8-7s8 2.6 8 7" stroke="rgba(230,225,215,.55)" stroke-width="1.5" stroke-linecap="round"/></svg>'
    icon_paw = '<svg width="18" height="18" viewBox="0 0 24 24" fill="none"><ellipse cx="12" cy="16" rx="5.5" ry="4.5" stroke="rgba(230,225,215,.55)" stroke-width="1.5"/><circle cx="5.5" cy="9" r="2.1" stroke="rgba(230,225,215,.55)" stroke-width="1.5"/><circle cx="10" cy="5.5" r="2.1" stroke="rgba(230,225,215,.55)" stroke-width="1.5"/><circle cx="14.5" cy="5.5" r="2.1" stroke="rgba(230,225,215,.55)" stroke-width="1.5"/><circle cx="18.5" cy="9" r="2.1" stroke="rgba(230,225,215,.55)" stroke-width="1.5"/></svg>'
    icon_tag = '<svg width="18" height="18" viewBox="0 0 24 24" fill="none"><path d="M11 3H5a2 2 0 00-2 2v6l10.6 10.6a2 2 0 002.8 0l5.2-5.2a2 2 0 000-2.8L11 3z" stroke="rgba(230,225,215,.55)" stroke-width="1.5" stroke-linejoin="round"/><circle cx="7.5" cy="7.5" r="1.4" stroke="rgba(230,225,215,.55)" stroke-width="1.3"/></svg>'
    icon_cal = '<svg width="18" height="18" viewBox="0 0 24 24" fill="none"><rect x="4" y="6" width="16" height="14" rx="2" stroke="rgba(230,225,215,.55)" stroke-width="1.5"/><path d="M8 3.5v4M16 3.5v4M4 10.5h16" stroke="rgba(230,225,215,.55)" stroke-width="1.5" stroke-linecap="round"/></svg>'

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{m.get('id','')} — R&J Grooming Membership</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,500;0,600;0,700;1,500&family=Montserrat:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a09;color:#f2ede2;font-family:'Montserrat',sans-serif;padding:28px 18px 50px;-webkit-font-smoothing:antialiased;min-height:100vh}}
  .wrap{{max-width:420px;margin:0 auto}}

  .flip-wrap{{perspective:1800px;margin-bottom:26px}}
  .flip-card{{position:relative;width:100%;height:88vh;min-height:520px;max-height:640px;transform-style:preserve-3d;transition:transform .75s cubic-bezier(.42,.15,.16,1);cursor:pointer}}
  .flip-card.flipped{{transform:rotateY(180deg)}}
  .face{{position:absolute;inset:0;backface-visibility:hidden;-webkit-backface-visibility:hidden;border-radius:20px;overflow:hidden;background:#0a0a09;border:1px solid rgba(200,195,180,.16)}}

  .face-front{{display:flex;align-items:center;justify-content:center}}
  .brand-frame{{position:absolute;inset:14px;border:1px solid rgba(200,195,180,.4);border-radius:12px}}
  .brand-inner{{position:absolute;inset:26px;border:1px solid rgba(200,195,180,.22);border-radius:6px;display:flex;flex-direction:column;align-items:center;justify-content:center}}
  .brand-logo-img{{width:58%;max-width:230px;height:auto;display:block;filter:drop-shadow(0 0 18px rgba(255,255,255,.06))}}
  .tap-hint{{position:absolute;bottom:52px;left:0;right:0;text-align:center;font-size:0.62rem;letter-spacing:.18em;text-transform:uppercase;color:rgba(205,199,181,.35)}}

  .face-back{{transform:rotateY(180deg);padding:24px 22px;overflow-y:auto;display:flex;flex-direction:column}}
  .back-head{{display:flex;align-items:center;justify-content:space-between;margin-bottom:18px}}
  .flip-back-btn{{background:none;border:1px solid rgba(255,255,255,.18);color:rgba(230,225,215,.7);font-size:0.62rem;letter-spacing:.12em;text-transform:uppercase;padding:6px 12px;border-radius:20px;cursor:pointer;font-family:'Montserrat',sans-serif}}
  .mem-id-label{{font-size:0.62rem;letter-spacing:.15em;text-transform:uppercase;color:rgba(230,225,215,.4)}}
  .mem-id{{font-family:'Playfair Display',serif;font-size:1.5rem;font-weight:600;color:#f2ede2}}

  .info-rows{{border-top:1px solid rgba(255,255,255,.08);margin-bottom:6px}}
  .info-row{{display:flex;align-items:center;gap:12px;padding:12px 0;border-bottom:1px solid rgba(255,255,255,.08)}}
  .info-row .ico{{flex-shrink:0;width:30px;height:30px;border-radius:50%;background:rgba(255,255,255,.05);display:flex;align-items:center;justify-content:center}}
  .info-row .info-label{{font-size:0.6rem;letter-spacing:.1em;text-transform:uppercase;color:rgba(230,225,215,.42);margin-bottom:2px}}
  .info-row .info-val{{font-size:0.92rem;color:#f2ede2;font-weight:500}}
  .info-row.two-col{{display:grid;grid-template-columns:1fr 1fr;gap:12px;align-items:start}}
  .info-row.two-col .info-sub{{display:flex;align-items:center;gap:10px}}

  .counter-card{{background:#131210;border:1px solid rgba(255,255,255,.08);border-radius:14px;padding:18px 16px;margin:16px 0}}
  .counter-label{{font-size:0.6rem;letter-spacing:.2em;text-transform:uppercase;color:rgba(200,195,180,.6);margin-bottom:14px;text-align:center}}
  .visit-circles{{display:flex;flex-wrap:wrap;gap:8px;justify-content:center;margin-bottom:16px}}
  .visit-circle{{width:32px;height:32px;border-radius:50%;border:1.5px solid rgba(200,195,180,.35);display:flex;align-items:center;justify-content:center;font-size:0.78rem;color:rgba(230,225,215,.4);font-family:'Playfair Display',serif}}
  .visit-circle.filled{{background:#e6e1d5;border-color:#e6e1d5;color:#0a0a09;font-weight:700}}
  .counter-stats{{display:flex;justify-content:space-around;text-align:center;padding-top:12px;border-top:1px solid rgba(255,255,255,.07)}}
  .counter-stat .n{{font-family:'Playfair Display',serif;font-size:1.25rem;font-weight:600;color:#f2ede2}}
  .counter-stat .l{{font-size:0.58rem;letter-spacing:.06em;text-transform:uppercase;color:rgba(230,225,215,.45);margin-top:2px}}

  .list-label{{font-size:0.62rem;letter-spacing:.18em;text-transform:uppercase;color:rgba(230,225,215,.5);margin:6px 0 10px}}
  .hist-row{{display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid rgba(255,255,255,.06);font-size:0.82rem}}
  .hist-date{{color:rgba(230,225,215,.5)}}
  .hist-note{{color:#f2ede2}}
  .hist-empty{{text-align:center;padding:16px 0;color:rgba(230,225,215,.35);font-size:0.8rem}}

  .completed-banner{{text-align:center;background:#131210;border:1px solid rgba(255,255,255,.1);border-radius:14px;padding:20px 18px;margin:16px 0}}
  .completed-title{{font-family:'Playfair Display',serif;font-size:1.05rem;margin-bottom:12px}}
  .new-membership-btn{{display:inline-block;background:#e6e1d5;color:#0a0a09;text-decoration:none;padding:11px 22px;border-radius:8px;font-size:0.82rem;font-weight:600}}
  .wallet-btn{{display:block;text-align:center;background:#000000;color:#ffffff;text-decoration:none;padding:12px 20px;border-radius:10px;font-size:0.85rem;font-weight:600;border:1px solid rgba(255,255,255,.15);margin:16px 0}}

  .qr-section{{text-align:center;margin-top:auto;padding-top:18px}}
  .qr-section img{{border-radius:10px;border:1px solid rgba(255,255,255,.1)}}
  .qr-caption{{font-size:0.64rem;color:rgba(230,225,215,.4);margin-top:8px}}
</style>
</head>
<body>
<div class="wrap">

  <div class="flip-wrap">
    <div class="flip-card" id="flipCard">

      <div class="face face-front" id="cardFront">
        <div class="brand-frame"></div>
        <div class="brand-inner">
          <img class="brand-logo-img" src="data:image/png;base64,{_get_logo_b64()}" alt="R&amp;J Grooming">
        </div>
        <div class="tap-hint">Нажмите, чтобы открыть абонемент</div>
      </div>

      <div class="face face-back">
        <div class="back-head">
          <div>
            <div class="mem-id-label">Абонемент</div>
            <div class="mem-id">{m.get('id','')}</div>
          </div>
          <button class="flip-back-btn" id="flipBackBtn">← Назад</button>
        </div>

        <div class="info-rows">
          <div class="info-row"><div class="ico">{icon_person}</div><div><div class="info-label">Владелец</div><div class="info-val">{m.get('client_name','—')}</div></div></div>
          <div class="info-row"><div class="ico">{icon_paw}</div><div><div class="info-label">Питомец</div><div class="info-val">{m.get('pet_name','—')} · {m.get('pet_type','—')}</div></div></div>
          <div class="info-row"><div class="ico">{icon_tag}</div><div><div class="info-label">Абонемент</div><div class="info-val">{m.get('plan_name','—')}</div></div></div>
          <div class="info-row two-col">
            <div class="info-sub"><div class="ico">{icon_cal}</div><div><div class="info-label">Приобретён</div><div class="info-val">{m.get('purchase_date','—')}</div></div></div>
            <div class="info-sub"><div class="ico">{icon_cal}</div><div><div class="info-label">До</div><div class="info-val">{m.get('expiry_date','—')}</div></div></div>
          </div>
        </div>

        <div class="counter-card">
          <div class="counter-label">Посещения</div>
          <div class="visit-circles">{circles_html}</div>
          <div class="counter-stats">
            <div class="counter-stat"><div class="n">{used}</div><div class="l">использовано</div></div>
            <div class="counter-stat"><div class="n">{remaining}</div><div class="l">осталось</div></div>
            <div class="counter-stat"><div class="n">{total}</div><div class="l">всего</div></div>
          </div>
        </div>

        {wallet_button_html}

        {completed_banner}

        <div class="list-label">История посещений</div>
        {history_html}

        <div class="qr-section">
          <div class="qr-caption">{m.get('id','')} · rjgrooming.salon</div>
        </div>
      </div>

    </div>
  </div>

</div>
<script>
var flipCard = document.getElementById('flipCard');
var flipBackBtn = document.getElementById('flipBackBtn');
document.getElementById('cardFront').addEventListener('click', function(){{
  flipCard.classList.add('flipped');
}});
flipBackBtn.addEventListener('click', function(e){{
  e.stopPropagation();
  flipCard.classList.remove('flipped');
}});
</script>
</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/admin/migrate-client-data")
def admin_migrate_client_data():
    """Разовая миграция: старые отдельные файлы rjgrooming/client_data/*.json
    переносятся в единый rjgrooming/client_data_index.json."""
    import base64 as _b64
    auth = _b64.b64encode(f"{CLOUDINARY_API_KEY}:{CLOUDINARY_API_SECRET}".encode()).decode()
    headers = {"Authorization": f"Basic {auth}"}

    try:
        r = requests.get(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/resources/raw",
            headers=headers,
            params={"prefix": "rjgrooming/client_data/", "type": "upload", "max_results": 500},
            timeout=20
        )
        listing = r.json()
    except Exception as e:
        return jsonify({"success": False, "error": f"list failed: {e}"}), 500

    resources = listing.get("resources", [])
    all_data = _load_all_client_data()
    migrated = []
    errors = []

    for res in resources:
        public_id = res.get("public_id", "")
        key = public_id.split("/")[-1]
        secure_url = res.get("secure_url")
        if not secure_url:
            continue
        try:
            fr = requests.get(secure_url, timeout=10)
            if fr.status_code == 200:
                all_data[key] = fr.json()
                migrated.append(key)
        except Exception as e:
            errors.append(f"{key}: {e}")

    ok = _save_all_client_data(all_data)
    return jsonify({
        "success": ok,
        "found_old_files": len(resources),
        "migrated_count": len(migrated),
        "migrated_keys": migrated,
        "errors": errors,
        "total_in_new_index": len(all_data)
    })

@app.route("/api/breed-prices")
def api_breed_prices():
    import base64 as _b64_mod
    try:
        html = _b64_mod.b64decode(BOOKING_HTML_B64).decode("utf-8")
        m = re.search(r"var DATA = (\[.*?\]);", html, re.DOTALL)
        if not m:
            return jsonify({"success": False, "error": "Не удалось найти данные о ценах"}), 500
        data = json.loads(m.group(1))
        result = [{"breed": d["breed"], "services": d["services"]} for d in data]
        return jsonify({"success": True, "breeds": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/admin/debug-wallet-update")
def admin_debug_wallet_update():
    """Диагностика: показывает, что реально хранится у абонемента про wallet, и пробует PUT-обновление."""
    mid = request.args.get("id", "").strip()
    if not mid:
        return jsonify({"error": "укажи ?id=RJ-00X в ссылке"}), 400
    memberships = _load_memberships()
    m = memberships.get(mid)
    if not m:
        return jsonify({"error": f"абонемент {mid} не найден"}), 404

    result = {
        "membership_id": mid,
        "used_visits": m.get("used_visits"),
        "total_visits": m.get("total_visits"),
        "wallet_serial_stored": m.get("wallet_serial"),
        "wallet_share_url_stored": m.get("wallet_share_url"),
        "WALLETWALLET_API_KEY_present": bool(WALLETWALLET_API_KEY),
    }

    serial = m.get("wallet_serial")
    if not serial:
        result["note"] = "У этого абонемента нет сохранённого wallet_serial — пасс либо не создавался, либо создание не удалось при создании абонемента."
        return jsonify(result)

    payload = _wallet_pass_payload(m)
    result["payload_sent"] = payload
    try:
        r = requests.put(
            f"https://api.walletwallet.dev/api/passes/{serial}",
            headers={"Authorization": f"Bearer {WALLETWALLET_API_KEY}", "Content-Type": "application/json"},
            json=payload,
            timeout=15,
        )
        result["status_code"] = r.status_code
        try:
            result["response_json"] = r.json()
        except Exception:
            result["response_text"] = r.text[:1000]
    except Exception as e:
        result["exception"] = str(e)
    return jsonify(result)

@app.route("/admin/debug-wallet-test")
def admin_debug_wallet_test():
    """Диагностика: пробует создать тестовый Wallet-пасс и показывает точный ответ WalletWallet API."""
    key_present = bool(WALLETWALLET_API_KEY)
    key_preview = (WALLETWALLET_API_KEY[:12] + "...") if WALLETWALLET_API_KEY else "(не задан)"
    fake_m = {
        "id": "TEST-DEBUG", "client_name": "Тест Тестов", "pet_name": "Тест",
        "total_visits": 5, "used_visits": 1, "expiry_date": "01.01.2027",
        "plan_name": "Тестовый абонемент", "purchase_date": "01.01.2026",
    }
    payload = _wallet_pass_payload(fake_m)
    result = {"WALLETWALLET_API_KEY_present": key_present, "key_preview": key_preview, "payload_sent": payload}
    if not key_present:
        result["error"] = "WALLETWALLET_API_KEY не установлена в Railway Variables"
    else:
        try:
            r = requests.post(
                "https://api.walletwallet.dev/api/passes",
                headers={"Authorization": f"Bearer {WALLETWALLET_API_KEY}", "Content-Type": "application/json"},
                json=payload,
                timeout=15,
            )
            result["status_code"] = r.status_code
            try:
                result["response_json"] = r.json()
            except Exception:
                result["response_text"] = r.text[:1000]
        except Exception as e:
            result["exception"] = str(e)
    return jsonify(result)

@app.route("/admin/debug-memberships")
def admin_debug_memberships():
    import base64 as _b64_mod
    debug = {}

    debug["env_check"] = {
        "cloud_name": CLOUDINARY_CLOUD_NAME,
        "api_key_set": bool(CLOUDINARY_API_KEY),
        "api_secret_set": bool(CLOUDINARY_API_SECRET),
    }

    try:
        auth = _b64_mod.b64encode(f"{CLOUDINARY_API_KEY}:{CLOUDINARY_API_SECRET}".encode()).decode()
        meta_r = requests.get(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/resources/raw/upload/{_MEMBERSHIP_INDEX_PUBLIC_ID}",
            headers={"Authorization": f"Basic {auth}"},
            timeout=8
        )
        debug["admin_api_status"] = meta_r.status_code
        debug["admin_api_response"] = meta_r.json()
    except Exception as e:
        debug["admin_api_error"] = str(e)

    debug["load_memberships_result"] = _load_memberships()
    debug["load_memberships_keys"] = list(debug["load_memberships_result"].keys())

    return jsonify(debug)

@app.route("/admin")
def admin_hub():

    due_badge = ""
    try:
        rows, _ = _get_reminder_dashboard_rows()
        due_now = sum(1 for r in rows if (r["status"] == "stage1" and not r["stage1_done"]) or (r["status"] == "done" and not r["stage2_done"]))
        if due_now:
            due_badge = f'<span class="card-badge">{due_now}</span>'
    except Exception:
        pass

    P = "?pass=anza1985"
    import urllib.parse as _urlp_hub
    _calendar_id = "50218a0d445e1be8f510d21128b82c0d33b6a78c57bb5a4ea9d14eb2fbcfeaa6@group.calendar.google.com"
    gcal_link = f"https://calendar.google.com/calendar/u/0/r?cid={_urlp_hub.quote(_calendar_id)}"
    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>R&J Grooming — Панель администратора</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,500;0,600;0,700;1,500&family=Montserrat:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a09;color:#f2ede2;font-family:'Montserrat',sans-serif;padding:44px 20px 80px}}
  .wrap{{max-width:680px;margin:0 auto}}
  .eyebrow{{font-size:0.68rem;letter-spacing:.3em;text-transform:uppercase;color:rgba(242,237,226,.4);margin-bottom:10px}}
  h1{{font-family:'Playfair Display',serif;font-weight:600;font-size:2.3rem;margin-bottom:4px}}
  .sub{{font-size:0.8rem;color:rgba(242,237,226,.5);margin-bottom:40px}}
  .section-label{{font-size:0.66rem;letter-spacing:.2em;text-transform:uppercase;color:#c9a05a;margin:28px 0 14px}}
  .cards{{display:flex;flex-direction:column;gap:12px}}
  .card{{position:relative;display:flex;align-items:center;gap:16px;background:#141310;border:1px solid rgba(201,160,90,.18);border-radius:14px;padding:20px 22px;text-decoration:none;color:#f2ede2;transition:border-color .15s}}
  .card:active{{border-color:rgba(201,160,90,.5)}}
  .card-icon{{font-size:1.6rem;width:48px;height:48px;border-radius:12px;background:rgba(201,160,90,.12);display:flex;align-items:center;justify-content:center;flex-shrink:0}}
  .card-txt{{flex:1}}
  .card-name{{font-family:'Playfair Display',serif;font-size:1.25rem;font-weight:600}}
  .card-desc{{font-size:0.78rem;color:rgba(242,237,226,.5);margin-top:3px}}
  .card-arrow{{color:rgba(201,160,90,.6);font-size:1.2rem}}
  .card-badge{{position:absolute;top:-8px;right:-8px;background:#e0824a;color:#0a0a09;font-size:0.72rem;font-weight:700;min-width:22px;height:22px;border-radius:12px;display:flex;align-items:center;justify-content:center;padding:0 6px}}
  .secondary{{display:flex;gap:10px;flex-wrap:wrap}}
  .schip{{flex:1;min-width:140px;text-align:center;background:#141310;border:1px solid rgba(255,255,255,.08);border-radius:10px;padding:12px 10px;text-decoration:none;color:rgba(242,237,226,.75);font-size:0.78rem}}
</style>
</head>
<body>
<div class="wrap">
  <div class="eyebrow">R&J Grooming</div>
  <h1>Панель администратора</h1>
  <div class="sub">Бронирование, статистика и напоминания — в одном месте</div>

  <div class="section-label">Модули</div>
  <div class="cards">
    <a class="card" href="/app{P}">
      <div class="card-icon">📅</div>
      <div class="card-txt">
        <div class="card-name">Бронирование</div>
        <div class="card-desc">Виджет онлайн-записи для клиентов</div>
      </div>
      <div class="card-arrow">→</div>
    </a>
    <a class="card" href="/stats{P}">
      <div class="card-icon">📊</div>
      <div class="card-txt">
        <div class="card-name">Статистика</div>
        <div class="card-desc">Выручка, записи, доли по мастерам</div>
      </div>
      <div class="card-arrow">→</div>
    </a>
    <a class="card" href="/admin/reminders{P}">
      {due_badge}
      <div class="card-icon">🔔</div>
      <div class="card-txt">
        <div class="card-name">Напоминания</div>
        <div class="card-desc">Клиенты без визита 35+ дней</div>
      </div>
      <div class="card-arrow">→</div>
    </a>
    <a class="card" href="/admin/clients{P}">
      <div class="card-icon">👥</div>
      <div class="card-txt">
        <div class="card-name">Клиенты</div>
        <div class="card-desc">Вся база — история визитов и контакты</div>
      </div>
      <div class="card-arrow">→</div>
    </a>
    <a class="card" href="/admin/memberships{P}">
      <div class="card-icon">🎫</div>
      <div class="card-txt">
        <div class="card-name">Абонементы</div>
        <div class="card-desc">Цифровые карты посещений</div>
      </div>
      <div class="card-arrow">→</div>
    </a>
    <a class="card" href="{gcal_link}" target="_blank" rel="noopener">
      <div class="card-icon">🗓️</div>
      <div class="card-txt">
        <div class="card-name">Google Calendar</div>
        <div class="card-desc">Сам календарь салона — все брони</div>
      </div>
      <div class="card-arrow">→</div>
    </a>
  </div>
</div>
</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/admin/search")
def admin_client_search():

    query = (request.args.get("q") or "").strip()
    results = []
    error = None

    if query:
        try:
            today = datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date()
            bookings = _fetch_full_history_bookings()

            q_lower = query.lower()
            q_digits = re.sub(r"[^\d]", "", query)

            by_phone = {}
            for b in bookings:
                name = (b.get("clientName") or "")
                phone = (b.get("clientPhone") or "")
                pet = (b.get("petName") or "")
                phone_digits = re.sub(r"[^\d]", "", phone)
                match = (
                    (q_lower and (q_lower in name.lower() or q_lower in pet.lower())) or
                    (q_digits and q_digits in phone_digits)
                )
                if not match:
                    continue
                key = phone or name
                try:
                    d = datetime.strptime(b.get("date", ""), "%d.%m.%Y").date()
                except Exception:
                    d = None
                if key not in by_phone or (d and (not by_phone[key]["_d"] or d > by_phone[key]["_d"])):
                    by_phone[key] = {
                        "_d": d, "name": name, "phone": phone, "pet": pet,
                        "breed": b.get("breed", ""), "master": b.get("master", ""),
                        "date": b.get("date", ""), "service": b.get("service", "")
                    }
            all_client_data = _load_all_client_data()
            for key, entry in by_phone.items():
                saved = all_client_data.get(_client_data_key(entry["phone"]), {}) if entry["phone"] else {}
                if saved.get("name"):
                    entry["name"] = saved["name"]
                entry["display_phone"] = _normalize_phone(saved.get("phone_override") or entry["phone"])

            results = sorted(by_phone.values(), key=lambda x: x["_d"] or datetime.min.date(), reverse=True)
        except Exception as e:
            error = str(e)

    def result_html(r):
        import urllib.parse as _urlp_s
        display_phone = r.get("display_phone") or r["phone"]
        wa_digits = re.sub(r"[^\d]", "", display_phone or "")
        phone_encoded = _urlp_s.quote(r["phone"] or "")
        return f"""
        <a class="row" href="/admin/client?phone={phone_encoded}&pass=anza1985">
          <div class="row-top">
            <div class="who">
              <span class="name">{r['name'] or '—'}</span>
              <span class="pet">{r['pet'] or '—'} · {r['breed'] or '—'}</span>
            </div>
            <span class="row-arrow">→</span>
          </div>
          <div class="row-bottom">
            <span>Последний визит: {r['date'] or '—'} · {r['master'] or '—'}</span>
          </div>
          <div class="row-bottom" style="margin-top:6px">
            <span class="contacts">
              <span onclick="event.stopPropagation();location.href='tel:{display_phone}'" style="color:#c9a05a">{display_phone or '—'}</span>
              <span onclick="event.stopPropagation();window.open('https://wa.me/{wa_digits}','_blank')" class="wa">WhatsApp</span>
            </span>
          </div>
          <div class="row-hint">Скопируй имя выше и вставь в поиск Instagram Direct, чтобы найти переписку</div>
        </a>"""

    if query and not results and not error:
        results_html = '<div class="empty">Ничего не найдено — проверь написание или попробуй телефон/кличку</div>'
    elif error:
        results_html = f'<div class="empty" style="color:#e0824a">Ошибка: {error}</div>'
    elif not query:
        results_html = '<div class="empty">Введи имя, телефон или кличку питомца</div>'
    else:
        results_html = "".join(result_html(r) for r in results)

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>R&J Grooming — Поиск клиента</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,500;0,600;0,700;1,500&family=Montserrat:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a09;color:#f2ede2;font-family:'Montserrat',sans-serif;padding:36px 20px 80px}}
  .wrap{{max-width:640px;margin:0 auto}}
  .eyebrow{{font-size:0.68rem;letter-spacing:.3em;text-transform:uppercase;color:rgba(242,237,226,.4);margin-bottom:10px}}
  .back-link{{display:inline-block;font-size:0.74rem;color:rgba(201,160,90,.75);text-decoration:none;margin-bottom:16px}}
  h1{{font-family:'Playfair Display',serif;font-weight:600;font-size:2.1rem;margin-bottom:20px}}
  form{{display:flex;gap:8px;margin-bottom:28px}}
  input[type=text]{{flex:1;background:#151310;border:1px solid rgba(201,160,90,.3);border-radius:10px;padding:14px 16px;color:#f2ede2;font-family:'Montserrat',sans-serif;font-size:0.95rem}}
  input[type=text]:focus{{outline:none;border-color:#c9a05a}}
  button{{background:#c9a05a;color:#0a0a09;border:none;border-radius:10px;padding:0 22px;font-weight:600;font-size:0.9rem;cursor:pointer}}
  .row{{display:block;background:#131210;border:1px solid rgba(255,255,255,.06);border-radius:12px;padding:16px 18px;margin-bottom:10px;text-decoration:none;color:inherit}}
  .row-top{{display:flex;justify-content:space-between;align-items:flex-start;gap:10px;margin-bottom:10px}}
  .who{{display:flex;flex-direction:column}}
  .name{{font-family:'Playfair Display',serif;font-size:1.2rem;font-weight:600}}
  .pet{{font-size:0.76rem;color:rgba(242,237,226,.5);margin-top:2px}}
  .row-arrow{{color:rgba(201,160,90,.6)}}
  .row-bottom{{font-size:0.78rem;color:rgba(242,237,226,.55)}}
  .contacts{{display:flex;align-items:center;gap:10px}}
  .wa{{color:#4ade80;font-size:0.72rem;font-weight:600;border:1px solid rgba(74,222,128,.35);border-radius:20px;padding:3px 10px;cursor:pointer}}
  .row-hint{{margin-top:10px;padding-top:10px;border-top:1px solid rgba(255,255,255,.06);font-size:0.7rem;color:rgba(201,160,90,.7);font-style:italic}}
  .empty{{text-align:center;padding:40px 0;color:rgba(242,237,226,.4);font-size:0.85rem}}
</style>
</head>
<body>
<div class="wrap">
  <a href="/admin?pass=anza1985" class="back-link">← Админ-панель</a>
  <div class="eyebrow">R&J Grooming · Поиск</div>
  <h1>Найти клиента</h1>
  <form method="get">
    <input type="hidden" name="pass" value="anza1985">
    <input type="text" name="q" placeholder="Имя, телефон или кличка питомца" value="{query}" autofocus>
    <button type="submit">Искать</button>
  </form>
  {results_html}
</div>
</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/admin/reminders")
def admin_reminders_dashboard():

    try:
        rows, today = _get_reminder_dashboard_rows()
        error = None
    except Exception as e:
        rows, today, error = [], datetime.now(_REMINDER_TZ).date() if _REMINDER_TZ else datetime.utcnow().date(), str(e)

    stage1_count = sum(1 for r in rows if r["status"] == "stage1" and not r["stage1_done"])
    stage2_count = sum(1 for r in rows if r["status"] == "done" and not r["stage2_done"])
    pending_count = sum(1 for r in rows if r["status"] == "pending")

    STATUS_META = {
        "pending": {"color": "#8a8578", "label": "ожидание"},
        "stage1":  {"color": "#c9a05a", "label": "напоминание #1"},
        "done":    {"color": "#e0824a", "label": "напоминание #2"},
    }

    def bar_color(days_since):
        if days_since < 35:
            return "#8a8578"
        elif days_since == 35:
            return "#4ade80"
        else:
            return "#e0524a"

    def row_html(r):
        import urllib.parse as _urlp
        pct = min(100, round(r["days_since"] / 42 * 100))
        meta = STATUS_META[r["status"]]
        bar = bar_color(r["days_since"])
        email_html = f'<a class="email" href="mailto:{r["email"]}">{r["email"]}</a>' if r.get("email") else '<span class="email-empty">email не указан</span>'
        wa_digits = re.sub(r"[^\d]", "", r.get("display_phone") or r["phone"] or "")
        date_str = r["date"].strftime("%d.%m.%Y")

        first_name = (r["name"] or "").split()[0] if r["name"] else ""
        pet = r["pet"] or "питомец"
        greeting = f"Здравствуйте, {first_name}! 🐾" if first_name else "Здравствуйте! 🐾"
        booking_link = "https://rjgrooming.up.railway.app/app"

        def _days_word(n):
            n10, n100 = n % 10, n % 100
            if 11 <= n100 <= 14:
                return "дней"
            if n10 == 1:
                return "день"
            if 2 <= n10 <= 4:
                return "дня"
            return "дней"

        days_word = _days_word(r["days_since"])
        if r["days_since"] >= 42:
            msg = (f"{greeting} {pet} давно не был{'а' if pet.endswith(('а','я')) else ''} у нас — "
                   f"прошло уже {r['days_since']} {days_word} с последнего визита. "
                   f"Будем рады снова вас видеть в R&J Grooming! Записаться можно здесь: {booking_link}")
        elif r["days_since"] >= 35:
            msg = (f"{greeting} Прошло {r['days_since']} {days_word} с последнего визита {pet} к нам — "
                   f"самое время подумать о следующем груминге. Будем рады видеть вас снова! "
                   f"Записаться можно здесь: {booking_link}")
        else:
            msg = ""
        wa_href = f"https://wa.me/{wa_digits}?text={_urlp.quote(msg)}" if msg else f"https://wa.me/{wa_digits}"

        checks = ""
        if r["days_since"] >= 35:
            c1 = "checked" if r["stage1_done"] else ""
            checks += f'<label class="chk"><input type="checkbox" data-phone="{r["phone"]}" data-date="{date_str}" data-stage="1" {c1}> Напоминание #1 отправлено</label>'
        if r["days_since"] >= 42:
            c2 = "checked" if r["stage2_done"] else ""
            checks += f'<label class="chk"><input type="checkbox" data-phone="{r["phone"]}" data-date="{date_str}" data-stage="2" {c2}> Напоминание #2 отправлено</label>'
        checks_html = f'<div class="row-checks">{checks}</div>' if checks else ""

        if r["days_since"] >= 42:
            needs_action = not r["stage2_done"]
        elif r["days_since"] >= 35:
            needs_action = not r["stage1_done"]
        else:
            needs_action = False
        pulse_class = " pulse" if needs_action else ""

        return f"""
        <div class="row{pulse_class}">
          <div class="row-top">
            <div class="who">
              <span class="name">{r['name'] or '—'}</span>
              <span class="pet">{r['pet'] or '—'} · {r['breed'] or '—'}</span>
            </div>
            <div class="badge" style="color:{meta['color']};border-color:{meta['color']}55;background:{meta['color']}14">{meta['label']}</div>
          </div>
          <div class="track">
            <div class="fill" style="width:{pct}%;background:{bar}"></div>
            <div class="tick" style="left:{35/42*100:.2f}%"></div>
            <div class="tick" style="left:100%"></div>
          </div>
          <div class="row-bottom">
            <span>Визит {date_str} · {r['master'] or '—'}</span>
            <span class="days">{r['days_since']} дн.</span>
            <span class="contacts">
              <a class="phone" href="tel:{r.get('display_phone') or r['phone']}">{r.get('display_phone') or r['phone']}</a>
              <a class="wa" href="{wa_href}" target="_blank" rel="noopener">WhatsApp{' с текстом' if msg else ''}</a>
            </span>
          </div>
          <div class="row-contact">{email_html}</div>
          {checks_html}
        </div>"""

    rows_html = "".join(row_html(r) for r in rows) if rows else '<div class="empty">За последние 40 дней визитов не найдено</div>'
    error_html = f'<div class="empty" style="color:#e0824a">Ошибка загрузки данных: {error}</div>' if error else ""

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>R&J Grooming — Напоминания</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,500;0,600;0,700;1,500&family=Montserrat:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0a0a09;color:#f2ede2;font-family:'Montserrat',sans-serif;padding:36px 20px 80px}}
  .wrap{{max-width:720px;margin:0 auto}}
  .eyebrow{{font-size:0.68rem;letter-spacing:.3em;text-transform:uppercase;color:rgba(242,237,226,.4);margin-bottom:10px}}
  .back-link{{display:inline-block;font-size:0.74rem;color:rgba(201,160,90,.75);text-decoration:none;margin-bottom:16px}}
  h1{{font-family:'Playfair Display',serif;font-weight:600;font-size:2.1rem;margin-bottom:6px}}
  .sub{{font-size:0.78rem;color:rgba(242,237,226,.5);margin-bottom:32px}}
  .stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:36px}}
  .stat{{background:#151310;border:1px solid rgba(201,160,90,.18);border-radius:12px;padding:18px 14px}}
  .stat .n{{font-family:'Playfair Display',serif;font-size:1.9rem;font-weight:600;color:#c9a05a}}
  .stat .l{{font-size:0.66rem;letter-spacing:.08em;text-transform:uppercase;color:rgba(242,237,226,.5);margin-top:4px}}
  .list-label{{font-size:0.68rem;letter-spacing:.2em;text-transform:uppercase;color:#c9a05a;margin-bottom:14px}}
  .row{{background:#131210;border:1px solid rgba(255,255,255,.06);border-radius:12px;padding:16px 18px;margin-bottom:10px}}
  .row.pulse{{animation:rowPulse 1.8s ease-in-out infinite}}
  @keyframes rowPulse{{
    0%,100%{{border-color:rgba(224,82,74,.25);box-shadow:0 0 0 0 rgba(224,82,74,.15)}}
    50%{{border-color:rgba(224,82,74,.75);box-shadow:0 0 0 4px rgba(224,82,74,.08)}}
  }}
  .row-top{{display:flex;justify-content:space-between;align-items:flex-start;gap:10px;margin-bottom:12px}}
  .who{{display:flex;flex-direction:column}}
  .name{{font-family:'Playfair Display',serif;font-size:1.15rem;font-weight:600}}
  .pet{{font-size:0.76rem;color:rgba(242,237,226,.5);margin-top:2px}}
  .badge{{font-size:0.62rem;letter-spacing:.06em;text-transform:uppercase;padding:5px 10px;border-radius:20px;border:1px solid;white-space:nowrap;flex-shrink:0}}
  .track{{position:relative;height:6px;background:rgba(255,255,255,.08);border-radius:4px;margin-bottom:10px}}
  .fill{{position:absolute;top:0;left:0;height:100%;border-radius:4px;transition:width .3s}}
  .tick{{position:absolute;top:-3px;width:2px;height:12px;background:rgba(242,237,226,.25)}}
  .row-bottom{{display:flex;justify-content:space-between;align-items:center;font-size:0.74rem;color:rgba(242,237,226,.55)}}
  .row-bottom .days{{font-weight:600;color:#f2ede2}}
  .phone{{color:#c9a05a;text-decoration:none}}
  .contacts{{display:flex;align-items:center;gap:10px}}
  .wa{{color:#4ade80;text-decoration:none;font-size:0.72rem;font-weight:600;border:1px solid rgba(74,222,128,.35);border-radius:20px;padding:3px 10px}}
  .row-contact{{margin-top:6px;font-size:0.74rem}}
  .row-contact .email{{color:rgba(242,237,226,.65);text-decoration:none}}
  .row-contact .email-empty{{color:rgba(242,237,226,.3);font-style:italic}}
  .row-checks{{margin-top:12px;padding-top:10px;border-top:1px solid rgba(255,255,255,.06);display:flex;flex-direction:column;gap:8px}}
  .chk{{display:flex;align-items:center;gap:8px;font-size:0.78rem;color:rgba(242,237,226,.75);cursor:pointer;user-select:none}}
  .chk input{{width:16px;height:16px;accent-color:#c9a05a;cursor:pointer}}
  .empty{{text-align:center;padding:40px 0;color:rgba(242,237,226,.4);font-size:0.85rem}}
  @media(max-width:480px){{
    .row-bottom{{flex-direction:column;align-items:flex-start;gap:4px}}
  }}
</style>
</head>
<body>
<div class="wrap">
  <a href="/admin?pass=anza1985" class="back-link">← Админ-панель</a>
  <div class="eyebrow">R&J Grooming · Напоминания</div>
  <h1>Клиенты без визита</h1>
  <div class="sub">Окно 40 дней от последнего визита · {today.strftime('%d.%m.%Y')}</div>

  <div class="stats">
    <div class="stat"><div class="n" id="statStage1">{stage1_count}</div><div class="l">напоминание #1</div></div>
    <div class="stat"><div class="n" id="statStage2">{stage2_count}</div><div class="l">напоминание #2</div></div>
    <div class="stat"><div class="n">{pending_count}</div><div class="l">ожидают</div></div>
  </div>

  <div class="list-label">Все клиенты ({len(rows)})</div>
  {error_html}
  {rows_html}
</div>
<script>
document.addEventListener('change', function(e){{
  if(e.target && e.target.matches('.chk input[type="checkbox"]')){{
    var el = e.target;
    var stage = parseInt(el.getAttribute('data-stage'), 10);
    var counterEl = document.getElementById(stage === 1 ? 'statStage1' : 'statStage2');
    if(counterEl){{
      var n = parseInt(counterEl.textContent, 10) || 0;
      counterEl.textContent = el.checked ? Math.max(0, n - 1) : n + 1;
    }}
    fetch('/api/mark-reminder', {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify({{
        phone: el.getAttribute('data-phone'),
        date: el.getAttribute('data-date'),
        stage: stage,
        done: el.checked
      }})
    }}).catch(function(err){{ console.error('mark-reminder failed', err); }});
  }}
}});
</script>
</body>
</html>"""
    return html, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/admin/export-clients")
def admin_export_clients():
    from io import BytesIO
    from flask import send_file
    try:
        content, filename = _build_client_export_workbook()
    except Exception as e:
        return f"Ошибка получения данных: {e}", 500
    return send_file(
        BytesIO(content),
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def send_weekly_client_export_email():
    resend_key = os.environ.get("RESEND_API_KEY")
    if not resend_key:
        print("WEEKLY EXPORT: RESEND_API_KEY не задан", flush=True)
        return
    try:
        import base64 as _b64
        content, filename = _build_client_export_workbook()
        requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
            json={
                "from": "booking@rjgrooming.salon",
                "to": ["myrnj1@gmail.com"],
                "subject": f"R&J Grooming — еженедельная выгрузка клиентов ({datetime.now().strftime('%d.%m.%Y')})",
                "html": "<p>Во вложении — актуальная выгрузка всех клиентов и лист с напоминаниями (+35/+42 дня).</p>",
                "attachments": [{
                    "filename": filename,
                    "content": _b64.b64encode(content).decode("ascii")
                }]
            },
            timeout=30
        )
        print("WEEKLY EXPORT: письмо отправлено", flush=True)
    except Exception as e:
        print(f"WEEKLY EXPORT ERROR: {e}", flush=True)

def _weekly_export_scheduler_loop():
    while True:
        try:
            now = datetime.now(_REMINDER_TZ) if _REMINDER_TZ else datetime.utcnow()
            next_run = now.replace(hour=9, minute=0, second=0, microsecond=0)
            days_ahead = (0 - next_run.weekday()) % 7  # 0 = понедельник
            if days_ahead == 0 and next_run <= now:
                days_ahead = 7
            next_run += timedelta(days=days_ahead)
            sleep_seconds = (next_run - now).total_seconds()
            _time.sleep(max(sleep_seconds, 60))
            send_weekly_client_export_email()
        except Exception as e:
            print(f"WEEKLY EXPORT SCHEDULER ERROR: {e}", flush=True)
            _time.sleep(3600)

threading.Thread(target=_weekly_export_scheduler_loop, daemon=True).start()

@app.route("/api/send-weekly-export")
def api_send_weekly_export():
    send_weekly_client_export_email()
    return jsonify({"success": True})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
